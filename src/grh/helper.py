# helper.py

import datetime
import secrets
from time import timezone
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse
import openpyxl
from inov import settings
from django.template.loader import get_template
from django.utils.html import strip_tags

from production.models import Aliment, AlimentFormule, Police
from shared.enum import Statut, StatutValidite

from django.utils import timezone





def generate_uiid(request):
    uiid_length = 20
    uiid = secrets.token_urlsafe(uiid_length)
    return uiid


def send_email(url, libelle_compagne, date_debut, date_fin, email, uiid, aliment):
    subject = 'INOV : INVITATION À VOUS INSCRIRE'
    template_name = "grh/enrolement/email_to_enrol.html"
    template = get_template(template_name)
    context = {
        'libelle_compagne': libelle_compagne,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'url': url,
        'uiid': uiid,
        'aliment': aliment,
    }
    html_message = template.render(context)
    plain_message = strip_tags(html_message)
    from_email = settings.EMAIL_HOST_USER
    try:
        send_mail(subject, plain_message, from_email, [email], html_message=html_message, fail_silently=False)
    except Exception as e:
        print(f"Error sending email to {email}: {e}")


def format_phone_number(raw_phone, selected_country_dial_code):
    raw_phone = raw_phone.replace(" ", "")
    formatted_phone = f"+{selected_country_dial_code}{raw_phone}"
    return formatted_phone


def export_beneficiaire(request, police_id):
    police = Police.objects.get(id=police_id)
    today = datetime.datetime.now(tz=timezone.utc)
    
    if police:
        #queryset = Aliment.objects.select_related('qualite_beneficiaire', 'police__client').filter(formulegarantie__police=police=police).order_by('-id')
        # Optimisation de la recuperation des beneficiaires
        #queryset = AlimentFormule.objects.filter(formule_id__in=[p.id for p in police.formules], statut=Statut.ACTIF).order_by('-id')

        # Récupérer les IDs des aliments dans aliment_formule
        aliment_ids = AlimentFormule.objects.filter(
            formule_id__in=[p.id for p in police.formules],
            statut=Statut.ACTIF,
            statut_validite=StatutValidite.VALIDE
        ).values_list('aliment_id', flat=True)

        # Récupérer les aliments correspondants à ces IDs
        queryset = Aliment.objects.filter(id__in=aliment_ids).order_by('-adherent_principal_id')

        print("queryset")
        print(queryset)
        print(queryset.count())

        # Exportation excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_DES_BENEFICIAIRES-POLICE_'+str(police.numero)+'.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'LISTE DES BENEFICIAIRES'

        # Write header row
        header = [
            'NUMERO_POLICE',
            'NUMERO_CARTE',
            'NOM',
            'PRENOMS',
            'DATE_NAISSANCE',
            'GENRE',
            'MATRICULE',
            'ADHERENT_PRINCIPAL',
            'LIEN',
            'NUMERO_FAMILLE',
            'DATE_ENTREE',
            'DATE_SORTI',
            'FORMULE',
            'TELEPHONE',
            'ETAT',
        ]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        data = []
        liste_aliments_ajoutes = []

        for aliment in queryset:

            if aliment.id not in liste_aliments_ajoutes:
                liste_aliments_ajoutes.append(aliment.id)

            numero_carte = aliment.carte_active().numero if aliment.carte_active() else None

            derniere_formule = None
            derniere_formule_libelle = ''
            etat_beneficiaire = aliment.etat_beneficiaire_atdate(today)
            if aliment.date_sortie: #and af.date_sortie <= today:
                #etat_beneficiaire = 'SORTI'
                derniere_formule = AlimentFormule.objects.filter(aliment=aliment).order_by('-id').first()
                if derniere_formule and derniere_formule.formule:
                    derniere_formule_libelle = derniere_formule.formule.libelle


            #nom_client = f'{police.client.nom} {police.client.prenoms}'
            sa_formule = aliment.formule_atdate(today)
            sa_formule_libelle = sa_formule.libelle if sa_formule else ''


            data_iten = [
                police.numero,
                numero_carte,
                aliment.nom,
                aliment.prenoms,
                aliment.date_naissance.strftime("%d/%m/%Y") if aliment.date_naissance else '',
                aliment.genre,
                aliment.matricule_employe if aliment.matricule_employe else '',
                aliment.adherent_principal.nom + ' ' + aliment.adherent_principal.prenoms,
                aliment.qualite_beneficiaire.libelle if aliment.qualite_beneficiaire else '',
                aliment.adherent_principal.numero_famille if aliment.adherent_principal else aliment.numero_famille,
                aliment.date_affiliation.strftime("%d/%m/%Y") if aliment.date_affiliation else '',
                aliment.date_sortie.strftime("%d/%m/%Y") if aliment.date_sortie else '',
                sa_formule_libelle if sa_formule_libelle else derniere_formule_libelle,
                aliment.telephone_mobile,
                etat_beneficiaire,
            ]
            data.append(data_iten)

        for row_num, row in enumerate(data, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    else:
        return JsonResponse({
            "message": "Police non trouvée"
        }, status=404)


# 2fa 

def send_otp_mail(email, otp):
    subject = 'Code de Vérification'
    message = f'''<!DOCTYPE html>
                    <html lang="fr">
                    <head>
                        <meta charset="UTF-8">
                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                        <title>Code de Vérification</title>
                    </head>
                    <body>
                       <p style="color:red">Si cette action n'a pas été initiée par vous, veuillez ignorer ce message.</p>

                        <p>Bonjour,</p>
                        <p>Veuillez trouver ci-dessous votre code de vérification unique :</p>
                        
                        <p>Code de vérification : <strong>{otp}</strong></p>
                        
                        <p>Ce code est valable pour 5 minutes.</p>
                        <p>Cordialement,</p>
                        <p>L'équipe INOV</p>
                    </body>
                    </html>
                '''
    send_mail(
        subject=subject,
        message="",
        html_message=message,
        from_email=None,
        # auth_user=bureau.email_smtp,
        # auth_password=bureau.mot_de_passe_smtp,
        recipient_list=[email],
        fail_silently=True,
    )
