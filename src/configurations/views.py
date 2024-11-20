# Create your views here.
import datetime
import os
from ast import literal_eval
from decimal import Decimal
from pprint import pprint
from sqlite3 import Date

import pandas as pd
from django.contrib import admin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.models import Permission
from django.contrib.sessions.models import Session
from django.core import serializers
from django.core.cache import cache
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db.models import Q
from django.forms import model_to_dict
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.utils.translation import gettext as _
from django.views.generic import TemplateView
from django_dump_die.middleware import dd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from configurations.helper_config import verify_sql_query
from configurations.models import ActionLog, Prescripteur, PrescripteurPrestataire, Prestataire, Specialite, Secteur, \
    Bureau, TypeActe, \
    TypePrestataire, User, AuthGroup, TypeEtablissement, Tarif, Rubrique, RegroupementActe, Acte, ReseauSoin, \
    PrestataireReseauSoin, WsBoby, ParamWsBoby, Affection, BackgroundQueryTask, ParamProduitCompagnie, Compagnie, \
    AlimentMatricule, ParamActe
from inov import settings
# Create your views here.
from production.models import TarifPrestataireClient, Client, Aliment, AlimentFormule, Mouvement, MouvementAliment, \
    Carte, Quittance, Reglement
from production.templatetags.my_filters import money_field
from shared.enum import PasswordType, Statut, StatutValidite, BaseCalculTM, StatutPaiementSinistre, \
    SatutBordereauDossierSinistres, StatutSinistre
from shared.helpers import generate_numero_famille_for_existing_aliment, \
    generer_nombre_famille_du_mois_for_existing_aliment
from django.contrib import messages
from django.db import transaction

from sinistre.models import Sinistre, PaiementComptable, HistoriqueOrdonnancementSinistre, \
    HistoriquePaiementComptableSinistre, BordereauOrdonnancement
import json
import io


def generate_numero_famille_all(request):
    aliments = Aliment.objects.filter(qualite_beneficiaire__code="AD", numero_famille__isnull=True)
    if aliments:
        for aliment in aliments:
            numero_famille = generate_numero_famille_for_existing_aliment(aliment)
            aliment.numero_famille = numero_famille
            aliment.save()

        response = {
            'statut': 1,
            'message': str(aliments.count()) + ' NUMEROS DE FAMILLES GENERES AVEC SUCCÈS'
        }

    else:
        response = {
            'statut': 0,
            'message': 'AUCUN ADHERENT PRINCIPAL SANS NUMEROS DE FAMILLE TROUVE'
        }

    return JsonResponse(response)


def generer_nombre_famille_du_mois_all(request):
    aliments = Aliment.objects.filter(qualite_beneficiaire__code="AD", numero_famille_du_mois__isnull=True)

    if aliments:
        for aliment in aliments:
            aliment.numero_famille_du_mois = generer_nombre_famille_du_mois_for_existing_aliment(aliment)
            aliment.save()

        response = {
            'statut': 1,
            'message': str(aliments.count()) + ' NOMBRE FAMILLES DU MOIS GENERES AVEC SUCCÈS'
        }

    else:
        response = {
            'statut': 0,
            'message': 'AUCUN ADHERENT PRINCIPAL SANS NOMBRE FAMILLE DU MOIS TROUVE'
        }

    return JsonResponse(response)


#table mis en place après que les incorporations ai déjà été fait
#vider la table pour recréer toutes les lignes
def create_mouvements_incorporation_aliments(request):
    aliments_formules = AlimentFormule.objects.filter(statut_validite=StatutValidite.VALIDE)

    cpt = 0
    if aliments_formules:
        for af in aliments_formules:

            if af.formule and af.formule.police:

                police = af.formule.police
                aliment = af.aliment
                date_mouvement = af.date_debut
                motif = "Incorporation"
                created_by = af.created_by

                #vu que l'aliment peut avoir plusieurs formule,
                # on crée le mouvement uniquement si c'est une nouvelle police sur laquelle l'aliment se trouve
                # c a d s'il n'ay pas de ligne de mouvement avec l'aliment et la police

                #Vérifier s'il existe une ligne mouvement_aliment avec la police
                mouvement_aliment_existant = MouvementAliment.objects.filter(police=police, aliment=aliment)

                if not mouvement_aliment_existant:

                    # Créer le mouvement d'incorporation du bénéficiaire
                    mouvement = Mouvement.objects.filter(code="INCORPORATION").first()
                    mouvement_aliment = MouvementAliment.objects.create(created_by=created_by,
                                                                        aliment=aliment,
                                                                        mouvement=mouvement,
                                                                        police=police,
                                                                        date_effet=date_mouvement,
                                                                        motif=motif)
                    mouvement_aliment.save()

                    pprint("- mouvement d'incorporation crée - aliment_id: " + str(aliment.id))
                    cpt = cpt + 1


        response = {
            'statut': 1,
            'message': str(cpt) + " MOUVEMENTS D'INCORPORATION CREES AVEC SUCCÈS"
        }

    else:
        response = {
            'statut': 0,
            'message': "AUCUN ALIMENT SANS MOUVEMENT D'INCORPORATION TROUVE"
        }

    return JsonResponse(response)


#créer des mouvements de sortie pour les aliments dont la date sortie est renseigné
def create_mouvements_sortie_aliments(request):
    aliments_formules = AlimentFormule.objects.filter(statut_validite=StatutValidite.VALIDE)

    cpt = 0
    if aliments_formules:
        for af in aliments_formules:

            if af.formule and af.formule.police:

                if af.aliment.date_sortie:
                    police = af.formule.police
                    aliment = af.aliment
                    date_mouvement = af.aliment.date_sortie
                    motif = "Sortie effectuée par le système"
                    created_by = af.created_by


                    mouvement = Mouvement.objects.filter(code="SORTIE-BENEF").first()
                    #Vérifier s'il existe une ligne mouvement_aliment avec la police, l'aliment et le mouvement de sortie
                    mouvement_aliment_sortie_existant = MouvementAliment.objects.filter(police=police, aliment=aliment, mouvement=mouvement)

                    if not mouvement_aliment_sortie_existant:

                        # Créer le mouvement d'incorporation du bénéficiaire
                        mouvement_aliment = MouvementAliment.objects.create(created_by=created_by,
                                                                            aliment=aliment,
                                                                            mouvement=mouvement,
                                                                            police=police,
                                                                            date_effet=date_mouvement,
                                                                            motif=motif)
                        mouvement_aliment.save()

                        pprint("- mouvement de sortie crée - aliment_id: " + str(aliment.id))
                        cpt = cpt + 1


        response = {
            'statut': 1,
            'message': str(cpt) + " MOUVEMENTS DE SORTIE CREES AVEC SUCCÈS"
        }

    else:
        response = {
            'statut': 0,
            'message': "AUCUN ALIMENT SANS MOUVEMENT DE SORTIE TROUVE"
        }

    return JsonResponse(response)


#Recalculer les parts des
def recalculer_parts_sinistres_sucaf(request):

    sinistres = Sinistre.objects.filter(police_id=207, type_prefinancement_id=3, dossier_sinistre_id__isnull=False)

    #dd(sinistres)

    cpt = 0
    if sinistres:
        for sinistre in sinistres:

            taux_couverture = sinistre.formulegarantie.taux_couverture
            taux_tm = sinistre.formulegarantie.taux_tm
            part_compagnie = Decimal(taux_couverture/100) * sinistre.frais_reel - sinistre.depassement
            part_assure = Decimal(taux_tm/100) * sinistre.frais_reel + sinistre.depassement

            #
            sinistre.part_compagnie = part_compagnie
            sinistre.part_assure = part_assure
            sinistre.type_prefinancement_id = 1
            sinistre.observation = 'Parts recalculées, et sinistre marqué comme préfinancé'
            sinistre.save()

            cpt = cpt + 1


        response = {
            'statut': 1,
            'message': str(cpt) + " SINISTRES ONT ETE RECALCULES AVEC SUCCÈS"
        }

    else:
        response = {
            'statut': 0,
            'message': "AUCUN SINISTRE TROUVE"
        }

    return JsonResponse(response)



def corriger_param_produit_compagnie(request):
    dd("DEJA EXECUTE, PAS BESOIN DE REEXECUTER, CONSERVER POUR UN USAGE ULTERIEUR")

    '''
    #marquer les compagnies qui ont des params taux différents
    compagnies = Compagnie.objects.all().order_by('id')
    cpt = 0
    for compagnie in compagnies:

        pprint(f'------ TRAITEMENT DES PARAMS DE LA COMPAGNIE: {compagnie.nom} (id={compagnie.id}) ------')
        params = ParamProduitCompagnie.objects.filter(compagnie=compagnie).order_by('id')

        pprint('params.count()')
        pprint(f'{params.count()} lignes de paramètres trouvés')

        if params:
            taux_com_gestion_ref = params.first().taux_com_gestion
            taux_com_courtage_ref = params.first().taux_com_courtage
            taux_com_courtage_terme_ref = params.first().taux_com_courtage_terme

            pprint("taux_com_gestion_ref")
            pprint(taux_com_gestion_ref)
            pprint("taux_com_courtage_ref")
            pprint(taux_com_courtage_ref)
            pprint("taux_com_courtage_terme_ref")
            pprint(taux_com_courtage_terme_ref)

            print("COMPARAISON DES LIGNES")
            cpt = 0
            for param in params:
                cpt = cpt + 1
                taux_com_gestion = param.taux_com_gestion
                taux_com_courtage = param.taux_com_courtage
                taux_com_courtage_terme = param.taux_com_courtage_terme

                print(f"LIGNE DE PARAM PRODUIT _N° {cpt}")
                pprint("taux_com_gestion")
                pprint(taux_com_gestion)
                pprint("taux_com_courtage")
                pprint(taux_com_courtage)
                pprint("taux_com_courtage_terme")
                pprint(taux_com_courtage_terme)

                if taux_com_gestion != taux_com_gestion_ref or taux_com_courtage != taux_com_courtage_ref or taux_com_courtage_terme != taux_com_courtage_terme_ref:
                    compagnie.has_taux_multiple = True
                    compagnie.save()
                    continue

                    pprint("DIFFERENTS DE LA PREMIERE LIGNE")

                else:
                    #dd(compagnie)
                    pprint ("IDENTIQUE A LA PREMIERE LIGNE")

            pprint(f'{params.count()} LIGNE(S) DE PARAM PRODUIT TROUVEE(S)')

        else:
            pprint("AUCUNE LIGNE DE PARAM PRODUIT TROUVEE")


        pprint(f'------ FIN DE TRAITEMENT POUR LA COMPAGNIE: {compagnie.nom} (id={compagnie.id}) ------')



    #désactiver les lignes doublons des lignes params qui n'ont pas changé
    compagnies = Compagnie.objects.filter(has_taux_multiple=False).order_by('id')
    for compagnie in compagnies:

        pprint(f'------ DESCATIVATION DES DOUBLONS DE PARAMS PRODUIT DE LA COMPAGNIE: {compagnie.nom} (id={compagnie.id}) ------')
        params = ParamProduitCompagnie.objects.filter(compagnie=compagnie).order_by('id')

        pprint('params.count()')
        pprint(f'{params.count()} lignes de paramètres trouvés')

        #si plusieurs lignes, désactivers garder la première et désactiver les autres
        if params.count() > 1:
            cpt_ligne_param = 0
            for param in params:
                cpt_ligne_param = cpt_ligne_param + 1
                if cpt_ligne_param > 1:

                    param.status = False
                    param.save()
                else:
                    pprint("LIGNE A CONSERVER")
                    pprint(param)

        pprint(f'------ FIN DESCATIVATION DES DOUBLONS DE PARAMS PRODUIT DE LA COMPAGNIE ------')

    response = {
        'statut': 1,
        'message': "OPERATION EFFECTUEE AVEC SUCCES"
    }

    return JsonResponse(response)
    
    '''


def update_matricule(request):
    #marquer les compagnies qui ont des params taux différents
    aliments_matricules = AlimentMatricule.objects.all().order_by('id')

    if aliments_matricules:

        for am in aliments_matricules:

            carte = Carte.objects.filter(numero=am.NUMERO_CARTE).order_by('id')

            aliment = carte.aliment
            aliment.matricule_employe = am.MATRICULE
            aliment.save()

        response = {
            'statut': 1,
            'message': "OPERATION EFFECTUEE AVEC SUCCES"
        }

    else:
        response = {
            'statut': 0,
            'message': "AUCUNE DONNEE"
        }


    return JsonResponse(response)



def correction_affections(request):

    affections = Affection.objects.all()

    #dd(affections)

    cpt = 0
    if affections:
        for affection in affections:

            cpt = cpt + 1


        response = {
            'statut': 1,
            'message': str(cpt) + " AFFECTIONS CORRIGEES AVEC SUCCÈS"
        }

    else:
        response = {
            'statut': 0,
            'message': "AUCUN AFFECTION TROUVE"
        }

    return JsonResponse(response)



def disponibilite_upd(request):
    message = {}
    message['heure'] = "LE SERVEUR PREND EN COMPTES LES CHANGEMENTS EFFECTUÉS DANS LE CODE - 19/12/2023 21:00"
    dd(message)


def redirecttohome(request):
    return redirect('/')


def clear_cache(request):
    cache.clear()

    return redirect('/')


@login_required
def custom_password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Additional logic after password change if needed
            user.password_type = PasswordType.CUSTOM
            user.save()
            pprint("Password has changed")

            return redirect('/')  # Redirect to password change success page
    else:
        form = PasswordChangeForm(user=request.user)

    pprint("custom_password_change")
    return render(request, 'registration/password_change_form.html', {'form': form})



def set_bureau(request):
    if request.method == 'POST':
        user = request.user

        if user.is_superuser or user.is_admin_group:
            bureau_id = request.POST.get('bureau_id')
            bureau = Bureau.objects.get(pk=bureau_id)
            user.bureau = bureau
            user.save()

    # Get the previous URL from the 'HTTP_REFERER' header, if available
    previous_page = request.META.get('HTTP_REFERER')

    # If the 'HTTP_REFERER' header is not available or points to the same URL,
    # redirect to a default URL (e.g., homepage).
    if not previous_page or previous_page == request.build_absolute_uri():
        return HttpResponseRedirect(reverse('home'))  # Replace 'home' with the name of your homepage URL pattern.
    else:
        return HttpResponseRedirect(previous_page)


class TarifsView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_prestataire"
    template_name = 'tarifs/tarifs.html'
    model = Prestataire

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        rubriques = Rubrique.objects.filter(status=True)
        regroupements_actes = RegroupementActe.objects.filter(status=True)
        #
        tarifs_exists = Tarif.objects.filter(bureau=self.request.user.bureau).exists()

        context_perso = {
            'rubriques': rubriques,
            'regroupements_actes': regroupements_actes,
            'tarifs_exists': tarifs_exists,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def tarifs_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_rubrique = request.GET.get('search_rubrique', '')
    search_code = request.GET.get('search_code', '')
    search_libelle = request.GET.get('search_libelle', '')
    search_value = request.GET.get('search[value]', '')

    user = request.user
    queryset = Tarif.objects.filter(bureau=user.bureau, statut=Statut.ACTIF, prestataire__isnull=True)

    if search_libelle:
        queryset = queryset.filter(
            Q(acte__libelle__icontains=search_libelle)
        )

    if search_code:
        queryset = queryset.filter(
            Q(acte__code__icontains=search_code)
        )

    if search_rubrique:
        queryset = queryset.filter(
            Q(acte__rubrique_id=search_rubrique)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'acte__code',
        1: 'acte__libelle',
        2: 'acte__lettre_cle',
        3: 'acte__rubrique__libelle',
        4: 'acte__regroupement_acte__libelle',
        5: 'cout_classique',
        6: 'cout_mutuelle',
        7: 'cout_public_hg',
        8: 'cout_public_chu',
        9: 'cout_public_ica',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('popup_detail_tarif', args=[p.id])  # URL to the detail view

        actions_html = (
            f'<span style="cursor:pointer;" class=" btn-popup_details_tarif badge btn-sm btn-details rounded-pill" data-href="{detail_url}"<span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span>')

        data.append({
            "id": p.id,
            "code_acte": p.acte.code,
            "libelle_acte": p.acte.libelle,
            "lettre_cle_acte": p.acte.lettre_cle,
            "rubrique": p.acte.rubrique.libelle if p.acte.rubrique else "",
            "regroupement": p.acte.regroupement_acte.libelle if p.acte.regroupement_acte else "",
            "cout_classique": f'<span style="text-align:right;">' + money_field(p.cout_classique) + '</span>',
            "cout_mutuelle": f'<span style="text-align:right;">' + money_field(p.cout_mutuelle) + '</span>',
            "cout_public_hg": f'<span style="text-align:right;">' + money_field(p.cout_public_hg) + '</span>',
            "cout_public_chu": f'<span style="text-align:right;">' + money_field(p.cout_public_chu) + '</span>',
            "cout_public_ica": f'<span style="text-align:right;">' + money_field(p.cout_public_ica) + '</span>',
            "statut": p.statut,
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def popup_detail_tarif(request, tarif_id):
    tarif = Tarif.objects.get(id=tarif_id)

    return render(request, 'tarifs/modal_details_tarif.html', {'tarif': tarif})


def add_tarif(request):
    '''
    if request.method == 'POST':
        name = request.POST.get('name')

        lettre_cle_public_hg = models.CharField(max_length=50, blank=True, null=True)
        coef_public_hg = models.CharField(max_length=50, blank=True, null=True)
        pu_public_hg = models.IntegerField(null=True)
        cout_public_hg = models.IntegerField(null=True)

        lettre_cle_public_chu = models.CharField(max_length=50, blank=True, null=True)
        coef_public_chu = models.CharField(max_length=50, blank=True, null=True)
        pu_public_chu = models.IntegerField(null=True)
        cout_public_chu = models.IntegerField(null=True)

        lettre_cle_public_ica = models.CharField(max_length=50, blank=True, null=True)
        coef_public_ica = models.CharField(max_length=50, blank=True, null=True)
        pu_public_ica = models.IntegerField(null=True)
        cout_public_ica = models.IntegerField(null=True)

        lettre_cle_mutuelle = models.CharField(max_length=50, blank=True, null=True)
        coef_mutuelle = models.CharField(max_length=50, blank=True, null=True)
        pu_mutuelle = models.IntegerField(null=True)
        cout_mutuelle = models.IntegerField(null=True)

        lettre_cle_classique = models.CharField(max_length=50, blank=True, null=True)
        coef_classique = models.CharField(max_length=50, blank=True, null=True)
        pu_classique = models.IntegerField(null=True)
        cout_classique = models.IntegerField(null=True)

        tarif = Tarif.objects.create(
            bureau=request.user.bureau,
            acte=type_prestataire_id,
            name=name,
        )

        tarif.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
            }
        }


    else:


    '''
    response = {
        'statut': 0,
        'message': "Methode non autorisée !",
        'data': {}
    }
    return JsonResponse(response)


def tarifs_prestataire_datatable(request, prestataire_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_rubrique = request.GET.get('search_rubrique', '')
    search_code = request.GET.get('search_code', '')
    search_libelle = request.GET.get('search_libelle', '')
    search_value = request.GET.get('search[value]', '')
    search = request.GET.get('search[value]', '')

    user = request.user
    queryset = Tarif.objects.filter(bureau=user.bureau, prestataire_id=prestataire_id, statut=Statut.ACTIF)

    if search_libelle:
        queryset = queryset.filter(
            Q(acte__libelle__icontains=search_libelle)
        )

    if search_code:
        queryset = queryset.filter(
            Q(acte__code__icontains=search_code)
        )

    if search_rubrique:
        queryset = queryset.filter(
            Q(acte__rubrique_id=search_rubrique)
        )

    if search:
        queryset = queryset.filter(
            Q(acte__libelle__icontains=search) |
            Q(acte__code__icontains=search) |
            Q(acte__rubrique__libelle__icontains=search) |
            Q(acte__regroupement_acte__libelle__icontains=search)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'acte__code',
        1: 'acte__libelle',
        2: 'acte__rubrique__libelle',
        3: 'acte__regroupement_acte__libelle',
        4: 'cout_classique',
        5: 'cout_mutuelle',
        6: 'cout_public_hg',
        7: 'cout_public_chu',
        8: 'cout_public_ica',
        9: 'coef_prestataire',
        10: 'pu_prestataire',
        11: 'cout_prestataire',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('popup_detail_tarif', args=[p.id])  # URL to the detail view

        actions_html = (
            f'<span style="cursor:pointer;" class=" btn-popup_details_tarif badge btn-sm btn-details rounded-pill" data-href="{detail_url}"<span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span>')

        data.append({
            "id": p.id,
            "code_acte": p.acte.code,
            "libelle_acte": p.acte.libelle,
            "lettre_cle_acte": p.acte.lettre_cle,
            "rubrique": p.acte.rubrique.libelle if p.acte.rubrique else "",
            "regroupement": p.acte.regroupement_acte.libelle if p.acte.regroupement_acte else "",
            "cout_classique": f'<span style="text-align:right;">' + money_field(p.cout_classique) + '</span>',
            "cout_mutuelle": f'<span style="text-align:right;">' + money_field(p.cout_mutuelle) + '</span>',
            "cout_public_hg": f'<span style="text-align:right;">' + money_field(p.cout_public_hg) + '</span>',
            "cout_public_chu": f'<span style="text-align:right;">' + money_field(p.cout_public_chu) + '</span>',
            "cout_public_ica": f'<span style="text-align:right;">' + money_field(p.cout_public_ica) + '</span>',
            "coef_prestataire": f'<span style="text-align:right;">' + money_field(p.coef_prestataire) + '</span>',
            "pu_prestataire": f'<span style="text-align:right;">' + money_field(p.pu_prestataire) + '</span>',
            "cout_prestataire": f'<span style="text-align:right;">' + money_field(p.cout_prestataire) + '</span>',
            "statut": p.statut,
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# Génère un fichier model avec les actes déjà en base pour que le gestionnaire puisse renseigner les coûts des actes
def generate_modele_tarifs_excel(request, prestataire_id):
    prestataire = Prestataire.objects.get(id=prestataire_id)

    # Données à inclure dans le DataFrame
    actes = Acte.objects.filter(type_acte__code="acte", status=True).order_by('rubrique_id')

    # Créer un Workbook et accéder à la première feuille
    wb = Workbook()
    ws = wb.active

    # Données à inclure dans le DataFrame
    data = {
        'CODE_RUBRIQUE': [],
        'LIBELLE_ACTE': [],
        'CODE_ACTE': [],
        'LETTRE_CLE': [],
        # 'LETTRE_CLE_CLASSIQUE': [],
        'COEF_CLASSIQUE': [],
        'PRIX_UNIT_CLASSIQUE': [],
        'TARIF_CLASSIQUE': [],
        # 'LETTRE_CLE_MUTUELLE': [],
        'COEF_MUTUELLE': [],
        'PRIX_UNIT_MUTUELLE': [],
        'TARIF_MUTUELLE': [],
        # 'LETTRE_CLE_HG': [],
        'COEF_HG': [],
        'PRIX_UNIT_HG': [],
        'TARIF_HG': [],
        # 'LETTRE_CLE_CHU': [],
        'COEF_CHU': [],
        'PRIX_UNIT_CHU': [],
        'TARIF_CHU': [],
        # 'LETTRE_CLE_ICA': [],
        'COEF_ICA': [],
        'PRIX_UNIT_ICA': [],
        'TARIF_ICA': [],
        # 'LETTRE_CLE_PRESTATAIRE': [],
        'COEF_PRESTATAIRE': [],
        'PRIX_UNIT_PRESTATAIRE': [],
        'TARIF_PRESTATAIRE': [],
    }

    # Ajouter les actes au DataFrame
    for acte in actes:
        # renseigner avec le tarif existant de ce prestataire --
        tarif_existant = Tarif.objects.filter(acte__code=acte.code, bureau=request.user.bureau,
                                              statut=Statut.ACTIF).first()
        tarif_existant_prestataire = Tarif.objects.filter(acte__code=acte.code, bureau=request.user.bureau,
                                                          prestataire=prestataire, statut=Statut.ACTIF).first()

        data['CODE_RUBRIQUE'].append(acte.rubrique.libelle)
        data['LIBELLE_ACTE'].append(acte.libelle)
        data['CODE_ACTE'].append(acte.code)
        data['LETTRE_CLE'].append(acte.lettre_cle)

        # data['LETTRE_CLE_CLASSIQUE'].append(tarif_existant.lettre_cle_classique)
        data['COEF_CLASSIQUE'].append(tarif_existant.coef_classique if tarif_existant else 1)
        data['PRIX_UNIT_CLASSIQUE'].append(tarif_existant.pu_classique if tarif_existant else 0)
        data['TARIF_CLASSIQUE'].append(tarif_existant.cout_classique if tarif_existant else 0)

        # data['LETTRE_CLE_MUTUELLE'].append(tarif_existant.lettre_cle_classique)
        data['COEF_MUTUELLE'].append(tarif_existant.coef_mutuelle if tarif_existant else 1)
        data['PRIX_UNIT_MUTUELLE'].append(tarif_existant.pu_mutuelle if tarif_existant else 0)
        data['TARIF_MUTUELLE'].append(tarif_existant.cout_mutuelle if tarif_existant else 0)

        # data['LETTRE_CLE_HG'].append(tarif_existant.lettre_cle_public_hg if tarif_existant else 0)
        data['COEF_HG'].append(tarif_existant.coef_public_hg if tarif_existant else 0)
        data['PRIX_UNIT_HG'].append(tarif_existant.pu_public_hg if tarif_existant else 0)
        data['TARIF_HG'].append(tarif_existant.cout_public_hg if tarif_existant else 0)

        # data['LETTRE_CLE_CHU'].append(tarif_existant.lettre_cle_public_chu if tarif_existant else 0)
        data['COEF_CHU'].append(tarif_existant.coef_public_chu if tarif_existant else 0)
        data['PRIX_UNIT_CHU'].append(tarif_existant.pu_public_chu if tarif_existant else 0)
        data['TARIF_CHU'].append(tarif_existant.cout_public_chu if tarif_existant else 0)

        # data['LETTRE_CLE_ICA'].append(tarif_existant.lettre_cle_public_ica if tarif_existant else 0)
        data['COEF_ICA'].append(tarif_existant.coef_public_ica if tarif_existant else 0)
        data['PRIX_UNIT_ICA'].append(tarif_existant.pu_public_ica if tarif_existant else 0)
        data['TARIF_ICA'].append(tarif_existant.cout_public_ica if tarif_existant else 0)

        # data['LETTRE_CLE_PRESTATAIRE'].append('')
        data['COEF_PRESTATAIRE'].append(
            tarif_existant_prestataire.coef_prestataire if tarif_existant_prestataire else 1)
        data['PRIX_UNIT_PRESTATAIRE'].append(
            tarif_existant_prestataire.pu_prestataire if tarif_existant_prestataire else 0)
        data['TARIF_PRESTATAIRE'].append(
            tarif_existant_prestataire.cout_prestataire if tarif_existant_prestataire else 0)

    # Créer un DataFrame avec Pandas
    df = pd.DataFrame(data)

    filename = 'INOV_V1-TARIF_DU_PRESTATAIRE_' + slugify(str(prestataire.name)).upper() + '.xlsx'

    # Créer une réponse HTTP avec le type MIME approprié
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename

    # Enregistrer le DataFrame dans le fichier Excel
    df.to_excel(response, index=False, engine='openpyxl')

    return response


# une que le gestionnaire à renseigné les coûts des actes, on l'importe
def import_tarif_pestataire(request, prestataire_id):
    # try:

    # Charger le fichier Excel depuis la requête
    file = request.FILES['fichier_import_tarif']

    # Lire le fichier Excel avec pandas
    df = pd.read_excel(file)

    rows_count = 0
    # Parcourir les lignes du DataFrame
    for index, row in df.iterrows():

        # Récupérer le code de l'acte
        code_acte = row['CODE_ACTE']

        # Récupérer l'objet Acte correspondant au code
        acte = Acte.objects.filter(code=code_acte).first()

        if acte:
            pprint("L'acte existe, donc on insère le tarif")

            # désactiver l'ancien tarif si existant
            Tarif.objects.filter(acte=acte, prestataire_id=prestataire_id).update(statut=Statut.INACTIF,
                                                                                  statut_validite=StatutValidite.CLOTURE)

            # Insérer une ligne dans la table Tarif avec les coefficients, prix unitaire, etc.
            tarif = Tarif.objects.create(
                acte=acte,
                code_acte=code_acte,
                coef_prestataire=row['COEF_PRESTATAIRE'],
                pu_prestataire=row['PRIX_UNIT_PRESTATAIRE'],
                cout_prestataire=row['TARIF_PRESTATAIRE'],
                statut=Statut.ACTIF,
                statut_validite=StatutValidite.VALIDE,
                bureau=request.user.bureau,
                prestataire_id=prestataire_id,
                created_by=request.user
            )

    # Indiquer que le prestataire a son propre tarif
    Prestataire.objects.filter(id=prestataire_id).update(has_tarif_prestataire=True)

    response = {
        'statut': 1,
        'message': "Tarifs importés avec succès !",
        'data': {}
    }

    '''except Exception as e:
        response = {
            'statut': 0,
            'message': f"Erreur lors de l'importation des tarifs : {str(e)}",
            'data': {}
        }
    '''

    return JsonResponse(response)


class PrestatairesView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_prestataire"
    template_name = 'prestataires/prestataires.html'
    model = Prestataire

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        secteurs = Secteur.objects.all()
        bureaux = Bureau.objects.filter(id=request.user.bureau.pk)
        types_prestataires = TypePrestataire.objects.all()
        types_etablissements = TypeEtablissement.objects.all()
        reseaux_soins = ReseauSoin.objects.filter(bureau=request.user.bureau, status=True)

        context_perso = {
            'bureaux': bureaux,
            'secteurs': secteurs,
            'types_prestataires': types_prestataires,
            'types_etablissements': types_etablissements,
            'reseaux_soins': reseaux_soins,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def prestataires_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_nom = request.GET.get('search_nom', '')
    search_code = request.GET.get('search_code', '')
    search_type = request.GET.get('search_type', '')
    search_value = request.GET.get('search[value]', '')

    user = request.user
    if request.user.is_superuser:
        queryset = Prestataire.objects.filter(bureau_id=user.bureau_id)
    else:
        queryset = Prestataire.objects.filter(status=True, bureau_id=user.bureau_id)

    if search_nom:
        queryset = queryset.filter(
            Q(name__icontains=search_nom)
        )

    if search_code:
        queryset = queryset.filter(
            Q(code__icontains=search_code)
        )

    if search_type:
        queryset = queryset.filter(
            Q(type_prestataire_id=search_type)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'name',
        1: 'code',
        2: 'type_prestataire__name',
        3: 'secteur__libelle',
        4: 'telephone',
        5: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('detail_prestataire', args=[p.id])  # URL to the detail view
        update_url = reverse('popup_modifier_prestataire', args=[p.id])  # URL to the detail view

        actions_html = (
            f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> {_("Détails")}</span></a>&nbsp;'
            f'<span style="cursor:pointer;" class="btn_modifier_prestataire badge btn-sm btn-modifier rounded-pill" data-href="{update_url}"><i class="fa fa-edit"></i> {_("Modifier")}</span>')

        data.append({
            "id": p.id,
            "name": p.name,
            "code": p.code,
            "type_prestataire": p.type_prestataire.name if p.type_prestataire else "",
            "secteur": p.secteur.libelle if p.secteur else "",
            "ville": p.ville,
            "telephone": p.telephone,
            "statut": "ACTIF" if p.status else " INACTIF",
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def export_prestaitaires(request):
    if request.method == 'POST':
        # Récupérer les paramètres de filtre

        search_nom = request.GET.get('search_nom', '')
        search_code = request.GET.get('search_code', '')
        search_type = request.GET.get('search_type', '')
        search_value = request.GET.get('search[value]', '')

        # Filtrer les données selon les paramètres
        prestataires = Prestataire.objects.filter(bureau=request.user.bureau)

        if search_nom:
            prestataires = prestataires.filter(name__icontains=search_nom)
        if search_code:
            prestataires = prestataires.filter(code__icontains=search_code)
        if search_type:
            prestataires = prestataires.filter(type_prestataire_id=search_type)

        # Préparation de l'exportation Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Liste_Prestataire.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'LISTE PRESTATAIRE'

        # Définir les en-têtes

        # Personnalisation de l'en-tête
        bold_font = Font(bold=True, color="FF0000")  # Rouge et en gras
        header_fill = PatternFill(start_color='FFEBCD', end_color='FFEBCD',
                                  fill_type='solid')  # Couleur de remplissage pour les entêtes
        headers = ['Id Prestataire', 'Nom prestataire', 'Type de prestataire', 'Ville', 'Contact', 'Date de création',
                   'E-mail']
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = bold_font
            cell.fill = header_fill

        # Remplissage du tableau avec les données
        for row_num, prestataire in enumerate(prestataires, 2):
            row = [
                prestataire.code,
                prestataire.name,
                prestataire.type_prestataire.name if prestataire.type_prestataire else '',
                prestataire.ville,
                prestataire.telephone,
                prestataire.created_at.strftime("%d/%m/%Y %H:%M"),
                prestataire.email,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Ajuster la largeur des colonnes en fonction du contenu pour faciliter la lecture du tableau
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(response)
        return response
    else:
        return HttpResponse(status=405)

def add_prestataire(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        rb_ordre = request.POST.get('rb_ordre')
        code = request.POST.get('code')
        telephone = request.POST.get('telephone')
        fax = request.POST.get('fax')
        email = request.POST.get('email')
        fax = request.POST.get('fax')
        ville = request.POST.get('ville')
        addresse = request.POST.get('addresse')
        secteur_id = request.POST.get('secteur_id')
        type_prestataire_id = request.POST.get('type_prestataire_id')
        reseaux_soins_ids = request.POST.getlist('reseaux_soins_ids')
        reseaux_soins = ReseauSoin.objects.filter(id__in=reseaux_soins_ids)

        latitude = None
        longitude = None
        try:
            latitude = float(request.POST.get('latitude').replace(",", ".").replace(" ", ""))
            longitude = float(request.POST.get('longitude').replace(",", ".").replace(" ", ""))
        except Exception as e:
            print(e)

        pprint(reseaux_soins)

        prestataire = Prestataire.objects.filter(bureau=request.user.bureau).latest('id')

        code = "P" + str(prestataire.pk + 1)

        if request.user.bureau:
            bureau = request.user.bureau

            fs = FileSystemStorage()

            '''
            logo = request.FILES['logo']
            if logo:
                fichier = request.FILES['logo']
                file_name_renamed = fichier.name.replace(" ", "_")
                logo_filename = fs.save(file_name_renamed, fichier)
    
            '''

            prestataire = Prestataire.objects.create(
                name=name,
                rb_ordre=rb_ordre,
                code=code,
                telephone=telephone,
                fax=fax,
                email=email,
                ville=ville,
                addresse=addresse,
                # logo=logo_filename,
                status=True,
                bureau_id=bureau.pk,
                type_prestataire_id=type_prestataire_id,
                secteur_id=secteur_id,
                latitude=latitude,
                longitude=longitude,
            )

            # Mettre a jour le code
            code_bureau = request.user.bureau.code
            prestataire.code = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(prestataire.pk).zfill(
                7) + '-P'
            # prestataire.code = 'P' + str(prestataire.pk).zfill(6)
            prestataire.save()

            # enregistrer ses réseaux de soins
            for reseau_soin in reseaux_soins:
                prs = PrestataireReseauSoin(reseau_soin=reseau_soin, prestataire=prestataire, created_by=request.user,
                                            date_integration=datetime.datetime.now(tz=timezone.utc))
                prs.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {
                }
            }


        else:
            response = {
                'statut': 0,
                'message': "Vous n'êtes lié à aucun bureau !",
                'data': {}
            }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def popup_modifier_prestataire(request, prestataire_id):
    prestataire = Prestataire.objects.get(id=prestataire_id)

    secteurs = Secteur.objects.all()
    types_prestataires = TypePrestataire.objects.all()
    types_etablissements = TypeEtablissement.objects.all()

    return render(request, 'prestataires/modal_modifier_prestataire.html', {
        'prestataire': prestataire,
        'secteurs': secteurs,
        'types_prestataires': types_prestataires,
        'types_etablissements': types_etablissements
    })


# TO_DO_ISMAEL
def update_prestataire(request, prestataire_id):
    if request.method == 'POST':

        prestataire = Prestataire.objects.filter(id=prestataire_id).first()

        type_prestataire_id = request.POST.get('type_prestataire_id')
        secteur_id = request.POST.get('secteur_id')
        type_etablissement_id = request.POST.get('type_etablissement_id')

        prestataire.name = request.POST.get('name')
        prestataire.rb_ordre = request.POST.get('rb_ordre')
        prestataire.telephone = request.POST.get('telephone')
        prestataire.email = request.POST.get('email')
        prestataire.fax = request.POST.get('fax')
        prestataire.addresse = request.POST.get('addresse')
        prestataire.ville = request.POST.get('ville')

        prestataire.secteur_id = secteur_id
        prestataire.type_prestataire_id = type_prestataire_id
        prestataire.type_etablissement_id = type_etablissement_id

        try:
            prestataire.latitude = float(request.POST.get('latitude').replace(",", ".").replace(" ", ""))
            prestataire.longitude = float(request.POST.get('longitude').replace(",", ".").replace(" ", ""))
        except Exception as e:
            print(e)
            prestataire.latitude = None
            prestataire.longitude = None

        if request.user.bureau:
            prestataire.bureau = request.user.bureau

            fs = FileSystemStorage()

            '''
            logo = request.FILES['logo']
            fichier_tarification = request.FILES['fichier_tarification']
    
            if logo:
                fichier = request.FILES['logo']
                file_name_renamed = fichier.name.replace(" ", "_")
                logo_filename = fs.save(file_name_renamed, fichier)
    
            if fichier_tarification:
                fichier = request.FILES['fichier_tarification']
                file_name_renamed = fichier.name.replace(" ", "_")
                fichier_tarification_filename = fs.save(file_name_renamed, fichier)
            '''

            prestataire.save()
            response = {
                'statut': 1,
                'message': "Modification du prestataire effectuée avec succès !",
                'data': {
                }
            }


        else:

            response = {
                'statut': 0,
                'message': "Vous n'êtes lié à aucun bureau !",
                'data': {}
            }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def add_reseau_soin_prestataire(request, prestataire_id):
    if request.method == 'POST':
        reseaux_soins_ids = request.POST.getlist('reseaux_soins_ids')
        reseaux_soins = ReseauSoin.objects.filter(id__in=reseaux_soins_ids)

        # enregistrer ses réseaux de soins
        for reseau_soin in reseaux_soins:
            prs = PrestataireReseauSoin(reseau_soin=reseau_soin, prestataire_id=prestataire_id, created_by=request.user,
                                        date_integration=datetime.datetime.now(tz=timezone.utc))
            prs.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def retirer_reseau_soin_prestataire(request, prs_id):
    if request.method == 'POST':
        # enregistrer ses réseaux de soins
        prs = PrestataireReseauSoin.objects.get(id=prs_id)
        prs.statut_validite = StatutValidite.CLOTURE
        prs.date_retrait = datetime.datetime.now(tz=timezone.utc)
        prs.removed_by = request.user
        prs.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def add_prescripteur(request):
    if request.method == 'POST':

        pprint(request.POST)

        nom = request.POST.get('nom')
        prenoms = request.POST.get('prenoms')
        numero_ordre = request.POST.get('numero_ordre')
        email = request.POST.get('email')
        telephone = request.POST.get('telephone')
        prestataire_id = request.POST.get('prestataire_id')

        specialite_id = request.POST.get('specialite_id')
        specialite = Specialite.objects.get(id=specialite_id)

        prestataire = Prestataire.objects.get(id=prestataire_id)

        # dd(prestataire)
        prescripteur = Prescripteur.objects.create(
            nom=nom,
            prenoms=prenoms,
            telephone=telephone,
            numero_ordre=numero_ordre,
            email=email,
            specialite_id=specialite.pk,
        )

        PrescripteurPrestataire.objects.create(
            prescripteur_id=prescripteur.pk,
            prestataire_id=prestataire.pk
        )

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': prescripteur.pk,
                'nom': prescripteur.nom,
                'prenoms': prescripteur.prenoms,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)

#@transaction.atomic
def import_prescripteurs(request, prestataire_id):
    if request.method == 'POST':
        # get file and use excel to import in prescripteur table

        fichier = request.FILES['fichier_import_prescripteurs']

        fs = FileSystemStorage(location='prestataires/upload_prescripteurs/')
        file_name_renamed = fichier.name.replace(" ", "_")

        filename = fs.save(file_name_renamed, fichier)
        uploaded_file_url = fs.url(filename)
        uploaded_file_full_path = fs.path(filename)
        prestataire = Prestataire.objects.get(id=prestataire_id)

        cpt_all = 0
        cpt_success = 0

        # df = pd.read_excel("." + uploaded_file_url)
        df = pd.read_excel(uploaded_file_full_path)
        for index, row in df.iterrows():
            cpt_all = cpt_all  + 1
            # Valeur du tableau
            # Access row values using column names
            #try:
            numero_ordre = row['NUMERO_ORDRE_MEDECIN']
            nom = row['NOM']
            prenoms = row['PRENOMS']
            telephone = row['TELEPHONE']
            code_specialite = row['CODE_SPECIALITE']

            specialite = Specialite.objects.filter(code=code_specialite).first()

            if specialite:
                if Prescripteur.objects.filter(numero_ordre=numero_ordre, bureau=request.user.bureau).exists():
                    # Ne retournera pas d'erreur si le prescripteur existe
                    prescripteur = Prescripteur.objects.filter(numero_ordre=numero_ordre, bureau=request.user.bureau).first()
                    # dd(prescripteur)
                else:
                    # Le prescripteur n'existe pas, on le créé
                    prescripteur = Prescripteur.objects.create(
                        nom=nom,
                        prenoms=prenoms,
                        numero_ordre=numero_ordre,
                        telephone=telephone,
                        specialite_id=specialite.pk,
                        bureau=request.user.bureau,
                    )
                    # dd(prescripteur)

                # On tente de trouver l'enregistrement de du prescripteur sinon on l'enregistre
                prescripteur_prestataire = PrescripteurPrestataire.objects.filter(prescripteur_id=prescripteur.id,
                                                                                  prestataire_id=prestataire.pk).first()

                if not prescripteur_prestataire:
                    PrescripteurPrestataire.objects.create(
                        prescripteur_id=prescripteur.pk,
                        prestataire_id=prestataire.pk,
                        created_at=datetime.datetime.now(tz=timezone.utc)
                    )

                cpt_success = cpt_success + 1

            else:
                response = {
                    'statut': 0,
                    'message': "Veilleez entrez une specialité existant dans la base de données !",
                    'data': {
                    }
                }

            #except KeyError:
            #    response = {
            #        'statut': 0,
            #        'message': "Erreur sur le mot clé " + KeyError.args.index,
            #        'data': {
            #        }
            #    }
        if cpt_success == cpt_all:
            response = {
                'statut': 1,
                'message': "Importation effectuée avec succès",
                'data': {
                }
            }
        else:
            response = {
                'statut': 0,
                'message': "Echec de l'importation, veuillez renseigner correctement le fichier svp",
                'data': {
                }
            }

        return JsonResponse(response)


# get Prestataire detail
class DetailsPrestatairesView(TemplateView):
    permission_required = "production.view_prestataires"
    template_name = 'prestataires/prestataire_details.html'
    model = Prestataire

    def get(self, request, prestataire_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        prestataire = Prestataire.objects.filter(id=prestataire_id, bureau=request.user.bureau).first()
        if prestataire:

            clients = Client.objects.all()
            specialities = Specialite.objects.all()

            prescripteurs = PrescripteurPrestataire.objects.filter(prestataire_id=prestataire.pk,
                                                                   statut_validite=StatutValidite.VALIDE)
            utilisateurs = User.objects.filter(prestataire_id=prestataire.pk)

            tarifs_prestataire_clients = TarifPrestataireClient.objects.filter(prestataire_id=prestataire.pk)

            # reseaux_soins = ReseauSoin.objects.filter(status=True)
            reseaux_soins_prestataire = PrestataireReseauSoin.objects.filter(prestataire=prestataire,
                                                                             statut_validite=StatutValidite.VALIDE)

            prestataire_reseausoin_ids = PrestataireReseauSoin.objects.filter(prestataire=prestataire,
                                                                              statut_validite=StatutValidite.VALIDE).values_list(
                'reseau_soin_id', flat=True).order_by('-id')
            reseaux_soins_restants = ReseauSoin.objects.filter(bureau=request.user.bureau).exclude(id__in=prestataire_reseausoin_ids)

            rubriques = Rubrique.objects.filter(status=True)
            regroupements_actes = RegroupementActe.objects.filter(status=True)

            context_perso = {
                'clients': clients,
                'prestataire': prestataire,
                'utilisateurs': utilisateurs,
                'prescripteurs': prescripteurs,
                'specialities': specialities,
                'reseaux_soins_restants': reseaux_soins_restants,
                'reseaux_soins_prestataire': reseaux_soins_prestataire,
                'tarifs_prestataire_clients': tarifs_prestataire_clients,
                'rubriques': rubriques,
                'regroupements_actes': regroupements_actes,
            }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect('')

    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required
def prescripteurs_by_prestataire(request, prestataire_id):
    prestataire_prescripteur = PrescripteurPrestataire.objects.filter(prestataire_id=prestataire_id)
    prescripteurs = []
    for pp in prestataire_prescripteur:
        if pp.prescripteur not in prescripteurs:
            prescripteurs.append(pp.prescripteur)
    prescripteurs_serialize = serializers.serialize('json', prescripteurs)
    return HttpResponse(prescripteurs_serialize, content_type='application/json')

# PRESCRIPTEUR

def prescripteurs_prestataires_datatable(request, prestataire_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    #   search_nom = request.GET.get('search_nom', '')
    search_numero_ordre = request.GET.get('search_numero_ordre', '')
    search_specialite = request.GET.get('search_specialite', '')
    search_value = request.GET.get('search[value]', '')

    prestataire_prescripteurs_ids = PrescripteurPrestataire.objects.filter(
        prestataire=prestataire_id,
        statut_validite=StatutValidite.VALIDE
    ).values_list('prescripteur_id', flat=True).order_by('id')

    queryset = Prescripteur.objects.filter(id__in=prestataire_prescripteurs_ids)

    # if search_nom:
    #     queryset = queryset.filter(
    #         Q(nom__icontains=search_nom)
    #     )

    if search_numero_ordre:
        queryset = queryset.filter(
            Q(numero_ordre__icontains=search_numero_ordre)
        )

    if search_specialite:
        queryset = queryset.filter(
            Q(specialite_id=search_specialite)
        )

    sort_columns = {
        0: 'nom',
        1: 'prenoms',
        2: 'numero_ordre',
        3: 'telephone',
        4: 'specialite__name'
    }

    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    data = []
    for p in page_obj:
        modifier_prescripteur_url = reverse('popup_modifier_prescripteur', args=[p.id])  # url modifier prescripteur
        retirer_prescripteur_url = reverse('retirer_prescripteur_prestataire', args=[prestataire_id, p.id])  # url retirer prescripteur

        actions_html = (
            f'<span style="cursor:pointer;" class="btn_modifier_prescripteur badge btn-sm btn-modifier rounded-pill" data-href="{modifier_prescripteur_url}"><i class="fa fa-edit"></i> Modifier</span>&nbsp;'
            f'<span style="cursor:pointer;" class="btn_retirer_prescripteur badge btn-sm btn-danger rounded-pill" data-href="{retirer_prescripteur_url}"><i class="fa fa-minus"></i> Retirer</span>')

        data.append({
            "id": p.id,
            "nom": p.nom,
            "prenoms": p.prenoms,
            "numero_ordre": p.numero_ordre,
            "telephone": p.telephone,
            "specialite": p.specialite.name.upper() if p.specialite else "",
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def popup_modifier_prescripteur(request, prescripteur_id):
    prescripteur = Prescripteur.objects.filter(id=prescripteur_id).first()

    specialites = Specialite.objects.all()
    prestataires = Prestataire.objects.filter(status=True, bureau=request.user.bureau)

    return render(request, 'prestataires/modal_modifier_prescripteur.html', {
        'prescripteur':prescripteur,
        'specialites':specialites,
        'prestataires': prestataires
    })


def update_prescripteur(request, prescripteur_id):
    if request.method == 'POST':

        prescripteur = Prescripteur.objects.filter(id=prescripteur_id).first()

        #   email = request.POST.get('email')
        specialite_id = request.POST.get('specialite_id')
        
        
        prescripteur.nom = request.POST.get('nom')
        prescripteur.prenoms = request.POST.get('prenoms')
        prescripteur.numero_ordre = request.POST.get('numero_ordre')
        prescripteur.telephone = request.POST.get('telephone')
        prescripteur.specialite_id = specialite_id
        
        prescripteur.save()

        response = {
            'statut': 1,
            'message': "Modification effectuée ss avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def retirer_prescripteur_prestataire(request, prestataire_id, prescripteur_id):
    if request.method == 'POST':
        prescripteur_prestataire = PrescripteurPrestataire.objects.filter(prescripteur_id=prescripteur_id,
                                                                  prestataire_id=prestataire_id,
                                                                  statut_validite=StatutValidite.VALIDE).update(
            statut_validite=StatutValidite.CLOTURE,
            deleted_by=request.user,
            deleted_at=datetime.datetime.now(tz=timezone.utc),
        )

        response = {
            'statut': 1,
            'message': "Prescripteur retiré avec succès !",
            'data': {
            }
        }

    return JsonResponse(response)


# FIN PRESCRIPTEUR


# TODO RESEAU DE SOINS

class ReseauxSoinsView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_prestataire"
    template_name = 'reseaux_soins/reseaux_soins.html'
    model = Prestataire

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        secteurs = Secteur.objects.all()
        bureaux = Bureau.objects.filter(id=request.user.bureau.pk)
        types_prestataires = TypePrestataire.objects.all()
        types_etablissements = TypeEtablissement.objects.all()

        context_perso = {
            'bureaux': bureaux,
            'secteurs': secteurs,
            'types_prestataires': types_prestataires,
            'types_etablissements': types_etablissements,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def reseauxsoins_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_value = request.GET.get('search[value]', '')

    queryset = ReseauSoin.objects.filter(status=True, bureau=request.user.bureau)

    if search_value:
        queryset = queryset.filter(
            Q(nom__icontains=search_value)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'nom',
        1: 'code',
        5: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('detail_reseau_soin', args=[p.id])  # URL to the detail view
        update_url = reverse('popup_modifier_reseau_soin', args=[p.id])  # URL to the detail view

        actions_html = (
            f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;'
            f'<span style="cursor:pointer;" class="btn_modifier_reseau_soin badge btn-sm btn-modifier rounded-pill" data-href="{update_url}"><i class="fa fa-edit"></i> Modifier</span>')

        data.append({
            "id": p.id,
            "name": p.nom,
            "code": p.code,
            "statut": "ACTIF" if p.status else " INACTIF",
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def add_reseau_soin(request):
    if request.method == 'POST':
        name = request.POST.get('name')

        if request.user.bureau:

            cpt_reseau_soin_bureau = ReseauSoin.objects.filter(bureau=request.user.bureau).count() + 1
            # Mettre a jour le code
            code_bureau = request.user.bureau.code
            code = str(code_bureau) + 'RXS' + '-' + str(cpt_reseau_soin_bureau).zfill(4)

            reseau_soin = ReseauSoin.objects.create(
                created_by=request.user,
                bureau=request.user.bureau,
                nom=name,
                code=code
            )

            reseau_soin.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {
                }
            }


        else:
            response = {
                'statut': 0,
                'message': "Vous n'êtes lié à aucun bureau !",
                'data': {}
            }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def popup_modifier_reseau_soin(request, reseau_soin_id):
    reseau_soin = ReseauSoin.objects.get(id=reseau_soin_id)

    prestataires = Prestataire.objects.filter(bureau=request.user.bureau)

    return render(request, 'reseaux_soins/modal_modifier_reseau_soin.html', {
        'reseau_soin': reseau_soin,
        'prestataires': prestataires
    })


# TO_DO_ISMAEL
def update_reseau_soin(request, reseau_soin_id):
    if request.method == 'POST':

        reseau_soin = ReseauSoin.objects.filter(id=reseau_soin_id).update(nom=request.POST.get('name'))

        response = {
            'statut': 1,
            'message': "Modification effectuée ss avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


class DetailsReseauSoinView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_prestataire"
    template_name = 'reseaux_soins/reseau_soin_details.html'
    model = ReseauSoin

    def get(self, request, reseau_soin_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        reseau_soin = ReseauSoin.objects.get(id=reseau_soin_id, status=True)
        prestataires = PrestataireReseauSoin.objects.filter(reseau_soin_id=reseau_soin_id)
        types_prestataires = TypePrestataire.objects.all()

        context_perso = {
            'reseau_soin': reseau_soin,
            'prestataires': prestataires,
            'types_prestataires': types_prestataires,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def reseau_soin_prestataires_datatable(request, reseau_soin_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_nom = request.GET.get('search_nom', '')
    search_code = request.GET.get('search_code', '')
    search_type = request.GET.get('search_type', '')
    search_value = request.GET.get('search[value]', '')

    prestataire_reseausoin_ids = PrestataireReseauSoin.objects.filter(reseau_soin_id=reseau_soin_id,
                                                                      statut_validite=StatutValidite.VALIDE).values_list(
        'prestataire_id', flat=True).order_by('-id')
    queryset = Prestataire.objects.filter(status=True, bureau_id=request.user.bureau_id).filter(
        id__in=prestataire_reseausoin_ids)

    if search_nom:
        queryset = queryset.filter(
            Q(name__icontains=search_nom)
        )

    if search_code:
        queryset = queryset.filter(
            Q(code__icontains=search_code)
        )

    if search_type:
        queryset = queryset.filter(
            Q(type_prestataire_id=search_type)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'name',
        1: 'code',
        2: 'type_prestataire__name',
        3: 'secteur__libelle',
        4: 'telephone',
        5: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('detail_prestataire', args=[p.id])  # URL to the detail view
        retirer_prestataire_url = reverse('retirer_prestataire_reseau',
                                          args=[reseau_soin_id, p.id])  # URL to the detail view

        actions_html = (
            f'<!--a href="{detail_url}" target="_blank"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;-->'
            f'<span style="cursor:pointer;" class="btn_retirer_prestataire badge btn-sm btn-danger rounded-pill" data-href="{retirer_prestataire_url}"><i class="fa fa-minus"></i> Retirer</span>')

        data.append({
            "id": p.id,
            "name": p.name,
            "code": p.code,
            "type_prestataire": p.type_prestataire.name if p.type_prestataire else "",
            "secteur": p.secteur.libelle if p.secteur else "",
            "ville": p.ville,
            "telephone": p.telephone,
            "statut": "ACTIF" if p.status else " INACTIF",
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# POPUP JOINDRE DES PRESTATAIRES
def popup_joindre_prestataires(request, reseau_soin_id):
    reseau_soin = ReseauSoin.objects.filter(id=reseau_soin_id, bureau=request.user.bureau).first()
    types_prestataires = TypePrestataire.objects.all().order_by('name')

    return render(request, 'reseaux_soins/popup_joindre_prestataires.html',
                  {'reseau_soin': reseau_soin, 'types_prestataires': types_prestataires})


def reseau_soin_prestataires_restants_datatable(request, reseau_soin_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_nom = request.GET.get('search_nom', '')
    search_code = request.GET.get('search_code', '')
    search_type = request.GET.get('search_type', '')
    search_value = request.GET.get('search[value]', '')

    prestataire_reseausoin_ids = PrestataireReseauSoin.objects.filter(reseau_soin_id=reseau_soin_id,
                                                                      statut_validite=StatutValidite.VALIDE).values_list(
        'prestataire_id', flat=True).order_by('-id')
    queryset = Prestataire.objects.filter(status=True, bureau_id=request.user.bureau_id).exclude(
        id__in=prestataire_reseausoin_ids)

    if search_nom:
        queryset = queryset.filter(
            Q(name__icontains=search_nom)
        )

    if search_code:
        queryset = queryset.filter(
            Q(code__icontains=search_code)
        )

    if search_type:
        queryset = queryset.filter(
            Q(type_prestataire_id=search_type)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'name',
        1: 'code',
        2: 'type_prestataire__name',
        3: 'secteur__libelle',
        4: 'telephone',
        5: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for p in page_obj:
        detail_url = reverse('detail_prestataire', args=[p.id])  # URL to the detail view
        integrer_prestataire_url = reverse('joindre_prestataire_reseau',
                                           args=[reseau_soin_id, p.id, ])  # URL to the detail view

        actions_html = (
            f'<span style="cursor:pointer;font-weight:normal;" class="btn_integrer_prestataire badge btn-sm btn-warning rounded-pill" data-href="{integrer_prestataire_url}"><i class="fa fa-plus" ></i> Intégrer</span>')

        checkbox_button = f'<input type="checkbox" name="checkbox_button" class="checkbox_button" value="' + str(
            p.id) + '"/>'

        data.append({
            "id": p.id,
            "checkbox": checkbox_button,
            "name": p.name,
            "code": p.code,
            "type_prestataire": p.type_prestataire.name if p.type_prestataire else "",
            "secteur": p.secteur.libelle if p.secteur else "",
            "ville": p.ville,
            "telephone": p.telephone,
            "statut": "ACTIF" if p.status else " INACTIF",
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def joindre_prestataires_reseau(request, reseau_soin_id):
    if request.method == 'POST':

        reseau_soin = ReseauSoin.objects.filter(id=reseau_soin_id).first()

        prestataires_ids = literal_eval(request.POST.get('selectedItems'))
        print("prestataires_ids")
        print(prestataires_ids)

        prestataires = Prestataire.objects.filter(id__in=prestataires_ids)

        for prestataire in prestataires:
            prestataire_reseau = PrestataireReseauSoin.objects.create(reseau_soin=reseau_soin,
                                                                      prestataire=prestataire,
                                                                      created_by=request.user,
                                                                      date_integration=datetime.datetime.now(
                                                                          tz=timezone.utc))
            prestataire_reseau.save()

        response = {
            'statut': 1,
            'message': "Prestataires intégrés au réseau de soins avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Not found",
            'data': {}
        }

    return JsonResponse(response)


def joindre_prestataire_reseau(request, reseau_soin_id, prestataire_id):
    if request.method == 'POST':

        reseau_soin = ReseauSoin.objects.filter(id=reseau_soin_id).first()
        prestataire = Prestataire.objects.filter(id=prestataire_id).first()

        if reseau_soin and prestataire:
            prestataire_reseau = PrestataireReseauSoin.objects.create(reseau_soin=reseau_soin,
                                                                      prestataire=prestataire,
                                                                      created_by=request.user,
                                                                      date_integration=datetime.datetime.now(
                                                                          tz=timezone.utc)
                                                                      )
            prestataire_reseau.save()

            response = {
                'statut': 1,
                'message': "Prestataire intégré au réseau de soins avec succès !",
                'data': {
                }
            }


        else:

            response = {
                'statut': 0,
                'message': "Prestataire ou réseau de soins introuvable",
                'data': {}
            }

    return JsonResponse(response)


def retirer_prestataire_reseau(request, reseau_soin_id, prestataire_id):
    if request.method == 'POST':
        prestataire_reseau = PrestataireReseauSoin.objects.filter(reseau_soin_id=reseau_soin_id,
                                                                  prestataire_id=prestataire_id,
                                                                  statut_validite=StatutValidite.VALIDE).update(
            statut_validite=StatutValidite.CLOTURE,
            removed_by=request.user,
            date_retrait=datetime.datetime.now(tz=timezone.utc),
        )

        response = {
            'statut': 1,
            'message': "Prestataire intégré au réseau de soins avec succès !",
            'data': {
            }
        }

    return JsonResponse(response)


# FIN RESEAU DE SOINS

#
class GroupePermissionsView(TemplateView):
    template_name = 'groupes/groupes_permissions.html'
    model = Permission

    def get(self, request, groupe_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        groupes = AuthGroup.objects.all()
        permissions = Permission.objects.all().order_by('content_type_id')

        context_perso = {
            'groupes': groupes,
            'permissions': permissions
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required()
# modiifie le statut d'un prestataire depuis le toggle sur la page détail du prestataire
def change_prestataire_status(request, prestataire_id):
    response = None

    if request.method == 'POST':

        prestataire = Prestataire.objects.get(id=prestataire_id)

        if prestataire.status == True:
            prestataire.status = False

        else:
            prestataire.status = True
        prestataire.save()
        # gardons des traces
        if prestataire.status == False:
            ActionLog.objects.create(done_by=request.user, action="update",
                                     description="Désactivation d'un prestataire", table="prestataire",
                                     row=prestataire.pk,
                                     # data_before=json.dumps(model_to_dict(formule_before)),
                                     # data_after=json.dumps(model_to_dict(formule))
                                     )
        else:
            ActionLog.objects.create(done_by=request.user, action="update",
                                     description="Activation d'un prestataire", table="prestataire",
                                     row=prestataire.pk,
                                     # data_before=json.dumps(model_to_dict(formule_before)),
                                     # data_after=json.dumps(model_to_dict(formule))
                                     )

        response = {
            'statut': 1,
            'message': "Statut Prestataire changé avec succès !",
            'data': {
            }
        }
        print(prestataire.status)

    return JsonResponse(response)


class WsBobyView(TemplateView):
    # permission_required = "configurations.view_prestataire"
    template_name = 'ws_bobys/bobys.html'
    model = WsBoby

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)
        try:
            del request.session['name']
            del request.session['query']
            del request.session['params']
            del request.session['value_params']
        except:
            pass
        context = {**context_original}
        return self.render_to_response(context)


def ws_boby_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search = request.GET.get('search[value]', '')
    print('search')
    print(search)

    queryset = WsBoby.objects.all()

    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) |
            Q(request__icontains=search)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'id',
        1: 'name',
        2: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for ws_boby in page_obj:
        update_url = reverse('ws_boby_edite', args=[ws_boby.id])  # URL to the detail view

        actions_html = f'<a style="cursor:pointer;" class="badge btn-sm btn-modifier rounded-pill" href="{update_url}"><i class="fa fa-edit"></i> Modifier</a>'

        data.append({
            "id": ws_boby.id,
            "name": ws_boby.name,
            "status": '<span class="badge badge-actif">Actif</span>' if ws_boby.status else '<span class="badge badge-inactif">Inactif</span>',
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


class WsBobyCreateView(TemplateView):
    # permission_required = "configurations.view_prestataire"
    template_name = 'ws_bobys/add_boby.html'
    model = WsBoby

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

    def post(self, request, *args, **kwargs):
        print(request.POST)
        query_data = request.POST
        name = query_data.get('name', '')
        query = query_data.get('query', '')
        params = []
        params = query_data.getlist('params[]', [])
        value_params = query_data.getlist('value_params[]', [])

        print('params')
        print(params)

        request.session['name'] = name if name else ""
        request.session['query'] = query if query else ""
        request.session['params'] = params if params else []
        request.session['value_params'] = value_params if value_params else []
        with transaction.atomic():
            try:
                verify_sql_query(query)

                ws_boby = WsBoby.objects.create(name=name, request=query, status=True)
                ws_boby.save()

                for i in range(len(params)):
                    param_ws_boby = ParamWsBoby.objects.create(ws_boby=ws_boby, name=params[i], value=value_params[i])
                    param_ws_boby.save()

                # request.session['message'] = "Bobys ajouté avec succès !"
                del request.session['name']
                del request.session['query']
                del request.session['params']
                del request.session['value_params']
                messages.success(request, "Boby ajouté avec succès !")
                return redirect('ws_bobys')

            except Exception as e:
                print(e)
                # request.session['message'] =
                messages.error(request, "ERREUR: " + str(e))
                return redirect('ws_boby_create')


class WsBobyEditeView(TemplateView):
    # permission_required = "configurations.view_prestataire"
    template_name = 'ws_bobys/edite_boby.html'
    model = WsBoby

    def get(self, request, ws_boby_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        ws_boby = WsBoby.objects.get(id=ws_boby_id)

        context_perso = {
            'ws_boby': ws_boby
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self, request, ws_boby_id, *args, **kwargs):
        print(request.POST)
        query_data = request.POST
        name = query_data.get('name', '')
        query = query_data.get('query', '')
        status = query_data.get('status', False)
        params = query_data.getlist('params[]', [])
        value_params = query_data.getlist('value_params[]', [])

        print('params')
        print(params)

        request.session['name'] = name if name else ""
        request.session['query'] = query if query else ""
        request.session['params'] = params if params else []
        request.session['value_params'] = value_params if value_params else []
        request.session['status'] = status
        with transaction.atomic():
            try:
                verify_sql_query(query)

                ws_boby = WsBoby.objects.get(id=ws_boby_id)
                ws_boby.name = name
                ws_boby.request = query
                ws_boby.status = status
                ws_boby.save()

                for j in ws_boby.paramwsboby_set.all():
                    j.delete()

                for i in range(len(params)):
                    param_ws_boby = ParamWsBoby.objects.create(ws_boby=ws_boby, name=params[i], value=value_params[i])
                    param_ws_boby.save()

                # request.session['message'] = "Bobys ajouté avec succès !"
                del request.session['name']
                del request.session['query']
                del request.session['params']
                del request.session['value_params']
                del request.session['status']

                messages.success(request, "Boby modifié avec succès !")
                return redirect('ws_bobys')

            except Exception as e:
                print(e)
                # request.session['message'] =
                messages.error(request, "ERREUR: " + str(e))
                return redirect('ws_boby_edite', ws_boby_id)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# ACTE
class ActesView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_acte"
    template_name = 'acte/actes.html'
    model = Acte

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        rubriques = Rubrique.objects.filter(status=True)
        liste_regroupements_actes = RegroupementActe.objects.filter(status=True)
        regroupements_actes = {
            rubrique.pk: [
                {'name': regroupements_acte.pk, 'value': regroupements_acte.libelle}
                for regroupements_acte in RegroupementActe.objects.filter(rubrique_id=rubrique.pk, status=True)
            ]
            for rubrique in rubriques
        }

        all_type_actes = TypeActe.objects.all()
        type_actes = json.dumps(list(all_type_actes.values('id', 'libelle')))

        base_calcul_tm_choices = BaseCalculTM.choices

        context_perso = {
            'rubriques': rubriques,
            'regroupements_actes': regroupements_actes,
            'liste_regroupements_actes': liste_regroupements_actes,
            'type_actes': type_actes,
            'base_calcul_tm_choices': base_calcul_tm_choices,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def actes_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_rubrique = request.GET.get('search_rubrique', '')
    search_code = request.GET.get('search_code', '')
    search_libelle = request.GET.get('search_libelle', '')
    search_entente_prealable = request.GET.get('search_entente_prealable', '')
    search_value = request.GET.get('search[value]', '')

    queryset = Acte.objects.filter(statut_validite=StatutValidite.VALIDE)

    if search_libelle:
        queryset = queryset.filter(
            Q(libelle__icontains=search_libelle)
        )

    if search_code:
        queryset = queryset.filter(
            Q(code__icontains=search_code)
        )

    if search_rubrique:
        queryset = queryset.filter(
            Q(rubrique_id=search_rubrique)
        )

    if search_entente_prealable:
        queryset = queryset.filter(
            Q(rubrique_id=search_entente_prealable)
        )

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: 'code',
        1: 'libelle',
        2: 'rubrique__libelle',
        3: 'regroupement_acte__libelle',
        4: 'lettre_cle',
        5: 'base_calcul_tm',
        7: 'delais_controle',
        8: 'entente_prealable',
        9: 'option_seance',
        10: 'specialiste_uniquement',
        11: 'status',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for acte in page_obj:
        parametre = ParamActe.objects.filter(bureau=request.user.bureau,acte=acte).first()
        detail_url = reverse('popup_detail_acte', args=[acte.id])  # url the detail view
        update_url = reverse('popup_modifier_acte', args=[acte.id])  # url to update view

        actions_html = (
            f'<span style="cursor:pointer;" class="btn-popup_details_acte badge btn-sm btn-details rounded-pill" data-href="{detail_url}"<span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span>&nbsp;'
            f'<span style="cursor:pointer;" class="btn_modifier_acte badge btn-sm btn-modifier rounded-pill" data-href="{update_url}"><i class="fa fa-edit"></i> Modifier</span>')

        data.append({
            "id": acte.id,
            "code": acte.code,
            "libelle": acte.libelle,
            "type": acte.type_acte.libelle if acte.type_acte else "",
            "rubrique": acte.rubrique.libelle if acte.rubrique else "",
            "regroupement": acte.regroupement_acte.libelle if acte.regroupement_acte else "",
            "lettre_cle": acte.lettre_cle,
            "base_calcul_tm": acte.base_calcul_tm,
            "delais_controle": parametre.delais_controle if parametre else 0,
            "entente_prealable": parametre.entente_prealable if parametre else None,
            "option_seance": acte.option_seance,
            "specialiste_uniquement": parametre.specialiste_uniquement if parametre else None,
            "status": acte.status,
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def popup_detail_acte(request, acte_id):
    acte = Acte.objects.get(id=acte_id)

    return render(request, 'acte/modal_details_acte.html', {'acte': acte})


def add_acte(request):

    if request.method == 'POST':

        bureau_id = request.POST.get('bureau_id')
        bureau = Bureau.objects.get(id=bureau_id)
        if bureau:

            rubrique_id = request.POST.get('rubrique_id')
            rubrique = Rubrique.objects.get(id=rubrique_id)
            regroupement_acte_id = request.POST.get('regroupement_acte_id')
            regroupement_acte = RegroupementActe.objects.get(id=regroupement_acte_id)
            type_acte_id = request.POST.get('type_acte')
            type_acte = TypeActe.objects.get(id=type_acte_id)
            libelle = request.POST.get('libelle', None)
            
            code = request.POST.get('code', None)
            if code == '':
                code = None

            lettre_cle = request.POST.get('lettre_cle', None)
            
            delais_carence = request.POST.get('delais_carence', None)
            if delais_carence:
                delais_carence = int(delais_carence.replace(' ', ''))
            else:
                delais_carence = None

            delais_controle = request.POST.get('delais_controle', None)
            if delais_controle:
                delais_controle = int(delais_controle.replace(' ', ''))
            else:
                delais_controle = None

            base_calcul_tm = request.POST.get('base_calcul_tm', None)
            specialiste_uniquement = 'specialiste_uniquement' in request.POST
            est_gratuit = 'est_gratuit' in request.POST
            status = 'status' in request.POST

            nouveau_acte = Acte.objects.create(
                rubrique = rubrique,
                regroupement_acte = regroupement_acte,
                type_acte = type_acte,
                libelle = libelle,
                code = code,
                lettre_cle = lettre_cle,
                delais_carence = delais_carence,
                delais_controle = delais_controle,
                base_calcul_tm = base_calcul_tm,
                specialiste_uniquement = specialiste_uniquement,
                est_gratuit = est_gratuit,
                status = status
            )
            nouveau_acte.save()

            #   coef_classique = request.POST.get('coef_classique', None)
            #   pu_classique = request.POST.get('pu_classique', None)
            #   cout_classique = request.POST.get('cout_classique', None)
            #   #
            #   coef_classique = int(coef_classique.replace(' ', '')) if coef_classique else None
            #   pu_classique = int(pu_classique.replace(' ', '')) if pu_classique else None
            #   cout_classique = int(cout_classique.replace(' ', '')) if cout_classique else None
#   
            #   coef_mutuelle = request.POST.get('coef_mutuelle', None)
            #   pu_mutuelle = request.POST.get('pu_mutuelle', None)
            #   cout_mutuelle = request.POST.get('cout_mutuelle', None)
            #   #
            #   coef_mutuelle = int(coef_mutuelle.replace(' ', '')) if coef_mutuelle else None
            #   pu_mutuelle = int(pu_mutuelle.replace(' ', '')) if pu_mutuelle else None
            #   cout_mutuelle = int(cout_mutuelle.replace(' ', '')) if cout_mutuelle else None
            #   
            #   coef_hg = request.POST.get('coef_hg', None)
            #   pu_hg = request.POST.get('pu_hg', None)
            #   cout_hg = request.POST.get('cout_hg', None)
            #   #
            #   coef_hg = int(coef_hg.replace(' ', '')) if coef_hg else None
            #   pu_hg = int(pu_hg.replace(' ', '')) if pu_hg else None
            #   cout_hg = int(cout_hg.replace(' ', '')) if cout_hg else None
            #   
            #   coef_chu = request.POST.get('coef_chu', None)
            #   pu_chu = request.POST.get('pu_chu', None)
            #   cout_chu = request.POST.get('cout_chu', None)
            #   #
            #   coef_chu = int(coef_chu.replace(' ', '')) if coef_chu else None
            #   pu_chu = int(pu_chu.replace(' ', '')) if pu_chu else None
            #   cout_chu = int(cout_chu.replace(' ', '')) if cout_chu else None
#   
            #   coef_ica = request.POST.get('coef_ica', None)
            #   pu_ica = request.POST.get('pu_ica', None)
            #   cout_ica = request.POST.get('cout_ica', None)
            #   #
            #   coef_ica = int(coef_ica.replace(' ', '')) if coef_ica else None
            #   pu_ica = int(pu_ica.replace(' ', '')) if pu_ica else None
            #   cout_ica = int(cout_ica.replace(' ', '')) if cout_ica else None
#   
            #   new_related_tarif = Tarif.objects.create(
            #       created_by = request.user,
            #       bureau = bureau,
            #       # #
            #       coef_classique = coef_classique,
            #       pu_classique = pu_classique,
            #       cout_classique = cout_classique,
            #       #
            #       coef_mutuelle = coef_mutuelle,
            #       pu_mutuelle = pu_mutuelle,
            #       cout_mutuelle = cout_mutuelle,
            #       #
            #       coef_public_hg = coef_hg,
            #       pu_public_hg = pu_hg,
            #       cout_public_hg = cout_hg,
            #       #
            #       coef_public_chu = coef_chu,
            #       pu_public_chu = pu_chu,
            #       cout_public_chu = cout_chu,
            #       #
            #       coef_public_ica = coef_ica,
            #       pu_public_ica = pu_ica,
            #       cout_public_ica = cout_ica
            #   )
            #   new_related_tarif.save()
            #   new_related_tarif.acte = nouveau_acte
            #   new_related_tarif.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {}
            }

        else:
            response = {
                'statut': 0,
                'message': "Vous n'êtes lié à aucun bureau !",
                'data': {}
            }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def popup_modifier_acte(request, acte_id):
    #
    user = request.user


    acte = Acte.objects.get(id=acte_id)
    # identique à celui de def update_acte
    if request.user.is_superuser: # Filtre les paramètres en fonction du statut d'administrateur
        parametres = ParamActe.objects.filter(acte=acte)  # Tous les paramètres pour les super admins
    else:
        parametres = ParamActe.objects.filter(acte=acte, bureau=request.user.bureau)  # Filtre par bureau pour les utilisateurs normaux
    tarif = Tarif.objects.filter(acte_id=acte_id).first()

    rubriques = Rubrique.objects.filter(status=True)

    selected_rubrique_id = acte.rubrique.id if acte.rubrique else None

    rubrique_regroupement_actes = []
    selected_regroupement_acte_id = None

    if selected_rubrique_id:
        selected_rubrique = Rubrique.objects.get(id=selected_rubrique_id)
        rubrique_regroupement_actes = RegroupementActe.objects.filter(rubrique=selected_rubrique)
        selected_regroupement_acte_id = acte.regroupement_acte.id if acte.regroupement_acte else None

    all_type_actes = TypeActe.objects.all()
    type_actes = json.dumps(list(all_type_actes.values('id', 'libelle')))

    base_calcul_tm_choices = BaseCalculTM.choices
    selected_base_calcul_tm = acte.base_calcul_tm
    
    
    related_tarifs = Tarif.objects.filter(acte_id=acte_id, bureau=user.bureau, prestataire__isnull=True, statut=Statut.ACTIF).all()

    return render(request, 'acte/modal_modifier_acte.html', {
        'acte': acte,
        'parametres': parametres,
        'tarif':tarif,
        'rubriques': rubriques,
        'selected_rubrique_id': selected_rubrique_id,
        'rubrique_regroupement_actes': rubrique_regroupement_actes,
        'selected_regroupement_acte_id': selected_regroupement_acte_id,
        'type_actes':type_actes,
        'base_calcul_tm_choices':base_calcul_tm_choices,
        'selected_base_calcul_tm': selected_base_calcul_tm,
        'related_tarifs':related_tarifs,
    })




def update_acte(request, acte_id):

    if request.method == 'POST':
        bureau = request.user.bureau

        acte = Acte.objects.filter(id=acte_id).first()
        tarif = Tarif.objects.filter(acte_id=acte_id).first()

        # update acte
        if bureau:

            rubrique_id = request.POST.get('rubrique_id')
            acte.rubrique = Rubrique.objects.get(id=rubrique_id)

            regroupement_acte_id = request.POST.get('regroupement_acte_id')
            acte.regroupement_acte = RegroupementActe.objects.get(id=regroupement_acte_id)

            type_acte_id = request.POST.get('type_acte')
            acte.type_acte = TypeActe.objects.get(id=type_acte_id)

            acte.libelle = request.POST.get('libelle', None)

            code = request.POST.get('code', None)
            if code == '':
                code = None
            #acte.code = code

            acte.lettre_cle = request.POST.get('lettre_cle', None)

            delais_controle = request.POST.get('delais_controle', None)
            if delais_controle:
                delais_controle = int(delais_controle.replace(' ', ''))
            else:
                delais_controle = None

            acte.save()
            
        #Enregistrer les paramètres (particularités) de l'acte pour le bureau de l'utilisateur connecté
        # excatement le même envoyer au modal (render modal_modifier_acte)
        if request.user.is_superuser: # Filtre les paramètres en fonction du statut d'administrateur
            parametres = ParamActe.objects.filter(acte=acte)  # Tous les paramètres pour les super admins
        else:
            parametres = ParamActe.objects.filter(acte=acte, bureau=request.user.bureau)  # Filtre par bureau pour les utilisateurs normaux
            
        for i in range(1, len(parametres) + 1):
            bureau_id = request.POST.get(f'bureau_id_{i}')
            delais_controle = request.POST.get(f'delais_controle_{i}', '').replace(' ', '')
            entente_prealable = f'entente_prealable_{i}' in request.POST
            specialiste_uniquement = f'specialiste_uniquement_{i}' in request.POST
            est_gratuit = f'est_gratuit_{i}' in request.POST
            status = f'status_{i}' in request.POST

            # Récupérer ou créer l'objet ParamActe
            param_acte = ParamActe.objects.filter(bureau_id=bureau_id, acte=acte).first()
            if param_acte:
                data_before = model_to_dict(param_acte)
                
                param_acte.delais_controle = delais_controle
                param_acte.entente_prealable = entente_prealable
                param_acte.specialiste_uniquement = specialiste_uniquement
                param_acte.est_gratuit = est_gratuit
                param_acte.status = status
                param_acte.updated_by = request.user
                param_acte.updated_at = datetime.datetime.now(tz=timezone.utc)
                param_acte.base_calcul_tm = request.POST.get('base_calcul_tm', None)
                param_acte.save()
                

                data_after = model_to_dict(param_acte)

                ActionLog.objects.create(
                    done_by=request.user, action="update",
                    description="Modification du paramétrage d'un acte", table="param_acte",
                    row=param_acte.pk,
                    data_before=json.dumps(data_before),
                    data_after=json.dumps(data_after)
                )

        '''
        if tarif:
            # update related tarif
            coef_classique = request.POST.get('coef_classique', None)
            pu_classique = request.POST.get('pu_classique', None)
            cout_classique = request.POST.get('cout_classique', None)
            #
            tarif.coef_classique = int(coef_classique.replace(' ', '')) if coef_classique else None
            tarif.pu_classique = int(pu_classique.replace(' ', '')) if pu_classique else None
            tarif.cout_classique = int(cout_classique.replace(' ', '')) if cout_classique else None

            coef_mutuelle = request.POST.get('coef_mutuelle', None)
            pu_mutuelle = request.POST.get('pu_mutuelle', None)
            cout_mutuelle = request.POST.get('cout_mutuelle', None)
            #
            tarif.coef_mutuelle = int(coef_mutuelle.replace(' ', '')) if coef_mutuelle else None
            tarif.pu_mutuelle = int(pu_mutuelle.replace(' ', '')) if pu_mutuelle else None
            tarif.cout_mutuelle = int(cout_mutuelle.replace(' ', '')) if cout_mutuelle else None
            
            coef_hg = request.POST.get('coef_hg', None)
            pu_hg = request.POST.get('pu_hg', None)
            cout_hg = request.POST.get('cout_hg', None)
            #
            tarif.coef_public_hg = int(coef_hg.replace(' ', '')) if coef_hg else None
            tarif.pu_public_hg = int(pu_hg.replace(' ', '')) if pu_hg else None
            tarif.cout_public_hg = int(cout_hg.replace(' ', '')) if cout_hg else None
            
            coef_chu = request.POST.get('coef_chu', None)
            pu_chu = request.POST.get('pu_chu', None)
            cout_chu = request.POST.get('cout_chu', None)
            #
            tarif.coef_public_chu = int(coef_chu.replace(' ', '')) if coef_chu else None
            tarif.pu_public_chu = int(pu_chu.replace(' ', '')) if pu_chu else None
            tarif.cout_public_chu = int(cout_chu.replace(' ', '')) if cout_chu else None

            coef_ica = request.POST.get('coef_ica', None)
            pu_ica = request.POST.get('pu_ica', None)
            cout_ica = request.POST.get('cout_ica', None)
            #
            tarif.coef_public_ica = int(coef_ica.replace(' ', '')) if coef_ica else None
            tarif.pu_public_ica = int(pu_ica.replace(' ', '')) if pu_ica else None
            tarif.cout_public_ica = int(cout_ica.replace(' ', '')) if cout_ica else None
            
            tarif.save()
        '''

        response = {
            'statut': 1,
            'message': "Modification de l'acte effectuée avec succès !",
            'data': {
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Methode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def add_acte_tarif(request, acte_id):
    if request.method == 'POST':
        #
        acte = Acte.objects.filter(id=acte_id).first()
        bureau = request.user.bureau

        if bureau:
            coef_classique = request.POST.get('coef_classique', None)
            pu_classique = request.POST.get('pu_classique', None)
            cout_classique = request.POST.get('cout_classique', None)
            #
            coef_classique = int(coef_classique.replace(' ', '')) if coef_classique else None
            pu_classique = int(pu_classique.replace(' ', '')) if pu_classique else None
            cout_classique = int(cout_classique.replace(' ', '')) if cout_classique else None

            coef_mutuelle = request.POST.get('coef_mutuelle', None)
            pu_mutuelle = request.POST.get('pu_mutuelle', None)
            cout_mutuelle = request.POST.get('cout_mutuelle', None)
            coef_mutuelle = int(coef_mutuelle.replace(' ', '')) if coef_mutuelle else None
            pu_mutuelle = int(pu_mutuelle.replace(' ', '')) if pu_mutuelle else None
            cout_mutuelle = int(cout_mutuelle.replace(' ', '')) if cout_mutuelle else None

            coef_hg = request.POST.get('coef_hg', None)
            pu_hg = request.POST.get('pu_hg', None)
            cout_hg = request.POST.get('cout_hg', None)
            coef_hg = int(coef_hg.replace(' ', '')) if coef_hg else None
            pu_hg = int(pu_hg.replace(' ', '')) if pu_hg else None
            cout_hg = int(cout_hg.replace(' ', '')) if cout_hg else None

            coef_chu = request.POST.get('coef_chu', None)
            pu_chu = request.POST.get('pu_chu', None)
            cout_chu = request.POST.get('cout_chu', None)
            coef_chu = int(coef_chu.replace(' ', '')) if coef_chu else None
            pu_chu = int(pu_chu.replace(' ', '')) if pu_chu else None
            cout_chu = int(cout_chu.replace(' ', '')) if cout_chu else None

            coef_ica = request.POST.get('coef_ica', None)
            pu_ica = request.POST.get('pu_ica', None)
            cout_ica = request.POST.get('cout_ica', None)
            coef_ica = int(coef_ica.replace(' ', '')) if coef_ica else None
            pu_ica = int(pu_ica.replace(' ', '')) if pu_ica else None
            cout_ica = int(cout_ica.replace(' ', '')) if cout_ica else None

            new_tarif = Tarif.objects.create(
                bureau=bureau,
                acte=acte,
                code_acte=acte.code,
                ##
                coef_classique=coef_classique,
                pu_classique=pu_classique,
                cout_classique=cout_classique,
                #
                coef_mutuelle=coef_mutuelle,
                pu_mutuelle=pu_mutuelle,
                cout_mutuelle=cout_mutuelle,
                #
                coef_public_hg=coef_hg,
                pu_public_hg=pu_hg,
                cout_public_hg=cout_hg,
                #
                coef_public_chu=coef_chu,
                pu_public_chu=pu_chu,
                cout_public_chu=cout_chu,
                #
                coef_public_ica=coef_ica,
                pu_public_ica=pu_ica,
                cout_public_ica=cout_ica
            )
            new_tarif.save()

            response = {
                'statut': 1,
                'message': "Ajout de tarif effectuée avec succès !",
                'data': {}
            }
        else:
            response = {
                'statut': 0,
                'message': "Vous n'êtes lié à aucun bureau !",
                'data': {}
            }
    else:
        response = {
            'statut': 0,
            'message': "Méthode non autorisée !",
            'data': {}
        }

    return JsonResponse(response)


def desactiver_tarif_acte(request, acte_id, tarif_id):
    if request.method == 'POST':
        try:
            tarif = Tarif.objects.get(id=tarif_id)
            tarif.statut = Statut.INACTIF
            tarif.statut_validite = StatutValidite.CLOTURE
            tarif.deleted_by = request.user
            tarif.updated_at = datetime.datetime.now(tz=timezone.utc)
            tarif.save()

            response = {
                'statut': 1,
                'message': "Tarif désactivé avec succès",
                'data': {}
            }
        except Tarif.DoesNotExist:
            response = {
                'statut': 0,
                'message': "Tarif non trouvé",
                'data': {}
            }
    else:
        response = {
            'statut': 0,
            'message': "Méthode non autorisée",
            'data': {}
        }

    return JsonResponse(response)


#  2fa


from .helper_config import send_verification_code
from django.contrib.auth import authenticate, login

#########################################
@login_required
def verify_code(request):
    if request.method == 'GET':
        print("User email:", request.user.email)

        request.session['is_verified'] = True
        return redirect('/')

        # if not request.user.email:
        #     request.session['is_verified'] = True
        #     return redirect('/')
        #
        # code = "123456" # send_verification_code(request, request.user.email)
        # if code:
        #     request.session['verification_code'] = code
        #     return render(request, '2fa/verify_code.html')


    elif request.method == 'POST':
        submitted_code = request.POST.get('code')
        stored_code = request.session.get('verification_code')

        print("Submitted code:", submitted_code)
        print("Stored code:", stored_code)

        if submitted_code == stored_code:
            del request.session['verification_code']
            request.session['is_verified'] = True
            messages.success(request, 'Verification successful. You are now logged in.')
            return redirect('/')
        else:
            messages.error(request, 'Invalid verification code. Please try again.')

    return render(request, '2fa/verify_code.html')

@login_required
def download_background_query_result(request, query_id):
    try:
        query = BackgroundQueryTask.objects.get(id=query_id)
        url = query.file.url
        query.delete()
        return redirect(url)
    except Exception as e:
        return redirect(reverse('admin:configurations_backgroundquerytask_changelist'))



    return JsonResponse(response)


## TÉLÉCHARGER MODEL DES TARIFS DU BUREAU

def generate_modele_tarifs_bureau(request):

    # Données à inclure dans le DataFrame
    actes = Acte.objects.filter(type_acte__code="acte", status=True).order_by('rubrique_id')

    # Créer un Workbook et accéder à la première feuille
    wb = Workbook()
    ws = wb.active

    # Données à inclure dans le DataFrame
    data = {
        'CODE_RUBRIQUE': [],
        'LIBELLE_ACTE': [],
        'ID_ACTE': [],
        'CODE_ACTE': [],
        'LETTRE_CLE': [],
        # 'LETTRE_CLE_CLASSIQUE': [],
        'COEF_CLASSIQUE': [],
        'PRIX_UNIT_CLASSIQUE': [],
        'TARIF_CLASSIQUE': [],
        # 'LETTRE_CLE_MUTUELLE': [],
        'COEF_MUTUELLE': [],
        'PRIX_UNIT_MUTUELLE': [],
        'TARIF_MUTUELLE': [],
        # 'LETTRE_CLE_HG': [],
        'COEF_HG': [],
        'PRIX_UNIT_HG': [],
        'TARIF_HG': [],
        # 'LETTRE_CLE_CHU': [],
        'COEF_CHU': [],
        'PRIX_UNIT_CHU': [],
        'TARIF_CHU': [],
        # 'LETTRE_CLE_ICA': [],
        'COEF_ICA': [],
        'PRIX_UNIT_ICA': [],
        'TARIF_ICA': [],
        # 'LETTRE_CLE_PRESTATAIRE': [],
        'COEF_PRESTATAIRE': [],
        'PRIX_UNIT_PRESTATAIRE': [],
        'TARIF_PRESTATAIRE': [],
    }

    # Ajouter les actes au DataFrame
    for acte in actes:
        # renseigner avec le tarif existant de ce prestataire --
        tarif_existant = Tarif.objects.filter(acte__code=acte.code, bureau=request.user.bureau, statut=Statut.ACTIF).first()
        tarif_existant_parametre = Tarif.objects.filter(bureau=request.user.bureau, acte=acte, prestataire_id__isnull=True, statut=Statut.ACTIF).first()

        data['CODE_RUBRIQUE'].append(acte.rubrique.libelle)
        data['LIBELLE_ACTE'].append(acte.libelle)
        data['ID_ACTE'].append(acte.id)
        data['CODE_ACTE'].append(acte.code)
        data['LETTRE_CLE'].append(acte.lettre_cle)
        
        # data['LETTRE_CLE_CLASSIQUE'].append(tarif_existant.lettre_cle_classique)
        data['COEF_CLASSIQUE'].append(tarif_existant.coef_classique if tarif_existant else 1)
        data['PRIX_UNIT_CLASSIQUE'].append(tarif_existant.pu_classique if tarif_existant else 0)
        data['TARIF_CLASSIQUE'].append(tarif_existant.cout_classique if tarif_existant else 0)

        # data['LETTRE_CLE_MUTUELLE'].append(tarif_existant.lettre_cle_classique)
        data['COEF_MUTUELLE'].append(tarif_existant.coef_mutuelle if tarif_existant else 1)
        data['PRIX_UNIT_MUTUELLE'].append(tarif_existant.pu_mutuelle if tarif_existant else 0)
        data['TARIF_MUTUELLE'].append(tarif_existant.cout_mutuelle if tarif_existant else 0)

        # data['LETTRE_CLE_HG'].append(tarif_existant.lettre_cle_public_hg if tarif_existant else 0)
        data['COEF_HG'].append(tarif_existant.coef_public_hg if tarif_existant else 0)
        data['PRIX_UNIT_HG'].append(tarif_existant.pu_public_hg if tarif_existant else 0)
        data['TARIF_HG'].append(tarif_existant.cout_public_hg if tarif_existant else 0)

        # data['LETTRE_CLE_CHU'].append(tarif_existant.lettre_cle_public_chu if tarif_existant else 0)
        data['COEF_CHU'].append(tarif_existant.coef_public_chu if tarif_existant else 0)
        data['PRIX_UNIT_CHU'].append(tarif_existant.pu_public_chu if tarif_existant else 0)
        data['TARIF_CHU'].append(tarif_existant.cout_public_chu if tarif_existant else 0)

        # data['LETTRE_CLE_ICA'].append(tarif_existant.lettre_cle_public_ica if tarif_existant else 0)
        data['COEF_ICA'].append(tarif_existant.coef_public_ica if tarif_existant else 0)
        data['PRIX_UNIT_ICA'].append(tarif_existant.pu_public_ica if tarif_existant else 0)
        data['TARIF_ICA'].append(tarif_existant.cout_public_ica if tarif_existant else 0)

        # data['LETTRE_CLE_PRESTATAIRE'].append('')
        data['COEF_PRESTATAIRE'].append(tarif_existant_parametre.coef_prestataire if tarif_existant_parametre else 1)
        data['PRIX_UNIT_PRESTATAIRE'].append(tarif_existant_parametre.pu_prestataire if tarif_existant_parametre else 0)
        data['TARIF_PRESTATAIRE'].append(tarif_existant_parametre.cout_prestataire if tarif_existant_parametre else 0)

    # Créer un DataFrame avec Pandas
    df = pd.DataFrame(data)
    
    filename = 'INOV_V1-TARIF_DU_BUREAU_' + slugify(str(request.user.bureau)).upper() + '.xlsx'
    pprint(filename)

    # Créer une réponse HTTP avec le type MIME approprié
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename

    # Enregistrer le DataFrame dans le fichier Excel
    df.to_excel(response, index=False, engine='openpyxl')

    return response

from django.core.exceptions import ObjectDoesNotExist
## IMPORTER LES TARIFS AU BUREAU
def import_tarifs_bureau(request):
    try:
        tarifs_existant = Tarif.objects.filter(bureau=request.user.bureau, prestataire__isnull=True, statut=Statut.ACTIF).first()

        if tarifs_existant:
            response = {
                'statut': 0,
                'message': "Vous avez déjà importé les tarifs du bureau !",
                'data': {}
            }
            return JsonResponse(response)

        file = request.FILES['fichier_import_tarif']
        df = pd.read_excel(file)
        rows_count = 0

        for index, row in df.iterrows():
            code_acte = row.get('CODE_ACTE')
            id_acte = row.get('ID_ACTE')

            try:
                acte = Acte.objects.get(id=id_acte)
                pprint("L'acte existe, donc on insère le tarif")

                # Désactiver l'ancien tarif si existant
                # Tarif.objects.filter(acte=acte).update(statut=Statut.INACTIF, statut_validite=StatutValidite.CLOTURE)

                # Insérer une ligne dans la table Tarif avec les coefficients, prix unitaire, etc.
                tarif = Tarif.objects.create(
                    acte=acte,
                    code_acte=code_acte,
                    bureau=request.user.bureau,
                    created_by=request.user,
                    statut=Statut.ACTIF,
                    statut_validite=StatutValidite.VALIDE,

                    coef_public_hg=row.get('COEF_HG'),
                    pu_public_hg=row.get('PRIX_UNIT_HG'),
                    cout_public_hg=row.get('TARIF_HG'),

                    coef_public_chu=row.get('COEF_CHU'),
                    pu_public_chu=row.get('PRIX_UNIT_CHU'),
                    cout_public_chu=row.get('TARIF_CHU'),

                    coef_public_ica=row.get('COEF_ICA'),
                    pu_public_ica=row.get('PRIX_UNIT_ICA'),
                    cout_public_ica=row.get('TARIF_ICA'),

                    coef_mutuelle=row.get('COEF_MUTUELLE'),
                    pu_mutuelle=row.get('PRIX_UNIT_MUTUELLE'),
                    cout_mutuelle=row.get('TARIF_MUTUELLE'),

                    coef_classique=row.get('COEF_CLASSIQUE'),
                    pu_classique=row.get('PRIX_UNIT_CLASSIQUE'),
                    cout_classique=row.get('TARIF_CLASSIQUE'),

                    coef_prestataire=row.get('COEF_PRESTATAIRE'),
                    pu_prestataire=row.get('PRIX_UNIT_PRESTATAIRE'),
                    cout_prestataire=row.get('TARIF_PRESTATAIRE'),
                )
            except ObjectDoesNotExist:
                continue

        response = {
            'statut': 1,
            'message': "Tarifs importés avec succès !",
            'data': {}
        }

    except Exception as e:
        response = {
            'statut': 0,
            'message': f"Erreur lors de l'importation des tarifs : {str(e)}",
            'data': {}
        }

    return JsonResponse(response)


class ConnectedUsersView(PermissionRequiredMixin, TemplateView):
    permission_required = "configurations.view_prestataire"
    template_name = 'users/connected_users.html'
    model = User

    def format_duration(self, duration):
        if duration is None:
            return "0:00:00"

        total_seconds = int(duration.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)

        return f"{hours}:{minutes:02}:{seconds:02}"

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        sessions = Session.objects.filter(expire_date__gt=timezone.now())
        active_user_ids = []
        session_data = {}

        for session in sessions:
            data = session.get_decoded()
            user_id = data.get('_auth_user_id')
            last_visited = data.get('last_visited', '')
            last_ip_address = data.get('last_ip_address', '')
            last_visit_time = data.get('last_visit_time', '')
            last_visit_time_dt = (
                timezone.make_aware(datetime.datetime.fromisoformat(last_visit_time))
                if last_visit_time else None
            )
            elapsed_time = (datetime.datetime.now(tz=timezone.utc) - last_visit_time_dt) if last_visit_time_dt else None

            if user_id:
                active_user_ids.append(user_id)
                session_data[user_id] = {
                    'id': user_id,
                    'last_visited': last_visited,
                    'last_ip_address': last_ip_address,
                    'last_visit_time': datetime.datetime.fromisoformat(last_visit_time) if last_visit_time else None,
                    'elapsed_time': self.format_duration(elapsed_time),
                }

        active_users = User.objects.filter(id__in=active_user_ids)

        #merge active_users_session_data and active_users
        active_users_session_data = []
        for user in active_users:
            session = session_data.get(str(user.id), {})
            active_users_session_data.append({
                'user_data': user,
                'session_data': session,
            })

        bureaux = Bureau.objects.filter(status=True)

        context_perso = {
            'bureaux': bureaux,
            'active_users_session_data': active_users_session_data,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def logout_user(request, user_id):
    try:
        # Récupérer l'utilisateur
        user = User.objects.get(id=user_id)

        # Récupérer toutes les sessions
        sessions = Session.objects.filter(expire_date__gte=timezone.now())

        # Parcourir les sessions
        for session in sessions:
            data = session.get_decoded()
            # Vérifier si l'utilisateur est dans cette session
            if data.get('_auth_user_id') == str(user.id):
                session.delete()  # Supprime la session
                break  # On peut arrêter après avoir trouvé la session

        return redirect(reverse('connectedusers'))

    except User.DoesNotExist:
        print("Utilisateur non trouvé.")


@method_decorator(login_required, name='dispatch')
class DbSuperAdminQueryView(TemplateView):
    template_name = 'db_query/db_super_admin_query.html'
    model = Aliment

    def get(self, request, *args, **kwargs):
        # TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau

        query_datas = [
            {
                "query_label": "MODIFICATION DE DATE D'ENTREE DE BÉNÉFICIAIRE",
                "query_name": "BENEF_ENTREE_MODIF",
            },
            {
                "query_label": "MODIFICATION DE DATE DE SORTIE DE BÉNÉFICIAIRE",
                "query_name": "BENEF_SORTI_MODIF",
            },
            {
                "query_label": "ANNULATION DE QUITTANCE SOLDÉE",
                "query_name": "ANNULATION_QUITTANCE_SOLDEE",
            },
            {
                "query_label": "ANNULATION DE BORDEREAU DE PAIEMENT",
                "query_name": "ANNULATION_BR_PAIEMENT",
            },
            {
                "query_label": "ANNULATION DE BORDEREAU D'ORDONNANCEMENT",
                "query_name": "ANNULATION_BR_ORDONNANCEMENT",
            }
            #
        ]



        context = self.get_context_data(**kwargs)
        context['query_datas'] = query_datas

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        print("----- fn= post -----")
        print(request.POST)
        print(request.POST.dict())

        query_name = request.POST.get('query_name')

        if query_name == "ANNULATION_QUITTANCE_SOLDEE":

            aq_numero = request.POST.get('aq_numero')
            aq_motif = request.POST.get('aq_motif')

            quittance = Quittance.objects.filter(numero=aq_numero, bureau=request.user.bureau).first()

            if quittance:
                # Verification si la quittance est déjà annulée
                if quittance.statut_validite == "ANNULEE":
                    return JsonResponse({
                        "message": "Erreur : Cette quittance a déjà été annulée."
                    }, status=400)

                # recuperation des reglements de la quittance
                reglement = Reglement.objects.filter(quittance=quittance)
                print("reglement", reglement)

                # Verification si reglement REVERSE ou NON
                revcomp = False
                for rgm in reglement:
                    if rgm.statut_reversement_compagnie == 'REVERSE':
                        revcomp = True
                        break

                if revcomp:
                    return JsonResponse({
                        "message": "Erreur : Impossible d'annuler la quittance car elle a déjà été reversée à la compagnie."
                    }, status=404)

                # Annulation de la quittance et de ses reglements
                date_du_jour = datetime.datetime.now()
                observation = f"Annulation de la quittance {quittance.numero} le {date_du_jour} pour motif : {aq_motif}"
                quittance.statut_validite = "ANNULEE"
                quittance.observation = observation
                quittance.save()

                # Annulation des reglements
                for rgm in reglement:
                    rgm.statut_validite = "ANNULE"
                    rgm.observation = observation
                    rgm.motif_annulation = aq_motif
                    rgm.save()

                return JsonResponse({
                    "message": f"Succès : La quittance {aq_numero} a été annulée avec succès."
                }, status=200)

            else:
                return JsonResponse({
                    "message": "Erreur : Quittance introuvable."
                }, status=404)

        ##
        elif query_name == "ANNULATION_BR_PAIEMENT":
            abrp_numero = request.POST.get('abrp_numero')
            abrp_motif = request.POST.get('abrp_motif')

            paiement_comptable = PaiementComptable.objects.filter(numero=abrp_numero, bureau=request.user.bureau).first()

            if paiement_comptable:
                # Verification si le paiement comptable est déjà annulé
                if paiement_comptable.statut_validite == StatutValidite.SUPPRIME:
                    return JsonResponse({
                        "message": "Erreur : Ce bordereau de paiement comptable a déjà été annulé."
                    }, status=400)

                date_du_jour = datetime.datetime.now()
                observation = f"Annulation de bordereau de paiement {paiement_comptable.numero} le {date_du_jour} pour motif : {abrp_motif}"

                # traitement br paiement comptable
                paiement_comptable.pc_deleted_by = request.user
                paiement_comptable.statut_validite = StatutValidite.SUPPRIME
                paiement_comptable.observation = observation
                paiement_comptable.save()

                # recuperation de borderereau associé
                paiement_comptable.bordereau_ordonnancement.statut_paiement = StatutPaiementSinistre.ORDONNANCE
                paiement_comptable.bordereau_ordonnancement.save()

                # récuperation sinistres associés
                sinistres = Sinistre.objects.filter(paiement_comptable=paiement_comptable)
                if sinistres:
                    for sinistre in sinistres:
                        sinistre.paiement_comptable = None
                        sinistre.date_paiement = None
                        sinistre.statut_paiement = StatutPaiementSinistre.ORDONNANCE
                        sinistre.save()
                        # mettre la facture à ordonnancer
                        sinistre.facture_prestataire.statut = SatutBordereauDossierSinistres.ORDONNANCE
                        sinistre.facture_prestataire.save()

                        # historiser les lignes qui étaient sur le bordereau
                        HistoriquePaiementComptableSinistre.objects.create(created_by=request.user,
                                                                        paiement_comptable=paiement_comptable,
                                                                        sinistre=sinistre,
                                                                        montant_paye=sinistre.montant_remb_accepte,
                                                                        observation=observation)

                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="annulation_bordereau_paiement",
                                             description="Annulation d'un bordereau de paiement",
                                             table="paiement_comptable",
                                             row=paiement_comptable.pk)

                return JsonResponse({
                    "message": f"Succès : Le bordereau de paiement {abrp_numero} a été annulé avec succès."
                }, status=200)

            else:
                return JsonResponse({
                    "message": "Erreur : Bordereau de paiement comptable introuvable."
                }, status=404)

        elif query_name == "ANNULATION_BR_ORDONNANCEMENT":

            abro_numero = request.POST.get('abro_numero')
            abro_motif = request.POST.get('abro_motif')

            # recuperation du bordereau d'ordonnancement associé au paiement comptable
            bordereau_ordonnancement = BordereauOrdonnancement.objects.filter(numero=abro_numero, bureau=request.user.bureau).first()

            if bordereau_ordonnancement:
                # Verification si le bordereau d'ordonnancement est déjà annulé
                if bordereau_ordonnancement.statut_validite == StatutValidite.SUPPRIME:
                    return JsonResponse({
                        "message": "Erreur : Ce bordereau d'ordonnancement a déjà été annulé."
                    }, status=400)

                date_du_jour = datetime.datetime.now()
                observation = f"Annulation de bordereau d'ordonnancement {bordereau_ordonnancement.numero} le {date_du_jour} pour motif : {abro_motif}"

                # traitement facture
                bordereau_ordonnancement.bo_deleted_by = request.user
                bordereau_ordonnancement.statut_paiement = StatutPaiementSinistre.ATTENTE
                bordereau_ordonnancement.statut_validite = StatutValidite.SUPPRIME
                bordereau_ordonnancement.observation = observation
                bordereau_ordonnancement.save()

                # récuperation sinistres associés
                sinistres = Sinistre.objects.filter(bordereau_ordonnancement=bordereau_ordonnancement)
                if sinistres:
                    for sinistre in sinistres:
                        sinistre.bordereau_ordonnancement = None
                        sinistre.statut_paiement = StatutPaiementSinistre.ATTENTE
                        # sinistre.observation = str(sinistre.observation)
                        sinistre.save()

                        # mettre la facture à traitée:: faire sortir de la boucle lorsque les factures seront directement liées aux bordereau d'ordonnancement
                        sinistre.facture_prestataire.statut = SatutBordereauDossierSinistres.VALIDE
                        sinistre.facture_prestataire.save()

                        # historiser les lignes qui étaient sur le bordereau
                        HistoriqueOrdonnancementSinistre.objects.create(created_by=request.user,
                                                                        bordereau_ordonnancement=bordereau_ordonnancement,
                                                                        sinistre=sinistre,
                                                                        montant_ordonnance=sinistre.montant_remb_accepte,
                                                                        observation=observation)

                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="annulation_bordereau_ordonnancement",
                                         description="Annulation d'un bordereau d'ordonnancement",
                                         table="bordereau_ordonnancement",
                                         row=bordereau_ordonnancement.pk)

                return JsonResponse({
                    "message": f"Succès : Le bordereau d'ordonnancement {abro_numero} a été annulé avec succès."
                }, status=200)
            else:
                return JsonResponse({
                    "message": "Erreur : Bordereau d'ordonnancement introuvable."
                }, status=404)
        elif query_name == "BENEF_ENTREE_MODIF":

            benefe_numero = request.POST.get('benefe_numero')
            benefe_date = request.POST.get('benefe_date')
            benefe_motif = request.POST.get('benefe_motif')



            carte = Carte.objects.filter(numero=benefe_numero, statut=Statut.ACTIF, aliment__bureau=request.user.bureau).first()

            if carte:
                aliment = carte.aliment
                aliment_formule = aliment.aliment_formule if aliment else None
                mouvement = aliment.last_mouvement if aliment else None

                # liste de sinistre lié à la carte
                sinistre = aliment.ses_sinistres.all().filter(
                    statut_validite=StatutValidite.VALIDE,
                    statut__in=[StatutSinistre.ACCORDE, StatutSinistre.ATTENTE],
                    date_survenance__lte=benefe_date
                ).order_by('-date_survenance').first() if aliment else None

                # verification si le bénéficiaire n'a pas sinistre a enterieur a la nouvelle date d'entree
                if sinistre:
                    return JsonResponse({
                        "message": "Erreur : Impossible de modifier la date d'entrée du bénéficiaire, car un sinistre antérieur à la nouvelle date d'entrée a été détecté."
                    }, status=404)


                date_du_jour = datetime.datetime.now()
                observation = f"Modification de date d'entrée de {carte.numero} le {date_du_jour} pour motif : {benefe_motif}"

                if aliment:
                    aliment.date_affiliation = benefe_date
                    aliment.save()

                    if aliment_formule:
                        aliment_formule.date_debut = benefe_date
                        aliment_formule.observation = observation
                        aliment_formule.save()

                    if mouvement:
                        if mouvement.mouvement_id == 7:  # INCORPORATION
                            mouvement.date_effet = benefe_date
                            mouvement.observation = observation
                            mouvement.save()
                else:
                    return JsonResponse({
                        "message": "Erreur : Bénéficiaire introuvable."
                    }, status=404)


                return JsonResponse({
                    "message": f"Succès : La date d'entrée du bénéficiaire {benefe_numero} a été modifié avec succès."
                }, status=200)


            else:
                return JsonResponse({
                    "message": "Erreur : Bénéficiaire introuvable."
                }, status=404)


        elif query_name == "BENEF_SORTI_MODIF":

            benefs_numero = request.POST.get('benefs_numero')
            benefs_date = request.POST.get('benefs_date')
            benefs_motif = request.POST.get('benefs_motif')

            carte = Carte.objects.filter(numero=benefs_numero, statut=Statut.ACTIF,
                                         aliment__bureau=request.user.bureau).first()

            if carte:
                aliment = carte.aliment
                aliment_formule = aliment.aliment_formule if aliment else None
                mouvement = aliment.last_mouvement if aliment else None

                # liste de sinistre lié à la carte
                sinistre = aliment.ses_sinistres.all().filter(
                    statut_validite=StatutValidite.VALIDE,
                    statut__in=[StatutSinistre.ACCORDE, StatutSinistre.ATTENTE],
                    date_survenance__gte=benefs_date
                ).order_by('-date_survenance').first() if aliment else None

                # verification si le bénéficiaire n'a pas sinistre a enterieur a la nouvelle date d'entree
                if sinistre:
                    return JsonResponse({
                        "message": "Erreur : Impossible de modifier la date d'entrée du bénéficiaire, car un sinistre posterieur à la nouvelle date de sortie a été détecté."
                    }, status=404)

                date_du_jour = datetime.datetime.now()
                observation = f"Modification de date de sortie de {carte.numero} le {date_du_jour} pour motif : {benefs_motif}"


                if aliment:
                    if aliment_formule.date_fin:
                        aliment.date_sortie = benefs_date
                        aliment.save()

                        aliment_formule.date_fin = benefs_date
                        aliment_formule.observation = observation
                        aliment_formule.save()

                        if mouvement.mouvement_id == 11:  # SORTIE-BENEF
                            mouvement.date_effet = benefs_date
                            mouvement.observation = observation
                            mouvement.save()

                        return JsonResponse({
                            "message": f"Succès : La date de sortie du bénéficiaire {benefs_numero} a été modifié avec succès."
                        }, status=200)

                    else:
                        return JsonResponse({
                            "message": "Erreur : Impossible de modifier la date de sortie du bénéficiaire, car il n'est pas encore sorti."
                        }, status=404)


                else:
                    return JsonResponse({
                        "message": "Erreur : Bénéficiaire introuvable."
                    }, status=404)

            else:
                return JsonResponse({
                    "message": "Erreur : Bénéficiaire introuvable."
                }, status=404)
        else:
            return JsonResponse({
                "message": "Erreur : Cette action n'est pas prise en charge."
            }, status=404)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



