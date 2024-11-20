import datetime
import json
import os
from ast import literal_eval
from collections import defaultdict
from copy import deepcopy
from datetime import datetime as datetimeJsdecode, timedelta
from decimal import Decimal
from functools import reduce
from pprint import pprint
from sqlite3 import Date

import PyPDF2
import openpyxl
import requests
from django.conf import settings
from django.contrib import admin
from django.contrib.auth.decorators import login_required
from django.core import serializers
# Create your views here.
from django.core.files.base import File
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count
from django.db.models import Q
from django.db.models import Sum
from django.db.models import Value, F
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.template.loader import get_template
#
from django.urls import reverse
from django.utils import timezone
from django.utils.datastructures import MultiValueDictKeyError
from django.utils.decorators import method_decorator
from django.utils.timezone import make_aware
from django.views.decorators.cache import never_cache
from django.views.generic import TemplateView, ListView
from num2words import num2words
from xhtml2pdf import pisa

from configurations.helper_config import execute_query, create_query_background_task
from configurations.models import Compagnie, User, Rubrique, Affection, Acte, Prescripteur, Prestataire, \
    TypePriseencharge, \
    JourFerie, ActionLog, PeriodeComptable, TypeRemboursement, ModeCreation, PrescripteurPrestataire, \
    BackgroundQueryTask, TypePrefinancement
from production.models import Carte, Aliment, Statut, TypeDocument, Bareme, Client
#
from production.models import Police, AlimentFormule
from production.templatetags.my_filters import money_field
from shared.enum import StatutPolice
from shared.enum import StatutSinistre, StatutSinistreBordereau, StatutSinistrePrestation, StatutValidite, \
    StatutRemboursement, StatutRemboursementSinistre, DesignationRemboursementSinistre, SatutBordereauDossierSinistres, \
    StatutPaiementSinistre, TypeBonConsultation
from shared.helpers import bool_plafond_atteint, get_tarif_acte_from_bareme, api_send_sms, render_pdf, link_callback, \
    is_jour_ferie, \
    as_money, get_ticket_moderateur_pharmacie, get_type_prefinancement_of_acte, \
    recalcule_montant_refacture_compagnie_et_client
from sinistre.helper_sinistre import exportation_en_excel_avec_style, \
    extraction_demandes_accords_prealables_traitees_par_medecins_conseil, extraction_des_sinistres_traites_valides, \
    requete_demandes_accords_prealables_traitees_par_les_medecins_conseil, requete_liste_des_sp_client_par_filiale, \
    requete_liste_paiement_sinistre_sante_entre_deux_dates, \
    requete_liste_sinistre_ordonnancee_par_period_par_beneficiaire, \
    requete_liste_sinistre_ordonnancee_par_period_par_prestataire, \
    requete_liste_sinistre_ordonnancee_par_period, requete_liste_sinistre_entre_2date, requete_analyse_prime_compta, \
    requete_liste_sinistre_saisies_entre_2date, requete_sinistres_traites_et_valides_par_les_gestionnaires, \
    requete_analyse_prime_compta_apporteur, get_retenue_selon_contexte
# Create your views here.
from sinistre.models import PaiementComptable, Sinistre, DossierSinistre, DocumentDossierSinistre, ProrogationSinistre, SinistreTemporaire, \
    FacturePrestataire, RemboursementSinistre, BordereauOrdonnancement, HistoriqueOrdonnancementSinistre


@method_decorator(login_required, name='dispatch')
class DossierSinistresView(TemplateView):
    template_name = 'liste_dossiers_sinistres.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        prestataires = Prestataire.objects.filter(
            bureau=request.user.bureau,
            status=True,
            id__in=DossierSinistre.objects.filter(
                statut_validite="VALIDE"
            ).values('prestataire_id')
        )

        today = timezone.now().date()
        context['today'] = today
        context['prestataires'] = prestataires
        context['breadcrumbs'] = [
            {'title': 'Prises en charges', 'url': ''},
            {'title': 'Traités', 'url': ''},
        ]

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def dossiersinistre_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_assure = request.GET.get('num_assure', '')
    search_numero_dossier_sinistre = request.GET.get('num_feuille_soins', '')
    search_date_survenance = request.GET.get('date_prestation', '')
    search_prestataire = request.GET.get('prestataire', '')
    date_reception_facture = request.GET.get('date_reception_facture', '')
    reference_facture = request.GET.get('reference_facture', '')
    statut_pec = request.GET.get('statut_pec', '')

    pprint("search_numero_assure")
    pprint(search_numero_assure)
    pprint("search_numero_dossier_sinistre")
    pprint(search_numero_dossier_sinistre)
    pprint("search_date_survenance")
    pprint(search_date_survenance)
    
    pprint("search_prestataire")
    pprint(search_prestataire)
    
    pprint("date_reception_facture")
    pprint(date_reception_facture)
    
    pprint("reference_facture")
    pprint(reference_facture)
    
    pprint("statut_pec")
    pprint(statut_pec)

    if request.user.is_med:  # inclure plus tard le code pays
        # queryset = [x for x in DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE, bureau=request.user.bureau).order_by('id') if x.has_sinistre_en_attente or x.has_prorogation_en_attente]

        # SOLUTION POUR OPTIMISER: Créer un champ statut_prorogation qui sera mis à jour à chaque qu'il yy a une prorpogation sur un sinistre
        # et un champ statut_entente qui est mis à jour quand les sinistres en attente existe sur le dossier
        if(search_numero_assure or search_numero_dossier_sinistre or search_date_survenance or search_prestataire or date_reception_facture or reference_facture or statut_pec):
            queryset = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE, bureau=request.user.bureau).filter(
                Q(statut_pec=StatutSinistre.ATTENTE) | Q(statut_prorogation=StatutSinistre.ATTENTE)).order_by('id')
                                                    
        else:
            queryset = DossierSinistre.objects.none()
            
        # pprint(queryset)


    elif request.user.is_pres or request.user.is_imag or request.user.is_optic or request.user.is_labo or request.user.is_dentaire:
        #cas de presta : pas besoin d'appliquer de filtre d'optimisation (données reduit)
        queryset = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE,
                                                  prestataire=request.user.prestataire).order_by('-id')

    elif request.user.is_pharm:  # Updated on 11102023: remove filtre , is_closed=True, updated on 25112023: filtrer les sinistres et regrouper pour trouver id des dossier_sinistres a afficher
        # queryset_dossier_sinistre_ids = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE, sinistres__prestataire=request.user.prestataire).values('id')
        #cas de pharmacie : pas besoin d'appliquer de filtre d'optimisation (données reduit)
        queryset_dossier_sinistre_ids = Sinistre.objects.filter(statut_validite=StatutValidite.VALIDE,
                                                                prestataire=request.user.prestataire).values(
            'dossier_sinistre_id')

        # Extracting a list of unique IDs from the queryset_uniq
        list_dossier_sinistre_ids = [item['dossier_sinistre_id'] for item in queryset_dossier_sinistre_ids]

        # Filtering DossierSinistre objects based on the unique IDs
        queryset = DossierSinistre.objects.filter(id__in=list_dossier_sinistre_ids).order_by('-id')


    else:
        if (search_numero_assure or search_numero_dossier_sinistre or search_date_survenance or search_prestataire or date_reception_facture or reference_facture or statut_pec):
            queryset = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE,
                                                  bureau=request.user.bureau).order_by('-id')
        else:
            queryset = DossierSinistre.objects.none()

    # la recherche
    if search_numero_assure:
        cartes = Carte.objects.filter(numero__contains=search_numero_assure)
        carte = cartes.first() if cartes else None
        aliment = carte.aliment if carte else None
        queryset = queryset.filter(aliment_id=aliment.pk) if aliment else queryset.filter(numero="nexisterajamais")

    if search_numero_dossier_sinistre:
        queryset = queryset.filter(numero=search_numero_dossier_sinistre)

    if search_date_survenance:
        queryset = queryset.filter(date_survenance__contains=search_date_survenance)

    if search_prestataire:
        queryset = queryset.filter(prestataire=search_prestataire)

    if date_reception_facture:
        queryset = queryset.filter(date_reception_facture__contains=date_reception_facture)


    if reference_facture:
        queryset = queryset.filter(reference_facture__contains=reference_facture)
        
    if statut_pec:
        queryset = queryset.filter(Q(statut_pec=statut_pec) | Q(statut_prorogation=statut_pec))

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'aliment__nom',
        2: 'statut',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    # filter les pec dont le statut est EN ATTENTE
    if request.user.is_med:
        queryset_without_accord = [x.id for x in queryset if x.statut == "EN ATTENTE" or x.statut_prorogation == "EN ATTENTE"]
        queryset = queryset.filter(id__in=queryset_without_accord)


    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_dossier_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'

        if request.user.is_pharm:
            type_or_numero_carte = c.aliment.carte_active().numero if c.aliment and c.aliment.carte_active() else ''
        else:
            if request.user.is_med and c.type_priseencharge.code == "CONSULT":
                type_or_numero_carte = "PHARMACIE"
            else:
                type_or_numero_carte = c.type_priseencharge.libelle if c.type_priseencharge else ''

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        if request.user.is_pharm:
            #total_frais_reel = c.total_frais_reel_medicament
            #total_part_compagnie = c.total_part_compagnie_medicament
            #total_part_assure = c.total_part_assure_medicament

            total_frais_reel = c.total_frais_reel_medicament
            total_part_compagnie = c.new_total_part_compagnie_medicament_prestataire
            total_part_assure = c.new_total_part_assure_medicament_prestataire
        elif request.user.is_prestataire:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_prestataire
            total_part_assure = c.new_total_part_assure_prestataire
        else:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_gestionnaire
            total_part_assure = c.new_total_part_assure_gestionnaire

        dossier_sinistre_statut = c.statut
        dossier_sinistre_statut_prorogation = c.statut_prorogation
        statut_html = f'<span class="badge badge-{c.statut.lower().replace(" ", "-")}">{dossier_sinistre_statut}</span>'
        # statut_html = f'<span class="badge badge-{c.statut_pec.lower().replace(" ","-")}">{c.statut_pec}</span>'

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        centre_prescripteur = c.centre_prescripteur.name if c.centre_prescripteur else ""
        nom_prestataire = c.prestataire.name if c.prestataire else ""
        nom_pharmacie = c.pharmacie.name if c.pharmacie else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_or_numero_carte": type_or_numero_carte,
            "nom": c.aliment.nom + ' ' + c.aliment.prenoms,
            "numero_carte": numero_carte,
            "centre_prescripteur": centre_prescripteur,
            "prestataire": nom_pharmacie if (request.user.is_med and c.type_priseencharge.code == "CONSULT") else nom_prestataire,
            "pharmacie": nom_pharmacie if nom_pharmacie else nom_prestataire,
            "total_frais_reel": money_field(total_frais_reel),
            "total_part_compagnie": money_field(total_part_compagnie),
            "total_part_assure": money_field(total_part_assure),
            "date_prestation": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else "",
            "statut": statut_html,
            "actions": actions_html,
        }

        if request.user.is_med:
            statut_prorogation_html = f'<span class="badge badge-{c.statut_prorogation.replace(" ", "-").lower()}">{c.statut_prorogation}</span>' if c.statut_prorogation else ""
            data_iten["statut_prorogation"] = statut_prorogation_html


        # if request.user.is_med:
        #    if dossier_sinistre_statut != "ACCORDE":
        #        data.append(data_iten)
        # else:
        #    data.append(data_iten)

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
class EntentesPrealablesView(TemplateView):
    template_name = 'liste_ententes_prealables.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        prestataires = Prestataire.objects.filter(
            bureau=request.user.bureau,
            status=True,
            id__in=DossierSinistre.objects.filter(
                statut_validite="VALIDE"
            ).values('prestataire_id')
        )

        today = timezone.now().date()
        context['today'] = today
        context['prestataires'] = prestataires
        context['breadcrumbs'] = [
            {'title': 'Prises en charges', 'url': ''},
            {'title': 'Traités', 'url': ''},
        ]

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def ententes_prealables_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_assure = request.GET.get('num_assure', '')
    search_numero_dossier_sinistre = request.GET.get('num_feuille_soins', '')
    search_date_survenance = request.GET.get('date_prestation', '')
    search_prestataire = request.GET.get('prestataire', '')
    date_reception_facture = request.GET.get('date_reception_facture', '')
    reference_facture = request.GET.get('reference_facture', '')
    statut_pec = request.GET.get('statut_pec', '')

    pprint("search_numero_assure")
    pprint(search_numero_assure)
    pprint("search_numero_dossier_sinistre")
    pprint(search_numero_dossier_sinistre)
    pprint("search_date_survenance")
    pprint(search_date_survenance)

    pprint("search_prestataire")
    pprint(search_prestataire)

    pprint("date_reception_facture")
    pprint(date_reception_facture)

    pprint("reference_facture")
    pprint(reference_facture)

    pprint("statut_pec")
    pprint(statut_pec)

    # SOLUTION POUR OPTIMISER: Créer un champ statut_prorogation qui sera mis à jour à chaque qu'il yy a une prorpogation sur un sinistre
    # et un champ statut_entente qui est mis à jour quand les sinistres en attente existe sur le dossier
    if (search_numero_assure or search_numero_dossier_sinistre or search_date_survenance or search_prestataire or date_reception_facture or reference_facture or statut_pec):
        queryset = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE,
                                                  bureau=request.user.bureau).filter(
            Q(statut_pec=StatutSinistre.ATTENTE) | Q(statut_prorogation=StatutSinistre.ATTENTE)).order_by('id')

    else:
        queryset = DossierSinistre.objects.none()

    # pprint(queryset)

    # la recherche
    if search_numero_assure:
        cartes = Carte.objects.filter(numero__contains=search_numero_assure)
        carte = cartes.first() if cartes else None
        aliment = carte.aliment if carte else None
        queryset = queryset.filter(aliment_id=aliment.pk) if aliment else queryset.filter(numero="nexisterajamais")

    if search_numero_dossier_sinistre:
        queryset = queryset.filter(numero=search_numero_dossier_sinistre)

    if search_date_survenance:
        queryset = queryset.filter(date_survenance__contains=search_date_survenance)

    if search_prestataire:
        queryset = queryset.filter(prestataire=search_prestataire)

    if statut_pec:
        queryset = queryset.filter(Q(statut_pec=statut_pec) | Q(statut_prorogation=statut_pec))

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'aliment__nom',
        2: 'statut',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    # filter les pec dont le statut est EN ATTENTE
    queryset_without_accord = [x.id for x in queryset if
                               x.statut == "EN ATTENTE" or x.statut_prorogation == "EN ATTENTE"]
    queryset = queryset.filter(id__in=queryset_without_accord)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_dossier_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'

        if request.user.is_med and c.type_priseencharge.code == "CONSULT":
            type_or_numero_carte = "PHARMACIE"
        else:
            type_or_numero_carte = c.type_priseencharge.libelle if c.type_priseencharge else ''

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        total_frais_reel = c.total_frais_reel + c.total_frais_reel_medicament
        total_part_compagnie = c.total_part_compagnie + c.total_part_compagnie_medicament
        total_part_assure = c.total_part_assure + c.total_part_assure_medicament

        dossier_sinistre_statut = c.statut
        dossier_sinistre_statut_prorogation = c.statut_prorogation
        statut_html = f'<span class="badge badge-{c.statut.lower().replace(" ", "-")}">{dossier_sinistre_statut}</span>'
        # statut_html = f'<span class="badge badge-{c.statut_pec.lower().replace(" ","-")}">{c.statut_pec}</span>'

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        centre_prescripteur = c.centre_prescripteur.name if c.centre_prescripteur else ""
        nom_prestataire = c.prestataire.name if c.prestataire else ""
        nom_pharmacie = c.pharmacie.name if c.pharmacie else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_or_numero_carte": type_or_numero_carte,
            "nom": c.aliment.nom + ' ' + c.aliment.prenoms,
            "numero_carte": numero_carte,
            "centre_prescripteur": centre_prescripteur,
            "prestataire": nom_pharmacie if (
                        request.user.is_med and c.type_priseencharge.code == "CONSULT") else nom_prestataire,
            "pharmacie": nom_pharmacie if nom_pharmacie else nom_prestataire,
            "total_frais_reel": money_field(total_frais_reel),
            "total_part_compagnie": money_field(total_part_compagnie),
            "total_part_assure": money_field(total_part_assure),
            "date_prestation": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else "",
            "statut": statut_html,
            "actions": actions_html,
        }

        if request.user.is_med:
            statut_prorogation_html = f'<span class="badge badge-{c.statut_prorogation.replace(" ", "-").lower()}">{c.statut_prorogation}</span>' if c.statut_prorogation else ""
            data_iten["statut_prorogation"] = statut_prorogation_html

        # if request.user.is_med:
        #    if dossier_sinistre_statut != "ACCORDE":
        #        data.append(data_iten)
        # else:
        #    data.append(data_iten)

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
class DossierSinistresTraitesView(TemplateView):
    template_name = 'liste_dossiers_traites.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        prestataires = Prestataire.objects.filter(
            bureau=request.user.bureau,
            status=True,
            id__in=DossierSinistre.objects.filter(
                statut_validite="VALIDE"
            ).values('prestataire_id')
        )

        today = timezone.now().date()
        context['today'] = today
        context['prestataires'] = prestataires
        context['breadcrumbs'] = [
            {'title': 'Prises en charges', 'url': ''},
            {'title': 'Traités', 'url': ''},
        ]

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def dossiersinistre_traites_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_assure = request.GET.get('num_assure', '')
    search_numero_dossier_sinistre = request.GET.get('num_feuille_soins', '')
    search_date_survenance = request.GET.get('date_prestation', '')
    search_prestataire = request.GET.get('prestataire', '')

    today = datetime.datetime.now(tz=timezone.utc)
    yesterday = datetime.datetime.now(tz=timezone.utc) - timedelta(days=3)
    queryset = DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE, bureau=request.user.bureau, has_sinistre_traite_bymedecin=True, date_traitement_sinistre_bymedecin__date__gte=yesterday).order_by('-id')
    # dd(queryset)

    # la recherche
    if search_numero_assure:
        cartes = Carte.objects.filter(numero=search_numero_assure)
        carte = cartes.first() if cartes else None
        aliment = carte.aliment if carte else None
        queryset = queryset.filter(aliment_id=aliment.pk) if aliment else queryset.filter(numero="nexisterajamais")

    if search_numero_dossier_sinistre:
        queryset = queryset.filter(numero__contains=search_numero_dossier_sinistre)

    if search_date_survenance:
        queryset = queryset.filter(date_survenance__contains=search_date_survenance)

    if search_prestataire:
        queryset = queryset.filter(prestataire=search_prestataire)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'aliment__nom',
        2: 'statut',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_dossier_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'

        if request.user.is_pharm:
            type_or_numero_carte = aliment.carte_active if aliment else ''
        else:
            if request.user.is_med and c.type_priseencharge.code == "CONSULT":
                type_or_numero_carte = "PHARMACIE"
            else:
                type_or_numero_carte = c.type_priseencharge.libelle if c.type_priseencharge else ''

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''


        if request.user.is_pharm:
            total_frais_reel = c.total_frais_reel_medicament
            total_part_compagnie = c.total_part_compagnie_medicament
            total_part_assure = c.total_part_assure_medicament
        elif request.user.is_prestataire:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_prestataire
            total_part_assure = c.new_total_part_assure_prestataire
        else:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_gestionnaire
            total_part_assure = c.new_total_part_assure_gestionnaire


        statut_html = f'<span class="badge badge-{c.statut.lower().replace(" ", "-")}">{c.statut}</span>'

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        nom_prestataire = c.prestataire.name if c.prestataire else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_or_numero_carte": type_or_numero_carte,
            "nom": c.aliment.nom + ' ' + c.aliment.prenoms,
            "numero_carte": numero_carte,
            "prestataire": nom_prestataire,
            "total_frais_reel": money_field(total_frais_reel),
            "total_part_compagnie": money_field(total_part_compagnie),
            "total_part_assure": money_field(total_part_assure),
            "date_prestation": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else "",
            "date_traitement_bymedecin": c.date_traitement_sinistre_bymedecin.strftime("%d/%m/%Y %H:%M") if c.date_traitement_sinistre_bymedecin else "",
            "statut": statut_html,
            "actions": actions_html,
        }

        if request.user.is_med:
            statut_prorogation_html = f'<span class="badge badge-{c.statut_prorogation.replace(" ", "-").lower()}">{c.statut_prorogation}</span>' if c.statut_prorogation else ""
            data_iten["statut_prorogation"] = statut_prorogation_html

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# GESTION DES SAISIES PAR LE GESTIONNAIRE SINISTRE


@method_decorator(login_required, name='dispatch')
class DossiersSinistresPhysiquesGestionnairesView(TemplateView):
    template_name = 'liste_dossiers_sinistres_physiques_gestionnaires.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        sinistres = []
        prestataires = Prestataire.objects.filter(bureau=request.user.bureau, status=True).exclude(
            type_prestataire__code="PRES02").exclude(type_prestataire__code__isnull=True)  # exclure les pharmacies
        centres_prescripteurs = Prestataire.objects.filter(type_prestataire__code="PRES01", bureau=request.user.bureau, status=True)
        prescripteurs = [p for p in Prescripteur.objects.filter(statut=True) if
                         p.prescripteurprestataire_set.filter(prestataire=request.user.prestataire).exists()]
        rubriques = Rubrique.objects.all()
        # dossiers_sinistres = [x for x in DossierSinistre.objects.all().order_by('-id') if x.sinistres.filter(statut=StatutSinistre.ATTENTE).exists()]

        # context['bureaux'] = bureaux
        context['sinistres'] = sinistres
        context['prestataires'] = prestataires
        context['centres_prescripteurs'] = centres_prescripteurs
        context['prescripteurs'] = prescripteurs
        context['rubriques'] = rubriques
        context['affections'] = Affection.objects.filter(status=True)
        context['types_remboursements'] = TypeRemboursement.objects.filter(status=True)

        context['yesterday'] = datetime.datetime.now(tz=timezone.utc) - datetime.timedelta(days=1)
        context['today'] = datetime.datetime.now(tz=timezone.utc)

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = [
            {'title': 'Prises en charges', 'url': ''},
            {'title': 'Traités', 'url': ''},
        ]

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def dossiersinistre_physique_gestionnaire_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_assure = request.GET.get('num_assure', '')
    search_numero_dossier_sinistre = request.GET.get('num_feuille_soins', '')
    search_date_survenance = request.GET.get('date_prestation', '')
    prestataire = request.GET.get('prestataire', '')

    pprint("search_numero_assure")
    pprint(search_numero_assure)

    pprint("search_numero_dossier_sinistre")
    pprint(search_numero_dossier_sinistre)

    pprint("search_date_survenance")
    pprint(search_date_survenance)

    queryset = DossierSinistre.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE, of_gestionnaire=1).order_by('id')
    # dd(queryset)

    if prestataire:
         queryset = queryset.filter(prestataire_id=prestataire)

    if search_numero_assure:
        cartes = Carte.objects.filter(numero=search_numero_assure)
        carte = cartes.first() if cartes else None
        aliment = carte.aliment if carte else None
        queryset = queryset.filter(aliment_id=aliment.pk) if aliment else queryset.filter(numero="nexisterajamais")

    if search_numero_dossier_sinistre:
        queryset = queryset.filter(numero__contains=search_numero_dossier_sinistre)
    if search_date_survenance:
        queryset = queryset.filter(created_at__contains=search_date_survenance)



    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'aliment__nom',
        2: 'statut',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_dossier_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'

        if request.user.is_pharm:
            type_or_numero_carte = aliment.carte_active if aliment else ''
        else:
            if request.user.is_med and c.type_priseencharge.code == "CONSULT":
                type_or_numero_carte = "PHARMACIE"
            else:
                type_or_numero_carte = c.type_priseencharge.libelle if c.type_priseencharge else ''

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''


        if request.user.is_pharm:
            total_frais_reel = c.total_frais_reel_medicament
            total_part_compagnie = c.total_part_compagnie_medicament
            total_part_assure = c.total_part_assure_medicament
        elif request.user.is_prestataire:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_prestataire
            total_part_assure = c.new_total_part_assure_prestataire
        else:
            total_frais_reel = c.new_total_frais_reel
            total_part_compagnie = c.new_total_part_compagnie_gestionnaire
            total_part_assure = c.new_total_part_assure_gestionnaire


        statut_html = f'<span class="badge badge-{c.statut.lower()}">{c.statut}</span>'

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_or_numero_carte": type_or_numero_carte,
            "nom": c.aliment.nom + ' ' + c.aliment.prenoms,
            "total_frais_reel": money_field(total_frais_reel),
            "total_part_compagnie": money_field(total_part_compagnie),
            "total_part_assure": money_field(total_part_assure),
            "date_prestation": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "statut": statut_html,
            "actions": actions_html,
        }

        if request.user.is_med:
            statut_prorogation_html = f'<span class="badge badge-{c.statut_prorogation.replace(" ", "-").lower()}">{c.statut_prorogation}</span>' if c.statut_prorogation else ''
            data_iten["statut_prorogation"] = statut_prorogation_html

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


#Liste des sinistres annulés
@method_decorator(login_required, name='dispatch')
class AnnulerSinistreGestionnairesView(TemplateView):
    template_name = 'annuler_sinistre.html'
    model = Sinistre

    #traitement à l'appel du lien en get
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        context['breadcrumbs'] = [
            {'title': 'Sinistres', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        return self.render_to_response(context)

    #traitement à l'appel du lien en post pour la recherche de dossier et la suppresion de dossier ou sinistre
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        #recuperation de tout ce qui peut venir en post que ca soit pour la recherche ou la suppression
        btn_recherche = self.request.POST.get('recherche', None)
        submit_delete_item = self.request.POST.get('submit_delete_item', None)
        type_item = self.request.POST.get('type_item', None)
        id_item = self.request.POST.get('id_item', None)
        motif_delete_item = self.request.POST.get('motif_delete_item', None)
        code_dossier_sinistre = self.request.POST.get('code_dossier_sinistre', None)
        context['breadcrumbs'] = [
            {'title': 'Sinistres', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        dossier_sinistre = None

        #cette condition précise que nous venons faire la recherche
        if btn_recherche and code_dossier_sinistre:
            print(code_dossier_sinistre)
            dossier_sinistre = DossierSinistre.objects.filter(numero=code_dossier_sinistre, bureau=request.user.bureau).exclude(statut_validite=StatutValidite.SUPPRIME).first()
            # dd(dossier_sinistre)if 
            #si on a trouver le dossier on récupère infos liées y compris les sinistres qui le composent
            if dossier_sinistre:
                sinistres = Sinistre.objects.filter(dossier_sinistre=dossier_sinistre).exclude(statut_validite=StatutValidite.SUPPRIME)
                total_frais_reel = dossier_sinistre.total_frais_reel + dossier_sinistre.total_frais_reel_medicament
                total_part_compagnie = dossier_sinistre.total_part_compagnie + dossier_sinistre.total_part_compagnie_medicament
                total_part_assure = dossier_sinistre.total_part_assure + dossier_sinistre.total_part_assure_medicament
                cartes = dossier_sinistre.aliment.cartes.filter(statut=Statut.ACTIF) if dossier_sinistre.aliment else None
                numero_carte = cartes.first().numero if cartes else None
                has_sinistre_on_facture = sinistres.filter(facture_prestataire__isnull=False).exists()
                context['sinistres'] = sinistres
                context['total_frais_reel'] = total_frais_reel
                context['total_part_compagnie'] = total_part_compagnie
                context['total_part_assure'] = total_part_assure
                context['numero_carte'] = numero_carte
                context['has_sinistre_on_facture'] = sinistres.filter(facture_prestataire__isnull=False).exists()

        #cette condition précise que nous venons faire la suppression de soit un dossier ou un sinistre
        if submit_delete_item and type_item and id_item:
            #dd(self.request.POST)
            if type_item == "dossier": #il s'agit de la suppresion d'un et les sinistres qui le composent
                dossier_sinistre = DossierSinistre.objects.filter(id=id_item, bureau=request.user.bureau).exclude(statut_validite=StatutValidite.SUPPRIME).first()
                if dossier_sinistre: #vérification selon données postées et suppression du dossier
                    sinistres = Sinistre.objects.filter(dossier_sinistre=dossier_sinistre)
                    code_dossier_sinistre = dossier_sinistre.numero
                    has_sinistre_on_facture = sinistres.filter(facture_prestataire__isnull=False).exists()

                    if has_sinistre_on_facture is False: #assure qu'un dossier est entiermeent annulé que si aucun de ses sinistres n'est sur une facture prestataire
                        dossier_sinistre.statut_validite = StatutValidite.SUPPRIME
                        dossier_sinistre.is_closed = True
                        #dd(dossier_sinistre)
                        dossier_sinistre.save()

                    for sinistre in sinistres: #suppression des sinistres qui composent le dossier
                        if sinistre.facture_prestataire is None: #assure que le sinistre n'est pas sur une facture prestataire
                            sinistre.statut_validite = StatutValidite.SUPPRIME
                            sinistre.statut = StatutValidite.SUPPRIME
                            sinistre.motif_suppression = motif_delete_item
                            sinistre.save()
            elif type_item == "selection_sinistre": #il s'agit de la suppresion d'une selection de sinistres du dossier recherché
                selected_sinistres = Sinistre.objects.filter(id__in=id_item.split(",")).exclude(statut_validite=StatutValidite.SUPPRIME, statut=StatutValidite.SUPPRIME)
                #dd(selected_sinistres.count())
                for sinistre in selected_sinistres:
                    if sinistre.facture_prestataire is None: #assure que le sinistre n'est pas sur une facture prestataire
                        sinistre.statut_validite = StatutValidite.SUPPRIME
                        sinistre.statut = StatutValidite.SUPPRIME
                        sinistre.motif_suppression = motif_delete_item
                        sinistre.save()
                        dossier_sinistre = sinistre.dossier_sinistre
                        code_dossier_sinistre = dossier_sinistre.numero
            else: #il s'agit de la suppresion d'un sinistre du dossier recherché
                sinistre = Sinistre.objects.filter(id=id_item).exclude(statut_validite=StatutValidite.SUPPRIME, statut=StatutValidite.SUPPRIME).first()
                if sinistre and sinistre.facture_prestataire is None: #vérification selon données postées et suppression
                    sinistre.statut_validite = StatutValidite.SUPPRIME
                    sinistre.statut = StatutValidite.SUPPRIME
                    sinistre.motif_suppression = motif_delete_item
                    sinistre.save()
                    dossier_sinistre = sinistre.dossier_sinistre
                    code_dossier_sinistre = dossier_sinistre.numero


        context['code_dossier_sinistre'] = code_dossier_sinistre
        context['dossier_sinistre'] = dossier_sinistre
            #print(code_dossier_sinistre)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required()
# modiifie le statut de verrouillage d'un dossier sinistre depuis le toggle sur la page d' annulation du dossier 
def change_dossier_closing_status(request, dossier_sinistre_id):

    response = None

    if request.method == 'POST':

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)

        if dossier_sinistre.is_closed == True:
            dossier_sinistre.is_closed = False

        else:
            dossier_sinistre.is_closed = True
        dossier_sinistre.save()
        #gardons des traces
        if dossier_sinistre.is_closed == False:
            ActionLog.objects.create(done_by=request.user, action="update",
                                     description="Déverrouillage d'un dossier sinistre", table="dossier_sinistre",
                                     row=dossier_sinistre.pk,
                                     )
        else:
            ActionLog.objects.create(done_by=request.user, action="update",
                                     description="Déverrouillage d'un dossier sinistre", table="dossier_sinistre",
                                     row=dossier_sinistre.pk,
                                     )

        response = {
            'statut': 1,
            'message': "Statut verrouillage dossier sinistre changé avec succès !",
            'data': {
            }
        }
        print(dossier_sinistre.is_closed)



    return JsonResponse(response)


#Liste des sinistres annulés
@method_decorator(login_required, name='dispatch')
class AnnulerFactureGestionnairesView(TemplateView):
    template_name = 'annuler_facture.html'
    model = Sinistre

    #traitement à l'appel du lien en get
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        return self.render_to_response(context)

    #traitement à l'appel du lien en post pour la recherche de dossier et la suppresion de dossier ou sinistre
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        #recuperation de tout ce qui peut venir en post que ca soit pour la recherche ou la suppression
        btn_recherche = self.request.POST.get('recherche', None)
        submit_delete_item = self.request.POST.get('submit_delete_item', None)
        id_item = self.request.POST.get('id_item', None)
        motif_delete_item = self.request.POST.get('motif_delete_item', None)
        code_bordereau = self.request.POST.get('code_bordereau', None)
        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        dossier_sinistre = None

        #cette condition précise que nous venons faire la recherche
        if btn_recherche and code_bordereau:
            facture = FacturePrestataire.objects.filter(numero=code_bordereau, bureau=request.user.bureau,
                        statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).first()

            beneficiaire_de_reglement = None

            #si on a trouver la facture dossier on récupère infos liées
            if facture:
                if facture.type_remboursement == TypeRemboursement.objects.get(code="TP"):
                    beneficiaire_de_reglement = facture.prestataire.name if facture.prestataire else ""
                elif facture.type_remboursement == TypeRemboursement.objects.get(code="RD") :
                    beneficiaire_de_reglement = (facture.adherent_principal.nom + " " + facture.adherent_principal.prenoms) if facture.adherent_principal else ""
            context['beneficiaire_de_reglement'] = beneficiaire_de_reglement
            context['code_bordereau'] = code_bordereau
            context['facture'] = facture

        #cette condition précise que nous venons faire l'annulation de la facture
        if submit_delete_item and id_item:

            facture = FacturePrestataire.objects.filter(id=id_item, bureau=request.user.bureau,
                        statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).first()

            if facture:
                # traitement facture
                facture.fp_deleted_by = request.user
                facture.statut = StatutValidite.SUPPRIME
                facture.observation = motif_delete_item
                facture.save()
                context['old_facture'] = facture.numero

                #

                # récuperation sinistres associés
                sinistres = Sinistre.objects.filter(facture_prestataire=facture)
                if sinistres:
                    for sinistre in sinistres:
                        sinistre.is_ges_processed = False
                        sinistre.statut_bordereau = StatutSinistreBordereau.ATTENTE
                        sinistre.observation = str(sinistre.observation) + "- old_facture_id=" + facture.numero if sinistre.observation else "old_facture_id=" + facture.numero
                        sinistre.facture_prestataire = None
                        sinistre.save()

                        #TODO : corriger récupprer les remboursement non encore annulés
                        # récuperation de remboursement_sinistre associé au sinistre
                        remboursement_sinistres = RemboursementSinistre.objects.filter(sinistre=sinistre, is_invalid=False)
                        for remboursement_sinistre in remboursement_sinistres:
                            remboursement_sinistre.is_invalid = True
                            remboursement_sinistre.is_invalid_by = request.user
                            remboursement_sinistre.observation = motif_delete_item
                            remboursement_sinistre.save()

                #code_bordereau = facture.numero


                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="suppression_facture_prestataire",
                                         description="Annulation d'une facture prestataire pour remettre les sinistre à l'étape de la génération de facture pour reprise du traitement par les gestionnaires sinistres", table="factures_prestataires",
                                         row=facture.pk)


            #print(code_dossier_sinistre)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@method_decorator(login_required, name='dispatch')
class AnnulerBordereauOrdonnancementView(TemplateView):
    template_name = 'annuler_bordereau_ordonnancement.html'
    model = BordereauOrdonnancement

    # traitement à l'appel du lien en get
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        return self.render_to_response(context)



    # traitement à l'appel du lien en post pour la recherche de dossier et la suppresion de dossier ou sinistre
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        # recuperation de tout ce qui peut venir en post que ca soit pour la recherche ou la suppression
        btn_recherche = self.request.POST.get('recherche', None)
        submit_delete_item = self.request.POST.get('submit_delete_item', None)
        id_item = self.request.POST.get('id_item', None)
        motif_delete_item = self.request.POST.get('motif_delete_item', None)
        code_bordereau = self.request.POST.get('code_bordereau', None)
        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        dossier_sinistre = None

        # cette condition précise que nous venons faire la recherche
        if btn_recherche and code_bordereau:
            bordereau_ordonnancement = BordereauOrdonnancement.objects.filter(numero=code_bordereau, bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE).first()

            context['code_bordereau'] = code_bordereau
            context['bordereau_ordonnancement'] = bordereau_ordonnancement

        # cette condition précise que nous venons faire l'annulation de la facture
        if submit_delete_item and id_item:

            bordereau_ordonnancement = BordereauOrdonnancement.objects.filter(id=id_item, bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE).first()

            if bordereau_ordonnancement:
                # traitement facture
                bordereau_ordonnancement.bo_deleted_by = request.user
                bordereau_ordonnancement.statut_paiement = StatutPaiementSinistre.ATTENTE
                bordereau_ordonnancement.statut = StatutValidite.SUPPRIME
                bordereau_ordonnancement.observation = motif_delete_item
                bordereau_ordonnancement.save()
                context['old_facture'] = bordereau_ordonnancement.numero

                # récuperation sinistres associés
                sinistres = Sinistre.objects.filter(bordereau_ordonnancement=bordereau_ordonnancement)
                if sinistres:
                    for sinistre in sinistres:
                        sinistre.bordereau_ordonnancement = None
                        sinistre.statut_paiement = StatutPaiementSinistre.ATTENTE
                        #sinistre.observation = str(sinistre.observation)
                        sinistre.save()

                        # mettre la facture à traitée:: faire sortir de la boucle lorsque les factures seront directement liées aux bordereau d'ordonnancement
                        sinistre.facture_prestataire.statut=SatutBordereauDossierSinistres.VALIDE
                        sinistre.facture_prestataire.save()

                        #historiser les lignes qui étaient sur le bordereau
                        HistoriqueOrdonnancementSinistre.objects.create(created_by=request.user, bordereau_ordonnancement=bordereau_ordonnancement, sinistre=sinistre, montant_ordonnance=sinistre.montant_remb_accepte, observation=motif_delete_item)


                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="annulation_bordereau_ordonnancement",
                                         description="Annulation d'un bordereau d'ordonnancement",
                                         table="bordereau_ordonnancement",
                                         row=bordereau_ordonnancement.pk)

            # print(code_dossier_sinistre)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



# model 2
@method_decorator(login_required, name='dispatch')
class SaisiePrestationGestionnairesView(TemplateView):
    template_name = 'form_saisie_prestation_gestionnaires.html'
    model = Sinistre

    def get(self, request, prestataire_id=None, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        prestataires_executants = Prestataire.objects.filter(id=prestataire_id)
        prestataire_executant = prestataires_executants.first() if prestataires_executants else None

        prestataires = Prestataire.objects.filter(bureau=request.user.bureau, status=True).exclude(
            type_prestataire__code="PRES02").exclude(type_prestataire__code__isnull=True)  # exclure les pharmacies
        centres_prescripteurs = Prestataire.objects.filter(bureau=request.user.bureau, type_prestataire__code__in=["PRES01", "PRES04"], status=True)
        # prescripteurs = [p for p in Prescripteur.objects.all() if p.prescripteurprestataire_set.filter(prestataire=request.user.prestataire).exists()]
        #prescripteurs = [p for p in Prescripteur.objects.all() if p.prescripteurprestataire_set.filter(prestataire=prestataire_executant).exists()]
        pps = PrescripteurPrestataire.objects.filter(prestataire=prestataire_executant, statut_validite=StatutValidite.VALIDE)
        prescripteurs = Prescripteur.objects.filter(id__in=[pp.prescripteur_id for pp in pps]).order_by('nom')
        # prescripteurs = [pp.prescripteur for pp in pps]


        rubriques = Rubrique.objects.filter()
        types_priseencharges = TypePriseencharge.objects.filter(statut_selectable=True)
        # si centre optique, charger optique uniquement, ainsi de suite
        # dd(prestataire_executant.type_prestataire)
        if prestataire_executant:
            if prestataire_executant.type_prestataire.code == "PRES01":
                types_priseencharges = TypePriseencharge.objects.filter(statut_selectable=True).exclude(code='OPTIQUE')

            elif prestataire_executant.type_prestataire.code == "PRES03":
                types_priseencharges = TypePriseencharge.objects.filter(statut_selectable=True, code='OPTIQUE')

            else:
                types_priseencharges = (TypePriseencharge.objects.filter(statut_selectable=True)
                                        .exclude(code='OPTIQUE')
                                        .exclude(code='HOSPIT'))

        context['types_priseencharges'] = types_priseencharges
        context['rubriques'] = rubriques

        context['prestataire_executant'] = prestataire_executant
        context['prestataires'] = prestataires
        context['centres_prescripteurs'] = centres_prescripteurs
        context['prescripteurs'] = prescripteurs
        context['affections'] = Affection.objects.filter(status=True)
        context['types_remboursements'] = TypeRemboursement.objects.filter(status=True)

        context['yesterday'] = timezone.now().date()- datetime.timedelta(days=1)
        context['today'] =timezone.now().date()

        today = timezone.now().date()
        context['today'] = today
        context['reference_facture_origin'] = request.GET.get('reference_facture', "")
        context['date_reception_facture_origin'] = request.GET.get('date_reception_facture', None)
        context['breadcrumbs'] = [
            {'title': 'Prises en charges', 'url': ''},
            {'title': 'Traités', 'url': ''},
        ]

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# liste des médicaments déjà en session:
def popup_add_medicament_gestionnaire(request, dossier_sinistre_id):
    request.session['liste_medicaments2'] = []
    liste_medicaments_in_session = request.session.get('liste_medicaments2', [])

    dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)

    medicaments = Acte.objects.filter(type_acte_id=2, status=True, statut_validite=StatutValidite.VALIDE)
    pharmacies = Prestataire.objects.filter(bureau=request.user.bureau, type_prestataire__code="PRES02", status=True)

    today = datetime.datetime.now(tz=timezone.utc)

    return render(request, 'modal_add_medicament_gestionnaire.html',
                  {'pharmacies': pharmacies, 'liste_medicaments_in_session': liste_medicaments_in_session,
                   'medicaments': medicaments, 'dossier_sinistre': dossier_sinistre, 'today': today})


@login_required()
def add_medicament_gestionnaire_todossiersinistre(request, dossier_sinistre_id):
    if request.method == 'POST':

        pprint("request.POST")
        pprint(request.POST)

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)

        if dossier_sinistre:

            liste_medicaments = request.session.get('liste_medicaments2', [])

            # try:
            # vider toute la session de calcul de prise en charge pour reprendre
            session_pec = request.user.id
            vider_sinistres_temporaires(session_pec)

            pharmacie_id = request.POST.get('pharmacie')
            prescripteur_id = request.POST.get('prescripteur_id')
            medicament_id = request.POST.get('medicament_id')
            qte = request.POST.get('quantite').replace(" ", "")
            prix_unitaire = request.POST.get('prix_unitaire').replace(" ", "")

            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            aliment = dossier_sinistre.aliment
            centre_prescripteur_id = dossier_sinistre.prestataire_id

            medicament = Acte.objects.get(id=medicament_id)

            if prix_unitaire == '': prix_unitaire = 0
            prix_unitaire = int(prix_unitaire)
            frais_reel = int(
                qte) * prix_unitaire  # comment on 22112023: déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé,
            # frais_reel = prix_unitaire #Added on 22112023: faire entrer pu et la quantité séparément

            '''
            nouvelle version: actes et medicaments fusionnés
            '''
            # récupérer ses consommations individuel et par famille
            formule = aliment.formule_atdate(date_survenance)
            periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

            consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, adherent_principal_id=aliment.adherent_principal.id,
                                                           statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            type_prise_en_charge_code = "PHARM"
            acte_id = medicament.id
            prescripteur_id = None  # pour ne pas qu'il vérifie si prescripteur autoriser à faire de la spécialité
            cout_acte = frais_reel
            nombre_jours = 0  # comment on 22112023: frais_reel = qte* pu (déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé)
            # nombre_jours = qte #comment on 22112023: frais_reel = qte* pu (déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé)

            # faire passer par l'algo de tarification
            infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id, pharmacie_id,
                                                    prescripteur_id, aliment.id, cout_acte, nombre_jours,
                                                    consommation_individuelle, consommation_famille,
                                                    session_pec)

            # pprint("---/__/---")
            pprint("infos_acte")
            pprint(infos_acte)
            if infos_acte['statut'] == 1:

                acte_libelle = infos_acte['data']['libelle']
                frais_reel = infos_acte['data']['frais_reel']
                part_compagnie = infos_acte['data']['part_compagnie']
                part_assure = infos_acte['data']['part_assure']
                ticket_moderateur = infos_acte['data']['ticket_moderateur']
                taux_couverture = infos_acte['data']['taux_couverture']
                depassement = infos_acte['data']['depassement']

                garanti = infos_acte['data']['garanti']
                bareme_id = infos_acte['data']['bareme_id']
                taux_tm = infos_acte['data']['taux_franchise']
                prestataire_id = infos_acte['data']['prestataire_id']
                prescripteur_id = infos_acte['data']['prescripteur_id']
                aliment_id = infos_acte['data']['aliment_id']

                statut = "ACCORDE"
                reviewed_at = datetime.datetime.now(tz=timezone.utc)
                approuved_by = None

                remove_medicament_url = reverse('remove_medicament_session_gestionnaire',
                                                args=[acte_id])  # URL to the detail view
                actions_html = f'<span style="cursor:pointer;color:red;" class="btn btn-sm btn-default btnSupprimerLigneMedicamentGestionnaire" data-href="{remove_medicament_url}"><i class="fa fa-times" style="font-size:12px;"></i> Supprimer</span>'

                # Renseigner une table de session pour conserver les médicaments ajouté au sinistre en cours de création
                # Ajouter le nouvel élément à la liste
                medicament_session = {'dossier_sinistre_id': dossier_sinistre_id,
                                      'pharmacie_id': pharmacie_id,
                                      'medicament_id': medicament.id,
                                      'libelle': acte_libelle,
                                      'qte_demande': qte,
                                      'qte_servie': qte,
                                      'prix_unitaire': float(prix_unitaire),
                                      'prix_total': float(frais_reel),
                                      'taux_couverture': float(taux_couverture),
                                      'part_assureur': float(part_compagnie),
                                      'depassement': float(depassement),
                                      'part_assure': float(part_assure),
                                      'statut': statut,
                                      'actions_html': actions_html
                                      }

                liste_medicaments.append(medicament_session)

                # Mettre à jour les données de session
                request.session['liste_medicaments2'] = liste_medicaments
                pprint("liste_medicaments2")
                # pprint(liste_medicaments)

                response = {
                    "statut": 1,
                    "message": infos_acte['message'],
                    "data": medicament_session
                }

            else:
                response = {
                    "statut": 0,
                    "message": infos_acte['message'],
                    "data": {}
                }

        else:
            response = {
                "statut": 0,
                "message": "Bon de soins non trouvé",
                "data": {}
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

    '''except:
        print("EXCEPT RUNNING ")
        response = {
            "statut": 0,
            "message": "Erreur survenu lors de l'ajout du medicament",
            "data": {}
        }'''

    return JsonResponse(response)


def remove_medicament_gestionnaire(request, medicament_id):
    if request.method == 'POST':

        liste_medicaments = request.session.get('liste_medicaments2', [])

        pprint(medicament_id)
        # Utiliser une boucle ou une compréhension de liste pour supprimer l'élément avec l'acte_id correspondant
        liste_medicaments = [item for item in liste_medicaments if item['medicament_id'] != medicament_id]
        pprint(liste_medicaments)

        # Mettre à jour les données de session
        request.session['liste_medicaments2'] = liste_medicaments

        response = {
            "statut": 1,
            "message": "Suppression effectuée avec succès"
        }

    else:

        response = {
            "statut": 0,
            "message": "Erreur survenu lors de la suppression du medicament",
            "data": {}
        }

    return JsonResponse(response)


# enregistrer sur le dossier

@login_required()
def add_sinistre_medicament_gestionnaire(request, dossier_sinistre_id):
    if request.method == 'POST':

        # try:
        pprint("request.POST")
        pprint(request.POST)

        date_survenance = request.POST.get('date_survenance')
        date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
        date_survenance = make_aware(date_survenance)

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)
        if dossier_sinistre:

            type_prefinancement = dossier_sinistre.type_prefinancement
            affection_id = dossier_sinistre.affection_id
            aliment = dossier_sinistre.aliment
            aliment_id = aliment.id
            type_prise_en_charge_code = "PHARM"
            formule = aliment.formule_atdate(date_survenance)

            if formule:
                periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

                liste_created_sinistres = []

                # récupérer les médicaments en session sur le dossier_sinistre
                liste_medicaments = request.session.get('liste_medicaments2', [])
                if liste_medicaments:

                    # parcourrir les médicaments en session
                    for med in liste_medicaments:
                        # refaire le calcul, vu qu'on vient d'enregistrer la consultation, la conssommation a augmenter
                        if med['dossier_sinistre_id'] == dossier_sinistre_id:
                            med_acte_id = med['medicament_id']
                            med_prestataire_id = med['pharmacie_id']
                            med_quantite = med['qte_servie']
                            med_prix_unitaire = med['prix_unitaire']
                            med_frais_reel = med['prix_total']

                            cout_acte = med_frais_reel
                            quantite = 0  # juste pour l'algo, on fait pu*qte

                            acte = Acte.objects.get(id=med_acte_id)
                            # Added on 17062024: mise en oeuvre du tpg
                            #A voir avec Marius comment gérer le type pref dans la cas d'ajout de medicament sur un dossier_sinistre existant
                            type_prefinancement = get_type_prefinancement_of_acte(acte, formule)

                            # Nouvelle idée: faire le calcul des consos dans la fonction get_tarif_acte_from_bareme
                            # récupérer ses consommations individuel et par famille
                            consommation_individuelle = Sinistre.objects.filter(
                                periode_couverture_id=periode_couverture_encours.pk,
                                aliment_id=aliment.id,
                                statut=StatutSinistre.ACCORDE,
                                statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT],
                                statut_validite=StatutValidite.VALIDE
                            ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                            consommation_famille = Sinistre.objects.filter(
                                periode_couverture_id=periode_couverture_encours.pk,
                                adherent_principal_id=aliment.adherent_principal.id,
                                statut=StatutSinistre.ACCORDE,
                                statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT],
                                statut_validite=StatutValidite.VALIDE
                            ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                            prescripteur_id = None  # pour ne pas qu'il controle si généraliste peut faire acte spécialiste
                            # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                            infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, med_acte_id,
                                                                    med_prestataire_id, prescripteur_id, aliment_id,
                                                                    cout_acte,
                                                                    quantite,
                                                                    consommation_individuelle, consommation_famille,
                                                                    session_pec)

                            med_frais_reel = infos_acte['data']['frais_reel']
                            med_part_compagnie = infos_acte['data']['part_compagnie']
                            med_part_assure = infos_acte['data']['part_assure']
                            med_ticket_moderateur = infos_acte['data']['ticket_moderateur']
                            med_depassement = infos_acte['data']['depassement']

                            med_plafond_acte = infos_acte['data']['plafond_acte']
                            med_nombre_acte = infos_acte['data']['nombre_acte']
                            med_garanti = infos_acte['data']['garanti']
                            med_bareme_id = infos_acte['data']['bareme_id']
                            med_taux_tm = infos_acte['data']['taux_franchise']

                            med_statut = StatutSinistre.ACCORDE
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)

                            sinistre = Sinistre.objects.create(type_sinistre="medicament",
                                                               created_by=request.user,
                                                               prestataire_id=med_prestataire_id,
                                                               aliment_id=aliment.id,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               police_id=formule.police.id,
                                                               periode_couverture_id=periode_couverture_encours.pk,
                                                               formulegarantie_id=formule.id,
                                                               bareme_id=med_bareme_id,
                                                               compagnie_id=formule.police.compagnie.id,
                                                               prescripteur_id=prescripteur_id,
                                                               affection_id=affection_id,
                                                               acte_id=med_acte_id,
                                                               frais_reel=med_frais_reel,
                                                               part_compagnie=med_part_compagnie,
                                                               part_assure=med_part_assure,
                                                               ticket_moderateur=med_ticket_moderateur,
                                                               depassement=med_depassement,
                                                               taux_tm=med_taux_tm,

                                                               prix_unitaire=med_prix_unitaire,
                                                               nombre_demande=med_quantite,
                                                               nombre_accorde=med_quantite,

                                                               montant_base_remboursement=med_frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else med_part_compagnie,
                                                               montant_remboursement_accepte=med_frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else med_part_compagnie, #sera actualisé au traitement des remboursements sinistre par les gestionnaires,
                                                               montant_remboursement_refuse=0,

                                                               montant_refacture_compagnie=med_part_compagnie, # sera actualisé au traitement des remboursements sinistre par les gestionnaires,
                                                               montant_refacture_client=med_part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0, # sera actualisé au traitement des remboursements sinistre par les gestionnaires,

                                                               taux_retenue = get_retenue_selon_contexte(med_prestataire_id) if dossier_sinistre.type_remboursement.code == "TP" else 0,

                                                               date_survenance=date_survenance,
                                                               statut=med_statut,
                                                               dossier_sinistre_id=dossier_sinistre.pk,
                                                               type_prefinancement=type_prefinancement,
                                                               reviewed_at=reviewed_at
                                                               )

                            sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                            code_bureau = request.user.bureau.code
                            sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                                sinistre_created.pk).zfill(7) + '-SP'
                            sinistre_created.save()

                            liste_created_sinistres.append(sinistre_created)

                    # vider la session de medicaments
                    request.session['liste_medicaments2'] = []

                    response = {
                        "statut": 1,
                        "message": "Enregistrement effectuée avec suuès",
                        "data": {
                            'liste_created_sinistres': serializers.serialize('json', liste_created_sinistres),
                        }
                    }

                else:
                    response = {
                        "statut": 0,
                        "message": "Aucun médicament trouvé",
                        "data": {}
                    }
            else:
                response = {
                    "statut": 0,
                    "message": "Aucune formule de garantie trouvée",
                    "data": {}
                }
        else:
            response = {
                "statut": 0,
                "message": "Dossier sinistre non trouvé",
                "data": {}
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

    else:
        response = {
            "statut": 0,
            "message": "Methode non suportée",
            "data": {}
        }

    '''except:
        print("EXCEPT RUNNING ")
        response = {
            "statut": 0,
            "message": "Erreur survenu lors de l'ajout du medicament",
            "data": {}
        }'''

    return JsonResponse(response)


# liste des médicaments déjà en session:
def popup_add_medicament_session_gestionnaire(request, acte_id, aliment_id, prestataire_id, prescripteur_id):
    liste_medicaments_in_session = request.session.get('liste_medicaments', [])

    medicaments = Acte.objects.filter(type_acte_id=2, status=True, statut_validite=StatutValidite.VALIDE)
    pharmacies = Prestataire.objects.filter(bureau=request.user.bureau, type_prestataire__code="PRES02", status=True, )

    return render(request, 'modal_add_medicament_session_gestionnaire.html',
                  {'pharmacies': pharmacies, 'liste_medicaments_in_session': liste_medicaments_in_session,
                   'medicaments': medicaments, 'acte_id': acte_id, 'aliment_id': aliment_id,
                   'prestataire_id': prestataire_id, 'prescripteur_id': prescripteur_id})


@login_required()
def add_medicament_session_gestionnaire(request):
    if request.method == 'POST':

        liste_medicaments = request.session.get('liste_medicaments', [])

        # try:
        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint("request.POST")
        pprint(request.POST)
        aliment_id = request.POST.get('aliment_id')
        prestataire_id = request.POST.get('prestataire_id')
        pharmacie_id = request.POST.get('pharmacie')
        prescripteur_id = request.POST.get('prescripteur_id')
        medicament_id = request.POST.get('medicament_id')
        qte = request.POST.get('quantite').replace(" ", "")
        prix_unitaire = request.POST.get('prix_unitaire').replace(" ", "")

        date_survenance = request.POST.get('date_survenance')
        date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
        date_survenance = make_aware(date_survenance)

        aliment = Aliment.objects.get(id=aliment_id)
        medicament = Acte.objects.get(id=medicament_id)

        if prix_unitaire == '': prix_unitaire = 0
        prix_unitaire = int(prix_unitaire)
        frais_reel = int(
            qte) * prix_unitaire  # comment on 22112023: déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé,
        # frais_reel = prix_unitaire #Added on 22112023: faire entrer pu et la quantité séparément

        '''
        nouvelle version: actes et medicaments fusionnés
        '''
        # récupérer ses consommations individuel et par famille
        formule = aliment.formule_atdate(date_survenance)

        if formule:
            periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

            consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            type_prise_en_charge_code = "PHARM"
            acte_id = medicament.id
            # prescripteur_id = 1
            cout_acte = frais_reel
            nombre_jours = 0  # comment on 22112023: frais_reel = qte* pu (déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé)
            # nombre_jours = qte #comment on 22112023: frais_reel = qte* pu (déjà multiplié donc la quantité ne doit pas rentrer dans l'algo - sinon le montant sera doublé)

            # faire passer par l'algo de tarification
            infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id, pharmacie_id,
                                                    prescripteur_id, aliment_id, cout_acte, nombre_jours,
                                                    consommation_individuelle, consommation_famille,
                                                    session_pec)

            # pprint("---/__/---")
            pprint("infos_acte")
            pprint(infos_acte)
            if infos_acte['statut'] == 1:

                acte_libelle = infos_acte['data']['libelle']
                frais_reel = infos_acte['data']['frais_reel']
                part_compagnie = infos_acte['data']['part_compagnie']
                part_assure = infos_acte['data']['part_assure']
                ticket_moderateur = infos_acte['data']['ticket_moderateur']
                taux_couverture = infos_acte['data']['taux_couverture']
                depassement = infos_acte['data']['depassement']

                garanti = infos_acte['data']['garanti']
                bareme_id = infos_acte['data']['bareme_id']
                taux_tm = infos_acte['data']['taux_franchise']
                prestataire_id = infos_acte['data']['prestataire_id']
                prescripteur_id = infos_acte['data']['prescripteur_id']
                aliment_id = infos_acte['data']['aliment_id']

                statut = "ACCORDE"
                reviewed_at = datetime.datetime.now(tz=timezone.utc)
                approuved_by = None

                remove_medicament_url = reverse('remove_medicament_session_gestionnaire',
                                                args=[acte_id])  # URL to the detail view
                actions_html = f'<span style="cursor:pointer;color:red;" class="btn btn-sm btn-default btnSupprimerLigneMedicamentGestionnaire" data-href="{remove_medicament_url}"><i class="fa fa-times" style="font-size:12px;"></i> Supprimer</span>'

                # Renseigner une table de session pour conserver les médicaments ajouté au sinistre en cours de création
                # Ajouter le nouvel élément à la liste
                medicament_session = {'pharmacie_id': pharmacie_id,
                                      'medicament_id': medicament.id,
                                      'libelle': acte_libelle,
                                      'qte_demande': qte,
                                      'qte_servie': qte,
                                      'prix_unitaire': float(prix_unitaire),
                                      'prix_total': float(frais_reel),
                                      'taux_couverture': float(taux_couverture),
                                      'part_assureur': float(part_compagnie),
                                      'depassement': float(depassement),
                                      'part_assure': float(part_assure),
                                      'statut': statut,
                                      'actions_html': actions_html
                                      }

                liste_medicaments.append(medicament_session)

                # Mettre à jour les données de session
                request.session['liste_medicaments'] = liste_medicaments
                pprint("liste_medicaments")
                # pprint(liste_medicaments)

                response = {
                    "statut": 1,
                    "message": infos_acte['message'],
                    "data": medicament_session
                }

            else:
                response = {
                    "statut": 0,
                    "message": infos_acte['message'],
                    "data": {}
                }
        else:
            response = {
                "statut": 0,
                "message": "Aucune formule trouvé à cette date",
                "data": {}
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

    '''except:
        print("EXCEPT RUNNING ")
        response = {
            "statut": 0,
            "message": "Erreur survenu lors de l'ajout du medicament",
            "data": {}
        }'''

    return JsonResponse(response)


def remove_medicament_session_gestionnaire(request, medicament_id):
    if request.method == 'POST':

        liste_medicaments = request.session.get('liste_medicaments', [])

        pprint(medicament_id)
        # Utiliser une boucle ou une compréhension de liste pour supprimer l'élément avec l'acte_id correspondant
        liste_medicaments = [item for item in liste_medicaments if item['medicament_id'] != medicament_id]
        pprint(liste_medicaments)

        # Mettre à jour les données de session
        request.session['liste_medicaments'] = liste_medicaments

        response = {
            "statut": 1,
            "message": "Suppression effectuée avec succès"
        }

    else:

        response = {
            "statut": 0,
            "message": "Erreur survenu lors de la suppression du medicament",
            "data": {}
        }

    return JsonResponse(response)


# FIN GESTIONNAIRES SINISTRE


@method_decorator(login_required, name='dispatch')
class FacturesPrestataireView(TemplateView):
    template_name = 'liste_factures_prestataires.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.filter(status=True).order_by('libelle')
        adherent_principal_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau, statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).values_list('adherent_principal_id', flat=True).order_by('-id')
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids)

        #prestataires = [x.prestataire for x in FacturePrestataire.objects.filter(prestataire__bureau=request.user.bureau, statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE])]
        prestataires_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau, statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).values_list('prestataire_id', flat=True).distinct()
        prestataires = Prestataire.objects.filter(id__in=prestataires_ids)

        assures_ids = FacturePrestataire.objects.filter(assure_id__isnull=False, bureau=request.user.bureau, statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).values_list('assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()
        # facture_prestataires = FacturePrestataire.objects.all()
        context = self.get_context_data(**kwargs)
        # context['dossiers_sinistres'] = dossiers_sinistres
        context['type_remboursements'] = type_remboursements
        context['adherents'] = adhs
        context['prestataires'] = prestataires
        context['assures'] = assures
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = 'today'


        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def factures_prestataire_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')

    queryset = FacturePrestataire.objects.filter(bureau=request.user.bureau,
        statut__in=[SatutBordereauDossierSinistres.ATTENTE, SatutBordereauDossierSinistres.REJETE]).order_by('-id')


    # la recherche
    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_type_remboursement:
        type_remboursement = TypeRemboursement.objects.get(id=search_type_remboursement)
        queryset = queryset.filter(type_remboursement=type_remboursement.pk)

        # if type_remboursement.code == "TP": // déjà gérer plus haut
        #     queryset = queryset.filter(prestataire__id=search_adherent_principal)

        # if search_type_remboursement == "RD":
        #     queryset = queryset.filter(adherent__id=search_adherent_principal)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_assure:
        queryset = queryset.filter(assure__id=search_assure)

    if search_adherent_principal:
        queryset = queryset.filter(adherent_principal=search_adherent_principal)


    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'beneficiaire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('detail_facture', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>'

        if c.type_remboursement == TypeRemboursement.objects.get(code="TP"):
            beneficiaire_de_reglement = c.prestataire.name if c.prestataire else ""
        elif c.type_remboursement == TypeRemboursement.objects.get(code="RD") :
            if c.assure:
                beneficiaire_de_reglement = (c.assure.nom + " " + (c.assure.prenoms if c.assure.prenoms else "")) if c.assure else ""
            else:
                beneficiaire_de_reglement = (c.adherent_principal.nom + " " + c.adherent_principal.prenoms) if c.adherent_principal else ""

        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""

        if c.statut == SatutBordereauDossierSinistres.REJETE:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-danger" style="font-size:10px;color:white;">{c.statut}</span>'
        else:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-en-attente" style="font-size:10px;color:white;">{c.statut}</span>'

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "beneficiaire__name": beneficiaire_de_reglement,
            "periode_comptable": periode_comptable_libelle,
            "net_a_payer": money_field(c.net_a_payer),
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "statut": statut_html,
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
class FacturesPrestataireTraiteView(TemplateView):
    template_name = 'liste_factures_prestataires_traitees.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.filter(status=True).order_by('libelle')
        adherent_principal_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE).values_list('adherent_principal_id', flat=True).order_by('-id')
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids)

        #prestataires = [x.prestataire for x in FacturePrestataire.objects.filter(prestataire__bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE)]

        #prestataires_ids = FacturePrestataire.objects.filter(prestataire__bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE).values_list('prestataire_id', flat=True).distinct()
        prestataires_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE).values_list('prestataire_id', flat=True).distinct()
        prestataires = Prestataire.objects.filter(id__in=prestataires_ids)

        assures_ids = FacturePrestataire.objects.filter(assure_id__isnull=False, bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE).values_list('assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()
        # facture_prestataires = FacturePrestataire.objects.all()
        context = self.get_context_data(**kwargs)
        # context['dossiers_sinistres'] = dossiers_sinistres

        context['type_remboursements'] = type_remboursements
        context['adherents'] = adhs
        context['prestataires'] = prestataires
        context['assures'] = assures
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def factures_prestataires_traitees_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')

    queryset = FacturePrestataire.objects.filter(bureau=request.user.bureau, statut=SatutBordereauDossierSinistres.VALIDE).order_by('-id')

    if search_type_remboursement:
        type_remboursement = TypeRemboursement.objects.get(id=search_type_remboursement)
        queryset = queryset.filter(type_remboursement__code=type_remboursement.code)

    # la recherche
    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_assure:
        queryset = queryset.filter(assure__id=search_assure)

    if search_adherent_principal:
        queryset = queryset.filter(adherent_principal=search_adherent_principal)

    if search_type_remboursement:
        queryset = queryset.filter(type_remboursement__id=search_type_remboursement)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'prestataire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('detail_facture', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>'

        statut_html = f'<span class="text-bold col-sm-12 badge badge-success" style="font-size:10px;color:white;">{c.statut}</span>'

        liste_sinistres_bordereau = Sinistre.objects.filter(facture_prestataire=c)
        montant_remb_total = c.net_a_payer
        montant_rejet_total = 0
        montant_net_a_payer = 0
        is_all_processed = True

        for sinistre in liste_sinistres_bordereau:
            montant_rejet_total = montant_rejet_total + (
                sinistre.montant_remb_refuse if sinistre.montant_remb_refuse else 0)

        for sinistre in liste_sinistres_bordereau:
            montant_net_a_payer = montant_net_a_payer + (
                sinistre.montant_remb_accepte if sinistre.montant_remb_accepte else 0)

        nom_beneficiaire = ""


        if c.type_remboursement.code == "RD":
            if c.assure:
                nom_beneficiaire = (c.assure.nom + " " + (c.assure.prenoms if c.assure.prenoms else "")) if c.assure else ""
            else:
                nom_beneficiaire = c.adherent_principal.nom + " " + c.adherent_principal.prenoms
        elif c.type_remboursement.code == "TP":
            nom_beneficiaire = c.prestataire.name if c.prestataire else ""

        nom_prestataire = nom_beneficiaire

        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""
        # rembourse = c.net_a_payer - float(montant_rejet_total)


        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "prestataire__name": nom_prestataire,
            "periode_comptable": periode_comptable_libelle,
            "net_a_payer": money_field(c.net_a_payer),
            "rejet": money_field(montant_rejet_total),
            "rembourse": money_field(montant_net_a_payer),
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "statut": statut_html,
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
class RemboursementAppliMobile(TemplateView):
    template_name = 'liste_dossiers_remboursement_direct_appli_mobile.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        dossiers_sinistres = [x for x in
                              DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE).order_by('id') if
                              x.sinistres.filter(approuved_by__bureau=request.user.bureau).exists()]
        pprint(dossiers_sinistres)

        prestataires = Prestataire.objects.filter(bureau=request.user.bureau, status=True)
        bordereaux = FacturePrestataire.objects.all()
        context = self.get_context_data(**kwargs)
        context['dossiers_sinistres'] = dossiers_sinistres
        context['prestataires'] = prestataires
        context['bordereaux'] = bordereaux
        context['user'] = user

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



@login_required
@never_cache
def search_assure(request):
    if request.method == "POST":

        # Récupérer la valeur de session_pec ou générer une nouvelle UUID
        session_pec = request.user.id

        numero_carte = request.POST.get('numero_carte')
        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        type_prise_en_charge = TypePriseencharge.objects.get(id=type_prise_en_charge_id)

        #quand on va commencer le TP Multipays, on enlevera le filtre bureau
        carte = Carte.objects.filter(numero=numero_carte, statut=Statut.ACTIF, aliment__bureau=request.user.bureau).first()

        if request.user.is_prestataire:
            date_survenance = datetime.datetime.now(tz=datetime.timezone.utc).date()

        else:
            prestataire_id = request.POST.get('prestataire_id')
            date_survenance_str = request.POST.get('date_survenance')
            # convertir en objet date
            date_survenance = datetimeJsdecode.strptime(date_survenance_str, '%Y-%m-%d').date()

            pprint("date_survenance renseigné")
            pprint(date_survenance)

        if carte:

            aliment = carte.aliment

            if aliment:

                #vérifier l'etat du bénéficiaire à cette date est sorti ou suspendu
                etat_beneficiaire = aliment.etat_beneficiaire_atdate(date_survenance)
                if etat_beneficiaire == "ACTIF":

                    # recherche en fonction de la date de survenance, date systeme pour le prestataire
                    formule = aliment.formule_atdate(date_survenance)
                    pprint("- Formule à la date du " + str(date_survenance))
                    pprint(formule)

                    if formule:
                        # dd(formule.libelle)

                        #Added on 20012023: intégration du réseau de soin, s'assurer que le prestataire est dans le réseau de soin de l'assuré
                        reseau_soin = formule.reseau_soin
                        if not reseau_soin or (reseau_soin and reseau_soin.has_prestataire(request.user.prestataire)):

                            # Added on 11102023: si la police a le statut ACTIF
                            # Quand on va faire les saisies antidatée, intégrer le code pour tenir compte de la date de suspension ...
                            if formule.police.statut == Statut.ACTIF:
                                etat_police = formule.police.etat_police
                                if etat_police == "En cours":

                                    # Added on 03112023: si la période est en cours
                                    periode_couverture = formule.police.periode_couverture_encours_atdate(date_survenance)

                                    if periode_couverture and periode_couverture.date_debut_effet.date() <= date_survenance and date_survenance <= periode_couverture.date_fin_effet.date():

                                        plafond_consommation_famille = formule.plafond_conso_famille
                                        plafond_consommation_individuelle = formule.plafond_conso_individuelle
                                        taux_couverture = formule.taux_couverture if formule.taux_couverture else 0

                                        # Added on 05102023: Préfiancenement du tm
                                        if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
                                            taux_couverture = 100

                                        # TODO : Tenier compte de la date de soins pour récupérer le plafond chambre et plafond hospit
                                        # son plafond chambre et plafond hospit
                                        bareme_plafond_chambre = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66023CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                        plafond_chambre = bareme_plafond_chambre.first().plafond_acte if bareme_plafond_chambre else 0

                                        bareme_plafond_hospitalisation = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66027CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                        plafond_hospitalisation = bareme_plafond_hospitalisation.first().plafond_acte if bareme_plafond_hospitalisation else 0


                                        bareme_plafond_accouchement = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66049CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                        plafond_accouchement = bareme_plafond_accouchement.first().plafond_acte if bareme_plafond_accouchement else 0


                                        # filtrer les actes en fonciton du type de prise en charge et du prestataire
                                        if type_prise_en_charge.code == 'AMBULAT':
                                            actes_garantis = Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE).exclude(
                                                rubrique__type_priseencharge__code="CONSULT").exclude(
                                                rubrique__type_priseencharge__code="HOSPIT").values('id', 'libelle', 'code',
                                                                                                    'option_seance')

                                        elif type_prise_en_charge.code == "CONSULT":

                                            actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE)
                                                              .filter(
                                                Q(rubrique__type_priseencharge_id=type_prise_en_charge_id) | Q(code="G79489CI01"))
                                                              .values('id', 'libelle', 'code', 'option_seance'))

                                            # si cabinet dentaire qui fait la consultation (à travers l'écran consultation): afficher uniquement l'acte consultation dentaire
                                            if request.user.prestataire and request.user.prestataire.type_prestataire.veos_code == 'DENTA':
                                                actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE)
                                                                  .filter(code="G79489CI01")
                                                                  .values('id', 'libelle', 'code', 'option_seance'))


                                        else:
                                            actes_garantis = Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE).filter(
                                                rubrique__type_priseencharge_id=type_prise_en_charge_id).values('id', 'libelle',
                                                                                                                'code',
                                                                                                                'option_seance')

                                            # dd(actes_garantis)

                                        # remove if condition: pour permettre au centre de soins de faire des actes optiques.
                                        # vu qu'on exclus les actes d'optiques qui ne respectent pas sur le barème, plus besoin de la condition if
                                        # si optique, affiche uniquement ceux qui respecteconditions de prise en charge

                                        baremes = Bareme.objects.filter(formulegarantie_id=formule.id,
                                                                        rubrique__type_priseencharge__code="OPTIQUE")
                                        baremes_a_retirer_ids = []

                                        pprint("Les baremes de la formule et qui concernent des actes d'optique")
                                        pprint(baremes)

                                        for b in baremes:
                                            bareme = Bareme.objects.get(id=b.id)

                                            if bareme.qualite_beneficiaire and bareme.qualite_beneficiaire != aliment.qualite_beneficiaire:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la qualite du beneficiaire")
                                                baremes_a_retirer_ids.append(bareme.id)

                                            if bareme.age_minimum and bareme.age_minimum > aliment.age:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la condition d'age minimum")
                                                baremes_a_retirer_ids.append(bareme.id)

                                            if bareme.age_maximum and bareme.age_maximum < aliment.age:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la condition d'age minimum")
                                                baremes_a_retirer_ids.append(bareme.id)

                                        pprint("baremes_a_retirer_ids")
                                        pprint(baremes_a_retirer_ids)

                                        # retirer les baremes qui ne respectent pas les conditions de prise en charge
                                        bon_baremes = baremes.exclude(id__in=baremes_a_retirer_ids)

                                        pprint("bon_baremes")
                                        pprint(bon_baremes)

                                        # les actes d'optiques dans le barème
                                        # filtre les actes enregistrés sur ces baremes
                                        pprint("filtre les actes enregistrés sur ces baremes")
                                        actes_optique_parametres = actes_garantis.filter(id__in=bon_baremes.values('acte_id'))
                                        pprint(actes_optique_parametres)

                                        pprint(type_prise_en_charge.code)
                                        # si opticien
                                        if type_prise_en_charge.code == "OPTIQUE":
                                            actes_garantis = actes_optique_parametres

                                        elif type_prise_en_charge.code == 'AMBULAT':

                                            # si centre de soins qui fait les actes d'optiques en soins ambulatoire,
                                            # complèter avec les actes d'optiques
                                            # Exclure tous les actes d'optiques pour ne rajouter que ceux qui correspondent au barème
                                            pprint(
                                                "Exclure tous les actes d'optiques pour ne rajouter que ceux qui correspondent au barème")
                                            actes_garantis = actes_garantis.exclude(rubrique__type_priseencharge__code="OPTIQUE")

                                            # rajouter ceux qui correspondent au barème et à l'assuré
                                            actes_garantis |= actes_optique_parametres

                                        # Si homme ou enfant, exclure les actes d'accouchement
                                        if aliment.genre == "M" or aliment.qualite_beneficiaire.code == "ENF":
                                            actes_garantis = actes_garantis.exclude(
                                                code__in=['G66044CI01', 'G66045CI01', 'G66046CI01', 'G66047CI01', 'G66048CI01',
                                                          'G66049CI01'])

                                        # garder l'acte consultation generaliste urgence garde uniquement entre 18H et 6h du matin, ainsi que les jours fériés
                                        jours_feries = JourFerie.objects.all().values('date')
                                        date_jour = datetime.datetime.now(tz=timezone.utc).date().strftime("%Y-%m-%d")
                                        heure_actuelle = datetime.datetime.now(tz=timezone.utc).time().hour
                                        # or date_jour not in jours_feries
                                        if (heure_actuelle > 7 and heure_actuelle < 18) or is_jour_ferie(date_jour):
                                            actes_garantis = actes_garantis.exclude(code__in=['G65913CI01', 'G65914CI01'])
                                            pprint(actes_garantis)

                                        pprint("jours_feries")
                                        pprint(jours_feries)

                                        if actes_garantis:
                                            actes_garantis_json = json.dumps(list(actes_garantis))

                                            photo_url = aliment.photo.url if aliment.photo else '/static/admin_custom/user.png'

                                            nom_client = f'{formule.police.client.nom} {formule.police.client.prenoms if formule.police.client.prenoms else "" }'

                                            response = {
                                                'statut': 1,
                                                'message': "ASSURÉ TROUVÉ",
                                                'data': {
                                                    'id': aliment.id,
                                                    'nom': aliment.nom,
                                                    'prenoms': aliment.prenoms,
                                                    'date_naissance': '/'.join(str(aliment.date_naissance).split('-')[
                                                                               ::-1]) if aliment.date_naissance else "",
                                                    'age': aliment.age,
                                                    'photo': photo_url,
                                                    'statut': aliment.statut,
                                                    'qualite_beneficiaire': aliment.qualite_beneficiaire.libelle,
                                                    'nom_client': nom_client,
                                                    'numero_police': formule.police.numero,
                                                    'formule': formule.type_tarif.libelle,
                                                    'formule_': formule.libelle + " - " + formule.type_tarif.libelle,
                                                    'taux': taux_couverture,
                                                    'nom_compagnie': formule.police.compagnie.nom,
                                                    'actes_garantis': actes_garantis_json,
                                                    'plafond_chambre': as_money(plafond_chambre),
                                                    'plafond_hospitalisation': as_money(plafond_hospitalisation),
                                                    'plafond_accouchement': as_money(plafond_accouchement),
                                                    'plafond_consommation_famille': as_money(
                                                        plafond_consommation_famille) + " " + formule.police.client.pays.devise.code,
                                                    'plafond_consommation_individuelle': as_money(
                                                        plafond_consommation_individuelle) + " " + formule.police.client.pays.devise.code,
                                                }
                                            }

                                            # vider toute la session de calcul de prise en charge pour reprendre
                                            SinistreTemporaire.objects.filter(
                                                session_pec=session_pec,
                                                aliment_id=aliment.id,
                                            ).delete()


                                        else:
                                            response = {
                                                'statut': 0,
                                                'message': "AUCUN ACTE DISPONIPLE",
                                            }

                                    else:
                                        response = {
                                            'statut': 0,
                                            'message': "L'ASSURÉ N'EST PAS COUVERT A CETTE DATE ",
                                        }

                                else:
                                    response = {
                                        'statut': 0,
                                        'message': "LA POLICE EST " + etat_police.upper() + "E",
                                    }

                            else:
                                response = {
                                    'statut': 0,
                                    'message': "LA POLICE N'EST PAS ACTIVE",
                                }

                        else:
                            response = {
                                'statut': 0,
                                'message': "CE CENTRE N'EST PAS DANS LE RÉSEAU DE SOINS DE L'ASSURÉ",
                            }

                    else:
                        response = {
                            'statut': 0,
                            'message': "L'ASSURÉ N'A AUCUNE FORMULE DE GARANTIE",
                        }

                else:
                    response = {
                        'statut': 0,
                        'message': "L'ASSURÉ EST " + str(etat_beneficiaire),
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "L'ASSURÉ EST " + str(aliment.statut),
                }

        else:

            response = {
                'statut': 0,
                'message': "CETTE CARTE N'EST PAS VALIDE",
            }

        return JsonResponse(response)


@login_required
@never_cache
def search_assure_bygestionnaire(request, prestataire_id):  # _bygestionnaire
    if request.method == "POST":

        # Récupérer la valeur de session_pec ou générer une nouvelle UUID
        session_pec = request.user.id

        numero_carte = request.POST.get('numero_carte')
        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')

        type_prise_en_charge_code = None
        if type_prise_en_charge_id:
            type_prise_en_charges = TypePriseencharge.objects.filter(id=type_prise_en_charge_id)
            type_prise_en_charge_code = type_prise_en_charges.first().code if type_prise_en_charges else None

        pprint("type_prise_en_charge_code")
        pprint(type_prise_en_charge_code)

        carte = Carte.objects.filter(numero=numero_carte, statut=Statut.ACTIF, aliment__bureau=request.user.bureau).first()

        #prestataire_id = request.POST.get('prestataire_id')
        prestataire = Prestataire.objects.get(id=prestataire_id)
        pprint(prestataire)

        date_survenance_str = request.POST.get('date_survenance')
        # convertir en objet date
        date_survenance = datetimeJsdecode.strptime(date_survenance_str, '%Y-%m-%d').date()

        pprint("date_survenance renseigné")
        pprint(date_survenance)

        if carte:

            aliment = carte.aliment

            if aliment:

                # vérifier l'etat du bénéficiaire à cette date est sorti ou suspendu
                etat_beneficiaire = aliment.etat_beneficiaire_atdate(date_survenance)
                if etat_beneficiaire == "ACTIF":

                    # recherche en fonction de la date de survenance, date systeme pour le prestataire
                    formule = aliment.formule_atdate(date_survenance)

                    if formule:
                        #Pas de besoin de vérification du réseau de soins quand c'est le gestionnaire qui saisi: cf reunion 23/01/2024

                        # Added on 11102023: si la police a le statut ACTIF
                        # Quand on va faire les saisies antidatée, intégrer le code pour tenir compte de la date de suspension ...
                        if formule.police.statut == Statut.ACTIF:
                            #etat_police = formule.police.etat_police
                            etat_police_atdate = formule.police.etat_police_atdate(date_survenance)

                            pprint("etat_police_atdate")
                            pprint(etat_police_atdate)
                            #pprint("etat_police")
                            #pprint(etat_police)

                            #if etat_police == "En cours":
                            if etat_police_atdate == "En cours":

                                # Added on 03112023: si la période est en cours
                                periode_couverture = formule.police.periode_couverture_encours_atdate(date_survenance)

                                if periode_couverture and periode_couverture.date_debut_effet.date() <= date_survenance and date_survenance <= periode_couverture.date_fin_effet.date():

                                    plafond_consommation_famille = formule.plafond_conso_famille
                                    plafond_consommation_individuelle = formule.plafond_conso_individuelle
                                    taux_couverture = formule.taux_couverture if formule.taux_couverture else 0

                                    # Added on 05102023: Préfiancenement du tm
                                    if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
                                        taux_couverture = 100

                                    # TODO : Tenier compte de la date de soins pour récupérer le plafond chambre et plafond hospit
                                    # son plafond chambre et plafond hospit
                                    bareme_plafond_chambre = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66023CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                    plafond_chambre = bareme_plafond_chambre.first().plafond_acte if bareme_plafond_chambre else 0

                                    bareme_plafond_hospitalisation = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66027CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                    plafond_hospitalisation = bareme_plafond_hospitalisation.first().plafond_acte if bareme_plafond_hospitalisation else 0

                                    bareme_plafond_accouchement = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66049CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                                    plafond_accouchement = bareme_plafond_accouchement.first().plafond_acte if bareme_plafond_accouchement else 0

                                    # Added condition on 22112023:
                                    actes_garantis = None

                                    if type_prise_en_charge_code:
                                        # Commented on 15112023: vu que c'est le gestionnaire qui saisie, afficher tous les actes
                                        # filtrer les actes en fonciton du type de prise en charge et du prestataire
                                        if type_prise_en_charge_code == 'AMBULAT':
                                            actes_garantis = Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE).exclude(
                                                rubrique__type_priseencharge__code="CONSULT").exclude(
                                                rubrique__type_priseencharge__code="HOSPIT").values('id', 'libelle', 'code',
                                                                                                    'option_seance')

                                        elif type_prise_en_charge_code == "CONSULT":

                                            actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE)
                                                              .filter(
                                                Q(rubrique__type_priseencharge_id=type_prise_en_charge_id) | Q(
                                                    code="G79489CI01"))
                                                              .values('id', 'libelle', 'code', 'option_seance'))

                                            # si cabinet dentaire qui fait la consultation (à travers l'écran consultation): afficher uniquement l'acte consultation dentaire
                                            if request.user.prestataire and request.user.prestataire.type_prestataire.veos_code == 'DENTA':
                                                actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE)
                                                                  .filter(code="G79489CI01")
                                                                  .values('id', 'libelle', 'code', 'option_seance'))


                                        else:
                                            actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1, statut_validite=StatutValidite.VALIDE)
                                                              .filter(rubrique__type_priseencharge_id=type_prise_en_charge_id)
                                                              .values('id', 'libelle', 'code', 'option_seance', 'rubrique'))

                                            # dd(actes_garantis)

                                        # remove if condition: pour permettre au centre de soins de faire des actes optiques.
                                        # vu qu'on exclus les actes d'optiques qui ne respectent pas sur le barème, plus besoin de la condition if
                                        # si optique, affiche uniquement ceux qui respecteconditions de prise en charge

                                        baremes = Bareme.objects.filter(formulegarantie_id=formule.id, rubrique__type_priseencharge__code="OPTIQUE")
                                        baremes_a_retirer_ids = []

                                        pprint("Les baremes de la formule et qui concernent des actes d'optique")
                                        pprint(baremes)

                                        for b in baremes:
                                            bareme = Bareme.objects.get(id=b.id)

                                            if bareme.qualite_beneficiaire and bareme.qualite_beneficiaire != aliment.qualite_beneficiaire:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la qualite du beneficiaire")
                                                baremes_a_retirer_ids.append(bareme.id)

                                            if bareme.age_minimum and bareme.age_minimum > aliment.age:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la condition d'age minimum")
                                                baremes_a_retirer_ids.append(bareme.id)

                                            if bareme.age_maximum and bareme.age_maximum < aliment.age:
                                                pprint("- Le bareme #" + str(
                                                    bareme.id) + " ne respecte pas la condition d'age minimum")
                                                baremes_a_retirer_ids.append(bareme.id)

                                        pprint("baremes_a_retirer_ids")
                                        pprint(baremes_a_retirer_ids)

                                        # retirer les baremes qui ne respectent pas les conditions de prise en charge
                                        bon_baremes = baremes.exclude(id__in=baremes_a_retirer_ids)

                                        pprint("bon_baremes")
                                        pprint(bon_baremes)

                                        # les actes d'optiques dans le barème
                                        # filtre les actes enregistrés sur ces baremes
                                        pprint("filtre les actes enregistrés sur ces baremes")
                                        actes_optique_parametres = actes_garantis.filter(id__in=bon_baremes.values('acte_id'))
                                        pprint(actes_optique_parametres)

                                        pprint(type_prise_en_charge_code)
                                        # si opticien
                                        if type_prise_en_charge_code == "OPTIQUE":
                                            actes_garantis = actes_optique_parametres

                                        elif type_prise_en_charge_code == 'AMBULAT':

                                            # si centre de soins qui fait les actes d'optiques en soins ambulatoire,
                                            # complèter avec les actes d'optiques
                                            # Exclure tous les actes d'optiques pour ne rajouter que ceux qui correspondent au barème
                                            pprint(
                                                "Exclure tous les actes d'optiques pour ne rajouter que ceux qui correspondent au barème")
                                            actes_garantis = actes_garantis.exclude(
                                                rubrique__type_priseencharge__code="OPTIQUE")

                                            # rajouter ceux qui correspondent au barème et à l'assuré
                                            actes_garantis |= actes_optique_parametres

                                        # Si homme ou enfant, exclure les actes d'accouchement
                                        if aliment.genre == "M" or aliment.qualite_beneficiaire.code == "ENF":
                                            actes_garantis = actes_garantis.exclude(
                                                code__in=['G66044CI01', 'G66045CI01', 'G66046CI01', 'G66047CI01', 'G66048CI01',
                                                          'G66049CI01'])

                                        # garder l'acte consultation generaliste urgence garde uniquement entre 18H et 6h du matin, ainsi que les jours fériés
                                        #jours_feries = JourFerie.objects.all().values('date')
                                        #date_jour = datetime.datetime.now(tz=timezone.utc).date().strftime("%Y-%m-%d")
                                        #heure_actuelle = datetime.datetime.now(tz=timezone.utc).time().hour
                                        # or date_jour not in jours_feries
                                        #if (heure_actuelle > 7 and heure_actuelle < 18) or is_jour_ferie(date_jour):
                                        #    actes_garantis = actes_garantis.exclude(code__in=['G65913CI01', 'G65914CI01'])
                                        #    pprint(actes_garantis)

                                        #pprint("jours_feries")
                                        #pprint(jours_feries)

                                        # commented on 20112023:reintégré cette notion de type_pec
                                        # Added on 15112023: Afficher tous les actes vu que c'est le gestionnaire qui saisie
                                        # actes_garantis = (Acte.objects.filter(type_acte__code="ACTE", status=1)
                                        # .filter(rubrique__type_priseencharge_id=type_prise_en_charge_id)
                                        #                  .values('id', 'libelle', 'code', 'option_seance', 'rubrique__code', 'rubrique__type_priseencharge__code'))

                                    if actes_garantis:
                                        actes_garantis_json = json.dumps(list(actes_garantis))

                                    else:
                                        actes_garantis_json = json.dumps([])

                                    photo_url = aliment.photo.url if aliment.photo else '/static/admin_custom/user.png'

                                    nom_client = f'{formule.police.client.nom} {formule.police.client.prenoms if formule.police.client.prenoms else "" }'

                                    # FIX : erreur "ERREUR LORS DE LA RECHERCHE" sur le bureau Gabon
                                    try:
                                        response = {
                                            'statut': 1,
                                            'message': "ASSURÉ TROUVÉ",
                                            'data': {
                                                'id': aliment.id,
                                                'nom': aliment.nom,
                                                'prenoms': aliment.prenoms,
                                                'date_naissance': '/'.join(str(aliment.date_naissance).split('-')[::-1]) if aliment.date_naissance else "",
                                                'age': aliment.age,
                                                'photo': photo_url,
                                                'statut': aliment.statut,
                                                'qualite_beneficiaire': aliment.qualite_beneficiaire.libelle if aliment.qualite_beneficiaire else "",
                                                'numero_police': formule.police.numero if formule.police else "",
                                                'nom_client': nom_client,
                                                'formule': formule.type_tarif.libelle if formule.type_tarif else "",
                                                'formule_': f"{formule.libelle} - {formule.type_tarif.libelle}" if formule.type_tarif else "",
                                                'taux': taux_couverture,
                                                'nom_compagnie': formule.police.compagnie.nom if formule.police and formule.police.compagnie else "",
                                                'actes_garantis': actes_garantis_json,
                                                'plafond_chambre': as_money(plafond_chambre),
                                                'plafond_hospitalisation': as_money(plafond_hospitalisation),
                                                'plafond_accouchement': as_money(plafond_accouchement),
                                                'plafond_consommation_famille': f"{as_money(plafond_consommation_famille)} {formule.police.client.pays.devise.code}" if formule.police and formule.police.client and formule.police.client.pays else "",
                                                'plafond_consommation_individuelle': f"{as_money(plafond_consommation_individuelle)} {formule.police.client.pays.devise.code}" if formule.police and formule.police.client and formule.police.client.pays else "",
                                            }
                                        }
                                    except Exception as e:
                                        response = {
                                            'statut': 0,
                                            'message': f"ERREUR LORS DE LA RÉCUPÉRATION DES DONNÉES : {str(e)}",
                                        }

                                    # vider toute la session de calcul de prise en charge pour reprendre
                                    SinistreTemporaire.objects.filter(
                                        session_pec=session_pec,
                                        aliment_id=aliment.id,
                                    ).delete()


                                else:
                                    response = {
                                        'statut': 0,
                                        'message': "L'ASSURÉ N'EST PAS COUVERT A CETTE DATE ",
                                    }

                            else:
                                response = {
                                    'statut': 0,
                                    'message': "LA POLICE EST " + etat_police_atdate.upper() + "E",
                                }

                        else:
                            response = {
                                'statut': 0,
                                'message': "LA POLICE N'EST PAS ACTIVE",
                            }

                    else:
                        response = {
                            'statut': 0,
                            'message': "L'ASSURÉ N'A AUCUNE FORMULE DE GARANTIE",
                        }

                else:
                    response = {
                        'statut': 0,
                        'message': "L'ASSURÉ EST " + str(etat_beneficiaire) + " A CETTE DATE",
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "L'ASSURÉ EST " + str(aliment.statut),
                }

        else:

            response = {
                'statut': 0,
                'message': "CETTE CARTE N'EST PAS VALIDE",
            }

        return JsonResponse(response)


def pharmacie_details(request, aliment_id):
    aliment = Aliment.objects.get(id=aliment_id)

    if aliment and aliment.formule:
        dossiers_sinistres = DossierSinistre.objects.filter(aliment_id=aliment.id, type_priseencharge__code="CONSULT", is_closed=False)

        # Added on 11102023: Permettre au pharmacien d'utiliser la v2 même si la consultation a été faite sur la v2
        if not dossiers_sinistres:
            # Added on 05102023: add tm préfinancé
            type_prefinancement = aliment.formule.police.type_prefinancement if aliment.formule.police.type_prefinancement else None

            # dd(dossiers_sinistres)
            prestataires_ph_direct = Prestataire.objects.filter(bureau=request.user.bureau, status=True, code__contains="PH_DIRECT")
            prestataire_ph_direct_id = prestataires_ph_direct.first().id if prestataires_ph_direct else None

            dossier_sinistre = DossierSinistre.objects.create(created_by=request.user,
                                                              bureau_id=request.user.bureau.id,
                                                              prestataire_id=prestataire_ph_direct_id,
                                                              centre_prescripteur_id=prestataire_ph_direct_id,# a confirmer
                                                              pharmacie_id=request.user.prestataire.id,
                                                              aliment_id=aliment.id,
                                                              formulegarantie_id=aliment.formule.id,
                                                              police_id=aliment.formule.police.id,
                                                              compagnie_id=aliment.formule.police.compagnie.id,
                                                              prescripteur_id=None,
                                                              type_priseencharge_id=1,  # consultation
                                                              type_remboursement_id=1,
                                                              date_survenance=datetime.datetime.now(tz=timezone.utc),
                                                              renseignement_clinique="",
                                                              plafond_chambre=0,
                                                              plafond_hospit=0,
                                                              affection_id=None,
                                                              #type_prefinancement=type_prefinancement
                                                              #statut_pec=StatutSinistre.ACCORDE #À VÉRIFIER
                                                              )
            dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre.pk)

            # mettre a jour le numéro du dossier
            code_bureau = request.user.bureau.code
            dossier_sinistre.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                dossier_sinistre.pk).zfill(7) + '-BS'
            # dossier_sinistre.numero = 'BS' + str(Date.today().year) + str(dossier_sinistre.pk).zfill(6)
            dossier_sinistre.save()

        dossiers_sinistres = DossierSinistre.objects.filter(aliment_id=aliment.id, type_priseencharge__code="CONSULT",
                                                            is_closed=False)

        return render(request, 'details_bon_de_soin.html',
                      {'aliment': aliment, 'dossiers_sinistres': dossiers_sinistres})

    else:
        admin_home_url = reverse('admin:index')
        return redirect(admin_home_url)


@login_required
@never_cache
def search_assure_pharmacie(request):
    if request.method == "POST":

        numero_carte = request.POST.get('numero_carte')

        # quand on va commencer le TP Multipays, on enlevera le filtre bureau
        carte = Carte.objects.filter(numero=numero_carte, statut=Statut.ACTIF, aliment__bureau=request.user.bureau).first()

        if carte:

            aliment = carte.aliment
            formule = aliment.formule

            if formule:

                etat_police = formule.police.etat_police
                if etat_police == "En cours":
                    response = {
                        'statut': 1,
                        'redirect': reverse('pharmacie_details', kwargs={'aliment_id': aliment.pk}),
                    }

                else:
                    response = {
                        'statut': 0,
                        'message': "LA POLICE EST " + etat_police.upper() + "E",
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "L'ASSURÉ N'A AUCUNE FORMULE DE GARANTIE",
                }

        else:

            response = {
                'statut': 0,
                'message': "CETTE CARTE N'EST PAS VALIDE",
            }

        return JsonResponse(response)


@login_required
@never_cache
def statuer_acte(request):
    if request.method == "POST":

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        sinistre_id = request.POST.get('sinistre_id')
        type_operation = request.POST.get('type_operation')
        nombre_accorde = request.POST.get('nombre_accorde')
        # dd(nombre_accorde)
        motif_rejet = request.POST.get('motif_rejet')

        nombre_accorde = int(nombre_accorde) if nombre_accorde else 1

        # Retrieve the Sinistre object based on the provided sinistre_id
        sinistre = get_object_or_404(Sinistre, id=sinistre_id)

        # Update the status of the Acte
        if type_operation == "Confirm":

            # Added on 28092023: refaire le calcul pour controler les plafonds
            # vider toute la session de calcul de prise en charge pour reprendre
            session_pec = request.user.id
            vider_sinistres_temporaires(session_pec)

            pprint(request.POST)

            type_prise_en_charge_code = sinistre.acte.rubrique.type_priseencharge.code
            pprint("type_prise_en_charge_code")
            pprint(type_prise_en_charge_code)
            cout_acte = sinistre.frais_reel  # A bien vérifier , en principe si le cout avait été saisi manuelement , ...
            prestataire = sinistre.prestataire
            prescripteur_id = sinistre.prescripteur.pk if sinistre.prescripteur else None
            aliment_id = sinistre.aliment.pk
            aliment = sinistre.aliment
            date_survenance = sinistre.date_survenance

            acte = sinistre.acte
            acte_id = acte.pk

            # récupérer ses consommations individuel et par famille
            periode_couverture_encours = aliment.formule.police.periode_couverture_encours
            consommation_individuelle = Sinistre.objects.filter(
                periode_couverture_id=periode_couverture_encours.pk,
                aliment_id=aliment.id,
                statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE
            ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            consommation_famille = Sinistre.objects.filter(
                periode_couverture_id=periode_couverture_encours.pk,
                adherent_principal_id=aliment.adherent_principal.id,
                statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE
            ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
            infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id, prestataire.id,
                                                    prescripteur_id, aliment_id, cout_acte, nombre_accorde,
                                                    consommation_individuelle, consommation_famille, session_pec)

            pprint(infos_acte['statut'])

            if infos_acte['statut'] == 0:
                response = {
                    'statut': 0,
                    'message': infos_acte['message'],
                    'data': {}
                }
                return JsonResponse(response)

            else:

                '''if acte.option_seance:
                    frais_reel = infos_acte['data']['frais_reel'] / int(nombre_accorde)
                    part_compagnie = infos_acte['data']['part_compagnie'] / int(nombre_accorde)
                    part_assure = infos_acte['data']['part_assure'] / int(nombre_accorde)
                    ticket_moderateur = infos_acte['data']['ticket_moderateur'] / int(nombre_accorde)
                    depassement = infos_acte['data']['depassement'] / int(nombre_accorde)
                else:'''
                frais_reel = infos_acte['data']['frais_reel']
                part_compagnie = infos_acte['data']['part_compagnie']
                part_assure = infos_acte['data']['part_assure']
                ticket_moderateur = infos_acte['data']['ticket_moderateur']
                depassement = infos_acte['data']['depassement']

                plafond_acte = infos_acte['data']['plafond_acte']
                nombre_acte = infos_acte['data']['nombre_acte']
                frequence = infos_acte['data']['frequence']
                unite_frequence = infos_acte['data']['unite_frequence']
                garanti = infos_acte['data']['garanti']
                bareme_id = infos_acte['data']['bareme_id']

                # mettre à jour les parts après recalcul
                sinistre.frais_reel = frais_reel
                sinistre.part_compagnie = part_compagnie
                sinistre.part_assure = part_assure
                sinistre.ticket_moderateur = ticket_moderateur
                sinistre.depassement = depassement

                sinistre.statut = StatutSinistre.ACCORDE
                sinistre.nombre_accorde = nombre_accorde

                # SI HOSPIT, Mettre à jour la date de sortie
                if sinistre.dossier_sinistre.type_priseencharge.code == "HOSPIT":
                    date_entree = sinistre.date_entree.date()
                    date_sortie = date_entree + datetime.timedelta(days=nombre_accorde)
                    sinistre.date_sortie = date_sortie

                texte_opereation = "accordé"

                sinistre.approuved_by = request.user
                sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                sinistre.save()

                #signaler sur le dossier qu'il est traité par medecin conseil
                update_dossier_traitement_by_med_cons(sinistre)

                # dd(sinistre.nombre_accorde)

                # Notifier waspito si le prestataire est WASPITO INC
                notifier_waspito(sinistre)

                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="accorde",
                                         description="Accord d'un sinistre", table="sinistre",
                                         row=sinistre.pk)

                # si c'est un acte avec séance, dupliquer les lignes de séance en fonction du nombre accordé
                if sinistre.acte.option_seance:
                    cpt = 1
                    for _ in range(sinistre.nombre_accorde - 1):
                        new_sinistre = deepcopy(sinistre)  # Copie de l'objet

                        new_sinistre.id = None
                        new_sinistre.numero = None
                        new_sinistre.nombre_demande = 1
                        new_sinistre.nombre_accorde = 1
                        new_sinistre.date_survenance = None
                        new_sinistre.statut_prestation = StatutSinistrePrestation.ATTENTE
                        new_sinistre.approuved_by = request.user
                        new_sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                        new_sinistre.save()

                        # générer le numéro
                        code_bureau = sinistre.dossier_sinistre.bureau.code
                        new_sinistre.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                            new_sinistre.pk).zfill(7) + '-SP'
                        new_sinistre.save()

                        cpt += 1
                        pprint("duplicata de l'acte - séance N° " + str(cpt))

                        # Notifier waspito si le prestataire est WASPITO INC
                        # notifier_waspito(new_sinistre)

                        # enregistrer dans les log
                        ActionLog.objects.create(done_by=request.user, action="accorde",
                                                 description="Accord d'un sinistre", table="sinistre",
                                                 row=new_sinistre.pk)

                sinistre.motif_rejet = motif_rejet
                sinistre.save()



        else:
            sinistre.approuved_by = request.user
            sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
            sinistre.statut = StatutSinistre.REJETE
            sinistre.motif_rejet = motif_rejet
            texte_opereation = "rejeté"
            sinistre.save()

            # signaler sur le dossier qu'il est traité par medecin conseil
            update_dossier_traitement_by_med_cons(sinistre)

            # Notifier waspito si le prestataire est WASPITO INC
            notifier_waspito(sinistre)

            # enregistrer dans les log
            ActionLog.objects.create(done_by=request.user, action="rejete",
                                     description="Rejet d'un sinistre", table="sinistre",
                                     row=sinistre.pk)

        # si tous les sinistres ont été traité, rediriger à la liste des prises en charges, sinon actualiser la page
        sinistres_en_attente = Sinistre.objects.filter(dossier_sinistre_id=sinistre.dossier_sinistre_id,
                                                       statut=StatutSinistre.ATTENTE)
        redirectto = '' if sinistres_en_attente else reverse('admin:index')

        # mettre à jour la date validation du dossier_sinistre
        # Mettre à jour le statut_pec de dossier_sinistre
        update_statut_pec_dossier_sinistre(sinistre.dossier_sinistre)

        response = {
            'statut': 1,
            'message': "Acte " + texte_opereation + " avec succèss.",
            'redirectto': redirectto,
            'data': {
                'id': sinistre_id,
                'type': type_operation,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "Aucune demande d'approbation ou de reject lancée",
        }

    return JsonResponse(response)


@login_required
@never_cache
def approuver_liste_acte(request):
    response = {}  # Initialiser la variable response en dehors de la boucle

    if request.method == 'POST':
        for key in request.POST:
            if key.startswith('acte_'):
                sinistre_id = key.split('_')[1]
                print("mise à jour du sinistre : " + sinistre_id)
                # Retrieve the Sinistre object based on the provided sinistre_id
                sinistre = get_object_or_404(Sinistre, id=sinistre_id)
                if request.POST.getlist(key) == ['on']:
                    sinistre.statut = StatutSinistre.ACCORDE
                    sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                    sinistre.approuved_by = request.user

                    # accorder le nombre demandé
                    sinistre.nombre_accorde = sinistre.nombre_demande

                    # SI HOSPIT, Mettre à jour la date de sortie
                    if sinistre.dossier_sinistre.type_priseencharge.code == "HOSPIT":
                        date_entree = sinistre.date_entree.date()
                        date_sortie = date_entree + datetime.timedelta(days=sinistre.nombre_demande)
                        sinistre.date_sortie = date_sortie

                    sinistre.save()

                    #signaler sur le dossier qu'il est traité par medecin conseil
                    update_dossier_traitement_by_med_cons(sinistre)

                    # Notifier waspito si le prestataire est WASPITO INC
                    notifier_waspito(sinistre)

                    # enregistrer dans les log
                    ActionLog.objects.create(done_by=request.user, action="accorde",
                                             description="Accord d'un sinistre", table="sinistre",
                                             row=sinistre.pk)

                    # si c'est un acte avec séance, dupliquer les lignes de séance en fonction du nombre accordé
                    if sinistre.acte.option_seance:
                        cpt = 1  # vu qu'il y a déjà une ligne de séance sur laquelle on est
                        nombre_total = sinistre.nombre_accorde

                        for _ in range(sinistre.nombre_accorde - 1):
                            new_sinistre = deepcopy(sinistre)  # Copie de l'objet

                            new_sinistre.id = None
                            new_sinistre.numero = None
                            new_sinistre.nombre_demande = 1
                            new_sinistre.nombre_accorde = 1
                            new_sinistre.date_survenance = None
                            new_sinistre.statut_prestation = StatutSinistrePrestation.ATTENTE
                            new_sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                            new_sinistre.approuved_by = request.user
                            new_sinistre.save()

                            # générer le numéro
                            code_bureau = sinistre.dossier_sinistre.bureau.code
                            new_sinistre.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                                new_sinistre.pk).zfill(7) + '-SP'
                            new_sinistre.save()

                            cpt += 1

                            # Notifier waspito si le prestataire est WASPITO INC
                            notifier_waspito(new_sinistre)

                            # enregistrer dans les log
                            ActionLog.objects.create(done_by=request.user, action="accorde",
                                                     description="Accord d'un sinistre", table="sinistre",
                                                     row=new_sinistre.pk)

                            pprint("cpt")
                            pprint(cpt)
                            pprint("sinistre.nombre_demande")
                            pprint(sinistre.nombre_demande)

                    else:
                        print("L'acte n'est pas avec séance")

                    response = {
                        'statut': 1,
                        'message': "Tous les actes sélectionnés ont été approuvés avec succèss.",
                        'data': {
                        }
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "Aucun acte sélectionné pour approuver",
                    'data': {}
                }

        sinistres_en_attente = Sinistre.objects.filter(dossier_sinistre_id=sinistre.dossier_sinistre_id,
                                                       statut=StatutSinistre.ATTENTE)
        redirectto = '' if sinistres_en_attente else reverse('admin:index')

        # Mettre à jour le statut_pec de dossier_sinistre
        update_statut_pec_dossier_sinistre(sinistre.dossier_sinistre)

        response = {
            'statut': 1,
            'message': "Tous les actes sélectionnés ont été approuvés avec succèss.",
            'redirectto': redirectto,
            'data': {
            }
        }


    else:
        response = {
            'statut': 0,
            'message': "Erreur, non autorisé",
            'data': {
            }
        }

    return JsonResponse(response)


def update_statut_pec_dossier_sinistre(dossier_sinistre):
    pprint("update_statut_pec_dossier_sinistre")
    pprint(dossier_sinistre)

    sinistres_en_attente = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ATTENTE)
    sinistres_accordes = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ACCORDE)
    if sinistres_en_attente:
        dossier_sinistre.statut_pec = StatutSinistre.ATTENTE
    else:
        # si pas de accorde
        if sinistres_accordes:
            dossier_sinistre.statut_pec = StatutSinistre.ACCORDE
        else:
            dossier_sinistre.statut_pec = StatutSinistre.REJETE

    dossier_sinistre.save()


@login_required
@never_cache
def approuver_liste_acte_new(request):
    response = {}  # Initialiser la variable response en dehors de la boucle

    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        dossier_sinistre_id = request.GET.get('dossier_sinistre_id', "0")

        for key in request.POST:
            if key.startswith('acte_'):
                sinistre_id = key.split('_')[1]
                print("mise à jour du sinistre : " + sinistre_id)
                # Retrieve the Sinistre object based on the provided sinistre_id
                sinistre = get_object_or_404(Sinistre, id=sinistre_id)
                if request.POST.getlist(key) == ['on']:

                    # Added on 28092023: permettre de recalculer à la validation par le medecin

                    type_prise_en_charge_code = sinistre.acte.rubrique.type_priseencharge.code
                    cout_acte = sinistre.frais_reel  # A bien vérifier , en principe si le cout avait été saisi manuelement , ...
                    prestataire = sinistre.prestataire
                    prescripteur_id = sinistre.prescripteur.pk
                    aliment_id = sinistre.aliment.pk
                    aliment = sinistre.aliment
                    date_survenance = sinistre.date_survenance

                    acte = sinistre.acte
                    acte_id = acte.pk
                    nombre_accorde = sinistre.nombre_demande  # accorde ce qui est demandé

                    # récupérer ses consommations individuel et par famille
                    periode_couverture_encours = aliment.formule.police.periode_couverture_encours
                    consommation_individuelle = Sinistre.objects.filter(
                        periode_couverture_id=periode_couverture_encours.pk,
                        aliment_id=aliment.id,
                        statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE
                    ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                    consommation_famille = Sinistre.objects.filter(
                        periode_couverture_id=periode_couverture_encours.pk,
                        adherent_principal_id=aliment.adherent_principal.id,
                        statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE
                    ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                    # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                    infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                            prestataire.id,
                                                            prescripteur_id, aliment_id, cout_acte, nombre_accorde,
                                                            consommation_individuelle, consommation_famille,
                                                            session_pec)

                    '''if acte.option_seance:
                        frais_reel = infos_acte['data']['frais_reel'] / int(nombre_accorde)
                        part_compagnie = infos_acte['data']['part_compagnie'] / int(nombre_accorde)
                        part_assure = infos_acte['data']['part_assure'] / int(nombre_accorde)
                        ticket_moderateur = infos_acte['data']['ticket_moderateur'] / int(nombre_accorde)
                        depassement = infos_acte['data']['depassement'] / int(nombre_accorde)
                    else:'''
                    frais_reel = infos_acte['data']['frais_reel']
                    part_compagnie = infos_acte['data']['part_compagnie']
                    part_assure = infos_acte['data']['part_assure']
                    ticket_moderateur = infos_acte['data']['ticket_moderateur']
                    depassement = infos_acte['data']['depassement']

                    plafond_acte = infos_acte['data']['plafond_acte']
                    nombre_acte = infos_acte['data']['nombre_acte']
                    frequence = infos_acte['data']['frequence']
                    unite_frequence = infos_acte['data']['unite_frequence']
                    garanti = infos_acte['data']['garanti']
                    bareme_id = infos_acte['data']['bareme_id']

                    # mettre à jour les parts après recalcul
                    sinistre.frais_reel = frais_reel
                    sinistre.part_compagnie = part_compagnie
                    sinistre.part_assure = part_assure
                    sinistre.ticket_moderateur = ticket_moderateur
                    sinistre.depassement = depassement

                    sinistre.statut = StatutSinistre.ACCORDE
                    sinistre.approuved_by = request.user
                    sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)

                    # accorder le nombre demandé
                    sinistre.nombre_accorde = nombre_accorde

                    # SI HOSPIT, Mettre à jour la date de sortie
                    if sinistre.dossier_sinistre.type_priseencharge.code == "HOSPIT":
                        date_entree = sinistre.date_entree.date()
                        date_sortie = date_entree + datetime.timedelta(days=sinistre.nombre_demande)
                        sinistre.date_sortie = date_sortie

                    sinistre.save()

                    #signaler sur le dossier qu'il est traité par medecin conseil
                    update_dossier_traitement_by_med_cons(sinistre)

                    ActionLog.objects.create(done_by=request.user, action="accorde",
                                             description="Accord d'un sinistre", table="sinistre",
                                             row=sinistre.pk)

                    # si c'est un acte avec séance, dupliquer les lignes de séance en fonction du nombre accordé
                    if sinistre.acte.option_seance:
                        cpt = 1  # vu qu'il y a déjà une ligne de séance sur laquelle on est
                        nombre_total = sinistre.nombre_accorde

                        while cpt < nombre_total:  # si c'est approuver tout, on considère le nombre demandé
                            new_sinistre = sinistre
                            new_sinistre.__dict__.update(sinistre.__dict__)
                            new_sinistre.id = None
                            new_sinistre.nombre_demande = 1
                            new_sinistre.nombre_accorde = 1
                            new_sinistre.date_survenance = None
                            new_sinistre.statut_prestation = StatutSinistrePrestation.ATTENTE
                            new_sinistre.approuved_by = request.user
                            new_sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                            new_sinistre.save()
                            cpt += 1

                            ActionLog.objects.create(done_by=request.user, action="accorde",
                                                     description="Accord d'un sinistre", table="sinistre",
                                                     row=new_sinistre.pk)

                            pprint("cpt")
                            pprint(cpt)
                            pprint("sinistre.nombre_demande")
                            pprint(sinistre.nombre_demande)

                    else:
                        print("L'acte n'est pas avec séance")

                    response = {
                        'statut': 1,
                        'message': "Tous les actes sélectionnés ont été approuvés avec succèss.",
                        'data': {
                        }
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "Aucun acte sélectionné pour approuver",
                    'data': {}
                }

        sinistres_en_attente = Sinistre.objects.filter(dossier_sinistre_id=dossier_sinistre_id,
                                                       statut=StatutSinistre.ATTENTE)
        redirectto = '' if sinistres_en_attente else reverse('admin:index')

        # Mettre à jour le statut_pec de dossier_sinistre
        update_statut_pec_dossier_sinistre(sinistre.dossier_sinistre)

        response = {
            'statut': 1,
            'message': "Tous les actes sélectionnés ont été approuvés avec succèss.",
            'redirectto': redirectto,
            'data': {
            }
        }


    else:
        response = {
            'statut': 0,
            'message': "Erreur, non autorisé",
            'data': {
            }
        }

    return JsonResponse(response)


@login_required
@never_cache
def rejeter_liste_acte(request):
    response = {}  # Initialiser la variable response en dehors de la boucle

    if request.method == 'POST':
        motif_rejet = request.GET.get('motif_rejet', "")
        dossier_sinistre_id = request.GET.get('dossier_sinistre_id', "0")
        pprint("dossier_sinistre_id")
        pprint(dossier_sinistre_id)

        for key in request.POST:
            if key.startswith('acte_'):
                sinistre_id = key.split('_')[1]
                print("mise à jour du sinistre : " + sinistre_id)
                # Retrieve the Sinistre object based on the provided sinistre_id
                sinistre = get_object_or_404(Sinistre, id=sinistre_id)

                if request.POST.getlist(key) == ['on']:
                    sinistre.motif_rejet = motif_rejet
                    sinistre.statut = StatutSinistre.REJETE
                    sinistre.approuved_by = request.user
                    sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                else:
                    print("Nothing to do ")
                sinistre.save()

                #signaler sur le dossier qu'il est traité par medecin conseil
                update_dossier_traitement_by_med_cons(sinistre)

                # Notifier waspito si le prestataire est WASPITO INC
                notifier_waspito(sinistre)

                # Enregistre dans les logs
                ActionLog.objects.create(done_by=request.user, action="rejete",
                                         description="Rejet d'un sinistre", table="sinistre",
                                         row=sinistre.pk)

                response = {
                    'statut': 1,
                    'message': "Tous les actes sélectionnés ont été rejété avec succèss.",
                    'data': {
                    }
                }
            else:

                response = {
                    'statut': 0,
                    'message': "Aucun acte n'a été sélectionné pour rejéter",
                    'data': {}
                }

        sinistres_en_attente = Sinistre.objects.filter(dossier_sinistre_id=dossier_sinistre_id,
                                                       statut=StatutSinistre.ATTENTE)
        redirectto = '' if sinistres_en_attente else reverse('admin:index')

        # Mettre à jour le statut_pec de dossier_sinistre
        update_statut_pec_dossier_sinistre(sinistre.dossier_sinistre)

        response = {
            'statut': 1,
            'message': "Tous les actes sélectionnés ont été rejété avec succèss.",
            'redirectto': redirectto,
            'data': {
            }
        }


    else:
        response = {
            'statut': 0,
            'message': "Erreur, non autorisé",
            'data': {
            }
        }

    return JsonResponse(response)


# signaler sur le dossier qu'il est traité par medecin conseil
def update_dossier_traitement_by_med_cons(sinistre):
    if sinistre.dossier_sinistre.has_sinistre_traite_bymedecin is False:
        dossier_sinistre = sinistre.dossier_sinistre
        dossier_sinistre.has_sinistre_traite_bymedecin = True
        dossier_sinistre.date_traitement_sinistre_bymedecin = timezone.now()
        dossier_sinistre.save()
        

def notifier_waspito(sinistre):
    if sinistre.prestataire.code == "9537":

        # Les données que vous souhaitez envoyer
        body = {
            "statut_sinistre": sinistre.statut,
            "code_acte": sinistre.acte.code,
            "libelle_acte": sinistre.acte.libelle,
            "frais_reel": float(sinistre.frais_reel),
            "part_assure": float(sinistre.part_assure),
            "part_assureur": float(sinistre.part_compagnie),
            "numero_sinistre": sinistre.numero,
            "numero_dossier_sinistre": sinistre.dossier_sinistre.numero,
            "numero_police": sinistre.police.numero,
            "motif": sinistre.motif_rejet
        }

        # Convertir les données en format JSON
        donnees_json = json.dumps(body)

        # URL de l'API
        url = "https://waspito.com/apps/webhooks/insurances/claim-status-updated"

        # En-têtes de la requête
        headers = {
            "Content-Type": "application/json",  # Indique que vous envoyez des données JSON
            "w-insurance-verify-hash": "8d6oMBWLjD0qC9PK2NsA2XvK6CpPKHj9bQRgab59bJHO7me55jfgjtEjGWEwSTq4"
        }

        # Effectuer la requête POST
        response = requests.post(url, data=donnees_json, headers=headers)

        # Vérifier la réponse du serveur
        if response.status_code == 200:
            print("La requête POST a été effectuée avec succès.")
        else:
            print("Erreur lors de la requête POST. Code d'état :", response.status_code)
            print(response.text)  # Affiche la réponse du serveur en cas d'erreur

        if sinistre.prestataire and sinistre.prestataire.code == "9537":
            pass
            # write your request.post data here


        else:
            pass


@login_required
@never_cache
def update_date_sortie_sinistre(request):
    if request.method == 'POST':
        sinistre_id = request.POST.get("id_sinistre")
        date_sortie = request.POST.get("date_sortie")
        sinistre = get_object_or_404(Sinistre, id=sinistre_id)
        sinistre.date_sortie = date_sortie
        sinistre.save()

        response = {
            'statut': 1,
            'message': "La date de sortie a été mise à jour !",
            'data': {
            }
        }
    else:
        response = {
            'statut': 0,
            'message': "Erreur la date de sortie n'a pas été mise à jour :/",
            'data': {
            }
        }

    return JsonResponse(response)


@login_required
@never_cache
def update_date_sortie_nb_jour(request):
    if request.method == 'POST':
        sinistre_id = request.POST.get("id_sinistre")
        date_sortie_encoded = request.POST.get("date_sortie")

        date_object = datetimeJsdecode.strptime(date_sortie_encoded, '%d/%m/%Y')
        # pri(date_sortie)
        formatted_date = date_object.strftime('%Y-%m-%d')
        sinistre = get_object_or_404(Sinistre, id=sinistre_id)
        sinistre.date_sortie = formatted_date
        sinistre.statut = StatutSinistre.ACCORDE
        sinistre.save()

        response = {
            'statut': 1,
            'message': "La date de sortie a été mise à jour !",
            'data': {
            }
        }
    else:
        response = {
            'statut': 0,
            'message': "Erreur la date de sortie n'a pas été mise à jour :/",
            'data': {
            }
        }

    return JsonResponse(response)


@login_required
@never_cache
def update_nombre_accorde_sinistre(request):
    if request.method == 'POST':
        sinistre_id = request.POST.get("id_sinistre")
        nombre_accorde = request.POST.get("nombre_accorde")
        sinistre = get_object_or_404(Sinistre, id=sinistre_id)
        sinistre.nombre_accorde = nombre_accorde
        sinistre.save()

        response = {
            'statut': 1,
            'message': "Le nombre accordé a été mise à jour !",
            'data': {
            }
        }

    else:
        response = {
            'statut': 0,
            'message': "Erreur le nombre accordé n'a pas été mise à jour :/",
            'data': {
            }
        }

    return JsonResponse(response)


@login_required
def get_infos_selected_actes(request):
    if request.method == "POST":

        # pprint(request.POST)

        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        type_prise_en_charge = request.POST.get('type_prise_en_charge')
        type_prise_en_charge_code = request.POST.get('type_prise_en_charge_code')
        actes = request.POST.getlist('selected_actes[]')
        cout_acte = request.POST.get('cout_acte[]').replace(' ', '') if request.POST.get('cout_acte[]') else 0

        if request.user.is_prestataire:
            type_remboursement = 'TP'
            date_survenance = datetime.datetime.now(tz=timezone.utc)
            prestataire_id = request.user.prestataire.id if request.user.prestataire else None

        else:
            type_remboursement = request.POST.get('type_remboursement')
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire)
            prestataire_id = prestataires.first().id if prestataires else None

        cout_acte = 0  # PAS BESOINS DE SAISIR MANUELEMENT LES COUTS POUR LE MOMENT POUR LES CONSULTATION - LES COUTS SONT DEJA ENREGISTRE
        quantite = 1

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        pprint("- Formule à la date du " + str(date_survenance))
        pprint(formule)

        if formule:
            periode_couverture_encours = formule.police.periode_couverture_encours

            pprint(periode_couverture_encours)

            consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                        statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            pprint([float(consommation_individuelle), float(consommation_famille)])

            infos = [
                get_tarif_acte_from_bareme("CONSULT", date_survenance, acte_id, prestataire_id, prescripteur_id,
                                           aliment_id,
                                           cout_acte, quantite, consommation_individuelle, consommation_famille) for
                acte_id
                in actes if acte_id]

            print(infos)
            for acte in infos:
                if acte['statut'] == 0:
                    response = {
                        'statut': 0,
                        'message': acte['message'],
                        'data': {}
                    }
                    return JsonResponse(response)

            total_frais_reel = sum(acte['data']['frais_reel'] for acte in infos if acte['statut'] == 1)
            total_part_assure = sum(acte['data']['part_assure'] for acte in infos if acte['statut'] == 1)
            total_part_compagnie = sum(acte['data']['part_compagnie'] for acte in infos if acte['statut'] == 1)

            # Added on 05102023: add tm préfinancé
            tm_prefinanced = False
            if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
                tm_prefinanced = True if request.user.is_prestataire else False


            return render(request, 'infos_selected_actes.html',
                          {'actes': infos,
                           'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                           'total_part_assure': total_part_assure, })

        else:
            response = {
                'statut': 0,
                'message': "FORMULE DE GARANTIE NON DÉFINI POUR CET ASSURÉ !",
                'data': {}
            }

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

    return JsonResponse(response)


@login_required
def get_infos_selected_actes_consultation_ls(request):
    if request.method == "POST":

        # pprint(request.POST)

        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        type_prise_en_charge = request.POST.get('type_prise_en_charge')
        type_prise_en_charge_code = request.POST.get('type_prise_en_charge_code')
        actes = request.POST.getlist('selected_actes[]')
        cout_acte = request.POST.get('cout_acte').replace(' ', '') if request.POST.get('cout_acte') else 0

        if request.user.is_prestataire:
            type_remboursement = 'TP'
            date_survenance = datetime.datetime.now(tz=timezone.utc)
            prestataire_id = request.user.prestataire.id if request.user.prestataire else None

        else:
            type_remboursement = request.POST.get('type_remboursement')
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire)
            prestataire_id = prestataires.first().id if prestataires else None

        #cout_acte = 0  # PAS BESOINS DE SAISIR MANUELEMENT LES COUTS POUR LE MOMENT POUR LES CONSULTATION - LES COUTS SONT DEJA ENREGISTRE
        quantite = 1

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        pprint("- Formule à la date du " + str(date_survenance))
        pprint(formule)

        if formule:
            periode_couverture_encours = formule.police.periode_couverture_encours

            pprint(periode_couverture_encours)

            consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                        statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

            pprint([float(consommation_individuelle), float(consommation_famille)])

            infos = [
                get_tarif_acte_from_bareme("CONSULT", date_survenance, acte_id, prestataire_id, prescripteur_id,
                                           aliment_id,
                                           cout_acte, quantite, consommation_individuelle, consommation_famille) for
                acte_id
                in actes if acte_id]

            print(infos)
            for acte in infos:
                if acte['statut'] == 0:
                    response = {
                        'statut': 0,
                        'message': acte['message'],
                        'data': {}
                    }
                    return JsonResponse(response)

            total_frais_reel = sum(acte['data']['frais_reel'] for acte in infos if acte['statut'] == 1)
            total_part_assure = sum(0.0 if acte['data']['tm_prefinanced'] else acte['data']['part_assure'] for acte in infos if acte['statut'] == 1)
            total_part_compagnie = sum(acte['data']['frais_reel'] if acte['data']['tm_prefinanced'] else acte['data']['part_compagnie'] for acte in infos if acte['statut'] == 1)

            # Added on 05102023: add tm préfinancé
            tm_prefinanced = False
            if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
                tm_prefinanced = True if request.user.is_prestataire else False

            return render(request, 'infos_selected_actes_consultation_ls.html',
                          {'actes': infos,
                           'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                           'total_part_assure': total_part_assure, })

        else:
            response = {
                'statut': 0,
                'message': "FORMULE DE GARANTIE NON DÉFINI POUR CET ASSURÉ !",
                'data': {}
            }

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

    return JsonResponse(response)


@login_required
def get_infos_selected_actes_soins_ambulatoires(request):
    if request.method == "POST":

        # pprint(request.POST)

        session_pec = request.user.id
        pprint(session_pec)

        # vider toute la session de calcul de prise en charge pour reprendre
        SinistreTemporaire.objects.filter(
            session_pec=session_pec
        ).delete()

        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        actes = request.POST.getlist('selected_actes[]')
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')

        if request.user.is_prestataire:
            type_remboursement = 'TP'
            date_survenance = datetime.datetime.now(tz=timezone.utc)
            prestataire_id = request.user.prestataire.id if request.user.prestataire else None

        else:
            type_remboursement = request.POST.get('type_remboursement')
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire)
            prestataire_id = prestataires.first().id if prestataires else None

        # actes = actes + actes_du_tableau
        actes_data = actes + actes_du_tableau
        # pprint("actes_data")
        # pprint(actes_data)

        cout_acte = 0
        quantite = 1  # A revoir pour les actes avec seance

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

        consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                       adherent_principal_id=aliment.adherent_principal.id,
                                                       statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        infos = []
        for ad in actes_data:
            if ad:
                acte_id, quantite, cout_acte = ad.split(':') if ':' in ad else (ad, 1, 0)

                quantite = quantite.replace(' ', '') if len(str(quantite)) > 3 else quantite
                cout_acte = cout_acte.replace(' ', '') if len(str(cout_acte)) > 3 else cout_acte

                pprint("[acte_id, quantite, cout_acte]")
                pprint([acte_id, quantite, cout_acte])
                acte_info = get_tarif_acte_from_bareme("AMBULAT", date_survenance, acte_id, prestataire_id,
                                                       prescripteur_id, aliment_id, cout_acte, quantite,
                                                       consommation_individuelle, consommation_famille,
                                                       session_pec) if acte_id else None
                infos.append(acte_info)

        pprint(infos)
        for acte in infos:
            if acte and acte['statut'] == 0:
                response = {
                    'statut': 0,
                    'message': acte['message'],
                    'data': {}
                }
                return JsonResponse(response)

        total_frais_reel = sum(acte['data']['frais_reel']*acte['data']['quantite_demande'] for acte in infos if acte and acte['statut'] == 1)
        total_part_assure = sum(acte['data']['part_assure']*acte['data']['quantite_demande'] for acte in infos if acte and acte['statut'] == 1)
        total_part_compagnie = sum(acte['data']['part_compagnie']*acte['data']['quantite_demande'] for acte in infos if acte and acte['statut'] == 1)

        # Added on 05102023: add tm préfinancé
        tm_prefinanced = False
        if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        return render(request, 'infos_selected_actes_soins_ambulatoires.html',
                      {'actes': infos,
                       'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                       'total_part_assure': total_part_assure, })

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

        return JsonResponse(response)


@login_required
def get_infos_selected_actes_hospitalisation(request):
    if request.method == "POST":

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)

        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        actes = request.POST.getlist('selected_actes[]')
        cout_acte = request.POST.get('cout_acte').replace(' ', '')
        nombre_jours = request.POST.get('nombre_jours').replace(' ', '')

        if request.user.is_prestataire:
            type_remboursement = 'TP'
            date_survenance = datetime.datetime.now(tz=timezone.utc)
            prestataire_id = request.user.prestataire.id if request.user.prestataire else None

        else:
            type_remboursement = request.POST.get('type_remboursement')
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire)
            prestataire_id = prestataires.first().id if prestataires else None

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

        consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                       adherent_principal_id=aliment.adherent_principal.id,
                                                       statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        tab_existants = []
        infos = []
        for acte_id in actes:

            if acte_id not in tab_existants:
                tab_existants.append(acte_id)

                # voir si quantité applicable
                acte = Acte.objects.get(id=acte_id)
                quantite = nombre_jours if acte.option_quantite else 1

                # cout_acte = cout_acte.replace(' ', '') if len(str(cout_acte)) > 3 else cout_acte

                acte_info = get_tarif_acte_from_bareme("HOSPIT", date_survenance, acte_id, prestataire_id,
                                                       prescripteur_id, aliment_id, cout_acte, quantite,
                                                       consommation_individuelle, consommation_famille,
                                                       session_pec) if acte_id else None
                infos.append(acte_info)

        # print(infos)
        for acte in infos:
            if acte['statut'] == 0:
                response = {
                    'statut': 0,
                    'message': acte['message'],
                    'data': {}
                }
                return JsonResponse(response)

        total_frais_reel = sum(acte['data']['frais_reel'] for acte in infos if acte['statut'] == 1)
        total_part_assure = sum(acte['data']['part_assure'] for acte in infos if acte['statut'] == 1)
        total_part_compagnie = sum(acte['data']['part_compagnie'] for acte in infos if acte['statut'] == 1)

        # Added on 05102023: add tm préfinancé
        tm_prefinanced = False
        if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        return render(request, 'infos_selected_actes_hospitalisation.html',
                      {'actes': infos,
                       'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                       'total_part_assure': total_part_assure, })

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

        return JsonResponse(response)


@login_required
def get_infos_selected_actes_optiques(request):
    pprint(request)
    if request.method == "POST":

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)
        type_priseencharge_code = request.POST.get('type_prise_en_charge_code')
        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        actes = request.POST.getlist('selected_actes[]')
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')
        pprint("actes_du_tableau")
        pprint(actes_du_tableau)

        if request.user.is_prestataire:
            type_remboursement = 'TP'
            date_survenance = datetime.datetime.now(tz=timezone.utc)
            prestataire_id = request.user.prestataire.id if request.user.prestataire else None

        else:
            type_remboursement = request.POST.get('type_remboursement')
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire)
            prestataire_id = prestataires.first().id if prestataires else None

        # actes = actes + actes_du_tableau
        actes_data = actes + actes_du_tableau
        pprint("actes_data")
        pprint(actes_data)

        cout_acte = 0
        quantite = 1

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

        consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                       adherent_principal_id=aliment.adherent_principal.id,
                                                       statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        tab_existants = []
        infos = []
        for ad in actes_data:
            if ad:
                acte_id, quantite, cout_acte = ad.split(':') if ':' in ad else (ad, 1, 0)

                if acte_id not in tab_existants:
                    tab_existants.append(acte_id)

                    cout_acte = cout_acte.replace(' ', '') if len(str(cout_acte)) > 3 else cout_acte

                    acte_info = get_tarif_acte_from_bareme(type_priseencharge_code, date_survenance, acte_id,
                                                           prestataire_id, prescripteur_id, aliment_id, cout_acte,
                                                           quantite, consommation_individuelle, consommation_famille,
                                                           session_pec) if acte_id else None
                    infos.append(acte_info)

        # pprint(infos)
        for acte in infos:
            if acte and acte['statut'] == 0:
                response = {
                    'statut': 0,
                    'message': acte['message'],
                    'data': {}
                }
                return JsonResponse(response)

        total_frais_reel = sum(acte['data']['frais_reel'] for acte in infos if acte and acte['statut'] == 1)
        total_part_assure = sum(acte['data']['part_assure'] for acte in infos if acte and acte['statut'] == 1)
        total_part_compagnie = sum(acte['data']['part_compagnie'] for acte in infos if acte and acte['statut'] == 1)

        # Added on 05102023: add tm préfinancé
        tm_prefinanced = False
        if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        return render(request, 'infos_selected_actes_optiques.html',
                      {'actes': infos,
                       'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                       'total_part_assure': total_part_assure, })

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

        return JsonResponse(response)


@login_required
def get_infos_selected_actes_optiques_search(request):
    pass


@login_required
def get_infos_selected_actes_gestionnaire(request):
    if request.method == "POST":

        pprint(request.POST)

        session_pec = request.user.id
        pprint(session_pec)

        # vider toute la session de calcul de prise en charge pour reprendre
        SinistreTemporaire.objects.filter(
            session_pec=session_pec
        ).delete()

        # vider la session des medicaments de la consultations
        request.session['liste_medicaments'] = []

        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        type_prise_en_charges = TypePriseencharge.objects.filter(id=type_prise_en_charge_id)
        type_prise_en_charge_code = type_prise_en_charges.first().code if type_prise_en_charges else None

        prescripteur_id = request.POST.get('prescripteur')
        aliment_id = request.POST.get('current_searched_aliment_id')
        actes = request.POST.getlist('selected_actes[]')
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')

        type_remboursement = request.POST.get('type_remboursement')
        date_survenance = request.POST.get('date_survenance')
        date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
        date_survenance = make_aware(date_survenance)

        prestataire = request.POST.get('prestataire_executant')
        prestataires = Prestataire.objects.filter(id=prestataire)
        prestataire_id = prestataires.first().id if prestataires else None

        # actes = actes + actes_du_tableau
        actes_data = actes + actes_du_tableau
        # pprint("actes_data")
        # pprint(actes_data)

        cout_acte = 0
        quantite = 1  # A revoir pour les actes avec seance

        aliment = Aliment.objects.filter(id=aliment_id).first()

        formule = aliment.formule_atdate(date_survenance)
        periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

        consommation_individuelle = \
        Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                       adherent_principal_id=aliment.adherent_principal.id,
                                                       statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        has_consultation_or_hospit = False
        infos = []
        for ad in actes_data:
            if ad:
                acte_id, quantite, cout_acte = ad.split(':') if ':' in ad else (ad, 1, 0)

                quantite = quantite.replace(' ', '') if len(str(quantite)) > 3 else quantite
                cout_acte = cout_acte.replace(' ', '') if len(str(cout_acte)) > 3 else cout_acte

                pprint("[acte_id, quantite, cout_acte]")
                pprint([acte_id, quantite, cout_acte])

                # si l'acte est consult ou hospit
                actes = Acte.objects.filter(id=acte_id)
                acte = actes.first() if actes else None

                acte_info = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                       prestataire_id,
                                                       prescripteur_id, aliment_id, cout_acte, quantite,
                                                       consommation_individuelle, consommation_famille,
                                                       session_pec) if acte_id else None

                # Commented on 22112023
                # TODO Ne marche pas bien: a revoir plus tard, non blocant
                '''#si consultation ou hospit, on choisi un seul acte et on sort de la boucle
                pprint("-----: " + acte.rubrique.libelle)
                if acte.rubrique.libelle == "CONSULTATION" or acte.rubrique.libelle == "HOSPITLISATION":
                    infos = []

                    infos.append(acte_info)
                    has_consultation_or_hospit = True
                    pprint("une consult ou hospit")
                    pprint(infos)

                else:
                    #continue
                    infos.append(acte_info)
                    pprint("acte.rubrique")
                    pprint(acte.rubrique)
                '''

                # appliquer à tout: consultation comme autre:
                infos.append(acte_info)
                pprint("acte.rubrique")
                pprint(acte.rubrique)

        # Commented on 22112023
        # TODO Ne marche pas bien: a revoir plus tard, non blocant
        ''' 
        if has_consultation_or_hospit:
            # Conserver uniquement le dernier élément de la rubrique_code consultation ou hospit, il se trouve dans le select
            # Filter infos based on rubrique code
            reversed_infos = reversed(infos)
            pprint('reversed_infos')
            pprint(reversed_infos)
            filtered_infos = [info for info in infos if info and info.get('data') and info['data'].get('rubrique') in ["CONSULTATION", "HOSPITALISATION"]]
            pprint("filtered_infos")
            pprint(filtered_infos)

            if filtered_infos:
                # If there are elements with the specified rubrique codes, keep only the last one
                infos = [filtered_infos[-1]]
                pprint("last_info")
                pprint(infos)
        '''

        pprint("NOMBRE DE LIGNES")
        # pprint(len(infos))
        # pprint(infos)
        for acte in infos:
            if acte and acte['statut'] == 0:
                response = {
                    'statut': 0,
                    'message': acte['message'],
                    'data': {}
                }
                return JsonResponse(response)

        total_frais_reel = sum(acte['data']['frais_reel']*acte['data']['quantite_demande'] for acte in infos if acte)
        total_part_assure = sum(acte['data']['part_assure']*acte['data']['quantite_demande'] for acte in infos if acte)
        total_part_compagnie = sum(acte['data']['part_compagnie']*acte['data']['quantite_demande'] for acte in infos if acte)

        # Added on 05102023: add tm préfinancé
        tm_prefinanced = False
        if formule.police.type_prefinancement and formule.police.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        # A bien penser et ajouter: si c'est une consultation ajouter le cout des médicaments au total
        '''if type_prise_en_charge_code == "CONSULT":
            # Réccupérer les médicaments en session
            liste_medicaments = request.session.get('liste_medicaments', [])

            total_frais_reel += sum(med['prix_total'] for med in liste_medicaments if med.medicament_id==acte_id)
            total_part_assure += sum(0 for med in liste_medicaments)
            total_part_compagnie += sum(0 for med in liste_medicaments)
        '''

        pprint(infos)
        return render(request, 'infos_selected_actes_gestionnaire.html',
                      {'actes': infos,
                       'total_frais_reel': total_frais_reel, 'total_part_compagnie': total_part_compagnie,
                       'total_part_assure': total_part_assure,
                       'type_prise_en_charge_code': type_prise_en_charge_code})

    else:
        response = {
            'statut': 0,
            'message': "Aucune données !",
            'data': {}
        }

        return JsonResponse(response)


# quand on utilise les actes du multiselect
@login_required
def add_sinistre(request):
    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)
        pprint("DEBUT CONSULTATION")
        pprint(session_pec)
        pprint("DEBUT CONSULTATION")

        pprint(request.POST)

        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        type_prise_en_charge_code = request.POST.get('type_prise_en_charge_code')
        # date_prestation = datetime.datetime.now(tz=timezone.utc)
        date_entree = request.POST.get('date_entree')
        date_sortie = request.POST.get('date_sortie')
        actes = request.POST.getlist('selected_actes[]')
        cout_acte = request.POST.get('cout_acte')
        nombre_jours = request.POST.get('nombre_jours')
        
        soins_a_l_entrange = request.POST.get('soins_a_l_entrange', None)

        prescripteur_id = request.POST.get('prescripteur')
        affection_id = request.POST.get('affection')
        renseignement_clinique = request.POST.get('renseignement_clinique')
        aliment_id = request.POST.get('current_searched_aliment_id')
        aliment = Aliment.objects.get(id=aliment_id)

        if request.user.is_prestataire:
            date_survenance = datetime.datetime.now(tz=timezone.utc)

            prestataire = request.user.prestataire

            type_remboursements = TypeRemboursement.objects.filter(code='TP')
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = False


        else:
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire_executant_id = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire_executant_id)
            prestataire = prestataires.first() if prestataires else None

            type_remboursement_input = request.POST.get('type_remboursement')
            type_remboursements = TypeRemboursement.objects.filter(id=type_remboursement_input)
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = True

        type_prise_en_charge = TypePriseencharge.objects.filter(id=type_prise_en_charge_id).first()
        statut = "ACCORDE" if type_prise_en_charge.code == "CONSULT" else "EN ATTENTE"

        if cout_acte:
            cout_acte = cout_acte.replace(' ', '')
        if nombre_jours:
            nombre_jours = nombre_jours.replace(' ', '')

        if aliment:

            formule = aliment.formule_atdate(date_survenance)

            if formule:

                # Added on 05102023: add tm préfinancé
                type_prefinancement = formule.police.type_prefinancement if formule.police.type_prefinancement else None

                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

                consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                            statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                # son plafond chambre et plafond hospit
                bareme_plafond_chambre = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66023CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                plafond_chambre = bareme_plafond_chambre.first().plafond_acte if bareme_plafond_chambre else 0

                bareme_plafond_hospitalisation = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66027CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                plafond_hospitalisation = bareme_plafond_hospitalisation.first().plafond_acte if bareme_plafond_hospitalisation else 0

                #bareme_plafond_accouchement = Bareme.objects.filter(formulegarantie_id=formule.id, acte__code="G66049CI01", is_garanti=True, date_debut__lte=date_survenance, statut=Statut.ACTIF)
                #plafond_accouchement = bareme_plafond_accouchement.first().plafond_acte if bareme_plafond_accouchement else 0

                if len(actes) > 0:

                    dossier_sinistre = DossierSinistre.objects.create(created_by=request.user,
                                                                      bureau_id=request.user.bureau.id,
                                                                      prestataire_id=prestataire.id,
                                                                      centre_prescripteur_id=prestataire.id,
                                                                      # a confirmer
                                                                      aliment_id=aliment.id,
                                                                      formulegarantie_id=formule.id,
                                                                      police_id=formule.police.id,
                                                                      compagnie_id=formule.police.compagnie.id,
                                                                      prescripteur_id=prescripteur_id,
                                                                      type_priseencharge_id=type_prise_en_charge_id,
                                                                      renseignement_clinique=renseignement_clinique,
                                                                      plafond_chambre=plafond_chambre,
                                                                      plafond_hospit=plafond_hospitalisation,
                                                                      #plafond_accouchement=0, #mettre a jour plus tard: plafond_accouchement,
                                                                      affection_id=affection_id,
                                                                      #type_prefinancement=type_prefinancement,
                                                                      type_remboursement=type_remboursement,
                                                                      of_gestionnaire=of_gestionnaire,
                                                                      date_survenance=date_survenance,
                                                                      soins_a_l_entrange=soins_a_l_entrange

                                                                      )
                    dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre.pk)

                    # créer les sinistres
                    i = 0
                    sinistre_principal_id = None
                    for acte_id in actes:

                        acte = Acte.objects.get(id=acte_id)

                        #Added on 17062024: mise en oeuvre du tpg
                        type_prefinancement = get_type_prefinancement_of_acte(acte, formule)

                        # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                        infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                                prestataire.id, prescripteur_id, aliment_id, cout_acte,
                                                                nombre_jours, consommation_individuelle,
                                                                consommation_famille, session_pec)

                        frais_reel = infos_acte['data']['frais_reel']
                        part_compagnie = infos_acte['data']['part_compagnie']
                        part_assure = infos_acte['data']['part_assure']
                        ticket_moderateur = infos_acte['data']['ticket_moderateur']
                        depassement = infos_acte['data']['depassement']

                        plafond_acte = infos_acte['data']['plafond_acte']
                        nombre_acte = infos_acte['data']['nombre_acte']
                        frequence = infos_acte['data']['frequence']
                        unite_frequence = infos_acte['data']['unite_frequence']
                        garanti = infos_acte['data']['garanti']
                        bareme_id = infos_acte['data']['bareme_id']
                        taux_tm = infos_acte['data']['taux_franchise']

                        statut = "ACCORDE" if (type_prise_en_charge.code == "CONSULT" or acte.entente_prealable == 0) else "EN ATTENTE"
                        if (type_prise_en_charge.code == "CONSULT" or acte.entente_prealable == 0):
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)
                            approuved_by = None
                        else:
                            reviewed_at = None
                            approuved_by = None

                        sinistre = Sinistre.objects.create(type_sinistre="acte",
                                                           created_by=request.user,
                                                           prestataire_id=prestataire.id,
                                                           aliment_id=aliment.id,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           police_id=formule.police.id,
                                                           periode_couverture_id=periode_couverture_encours.pk,
                                                           formulegarantie_id=formule.id,
                                                           bareme_id=bareme_id,
                                                           compagnie_id=formule.police.compagnie.id,
                                                           prescripteur_id=prescripteur_id,
                                                           affection_id=affection_id,
                                                           acte_id=acte_id,
                                                           frais_reel=frais_reel,
                                                           part_compagnie=part_compagnie,
                                                           part_assure=part_assure,
                                                           ticket_moderateur=ticket_moderateur,
                                                           depassement=depassement,
                                                           taux_tm=taux_tm,

                                                           montant_plafond=plafond_acte,
                                                           nombre_plafond=nombre_acte,
                                                           frequence=frequence,
                                                           unite_frequence=unite_frequence,

                                                           montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                           montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie, #sera actualisé au traitement des remboursements sinistre par les gestionnaires
                                                           montant_remboursement_refuse=0, #sera actualisé au traitement des remboursements sinistre par les gestionnaires

                                                           montant_refacture_compagnie=part_compagnie, # sera actualisé au traitement des remboursements sinistre par les gestionnaires,
                                                           montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0, # sera actualisé au traitement des remboursements sinistre par les gestionnaires,

                                                           taux_retenue = get_retenue_selon_contexte(prestataire.id),

                                                           date_survenance=date_survenance,
                                                           date_entree=date_entree,
                                                           date_sortie=date_sortie,
                                                           nombre_demande=nombre_jours,  # en hospit
                                                           statut=statut,
                                                           dossier_sinistre_id=dossier_sinistre.pk,
                                                           type_prefinancement=type_prefinancement,
                                                           reviewed_at=reviewed_at,
                                                           soins_a_l_entrange=soins_a_l_entrange
                                                           )

                        sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                        code_bureau = request.user.bureau.code
                        sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                            sinistre_created.pk).zfill(7) + '-SP'
                        sinistre_created.save()

                        i = i + 1

                        # vider toute la session de calcul de prise en charge pour reprendre
                        vider_sinistres_temporaires(request.user.id)

                # Mettre a jour le statut_pec du dossier
                sinistres_en_attente = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ATTENTE)
                dossier_sinistre.statut_pec = 'EN ATTENTE' if sinistres_en_attente else 'ACCORDE'

                #si hospitalisation (accouchement), mettre à jour le plafond accouchement
                if type_prise_en_charge.code == "HOSPIT":
                    acte = dossier_sinistre.sinistres.first().acte #en hospit c'est un seul acte par dossier
                    bareme_plafond_accouchement = Bareme.objects.filter(formulegarantie_id=formule.id,
                                                                        acte=acte, is_garanti=True,
                                                                        date_debut__lte=date_survenance,
                                                                        statut=Statut.ACTIF)
                    plafond_accouchement = bareme_plafond_accouchement.first().plafond_acte if bareme_plafond_accouchement else 0
                    dossier_sinistre.plafond_accouchement = plafond_accouchement


                dossier_sinistre.numero = 'BS' + str(Date.today().year) + str(dossier_sinistre.id).zfill(6)
                dossier_sinistre.save()

                # supprimer tout sinistre temporaire portant l'acte et l'assuré, avant de recréer le sinistre
                vider_sinistres_temporaires(session_pec)

                pprint("FIN CONSULTATION")
                pprint(session_pec)
                pprint("FIN CONSULTATION")

                # afficher le bon
                pprint("afficher le bon")


                numero_telephone_sms = aliment.adherent_principal.telephone_mobile
                if aliment.adherent_principal.sms_active and numero_telephone_sms:
                    numero_telephone_sms = numero_telephone_sms.replace(" ", "").replace("-", "").replace("/","").replace("(", "").replace(")", "").replace("+", "")

                    if len(numero_telephone_sms) >= 10:
                        date_survenance_fr = date_survenance.strftime('%d/%m/%Y à %H:%M')
                        message = "Un acte de " + sinistre.acte.libelle + " vient d’être effectué à " + sinistre.prestataire.name + ".\n Date : " + date_survenance_fr + " \n PATIENT : " + str(
                            sinistre.aliment) + " \n N° CARTE : " + sinistre.aliment.carte_active().numero + " \n INOV vous souhaite un prompt rétablissement."
                        statut_sms = api_send_sms(message, numero_telephone_sms)


                if type_prise_en_charge.code == "CONSULT":
                    pprint("cas d'une consultation")

                    pprint("afficher le bon")
                    if request.user.bureau.type_bon_consultation == TypeBonConsultation.AUTO_CARBONE:
                        return render_pdf_view(request, dossier_sinistre.pk)
                    else:
                        return render_pdf_view_general(request, dossier_sinistre.pk)

                else:

                    pprint("cas autre que consultation: afficher le détails du dossier")

                    return redirect(
                        reverse('details_dossier_sinistre', kwargs={'dossier_sinistre_id': dossier_sinistre.pk}))



            else:

                response = {
                    'statut': 0,
                    'message': "Formule non trouvé !",
                    'data': {}
                }

        else:

            response = {
                'statut': 0,
                'message': "Assuré non trouvé !",
                'data': {
                }
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(session_pec)

        return JsonResponse(response)

    else:

        return redirect('/')


def add_sinistre_waspito(sinistre):
    # dossier_sinistre = DossierSinistre.objects.create(sinistre)

    # dd(sinistre)
    return sinistre


@login_required
def update_sinistre_hospitalisation(request, dossier_sinistre_id):
    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)
        cout_acte = request.POST.get('cout_final_sinistre')

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)
        type_prise_en_charge_code = dossier_sinistre.type_priseencharge.code

        sinistre = Sinistre.objects.get(dossier_sinistre_id=dossier_sinistre_id)
        prescripteur = sinistre.prescripteur

        if request.user.is_prestataire:
            date_survenance = datetime.datetime.now(tz=timezone.utc)

            prestataire = request.user.prestataire

            type_remboursements = TypeRemboursement.objects.filter(code='TP')
            type_remboursement = type_remboursements.first() if type_remboursements else None

        else:
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire_executant_id = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire_executant_id)
            prestataire = prestataires.first() if prestataires else None

            type_remboursement_input = request.POST.get('type_remboursement')
            type_remboursements = TypeRemboursement.objects.filter(id=type_remboursement_input)
            type_remboursement = type_remboursements.first() if type_remboursements else None

        if cout_acte:
            cout_acte = cout_acte.replace(' ', '')

        aliment = dossier_sinistre.aliment
        acte = sinistre.acte

        if aliment:

            if aliment.formule:

                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = aliment.formule.police.periode_couverture_encours
                consommation_individuelle = Sinistre.objects.filter(
                    periode_couverture_id=periode_couverture_encours.pk,
                    aliment_id=aliment.id,
                    statut=StatutSinistre.ACCORDE,
                    statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT],
                    statut_validite=StatutValidite.VALIDE
                ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                consommation_famille = Sinistre.objects.filter(
                    periode_couverture_id=periode_couverture_encours.pk,
                    adherent_principal_id=aliment.adherent_principal.id,
                    statut=StatutSinistre.ACCORDE,
                    statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT],
                    statut_validite=StatutValidite.VALIDE
                ).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                formule = aliment.formule
                # son plafond chambre et plafond hospit
                bareme_plafond_chambre = Bareme.objects.filter(formulegarantie_id=aliment.formule.id,
                                                               acte__code="G66023CI01")
                plafond_chambre = bareme_plafond_chambre.first().plafond_acte if bareme_plafond_chambre else 0
                bareme_plafond_hospitalisation = Bareme.objects.filter(formulegarantie_id=aliment.formule.id,
                                                                       acte__code="G66027CI01")
                plafond_hospitalisation = bareme_plafond_hospitalisation.first().plafond_acte if bareme_plafond_hospitalisation else 0

                # créer les sinistres
                # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                quantite = 1
                infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte.id,
                                                        prestataire.id, prescripteur.id, aliment.id, cout_acte,
                                                        quantite, consommation_individuelle,
                                                        consommation_famille, session_pec)

                if infos_acte['statut'] == 1:
                    frais_reel = infos_acte['data']['frais_reel']
                    part_compagnie = infos_acte['data']['part_compagnie']
                    part_assure = infos_acte['data']['part_assure']
                    ticket_moderateur = infos_acte['data']['ticket_moderateur']
                    depassement = infos_acte['data']['depassement']

                    plafond_acte = infos_acte['data']['plafond_acte']
                    nombre_acte = infos_acte['data']['nombre_acte']
                    frequence = infos_acte['data']['frequence']
                    unite_frequence = infos_acte['data']['unite_frequence']
                    garanti = infos_acte['data']['garanti']
                    bareme_id = infos_acte['data']['bareme_id']

                    type_prefinancement = sinistre.type_prefinancement

                    Sinistre.objects.filter(id=sinistre.id).update(updated_price_by=request.user,
                                                                   bareme_id=bareme_id,
                                                                   compagnie_id=aliment.formule.police.compagnie.id,
                                                                   frais_reel=frais_reel,
                                                                   part_compagnie=part_compagnie,
                                                                   part_assure=part_assure,
                                                                   ticket_moderateur=ticket_moderateur,
                                                                   depassement=depassement,

                                                                   montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                                   montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                                   montant_remboursement_refuse=0,

                                                                   montant_refacture_compagnie=part_compagnie,
                                                                   montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0,

                                                                   montant_plafond=plafond_acte,
                                                                   nombre_plafond=nombre_acte,
                                                                   frequence=frequence,
                                                                   unite_frequence=unite_frequence,
                                                                   )

                    response = {
                        'statut': 1,
                        'message': "Le coût total de l'hospitalisation a été mis à jour !",
                        'data': {}
                    }

                else:

                    response = {
                        'statut': 0,
                        'message': infos_acte['message'],
                        'data': {}
                    }


            else:

                response = {
                    'statut': 0,
                    'message': "Formule non trouvé !",
                    'data': {}
                }

        else:

            response = {
                'statut': 0,
                'message': "Assuré non trouvé !",
                'data': {
                }
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

        return JsonResponse(response)

    else:

        return redirect('/')


# demande de prorogation
@login_required
def demande_prorogation(request, sinistre_id):
    if request.method == 'POST':

        pprint(request.POST)

        jour_demande = request.POST.get('nombre_jours_prorogation')
        motif = request.POST.get('motif_prorogation')

        date_entree = datetime.datetime.now(tz=timezone.utc)
        date_sortie = datetime.datetime.now(tz=timezone.utc)

        ProrogationSinistre.objects.create(created_by=request.user,
                                           sinistre_id=sinistre_id,
                                           jour_demande=jour_demande,
                                           jour_accorde=0,
                                           motif_demande=motif,
                                           date_entree=date_entree,
                                           date_sortie=date_sortie,
                                           )

        # mettre à jour le statut_prorogation du dossier_sinistre
        sinistre = Sinistre.objects.get(id=sinistre_id)
        dossier_sinistre = sinistre.dossier_sinistre
        dossier_sinistre.statut_prorogation = StatutSinistre.ATTENTE
        dossier_sinistre.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de prorogation effectuée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
def approuver_prorogation(request):
    if request.method == 'POST':

        pprint(request.POST)

        prorogation_id = request.POST.get('prorogation_id')
        jour_accorde = request.POST.get('jour_accorde')
        motif = request.POST.get('motif')

        date_entree = datetime.datetime.now(tz=timezone.utc)
        date_sortie = datetime.datetime.now(tz=timezone.utc)

        ProrogationSinistre.objects.filter(id=prorogation_id).update(reviewed_by=request.user,
                                                                     jour_accorde=jour_accorde,
                                                                     motif_rejet=motif,
                                                                     statut=StatutSinistre.ACCORDE
                                                                     )

        # mettre à jour le statut_prorogation du dossier_sinistre

        prorogation_sinistre = ProrogationSinistre.objects.get(id=prorogation_id)
        sinistre = prorogation_sinistre.sinistre
        dossier_sinistre = sinistre.dossier_sinistre

        # si pas d'autre demandes de prorogation en attente: le dossier d'hospit a un seul sinistre
        prorogations_en_attentes = ProrogationSinistre.objects.filter(sinistre=sinistre, statut=StatutSinistre.ATTENTE)
        if not prorogations_en_attentes:
            dossier_sinistre.statut_prorogation = StatutSinistre.ACCORDE
            dossier_sinistre.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de prorogation approuvée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
def rejeter_prorogation(request):
    if request.method == 'POST':

        pprint(request.POST)

        prorogation_id = request.POST.get('prorogation_id')
        motif = request.POST.get('motif')

        date_entree = datetime.datetime.now(tz=timezone.utc)
        date_sortie = datetime.datetime.now(tz=timezone.utc)

        ProrogationSinistre.objects.filter(id=prorogation_id).update(reviewed_by=request.user,
                                                                     jour_accorde=0,
                                                                     motif_rejet=motif,
                                                                     statut=StatutSinistre.REJETE
                                                                     )

        # mettre à jour le statut_prorogation du dossier_sinistre
        prorogation_sinistre = ProrogationSinistre.objects.get(id=prorogation_id)
        sinistre = prorogation_sinistre.sinistre
        dossier_sinistre = sinistre.dossier_sinistre

        # si pas d'autre demandes de prorogation en attente: le dossier d'hospit a un seul sinistre
        prorogations_en_attentes = ProrogationSinistre.objects.filter(sinistre=sinistre, statut=StatutSinistre.ATTENTE)
        if not prorogations_en_attentes:
            dossier_sinistre.statut_prorogation = StatutSinistre.ACCORDE
            dossier_sinistre.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de prorogation rejetée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


# TODO : accepter_remboursement
@login_required
def accepter_remboursement(request, sinistre_id):
    if request.method == 'POST':

        pprint(request.POST)

        sinistre = Sinistre.objects.get(id=sinistre_id)

        motif = request.POST.get('motif')
        montant_accepte = request.POST.get('montant_accepte').replace(' ', '')
        montant_refuse = request.POST.get('montant_refuse').replace(' ', '')
                    
        # if montant_accepte == '': montant_accepte = 0
        if montant_refuse == '': montant_refuse = '0'

        montant_refuse = int(montant_refuse)
        if sinistre.tm_prefinanced:
            montant_accepte = int(sinistre.frais_reel) - int(montant_refuse)
        else:
            montant_accepte = int(sinistre.part_compagnie) - int(montant_refuse)

        montant_tps=int((montant_accepte*sinistre.taux_retenue)/100) if sinistre.taux_retenue and sinistre.facture_prestataire.type_remboursement.code == "TP" else 0

        rb_sinistre = RemboursementSinistre.objects.filter(sinistre=sinistre, is_invalid=False).first()

        if rb_sinistre is None:
            # save remboursement sinistre accepted
            RemboursementSinistre.objects.create(created_by=request.user,
                                                 designation=DesignationRemboursementSinistre.NET_A_PAYER,
                                                 sinistre=sinistre,
                                                 montant=montant_accepte,
                                                 statut=StatutRemboursementSinistre.ACCEPTE)
            # save remboursement sinistre refused
            RemboursementSinistre.objects.create(created_by=request.user,
                                                 designation=DesignationRemboursementSinistre.MONTANT_REFUSE,
                                                 sinistre=sinistre,
                                                 montant=montant_refuse,
                                                 motif=motif,
                                                 statut=StatutRemboursementSinistre.REFUSE)
            if sinistre.taux_retenue is not None and sinistre.facture_prestataire.type_remboursement.code == "TP": 
                RemboursementSinistre.objects.create(created_by=request.user,
                                                    designation=DesignationRemboursementSinistre.TAXT,
                                                    sinistre=sinistre,
                                                    montant=montant_tps,
                                                    motif=motif,
                                                    statut=StatutRemboursementSinistre.TAXT)
            # Added on 04052024:mettre à jour le montant_remboursement_accepte
            sinistre.montant_remboursement_accepte = montant_accepte
            sinistre.montant_remboursement_refuse = montant_refuse
            sinistre.tps = montant_tps 

            # Added on 07062024: mettre à jour le statut remboursement dont on tiendra compte dans le calcul de la conso
            if montant_accepte == 0:
                sinistre.statut_remboursement = StatutRemboursement.REFUSE
            elif montant_refuse == 0:
                sinistre.statut_remboursement = StatutRemboursement.ACCEPTE
            else:
                sinistre.statut_remboursement = StatutRemboursement.ACCEPTE_PARTIELLEMENT

            sinistre.is_ges_processed = True
            sinistre.save()


            #si tm_prefinancé, recalculer les parts à refacturer
            recalcule_montant_refacture_compagnie_et_client(sinistre)


        # notifier du succès
        response = {
            'statut': 1,
            'message': f'Demande de remboursement acceptée avec succès !',
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


# TODO : refuser_remboursement
@login_required
def refuser_remboursement(request, sinistre_id):
    if request.method == 'POST':

        pprint(request.POST)

        motif = request.POST.get('motif')

        sinistre = Sinistre.objects.get(id=sinistre_id)

        if sinistre.tm_prefinanced:
            montant_refuse = int(sinistre.frais_reel)
        else:
            montant_refuse = int(sinistre.part_compagnie)

        rb_sinistre = RemboursementSinistre.objects.filter(sinistre=sinistre, is_invalid=False).first()

        if rb_sinistre is None:
            # save remboursement sinistre accepted
            RemboursementSinistre.objects.create(created_by=request.user,
                                                 designation=DesignationRemboursementSinistre.NET_A_PAYER,
                                                 sinistre=sinistre,
                                                 montant=0,
                                                 statut=StatutRemboursementSinistre.ACCEPTE)
            # save remboursement sinistre refused
            RemboursementSinistre.objects.create(created_by=request.user,
                                                 designation=DesignationRemboursementSinistre.MONTANT_REFUSE,
                                                 sinistre=sinistre,
                                                 montant=montant_refuse,
                                                 motif=motif,
                                                 statut=StatutRemboursementSinistre.REFUSE)

            # Added on 04052024:mettre à jour le montant_remboursement_accepte
            sinistre.montant_remboursement_accepte = 0
            sinistre.montant_remboursement_refuse = montant_refuse

            #Added on 07062024: mettre à jour le statut remboursement dont on tiendra compte dans le calcul de la conso
            sinistre.statut_remboursement = StatutRemboursement.REFUSE

            sinistre.is_ges_processed = True
            sinistre.save()

            # si tm_prefinancé, recalculer les parts à refacturer
            recalcule_montant_refacture_compagnie_et_client(sinistre)


        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement rejetée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
@never_cache
def traiter_liste_remboursement(request):
    if request.method == 'POST':

        pprint(request.POST)
        sinistres_ids = request.POST.getlist('sinistres_ids[]')
        pprint(sinistres_ids)

        for sinistre_id in sinistres_ids:
            sinistre = Sinistre.objects.get(id=sinistre_id)

            motif = "RAS"
            if sinistre.tm_prefinanced:
                montant_accepte = int(sinistre.frais_reel)
            else:
                montant_accepte = int(sinistre.part_compagnie)

            sinistre = Sinistre.objects.get(id=sinistre_id)

            rb_sinistre = RemboursementSinistre.objects.filter(sinistre=sinistre, is_invalid=False).first()

            if rb_sinistre is None:

                montant_tps=int((montant_accepte*sinistre.taux_retenue)/100) if sinistre.taux_retenue and sinistre.facture_prestataire.type_remboursement.code == "TP" else 0

                # save remboursement sinistre accepted
                RemboursementSinistre.objects.create(created_by=request.user,
                                                     designation=DesignationRemboursementSinistre.NET_A_PAYER,
                                                     sinistre=sinistre,
                                                     montant=montant_accepte,
                                                     motif=motif,
                                                     statut=StatutRemboursementSinistre.ACCEPTE)
                if sinistre.taux_retenue is not None and sinistre.facture_prestataire.type_remboursement.code == "TP":
                    RemboursementSinistre.objects.create(created_by=request.user,
                                                        designation=DesignationRemboursementSinistre.TAXT,
                                                        sinistre=sinistre,
                                                        montant=montant_tps,
                                                        motif=motif,
                                                        statut=StatutRemboursementSinistre.TAXT)

                # Added on 04052024:mettre à jour le montant_remboursement_accepte
                sinistre.montant_remboursement_accepte = montant_accepte
                sinistre.montant_remboursement_refuse = 0
                sinistre.tps = montant_tps

                # Added on 07062024: mettre à jour le statut remboursement dont on tiendra compte dans le calcul de la conso
                sinistre.statut_remboursement = StatutRemboursement.ACCEPTE

                sinistre.is_ges_processed = True
                sinistre.save()

                # si tm_prefinancé, recalculer les parts à refacturer
                recalcule_montant_refacture_compagnie_et_client(sinistre)

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement acceptée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
@never_cache
def refuser_liste_remboursement(request):
    if request.method == 'POST':

        pprint(request.POST)
        sinistres_ids = request.POST.getlist('sinistres_ids[]')
        pprint(sinistres_ids)
        motif = request.POST.get('motif')
        pprint(motif)

        for sinistre_id in sinistres_ids:
            sinistre = Sinistre.objects.get(id=sinistre_id)

            # motif = "RAS"
            if sinistre.tm_prefinanced:
                montant_refuser = int(sinistre.frais_reel)
            else:
                montant_refuser = int(sinistre.part_compagnie)

            sinistre = Sinistre.objects.get(id=sinistre_id)

            rb_sinistre = RemboursementSinistre.objects.filter(sinistre=sinistre, is_invalid=False).first()
            if rb_sinistre is None:
                # save remboursement sinistre accepted
                RemboursementSinistre.objects.create(created_by=request.user,
                                                     designation=DesignationRemboursementSinistre.NET_A_PAYER,
                                                     sinistre=sinistre,
                                                     montant=0,
                                                     statut=StatutRemboursementSinistre.ACCEPTE)

                # save remboursement sinistre refused
                RemboursementSinistre.objects.create(created_by=request.user,
                                                     designation=DesignationRemboursementSinistre.MONTANT_REFUSE,
                                                     sinistre=sinistre,
                                                     montant=montant_refuser,
                                                     motif=motif,
                                                     statut=StatutRemboursementSinistre.REFUSE)

                # Added on 04052024:mettre à jour le montant_remboursement_accepte
                sinistre.montant_remboursement_accepte = 0
                sinistre.montant_remboursement_refuse = montant_refuser

                # Added on 07062024: mettre à jour le statut remboursement dont on tiendra compte dans le calcul de la conso
                sinistre.statut_remboursement = StatutRemboursement.REFUSE

                sinistre.is_ges_processed = True
                sinistre.save()

                # si tm_prefinancé, recalculer les parts à refacturer
                recalcule_montant_refacture_compagnie_et_client(sinistre)

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement acceptée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
@never_cache
def valider_facture_remboursement(request, facture_prestataire_id):
    if request.method == 'GET':

        facture_prestataire = FacturePrestataire.objects.get(id=facture_prestataire_id)

        facture_prestataire.statut = SatutBordereauDossierSinistres.VALIDE
        facture_prestataire.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement acceptée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


#
@login_required
def add_sinistre_soins_ambulatoire(request):
    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)

        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        type_prise_en_charge_code = "AMBULAT"
        # date_prestation = datetime.datetime.now(tz=timezone.utc)
        # date_survenance = datetime.datetime.now(tz=timezone.utc)
        date_entree = request.POST.get('date_entree')
        date_sortie = request.POST.get('date_sortie')
        # actes = request.POST.getlist('selected_actes[]')
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')
        actes = request.POST.getlist('selected_acte_info[]')
        nombre_seances = request.POST.getlist('nombre_seance[]')
        couts_actes = request.POST.getlist('cout_acte[]')
        # prestataire = request.user.prestataire
        prescripteur_id = request.POST.get('prescripteur')
        centre_prescripteur_id = request.POST.get('centre_prescripteur')
        affection_id = request.POST.get('affection')
        renseignement_clinique = request.POST.get('renseignement_clinique')
        aliment_id = request.POST.get('current_searched_aliment_id')
        aliment = Aliment.objects.get(id=aliment_id)
        pprint(nombre_seances)
        
        soins_a_l_entrange = request.POST.get('soins_a_l_entrange', None)

        # Added on 15112023: si date entrée non remseigné,
        if date_entree == '': date_entree = None
        if date_sortie == '': date_sortie = None

        if request.user.is_prestataire:
            date_survenance = datetime.datetime.now(tz=timezone.utc)

            prestataire = request.user.prestataire

            type_remboursements = TypeRemboursement.objects.filter(code='TP')
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = False

        else:
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire_executant_id = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire_executant_id)
            prestataire = prestataires.first() if prestataires else None

            type_remboursement_input = request.POST.get('type_remboursement')
            type_remboursements = TypeRemboursement.objects.filter(id=type_remboursement_input)
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = True

        # actes = actes_du_tableau

        if aliment:

            formule = aliment.formule_atdate(date_survenance)
            if formule:

                # Added on 05102023: add tm préfinancé
                #type_prefinancement = formule.police.type_prefinancement if formule.police.type_prefinancement else None

                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)
                consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                            statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                # vérifier si des actes ont étés envoyer pour êtree effectuer
                if len(actes) > 0:

                    # crée le dossier sinistre
                    dossier_sinistre = DossierSinistre.objects.create(created_by=request.user,
                                                                      bureau_id=request.user.bureau.id,
                                                                      prestataire_id=prestataire.id,
                                                                      centre_prescripteur_id=centre_prescripteur_id,
                                                                      aliment_id=aliment.id,
                                                                      formulegarantie_id=formule.id,
                                                                      police_id=formule.police.id,
                                                                      compagnie_id=formule.police.compagnie.id,
                                                                      prescripteur_id=prescripteur_id,
                                                                      type_priseencharge_id=type_prise_en_charge_id,
                                                                      renseignement_clinique=renseignement_clinique,
                                                                      affection_id=affection_id,
                                                                      #type_prefinancement=type_prefinancement,
                                                                      type_remboursement=type_remboursement,
                                                                      of_gestionnaire=of_gestionnaire,
                                                                      date_survenance=date_survenance,
                                                                      soins_a_l_entrange=soins_a_l_entrange
                                                                      )
                    dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre.pk)

                    # créer les sinistres
                    i = 0
                    sinistre_principal_id = None
                    for acte_id in actes:

                        acte = Acte.objects.get(id=acte_id)

                        # Added on 17062024: mise en oeuvre du tpg
                        type_prefinancement = get_type_prefinancement_of_acte(acte, formule)

                        nombre_demande = int(nombre_seances[i])
                        cout_acte = 0
                        # si actes d'optique ou dentaire dont le cout est saisi manuelement
                        cout_acte = int(couts_actes[i].replace(' ', '')) if len(str(couts_actes[i])) > 3 else int(
                            couts_actes[i])
                        quantite = nombre_demande

                        # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                        infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                                prestataire.id, prescripteur_id, aliment_id, cout_acte,
                                                                quantite,
                                                                consommation_individuelle, consommation_famille,
                                                                session_pec)

                        '''if acte.option_seance:
                            frais_reel = infos_acte['data']['frais_reel'] / int(nombre_demande)
                            part_compagnie = infos_acte['data']['part_compagnie'] / int(nombre_demande)
                            part_assure = infos_acte['data']['part_assure'] / int(nombre_demande)
                            ticket_moderateur = infos_acte['data']['ticket_moderateur'] / int(nombre_demande)
                            depassement = infos_acte['data']['depassement'] / int(nombre_demande)
                        else:'''
                        frais_reel = infos_acte['data']['frais_reel']
                        part_compagnie = infos_acte['data']['part_compagnie']
                        part_assure = infos_acte['data']['part_assure']
                        ticket_moderateur = infos_acte['data']['ticket_moderateur']
                        depassement = infos_acte['data']['depassement']

                        plafond_acte = infos_acte['data']['plafond_acte']
                        nombre_acte = infos_acte['data']['nombre_acte']
                        frequence = infos_acte['data']['frequence']
                        unite_frequence = infos_acte['data']['unite_frequence']
                        garanti = infos_acte['data']['garanti']
                        bareme_id = infos_acte['data']['bareme_id']
                        taux_tm = infos_acte['data']['taux_franchise']

                        statut = "ACCORDE" if acte.entente_prealable == 0 else "EN ATTENTE"
                        if acte.entente_prealable == 0:
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)
                            approuved_by = None
                        else:
                            reviewed_at = None
                            approuved_by = None

                        # Added on 15112023: si le gestionnaire qui saisie, on accorde tous par défaut et il vont gérer au remboursement
                        if not request.user.is_prestataire:
                            statut = "ACCORDE"
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)

                        sinistre = Sinistre.objects.create(type_sinistre="acte",
                                                           created_by=request.user,
                                                           prestataire_id=prestataire.id,
                                                           aliment_id=aliment.id,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           police_id=formule.police.id,
                                                           periode_couverture_id=periode_couverture_encours.pk,
                                                           formulegarantie_id=formule.id,
                                                           bareme_id=bareme_id,
                                                           compagnie_id=formule.police.compagnie.id,
                                                           prescripteur_id=prescripteur_id,
                                                           affection_id=affection_id,
                                                           acte_id=acte_id,
                                                           frais_reel=frais_reel,
                                                           part_compagnie=part_compagnie,
                                                           part_assure=part_assure,
                                                           ticket_moderateur=ticket_moderateur,
                                                           depassement=depassement,
                                                           taux_tm=taux_tm,

                                                           montant_plafond=plafond_acte,
                                                           nombre_plafond=nombre_acte,
                                                           frequence=frequence,
                                                           unite_frequence=unite_frequence,

                                                           nombre_demande=nombre_demande,
                                                           nombre_accorde=0,

                                                           montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                           montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                           montant_remboursement_refuse=0,

                                                           montant_refacture_compagnie=part_compagnie,
                                                           montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0,

                                                           taux_retenue = get_retenue_selon_contexte(prestataire.id),

                                                           date_survenance=date_survenance,
                                                           date_entree=date_entree,
                                                           date_sortie=date_sortie,
                                                           statut=statut,
                                                           dossier_sinistre_id=dossier_sinistre.pk,
                                                           type_prefinancement=type_prefinancement,
                                                           reviewed_at=reviewed_at,
                                                           soins_a_l_entrange=soins_a_l_entrange
                                                           )

                        sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                        code_bureau = request.user.bureau.code
                        sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                            sinistre_created.pk).zfill(7) + '-SP'
                        sinistre_created.save()

                        #added on 05102024: si acte avec séance et non soumis à entente préalable, dupliquer le sinistre qui vient d'êrre crée
                        # si acte avec séance dupliquer au nombre demandé
                        if acte.option_seance and acte.entente_prealable == 0:
                            dupliquer = True
                            pprint("DUPLIQUER LE SINISTRE SELON LE NOMBRE DEMANDE")
                            cpt = 1  # vu qu'il y a déjà une ligne de séance sur laquelle on est
                            nombre_total = int(sinistre_created.nombre_demande)

                            for _ in range(int(sinistre_created.nombre_demande) - 1):
                                new_sinistre = deepcopy(sinistre_created)  # Copie de l'objet

                                new_sinistre.id = None
                                new_sinistre.numero = None
                                new_sinistre.nombre_demande = 1
                                new_sinistre.nombre_accorde = 1
                                new_sinistre.date_survenance = None
                                new_sinistre.statut_prestation = StatutSinistrePrestation.ATTENTE
                                # new_sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                                # new_sinistre.approuved_by = request.user
                                new_sinistre.save()

                                # générer le numéro
                                code_bureau = sinistre_created.dossier_sinistre.bureau.code
                                new_sinistre.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                                    new_sinistre.pk).zfill(7) + '-SP'
                                new_sinistre.save()

                                cpt += 1

                                pprint("cpt")
                                pprint(cpt)
                                pprint("sinistre.nombre_demande")
                                pprint(sinistre.nombre_demande)


                        i = i + 1

                        # vider toute la session de calcul de prise en charge pour reprendre
                        vider_sinistres_temporaires(request.user.id)


                # mettre a jour le total frais reel
                # Mettre a jour le statut_pec du dossier
                sinistres_en_attente = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ATTENTE)
                dossier_sinistre.statut_pec = 'EN ATTENTE' if sinistres_en_attente else 'ACCORDE'

                dossier_sinistre.numero = 'BS' + str(Date.today().year) + str(dossier_sinistre.pk).zfill(6)
                dossier_sinistre.save()


                # Envoi du SMS d'attente prealable #SMS
                if sinistres_en_attente:
                    numero_telephone_sms = aliment.adherent_principal.telephone_mobile
                    if aliment.adherent_principal.sms_active and numero_telephone_sms:
                        numero_telephone_sms = numero_telephone_sms.replace(" ", "").replace("-", "").replace("/","").replace("(", "").replace(")", "")

                        if len(numero_telephone_sms) >= 10:
                            date_survenance_fr = date_survenance.strftime('%d/%m/%Y à %H:%M')
                            message = "Une demande d'accord préalable vient d'être effectué à " + sinistre.prestataire.name + ".\n Date : " + date_survenance_fr + " \n PATIENT : " + str(
                                sinistre.aliment) + " \n N° CARTE : " + sinistre.aliment.carte_active().numero + " \n INOV vous souhaite un prompt rétablissement."
                            statut_sms = api_send_sms(message, numero_telephone_sms)

                            if statut_sms:
                                pprint("SMS demande préalable envoyé ")
                            else:
                                pprint("SMS demande préalable non envoyé ")


                # Imprimer le bon

                # notifier du succès
                response = {
                    'statut': 1,
                    'message': "Enregistrement effectuée avec succès !",
                    'data': sinistre_created.numero
                }

                # afficher le bon
                return redirect(
                    reverse('details_dossier_sinistre', kwargs={'dossier_sinistre_id': dossier_sinistre.pk}))

            else:

                response = {
                    'statut': 0,
                    'message': "Formule non trouvé !",
                    'data': {}
                }

        else:

            response = {
                'statut': 0,
                'message': "Assuré non trouvé !",
                'data': {
                }
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

        return JsonResponse(response)

    else:

        return redirect('/')


@login_required
def add_sinistre_optique(request):
    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)

        type_prise_en_charge_code = request.POST.get('type_prise_en_charge_code')
        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        date_prestation = datetime.datetime.now(tz=timezone.utc)
        date_survenance = datetime.datetime.now(tz=timezone.utc)
        date_entree = request.POST.get('date_entree')
        date_sortie = request.POST.get('date_sortie')
        # actes = request.POST.getlist('selected_actes[]')
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')
        actes = request.POST.getlist('selected_acte_info[]')
        nombre_seances = request.POST.getlist('nombre_seance[]')
        couts_actes = request.POST.getlist('cout_acte[]')
        prestataire = request.user.prestataire
        prescripteur_id = request.POST.get('prescripteur')
        centre_prescripteur_id = request.POST.get('centre_prescripteur')
        affection_id = request.POST.get('affection')
        renseignement_clinique = request.POST.get('renseignement_clinique')
        aliment_id = request.POST.get('current_searched_aliment_id')
        aliment = Aliment.objects.get(id=aliment_id)
        soins_a_l_entrange = request.POST.get('soins_a_l_entrange', None)

        if request.user.is_prestataire:
            date_survenance = datetime.datetime.now(tz=timezone.utc)

            prestataire = request.user.prestataire

            type_remboursements = TypeRemboursement.objects.filter(code='TP')
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = False

        else:
            date_survenance = request.POST.get('date_survenance')
            date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d')
            date_survenance = make_aware(date_survenance)

            prestataire_executant_id = request.POST.get('prestataire_executant')
            prestataires = Prestataire.objects.filter(id=prestataire_executant_id)
            prestataire = prestataires.first() if prestataires else None

            type_remboursement_input = request.POST.get('type_remboursement')
            type_remboursements = TypeRemboursement.objects.filter(id=type_remboursement_input)
            type_remboursement = type_remboursements.first() if type_remboursements else None

            of_gestionnaire = True

        # actes = actes_du_tableau

        if aliment:

            formule = aliment.formule_atdate(date_survenance)
            if formule:

                # Added on 05102023: add tm préfinancé
                #type_prefinancement = formule.police.type_prefinancement if formule.police.type_prefinancement else None

                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)
                print("periode_couverture_encours")
                print(periode_couverture_encours)
                consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                            statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                if len(actes) > 0:

                    # Added on 19102023: changer les imageries, optiques, dentaires en soins ambulatoire

                    if type_prise_en_charge_code != "OPTIQUE":
                        type_priseencharges = TypePriseencharge.objects.filter(code="AMBULAT")
                        type_prise_en_charge_id = type_priseencharges.first().id if type_priseencharges else None

                    # crée le dossier sinistre
                    dossier_sinistre = DossierSinistre.objects.create(created_by=request.user,
                                                                      bureau_id=request.user.bureau.id,
                                                                      prestataire_id=prestataire.id,
                                                                      centre_prescripteur_id=centre_prescripteur_id,
                                                                      aliment_id=aliment.id,
                                                                      formulegarantie_id=formule.id,
                                                                      police_id=formule.police.id,
                                                                      compagnie_id=formule.police.compagnie.id,
                                                                      prescripteur_id=prescripteur_id,
                                                                      type_priseencharge_id=type_prise_en_charge_id,
                                                                      renseignement_clinique=renseignement_clinique,
                                                                      affection_id=affection_id,
                                                                      #type_prefinancement=type_prefinancement,
                                                                      type_remboursement=type_remboursement,
                                                                      of_gestionnaire=of_gestionnaire,
                                                                      date_survenance=date_survenance,
                                                                      soins_a_l_entrange=soins_a_l_entrange
                                                                      )
                    dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre.pk)

                    # créer les sinistres
                    i = 0
                    sinistre_principal_id = None
                    for acte_id in actes:

                        acte = Acte.objects.get(id=acte_id)

                        # Added on 17062024: mise en oeuvre du tpg
                        type_prefinancement = get_type_prefinancement_of_acte(acte, formule)

                        nombre_demande = 1  # int(nombre_seances[i])
                        cout_acte = 0
                        cout_acte = int(couts_actes[i].replace(' ', '')) if len(str(couts_actes[i])) > 3 else int(
                            couts_actes[i])
                        quantite = nombre_demande

                        # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                        infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                                prestataire.id,
                                                                prescripteur_id, aliment_id, cout_acte, quantite,
                                                                consommation_individuelle, consommation_famille,
                                                                session_pec)

                        frais_reel = infos_acte['data']['frais_reel']
                        part_compagnie = infos_acte['data']['part_compagnie']
                        part_assure = infos_acte['data']['part_assure']
                        ticket_moderateur = infos_acte['data']['ticket_moderateur']
                        depassement = infos_acte['data']['depassement']

                        plafond_acte = infos_acte['data']['plafond_acte']
                        nombre_acte = infos_acte['data']['nombre_acte']
                        frequence = infos_acte['data']['frequence']
                        unite_frequence = infos_acte['data']['unite_frequence']
                        garanti = infos_acte['data']['garanti']
                        bareme_id = infos_acte['data']['bareme_id']
                        taux_tm = infos_acte['data']['taux_franchise']

                        statut = "ACCORDE" if acte.entente_prealable == 0 else "EN ATTENTE"
                        if acte.entente_prealable == 0:
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)
                            approuved_by = None
                        else:
                            reviewed_at = None
                            approuved_by = None

                        sinistre = Sinistre.objects.create(type_sinistre="acte",
                                                           created_by=request.user,
                                                           prestataire_id=prestataire.id,
                                                           aliment_id=aliment.id,
                                                           adherent_principal_id=aliment.adherent_principal.id,
                                                           police_id=formule.police.id,
                                                           periode_couverture_id=periode_couverture_encours.pk,
                                                           formulegarantie_id=formule.id,
                                                           bareme_id=bareme_id,
                                                           compagnie_id=formule.police.compagnie.id,
                                                           prescripteur_id=prescripteur_id,
                                                           affection_id=affection_id,
                                                           acte_id=acte_id,
                                                           frais_reel=frais_reel,
                                                           part_compagnie=part_compagnie,
                                                           part_assure=part_assure,
                                                           ticket_moderateur=ticket_moderateur,
                                                           depassement=depassement,
                                                           taux_tm=taux_tm,

                                                           montant_plafond=plafond_acte,
                                                           nombre_plafond=nombre_acte,
                                                           frequence=frequence,
                                                           unite_frequence=unite_frequence,

                                                           nombre_demande=nombre_demande,
                                                           nombre_accorde=0,

                                                           montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                           montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                           montant_remboursement_refuse=0,

                                                           montant_refacture_compagnie=part_compagnie,
                                                           montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0,

                                                           taux_retenue = get_retenue_selon_contexte(prestataire.id),

                                                           date_survenance=date_prestation,
                                                           date_entree=date_entree,
                                                           date_sortie=date_sortie,
                                                           statut=statut,
                                                           dossier_sinistre_id=dossier_sinistre.pk,
                                                           type_prefinancement=type_prefinancement,
                                                           reviewed_at=reviewed_at,
                                                           soins_a_l_entrange=soins_a_l_entrange
                                                           )

                        sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                        code_bureau = request.user.bureau.code
                        sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                            sinistre_created.pk).zfill(7) + '-SP'
                        sinistre_created.save()

                        i = i + 1

                        # vider toute la session de calcul de prise en charge pour reprendre
                        vider_sinistres_temporaires(request.user.id)

                # mettre a jour le total frais reel
                # Mettre a jour le statut_pec du dossier
                sinistres_en_attente = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ATTENTE)
                dossier_sinistre.statut_pec = 'EN ATTENTE' if sinistres_en_attente else 'ACCORDE'

                dossier_sinistre.numero = 'BS' + str(Date.today().year) + str(dossier_sinistre.pk).zfill(6)
                dossier_sinistre.save()


                # Envoi du SMS d'attente prealable #SMS
                if sinistres_en_attente:
                    numero_telephone_sms = aliment.adherent_principal.telephone_mobile
                    if aliment.adherent_principal.sms_active and numero_telephone_sms:
                        numero_telephone_sms = numero_telephone_sms.replace(" ", "").replace("-", "").replace("/","").replace("(", "").replace(")", "")

                        if len(numero_telephone_sms) >= 10:
                            date_survenance_fr = date_survenance.strftime('%d/%m/%Y à %H:%M')
                            message = "Une demande d'accord préalable vient d'être effectué à " + sinistre.prestataire.name + ".\n Date : " + date_survenance_fr + " \n PATIENT : " + str(
                                sinistre.aliment) + " \n N° CARTE : " + sinistre.aliment.carte_active().numero + " \n INOV vous souhaite un prompt rétablissement."
                            statut_sms = api_send_sms(message, numero_telephone_sms)

                            if statut_sms:
                                pprint("SMS demande préalable envoyé ")
                            else:
                                pprint("SMS demande préalable non envoyé ")


                # Imprimer le bon

                # notifier du succès
                response = {
                    'statut': 1,
                    'message': "Enregistrement effectuée avec succès !",
                    'data': sinistre_created.numero
                }

                # afficher le bon
                return redirect(
                    reverse('details_dossier_sinistre', kwargs={'dossier_sinistre_id': dossier_sinistre.pk}))

            else:

                response = {
                    'statut': 0,
                    'message': "Formule non trouvé !",
                    'data': {}
                }

        else:

            response = {
                'statut': 0,
                'message': "Assuré non trouvé !",
                'data': {
                }
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

        return JsonResponse(response)

    else:

        return redirect('/')


def popup_choose_prestataire(request):
    prestataires = Prestataire.objects.filter(bureau=request.user.bureau, type_prestataire_code="PRES002", status=True)

    return render(request, 'modal_choose_prestataire.html',
                  {'prestataires': prestataires})


#
def search_benef_by_name_datatable(request):
    items_per_page = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    page_number = start // items_per_page + 1  # Calculate page number correctly
    search_beneficiaire = request.GET.get('search_beneficiaire', '')

    # Get active polices and related formule IDs
    current_bureau = request.user.bureau
    active_polices = Police.objects.filter(
        bureau=current_bureau,
        statut=StatutPolice.ACTIF,
        statut_validite=StatutValidite.VALIDE
    )
    formule_ids = [formule.id for police in active_polices for formule in police.formules.all()]

    # Get aliment IDs related to active formules
    aliment_ids = AlimentFormule.objects.filter(
        formule_id__in=formule_ids,
        statut=Statut.ACTIF,
        statut_validite=StatutValidite.VALIDE
    ).values_list('aliment_id', flat=True)

    queryset = Aliment.objects.filter(id__in=aliment_ids)

    # Filter based on search_beneficiaire
    if search_beneficiaire:
        queryset = queryset.annotate(full_name=Concat(F('nom'), Value(' '), F('prenoms'))).filter(
            Q(full_name__icontains=search_beneficiaire)
        ).distinct()

    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)

    data = [
        {
            "id": c.id,
            "aliment__cartes__numero": f'{c.carte_active()}' if c.carte_active else "",
            "aliment__nom": c.nom or "",
            "aliment__prenoms": c.prenoms or "",
            "actions": f'<span style="cursor:pointer;" class="btn_select_beneficiaire" data-numero_carte="{c.carte_active()}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-plus"></i> Sélectionner</span></span>',
        }
        for c in page_obj
    ]

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })

@login_required
def add_sinistre_gestionnaire(request):
    if request.method == 'POST':

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint(request.POST)
        mode_creation = ModeCreation.objects.get(code="SAISIE-GESTIONNAIRE")

        type_prise_en_charge_id = request.POST.get('type_prise_en_charge_id')
        date_entree = request.POST.get('date_entree')
        date_sortie = request.POST.get('date_sortie')
        nombre_jours = request.POST.get('nombre_jours') #
        actes_du_tableau = request.POST.get('actes_du_tableau').rstrip(',').split(',')
        actes = request.POST.getlist('selected_acte_info[]')
        nombre_seances = request.POST.getlist('nombre_seance[]')
        couts_actes = request.POST.getlist('cout_acte[]')
        prescripteur_id = request.POST.get('prescripteur')
        centre_prescripteur_id = request.POST.get('centre_prescripteur')
        affection_id = request.POST.get('affection')
        renseignement_clinique = request.POST.get('renseignement_clinique_gestionnaire')
        commentaire = request.POST.get('commentaire_gestionnaire')
        aliment_id = request.POST.get('current_searched_aliment_id')
        aliment = Aliment.objects.get(id=aliment_id)
        
        soins_a_l_entrange = request.POST.get('soins_a_l_entrange', None)
        
        # pprint(nombre_seances)

        if nombre_jours: nombre_jours = nombre_jours.replace(' ', '')
        # Added on 15112023: si date entrée non remseigné,
        if date_entree == '': date_entree = None
        if date_sortie == '': date_sortie = None
        type_prise_en_charges = TypePriseencharge.objects.filter(id=type_prise_en_charge_id)
        type_prise_en_charge_code = type_prise_en_charges.first().code if type_prise_en_charges else None

        heure_survenance = request.POST.get('heure_survenance')
        heure_survenance = heure_survenance + ':00' if heure_survenance else '00:00:00'
        date_survenance = request.POST.get('date_survenance')
        date_survenance = date_survenance + ' ' + heure_survenance
        date_survenance = datetimeJsdecode.strptime(date_survenance, '%Y-%m-%d %H:%M:%S')
        date_survenance = make_aware(date_survenance)


        prestataire_executant_id = request.POST.get('prestataire_executant')
        prestataires = Prestataire.objects.filter(id=prestataire_executant_id)
        prestataire = prestataires.first() if prestataires else None

        type_remboursement_input = request.POST.get('type_remboursement')
        type_remboursements = TypeRemboursement.objects.filter(id=type_remboursement_input)
        type_remboursement = type_remboursements.first() if type_remboursements else None

        of_gestionnaire = True

        reference_facture_input = request.POST.get('reference_facture', None)
        date_reception_facture_input = request.POST.get('date_reception_facture', None)
        date_reception_facture_input = date_reception_facture_input if date_reception_facture_input is not None and date_reception_facture_input != "" else None

        # actes = actes_du_tableau

        if aliment:

            formule = aliment.formule_atdate(date_survenance)
            if formule:

                # Added on 05102023: add tm préfinancé
                type_prefinancement = formule.police.type_prefinancement if formule.police.type_prefinancement else None

                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = formule.police.periode_couverture_encours_atdate(date_survenance)

                if periode_couverture_encours:
                    pprint("periode_couverture_encours ------")
                    pprint(periode_couverture_encours)

                    consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                                statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                    consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                                   adherent_principal_id=aliment.adherent_principal.id,
                                                                   statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                    # vérifier si des actes ont étés envoyer pour êtree effectuer
                    if len(actes) > 0:

                        # crée le dossier sinistre
                        dossier_sinistre = DossierSinistre.objects.create(created_by=request.user,
                                                                          bureau_id=request.user.bureau.id,
                                                                          prestataire_id=prestataire.id,
                                                                          centre_prescripteur_id=centre_prescripteur_id,
                                                                          aliment_id=aliment.id,
                                                                          formulegarantie_id=formule.id,
                                                                          police_id=formule.police.id,
                                                                          compagnie_id=formule.police.compagnie.id,
                                                                          prescripteur_id=prescripteur_id,
                                                                          type_priseencharge_id=type_prise_en_charge_id,
                                                                          renseignement_clinique=renseignement_clinique,
                                                                          commentaire=commentaire,
                                                                          affection_id=affection_id,
                                                                          #type_prefinancement=type_prefinancement,
                                                                          type_remboursement=type_remboursement,
                                                                          of_gestionnaire=of_gestionnaire,
                                                                          mode_creation=mode_creation,
                                                                          date_survenance=date_survenance,
                                                                          reference_facture=reference_facture_input,
                                                                          date_reception_facture=date_reception_facture_input,
                                                                          soins_a_l_entrange=soins_a_l_entrange
                                                                          )
                        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre.pk)

                        # créer les sinistres
                        i = 0
                        sinistre_principal_id = None
                        for acte_id in actes:

                            acte = Acte.objects.get(id=acte_id)

                            # Added on 17062024: mise en oeuvre du tpg
                            # Added on 09102024: pas de prefinancement en RD
                            if type_remboursement and type_remboursement.code == "RD":
                                type_prefinancement = TypePrefinancement.objects.filter(code='NON_PREF').first()
                            else:
                                type_prefinancement = get_type_prefinancement_of_acte(acte, formule)

                            nombre_demande = int(nombre_seances[i])

                            if type_prise_en_charge_code == "HOSPIT":
                                nombre_demande = nombre_jours

                            cout_acte = 0
                            # si actes d'optique ou dentaire dont le cout est saisi manuelement
                            cout_acte = int(couts_actes[i].replace(' ', '')) if len(str(couts_actes[i])) > 3 else int(
                                couts_actes[i])
                            quantite = nombre_demande

                            # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                            infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                                    prestataire.id, prescripteur_id, aliment_id, cout_acte,
                                                                    quantite,
                                                                    consommation_individuelle, consommation_famille,
                                                                    session_pec)

                            '''if acte.option_seance:
                                frais_reel = infos_acte['data']['frais_reel'] / int(nombre_demande)
                                part_compagnie = infos_acte['data']['part_compagnie'] / int(nombre_demande)
                                part_assure = infos_acte['data']['part_assure'] / int(nombre_demande)
                                ticket_moderateur = infos_acte['data']['ticket_moderateur'] / int(nombre_demande)
                                depassement = infos_acte['data']['depassement'] / int(nombre_demande)

                            else:'''
                            frais_reel = infos_acte['data']['frais_reel']
                            part_compagnie = infos_acte['data']['part_compagnie']
                            part_assure = infos_acte['data']['part_assure']
                            ticket_moderateur = infos_acte['data']['ticket_moderateur']
                            depassement = infos_acte['data']['depassement']

                            plafond_acte = infos_acte['data']['plafond_acte']
                            nombre_acte = infos_acte['data']['nombre_acte']
                            frequence = infos_acte['data']['frequence']
                            unite_frequence = infos_acte['data']['unite_frequence']
                            garanti = infos_acte['data']['garanti']
                            bareme_id = infos_acte['data']['bareme_id']
                            taux_tm = infos_acte['data']['taux_franchise']

                            statut = "ACCORDE"
                            reviewed_at = datetime.datetime.now(tz=timezone.utc)

                            sinistre = Sinistre.objects.create(type_sinistre="acte",
                                                               created_by=request.user,
                                                               prestataire_id=prestataire.id,
                                                               aliment_id=aliment.id,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               police_id=formule.police.id,
                                                               periode_couverture_id=periode_couverture_encours.pk,
                                                               formulegarantie_id=formule.id,
                                                               bareme_id=bareme_id,
                                                               compagnie_id=formule.police.compagnie.id,
                                                               prescripteur_id=prescripteur_id,
                                                               affection_id=affection_id,
                                                               acte_id=acte_id,
                                                               frais_reel=frais_reel,
                                                               part_compagnie=part_compagnie,
                                                               part_assure=part_assure,
                                                               ticket_moderateur=ticket_moderateur,
                                                               depassement=depassement,
                                                               taux_tm=taux_tm,

                                                               montant_plafond=plafond_acte,
                                                               nombre_plafond=nombre_acte,
                                                               frequence=frequence,
                                                               unite_frequence=unite_frequence,

                                                               statut_prestation="EFFECTUE",
                                                               nombre_demande_initial=nombre_demande,
                                                               nombre_demande=1, #nombre_demande,
                                                               nombre_accorde=1, #nombre_demande, #0, accorder le nombre demandé vu que saisi par le gestionnaire, ils vont rejeter lors du traitement de la facture

                                                               montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                               montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                                               montant_remboursement_refuse=0,

                                                               montant_refacture_compagnie=part_compagnie,
                                                               montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0,

                                                               taux_retenue = get_retenue_selon_contexte(prestataire.id) if dossier_sinistre.type_remboursement.code == "TP" else 0,

                                                               date_survenance=date_survenance,
                                                               date_entree=date_entree,
                                                               date_sortie=date_sortie,
                                                               reference_facture=reference_facture_input,
                                                               date_reception_facture=date_reception_facture_input,
                                                               statut=statut,
                                                               dossier_sinistre_id=dossier_sinistre.pk,
                                                               type_prefinancement=type_prefinancement,
                                                               reviewed_at=reviewed_at,
                                                               soins_a_l_entrange=soins_a_l_entrange
                                                               )

                            sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                            code_bureau = request.user.bureau.code
                            sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(sinistre_created.pk).zfill(7) + '-SP'
                            sinistre_created.save()

                            #si acte avec séance dupliquer au nombre demandé
                            if acte.option_seance:
                                dupliquer = True
                                pprint("DUPLIQUER LE SINISTRE SELON LE NOMBRE DEMANDE")
                                cpt = 1  # vu qu'il y a déjà une ligne de séance sur laquelle on est
                                nombre_total = int(sinistre.nombre_demande_initial)

                                for _ in range(int(sinistre.nombre_demande_initial) - 1):
                                    new_sinistre = deepcopy(sinistre)  # Copie de l'objet

                                    new_sinistre.id = None
                                    new_sinistre.numero = None
                                    new_sinistre.nombre_demande = 1
                                    new_sinistre.nombre_accorde = 1
                                    new_sinistre.date_survenance = None
                                    new_sinistre.statut_prestation = StatutSinistrePrestation.EFFECTUE
                                    #new_sinistre.reviewed_at = datetime.datetime.now(tz=timezone.utc)
                                    #new_sinistre.approuved_by = request.user
                                    new_sinistre.save()

                                    # générer le numéro
                                    code_bureau = sinistre.dossier_sinistre.bureau.code
                                    new_sinistre.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(new_sinistre.pk).zfill(7) + '-SP'
                                    new_sinistre.save()

                                    cpt += 1

                                    # Notifier waspito si le prestataire est WASPITO INC
                                    #notifier_waspito(new_sinistre)

                                    # enregistrer dans les log
                                    #ActionLog.objects.create(done_by=request.user, action="creation", description="Création d'un sinistre", table="sinistre", row=new_sinistre.pk)

                                    pprint("cpt")
                                    pprint(cpt)
                                    pprint("sinistre.nombre_demande")
                                    pprint(sinistre.nombre_demande)



                            i = i + 1

                            # vider toute la session de calcul de prise en charge pour reprendre
                            vider_sinistres_temporaires(request.user.id)


                        # mettre a jour le total frais reel
                        # Mettre a jour le statut_pec du dossier
                        dossier_sinistre.statut_pec = 'ACCORDE'

                        #added on 29042024
                        # si hospitalisation (accouchement), mettre à jour le plafond accouchement
                        if type_prise_en_charge_code == "HOSPIT":
                            acte = dossier_sinistre.sinistres.first().acte  # en hospit c'est un seul acte par dossier
                            bareme_plafond_accouchement = Bareme.objects.filter(formulegarantie_id=formule.id,
                                                                                acte=acte, is_garanti=True,
                                                                                date_debut__lte=date_survenance,
                                                                                statut=Statut.ACTIF)
                            plafond_accouchement = bareme_plafond_accouchement.first().plafond_acte if bareme_plafond_accouchement else 0
                            dossier_sinistre.plafond_accouchement = plafond_accouchement

                        dossier_sinistre.numero = 'BS' + str(Date.today().year) + str(dossier_sinistre.pk).zfill(6)
                        dossier_sinistre.save()

                        # Si c'est une consultation enregistrer les médicaments ratachés
                        if type_prise_en_charge_code == "CONSULT":

                            # Réccupérer les médicaments en session
                            liste_medicaments = request.session.get('liste_medicaments', [])

                            dossier_sinistre_pharmacie_id = None

                            for med in liste_medicaments:
                                # refaire le calcul, vu qu'on vient d'enregistrer la consultation, la conssommation a augmenter

                                dossier_sinistre_pharmacie_id = med['pharmacie_id']

                                med_acte_id = med['medicament_id']
                                med_prestataire_id = med['pharmacie_id']
                                med_quantite = med['qte_servie']
                                med_prix_unitaire = med['prix_unitaire']
                                med_frais_reel = med['prix_total']

                                cout_acte = med_frais_reel
                                quantite = 0  # juste pour l'algo, on fait pu*qte

                                # Added on 17062024: mise en oeuvre du tpg
                                med_acte = Acte.objects.get(id=med_acte_id)
                                # Added on 30102024: pas de prefinancement en RD
                                if type_remboursement and type_remboursement.code == "RD":
                                    med_type_prefinancement = TypePrefinancement.objects.filter(code='NON_PREF').first()
                                else:
                                    med_type_prefinancement = get_type_prefinancement_of_acte(med_acte, formule)

                                # Nouvelle idée: faire le calcul des consos dans la fonction get_tarif_acte_from_bareme
                                # récupérer ses consommations individuel et par famille
                                consommation_individuelle = Sinistre.objects.filter(
                                    periode_couverture_id=periode_couverture_encours.pk,
                                    aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                                consommation_famille = Sinistre.objects.filter(
                                    periode_couverture_id=periode_couverture_encours.pk,
                                    adherent_principal_id=aliment.adherent_principal.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                                # pour plus de sécurité, récupérer les tarif à partir du fichier excel et nom du formulaire utilisateur
                                infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance,
                                                                        med_acte_id,
                                                                        med_prestataire_id, prescripteur_id, aliment_id,
                                                                        cout_acte,
                                                                        quantite,
                                                                        consommation_individuelle, consommation_famille,
                                                                        session_pec)

                                med_frais_reel = infos_acte['data']['frais_reel']
                                med_part_compagnie = infos_acte['data']['part_compagnie']
                                med_part_assure = infos_acte['data']['part_assure']
                                med_ticket_moderateur = infos_acte['data']['ticket_moderateur']
                                med_depassement = infos_acte['data']['depassement']

                                med_plafond_acte = infos_acte['data']['plafond_acte']
                                med_nombre_acte = infos_acte['data']['nombre_acte']
                                med_garanti = infos_acte['data']['garanti']
                                med_bareme_id = infos_acte['data']['bareme_id']
                                med_taux_tm = infos_acte['data']['taux_franchise']

                                med_statut = StatutSinistre.ACCORDE
                                reviewed_at = datetime.datetime.now(tz=timezone.utc)

                                sinistre = Sinistre.objects.create(type_sinistre="medicament",
                                                                   created_by=request.user,
                                                                   prestataire_id=med_prestataire_id,
                                                                   aliment_id=aliment.id,
                                                                   adherent_principal_id=aliment.adherent_principal.id,
                                                                   police_id=formule.police.id,
                                                                   periode_couverture_id=periode_couverture_encours.pk,
                                                                   formulegarantie_id=formule.id,
                                                                   bareme_id=med_bareme_id,
                                                                   compagnie_id=formule.police.compagnie.id,
                                                                   prescripteur_id=prescripteur_id,
                                                                   affection_id=affection_id,
                                                                   acte_id=med_acte_id,
                                                                   frais_reel=med_frais_reel,
                                                                   part_compagnie=med_part_compagnie,
                                                                   part_assure=med_part_assure,
                                                                   ticket_moderateur=med_ticket_moderateur,
                                                                   depassement=med_depassement,
                                                                   taux_tm=med_taux_tm,

                                                                   prix_unitaire=med_prix_unitaire,
                                                                   nombre_demande=med_quantite,
                                                                   nombre_accorde=med_quantite,

                                                                   montant_base_remboursement=med_frais_reel if (med_type_prefinancement and med_type_prefinancement.code == "PREF_TOUT") else med_part_compagnie,
                                                                   montant_remboursement_accepte=med_frais_reel if (med_type_prefinancement and med_type_prefinancement.code == "PREF_TOUT") else med_part_compagnie,
                                                                   montant_remboursement_refuse=0,

                                                                   taux_retenue = get_retenue_selon_contexte(med_prestataire_id),

                                                                   date_survenance=date_survenance,
                                                                   reference_facture=reference_facture_input,
                                                                   date_reception_facture=date_reception_facture_input,
                                                                   statut=med_statut,
                                                                   dossier_sinistre_id=dossier_sinistre.pk,
                                                                   type_prefinancement=med_type_prefinancement,
                                                                   reviewed_at=reviewed_at,
                                                                   soins_a_l_entrange=soins_a_l_entrange
                                                                   )

                                sinistre_created = Sinistre.objects.get(id=sinistre.pk)

                                code_bureau = request.user.bureau.code
                                sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                                    sinistre_created.pk).zfill(7) + '-SP'
                                sinistre_created.save()

                                pprint("sinistre_created")
                                pprint(sinistre_created)

                            #renseigner la pharmacie sur le dossier_sinistre
                            dossier_sinistre.pharmacie_id = dossier_sinistre_pharmacie_id
                            dossier_sinistre.save()

                        # notifier du succès
                        response = {
                            'statut': 1,
                            'message': "Enregistrement effectuée avec succès !",
                            'data': sinistre_created.numero
                        }

                    else:

                        response = {
                            'statut': 0,
                            'message': "Aucun acte sélectionné !",
                            'data': {}
                        }

                else:

                    response = {
                        'statut': 0,
                        'message': "Période de couverture non trouvé !",
                        'data': {}
                    }

            else:

                response = {
                    'statut': 0,
                    'message': "Formule non trouvé !",
                    'data': {}
                }

        else:

            response = {
                'statut': 0,
                'message': "Assuré non trouvé !",
                'data': {
                }
            }

        return JsonResponse(response)

    else:
        response = {
            'statut': 0,
            'message': "Aucune donnée reçue",
            'data': {
            }
        }

    return JsonResponse(response)


def vider_sinistres_temporaires(session_pec):
    pprint("session_pec")
    pprint(session_pec)
    SinistreTemporaire.objects.filter(session_pec=session_pec).delete()


@method_decorator(login_required, name='dispatch')
class DetailsDossierSinistreView(TemplateView):
    # permission_required = "sinistre.view_sinistre"
    template_name = 'details_dossier_sinistre.html'
    model = Sinistre

    def get(self, request, dossier_sinistre_id, *args, **kwargs):

        if request.user.is_pres or request.user.is_imag or request.user.is_optic or request.user.is_labo or request.user.is_dentaire:
            dossier_sinistre = DossierSinistre.objects.filter(id=dossier_sinistre_id, prestataire=request.user.prestataire, bureau=request.user.bureau).first()

        elif request.user.is_pharm:
            dossier_sinistre = DossierSinistre.objects.filter(id=dossier_sinistre_id, type_priseencharge__code="CONSULT", bureau=request.user.bureau).first()
            # print(f'@@ ville by prestataire : {dossier_sinistre.prestataire.ville} @@')
            # print(f'@@ ville by bureau : {dossier_sinistre.bureau.ville} @@')

        else:
            dossier_sinistre = DossierSinistre.objects.filter(id=dossier_sinistre_id, bureau=request.user.bureau).first()


        if dossier_sinistre:
            sinistres = Sinistre.objects.filter(dossier_sinistre_id=dossier_sinistre_id, type_sinistre="acte",
                                                statut_validite=StatutValidite.VALIDE)
            # commented on 25112023: medicaments = Sinistre.objects.filter(dossier_sinistre_id=dossier_sinistre_id, statut_validite=StatutValidite.VALIDE).filter(type_sinistre='medicament').filter(Q(served_by=request.user) if (not request.user.is_superuser and not request.user.is_med) else Q())
            medicaments = Sinistre.objects.filter(dossier_sinistre_id=dossier_sinistre_id,
                                                  statut_validite=StatutValidite.VALIDE).filter(
                type_sinistre='medicament').filter(Q(prestataire=request.user.prestataire) if (
                        not request.user.is_superuser and not request.user.is_med and not request.user.is_ges) else Q())

            medicaments_total_frais_reel = sum(sm.total_frais_reel for sm in medicaments.exclude(statut="REJETE"))
            medicaments_total_part_assure = sum(sm.total_part_assure for sm in medicaments.exclude(statut="REJETE"))
            medicaments_total_part_compagnie = sum(
                sm.total_part_compagnie for sm in medicaments.exclude(statut="REJETE"))

            types_documents = TypeDocument.objects.filter(is_sinistre=True)
            affections = Affection.objects.filter(status=True)

            # dd(medicaments)

            '''
            #gestion des medicaments à afficher
            liste_medicaments = Medicament.objects.all()
            
            #retirer ceux dont la rubrique n'est pas garantie dans le bareme
            rubriques_pharmacie_exclu = Bareme.objects.filter(rubrique__type_priseencharge__code="PHARM", formulegarantie_id=dossier_sinistre.aliment.formule.pk, is_garanti=False)
            pprint("rubriques_pharmacie_exclu")
            pprint(rubriques_pharmacie_exclu)

            # retirer les médicament dont la rubrique est exclu
            liste_medicaments = liste_medicaments.exclude(rubrique__in=rubriques_pharmacie_exclu.values_list('rubrique', flat=True))
            pprint("liste_medicaments")
            pprint(liste_medicaments)
            '''

            # version avec actes et medicaments fusionnés
            liste_medicaments = Acte.objects.filter(type_acte__code="MEDICAMENT", status=1, statut_validite=StatutValidite.VALIDE)
            pprint("liste_medicaments")
            pprint(liste_medicaments)

            #
            documents = dossier_sinistre.documents.filter(statut=Statut.ACTIF)  # .distinct('type_document')
            document_dict = {}
            unique_documents = []

            for document in documents:
                if document.type_document not in document_dict:
                    document_dict[document.type_document] = document
                    unique_documents.append(document)

            context = self.get_context_data(**kwargs)
            context['dossier_sinistre'] = dossier_sinistre
            context['sinistres'] = sinistres
            context['liste_medicaments'] = liste_medicaments
            context['medicaments'] = medicaments
            context['affections'] = affections
            context['types_documents'] = types_documents
            context['documents'] = unique_documents
            context['total_frais_reel'] = dossier_sinistre.total_frais_reel
            context['total_part_assure'] = dossier_sinistre.total_part_assure
            context['total_part_compagnie'] = dossier_sinistre.total_part_compagnie
            context['medicaments_total_frais_reel'] = medicaments_total_frais_reel
            context['medicaments_total_part_assure'] = medicaments_total_part_assure
            context['medicaments_total_part_compagnie'] = medicaments_total_part_compagnie

            return self.render_to_response(context)

        else:
            return redirect('/')

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def dossier_sinistre_add_document(request, dossier_sinistre_id):
    response = {
        'statut': 0,
        'message': "Veuillez renseigner correctement le formulaire !",
        # 'errors': form.errors,
    }

    if request.method == "POST":

        # form = DocumentForm(request.POST, request.FILES)

        # files = request.FILES['fichiers']
        # files = []
        # type_documents = []
        documents = []

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)

        # print(request.POST)
        # print(request.FILES)
        types_documents = TypeDocument.objects.all()

        for i in range(types_documents.count()):
            if request.FILES.get(f'fichier_{i}'):

                type_document_id = request.POST.get(f'type_document_{i}')

                #
                try:

                    file = request.FILES[f'fichier_{i}']

                    fs = FileSystemStorage()
                    # file_name_renamed = 'doc_' + str(dossier_sinistre_id) + '_' + str(uuid.uuid4()) +'_'+ file.name.replace(" ", "_")
                    file_name_renamed = 'doc_' + str(dossier_sinistre_id) + '_' + file.name.replace(" ", "_")
                    file_upload_path = 'dossiers_sinistres/documents/' + file_name_renamed

                    fs.save(file_upload_path, file)

                    type_document = TypeDocument.objects.get(id=type_document_id)
                    document = DocumentDossierSinistre.objects.create(dossier_sinistre=dossier_sinistre,
                                                                      type_document=type_document,
                                                                      fichier=file_upload_path)

                    print(vars(document))

                    documents.append({
                        'id': document.pk,
                        # 'nom': document.nom,
                        'fichier': '<a href="' + document.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                        'type_document': document.type_document.libelle,
                        # 'confidentialite': document.confidentialite,
                    })


                except MultiValueDictKeyError:
                    file_upload_path = ''

                #

        # print(files)
        # print(type_documents)
        response = {
            # 'files':files,
            # 'type_documents':type_documents,
            'statut': 1,
            'message': "Enregistrement effectué avec succès !",
            'documents': list(documents),
        }

    return JsonResponse(response)


def handle_uploaded_document(f, filename):
    path_ot_db = '/dossiers_sinistres/documents/'
    dirname = settings.MEDIA_URL.replace('/', '') + path_ot_db
    path = os.path.join(dirname)

    if not os.path.exists(path):
        os.makedirs(path)

    with open(dirname + '/' + filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

    return path_ot_db + '/' + filename


def supprimer_document(request):
    if request.method == "POST":

        document_id = request.POST.get('document_id')

        document = DocumentDossierSinistre.objects.get(id=document_id)
        if document.pk is not None:
            document.delete()

            response = {
                'statut': 1,
                'message': "Document supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Document non trouvé !",
            }

        return JsonResponse(response)


# TODO : popup_details_sinistre
def popup_details_sinistre(request, sinistre_id):
    sinistre = Sinistre.objects.get(id=sinistre_id)
    sinistres_historique_acte = Sinistre.objects.filter(acte_id=sinistre.acte_id,
                                                        aliment_id=sinistre.aliment_id).exclude(
        id=sinistre.pk).order_by('-id')[:3]
    prorogations = ProrogationSinistre.objects.filter(sinistre_id=sinistre_id)

    return render(request, 'modal_details_sinistre.html',
                  {'sinistre': sinistre, 'sinistres_historique_acte': sinistres_historique_acte,
                   'prorogations': prorogations})


def popup_seance_done(request, sinistre_id):
    sinistre = Sinistre.objects.get(id=sinistre_id)
    today = datetime.datetime.now(tz=timezone.utc)

    if request.method == 'POST':
        date_survenance = request.POST.get('date_survenance')
        sinistre.statut_prestation = StatutSinistrePrestation.EFFECTUE
        sinistre.date_survenance = date_survenance
        sinistre.save()

        response = {
            "statut": 1,
            "message": "Séance marquée comme éffectuée",
            "data": {
            }
        }

        return JsonResponse(response)

    return render(request, 'modal_seance_done.html', {'sinistre': sinistre, 'today': today})


def popup_modifier_sinistre_medicament(request, sinistre_id):
    sinistre = Sinistre.objects.get(id=sinistre_id)

    return render(request, 'modal_details_sinistre.html', {'sinistre': sinistre})


@login_required()
def delete_sinistre_medicament(request, sinistre_id):
    if request.method == 'POST':

        sinistre = Sinistre.objects.get(id=sinistre_id)
        sinistre.statut = StatutValidite.SUPPRIME  # à analyser
        sinistre.statut_validite = StatutValidite.SUPPRIME
        sinistre.deleted_author = request.user
        sinistre.deleted_at = datetime.datetime.now(tz=timezone.utc)
        sinistre.save()

        response = {
            "statut": 1,
            "message": "Suppression effectué avec succès",
            "data": {
            }
        }

        return JsonResponse(response)

    else:
        return redirect('/')


#todo voir s'il ne faut pas remplacer aliment.formule par formule_atdate
@login_required()
def add_sinistre_medicament(request, dossier_sinistre_id):
    if request.method == 'POST':

        # try:

        # vider toute la session de calcul de prise en charge pour reprendre
        session_pec = request.user.id
        vider_sinistres_temporaires(session_pec)

        pprint("request.POST")
        pprint(request.POST)
        prestataire_id = request.user.prestataire.id
        medicament_id = request.POST.get('medicament')
        qte = request.POST.get('qte').replace(" ", "")
        prix_unitaire = int(request.POST.get('prix_unitaire').replace(" ", ""))
        date_prestation = datetime.datetime.now(tz=timezone.utc)

        medicament = Acte.objects.get(id=medicament_id)

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)
        taux_couverture_formule = dossier_sinistre.aliment.formule.taux_couverture
        
        # in case prestataire_id is NoneType
        #   prestataire_id = dossier_sinistre.prestataire.id

        # dd()
        aliment = dossier_sinistre.aliment

        frais_reel = int(qte) * prix_unitaire

        '''
        nouvelle version: actes et medicaments fusionnés
        '''
        # récupérer ses consommations individuel et par famille
        periode_couverture_encours = aliment.formule.police.periode_couverture_encours
        consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                    statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                       adherent_principal_id=aliment.adherent_principal.id,
                                                       statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

        acte = Acte.objects.get(id=medicament_id)
        # Added on 17062024: mise en oeuvre du tpg
        type_prefinancement = get_type_prefinancement_of_acte(acte, aliment.formule)

        type_prise_en_charge_code = "PHARM"
        date_survenance = datetime.datetime.now(tz=timezone.utc)
        acte_id = acte.id
        prescripteur_id = dossier_sinistre.prescripteur.id if dossier_sinistre.prescripteur else None
        aliment_id = dossier_sinistre.aliment.id
        cout_acte = frais_reel
        nombre_jours = 0
        affection = dossier_sinistre.affection if dossier_sinistre.affection else None

        # faire passer par l'algo de tarification
        infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id, prestataire_id,
                                                prescripteur_id, aliment_id, cout_acte,
                                                nombre_jours, consommation_individuelle, consommation_famille,
                                                session_pec)

        pprint("---/__/---")
        pprint("infos_acte")
        pprint(infos_acte)
        if infos_acte['statut'] == 1:

            frais_reel = infos_acte['data']['frais_reel']
            part_compagnie = infos_acte['data']['part_compagnie']
            part_assure = infos_acte['data']['part_assure']
            ticket_moderateur = infos_acte['data']['ticket_moderateur']
            depassement = infos_acte['data']['depassement']

            plafond_acte = infos_acte['data']['plafond_acte']
            nombre_acte = infos_acte['data']['nombre_acte']
            frequence = infos_acte['data']['frequence']
            unite_frequence = infos_acte['data']['unite_frequence']
            garanti = infos_acte['data']['garanti']
            bareme_id = infos_acte['data']['bareme_id']
            taux_tm = infos_acte['data']['taux_franchise']

            statut = "ACCORDE" if (acte.entente_prealable == 0) else "EN ATTENTE"
            if acte.entente_prealable == 0:
                reviewed_at = datetime.datetime.now(tz=timezone.utc)
                approuved_by = None
            else:
                reviewed_at = None
                approuved_by = None

            sinistre = Sinistre.objects.create(type_sinistre="medicament",
                                               created_by=request.user,
                                               served_by=request.user,
                                               prestataire_id=prestataire_id,
                                               aliment_id=aliment.id,
                                               adherent_principal_id=aliment.adherent_principal.id,
                                               police_id=aliment.formule.police.id,
                                               periode_couverture_id=aliment.formule.police.periode_couverture_encours.pk,
                                               formulegarantie_id=aliment.formule.id,
                                               bareme_id=bareme_id,
                                               compagnie_id=aliment.formule.police.compagnie.id,
                                               prescripteur_id=prescripteur_id,
                                               affection=affection,
                                               acte_id=acte_id,
                                               frais_reel=frais_reel,
                                               part_compagnie=part_compagnie,
                                               part_assure=part_assure,
                                               ticket_moderateur=ticket_moderateur,
                                               depassement=depassement,
                                               taux_tm=taux_tm,

                                               montant_plafond=plafond_acte,
                                               nombre_plafond=nombre_acte,
                                               frequence=frequence,
                                               unite_frequence=unite_frequence,

                                               prix_unitaire=prix_unitaire,
                                               nombre_demande=qte,
                                               nombre_accorde=qte,

                                               montant_base_remboursement=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                               montant_remboursement_accepte=frais_reel if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else part_compagnie,
                                               montant_remboursement_refuse=0,

                                               montant_refacture_compagnie=part_compagnie,
                                               montant_refacture_client=part_assure if (type_prefinancement and type_prefinancement.code == "PREF_TOUT") else 0,

                                               taux_retenue = get_retenue_selon_contexte(prestataire_id),

                                               date_survenance=date_prestation,
                                               statut=statut,
                                               dossier_sinistre_id=dossier_sinistre.pk,
                                               type_prefinancement=type_prefinancement,
                                               reviewed_at=reviewed_at
                                               )

            sinistre_created = Sinistre.objects.get(id=sinistre.pk)

            code_bureau = request.user.bureau.code
            sinistre_created.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(
                sinistre_created.pk).zfill(7) + '-SP'
            sinistre_created.save()

            # Mettre a jour le statut_pec du dossier
            sinistres_en_attente = dossier_sinistre.sinistres.filter(statut=StatutSinistre.ATTENTE)
            dossier_sinistre.statut_pec = 'EN ATTENTE' if sinistres_en_attente else 'ACCORDE'
            #Added on 01012023: renseigner la pharmacie qui a servi le médicament sur le dossier_sinistre
            dossier_sinistre.pharmacie=request.user.prestataire
            dossier_sinistre.save()

            response = {
                "statut": 1,
                "message": "Médicament ajouté avec succèss!",
                "data": {
                    'medicament': sinistre_created.acte.libelle,
                    'qte_demande': sinistre_created.nombre_demande,
                    'qte_servie': sinistre_created.nombre_accorde,
                    'prix_unitaire': float(sinistre_created.prix_unitaire),
                    'prix_total': float(sinistre_created.frais_reel),
                    'part_assureur': float(sinistre_created.frais_reel) if sinistre_created.tm_prefinanced else float(
                        sinistre_created.part_compagnie),
                    'part_assure': 0 if sinistre_created.tm_prefinanced else float(sinistre_created.part_assure),
                    'statut': sinistre_created.statut,
                }
            }

        else:
            response = {
                "statut": 0,
                "message": infos_acte['message'],
                "data": {}
            }

        # vider toute la session de calcul de prise en charge pour reprendre
        vider_sinistres_temporaires(request.user.id)

        '''except:
        print("EXCEPT RUNNING ")
        response = {
            "statut": 0,
            "message": "Erreur survenu lors de l'ajout du medicament",
            "data": {}
        }'''

    return JsonResponse(response)


def update_medicament_sinistre(request, sinistre_id):
    if request.method == 'POST':

        try:

            # vider toute la session de calcul de prise en charge pour reprendre
            session_pec = request.user.id
            vider_sinistres_temporaires(session_pec)

            sinistre = Sinistre.objects.get(id=sinistre_id)

            if sinistre:

                # dd(request.POST)
                prestataire_id = sinistre.prestataire_id
                medicament_id = request.POST.get('medicament')
                qte = request.POST.get('qte').replace(" ", "")
                prix_unitaire = int(request.POST.get('prix_unitaire').replace(" ", ""))
                date_prestation = datetime.datetime.now(tz=timezone.utc)

                dossier_sinistre = DossierSinistre.objects.get(id=sinistre.dossier_sinistre_id)
                taux_couverture_formule = dossier_sinistre.aliment.formule.taux_couverture

                # dd()
                aliment = dossier_sinistre.aliment

                frais_reel = int(qte) * prix_unitaire

                '''
                nouvelle version: actes et medicaments fusionnés
                '''
                # récupérer ses consommations individuel et par famille
                periode_couverture_encours = aliment.formule.police.periode_couverture_encours
                consommation_individuelle = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk, aliment_id=aliment.id,
                                            statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                consommation_famille = Sinistre.objects.filter(periode_couverture_id=periode_couverture_encours.pk,
                                                               adherent_principal_id=aliment.adherent_principal.id,
                                                               statut=StatutSinistre.ACCORDE, statut_remboursement__in=[StatutRemboursement.ATTENTE, StatutRemboursement.DEMANDE, StatutRemboursement.ACCEPTE, StatutRemboursement.ACCEPTE_PARTIELLEMENT], statut_validite=StatutValidite.VALIDE).aggregate(Sum('part_compagnie'))['part_compagnie__sum'] or 0

                acte = Acte.objects.get(id=medicament_id)
                type_prise_en_charge_code = "PHARM"
                date_survenance = datetime.datetime.now(tz=timezone.utc)
                acte_id = acte.id
                prescripteur_id = dossier_sinistre.prescripteur.id
                aliment_id = dossier_sinistre.aliment.id
                cout_acte = frais_reel
                nombre_jours = 0
                affection = dossier_sinistre.affection

                # faire passer par l'algo de tarification
                infos_acte = get_tarif_acte_from_bareme(type_prise_en_charge_code, date_survenance, acte_id,
                                                        prestataire_id,
                                                        prescripteur_id, aliment_id, cout_acte,
                                                        nombre_jours, consommation_individuelle, consommation_famille,
                                                        session_pec)

                pprint("infos_acte")
                pprint(infos_acte)
                if infos_acte['statut'] == 1:

                    frais_reel = infos_acte['data']['frais_reel']
                    part_compagnie = infos_acte['data']['part_compagnie']
                    part_assure = infos_acte['data']['part_assure']
                    ticket_moderateur = infos_acte['data']['ticket_moderateur']
                    depassement = infos_acte['data']['depassement']

                    plafond_acte = infos_acte['data']['plafond_acte']
                    nombre_acte = infos_acte['data']['nombre_acte']
                    frequence = infos_acte['data']['frequence']
                    unite_frequence = infos_acte['data']['unite_frequence']
                    garanti = infos_acte['data']['garanti']
                    bareme_id = infos_acte['data']['bareme_id']

                    statut = "ACCORDE" if (acte.entente_prealable == 0) else "EN ATTENTE"

                    Sinistre.objects.filter(pk=sinistre_id).update(type_sinistre="medicament",
                                                                   # created_by=request.user,
                                                                   served_by=request.user,
                                                                   prestataire_id=prestataire_id,
                                                                   aliment_id=aliment.id,
                                                                   adherent_principal_id=aliment.adherent_principal.id,
                                                                   police_id=aliment.formule.police.id,
                                                                   periode_couverture_id=aliment.formule.police.periode_couverture_encours.pk,
                                                                   formulegarantie_id=aliment.formule.id,
                                                                   bareme_id=bareme_id,
                                                                   compagnie_id=aliment.formule.police.compagnie.id,
                                                                   prescripteur_id=prescripteur_id,
                                                                   affection=affection,
                                                                   acte_id=acte_id,
                                                                   frais_reel=frais_reel,
                                                                   part_compagnie=part_compagnie,
                                                                   part_assure=part_assure,
                                                                   ticket_moderateur=ticket_moderateur,
                                                                   depassement=depassement,

                                                                   montant_plafond=plafond_acte,
                                                                   nombre_plafond=nombre_acte,
                                                                   frequence=frequence,
                                                                   unite_frequence=unite_frequence,

                                                                   prix_unitaire=prix_unitaire,
                                                                   nombre_demande=qte,
                                                                   nombre_accorde=qte,

                                                                   date_survenance=date_prestation,
                                                                   statut=statut,
                                                                   dossier_sinistre_id=dossier_sinistre.pk
                                                                   )

                    sinistre_updated = Sinistre.objects.get(pk=sinistre_id)

                    response = {
                        "statut": 1,
                        "message": "Médicament modifié avec succèss!",
                        "data": {
                            'medicament': sinistre_updated.acte.libelle,
                            'qte_demande': sinistre_updated.nombre_demande,
                            'qte_servie': sinistre_updated.nombre_accorde,
                            'prix_unitaire': float(sinistre_updated.prix_unitaire),
                            'prix_total': float(sinistre_updated.frais_reel),
                            'part_assureur': float(sinistre_updated.part_compagnie),
                            'part_assure': float(sinistre_updated.part_assure),
                            'statut': sinistre_updated.statut,
                        }
                    }

                else:
                    response = {
                        "statut": 0,
                        "message": infos_acte['message'],
                        "data": {}
                    }

            else:
                response = {
                    "statut": 0,
                    "message": "Sinistre non trouvé",
                    "data": {}
                }

            # vider toute la session de calcul de prise en charge pour reprendre
            vider_sinistres_temporaires(request.user.id)


        except Exception as e:
            print("EXCEPT RUNNING: Erreur lors de la mise à jour du medicament: " + str(e))
            response = {
                "statut": 0,
                "message": "Erreur lors de la mise à jour du medicament: " + str(e),
                "data": {}
            }

        return JsonResponse(response)


def update_add_affection(request, dossier_sinistre_id):
    if request.method == 'POST':
        code_affection = request.POST.get('code_affection')
        affection = Affection.objects.get(id=code_affection)

        dossier_sinistre = DossierSinistre.objects.get(id=dossier_sinistre_id)

        try:
            dossier_sinistre.affection = affection
            dossier_sinistre.updated_by = request.user
            dossier_sinistre.updated_at = datetime.datetime.now(tz=timezone.utc)
            dossier_sinistre.save()
            response = {
                "statut": 1,
                "message": "Le code affection a été enregistré.",
                "data": {}
            }
        except:
            print("EXCEPT RUNNING ")
            response = {
                "statut": 0,
                "message": "Erreur survenu lors de la mise à de l'affection.",
                "data": {}
            }

        return JsonResponse(response)


def close_dossier_medication(request, dossier_sinistre_id):
    print(dossier_sinistre_id)
    print("Le dossier sinistre")
    if request.method == 'POST':
        # sinistre = Sinistre.objects.get(numero=numero_sinistre)
        dossier_sinistre_id = request.POST.get('dossier_sinistre_id')
        try:
            print("TRYING TO UPDATE")
            dossier_sinistre = DossierSinistre.objects.get(
                pk=dossier_sinistre_id
            )
            dossier_sinistre.is_closed = True
            dossier_sinistre.save()

            response = {
                "statut": 1,
                "message": "Le dossier " + str(
                    dossier_sinistre.numero) + " à été cloturé. Vous ne pouver plus rien y ajouter.",
                "data": {
                }
            }
        except Exception as e:
            print("EXCEPT RUNNING ")
            print(e)
            response = {
                "statut": 0,
                "message": "Une erreur est survenue lors de la cloture du dossier",
                "data": {}
            }

        return JsonResponse(response)

#TODO remove 'tm_prefinanced': tm_prefinanced and get the option on sinistre object
def render_pdf_view(request, id):
    dossiers_sinistres = DossierSinistre.objects.filter(id=id, statut_validite=StatutValidite.VALIDE)


    if dossiers_sinistres:

        dossier_sinistre = dossiers_sinistres.first()

        # sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut='ACCORDE')
        sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut__in=('ACCORDE', 'REJETE'))

        # Added on 05102023: add tm préfinancé
        request_user_is_prestataire = request.user.is_prestataire

        tm_prefinanced = False
        if dossier_sinistre.type_prefinancement and dossier_sinistre.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        ticket_moderateur_consultation = 0
        ticket_moderateur_pharmacie = 0
        ticket_moderateur_pharmacie = get_ticket_moderateur_pharmacie(dossier_sinistre.aliment_id, dossier_sinistre.formulegarantie_id, dossier_sinistre.date_survenance)
        pprint("ticket_moderateur_pharmacie")
        pprint(ticket_moderateur_pharmacie)

        garantis_pharmacie = dossier_sinistre.formulegarantie.garantis_pharmacie
        plafond_atteint = False

        texte_nombre_demande = ""
        texte_nombre_accorde = ""

        if dossier_sinistre.type_priseencharge.code == "CONSULT":

            # vérification si vitamine et vaccins non garanti
            '''bareme_vitamine = Bareme.objects.filter(acte__code="VITAMINES",
                                                    formulegarantie_id=dossier_sinistre.aliment.formule.id, )
            bareme_vaccin = Bareme.objects.filter(acte__code="VACCIN",
                                                  formulegarantie_id=dossier_sinistre.aliment.formule.id, )
    
            if bareme_vitamine:
                vitamines_garanti = bareme_vitamine.is_garanti
    
            if bareme_vaccin:
                vaccins_garanti = bareme_vaccin.is_garanti
            '''

            plafond_atteint = bool_plafond_atteint(dossier_sinistre)

            if sinistres:
                sinistre = sinistres[0]
                ticket_moderateur_consultation = sinistre.bareme.taux_tm if sinistre.bareme else sinistre.formulegarantie.taux_tm
                #a revoir: prendre le plafond sur la pharmacie et non sur la consultation

            #Added on 16022024
            ticket_moderateur_consultation = ticket_moderateur_pharmacie

            template_path = 'sinistre_printer_consultation.html'
        elif dossier_sinistre.type_priseencharge.code == "AMBULAT" or dossier_sinistre.type_priseencharge.pk == 4 or dossier_sinistre.type_priseencharge.pk == 5 or dossier_sinistre.type_priseencharge.pk == 6:
            template_path = 'sinistre_printer_ambulatoire.html'
        elif dossier_sinistre.type_priseencharge.code == "HOSPIT":
            if sinistres:
                sinistre = sinistres[0]
                prorogations = ProrogationSinistre.objects.filter(sinistre_id=sinistre.id)
                for prorogation in prorogations:
                    texte_nombre_demande = texte_nombre_demande + " + " + str(prorogation.jour_demande)
                    texte_nombre_accorde = texte_nombre_accorde + " + " + str(prorogation.jour_accorde)

                    # dd(texte_nombre_demande)

            template_path = 'sinistre_printer_hospitalisation.html'
        else:
            return HttpResponse('Type de prise en charge inconnu.')

        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="bon_' + dossier_sinistre.numero + '.pdf"'
        # find the template and render it.
        template = get_template(template_path)
        context = {'id': id, "data": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10], "dossier_sinistre": dossier_sinistre,
                   "sinistres": sinistres, 'plafond_atteint': plafond_atteint, 'garantis_pharmacie': garantis_pharmacie,
                   'ticket_moderateur_consultation': ticket_moderateur_consultation,
                   'texte_nombre_demande': texte_nombre_demande, 'texte_nombre_accorde': texte_nombre_accorde,
                   'tm_prefinanced': tm_prefinanced, 'request_user_is_prestataire': request_user_is_prestataire, }
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

        # if pisa_status.err:
        #    return HttpResponse('We had some errors <pre>' + html + '</pre>')
        # return response

        pdf = render_pdf(template_path, context_dict=context)
        return HttpResponse(File(pdf), content_type='application/pdf')

    else:
        return redirect('/')


def render_pdf_view_general(request, id):
    dossiers_sinistres = DossierSinistre.objects.filter(id=id, statut_validite=StatutValidite.VALIDE)

    if dossiers_sinistres:

        dossier_sinistre = dossiers_sinistres.first()

        # sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut='ACCORDE')
        sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut__in=('ACCORDE', 'REJETE'))

        # Added on 05102023: add tm préfinancé
        request_user_is_prestataire = request.user.is_prestataire

        tm_prefinanced = False
        if dossier_sinistre.type_prefinancement and dossier_sinistre.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        ticket_moderateur_consultation = 0
        ticket_moderateur_pharmacie = 0
        ticket_moderateur_pharmacie = get_ticket_moderateur_pharmacie(dossier_sinistre.aliment_id,
                                                                      dossier_sinistre.formulegarantie_id,
                                                                      dossier_sinistre.date_survenance)
        pprint("ticket_moderateur_pharmacie")
        pprint(ticket_moderateur_pharmacie)

        garantis_pharmacie = dossier_sinistre.formulegarantie.garantis_pharmacie
        plafond_atteint = False

        texte_nombre_demande = ""
        texte_nombre_accorde = ""

        if dossier_sinistre.type_priseencharge.code == "CONSULT":

            # vérification si vitamine et vaccins non garanti
            '''bareme_vitamine = Bareme.objects.filter(acte__code="VITAMINES",
                                                    formulegarantie_id=dossier_sinistre.aliment.formule.id, )
            bareme_vaccin = Bareme.objects.filter(acte__code="VACCIN",
                                                  formulegarantie_id=dossier_sinistre.aliment.formule.id, )

            if bareme_vitamine:
                vitamines_garanti = bareme_vitamine.is_garanti

            if bareme_vaccin:
                vaccins_garanti = bareme_vaccin.is_garanti
            '''

            plafond_atteint = bool_plafond_atteint(dossier_sinistre)

            if sinistres:
                sinistre = sinistres[0]
                ticket_moderateur_consultation = sinistre.bareme.taux_tm if sinistre.bareme else sinistre.formulegarantie.taux_tm
                # a revoir: prendre le plafond sur la pharmacie et non sur la consultation

            # Added on 16022024
            ticket_moderateur_consultation = ticket_moderateur_pharmacie

            template_path = 'sinistre_printer_general.html'

        else:
            return HttpResponse('Type de prise en charge inconnu.')

        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="bon_' + dossier_sinistre.numero + '.pdf"'
        # find the template and render it.
        template = get_template(template_path)
        context = {'id': id, "data": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10], "dossier_sinistre": dossier_sinistre,
                   "sinistres": sinistres, 'plafond_atteint': plafond_atteint, 'garantis_pharmacie': garantis_pharmacie,
                   'ticket_moderateur_consultation': ticket_moderateur_consultation,
                   'texte_nombre_demande': texte_nombre_demande, 'texte_nombre_accorde': texte_nombre_accorde,
                   'tm_prefinanced': tm_prefinanced, 'request_user_is_prestataire': request_user_is_prestataire, }
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

        # if pisa_status.err:
        #    return HttpResponse('We had some errors <pre>' + html + '</pre>')
        # return response

        pdf = render_pdf(template_path, context_dict=context)
        return HttpResponse(File(pdf), content_type='application/pdf')

    else:
        return redirect('/')


## generation pdf dossier sinistre pour le profil pharmacien
def render_pdf_view_pharmacie(request, id):
    dossiers_sinistres = DossierSinistre.objects.filter(id=id, statut_validite=StatutValidite.VALIDE)

    if dossiers_sinistres:

        dossier_sinistre = dossiers_sinistres.first()

        # sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut='ACCORDE')
        #   sinistres = Sinistre.objects.filter(dossier_sinistre__id=id, statut__in=('ACCORDE', 'REJETE'))
        sinistres = Sinistre.objects.filter(acte__type_acte__libelle='MEDICAMENT', dossier_sinistre__id=id, statut__in=('ACCORDE', 'REJETE'))
        # for s in sinistres:
        #     print(s)
        #     print('--------------')

        # Added on 05102023: add tm préfinancé
        request_user_is_prestataire = request.user.is_prestataire

        tm_prefinanced = False
        if dossier_sinistre.type_prefinancement and dossier_sinistre.type_prefinancement.code == "PREF_TOUT":
            tm_prefinanced = True if request.user.is_prestataire else False

        ticket_moderateur_consultation = 0
        ticket_moderateur_pharmacie = 0
        ticket_moderateur_pharmacie = get_ticket_moderateur_pharmacie(dossier_sinistre.aliment_id,
                                                                      dossier_sinistre.formulegarantie_id,
                                                                      dossier_sinistre.date_survenance)
        pprint("ticket_moderateur_pharmacie")
        pprint(ticket_moderateur_pharmacie)

        garantis_pharmacie = dossier_sinistre.formulegarantie.garantis_pharmacie
        plafond_atteint = False

        texte_nombre_demande = ""
        texte_nombre_accorde = ""

        if dossier_sinistre.type_priseencharge.code == "CONSULT":

            # vérification si vitamine et vaccins non garanti
            '''bareme_vitamine = Bareme.objects.filter(acte__code="VITAMINES",
                                                    formulegarantie_id=dossier_sinistre.aliment.formule.id, )
            bareme_vaccin = Bareme.objects.filter(acte__code="VACCIN",
                                                  formulegarantie_id=dossier_sinistre.aliment.formule.id, )

            if bareme_vitamine:
                vitamines_garanti = bareme_vitamine.is_garanti

            if bareme_vaccin:
                vaccins_garanti = bareme_vaccin.is_garanti
            '''

            plafond_atteint = bool_plafond_atteint(dossier_sinistre)

            if sinistres:
                sinistre = sinistres[0]
                ticket_moderateur_consultation = sinistre.bareme.taux_tm if sinistre.bareme else sinistre.formulegarantie.taux_tm
                # a revoir: prendre le plafond sur la pharmacie et non sur la consultation

            # Added on 16022024
            ticket_moderateur_consultation = ticket_moderateur_pharmacie

            template_path = 'sinistre_printer_pharmacie.html'

        else:
            return HttpResponse('Type de prise en charge inconnu.')

        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="bon_' + dossier_sinistre.numero + '.pdf"'
        # find the template and render it.
        template = get_template(template_path)
        context = {'id': id, "data": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10], "dossier_sinistre": dossier_sinistre,
                   "sinistres": sinistres, 'plafond_atteint': plafond_atteint, 'garantis_pharmacie': garantis_pharmacie,
                   'ticket_moderateur_consultation': ticket_moderateur_consultation,
                   'texte_nombre_demande': texte_nombre_demande, 'texte_nombre_accorde': texte_nombre_accorde,
                   'tm_prefinanced': tm_prefinanced, 'request_user_is_prestataire': request_user_is_prestataire, }
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse('We had some errors <pre>' + html + '</pre>')
        return response

        # if pisa_status.err:
        #    return HttpResponse('We had some errors <pre>' + html + '</pre>')
        # return response

        pdf = render_pdf(template_path, context_dict=context)
        return HttpResponse(File(pdf), content_type='application/pdf')

    else:
        return redirect('/')


class GenerateFactureView(ListView):
    model = DossierSinistre
    template_name = 'generation_bordereau.html'
    context_object_name = 'dossiers_sinistres'

    def get_queryset(self):
        # print(self.request.user.prestataire)
        return []

    def get_context_data(self, **kwargs):
        # curent_data = datetime.datetime.now(tz=timezone.utc)
        # print(curent_data.month)
        # print(curent_data.year)

        type_remboursements = TypeRemboursement.objects.filter(status=True).order_by('libelle')
        periode_comptable = PeriodeComptable.objects.all().order_by('annee', 'mois')

        # liste_sinistres_sans_factures = Sinistre.objects.filter(police__bureau=self.request.user.bureau, dossier_sinistre__isnull=False, facture_prestataire__isnull=True, statut=StatutSinistre.ACCORDE, statut_validite=StatutValidite.VALIDE).values('prestataire_id')
        #
        # prestataire_ids = liste_sinistres_sans_factures.filter(dossier_sinistre__type_remboursement__code="TP").values_list('prestataire_id', flat=True)
        # prestataires = Prestataire.objects.filter(bureau=self.request.user.bureau, id__in=prestataire_ids).order_by('name')
        #
        # adherent_principal_ids = liste_sinistres_sans_factures.filter(dossier_sinistre__type_remboursement__code="RD").values_list('adherent_principal_id', flat=True)
        # adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')


        #sql_tp = '''
        #    SELECT distinct prestataire_id
        #    FROM sinistres
        #    INNER JOIN polices ON polices.id = sinistres.police_id
        #    WHERE polices.bureau_id = %s
        #    AND dossier_sinistre_id IS NOT NULL
        #    AND facture_prestataire_id IS NULL
        #    AND sinistres.statut = %s
        #    AND sinistres.statut_validite = %s
        #'''

        #params = [self.request.user.bureau_id, StatutSinistre.ACCORDE, StatutValidite.VALIDE]
        #data, columns = execute_query_with_params(sql_tp, params)

        #prestataire_ids = [row[0] for row in data]
        #prestataires = Prestataire.objects.filter(id__in=prestataire_ids).order_by('name')

        #adherent_principal_ids = [row[0] for row in data]
        #adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')

        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
            "periode_comptable": periode_comptable,
            # "prestataires": prestataires,
            "type_remboursements": type_remboursements,
            # "adhs": adhs
        }

def get_prestataires_generate_facture(request):
    prestataire_ids = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,dossier_sinistre__isnull=False,
                                                            facture_prestataire__isnull=True,
                                                            statut=StatutSinistre.ACCORDE,
                                                            dossier_sinistre__type_remboursement__code="TP",
                                                            statut_validite=StatutValidite.VALIDE).values('prestataire_id')

    prestataires = Prestataire.objects.filter(bureau=request.user.bureau, id__in=prestataire_ids).order_by('name')
    data = []

    for prestataire in prestataires:
        data.append({"id": prestataire.id, "name": prestataire.name})

    return JsonResponse({
        "data": data,
    })

def search_prestataires_generate_facture_by_name_datatable(request):
    items_per_page = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    page_number = start // items_per_page + 1  # Calculate page number correctly
    search_prestataire_m = request.GET.get('search_prestataire_m', '')



    # Filter based on search_prestataire_m
    if search_prestataire_m:
        # prestataire_ids = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        #                                           dossier_sinistre__isnull=False,
        #                                           facture_prestataire__isnull=True,
        #                                           statut=StatutSinistre.ACCORDE,
        #                                           dossier_sinistre__type_remboursement__code="TP",
        #                                           statut_validite=StatutValidite.VALIDE).values('prestataire_id')

        # queryset = Prestataire.objects.filter(Q(nom__icontains=search_prestataire_m) | Q(prenoms__icontains=search_prestataire_m) & Q(id__in=adherent_principal_ids)).distinct()

        queryset = Prestataire.objects.filter(Q(name__icontains=search_prestataire_m), bureau=request.user.bureau, status=1).order_by('name')
        #queryset = queryset.filter(Q(name__icontains=search_prestataire_m)).distinct()
    else:
        queryset = Prestataire.objects.filter(id__isnull=True)

    print("queryset")
    print(queryset)
    print(items_per_page)
    print(start)
    print(page_number)

    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)
    data = []

    # nom_prenom = f'{c.nom} {c.prenoms}' if c.nom else c.prenoms
    data = [
        {
            "id": c.id,
            "prestataire__code": f'{c.code}',
            "prestataire__name": c.name,
            "actions": f'<span style="cursor:pointer;" class="btn_select_prestataire" data-presta_id="{c.id}" data-presta_name="{c.name} ({c.code})"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-plus"></i> Sélectionner</span></span>',
        }
        for c in page_obj
    ]

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })

def search_assures_generate_facture_by_name_datatable(request):
    items_per_page = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    page_number = start // items_per_page + 1  # Calculate page number correctly
    search_assure_m = request.GET.get('search_assure_m', '')



    # Filter based on search_assure_m
    if search_assure_m:
        # prestataire_ids = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        #                                           dossier_sinistre__isnull=False,
        #                                           facture_prestataire__isnull=True,
        #                                           statut=StatutSinistre.ACCORDE,
        #                                           dossier_sinistre__type_remboursement__code="TP",
        #                                           statut_validite=StatutValidite.VALIDE).values('prestataire_id')

        # queryset = Prestataire.objects.filter(Q(nom__icontains=search_assure_m) | Q(prenoms__icontains=search_assure_m) & Q(id__in=adherent_principal_ids)).distinct()

        queryset = Client.objects.annotate(full_name=Concat(F('nom'), Value(' '), F('prenoms'))).filter(Q(full_name__icontains=search_assure_m) | Q(code__icontains=search_assure_m), bureau=request.user.bureau, statut=Statut.ACTIF).order_by('nom')
        #queryset = queryset.filter(Q(name__icontains=search_assure_m)).distinct()
    else:
        queryset = Client.objects.filter(id__isnull=True)

    print("queryset")
    print(queryset)
    print(items_per_page)
    print(start)
    print(page_number)

    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)
    data = []

    # full_name = f'{c.nom} {c.prenoms}' if c.nom else c.prenoms
    data = [
        {
            "id": c.id,
            "assure__code": f'{c.code}',
            "assure__name": f'{c.nom} {c.prenoms if c.prenoms else ""}',
            "actions": f'<span style="cursor:pointer;" class="btn_select_assure" data-assure_id="{c.id}" data-assure_name="{c.nom} {c.prenoms if c.prenoms else ""} ({c.code})"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-plus"></i> Sélectionner</span></span>',
        }
        for c in page_obj
    ]

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })




def search_adherents_generate_facture_by_name_datatable(request):
    items_per_page = int(request.GET.get('length', 10))
    start = int(request.GET.get('start', 0))
    page_number = start // items_per_page + 1  # Calculate page number correctly
    search_beneficiaire = request.GET.get('search_beneficiaire', '')



    # Filter based on search_beneficiaire
    if search_beneficiaire:
        # adherent_principal_ids = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        #                                                  dossier_sinistre__isnull=False,
        #                                                  facture_prestataire__isnull=True,
        #                                                  statut=StatutSinistre.ACCORDE,
        #                                                  dossier_sinistre__type_remboursement__code="RD",
        #                                                  statut_validite=StatutValidite.VALIDE).values('adherent_principal_id')

        # queryset = Aliment.objects.filter(Q(nom__icontains=search_beneficiaire) | Q(prenoms__icontains=search_beneficiaire) & Q(id__in=adherent_principal_ids)).distinct()

        # queryset = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')
        queryset = Aliment.objects.annotate(full_name=Concat(F('nom'), Value(' '), F('prenoms'))).filter(
            Q(full_name__icontains=search_beneficiaire),
            bureau=request.user.bureau,
            qualite_beneficiaire__code="AD"
        ).distinct().order_by('nom')

    else:
        queryset = Aliment.objects.filter(id__isnull=True)

    print("queryset")
    print(queryset)
    print(items_per_page)
    print(start)
    print(page_number)

    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)
    data = []

    # nom_prenom = f'{c.nom} {c.prenoms}' if c.nom else c.prenoms
    data = [
        {
            "id": c.id,
            "aliment__cartes__numero": f'{c.carte_active()}' if c.carte_active else "",
            "aliment__nom": c.nom or "",
            "aliment__prenoms": c.prenoms or "",
            "actions": f'<span style="cursor:pointer;" class="btn_select_beneficiaire" data-benef_id="{c.id}" data-benef_nom="{c.nom} {c.prenoms}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-plus"></i> Sélectionner</span></span>',
        }
        for c in page_obj
    ]

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def generate_facture_datatable(request):
    items_per_page = 25
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = request.GET.get('order[0][column]')
    sort_direction = request.GET.get('order[0][dir]')

    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_adh = request.GET.get('search_adh', '')
    search_assure = request.GET.get('search_assure', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    select_all = request.GET.get('all', '')
    search = request.GET.get('search[value]', '')



    if request.user.is_prestataire:
        search_prestataire = request.user.prestataire.pk
        search_type_remboursement = 1

    # typeRemboursement = None

    pprint("search_periode_comptable")
    pprint(search_periode_comptable)

    pprint("search_prestataire")
    pprint(search_prestataire)

    pprint("search_adh")
    pprint(search_adh)

    pprint("search_assure")
    pprint(search_assure)

    pprint("search_type_remboursement")
    pprint(search_type_remboursement)




    if search_periode_comptable and (search_prestataire or search_adh or search_assure) and search_type_remboursement:

        if search_type_remboursement:
            typeRemboursement = TypeRemboursement.objects.get(id=int(search_type_remboursement))

        if request.user.is_pharm:
            queryset = Sinistre.objects.filter(prestataire=request.user.prestataire,
                                               statut=StatutSinistre.ACCORDE,
                                               facture_prestataire__isnull=True,
                                               dossier_sinistre__isnull=False, #pour ne plus générer de factures pour les sinistres de la v1
                                               statut_bordereau=StatutSinistreBordereau.ATTENTE).order_by('-id')
            # print("pharm")
            # print(queryset)
            # if x.statut == StatutSinistre.ACCORDE and -- si on voulais que tout le bordereau soit accordé d'abord

        # autres prestataires
        elif request.user.is_pres or request.user.is_imag or request.user.is_optic or request.user.is_labo or request.user.is_dentaire or request.user.is_med:
            queryset = Sinistre.objects.filter(prestataire=request.user.prestataire,
                                               statut=StatutSinistre.ACCORDE,
                                               facture_prestataire__isnull=True,
                                               dossier_sinistre__isnull=False, #pour ne plus générer de factures pour les sinistres de la v1
                                               statut_bordereau=StatutSinistreBordereau.ATTENTE).order_by(
                '-id')

        else:
            queryset = Sinistre.objects.filter(statut=StatutSinistre.ACCORDE,
                                               # dossier_sinistre__mode_creation__code='SAISIE-GESTIONNAIRE',
                                               facture_prestataire__isnull=True,
                                               dossier_sinistre__isnull=False, #pour ne plus générer de factures pour les sinistres de la v1
                                               statut_bordereau=StatutSinistreBordereau.ATTENTE).order_by(
                '-id')



        #rendre le filtre par période obligatoire
        periode_comptable = PeriodeComptable.objects.get(id=int(search_periode_comptable))
        queryset = queryset.filter(created_at__month=periode_comptable.mois, created_at__year=periode_comptable.annee)

        if typeRemboursement:
            queryset = queryset.filter(dossier_sinistre__type_remboursement__code=typeRemboursement.code)

            if typeRemboursement.code == "TP":
                pprint("DANS LE CAS DU TIERS PAYANT")
                if search_prestataire:
                    queryset = queryset.filter(prestataire__id=int(search_prestataire))
            elif typeRemboursement.code == "RD":
                pprint("DANS LE CAS DU REMBOURSEMENT DIRECT")
                if search_adh:
                    queryset = queryset.filter(aliment__adherent_principal__id=int(search_adh))
                elif search_assure:
                    queryset = queryset.filter(police__client__id=int(search_assure))





        if search:
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(dossier_sinistre__numero__icontains=search) |
                Q(prestataire__name__icontains=search) |
                Q(aliment__nom__icontains=search) |
                Q(aliment__prenoms__icontains=search) |
                Q(aliment__cartes__numero__icontains=search) |
                Q(acte__libelle__icontains=search)
            )


        if select_all:
            data_ids = [x.id for x in queryset]
            print(data_ids)
            print(len(data_ids))
            return JsonResponse({
                "data": data_ids,
            })

    else:
        pprint("Pas de prestataire ni de période renseigné")
        queryset = queryset = Sinistre.objects.filter(id=0)

    # Map column index to corresponding model field for sorting
    # sort_columns = {
    #     0: '-numero',
    #     1: 'aliment__nom',
    #     2: 'statut',
    #     # Add more columns as needed
    # }

    # # Default sorting by 'id' if column index is not found
    # sort_column = sort_columns.get(sort_column_index, '-id')
    #
    # if sort_direction == 'desc':
    #     sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_dossier_sinistre_url = reverse('details_dossier_sinistre',
                                              args=[
                                                  c.dossier_sinistre.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_dossier_sinistre_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        statut_html = f'<span class="badge badge-{c.statut.lower()}">{c.statut}</span>'

        # total_facture = c.total_frais_reel if c.total_frais_reel else 0
        # total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        # total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie

        total_frais_reel = c.total_frais_reel
        total_part_compagnie = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        # checkbox = f'<input type="checkbox" name="sinistres_ids[]" value="{c.id}" data-sinistre_id="{c.id}" class ="select-row sinistres">'
        nom_beneficiaire = f'{c.aliment.nom if c.aliment.nom else ""} {c.aliment.prenoms if c.aliment.prenoms else ""}'

        data_iten = {
            "id": c.id,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "numero": c.numero,
            "dossier_sinistre__numero": c.dossier_sinistre.numero if c.dossier_sinistre else '',
            "prestataire": c.prestataire.name,
            "beneficiaire": nom_beneficiaire,
            "carte_active": numero_carte,
            "acte": c.acte.libelle,
            "total_frais_reel": money_field(total_frais_reel),
            "total_part_assure": money_field(total_part_assure),
            "total_part_compagnie": money_field(total_part_compagnie),
            "statut": statut_html,
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


class ListeBordereauView(TemplateView):
    model = FacturePrestataire
    template_name = 'liste_bordereau.html'
    #context_object_name = 'bordereaux'

    def get(self, request, *args, **kwargs):
        factures = FacturePrestataire.objects.filter(prestataire__isnull=False, prestataire=self.request.user.prestataire, statut_validite=StatutValidite.VALIDE).exclude(statut="ANNULER")

        context = self.get_context_data(**kwargs)
        context['factures_non_payees'] = factures.exclude(statut=SatutBordereauDossierSinistres.PAYE)
        context['factures_payees'] = factures.filter(statut=SatutBordereauDossierSinistres.PAYE)
        context['annees_comptables'] = PeriodeComptable.objects.values_list('annee', flat=True).distinct().order_by('-annee')

        #dd(factures.count())
        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def liste_facture_datatable(request):
    items_per_page = 25
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = request.GET.get('order[0][column]')
    sort_direction = request.GET.get('order[0][dir]')

    search_annee_comptable = request.GET.get('search_annee_comptable', '')

    select_all = request.GET.get('all', '')
    search = request.GET.get('search', '')

    #dd(search_annee_comptable)

    pprint(search)

    # si l'utilisateur n'est pas un prestataire
    if not request.user.prestataire:
        return JsonResponse({
            "data": [],
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "draw": int(request.GET.get('draw', 1)),
        })



    if search_annee_comptable:

        queryset = FacturePrestataire.objects.filter(periode_comptable__annee=search_annee_comptable, prestataire=request.user.prestataire, statut_validite=StatutValidite.VALIDE, statut=SatutBordereauDossierSinistres.PAYE).exclude(statut="ANNULER").order_by('-id')
            # print("pharm")
            # print(queryset)
            # if x.statut == StatutSinistre.ACCORDE and -- si on voulais que tout le bordereau soit accordé d'abord


        if search:
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(prestataire__name__icontains=search) #|
                #Q(net_a_payer__icontains=search) |
                #Q(created_at__icontains=search)
            )



        """
        if select_all:
            data_ids = [x.id for x in queryset]
            print(data_ids)
            print(len(data_ids))
            return JsonResponse({
                "data": data_ids,
            })
        """

    else:
        pprint("Pas de période renseignée")
        queryset = FacturePrestataire.objects.filter(statut="nevajamaisexister")


    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_facture_url = reverse('detail_facture',
                                              args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_facture_url}"><span class="btn btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détail</span></a>&nbsp;&nbsp;'

        data_iten = {
            "id": c.id,
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else '',
            "numero": c.numero,
            "prestataire": c.prestataire.name if c.prestataire else '',
            "net_a_payer": money_field(c.net_a_payer),
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# TODO : DetailBordereauView
@method_decorator(login_required, name='dispatch')
class DetailBordereauView(TemplateView):
    # permission_required = "sinistre.view_sinistre"
    template_name = 'detail_bordereau.html'
    model = Sinistre

    def get(self, request, bordereau_id, *args, **kwargs):
        
        bordereau = FacturePrestataire.objects.filter(id=bordereau_id, bureau=request.user.bureau).first()
        
        if not bordereau:
            raise ValueError("Bordereau non trouvé.")

        # Récupération de l'ordonnancement à partir du sinistre quelconque
        sinistre = Sinistre.objects.filter(facture_prestataire=bordereau).first()
        
        if not sinistre:
            raise ValueError("Pas de sinistre non trouvé pour le bordereau donné.")
        
        
        bordereau_ordonnancement_id = sinistre.bordereau_ordonnancement_id

        # Récupération des paiements comptables
        bordereau_payes = PaiementComptable.objects.filter(
            bordereau_ordonnancement_id=bordereau_ordonnancement_id,
        ).first()
        
        if bordereau:
            liste_sinistres_bordereau = Sinistre.objects.filter(facture_prestataire=bordereau)
            montant_remb_total = bordereau.net_a_payer
            montant_rejet_total = 0
            montant_accepte_total = 0
            is_all_processed = True

            for sinistre in liste_sinistres_bordereau:
                montant_rejet_total = montant_rejet_total + (
                    sinistre.montant_remb_refuse if sinistre.montant_remb_refuse else 0)
                montant_accepte_total = montant_accepte_total + (
                    sinistre.montant_remb_accepte if sinistre.montant_remb_accepte else 0)
                if not sinistre.is_ges_processed:
                    is_all_processed = False

            context = self.get_context_data(**kwargs)
            context['bordereau'] = bordereau
            context['bordereau_payes'] = bordereau_payes
            context['liste_sinistres_bordereau'] = liste_sinistres_bordereau
            context['montant_remb_total'] = montant_remb_total
            context['montant_rejet_total'] = montant_rejet_total
            context['montant_accepte_total'] = montant_accepte_total
            context['is_all_processed'] = is_all_processed
            context[
                'is_all_processed_validated'] = True if bordereau.statut == SatutBordereauDossierSinistres.VALIDE else False

            return self.render_to_response(context)

        else:

            return redirect("/")

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def details_bordereau_prestataire_datatable(request, bordereau_id, *args, **kwargs):
    items_per_page = 10
    page_number = request.GET.get('page')
    # start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search = request.GET.get('search[value]', '')
    select_all = request.GET.get('all', '')
    sort = ''
    print("sort_column_index")
    print(sort_column_index)
    print("sort_direction")
    print(sort_direction)



    bordereau = FacturePrestataire.objects.get(id=bordereau_id, bureau=request.user.bureau)
    queryset = Sinistre.objects.filter(facture_prestataire=bordereau)

    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) |
            Q(dossier_sinistre__numero__icontains=search) |
            Q(prestataire__name__icontains=search) |
            Q(aliment__nom__icontains=search) |
            Q(aliment__prenoms__icontains=search) |
            Q(aliment__cartes__numero__icontains=search) |
            Q(acte__libelle__icontains=search)
        ).distinct()

    if select_all:
        data_ids = [x.id for x in queryset if not x.is_ges_processed]
        print(data_ids)
        print(len(data_ids))
        return JsonResponse({
            "data": data_ids,
        })

    sort_columns = {
        0: 'id',
        12: 'is_ges_processed',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, '-id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    print('@@@ sort_column @@@')
    print(sort_column)
    queryset = queryset.order_by(sort_column)

    # collecte des montants totaux

    for x in queryset:
        total_base_remb = x.total_frais_reel if x.tm_prefinanced else x.total_part_compagnie
        total_rejet = x.montant_remb_refuse if x.montant_remb_refuse else 0
        total_net_payer = x.montant_remb_accepte if x.montant_remb_accepte else 0



    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        if c.is_ges_processed:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-success" style="font-size:10px;color:white;">TRAITÉ</span>'
        else:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-en-attente" style="font-size:10px;color:white;">NON TRAITÉ</span>'

        if not request.user.is_prestataire:
            popup_details_sinistre = reverse('popup_details_sinistre', args=[c.id])
            if c.is_ges_processed:
                action_html = f'<a href="javascript:0;"><span title="Traiter le remboursement" data-href ="{popup_details_sinistre}" class="btn-sm btn-details rounded-pill badge-en-attente btn-popup_details_sinistre" >Detail</span></a>'
            else:
                action_html = f'<a href="javascript:0;"><span title="Traiter le remboursement" data-href ="{popup_details_sinistre}" class="btn-sm btn-details rounded-pill btn-popup_details_sinistre" ><i class="fa fa-edit"></i>Traiter</span></a>'
        else:
            detail_dossier_url = reverse('details_dossier_sinistre', args=[c.dossier_sinistre.id])
            action_html = f'<a href="{detail_dossier_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>'


        total_facture = c.total_frais_reel if c.total_frais_reel else 0
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_rejet = c.montant_remb_refuse if c.montant_remb_refuse else 0
        total_net_payer = c.montant_remb_accepte if c.montant_remb_accepte else 0

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        # checkbox = f'<input type="checkbox" name="sinistres_ids[]" value="{c.id}" data-sinistre_id="{c.id}" class ="select-row sinistres">'
        nom_beneficiaire = f'{c.aliment.nom if c.aliment.nom else ""} {c.aliment.prenoms if c.aliment.prenoms else ""}'

        data_iten = {
            "id": c.id,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "numero": c.numero,
            "dossier_sinistre__numero": c.dossier_sinistre.numero if c.dossier_sinistre else "",
            # "is_ges_processed": c.is_ges_processed,
            "beneficiaire": nom_beneficiaire,
            "carte_active": numero_carte,
            "acte": c.acte.libelle,
            "total_facture": money_field(total_facture),
            "total_part_assure": money_field(total_part_assure),
            "total_base_remb": money_field(total_base_remb),
            "total_rejet": money_field(total_rejet),
            "total_net_payer": money_field(total_net_payer),
            "statut": statut_html,
            "actions": action_html,
            "is_ges_processed": c.is_ges_processed,
        }

        data.append(data_iten)

    data_ids = [x['id'] for x in data if not x['is_ges_processed']]

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
        "selectedItems": data_ids,
    })

# def details_bordereau_prestataire_datatable():
#     pass
#     # detail_sinistre_url = reverse('popup_details_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
    # actions_html = f'<span title="Traiter le remboursement" data-href ="{detail_sinistre_url}" class="btn badge btn-sm btn-details rounded-pill btn-popup_details_sinistre" >Traiter</span>'


def facture_pdf(request):
    data = Sinistre.objects.filter(id__in=['1154767', '1154766', '1154765', '1154764', '1154763', '1154762', '1154761', '1154760', '1154759', '1154758', '1154757', '1154756', '1154755', '1154754', '1154753', '1154752', '1154751', '1154750', '1154749', '1154748', '1154747', '1154746',1154736, 1154735, 1154734, 1154733, 1154732])
    facture = data.first().facture_prestataire
    pprint("facture")
    pprint(facture)
    net_a_payer = 0
    part_assure = 0
    montant_reel = 0
    sinistes_pdf = []
    sinistres_par_adherent = defaultdict(list)
    # A COMPLETER AVEC LE CAS OU LA FACTURE EST GÉNÉRÉE PAR UN GESTIONNAIRE

    for sinistre in data:
        sinistre.facture_prestataire = facture
        sinistre.statut_bordereau = StatutSinistreBordereau.AJOUTE_BORDEREAU
        sinistre.statut_remboursement = StatutRemboursement.DEMANDE
        sinistre.save()

        print("sinistre.adherent_principal")
        print(sinistre.adherent_principal)

        # if sinistre.type_prefinancement.code == "PREF_TOUT":
        if sinistre.tm_prefinanced:
            part_assure = part_assure + 0
            net_a_payer = net_a_payer + float(sinistre.total_frais_reel)

        else:
            part_assure = part_assure + float(sinistre.total_part_assure)
            net_a_payer = net_a_payer + float(sinistre.total_part_compagnie)

        # part_assure = part_assure + sinistre.total_part_assure
        montant_reel = montant_reel + float(sinistre.total_frais_reel)
        adherent_id = sinistre.adherent_principal.id
        sinistres_par_adherent[adherent_id].append(sinistre)
        sinistes_pdf.append(sinistre)

    facture.net_a_payer = net_a_payer

    net_a_payer_lettre = num2words(net_a_payer, lang="fr")

    pprint(net_a_payer_lettre)
    pprint("sinistres_par_adherent")
    pprint(sinistres_par_adherent)

    sinistres_par_adherent2 = []
    nombre_sinistres=0
    # Affichage du regroupement
    for adherent_id, sinistres in sinistres_par_adherent.items():
        print(f"Adhérent {adherent_id} :")

        net_a_payer2 = 0
        part_assure2 = 0
        montant_reel2 = 0

        for sinistre in sinistres:
            print(f"  - {sinistre}")
            if sinistre.tm_prefinanced:
                part_assure2 = part_assure2 + 0
                net_a_payer2 = net_a_payer2 + float(sinistre.total_frais_reel)

            else:
                part_assure2 = part_assure2 + float(sinistre.total_part_assure)
                net_a_payer2 = net_a_payer2 + float(sinistre.total_part_compagnie)

            # part_assure = part_assure + sinistre.total_part_assure
            montant_reel2 = montant_reel2 + float(sinistre.total_frais_reel)

        sinistres_par_adherent2.append({
            "adherent": sinistres[0].adherent_principal.nom_prenoms,
            "sinistres": sinistres,
            "nombre_sinistres": len(sinistres),
            "sous_net_a_payer": net_a_payer2,
        })
        nombre_sinistres = nombre_sinistres + len(sinistres)

    pprint("sinistres_par_adherent2")
    pprint(sinistres_par_adherent2)
    # Added on 06112023: add tm préfinancé
    request_user_is_prestataire = request.user.is_prestataire
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/bordereau_paiement_assure.html', context_dict={'sinistres': sinistes_pdf, 'nombre_sinistres': nombre_sinistres, 'sinistres_par_adherent':sinistres_par_adherent2, 'bordereau': facture,
                                               'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel,
                                               'net_a_payer_lettre': net_a_payer_lettre,
                                               'request_user_is_prestataire': request_user_is_prestataire,
                                               'currency_code': currency_code})
    # liste off 100 data
    # return render(request, 'courriers/bordereau_paiement.html', context={'data': data})
    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)
    print(nombre_pages)

    pdf = render_pdf('courriers/bordereau_paiement_assure.html', context_dict={'sinistres': sinistes_pdf, 'nombre_sinistres': nombre_sinistres, 'sinistres_par_adherent':sinistres_par_adherent2, 'bordereau': facture,
                                               'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel,
                                               'net_a_payer_lettre': net_a_payer_lettre,
                                               'nombre_pages': nombre_pages,
                                               'request_user_is_prestataire': request_user_is_prestataire,
                                               'currency_code': currency_code})

    return HttpResponse(File(pdf), content_type='application/pdf')


@transaction.atomic
def submit_generate_bordereau(request):
    # TODO : CODE_PAYS_YY_MM
    # try:
    with transaction.atomic():

        periode_comptable = PeriodeComptable.objects.get(id=int(request.POST.get('periode_id')))


        #print("periode " + request.POST.get('periode_id'))
        #print("search adh " + request.POST.get('search_adh'))
        #print("prestataire " + request.POST.get('prestataire_id'))
        #print("selectedItems" + request.POST.get('selectedItems'))
        #pprint("ADH ID")
        #print(request.POST.get('search_adh'))

        sinistres_ids = literal_eval(request.POST.get('selectedItems'))
        print("sinistres_ids")
        print(sinistres_ids)

        search_periode_comptable = request.GET.get('periode_id', '')
        search_prestataire = request.POST.get('prestataire_id', '')
        search_assure = request.POST.get('search_assure', '')
        search_type_remboursement = request.POST.get('type_remboursement_id', '')

        print("type_remboursement_id " + request.POST.get('type_remboursement_id'))

        is_assure = False

        endcode_facture = ""
        if request.user.is_prestataire:
            endcode_facture = "FP"

        else:
            if search_type_remboursement:
                typeRemboursement = TypeRemboursement.objects.get(id=int(search_type_remboursement))

                if typeRemboursement.code == "TP":
                    endcode_facture = "FP"
                elif typeRemboursement.code == "RD":
                    endcode_facture = "FRD"



        search_adh = request.POST.get('search_adh', '')


        sinistres = Sinistre.objects.filter(id__in=sinistres_ids)

        if sinistres:

            if not request.user.prestataire:
                pprint("ENTRE DANS ")
                pprint("VOICI SEARCH PRESTATAIRE " + search_prestataire )
                if search_prestataire != '' and typeRemboursement.code == "TP":
                    pprint("PRESTATAIRE")
                    prestataire = Prestataire.objects.get(id=int(request.POST.get('prestataire_id')))
                    facture = FacturePrestataire.objects.create(
                        created_by=request.user,
                        bureau=request.user.bureau,
                        periode_comptable=periode_comptable,
                        prestataire=prestataire,
                        type_remboursement=TypeRemboursement.objects.get(code="TP"),
                    )

                else:
                    if search_assure != '' and typeRemboursement.code == "RD":
                        pprint("REMBOURSEMENT DIRECT")
                        pprint("ASSURE ID =  " + search_assure)
                        assure = Client.objects.get(id=int(search_assure))
                        facture = FacturePrestataire.objects.create(
                            type_remboursement=TypeRemboursement.objects.get(code="RD"),
                            created_by=request.user,
                            bureau=request.user.bureau,
                            periode_comptable=periode_comptable,
                            assure=assure,

                        )
                        is_assure = True
                    else:
                        pprint("REMBOURSEMENT DIRECT")
                        pprint("ADHERENT PRINCIPAL ID =  " + search_adh)
                        adherent_principal = Aliment.objects.get(id=int(search_adh))
                        facture = FacturePrestataire.objects.create(
                            type_remboursement=TypeRemboursement.objects.get(code="RD"),
                            created_by=request.user,
                            bureau=request.user.bureau,
                            periode_comptable=periode_comptable,
                            adherent_principal=adherent_principal,

                        )
            else:
                facture = FacturePrestataire.objects.create(
                    created_by=request.user,
                    periode_comptable=periode_comptable,
                    prestataire=request.user.prestataire,
                    type_remboursement=TypeRemboursement.objects.get(code="TP"),
                    bureau=request.user.bureau
                )


            # facture.net_a_payer
            code_bureau = request.user.bureau.code
            facture.numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(facture.pk).zfill(7) + "-" +endcode_facture
            facture.save()

            net_a_payer = 0
            part_assure = 0
            montant_reel = 0
            sinistes_pdf = []
            sinistres_par_adherent = defaultdict(list)

            # A COMPLETER AVEC LE CAS OU LA FACTURE EST GÉNÉRÉE PAR UN GESTIONNAIRE

            for sinistre in sinistres:
                sinistre.facture_prestataire = facture
                sinistre.statut_bordereau = StatutSinistreBordereau.AJOUTE_BORDEREAU
                sinistre.statut_remboursement = StatutRemboursement.DEMANDE
                sinistre.save()

                # if sinistre.type_prefinancement.code == "PREF_TOUT":
                if sinistre.tm_prefinanced:
                    part_assure = part_assure + 0
                    net_a_payer = net_a_payer + float(sinistre.total_frais_reel)

                else:
                    part_assure = part_assure + float(sinistre.total_part_assure)
                    net_a_payer = net_a_payer + float(sinistre.total_part_compagnie)

                # part_assure = part_assure + sinistre.total_part_assure
                montant_reel = montant_reel + float(sinistre.total_frais_reel)
                sinistes_pdf.append(sinistre)
                adherent_id = sinistre.adherent_principal.id
                sinistres_par_adherent[adherent_id].append(sinistre)

            facture.net_a_payer = net_a_payer

            net_a_payer_lettre = num2words(net_a_payer, lang="fr")

            pprint(net_a_payer_lettre)

            # Added on 06112023: add tm préfinancé
            request_user_is_prestataire = request.user.is_prestataire
            #
            currency_code = request.user.bureau.pays.devise.code

            if is_assure:
                sinistres_par_adherent2 = []
                nombre_sinistres = 0
                for adherent_id, sinistres in sinistres_par_adherent.items():
                    print(f"Adhérent {adherent_id} :")

                    net_a_payer2 = 0
                    part_assure2 = 0
                    montant_reel2 = 0

                    for sinistre in sinistres:
                        print(f"  - {sinistre}")
                        if sinistre.tm_prefinanced:
                            part_assure2 = part_assure2 + 0
                            net_a_payer2 = net_a_payer2 + float(sinistre.total_frais_reel)

                        else:
                            part_assure2 = part_assure2 + float(sinistre.total_part_assure)
                            net_a_payer2 = net_a_payer2 + float(sinistre.total_part_compagnie)

                        # part_assure = part_assure + sinistre.total_part_assure
                        montant_reel2 = montant_reel2 + float(sinistre.total_frais_reel)

                    sinistres_par_adherent2.append({
                        "adherent": sinistres[0].adherent_principal.nom_prenoms,
                        "sinistres": sinistres,
                        "nombre_sinistres": len(sinistres),
                        "sous_net_a_payer": net_a_payer2,
                    })
                    nombre_sinistres = nombre_sinistres + len(sinistres)


                # generer le pdf
                pdf = render_pdf("courriers/bordereau_paiement_assure.html",
                                 context_dict={'sinistres': sinistes_pdf,
                                               'nombre_sinistres': nombre_sinistres,
                                               'sinistres_par_adherent': sinistres_par_adherent2,
                                               'bordereau': facture,
                                               'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel,
                                               'net_a_payer_lettre': net_a_payer_lettre,
                                               'request_user_is_prestataire': request_user_is_prestataire,
                                               'currency_code': currency_code})
                pdf_file = PyPDF2.PdfReader(pdf)
                nombre_pages = len(pdf_file.pages)

                pdf = render_pdf("courriers/bordereau_paiement_assure.html",
                                 context_dict={'sinistres': sinistes_pdf,
                                               'nombre_sinistres': nombre_sinistres,
                                               'sinistres_par_adherent': sinistres_par_adherent2,
                                               'bordereau': facture,
                                               'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel,
                                               'net_a_payer_lettre': net_a_payer_lettre,
                                               'nombre_pages': nombre_pages,
                                               'request_user_is_prestataire': request_user_is_prestataire,
                                               'currency_code': currency_code})
            else:
                # generer le pdf
                pdf = render_pdf("courriers/bordereau_paiement.html",
                                 context_dict={'sinistres': sinistes_pdf, 'bordereau': facture, 'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel, 'net_a_payer_lettre':net_a_payer_lettre,
                                               'request_user_is_prestataire': request_user_is_prestataire, 'currency_code': currency_code})
                pdf_file = PyPDF2.PdfReader(pdf)
                nombre_pages = len(pdf_file.pages)

                pdf = render_pdf("courriers/bordereau_paiement.html",
                                 context_dict={'sinistres': sinistes_pdf, 'bordereau': facture, 'net_a_payer': net_a_payer,
                                               'part_assure': part_assure, 'montant_reel': montant_reel,  'net_a_payer_lettre':net_a_payer_lettre,
                                               'nombre_pages': nombre_pages,
                                               'request_user_is_prestataire': request_user_is_prestataire, 'currency_code': currency_code})


            # Enregistrement du bordereau
            facture.fichier.save(f'bordereau_paiement_{facture.numero}.pdf', File(pdf))
            facture.save()

            return JsonResponse(
                {'statut': 1, 'message': 'Bordereau généré avec succès', 'bordereau_pdf': facture.fichier.url,
                 'bordereau_id': facture.pk}, status=200)

        else:
            return JsonResponse({'statut': 0, 'message': 'Aucun sinistre sélectionné'})

# except Exception as e:
#    return JsonResponse({'error': str(e)}, status=500)


def borderau_validation_pdf_old(request, liste_sinistre, beneficiaire, par_compagnie=False):

    liste_compagnies_concernes = liste_sinistre.values('compagnie_id').annotate(
        nombre_sinistres=Count('compagnie_id'),
    )

    results = []

    for groupe in liste_compagnies_concernes:
        sinistres = liste_sinistre.filter(compagnie_id=groupe['compagnie_id'])
        compagnie = Compagnie.objects.filter(id=groupe['compagnie_id']).first()

        # Montant net a payer est maintenant une propriété montant_remb_accepte
        montant_remb_accepte_par_compagnie = sum(s.montant_remb_accepte for s in sinistres)

        total_part_assure = sum(s.total_part_assure for s in sinistres)
        total_part_compagnie = sum(s.total_part_compagnie for s in sinistres)
        total_part_beneficiare = sum(s.total_part_assure for s in sinistres)
        total_frais_reel = sum(s.total_frais_reel for s in sinistres)

        total_base_remboursement = sum(s.total_part_compagnie or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)

        result = {
            'beneficiaire': beneficiaire,
            'compagnie': compagnie.nom,
            'total_part_assure': total_part_assure,
            'total_part_compagnie': total_part_compagnie,
            'total_part_beneficiare': total_part_beneficiare,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': total_base_remboursement,
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_nombre_sinistres': len(sinistres),
            'montant_remb_accepte_par_compagnie': montant_remb_accepte_par_compagnie,
            'sinistres': sinistres,
        }

        results.append(result)

    total_global_nombre_sinistres = sum(r['total_nombre_sinistres'] for r in results)
    total_global_part_assure = sum(r['total_part_assure'] for r in results)
    total_global_part_compagnie = sum(r['total_part_compagnie'] for r in results)
    total_global_part_beneficiare = sum(r['total_part_assure'] for r in results)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in results)
    total_global_rejete = sum(r['total_rejete'] for r in results)
    total_global_base_remboursement = sum(r['total_base_remboursement'] for r in results)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in results)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in results)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in results)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(r['montant_remb_accepte_par_compagnie'] for r in results)

    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_validation.html', {'sinistres_groupes': results, 'beneficiaire': beneficiaire, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': results,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_part_compagnie': total_global_part_compagnie,
        'total_global_part_beneficiare': total_global_part_beneficiare,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'beneficiaire': beneficiaire,
        'currency_code': currency_code
    }
    pdf = render_pdf('courriers/borderau_validation.html', contexte)

    return pdf

    # #AFFICHER DIRECTEMENT
    # return HttpResponse(File(pdf), content_type='application/pdf')


def borderau_validation_pdf(request, liste_sinistre, beneficiaire, par_compagnie=False):
    liste_compagnies_concernes = liste_sinistre.values('compagnie_id').annotate(
        nombre_sinistres=Count('compagnie_id'),
    )

    results = []

    for groupe in liste_compagnies_concernes:
        sinistres = liste_sinistre.filter(compagnie_id=groupe['compagnie_id'])
        compagnie = Compagnie.objects.filter(id=groupe['compagnie_id']).first()

        # Montant net a payer est maintenant une propriété montant_remb_accepte
        montant_remb_accepte_par_compagnie = sum(s.montant_remb_accepte for s in sinistres)

        total_frais_reel = sum(s.total_frais_reel for s in sinistres)
        total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

        total_base_remboursement = sum((s.total_frais_reel if s.tm_prefinanced else s.total_part_compagnie) or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes

        result = {
            'beneficiaire': beneficiaire,
            'compagnie': compagnie.nom,
            'total_part_assure': total_part_assure,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': float(total_base_remboursement),
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_nombre_sinistres': len(sinistres),
            #'montant_remb_accepte_par_compagnie': montant_remb_accepte_par_compagnie,
            'total_net_a_payer': total_net_a_payer,
            'sinistres': sinistres,
        }

        results.append(result)

    total_global_nombre_sinistres = sum(r['total_nombre_sinistres'] for r in results)
    total_global_part_assure = sum(r['total_part_assure'] for r in results)
    #total_global_part_beneficiare = sum(r['total_part_assure'] for r in results)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in results)
    total_global_rejete = sum(r['total_rejete'] for r in results)
    total_global_base_remboursement = sum(r['total_base_remboursement'] for r in results)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in results)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in results)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in results)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(r['total_net_a_payer'] for r in results)
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_validation.html', {'sinistres_groupes': results, 'beneficiaire': beneficiaire, 'resultats': results, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': results,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        #'total_global_part_beneficiare': total_global_part_beneficiare,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'beneficiaire': beneficiaire,
        'currency_code': currency_code
    }
    pdf = render_pdf('courriers/borderau_validation.html', contexte)

    return pdf

    # #AFFICHER DIRECTEMENT
    # return HttpResponse(File(pdf), content_type='application/pdf')


def borderau_ordonnancement_pdf(request, liste_sinistre, prestataire, bordereau_ordonnancement):
    from itertools import groupby

    pprint("prestataire à payer")
    pprint(prestataire)

    # Regrouper les sinistres par compagnie
    sinistres_par_compagnie = {}
    for compagnie, sinistres_groupe in groupby(liste_sinistre, key=lambda x: x.compagnie):
        sinistres_par_compagnie[compagnie] = list(sinistres_groupe)

    # Calculer les totaux pour chaque compagnie
    resultats = []
    for compagnie, sinistres in sinistres_par_compagnie.items():
        total_nombre_sinistres = len(sinistres)

        total_frais_reel = sum(s.total_frais_reel for s in sinistres)
        total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

        total_base_remboursement = sum((s.total_frais_reel if s.tm_prefinanced else s.total_part_compagnie) or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes


        # Ajouter les résultats pour cette compagnie à la liste de résultats
        resultats.append({
            'prestataire': prestataire.name,
            'compagnie': compagnie.nom,
            'total_nombre_sinistres': total_nombre_sinistres,
            'total_part_assure': total_part_assure,
            #'total_part_beneficiare': total_part_assure,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': float(total_base_remboursement),
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_net_a_payer': total_net_a_payer,
            'sinistres': sinistres,
        })

    # Calcul des totaux globaux
    total_global_nombre_sinistres = sum(resultat['total_nombre_sinistres'] for resultat in resultats)
    total_global_part_assure = sum(resultat['total_part_assure'] for resultat in resultats)
    #total_global_part_compagnie = sum(resultat['total_part_compagnie'] for resultat in resultats)
    #total_global_part_beneficiare = sum(resultat['total_part_beneficiare'] for resultat in resultats)
    total_global_frais_reel = sum(resultat['total_frais_reel'] for resultat in resultats)
    total_global_rejete = sum(resultat['total_rejete'] for resultat in resultats)
    total_global_base_remboursement = sum(resultat['total_base_remboursement'] for resultat in resultats)
    total_global_base_taxable = sum(resultat['total_base_taxable'] for resultat in resultats)
    total_global_taxe_tbs = sum(resultat['total_taxe_tbs'] for resultat in resultats)
    total_global_taxe_far = sum(resultat['total_taxe_far'] for resultat in resultats)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(resultat['total_net_a_payer'] for resultat in resultats)
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_ordonnancement.html', {'sinistres_groupes': resultats, 'prestataire': prestataire, 'resultats': resultats, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': resultats,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        #'total_global_part_compagnie': total_global_part_compagnie,
        #'total_global_part_beneficiare': total_global_part_beneficiare,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'prestataire': prestataire,
        'currency_code': currency_code,
        'bordereau_ordonnancement': bordereau_ordonnancement
    }
    pdf = render_pdf('courriers/borderau_ordonnancement.html', contexte)

    return pdf



def borderau_ordonnancement_rd_pdf(request, liste_sinistre, adherent_principal, bordereau_ordonnancement):
    from itertools import groupby

    pprint("adherent_principal à payer")
    pprint(adherent_principal)

    # Regrouper les sinistres par compagnie
    sinistres_par_compagnie = {}
    for compagnie, sinistres_groupe in groupby(liste_sinistre, key=lambda x: x.compagnie):
        sinistres_par_compagnie[compagnie] = list(sinistres_groupe)

    # Calculer les totaux pour chaque compagnie
    resultats = []
    for compagnie, sinistres in sinistres_par_compagnie.items():
        total_nombre_sinistres = len(sinistres)

        total_frais_reel = sum(s.total_frais_reel for s in sinistres)
        total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

        total_base_remboursement = sum((s.total_frais_reel if s.tm_prefinanced else s.total_part_compagnie) or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes


        # Ajouter les résultats pour cette compagnie à la liste de résultats
        resultats.append({
            'adherent_principal': adherent_principal,
            'compagnie': compagnie.nom,
            'total_nombre_sinistres': total_nombre_sinistres,
            'total_part_assure': total_part_assure,
            #'total_part_compagnie': total_part_compagnie,
            #'total_part_beneficiare': total_part_assure,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': float(total_base_remboursement),
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_net_a_payer': total_net_a_payer,
            'sinistres': sinistres,
        })

    # Calcul des totaux globaux
    total_global_nombre_sinistres = sum(resultat['total_nombre_sinistres'] for resultat in resultats)
    total_global_part_assure = sum(resultat['total_part_assure'] for resultat in resultats)
    #total_global_part_compagnie = sum(resultat['total_part_compagnie'] for resultat in resultats)
    #total_global_part_beneficiare = sum(resultat['total_part_beneficiare'] for resultat in resultats)
    total_global_frais_reel = sum(resultat['total_frais_reel'] for resultat in resultats)
    total_global_rejete = sum(resultat['total_rejete'] for resultat in resultats)
    total_global_base_remboursement = sum(resultat['total_base_remboursement'] for resultat in resultats)
    total_global_base_taxable = sum(resultat['total_base_taxable'] for resultat in resultats)
    total_global_taxe_tbs = sum(resultat['total_taxe_tbs'] for resultat in resultats)
    total_global_taxe_far = sum(resultat['total_taxe_far'] for resultat in resultats)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(resultat['total_net_a_payer'] for resultat in resultats)
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_ordonnancement_rd.html', {'sinistres_groupes': resultats, 'adherent_principal': adherent_principal, 'resultats': resultats, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': resultats,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'adherent_principal': adherent_principal,
        'currency_code': currency_code,
        'bordereau_ordonnancement': bordereau_ordonnancement
    }
    pdf = render_pdf('courriers/borderau_ordonnancement_rd.html', contexte)

    return pdf

def borderau_ordonnancement_rd_assure_pdf(request, liste_sinistre, assure, bordereau_ordonnancement):
    from itertools import groupby

    pprint("assure à payer")
    pprint(assure)

    # Regrouper les sinistres par compagnie
    sinistres_par_adherent = defaultdict(list)
    for sinistre in liste_sinistre:
        adherent_id = sinistre.adherent_principal.id
        sinistres_par_adherent[adherent_id].append(sinistre)

    # for compagnie, sinistres_groupe in groupby(liste_sinistre, key=lambda x: x.compagnie):
    #     sinistres_par_compagnie[compagnie] = list(sinistres_groupe)

    # Calculer les totaux pour chaque compagnie
    resultats = []
    for adherent_id, sinistres in sinistres_par_adherent.items():
        total_nombre_sinistres = len(sinistres)

        total_frais_reel = sum(s.total_frais_reel for s in sinistres)
        total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

        total_base_remboursement = sum((s.total_frais_reel if s.tm_prefinanced else s.total_part_compagnie) or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes


        # Ajouter les résultats pour cette compagnie à la liste de résultats
        resultats.append({
            'assure': assure,
            "adherent": sinistres[0].adherent_principal.nom_prenoms,
            'total_nombre_sinistres': total_nombre_sinistres,
            'total_part_assure': total_part_assure,
            #'total_part_compagnie': total_part_compagnie,
            #'total_part_beneficiare': total_part_assure,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': float(total_base_remboursement),
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_net_a_payer': total_net_a_payer,
            'sinistres': sinistres,
        })

    # Calcul des totaux globaux
    total_global_nombre_sinistres = sum(resultat['total_nombre_sinistres'] for resultat in resultats)
    total_global_part_assure = sum(resultat['total_part_assure'] for resultat in resultats)
    #total_global_part_compagnie = sum(resultat['total_part_compagnie'] for resultat in resultats)
    #total_global_part_beneficiare = sum(resultat['total_part_beneficiare'] for resultat in resultats)
    total_global_frais_reel = sum(resultat['total_frais_reel'] for resultat in resultats)
    total_global_rejete = sum(resultat['total_rejete'] for resultat in resultats)
    total_global_base_remboursement = sum(resultat['total_base_remboursement'] for resultat in resultats)
    total_global_base_taxable = sum(resultat['total_base_taxable'] for resultat in resultats)
    total_global_taxe_tbs = sum(resultat['total_taxe_tbs'] for resultat in resultats)
    total_global_taxe_far = sum(resultat['total_taxe_far'] for resultat in resultats)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(resultat['total_net_a_payer'] for resultat in resultats)
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_ordonnancement_rd_assure.html', {'sinistres_groupes': resultats, 'assure': assure, 'resultats': resultats, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': resultats,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'assure': assure,
        'currency_code': currency_code,
        'bordereau_ordonnancement': bordereau_ordonnancement
    }
    pdf = render_pdf('courriers/borderau_ordonnancement_rd_assure.html', contexte)

    return pdf


def regenerer_borderau_ordonnancement_rd_pdf(request, bordereau_ordonnancement_id):
    from itertools import groupby

    pprint("Regénération du bordereau d'ordonnancement RD")

    bordereau_ordonnancement = BordereauOrdonnancement.objects.get(id=bordereau_ordonnancement_id)
    if bordereau_ordonnancement.assure:
        liste_sinistre = Sinistre.objects.filter(bordereau_ordonnancement_id=bordereau_ordonnancement.id,
                                                 facture_prestataire__assure__id=bordereau_ordonnancement.assure.id)

        pdf = borderau_ordonnancement_rd_assure_pdf(request, liste_sinistre, bordereau_ordonnancement.assure)

    else:
        adherent_principal = bordereau_ordonnancement.adherent_principal

        liste_sinistre = Sinistre.objects.filter(bordereau_ordonnancement=bordereau_ordonnancement)





        # Regrouper les sinistres par compagnie
        sinistres_par_compagnie = {}
        for compagnie, sinistres_groupe in groupby(liste_sinistre, key=lambda x: x.compagnie):
            sinistres_par_compagnie[compagnie] = list(sinistres_groupe)

        # Calculer les totaux pour chaque compagnie
        resultats = []
        for compagnie, sinistres in sinistres_par_compagnie.items():
            total_nombre_sinistres = len(sinistres)

            total_frais_reel = sum(s.total_frais_reel for s in sinistres)
            total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

            total_base_remboursement = sum((s.total_frais_reel if s.tm_prefinanced else s.total_part_compagnie) or 0 for s in sinistres)
            total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
            total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

            total_base_taxable = total_accepte

            # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
            total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
            total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
            total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
            total_net_a_payer = total_base_taxable + total_taxes


            # Ajouter les résultats pour cette compagnie à la liste de résultats
            resultats.append({
                'adherent_principal': adherent_principal,
                'compagnie': compagnie.nom,
                'total_nombre_sinistres': total_nombre_sinistres,
                'total_part_assure': total_part_assure,
                #'total_part_compagnie': total_part_compagnie,
                #'total_part_beneficiare': total_part_assure,
                'total_frais_reel': total_frais_reel,
                'total_rejete': total_rejete,
                'total_base_remboursement': float(total_base_remboursement),
                'total_base_taxable': total_base_taxable,
                'total_taxe_far': total_taxe_far,
                'total_taxe_tbs': total_taxe_tbs,
                'total_taxes': total_taxes,
                'total_net_a_payer': total_net_a_payer,
                'sinistres': sinistres,
            })

        # Calcul des totaux globaux
        total_global_nombre_sinistres = sum(resultat['total_nombre_sinistres'] for resultat in resultats)
        total_global_part_assure = sum(resultat['total_part_assure'] for resultat in resultats)
        #total_global_part_compagnie = sum(resultat['total_part_compagnie'] for resultat in resultats)
        #total_global_part_beneficiare = sum(resultat['total_part_beneficiare'] for resultat in resultats)
        total_global_frais_reel = sum(resultat['total_frais_reel'] for resultat in resultats)
        total_global_rejete = sum(resultat['total_rejete'] for resultat in resultats)
        total_global_base_remboursement = sum(resultat['total_base_remboursement'] for resultat in resultats)
        total_global_base_taxable = sum(resultat['total_base_taxable'] for resultat in resultats)
        total_global_taxe_tbs = sum(resultat['total_taxe_tbs'] for resultat in resultats)
        total_global_taxe_far = sum(resultat['total_taxe_far'] for resultat in resultats)
        total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
        total_global_net_a_payer = sum(resultat['total_net_a_payer'] for resultat in resultats)
        #
        currency_code = request.user.bureau.pays.devise.code

        pdf = render_pdf('courriers/borderau_ordonnancement_rd.html', {'sinistres_groupes': resultats, 'adherent_principal': adherent_principal, 'resultats': resultats, 'currency_code': currency_code})

        pdf_file = PyPDF2.PdfReader(pdf)
        nombre_pages = len(pdf_file.pages)

        contexte = {
            'resultats': resultats,
            'total_global_nombre_sinistres': total_global_nombre_sinistres,
            'total_global_part_assure': total_global_part_assure,
            'total_global_frais_reel': total_global_frais_reel,
            'total_global_rejete': total_global_rejete,
            'total_global_base_remboursement': total_global_base_remboursement,
            'total_global_base_taxable': total_global_base_taxable,
            'total_global_taxe_tbs': total_global_taxe_tbs,
            'total_global_taxes': total_global_taxes,
            'total_global_net_a_payer': total_global_net_a_payer,
            'nombre_pages': nombre_pages,
            'adherent_principal': adherent_principal,
            'currency_code': currency_code,
            'bordereau_ordonnancement': bordereau_ordonnancement,
        }
        pdf = render_pdf('courriers/borderau_ordonnancement_rd.html', contexte)

        # Update bordereau data and save
        bordereau_ordonnancement.fichier.save(f'bordereau_ordonnancement_{bordereau_ordonnancement.numero}.pdf', File(pdf))
        bordereau_ordonnancement.montant_remb_total = total_global_frais_reel
        bordereau_ordonnancement.montant_rejet_total = total_global_rejete
        bordereau_ordonnancement.montant_accepte_total = total_global_net_a_payer
        bordereau_ordonnancement.save()


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')


def regenerer_borderau_ordonnancement_pdf(request, bordereau_ordonnancement_id):

    bordereau_ordonnancement = BordereauOrdonnancement.objects.get(id=bordereau_ordonnancement_id)

    prestataire = bordereau_ordonnancement.prestataire

    liste_sinistre = Sinistre.objects.filter(bordereau_ordonnancement=bordereau_ordonnancement)

    liste_compagnies_concernes = liste_sinistre.values('compagnie_id').annotate(
        nombre_sinistres=Count('compagnie_id'),
    )

    results = []

    for groupe in liste_compagnies_concernes:
        sinistres = liste_sinistre.filter(compagnie_id=groupe['compagnie_id'])
        compagnie = Compagnie.objects.filter(id=groupe['compagnie_id']).first()


        total_frais_reel = sum(s.total_frais_reel for s in sinistres)
        total_part_assure = sum((0 if s.tm_prefinanced else s.total_part_assure) for s in sinistres) #identique a total_part_beneficiare

        total_base_remboursement = sum((Decimal(s.total_frais_reel) if s.tm_prefinanced else Decimal(s.total_part_compagnie)) or 0 for s in sinistres)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres)
        total_accepte = sum(s.montant_remb_accepte or 0 for s in sinistres)

        total_base_taxable = total_accepte

        # total_base_taxable = sum(s.base_taxable or 0 for s in sinistres)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres)
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes


        result = {
            'prestataire': prestataire.name,
            'compagnie': compagnie.nom,
            'total_part_assure': total_part_assure,
            'total_frais_reel': total_frais_reel,
            'total_rejete': total_rejete,
            'total_base_remboursement': float(total_base_remboursement),
            'total_base_taxable': total_base_taxable,
            'total_taxe_far': total_taxe_far,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxes': total_taxes,
            'total_nombre_sinistres': len(sinistres),
            'total_net_a_payer': total_net_a_payer,
            'sinistres': sinistres,
        }

        results.append(result)

        #dd(results)

    total_global_nombre_sinistres = sum(r['total_nombre_sinistres'] for r in results)
    total_global_part_assure = sum(r['total_part_assure'] for r in results)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in results)
    total_global_rejete = sum(r['total_rejete'] for r in results)
    total_global_base_remboursement = sum(r['total_base_remboursement'] for r in results)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in results)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in results)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in results)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(r['total_net_a_payer'] for r in results)
    #
    currency_code = request.user.bureau.pays.devise.code

    pdf = render_pdf('courriers/borderau_ordonnancement.html', {'sinistres_groupes': results, 'prestataire': prestataire, 'resultats': results, 'currency_code': currency_code})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'resultats': results,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
        'nombre_pages': nombre_pages,
        'prestataire': prestataire,
        'bordereau_ordonnancement': bordereau_ordonnancement,
        'currency_code': currency_code
    }
    pdf = render_pdf('courriers/borderau_ordonnancement.html', contexte)

    # Update bordereau data and save
    bordereau_ordonnancement.fichier.save(f'bordereau_ordonnancement_{bordereau_ordonnancement.numero}.pdf', File(pdf))
    bordereau_ordonnancement.montant_remb_total = total_global_frais_reel
    bordereau_ordonnancement.montant_rejet_total = total_global_rejete
    bordereau_ordonnancement.montant_accepte_total = total_global_net_a_payer
    bordereau_ordonnancement.save()

    #dd(bordereau_ordonnancement)


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')




def borderau_ordonnancement_pdf_exemple(request):
    data = range(25)

    sinistres = Sinistre.objects.filter(statut_validite="VALIDE").order_by('prestataire')

    pdf = render_pdf('courriers/borderau_ordonnancement.html', {'sinistres': data})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    pdf = render_pdf('courriers/borderau_ordonnancement.html', {'sinistres': data, 'nombre_pages': nombre_pages})

    return HttpResponse(File(pdf), content_type='application/pdf')


# TODO : BORDEREAU DE VALIDATION
@method_decorator(login_required, name='dispatch')
class GenerationBrValidationView(TemplateView):
    # permission_required = "sinistre.view_sinistre"
    template_name = 'generation_br_validation.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        bordereau = FacturePrestataire.objects.all()

        # liste_sinistres_bordereau = [x for x in Sinistre.objects.filter(facture_prestataire__isnull=False) if x.is_processed]
        # prestataires = [x.prestataire for x in
        #                 FacturePrestataire.objects.filter(prestataire__bureau=request.user.bureau,
        #                     statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE])]

        # base des sinistres concernés pour en sortir les id de facture possible pour degager les autres entites de facon exacte
        base_sinistres = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        facture_prestataire__statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE],
        is_ges_processed=True).order_by('-id')
        factures_prestataires_ids =  base_sinistres.values_list('facture_prestataire_id', flat=True).distinct(),

        # Load distinct prestataires
        prestataires_ids = FacturePrestataire.objects.filter(id__in=factures_prestataires_ids, bureau=request.user.bureau,statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE]).values_list('prestataire_id', flat=True).distinct()
        prestataires = Prestataire.objects.filter(id__in=prestataires_ids)

        assures_ids = FacturePrestataire.objects.filter(id__in=factures_prestataires_ids, assure_id__isnull=False, bureau=request.user.bureau,statut__in=[SatutBordereauDossierSinistres.ATTENTE,SatutBordereauDossierSinistres.REJETE]).values_list('assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()

        type_remboursements = TypeRemboursement.objects.filter(status=True)#.order_by('libelle')

        # adherent_principal_ids = Sinistre.objects.filter(facture_prestataire__statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE], is_ges_processed=True).values_list('adherent_principal_id', flat=True).order_by('-id').distinct()
        # adhs = Aliment.objects.filter(id__in=adherent_principal_ids)

        adherent_principal_ids = FacturePrestataire.objects.filter(id__in=factures_prestataires_ids, adherent_principal_id__isnull=False, bureau=request.user.bureau, statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE]).values_list('adherent_principal_id', flat=True).order_by('-id').distinct()
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids)

        context = self.get_context_data(**kwargs)
        context['bordereau'] = bordereau
        # context['liste_sinistres_bordereau'] = liste_sinistres_bordereau
        context['montant_remb_total'] = 0
        context['montant_rejet_total'] = 0
        context['montant_accepte_total'] = 0
        context['is_all_processed'] = False
        context['prestataires'] = prestataires
        context['assures'] = assures
        context['periodes_comptables'] = periodes_comptables
        context['type_remboursements'] = type_remboursements
        context['adherent_principaux'] = adhs
        # context['is_all_processed_validated'] = True if bordereau.statut == SatutBordereauDossierSinistres.VALIDE else False

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def generation_br_validation_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    # start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    # sort_column_index = int(request.GET.get('order[0][column]'))
    # sort_direction = request.GET.get('order[0][dir]')
    search = request.GET.get('search[value]', '')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_reference_facture = request.GET.get('search_reference_facture', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')


    queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        facture_prestataire__statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE],
        is_ges_processed=True).order_by('-id')

    if search_numero_bordereau:
        queryset = queryset.filter(dossier_sinistre__numero__contains=search_numero_bordereau)

    if search_reference_facture:
        queryset = queryset.filter(reference_facture__contains=search_reference_facture)

    if search_type_remboursement:
        type_remboursement = TypeRemboursement.objects.get(id=search_type_remboursement)
        queryset = queryset.filter(dossier_sinistre__type_remboursement=type_remboursement)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=int(search_prestataire))

    if search_assure:
        queryset = queryset.filter(facture_prestataire__assure__id=search_assure)

    if search_adherent_principal:
            queryset = queryset.filter(adherent_principal__id=search_adherent_principal)

    if search_periode_comptable:
        queryset = queryset.filter(facture_prestataire__periode_comptable__id=int(search_periode_comptable))


    if search:
        queryset = queryset.filter(
            Q(numero__icontains=search) |
            Q(dossier_sinistre__numero__icontains=search) |
            Q(prestataire__name__icontains=search) |
            Q(aliment__nom__icontains=search) |
            Q(aliment__prenoms__icontains=search) |
            Q(aliment__cartes__numero__icontains=search) |
            Q(acte__libelle__icontains=search)
        )

    # collecte des montants totaux
    montant_remb_total = 0
    montant_rejet_total = 0
    montant_accepte_total = 0
    for x in queryset:
        total_base_remb = x.total_frais_reel if x.tm_prefinanced else x.total_part_compagnie
        total_rejet = x.montant_remb_refuse if x.montant_remb_refuse else 0
        total_net_payer = x.montant_remb_accepte if x.montant_remb_accepte else 0

        # collecte des montants totaux
        montant_remb_total = montant_remb_total + float(total_base_remb)
        montant_rejet_total = montant_rejet_total + float(total_rejet)
        montant_accepte_total = montant_accepte_total + float(total_net_payer)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        if c.is_ges_processed:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-success" style="font-size:10px;color:white;">TRAITÉ</span>'
        else:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-en-attente" style="font-size:10px;color:white;">NON TRAITÉ</span>'

        total_facture = c.total_frais_reel if c.total_frais_reel else 0
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_rejet = c.montant_remb_refuse if c.montant_remb_refuse else 0
        total_net_payer = c.montant_remb_accepte if c.montant_remb_accepte else 0

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        # checkbox = f'<input type="checkbox" name="sinistres_ids[]" value="{c.id}" data-sinistre_id="{c.id}" class ="select-row sinistres">'

        data_iten = {
            "id": c.id,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "numero": c.numero,
            "dossier_sinistre__numero": c.dossier_sinistre.numero,
            "reference_facture": c.reference_facture,
            "prestataire": c.prestataire.name,
            "beneficiaire": c.aliment.nom + ' ' + c.aliment.prenoms,
            "carte_active": numero_carte,
            "acte": c.acte.libelle,
            "total_facture": money_field(total_facture),
            "total_part_assure": money_field(total_part_assure),
            "total_base_remb": money_field(total_base_remb),
            "total_rejet": money_field(total_rejet),
            "total_net_payer": money_field(total_net_payer),
            "total_part_compagnie": money_field(total_base_remb),
            "statut": statut_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
        "montant_remb_total": money_field(montant_remb_total),
        "montant_rejet_total": money_field(montant_rejet_total),
        "montant_accepte_total": money_field(montant_accepte_total),
    })


@transaction.atomic
def submit_generation_br_validation(request):
    with transaction.atomic():
        # try:
        # RECUPERATION DES DONNEES DU FILTRE
        print(request.POST)
        print(request.POST.get('search_periode_comptable', 0))
        search_numero_bordereau = request.POST.get('search_numero_bordereau')
        search_reference_facture = request.POST.get('search_reference_facture')
        search_periode_comptable = request.POST.get('search_periode_comptable')
        search_type_remboursement = request.POST.get('search_type_remboursement')
        search_adherent_principal = request.POST.get('search_adherent_principal')
        search_assure = request.POST.get('search_assure')
        search_prestataire = request.POST.get('search_prestataire')
        montant_remb_total = float(request.POST.get('montant_remb_total', 0).replace(" ", ""))
        montant_rejet_total = float(request.POST.get('montant_rejet_total', 0).replace(" ", ""))
        montant_accepte_total = float(request.POST.get('montant_accepte_total', 0).replace(" ", ""))
        par_compagnie = request.POST.get('par_compagnie', '')
        par_compagnie = True if par_compagnie == 'true' else False

        beneficiaire_a_afficher_sur_le_brouillard_de_saisie = ""

        print("search prestataire " +  search_prestataire)
        print("search_type_remboursement " +  str(search_type_remboursement))
        print("search_adherent_principal " +  str(search_adherent_principal))
        print("search_assure " +  str(search_assure))

        prestataire = None
        aliment = None

        if search_adherent_principal:
            aliment = Aliment.objects.get(id=int(search_adherent_principal))
            beneficiaire_a_afficher_sur_le_brouillard_de_saisie = aliment.adherent_principal
        elif search_prestataire:
            prestataire = Prestataire.objects.get(id=int(search_prestataire))
            beneficiaire_a_afficher_sur_le_brouillard_de_saisie = prestataire.name
        elif search_assure:
            assure = Client.objects.get(id=int(search_assure))
            beneficiaire_a_afficher_sur_le_brouillard_de_saisie = assure.nom + ' ' + (assure.prenoms if assure.prenoms else '')

        print(search_numero_bordereau)
        print(search_reference_facture)
        print(search_periode_comptable)
        print(search_prestataire)
        print(montant_remb_total)
        print(montant_rejet_total)
        print(montant_accepte_total)
        print(par_compagnie)

        # periode_comptable = PeriodeComptable.objects.get(id=int(search_periode_comptable)) if PeriodeComptable.objects.get(id=int(search_periode_comptable)) else None

        if search_periode_comptable:
            periode_comptable = PeriodeComptable.objects.get(id=int(search_periode_comptable))
        else:
            return JsonResponse({'statut': 0, 'message': 'Aucune donnée trouvé'})

        # if search_prestataire:
        #     prestataires = Prestataire.objects.filter(id=int(search_prestataire))
        #     prestataire = prestataires.first() if prestataires else None
        #     print(prestataire)
        # else:
        #     prestataire = None

        print(periode_comptable)

        # FIN RECUPERATION DES DONNEES DU FILTRE

        # RECUPERATION DES SINISTRES TRAITES A PARTI DU FILTRE
        # queryset = [x for x in Sinistre.objects.filter(
        #     facture_prestataire__statut=SatutBordereauDossierSinistres.VALIDE, bordereau_validation__isnull=True) if x.is_processed]

        queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
            facture_prestataire__statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE],
            is_ges_processed=True,
        ) #.annotate(num_remboursements=Count('remboursements')).filter(num_remboursements__gt=0)

        if search_numero_bordereau:
            queryset = queryset.filter(
                dossier_sinistre__numero__contains=search_numero_bordereau,
            )

        if search_reference_facture:
            queryset = queryset.filter(
                reference_facture__contains=search_reference_facture,
            )

        if search_type_remboursement:
            type_remboursement = TypeRemboursement.objects.get(id=search_type_remboursement)
            queryset = queryset.filter(dossier_sinistre__type_remboursement=type_remboursement)

        if search_periode_comptable:
            queryset = queryset.filter(facture_prestataire__periode_comptable__id=int(search_periode_comptable))

        if search_prestataire:
            queryset = queryset.filter(prestataire__id=int(search_prestataire))

        if search_adherent_principal:
            queryset = queryset.filter(adherent_principal__id=int(search_adherent_principal))
        # FIN RECUPERATION DES SINISTRES TRAITES A PARTI DU FILTRE
        print(queryset.count())
        if queryset.count() > 0:
            # GENERATION DU BORDEREAU DE VALIDATION (FICHIER PDF)
            # generer le pdf
            pdf = borderau_validation_pdf(request, liste_sinistre=queryset, beneficiaire=beneficiaire_a_afficher_sur_le_brouillard_de_saisie, par_compagnie=par_compagnie)

            # [REMOVE] WAS JUST FOR TEST
            # return HttpResponse(File(pdf), content_type='application/pdf')

            # FIN GENERATION DU BORDEREAU DE VALIDATION (FICHIER PDF)

            file = File(pdf, f'bordereau_validation_remboursement_.pdf')

            # upload file on server and return its url
            file_url = uploaded_file_url(file)

            return JsonResponse(
                {'statut': 1, 'message': 'Bordereau généré avec succès',
                 'bordereau_pdf': file_url}, status=200)

        else:
            return JsonResponse({'statut': 0, 'message': 'Aucune donnée trouvé'})

        # except Exception as e:
        # return JsonResponse({'statut': 0, 'message': 'Erreur lors de la génération du bordereau de validation', 'error': str(e)})


def uploaded_file_url(file):
    # Simulation de l'upload du fichier sur le serveur
    fs = FileSystemStorage()
    file_path = fs.save(file.name, file)
    return fs.url(file_path)


# TODO : BORDEREAU DE ORDONNANCEMENT
@method_decorator(login_required, name='dispatch')
class GenerationBrOrdonnancementView(TemplateView):
    # permission_required = "sinistre.view_sinistre"
    template_name = 'generation_br_ordonnancement.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        bordereau = FacturePrestataire.objects.filter(bureau=request.user.bureau)

        # liste_sinistres_bordereau = [x for x in Sinistre.objects.filter(facture_prestataire__isnull=False) if x.is_processed]
        # prestataires = [x.prestataire for x in
        #                 FacturePrestataire.objects.filter(bureau=request.user.bureau,
        #                     statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE])]
        # # Load distinct prestataires
        # prestataires = reduce(lambda re, x: re + [x] if x not in re else re, prestataires, [])

        prestataires_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau,
                                                             statut=SatutBordereauDossierSinistres.VALIDE).values_list(
            'prestataire_id', flat=True).distinct()
        prestataires = Prestataire.objects.filter(id__in=prestataires_ids)


        assures_ids = FacturePrestataire.objects.filter(assure_id__isnull=False, bureau=request.user.bureau,
                                                        statut=SatutBordereauDossierSinistres.VALIDE).values_list(
            'assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        # liste_sinistres_bordereau = [x for x in Sinistre.objects.filter(facture_prestataire__isnull=False) if x.is_processed]
        # adhs = [x.adherent_principal for x in
        #                 FacturePrestataire.objects.filter(bureau=request.user.bureau,
        #                     statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE])]

        adherent_principal_ids = FacturePrestataire.objects.filter(bureau=request.user.bureau,
                                                  statut=SatutBordereauDossierSinistres.VALIDE).values_list('adherent_principal_id', flat=True).order_by('-id').distinct()
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids)



        # aliments = Aliment.objects.filter(id__in=aliment_ids)
        # Load distinct adherent_principal
        adhs = reduce(lambda re, x: re + [x] if x not in re else re, adhs, [])

        periodes_comptables = [] # PeriodeComptable.objects.all()

        type_remboursements = TypeRemboursement.objects.filter()

        context = self.get_context_data(**kwargs)
        context['bordereau'] = bordereau
        # context['liste_sinistres_bordereau'] = liste_sinistres_bordereau
        context['montant_remb_total'] = 0
        context['montant_rejet_total'] = 0
        context['montant_accepte_total'] = 0
        context['is_all_processed'] = False
        context['prestataires'] = prestataires
        context['assures'] = assures
        context['periodes_comptables'] = periodes_comptables
        context['type_remboursements'] = type_remboursements
        context['adhs'] = adhs
        # context['is_all_processed_validated'] = True if bordereau.statut == SatutBordereauDossierSinistres.VALIDE else False

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


#new code add
def export_sinistres_ordonnancement(request):
    try:
        # Retrieve the Sinistre instances based on filters and user's bureau
        queryset = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            bordereau_ordonnancement__isnull=True,
            facture_prestataire__statut__in=[SatutBordereauDossierSinistres.VALIDE, SatutBordereauDossierSinistres.REJETE],
            is_ges_processed=True
        ).order_by('-id')

        # Create an Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="LISTE_SINISTRES_BENEFICIAIRE.xlsx"'

        # Initialize the workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'SINISTRES'

        # Write header row
        header = [
        'DATE SOIN',
        'N° SINISTRE',
        'N° FEUILLE DE SOINS',
        'PRESTATAIRE',
        'MATRICULE',
        'NOM BÉNÉFICIAIRE',
        'PRENOMS BÉNÉFICIAIRE',
        'ACTE',
        'MONTANT FACTURÉ',
        'PART ASSURÉ',
        'BASE REMB',
        'REJET',
        'NET À PAYER',
        'ÉTAT'

        ]
        for col_num, column_title in enumerate(header, 1):
            worksheet.cell(row=1, column=col_num).value = column_title

        # Write data rows
        for row_num, sinistre in enumerate(queryset, 2):
            # Retrieve Aliment (beneficiary) data through the foreign key
            aliment = sinistre.aliment

            row = [
                sinistre.date_survenance.strftime("%d/%m/%Y %H:%M") if sinistre.date_survenance else "",
                sinistre.numero if sinistre.numero else "",
                sinistre.dossier_sinistre.numero if sinistre.dossier_sinistre else "",
                sinistre.prestataire.name if sinistre.prestataire and sinistre.prestataire.name else "",
                aliment.veos_numero_carte if aliment and aliment.veos_numero_carte else "",  # Matricule CIE from Aliment
                aliment.nom if aliment and aliment.nom else "",  # Nom from Aliment
                aliment.prenoms if aliment and aliment.prenoms else "",  # Nom from Aliment
                sinistre.acte.libelle if sinistre.acte and sinistre.acte.libelle else "",
                sinistre.total_frais_reel,
                sinistre.total_part_assure if sinistre.total_part_assure else "",
                sinistre.total_part_compagnie if sinistre.total_part_compagnie else "",
                sinistre.montant_remb_accepte if sinistre.montant_remb_accepte else "",
                sinistre.total_part_compagnie if sinistre.total_part_compagnie else "",
                sinistre.statut if sinistre.statut else ""


            ]

            # Write each cell value to the worksheet
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value

        # Save the workbook to the response
        workbook.save(response)
        return response

    except Aliment.DoesNotExist:
        return JsonResponse({"message": "Bénéficiaire non trouvé"}, status=404)

#######################################################

def generation_br_ordonnancement_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    # start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    # sort_column_index = int(request.GET.get('order[0][column]'))
    # sort_direction = request.GET.get('order[0][dir]')
    search = request.GET.get('search[value]', '')
    pprint(search)

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_facture = request.GET.get('search_facture', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')

    pprint("search_type_remboursement")
    pprint(search_type_remboursement)

    pprint("search_prestataire")
    pprint(search_prestataire)

    pprint("search_adherent_principal")
    pprint(search_adherent_principal)

    pprint("search_assure")
    pprint(search_assure)

    pprint("search_periode_comptable")
    pprint(search_periode_comptable)

    pprint("search_facture")
    pprint(search_facture)

    # select_all = request.GET.get('all', '')
    ordre_de = ""

    # initial queryset to nothing data
    queryset = Sinistre.objects.none()

    pprint("queryset")
    pprint(queryset)

    # if search_facture:
    #     queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
    #                                    facture_prestataire_id=int(search_facture),
    #                                    bordereau_ordonnancement__isnull=True,
    #                                    facture_prestataire__statut=SatutBordereauDossierSinistres.VALIDE,
    #                                    is_ges_processed=True, ).order_by('-id')
    # print(queryset)
    # if search_numero_bordereau:
    #     queryset = queryset.filter(dossier_sinistre__numero__contains=search_numero_bordereau)
    #

    if search_periode_comptable:
        queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
                                           facture_prestataire__periode_comptable__id=int(search_periode_comptable),
                                           bordereau_ordonnancement__isnull=True,
                                           facture_prestataire__statut=SatutBordereauDossierSinistres.VALIDE,
                                           is_ges_processed=True, ).order_by('-id')

    if search_facture and queryset:
        queryset = queryset.filter(facture_prestataire_id=int(search_facture))

    if search_type_remboursement and queryset:
        queryset = queryset.filter(facture_prestataire__type_remboursement__id=int(search_type_remboursement))

    if search_adherent_principal and queryset:
        adherent_principal = Aliment.objects.get(id=search_adherent_principal)
        ordre_de = adherent_principal.nom + " " + adherent_principal.prenoms if adherent_principal else ""
        queryset = queryset.filter(adherent_principal__id=int(search_adherent_principal))

    if search_assure and queryset:
        assure = Client.objects.get(id=int(search_assure))
        ordre_de = assure.nom + " " + (assure.prenoms if assure.prenoms else "")
        queryset = queryset.filter(facture_prestataire__assure_id=int(search_assure))

    if search_prestataire and queryset:
        prestataire = Prestataire.objects.get(id=int(search_prestataire))
        ordre_de = prestataire.rb_ordre if prestataire else ""
        queryset = queryset.filter(prestataire__id=int(search_prestataire))
        
        print(queryset)
        print(search_prestataire)

    if search and queryset:
        queryset = queryset.filter(
            Q(numero__icontains=search) |
            Q(dossier_sinistre__numero__icontains=search) |
            Q(prestataire__name__icontains=search) |
            Q(aliment__nom__icontains=search) |
            Q(aliment__prenoms__icontains=search) |
            Q(aliment__cartes__numero__icontains=search) |
            Q(acte__libelle__icontains=search)
        )

    # if select_all:
    #     data_ids = [x.id for x in queryset]
    #     return JsonResponse({
    #         "data": data_ids,
    #     })
    # collecte des montants totaux
    montant_remb_total = 0
    montant_rejet_total = 0
    montant_accepte_total = 0
    ordre_de = ordre_de if queryset else ""
    for x in queryset:
        total_base_remb = x.total_frais_reel if x.tm_prefinanced else x.total_part_compagnie
        total_rejet = x.montant_remb_refuse if x.montant_remb_refuse else 0
        total_net_payer = x.montant_remb_accepte if x.montant_remb_accepte else 0

        # collecte des montants totaux
        montant_remb_total = montant_remb_total + float(total_base_remb)
        montant_rejet_total = montant_rejet_total + float(total_rejet)
        montant_accepte_total = montant_accepte_total + float(total_net_payer)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        if c.is_ges_processed:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-success" style="font-size:10px;color:white;">TRAITÉ</span>'
        else:
            statut_html = f'<span class="text-bold col-sm-12 badge badge-en-attente" style="font-size:10px;color:white;">NON TRAITÉ</span>'

        rejet_popup_url = reverse('popup_rejet_ordonnancement_sinistre', args=[c.id])
        action_html = f'<a href="javascript:0;"><span class="badge btn-sm btn-danger rounded-pill btn-open_modal_rejeter_remboursement_ordonnancement" data-href ="{rejet_popup_url}" data-toggle="modal" data-target="#modal_rejeter_remboursement_ordonnancement"><i class="fa fa-times"></i> Réfuser</span></a>'

        total_facture = c.total_frais_reel if c.total_frais_reel else 0
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_rejet = c.montant_remb_refuse if c.montant_remb_refuse else 0
        total_net_payer = c.montant_remb_accepte if c.montant_remb_accepte else 0

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        # checkbox = f'<input type="checkbox" name="selected_items[]" value="{c.id}">'

        data_iten = {
            "id": c.id,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "numero": c.numero,
            "dossier_sinistre__numero": c.dossier_sinistre.numero,
            "prestataire": c.prestataire.name,
            "beneficiaire": c.aliment.nom + ' ' + c.aliment.prenoms,
            "carte_active": numero_carte,
            "acte": c.acte.libelle,
            "total_facture": money_field(total_facture),
            "total_part_assure": money_field(total_part_assure),
            "total_base_remb": money_field(total_base_remb),
            "total_rejet": money_field(total_rejet),
            "total_net_payer": money_field(total_net_payer),
            "statut": statut_html,
            "action": action_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
        "montant_remb_total": money_field(montant_remb_total),
        "montant_rejet_total": money_field(montant_rejet_total),
        "montant_accepte_total": money_field(montant_accepte_total),
        "ordre_de": ordre_de,
    })


def get_facture_br_ordonnancement(request):
    if request.method == 'GET':
        prestataire_id = request.GET.get('prestataire_id')
        assure_id = request.GET.get('assure_id')
        adherent_principal_id = request.GET.get('adherent_principal_id')
        periode_comptable_id = request.GET.get('periode_comptable_id')

        factures = []
        data_response = []
        if assure_id:
            factures = FacturePrestataire.objects.filter(assure_id=assure_id,
                                                         bureau=request.user.bureau,
                                                         periode_comptable_id=periode_comptable_id,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).order_by('-id')
            print("factures assure")
        elif adherent_principal_id:
            factures = FacturePrestataire.objects.filter(adherent_principal_id=adherent_principal_id,
                                                         bureau=request.user.bureau,
                                                         periode_comptable_id=periode_comptable_id,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).order_by('-id')
            print("factures adherent_principal")
        elif prestataire_id:
            factures = FacturePrestataire.objects.filter(prestataire_id=prestataire_id,
                                                         bureau=request.user.bureau,
                                                         periode_comptable_id=periode_comptable_id,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).order_by('-id')
            print("factures prestataire")

        for facture in factures:
            data_response.append({'id': facture.id, 'numero': facture.numero, 'periode_comptable': facture.periode_comptable.libelle})

        return JsonResponse(data_response, safe=False)

def get_periode_br_ordonnancement(request):
    if request.method == 'GET':
        prestataire_id = request.GET.get('prestataire_id')
        assure_id = request.GET.get('assure_id')
        adherent_principal_id = request.GET.get('adherent_principal_id')

        periode_comptable_ids = []
        data_response = []
        if assure_id:
            periode_comptable_ids = FacturePrestataire.objects.filter(assure_id=assure_id,
                                                         bureau=request.user.bureau,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).values_list(
            'periode_comptable_id', flat=True).distinct()
            print("factures assure")
        elif adherent_principal_id:
            periode_comptable_ids = FacturePrestataire.objects.filter(adherent_principal_id=adherent_principal_id,
                                                         bureau=request.user.bureau,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).values_list(
            'periode_comptable_id', flat=True).distinct()
            print("factures adherent_principal")
        elif prestataire_id:
            periode_comptable_ids = FacturePrestataire.objects.filter(prestataire_id=prestataire_id,
                                                         bureau=request.user.bureau,
                                                         statut=SatutBordereauDossierSinistres.VALIDE).values_list(
            'periode_comptable_id', flat=True).distinct()
            print("factures prestataire")

        periodes_comptables = PeriodeComptable.objects.filter(id__in=periode_comptable_ids).order_by('id')

        for periode_comptable in periodes_comptables:
            data_response.append({'id': periode_comptable.id, 'libelle': periode_comptable.libelle})

        return JsonResponse(data_response, safe=False)


def popup_rejet_ordonnancement_sinistre(request, sinistre_id):
    sinistre = Sinistre.objects.get(id=sinistre_id)

    return render(request, 'modal_rejet_ordonnancement_sinistre.html',
                  {'sinistre': sinistre})


def refuser_remboursement_ordonnancement(request, sinistre_id):
    if request.method == 'POST':

        pprint(request.POST)

        motif = request.POST.get('motif')

        sinistre = Sinistre.objects.get(id=sinistre_id)
        facture_sinistre = sinistre.facture_prestataire
        sinistre_rembs = sinistre.remboursements.filter(is_invalid=False)

        for remb_sinistre in sinistre_rembs:
            remb_sinistre.is_invalid = True
            remb_sinistre.is_invalid_by = request.user
            remb_sinistre.save()

        sinistre.motif_rejet_ordonnancement = motif
        sinistre.is_ges_processed = False
        sinistre.save()

        facture_sinistre.statut = SatutBordereauDossierSinistres.REJETE
        facture_sinistre.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement rejetée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


def annuler_remboursement_ordonnancement(request, sinistre_id):
    if request.method == 'POST':

        pprint(request.POST)

        motif = request.POST.get('motif')

        sinistre = Sinistre.objects.get(id=sinistre_id)
        facture_sinistre = sinistre.facture_prestataire
        sinistre_rembs = sinistre.remboursements.filter(is_invalid=False)

        for remb_sinistre in sinistre_rembs:
            remb_sinistre.is_invalid = True
            remb_sinistre.is_invalid_by = request.user
            remb_sinistre.save()

        sinistre.motif_rejet_ordonnancement = motif
        sinistre.is_ges_processed = False
        sinistre.facture_prestataire = None
        sinistre.statut_remboursement = StatutRemboursement.ANNULE
        sinistre.statut_validite = StatutValidite.SUPPRIME
        sinistre.save()

        sinistres = Sinistre.objects.filter(facture_prestataire=facture_sinistre)

        net_a_payer = 0
        part_assure = 0
        montant_reel = 0
        sinistes_pdf = []

        for sin in sinistres:

            # if sinistre.type_prefinancement.code == "PREF_TOUT":
            if sin.tm_prefinanced:
                part_assure = part_assure + 0
                net_a_payer = net_a_payer + float(sin.total_frais_reel)

            else:
                part_assure = part_assure + float(sin.total_part_assure)
                net_a_payer = net_a_payer + float(sin.total_part_compagnie)

            # part_assure = part_assure + sinistre.total_part_assure
            montant_reel = montant_reel + float(sin.total_frais_reel)
            sinistes_pdf.append(sin)

        facture_sinistre.net_a_payer = net_a_payer

        # Added on 06112023: add tm préfinancé
        request_user_is_prestataire = request.user.is_prestataire

        #
        currency_code = request.user.bureau.pays.devise.code

        # generer le pdf
        pdf = render_pdf("courriers/bordereau_paiement.html",
                         context_dict={'sinistres': sinistes_pdf, 'bordereau': facture_sinistre,
                                       'net_a_payer': net_a_payer,
                                       'part_assure': part_assure, 'montant_reel': montant_reel,
                                       'request_user_is_prestataire': request_user_is_prestataire, 'currency_code': currency_code})
        pdf_file = PyPDF2.PdfReader(pdf)
        nombre_pages = len(pdf_file.pages)

        pdf = render_pdf("courriers/bordereau_paiement.html",
                         context_dict={'sinistres': sinistes_pdf, 'bordereau': facture_sinistre,
                                       'net_a_payer': net_a_payer,
                                       'part_assure': part_assure, 'montant_reel': montant_reel,
                                       'nombre_pages': nombre_pages,
                                       'request_user_is_prestataire': request_user_is_prestataire, 'currency_code': currency_code})

        # Enregistrement du bordereau
        facture_sinistre.fichier.save(f'bordereau_paiement_{facture_sinistre.numero}.pdf', File(pdf))
        facture_sinistre.save()

        # notifier du succès
        response = {
            'statut': 1,
            'message': "Demande de remboursement rejetée avec succès !",
            'data': ""
        }

        return JsonResponse(response)

    else:

        return redirect('/')


@transaction.atomic
def submit_generation_br_ordonnancement(request):
    with transaction.atomic():
        # try:
        # RECUPERATION DES DONNEES DU FILTRE
        pprint("----- fn= submit_generation_br_ordonnancement -----")
        print(request.POST)
        print(request.POST.get('search_periode_comptable', 0))
        search_numero_bordereau = request.POST.get('search_numero_bordereau')
        search_facture = request.POST.get('search_facture')
        search_periode_comptable = request.POST.get('search_periode_comptable')
        search_prestataire = request.POST.get('search_prestataire', '')
        search_type_remboursement = request.POST.get('search_type_remboursement', 0)
        search_adherent_principal = request.POST.get('search_adherent_principal', '')
        search_assure = request.POST.get('search_assure', '')
        montant_remb_total = float(request.POST.get('montant_remb_total', '0').replace(' ',''))
        montant_rejet_total = float(request.POST.get('montant_rejet_total', '0').replace(' ',''))
        montant_accepte_total = float(request.POST.get('montant_accepte_total', '0').replace(' ',''))
        ordre_de = request.POST.get('ordre_de', '')
        par_compagnie = request.POST.get('par_compagnie', '')
        par_compagnie = True if par_compagnie == 'true' else False
        print("search_numero_bordereau " + str(search_numero_bordereau))
        print("search_facture " + str(search_facture))
        print("search_periode_comptable " + str(search_periode_comptable))
        print("search_prestataire " + str(search_prestataire))
        print("search_type_remboursement " + search_type_remboursement)
        print("search_adherent_principal " + search_adherent_principal)
        print("search_assure " + search_assure)
        print(montant_remb_total)
        print(montant_rejet_total)
        print(montant_accepte_total)
        print(ordre_de)
        print(par_compagnie)

        is_assure = False

        if search_prestataire:
            prestataire = Prestataire.objects.get(id=search_prestataire)

        if search_adherent_principal:
            adherent_principal = Aliment.objects.get(id=search_adherent_principal)
            pprint("adherent_principal")
            pprint(adherent_principal)
        else:
            adherent_principal = None

        if search_assure:
            assure = Client.objects.get(id=search_assure)
            pprint("assure")
            pprint(assure)
            is_assure = True
        else:
            assure = None

        periode_comptable = PeriodeComptable.objects.get(
            id=int(search_periode_comptable)) if PeriodeComptable.objects.get(
            id=int(search_periode_comptable)) else None

        # if search_type_remboursement == 1:
        #     prestataire = Prestataire.objects.get(id=int(search_prestataire)) if Prestataire.objects.get(id=int(search_prestataire)) else None
        #     print(prestataire)

        # if search_type_remboursement == 2:
        #     adherent_principal = Aliment.objects.get(id=int(search_adherent_principal)) if Aliment.objects.get(id=int(search_adherent_principal)) else None

        # print(periode_comptable)


        # [REMOVE] WAS JUST FOR TEST
        # return JsonResponse(
        #     {'statut': 0, 'message': 'Erreur lors de la génération du bordereau de validation', 'error': ''})

        # FIN RECUPERATION DES DONNEES DU FILTRE

        # RECUPERATION DES SINISTRES TRAITES A PARTI DU FILTRE
        pprint("RECUPERATION DES SINISTRES TRAITES A PARTI DU FILTRE")
        # initial queryset to nothing data
        queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
                                           facture_prestataire__periode_comptable__id=int(search_periode_comptable),
                                           bordereau_ordonnancement__isnull=True,
                                           facture_prestataire__statut=SatutBordereauDossierSinistres.VALIDE,
                                           is_ges_processed=True, ).order_by('-id')


        # if search_facture:
        #     facture = FacturePrestataire.objects.get(id=int(search_facture))
        #     periode_comptable = facture.periode_comptable
        #     queryset = Sinistre.objects.filter(prestataire__bureau=request.user.bureau,
        #                                        facture_prestataire_id=int(search_facture),
        #                                        bordereau_ordonnancement__isnull=True,
        #                                        facture_prestataire__statut=SatutBordereauDossierSinistres.VALIDE,
        #                                        is_ges_processed=True, ).order_by('-id')

        #queryset = queryset.filter(dossier_sinistre__numero__contains=search_numero_bordereau)

        # queryset = queryset.filter(facture_prestataire__periode_comptable__id=int(search_periode_comptable))

        if search_prestataire:
            queryset = queryset.filter(prestataire__id=int(search_prestataire))

        if search_adherent_principal:
            queryset = queryset.filter(adherent_principal__id=int(search_adherent_principal))

        if search_assure:
            queryset = queryset.filter(facture_prestataire__assure__id=int(search_assure))

        if search_facture and queryset:
            queryset = queryset.filter(facture_prestataire_id=int(search_facture))

        if search_type_remboursement:
            queryset = queryset.filter(facture_prestataire__type_remboursement__id=int(search_type_remboursement))
                # FIN RECUPERATION DES SINISTRES TRAITES A PARTI DU FILTRE

        if len(queryset) > 0:

            if search_type_remboursement:
                typeRemboursement = TypeRemboursement.objects.get(id=search_type_remboursement)


            # pdf = borderau_ordonnancement_pdf_exemple(request)
            # FIN GENERATION DU BORDEREAU D'ORDONNANCEMENT (FICHIER PDF)
            pprint("FIN GENERATION DU BORDEREAU D'ORDONNANCEMENT (FICHIER PDF)")

            # Initialisation(set montant a payer) pour le compable 
            montant_total_impaye = montant_accepte_total

            # ENREGISTREMENT DU BORDEREAU D'ORDONNANCEMENT
            if search_type_remboursement == '1':
                borderau_ordonnancement = BordereauOrdonnancement.objects.create(created_by=request.user,
                                                                             type_remboursement=typeRemboursement,
                                                                             prestataire=prestataire if prestataire else None,
                                                                             periode_comptable=periode_comptable,
                                                                             montant_remb_total=montant_remb_total,
                                                                             montant_rejet_total=montant_rejet_total,
                                                                             montant_accepte_total=montant_accepte_total,
                                                                             ordre_de=ordre_de,
                                                                             par_compagnie=par_compagnie,
                                                                             bureau=request.user.bureau,
                                                                             statut_paiement=StatutPaiementSinistre.ORDONNANCE,
                                                                             montant_total_impaye=montant_total_impaye)

            elif search_type_remboursement == '2':
                if  assure:
                    borderau_ordonnancement = BordereauOrdonnancement.objects.create(created_by=request.user,
                                                                                     type_remboursement=typeRemboursement,
                                                                                     assure=assure if assure else None,
                                                                                     periode_comptable=periode_comptable,
                                                                                     montant_remb_total=montant_remb_total,
                                                                                     montant_rejet_total=montant_rejet_total,
                                                                                     montant_accepte_total=montant_accepte_total,
                                                                                     ordre_de=ordre_de,
                                                                                     par_compagnie=par_compagnie,
                                                                                     bureau=request.user.bureau,
                                                                                     statut_paiement=StatutPaiementSinistre.ORDONNANCE,
                                                                                     montant_total_impaye=montant_total_impaye)
                else:
                    borderau_ordonnancement = BordereauOrdonnancement.objects.create(created_by=request.user,
                                                                             type_remboursement=typeRemboursement,
                                                                             adherent_principal=adherent_principal if adherent_principal else None,
                                                                             periode_comptable=periode_comptable,
                                                                             montant_remb_total=montant_remb_total,
                                                                             montant_rejet_total=montant_rejet_total,
                                                                             montant_accepte_total=montant_accepte_total,
                                                                             ordre_de=ordre_de,
                                                                             par_compagnie=par_compagnie,
                                                                             bureau=request.user.bureau,
                                                                             statut_paiement=StatutPaiementSinistre.ORDONNANCE,
                                                                             montant_total_impaye=montant_total_impaye)
            code_bureau = request.user.bureau.code
            borderau_ordonnancement.numero = str(code_bureau) + str(Date.today().day) + str(
                Date.today().month) + str(Date.today().year)[-2:] + '-' + str(borderau_ordonnancement.pk).zfill(
                7) + '-BO'
            borderau_ordonnancement.save()

            # GENERATION DU BORDEREAU D'ORDONNANCEMENT (FICHIER PDF)
            pprint("GENERATION DU BORDEREAU D'ORDONNANCEMENT (FICHIER PDF)")
            # generer le pdf

            if typeRemboursement.code == "TP":
                pprint("borderau_ordonnancement_pdf")
                pdf = borderau_ordonnancement_pdf(request, liste_sinistre=queryset, prestataire=prestataire, bordereau_ordonnancement=borderau_ordonnancement)
            elif typeRemboursement.code == "RD":
                pprint("borderau_ordonnancement_rd_pdf")
                if assure:
                    pdf = borderau_ordonnancement_rd_assure_pdf(request, liste_sinistre=queryset, assure=assure, bordereau_ordonnancement=borderau_ordonnancement)
                else:
                    pdf = borderau_ordonnancement_rd_pdf(request, liste_sinistre=queryset,
                                                         adherent_principal=adherent_principal, bordereau_ordonnancement=borderau_ordonnancement)


            # enregistrer le pdf generé
            borderau_ordonnancement.fichier.save(
                f'bordereau_ordonnancement_remboursement_{borderau_ordonnancement.numero}.pdf', File(pdf))
            borderau_ordonnancement.save()
            # FIN ENREGISTREMENT DU BORDEREAU DE VALIDATION

            # MISE A JOUR DES SINISTRES
            for sinistre in queryset:
                sinistre.bordereau_ordonnancement = borderau_ordonnancement
                sinistre.statut_paiement = StatutPaiementSinistre.ORDONNANCE
                sinistre.facture_prestataire.statut = SatutBordereauDossierSinistres.ORDONNANCE
                sinistre.facture_prestataire.save()
                sinistre.save()
            # FIN MISE A JOUR DES SINISTRES

            return JsonResponse(
                {'statut': 1, 'message': 'Bordereau généré avec succès',
                 'bordereau_pdf': borderau_ordonnancement.fichier.url,
                 'bordereau_id': borderau_ordonnancement.id}, status=200)
        else:
            pprint("AUCUN REMBOURSEMENT TROUVE")
            return JsonResponse({'statut': 0, 'message': 'Facture prestataire partiellement traitée'})

    # except Exception as e:
    #     return JsonResponse(
    #         {'statut': 0, 'message': 'Erreur lors de la génération du bordereau de validation', 'error': str(e)})


def bordereau_ordonnancement_pdf(request):
    bordereau = BordereauOrdonnancement.objects.get(id=15747)
    liste_sinistre = Sinistre.objects.filter(bordereau_ordonnancement_id=15747, facture_prestataire__assure__id=10130)

    pdf = borderau_ordonnancement_rd_assure_pdf(request, liste_sinistre, bordereau.assure)
    return HttpResponse(File(pdf), content_type='application/pdf')

@method_decorator(login_required, name='dispatch')
class BordereauOrdonnancementView(TemplateView):
    template_name = 'liste_bordereau_ordonnancement.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        # dossiers_sinistres = [x for x in
        #                       DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE).order_by('id') if
        #                       x.sinistres.filter(approuved_by__bureau=request.user.bureau).exists()]
        # pprint(dossiers_sinistres)

        prestataires = [x.prestataire for x in
                        BordereauOrdonnancement.objects.filter(prestataire__bureau=request.user.bureau)]
        # Load distinct prestataires
        prestataires = reduce(lambda re, x: re + [x] if x not in re else re, prestataires, [])

        periodes_comptables = PeriodeComptable.objects.all()
        # facture_prestataires = FacturePrestataire.objects.all()
        context = self.get_context_data(**kwargs)
        # context['dossiers_sinistres'] = dossiers_sinistres
        context['prestataires'] = prestataires
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def bordereau_ordonnancement_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')

    queryset = BordereauOrdonnancement.objects.filter(created_by__bureau=request.user.bureau).order_by('-id')

    # la recherche
    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'prestataire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('bordereau_ordonnancement_detail',
                             args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>'

        nom_prestataire = ""

        if c.type_remboursement:
            if c.type_remboursement.code == "TP":
                nom_prestataire = c.prestataire.name
            else:
                nom_prestataire = c.adherent_principal.nom + " " + c.adherent_principal.prenoms

        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "prestataire__name": nom_prestataire,
            "periode_comptable": periode_comptable_libelle,
            "net_a_payer": money_field(c.montant_remb_total),
            "rejet": money_field(c.montant_rejet_total),
            "rembourse": money_field(c.montant_accepte_total),
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })



@method_decorator(login_required, name='dispatch')
class BordereauOrdonnancementPayeView(TemplateView):
    template_name = 'liste_bordereau_ordonnancement_paye.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        # dossiers_sinistres = [x for x in
        #                       DossierSinistre.objects.filter(statut_validite=StatutValidite.VALIDE).order_by('id') if
        #                       x.sinistres.filter(approuved_by__bureau=request.user.bureau).exists()]
        # pprint(dossiers_sinistres)

        prestataires = [x.prestataire for x in
                        BordereauOrdonnancement.objects.filter(prestataire__bureau=request.user.bureau)]
        # Load distinct prestataires
        prestataires = reduce(lambda re, x: re + [x] if x not in re else re, prestataires, [])
        periodes_comptables = PeriodeComptable.objects.all()
        # facture_prestataires = FacturePrestataire.objects.all()
        context = self.get_context_data(**kwargs)
        # context['dossiers_sinistres'] = dossiers_sinistres
        context['prestataires'] = prestataires
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        today = datetime.datetime.now(tz=timezone.utc)
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def bordereau_ordonnancement_paye_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')

    queryset = BordereauOrdonnancement.objects.filter(created_by__bureau=request.user.bureau, numero="PAYE").order_by('-id') #A CORRIGER EN METTANT STATUT

    # la recherche
    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'prestataire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('bordereau_ordonnancement_detail',
                             args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>'

        nom_prestataire = c.prestataire.name if c.prestataire else ""
        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "prestataire__name": nom_prestataire,
            "periode_comptable": periode_comptable_libelle,
            "net_a_payer": money_field(c.montant_accepte_total),
            "rejet": money_field(c.montant_rejet_total),
            "rembourse": money_field(c.montant_remb_total),
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })



@method_decorator(login_required, name='dispatch')
class DetailBordereauOrdonnancementView(TemplateView):
    # permission_required = "sinistre.view_sinistre"
    template_name = 'detail_bordereau_ordonnancement.html'
    model = Sinistre

    def get(self, request, bordereau_id, *args, **kwargs):
        #TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau
        bordereau = BordereauOrdonnancement.objects.get(id=bordereau_id)

        liste_sinistres_bordereau = Sinistre.objects.filter(bordereau_ordonnancement=bordereau)

        montant_remb_total = bordereau.montant_remb_total
        montant_rejet_total = bordereau.montant_rejet_total
        montant_accepte_total = bordereau.montant_accepte_total

        prestataires = [x.prestataire for x in
                        FacturePrestataire.objects.filter(statut=SatutBordereauDossierSinistres.VALIDE)]

        context = self.get_context_data(**kwargs)
        context['bordereau'] = bordereau
        context['liste_sinistres_bordereau'] = liste_sinistres_bordereau
        context['montant_remb_total'] = montant_remb_total
        context['montant_rejet_total'] = montant_rejet_total
        context['montant_accepte_total'] = montant_accepte_total
        context['prestataires'] = prestataires

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@method_decorator(login_required, name='dispatch')
class ExecutionRequeteExcelView(TemplateView):
    template_name = 'execution_requete_excel.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        #TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau
        periode_comptable = PeriodeComptable.objects.all()
        query_datas = [
            {
                "query_label": "LISTE DES SINISTRES ORDONNANCÉS",
                "query_name": "SIN_ORDONNANCES",
                "query_param_data": periode_comptable,
                "query_param_label" :"Période comptable",
                "query_param_name" :"period_comptable",
                "query_param_type" :"select",
                "query_param_isRequired" : False
            },
            {
                "query_label": "LISTE DES SINISTRES ENTRE DEUX DATES",
                "query_name": "SIN_2DATES",
                "query_param_data": periode_comptable,
                "query_param_label": "Période comptable",
                "query_param_name": "period_comptable",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            #
            {
                "query_label": "LISTE DES SINISTRES SAISIES ENTRE DEUX DATES",
                "query_name": "SIN_SAISIES_2DATES",
                "query_param_data": periode_comptable,
                "query_param_label": "Période comptable",
                "query_param_name": "period_comptable",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            ##
            {
                "query_label": "ANALYSE DÉTAILLÉE DES PRIMES COMPTA",
                "query_name": "ANALYSE_PRIMES",
                "query_param_data": periode_comptable,
                "query_param_label" :"Période comptable",
                "query_param_name" :"period_comptable",
                "query_param_type" :"select",
                "query_param_isRequired" : False
            },
            {
                "query_label": "ANALYSE DÉTAILLÉE DES PRIMES COMPTA AVEC APPORTEURS",
                "query_name": "ANALYSE_PRIMES_APPORTEURS",
                "query_param_data": periode_comptable,
                "query_param_label": "Période comptable",
                "query_param_name": "period_comptable",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            {
                "query_label": "SUIVI DES S/P CLIENT SANTÉ PAR FILIALE",
                "query_name": "SUIVI_SP_CLIENT_PAR_FILIALE",
                "query_param_label" :"S/P À LA DATE DU",
                "query_param_name" :"sp_a_la_date_du",
                "query_param_type" :"date",
                "query_param_isRequired" : True
            },
            {
                "query_label": "PAIEMENT SINISTRE SANTE ENTRE DEUX DATES",
                "query_name": "PAIEMENT_SINISTRE_ENTRE_DEUX_DATES",
                "query_param_label" :"PAIEMENT SINISTRES SANTÉ ENTRE DEUX DATES",
                "query_param_name" :"paiement_sinistre_sante_entre_deux_dates",
                "query_param_type" :"date",
                "query_param_isRequired" : True
            },

            {
                "query_label": "EXTRACTION DES SINISTRES TRAITES ET VALIDES",
                "query_name": "SIN_TRAITES_VALIDES",
                "query_param_label": "EXTRACTION DES SINISTRES TRAITÉS ET VALIDÉS",
                "query_param_name": "extraction_des_sinistres_traites_valides",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            {
                "query_label": "EXTRACTION DES DEMANDES D’ACCORDS PREALABLES TRAITEES PAR LES MEDECINS CONSEIL",
                "query_name": "DEMANDES_ACCORDS_TRAITES_MEDECINS",
                "query_param_label": "EXTRACTION DES DEMANDES D’ACCORDS PRÉALABLES TRAITÉES PAR LES MÉDECINS CONSEIL",
                "query_param_name": "extraction_demandes_accords_prealables_traitees_par_medecins_conseil",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            '''
            ##
            {
                "query_label": "DEMANDES ACCORDS PREALABLES TRAITEES PAR LES MEDECINS CONSEIL",
                "query_name": "DEMANDES_ACCORDS_PREALABLES_TRAITEES",
                "query_param_label": "DEMANDES ACCORDS PRÉALABLES TRAITÉES",
                "query_param_name": "demandes_accords_prealables_traitees",
                "query_param_type": "date",
                "query_param_isRequired": True
            },
            {
                "query_label": "SINISTRES TRAITES ET VALIDES PAR LES GESTIONNAIRES",
                "query_name": "SINISTRES_TRAITES_ET_VALIDES_PAR_LES_GESTIONNAIRES",
                "query_param_label": "SINISTRES TRAITES ET VALIDES PAR LES GESTIONNAIRES",
                "query_param_name": "sinistres_traites_par_les_gestionnaires",
                "query_param_type": "date",
                "query_param_isRequired" : True
            }'''
            #
        ]

        #query_datas = []

        # liste_sinistres_bordereau = [x for x in Sinistre.objects.filter(facture_prestataire__isnull=False) if x.is_processed]
        prestataires = [x.prestataire for x in
                        BordereauOrdonnancement.objects.filter(prestataire__bureau=request.user.bureau)]
        # Load distinct prestataires
        prestataires = reduce(lambda re, x: re + [x] if x not in re else re, prestataires, [])


        beneficiaires = [x.adherent_principal for x in
                        BordereauOrdonnancement.objects.filter(adherent_principal__bureau=request.user.bureau).order_by("adherent_principal__nom", "adherent_principal__prenoms")]
        # Load distinct beneficiaires
        beneficiaires = reduce(lambda re, x: re + [x] if x not in re else re, beneficiaires, [])

        # dd(beneficiaires)

        context = self.get_context_data(**kwargs)
        context['query_datas'] = query_datas
        context['periode_comptable'] = periode_comptable
        context['prestataires'] = prestataires
        context['beneficiaires'] = beneficiaires

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        print("----- fn= post -----")
        print(request.POST)
        print(request.POST.dict())

        query_name = request.POST.get('query_name')
        period_comptable = request.POST.get("period_comptable")
        prestataire = request.POST.get("prestataire")
        beneficiaire = request.POST.get("beneficiaire")
        # dates = request.POST.get("dates").split(" - ")
        date_debut = request.POST.get("date_debut")
        date_fin = request.POST.get("date_fin")

        sp_a_la_date_du = request.POST.get("sp_a_la_date_du")

        date_debut_paiment_sinisre = request.POST.get("date_debut_paiment_sinisre")
        date_fin_paiment_sinisre = request.POST.get("date_fin_paiment_sinisre")

        date_debut_survenance_demandes_accords_prealables_traitees = request.POST.get("date_debut_survenance_demandes_accords_prealables_traitees")
        date_fin_survenance_demandes_accords_prealables_traitees = request.POST.get("date_fin_survenance_demandes_accords_prealables_traitees")

        date_debut_survenance_sinistres_traites_par_les_gestionnaires = request.POST.get("date_debut_survenance_sinistres_traites_par_les_gestionnaires")
        date_fin_survenance_sinistres_traites_par_les_gestionnaires = request.POST.get("date_fin_survenance_sinistres_traites_par_les_gestionnaires")



        if query_name == "SIN_ORDONNANCES":

            reference_facture = request.POST.get("reference_facture")

            if prestataire:
                query = requete_liste_sinistre_ordonnancee_par_period_par_prestataire(request.user.bureau.id, date_debut=date_debut, date_fin=date_fin, prestataire_id=int(prestataire), reference_facture=reference_facture)

                # creation de tach background
                create_query_background_task(name=f'LISTE DES SINISTRES ORDONNANCÉS {date_debut} _ {date_fin}', query=query, request=request)

                queryset, header = execute_query(query)
            elif beneficiaire:
                query = requete_liste_sinistre_ordonnancee_par_period_par_beneficiaire(request.user.bureau.id, date_debut=date_debut, date_fin=date_fin, beneficiaire_id=int(beneficiaire), reference_facture=reference_facture)

                # creation de tach background
                create_query_background_task(name=f'LISTE DES SINISTRES ORDONNANCÉS {date_debut} _ {date_fin}', query=query, request=request)

                queryset, header = execute_query(query)
            else:
                query = requete_liste_sinistre_ordonnancee_par_period(request.user.bureau.id, date_debut=date_debut, date_fin=date_fin, reference_facture=reference_facture)

                # creation de tach background
                create_query_background_task(name=f'LISTE DES SINISTRES ORDONNANCÉS {date_debut} _ {date_fin}', query=query, request=request)

                queryset, header = execute_query(query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)


            # Exportation excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'LISTE DES SINISTRES ORDONNANCÉS'


            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response

            # return exportation_en_excel_avec_style('LISTE DES SINISTRES ORDONNANCÉS', header, queryset)

        elif query_name == "SIN_2DATES":
            date_debut = request.POST.get("sin_2dates_date_debut")
            date_fin = request.POST.get("sin_2dates_date_fin")
            reference_facture = request.POST.get("sin_2dates_reference_facture")

            query = requete_liste_sinistre_entre_2date(bureau_id=request.user.bureau.id, date_debut=date_debut, date_fin=date_fin, reference_facture=reference_facture)

            # creation de tache background
            create_query_background_task(name=f'SINISTRES_ENTRE_2DATES {date_debut} _ {date_fin}', query=query,
                                         request=request)

            queryset, header = execute_query(query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)



            # Exportation excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'SINISTRES_ENTRE_2DATES'

            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response
            # return exportation_en_excel_avec_style('SINISTRES_ENTRE_2DATES', header, queryset)

        #
        elif query_name == "SIN_SAISIES_2DATES":
            date_debut = request.POST.get("sin_saisies_2dates_date_debut")
            date_fin = request.POST.get("sin_saisies_2dates_date_fin")
            reference_facture = request.POST.get("sin_saisies_2dates_reference_facture")

            query = requete_liste_sinistre_saisies_entre_2date(bureau_id=request.user.bureau.id, date_debut=date_debut, date_fin=date_fin, reference_facture=reference_facture)

            # creation de tache background
            create_query_background_task(name=f'SINISTRES_SAISIES_ENTRE_2DATES {date_debut} _ {date_fin}', query=query,
                                         request=request)

            queryset, header = execute_query(query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)



            # Exportation excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'SINISTRES_SAISIES_ENTRE_2DATES'

            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response
            # return exportation_en_excel_avec_style('SINISTRES_ENTRE_2DATES', header, queryset)
            
            
            
            
            
        #
        elif query_name == "ANALYSE_PRIMES":

            response = requete_analyse_prime_compta(request)
            return response

        elif query_name == "ANALYSE_PRIMES_APPORTEURS":
            # save log
            response = requete_analyse_prime_compta_apporteur(request)
            return response

        elif query_name == "SUIVI_SP_CLIENT_PAR_FILIALE":

            if not sp_a_la_date_du:
                return JsonResponse({
                "message": "Veuillez svp remplir le date de limitation S/P client."
            }, status=400)

            title = "SUIVI_SP_CLIENT_PAR_FILIALE"

            query = requete_liste_des_sp_client_par_filiale(request.user.bureau.code, sp_a_la_date_du)

            # creation de tache background
            create_query_background_task(name=f'SUIVI_SP_CLIENT_PAR_FILIALE {sp_a_la_date_du}', query=query,
                                         request=request)

            queryset, header = execute_query(query)



            return exportation_en_excel_avec_style(title, header, queryset)

        elif query_name == "PAIEMENT_SINISTRE_ENTRE_DEUX_DATES":

            reference_facture = request.POST.get("reference_facture_paiment_sinisre")
            numero_police = request.POST.get("numero_police")

            title = "PAIEMENT_SINISTRE_ENTRE_DEUX_DATES"

            query = requete_liste_paiement_sinistre_sante_entre_deux_dates(request.user.bureau.code, date_debut_paiment_sinisre, date_fin_paiment_sinisre, reference_facture, numero_police)

            # creation de tache background
            create_query_background_task(name=f'PAIEMENT_SINISTRE_ENTRE_DEUX_DATES {date_debut_paiment_sinisre} _ {date_fin_paiment_sinisre}', query=query, request=request)
            queryset, header = execute_query(query)

            pprint("query")
            pprint(query)


            return exportation_en_excel_avec_style(title, header, queryset)

        elif query_name == "DEMANDES_ACCORDS_PREALABLES_TRAITEES":

            title = "ACCORDS_PREALABLES_TRAITEES"

            query = requete_demandes_accords_prealables_traitees_par_les_medecins_conseil(request.user.bureau.id, date_debut_survenance_demandes_accords_prealables_traitees, date_fin_survenance_demandes_accords_prealables_traitees)

            # creation de tache background
            create_query_background_task(name=f'DEMANDES_ACCORDS_PREALABLES_TRAITEES {date_debut_survenance_demandes_accords_prealables_traitees} _ {date_fin_survenance_demandes_accords_prealables_traitees}', query=query, request=request)
            queryset, header = execute_query(query)

            pprint("query")
            pprint(query)

            return exportation_en_excel_avec_style(title, header, queryset)

        elif query_name == "SINISTRES_TRAITES_ET_VALIDES_PAR_LES_GESTIONNAIRES":
            reference_facture = request.POST.get("reference_facture_sinistres_traites_par_gestionnaires")

            title = "SINISTRES_TRAITES_PAR_GESTIONNAIRES"

            query = requete_sinistres_traites_et_valides_par_les_gestionnaires(request.user.bureau.id, date_debut_survenance_sinistres_traites_par_les_gestionnaires, date_fin_survenance_sinistres_traites_par_les_gestionnaires, reference_facture)

            # creation de tache background
            create_query_background_task(name=f'SINISTRES_TRAITES_ET_VALIDES_PAR_LES_GESTIONNAIRES {date_debut_survenance_sinistres_traites_par_les_gestionnaires} _ {date_fin_survenance_sinistres_traites_par_les_gestionnaires}', query=query, request=request)
            queryset, header = execute_query(query)

            pprint("query")
            pprint(query)

            return exportation_en_excel_avec_style(title, header, queryset)

        ##
        elif query_name == "SIN_TRAITES_VALIDES":
            date_debut = request.POST.get("sin_traites_valides_date_debut")
            date_fin = request.POST.get("sin_traites_valides_date_fin")

            query = extraction_des_sinistres_traites_valides(bureau_id=request.user.bureau.id, date_debut=date_debut, date_fin=date_fin)

            # creation de tache background
            create_query_background_task(name=f'SIN_TRAITES_VALIDES {date_debut} _ {date_fin}', query=query,
                                         request=request)

            queryset, header = execute_query(query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)

            # Exportation excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'SIN_TRAITES_VALIDES'

            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response

        ##
        elif query_name == 'DEMANDES_ACCORDS_TRAITES_MEDECINS':
            date_debut = request.POST.get("demandes_accords_traites_medecins_date_debut")
            date_fin = request.POST.get("demandes_accords_traites_medecins_date_fin")

            query = extraction_demandes_accords_prealables_traitees_par_medecins_conseil(bureau_id=request.user.bureau.id, date_debut=date_debut, date_fin=date_fin)

            # creation de tache background
            create_query_background_task(name=f'DEMANDES_ACCORDS_TRAITES_MEDECINS {date_debut} _ {date_fin}', query=query,
                                         request=request)

            queryset, header = execute_query(query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)

            # Exportation excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'DEMANDES_ACCORDS_TRAITES_MEDECINS'

            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num + 1, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response
        ##

        else:
            return JsonResponse({
                "message": "La requête n'est pas prise en charge"
            }, status=404)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def verif_background_requete_excel(request):
    task_id = request.session.get('task_id',None)
    task_event = request.POST.get('task_event', None)
    print(task_id)

    if task_id:
        try:
            task = BackgroundQueryTask.objects.get(id=task_id)
            if task_event:
                task.status = "ENCOURS"
                task.save()
            else:
                task.delete()

            return JsonResponse({
                "status": "OK",
                "task_id": task_id
            }, status=200)
        except BackgroundQueryTask.DoesNotExist:
            return JsonResponse({
                "status": "KO",
                "task_id": task_id
            }, status=404)

    return JsonResponse({
        "status": "KO",
        "task_id": task_id
    }, status=404)


#
# @method_decorator(login_required, name='dispatch')
# class ExecutionRequeteExcelView(TemplateView):
#     template_name = 'execution_requete_excel.html'
#     model = Sinistre
#
#     def get(self, request, args, *kwargs):
#         #TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau
#         periode_comptable = PeriodeComptable.objects.all()
#         query_datas = [
#             {
#                 "query_label": "LISTE DE SINISTRES ORDONNANCÉS",
#                 "query_name": "SIN_ORDONNANCES",
#                 "query_param_data": periode_comptable,
#                 "query_param_label" :"Période comptable",
#                 "query_param_name" :"period_comptable",
#                 "query_param_type" :"select",
#                 "query_param_isRequired" : False
#             },
#         ]
#
#         # liste_sinistres_bordereau = [x for x in Sinistre.objects.filter(facture_prestataire__isnull=False) if x.is_processed]
#         prestataires = [x.prestataire for x in
#                         BordereauOrdonnancement.objects.filter(prestataire__bureau=request.user.bureau)]
#         # Load distinct prestataires
#         prestataires = reduce(lambda re, x: re + [x] if x not in re else re, prestataires, [])
#
#         context = self.get_context_data(**kwargs)
#         context['query_datas'] = query_datas
#         context['periode_comptable'] = periode_comptable
#         context['prestataires'] = prestataires
#
#         return self.render_to_response(context)
#
#     def post(self, request, args, *kwargs):
#         print("----- fn= post -----")
#         print(request.POST)
#         query_name = request.POST.get('query_name')
#         period_comptable = request.POST.get("period_comptable")
#         prestataire = request.POST.get("prestataire")
#         # dates = request.POST.get("dates").split(" - ")
#         # date_debut = datetime.datetime.strptime(dates[0], "%d/%m/%Y").date()
#         # date_fin = datetime.datetime.strptime(dates[1], "%d/%m/%Y").date() + datetime.timedelta(days=1)
#         # print(dates)
#         # print(date_debut)
#         # print(date_fin)
#         print(query_name)
#         print(period_comptable)
#         print(prestataire)
#
#         if query_name == "SIN_ORDONNANCES":
#             # queryset = Sinistre.objects.filter(bordereau_ordonnancement__isnull=False).order_by('-id')
#             queryset = Sinistre.objects.filter(bordereau_ordonnancement__isnull=False, police__bureau=request.user.bureau, bordereau_ordonnancement__periode_comptable=int(period_comptable), bordereau_ordonnancement__prestataire=int(prestataire)).order_by('-id')
#             print("queryset")
#             print(queryset)
#             # Exportation excel
#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'
#
#             workbook = openpyxl.Workbook()
#             worksheet = workbook.active
#             worksheet.title = 'LISTE DE SINISTRES ORDONNANCÉS'
#
#             # Write header row
#
#             header = [
#                 'NOM_CLIENT',
#                 'ID_SIN',
#                 'NUMERO_DOSSIER',
#                 'NUMERO_POLICE',
#                 'DATS_SIN',
#                 'DATE_EFFET',
#                 'DATE_ECHEANCE',
#                 'NOM_CIE',
#                 'ADHERENT_PRINCIPAL',
#                 'NUMERO_FAMILLE',
#                 'FORMULE',
#                 'NUMERO_CARTE',
#                 'NOM_PATIENT',
#                 'PRENOM_PATIENT',
#                 'DATE_NAISSANCE_PATIENT',
#                 'LIEN_PATIENT',
#                 'DATE_SINISTRE',
#                 'DATE_SAISIE',
#                 'SAISI_PAR',
#                 'ACTE',
#                 'LIB_AFFECTION',
#                 'CODE_CIMDIS',
#                 'LIB_REGROUPEMENT',
#                 'PRESTATAIRE',
#                 'FRAIS_REEL',
#                 'PART_INOV',
#                 'PART_ASSURE',
#                 'DEPASSEMENT/EXCLUSION',
#                 'TICKET PREFINANCE',
#                 'PART COMPAGNIE',
#                 'REJET',
#                 'NET_REGLE',
#                 'DATE_REGLEMENT',
#                 'NUMERO_BORDEREAU',
#                 'BENEFICIAIRE_DU_REMB',
#                 'STATUT_PEC',
#                 'STATUT_DEMAND_REMB',
#                 'STATUT_SINISTRE',
#                 'DATE RECEPTION FACTURE',
#                 'PERIODE COMPTABLE',
#                 'DATE_GENERATION_BORDEREAU',
#             ]
#             for col_num, column_title in enumerate(header, 1):
#                 cell = worksheet.cell(row=1, column=col_num)
#                 cell.value = column_title
#
#             # Write data rows
#             data = []
#             for c in queryset:
#                 if not c.aliment:
#                     c.aliment.nom = ''
#                 if not c.aliment:
#                     c.aliment.prenoms = ''
#
#                 total_facture = c.total_frais_reel if c.total_frais_reel else 0
#                 total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
#                 total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
#                 total_rejet = c.montant_remb_refuse if c.montant_remb_refuse else 0
#                 total_net_payer = c.montant_remb_accepte if c.montant_remb_accepte else 0
#
#                 # collecte des montants totaux
#
#                 cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
#                 numero_carte = cartes.first().numero if cartes else None
#                 # ['Période comptable', 'Date soin', 'N˚ Sinistre', 'N˚ Bon', 'Prestataire', 'Bénéficiaire du soin', 'Matricule', 'Montant facturé', 'Par assuré', 'Base remb.', 'Rejet', 'Net à payer']
#                 # data_iten = [
#                 #     c.bordereau_ordonnancement.periode_comptable.libelle if c.bordereau_ordonnancement.periode_comptable else "",
#                 #     c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
#                 #     c.numero,
#                 #     c.dossier_sinistre.numero,
#                 #     c.prestataire.name,
#                 #     c.aliment.nom + ' ' + c.aliment.prenoms,
#                 #     numero_carte,
#                 #     money_field(total_facture),
#                 #     money_field(total_part_assure),
#                 #     money_field(total_base_remb),
#                 #     money_field(total_rejet),
#                 #     money_field(total_net_payer),
#                 # ]
#
#                 data_iten = [
#                     c.police.client.nom if c.police.client.nom else '' + ' ' + c.police.client.prenoms if c.police.client.prenoms else '',
#                     c.numero,
#                     c.dossier_sinistre.numero,
#                     c.police.numero,
#                     c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
#                     c.police.date_debut_effet.strftime("%d/%m/%Y") if c.police.date_debut_effet else '',
#                     c.police.date_fin_effet.strftime("%d/%m/%Y") if c.police.date_fin_effet else '',
#                     c.police.compagnie.nom,
#                     c.aliment.adherent_principal.nom + ' ' + c.aliment.adherent_principal.prenoms,
#                     c.aliment.adherent_principal.numero_famille,
#                     c.formulegarantie.libelle,
#                     numero_carte,
#                     c.aliment.nom,
#                     c.aliment.prenoms,
#                     c.aliment.date_naissance.strftime("%d/%m/%Y") if c.aliment.date_naissance else '',
#                     c.aliment.qualite_beneficiaire.libelle,
#                     c.date_survenance.strftime("%d/%m/%Y") if c.date_survenance else '',
#                     c.created_at.strftime("%d/%m/%Y") if c.created_at else '',
#                     c.created_by.first_name + ' ' + c.created_by.last_name,
#                     c.acte.libelle,
#                     c.affection.libelle if c.affection else '',
#                     c.affection.code_cim_10 if c.affection else '',
#                     c.acte.regroupement_acte.libelle if c.acte.regroupement_acte else '',
#                     c.prestataire.name,
#                     total_facture,
#                     total_base_remb,
#                     total_part_assure,
#                     c.depassement if c.depassement else 0,
#                     c.total_part_assure if c.tm_prefinanced else 0, # 'TICKET PREFINANCE' a discuter,
#                     total_net_payer,
#                     total_rejet,
#                     total_net_payer,
#                     c.bordereau_ordonnancement.created_at.strftime("%d/%m/%Y") if c.bordereau_ordonnancement.created_at else '',
#                     c.bordereau_ordonnancement.numero,
#                     c.bordereau_ordonnancement.ordre_de,
#                     c.statut,
#                     c.statut_remboursement,
#                     c.statut_validite,
#                     c.facture_prestataire.created_at.strftime("%d/%m/%Y") if c.facture_prestataire.created_at else '',
#                     c.bordereau_ordonnancement.periode_comptable.libelle if c.bordereau_ordonnancement.periode_comptable else "",
#                     c.bordereau_ordonnancement.created_at.strftime("%d/%m/%Y") if c.bordereau_ordonnancement.created_at else '',
#                 ]
#                 data.append(data_iten)
#
#             for row_num, row in enumerate(data, 1):
#                 for col_num, cell_value in enumerate(row, 1):
#                     cell = worksheet.cell(row=row_num + 1, column=col_num)
#                     cell.value = cell_value
#
#             workbook.save(response)
#             return response
#         else:
#             return JsonResponse({
#                 "message": "La requête n'est pas prise en charge"
#             }, status=404)
#
#     def get_context_data(self, **kwargs):
#         return {
#             **super().get_context_data(**kwargs),
#             **admin.site.each_context(self.request),
#             "opts": self.model._meta,
#         }
#


