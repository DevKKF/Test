from pprint import pprint

from comptabilite.models import EncaissementCommission
from configurations.helper_config import send_notification_background_task_mail, execute_query
from configurations.models import ActionLog, BackgroundQueryTask, CronLog, Prestataire, Retenue
from production.models import MouvementPolice, Reglement, Quittance, ApporteurPolice
from shared.enum import StatutReversementCompagnie, StatutEncaissementCommission, StatutValidite
from shared.veos import get_taux_euro_by_devise
import openpyxl
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.core.files import File
from datetime import datetime
import os
from django.db.models import Q, Sum


def requete_liste_sinistre_ordonnancee_par_period(bureau_id, date_debut, date_fin, reference_facture):
    print(date_debut)
    print(date_fin)

    requete = f"""
    SELECT
    CONCAT(COALESCE(clt.nom),' ', COALESCE(clt.prenoms, '')) 'CLIENT',
    doss_sin.numero 'NUMERO_DOSSIER',
    pl.numero 'NUMERO_POLICE',
    cart.numero 'NUMERO_CARTE',
    alm.nom 'NOM_PATIENT',
    alm.prenoms 'PRENOM_PATIENT',
    alm.matricule_employe 'MATRICULE',
    alm.date_naissance 'DATE_NAISSANCE_PATIENT',
    'ENTRÉ' as 'STATUT_PATIENT',
    CONCAT(alm_adh.nom,' ',alm_adh.prenoms) 'ADHERENT_PRINCIPAL',
    fg.libelle 'FORMULE',
    sin.numero 'NUMERO_SINISTRE',
    sin.date_survenance 'DATE_SINISTRE',
    f_pres.created_at 'DATE_DE_RECEPTION_FACTURE',
    sin.reference_facture 'REFERENCE_FACTURE',
    act.libelle 'ACTE',
    affect.libelle 'AFFECTION',
    affect.code_cim_10 'CODE_AFFECTION',
    CONCAT(prest.code,'|',prest.name) 'PRESTATAIRE',
    rss.nom as 'RESEAU_SOIN',
    tp_prest.name 'TYPE_PRESTATAIRE',
    centre_prest.name 'CENTRE_PRESCRIPTEUR',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_TRAITANT',
    spte.name 'SPECIALITE_MEDECIN_TRAITANT',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_PRESCRIPTEUR',
    spte.name 'SPECIALITE_MEDECIN_PRESCRIPTEUR',
    CASE  
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0 THEN sin.nombre_demande * sin.frais_reel
        ELSE sin.frais_reel
    END 'FRAIS_REEL',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie 
        ELSE sin.part_compagnie
    END 'PART_INOV',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure 
        ELSE sin.part_assure
    END 'PART_ASSURE',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'TPS / AIB / BNC',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'FAR',
    sin.depassement 'DEPASSEMENT/EXCLUSION',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        ELSE 0
    END 'TICKET PREFINANCE',
    (select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'PART COMPAGNIE',
    # (select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'NET_REGLE',
    ((select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id)) 'NET_REGLE',
    br_ord.created_at 'DATE_REG',
    '' as 'NUMERO_LCHQ',
    br_ord.numero 'NUMERO_BORDEREAU',
    prest.code 'NUM_BENEFICIAIRE_DU_REMB',
    CASE
        WHEN prest.rb_ordre IS NULL THEN prest.name 
        ELSE prest.rb_ordre 
    END 'NOM_BENEFICIAIRE_DU_REMB',
    cgni.nom 'COMPAGNIE',
    b.code 'CODE_SOCIETE',
    rg_act.libelle 'REGROUPEMENT_INOV',
    qtbf.libelle 'TYPE_ASSURE'
    
    
    FROM sinistres sin
    LEFT JOIN formulegarantie fg on fg.id = sin.formulegarantie_id
    LEFT JOIN aliments alm ON alm.id = sin.aliment_id
    LEFT JOIN aliments alm_adh ON alm_adh.id=alm.adherent_principal_id
    LEFT JOIN cartes cart on cart.aliment_id=alm.id AND cart.statut='ACTIF'
    LEFT JOIN qualite_beneficiaire qtbf on qtbf.id = alm.qualite_beneficiaire_id
    LEFT JOIN actes act on act.id = sin.acte_id
    LEFT JOIN regroupement_acte rg_act on rg_act.id  = act.regroupement_acte_id
    LEFT JOIN prescripteur presteur ON presteur.id = sin.prescripteur_id
    LEFT JOIN specialite spte on spte.id = presteur.specialite_id
    LEFT JOIN dossier_sinistre doss_sin on doss_sin.id = sin.dossier_sinistre_id
    LEFT JOIN prestataires centre_prest on centre_prest.id = doss_sin.centre_prescripteur_id
    LEFT JOIN polices pl on sin.police_id=pl.id
    LEFT JOIN compagnies cgni on sin.compagnie_id=cgni.id
    LEFT JOIN clients clt on clt.id=pl.client_id
    LEFT JOIN factures_prestataires f_pres on f_pres.id = sin.facture_prestataire_id
    LEFT JOIN affections affect on affect.id = sin.affection_id
    LEFT JOIN prestataires prest on sin.prestataire_id=prest.id
    LEFT JOIN bureau b on b.id=prest.bureau_id
    LEFT JOIN type_prestataires tp_prest on tp_prest.id=prest.type_prestataire_id
    LEFT JOIN bordereau_ordonnancement  br_ord on br_ord.id=sin.bordereau_ordonnancement_id
    LEFT JOIN prestataires br_ord_prest on br_ord_prest.id=br_ord.prestataire_id
    LEFT JOIN type_prefinancement typ_pref on typ_pref.id=sin.type_prefinancement_id
    LEFT JOIN reseaux_soins rss on rss.id=fg.reseau_soin_id
    WHERE sin.bordereau_ordonnancement_id IS NOT NULL 
    AND sin.dossier_sinistre_id is not null
    AND sin.statut_validite = 'VALIDE'
    AND prest.bureau_id={bureau_id}
    AND DATE(br_ord.created_at) BETWEEN '{date_debut}' AND '{date_fin}'
    """

    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete 
    
def requete_liste_sinistre_ordonnancee_par_period_par_prestataire(bureau_id, date_debut, date_fin, prestataire_id, reference_facture):
    
    requete = f"""
    SELECT
    CONCAT(COALESCE(clt.nom),' ', COALESCE(clt.prenoms, '')) 'CLIENT',
    doss_sin.numero 'NUMERO_DOSSIER',
    pl.numero 'NUMERO_POLICE',
    cart.numero 'NUMERO_CARTE',
    alm.nom 'NOM_PATIENT',
    alm.prenoms 'PRENOM_PATIENT',
    alm.matricule_employe 'MATRICULE',
    alm.date_naissance 'DATE_NAISSANCE_PATIENT',
    'ENTRÉ' as 'STATUT_PATIENT',
    CONCAT(alm_adh.nom,' ',alm_adh.prenoms) 'ADHERENT_PRINCIPAL',
    fg.libelle 'FORMULE',
    sin.numero 'NUMERO_SINISTRE',
    sin.date_survenance 'DATE_SINISTRE',
    f_pres.created_at 'DATE_DE_RECEPTION_FACTURE',
    sin.reference_facture 'REFERENCE_FACTURE',
    act.libelle 'ACTE',
    affect.libelle 'AFFECTION',
    affect.code_cim_10 'CODE_AFFECTION',
    CONCAT(prest.code,'|',prest.name) 'PRESTATAIRE',
    '' as 'RESEAU_SOIN',
    tp_prest.name 'TYPE_PRESTATAIRE',
    centre_prest.name 'CENTRE_PRESCRIPTEUR',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_TRAITANT',
    spte.name 'SPECIALITE_MEDECIN_TRAITANT',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_PRESCRIPTEUR',
    spte.name 'SPECIALITE_MEDECIN_PRESCRIPTEUR',
    CASE  
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0 THEN sin.nombre_demande * sin.frais_reel
        ELSE sin.frais_reel
    END 'FRAIS_REEL',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie 
        ELSE sin.part_compagnie
    END 'PART_INOV',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure 
        ELSE sin.part_assure
    END 'PART_ASSURE',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'TPS / AIB / BNC',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'FAR',
    sin.depassement 'DEPASSEMENT/EXCLUSION',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        ELSE 0
    END 'TICKET PREFINANCE',
    (select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'PART COMPAGNIE',
    ((select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id)) 'NET_REGLE',
    br_ord.created_at 'DATE_REG',
    rss.nom as 'NUMERO_LCHQ',
    br_ord.numero 'NUMERO_BORDEREAU',
    prest.code 'NUM_BENEFICIAIRE_DU_REMB',
    CASE
        WHEN prest.rb_ordre IS NULL THEN prest.name 
        ELSE prest.rb_ordre 
    END 'NOM_BENEFICIAIRE_DU_REMB',
    cgni.nom 'COMPAGNIE',
    b.code 'CODE_SOCIETE',
    rg_act.libelle 'REGROUPEMENT_INOV',
    qtbf.libelle 'TYPE_ASSURE'


    FROM sinistres sin
    LEFT JOIN formulegarantie fg on fg.id = sin.formulegarantie_id
    LEFT JOIN aliments alm ON alm.id = sin.aliment_id
    LEFT JOIN aliments alm_adh ON alm_adh.id=alm.adherent_principal_id
    LEFT JOIN cartes cart on cart.aliment_id=alm.id AND cart.statut='ACTIF'
    LEFT JOIN qualite_beneficiaire qtbf on qtbf.id = alm.qualite_beneficiaire_id
    LEFT JOIN actes act on act.id = sin.acte_id
    LEFT JOIN regroupement_acte rg_act on rg_act.id  = act.regroupement_acte_id
    LEFT JOIN prescripteur presteur ON presteur.id = sin.prescripteur_id
    LEFT JOIN specialite spte on spte.id = presteur.specialite_id
    LEFT JOIN dossier_sinistre doss_sin on doss_sin.id = sin.dossier_sinistre_id
    LEFT JOIN prestataires centre_prest on centre_prest.id = doss_sin.centre_prescripteur_id
    LEFT JOIN polices pl on sin.police_id=pl.id
    LEFT JOIN compagnies cgni on sin.compagnie_id=cgni.id
    LEFT JOIN clients clt on clt.id=pl.client_id
    LEFT JOIN factures_prestataires f_pres on f_pres.id = sin.facture_prestataire_id
    LEFT JOIN affections affect on affect.id = sin.affection_id
    LEFT JOIN prestataires prest on sin.prestataire_id=prest.id
    LEFT JOIN bureau b on b.id=prest.bureau_id
    LEFT JOIN type_prestataires tp_prest on tp_prest.id=prest.type_prestataire_id
    LEFT JOIN bordereau_ordonnancement  br_ord on br_ord.id=sin.bordereau_ordonnancement_id
    LEFT JOIN prestataires br_ord_prest on br_ord_prest.id=br_ord.prestataire_id
    LEFT JOIN type_prefinancement typ_pref on typ_pref.id=sin.type_prefinancement_id
    LEFT JOIN reseaux_soins rss on rss.id=fg.reseau_soin_id
    WHERE sin.bordereau_ordonnancement_id IS NOT NULL 
    AND sin.dossier_sinistre_id is not null
    AND sin.statut_validite = 'VALIDE'
    AND prest.bureau_id={bureau_id}
    AND DATE(br_ord.created_at) BETWEEN '{date_debut}' AND '{date_fin}'
    AND prest.id={prestataire_id}
    """
    
    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete 

def requete_liste_sinistre_ordonnancee_par_period_par_beneficiaire(bureau_id, date_debut, date_fin, beneficiaire_id, reference_facture):
    
    requete = f"""
    SELECT
    CONCAT(COALESCE(clt.nom),' ', COALESCE(clt.prenoms, '')) 'CLIENT',
    doss_sin.numero 'NUMERO_DOSSIER',
    pl.numero 'NUMERO_POLICE',
    cart.numero 'NUMERO_CARTE',
    alm.nom 'NOM_PATIENT',
    alm.prenoms 'PRENOM_PATIENT',
    alm.matricule_employe 'MATRICULE',
    alm.date_naissance 'DATE_NAISSANCE_PATIENT',
    'ENTRÉ' as 'STATUT_PATIENT',
    CONCAT(alm_adh.nom,' ',alm_adh.prenoms) 'ADHERENT_PRINCIPAL',
    fg.libelle 'FORMULE',
    sin.numero 'NUMERO_SINISTRE',
    sin.date_survenance 'DATE_SINISTRE',
    f_pres.created_at 'DATE_DE_RECEPTION_FACTURE',
    sin.reference_facture 'REFERENCE_FACTURE',
    act.libelle 'ACTE',
    affect.libelle 'AFFECTION',
    affect.code_cim_10 'CODE_AFFECTION',
    CONCAT(prest.code,'|',prest.name) 'PRESTATAIRE',
    '' as 'RESEAU_SOIN',
    tp_prest.name 'TYPE_PRESTATAIRE',
    centre_prest.name 'CENTRE_PRESCRIPTEUR',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_TRAITANT',
    spte.name 'SPECIALITE_MEDECIN_TRAITANT',
    CONCAT(presteur.nom,' ',presteur.prenoms) 'MEDECIN_PRESCRIPTEUR',
    spte.name 'SPECIALITE_MEDECIN_PRESCRIPTEUR',
    CASE  
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0 THEN sin.nombre_demande * sin.frais_reel
        ELSE sin.frais_reel
    END 'FRAIS_REEL',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.frais_reel 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_compagnie 
        ELSE sin.part_compagnie
    END 'PART_INOV',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * 0 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure 
        ELSE sin.part_assure
    END 'PART_ASSURE',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'TPS / AIB / BNC',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'FAR',
    sin.depassement 'DEPASSEMENT/EXCLUSION',
    CASE 
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NOT NULL) AND typ_pref.code = 'PREF_TOUT') THEN sin.nombre_demande * sin.part_assure
        WHEN (sin.acte_id IS NOT null) AND ((act.option_seance=1 or act.code='G66023CI01') and sin.nombre_demande IS NOT NULL and sin.nombre_demande > 0 and sin.statut='EN ATTENTE') AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        WHEN (sin.medicament_id IS NOT null) and sin.nombre_demande > 0  AND ((sin.type_prefinancement_id IS NULL) OR typ_pref.code <> 'PREF_TOUT') THEN sin.nombre_demande * 0
        ELSE 0
    END 'TICKET PREFINANCE',
    (select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'PART COMPAGNIE',
    # (select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'NET_REGLE',
    ((select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id)) 'NET_REGLE',
    br_ord.created_at 'DATE_REG',
    rss.nom as 'NUMERO_LCHQ',
    br_ord.numero 'NUMERO_BORDEREAU',
    prest.code 'NUM_BENEFICIAIRE_DU_REMB',
    CASE
        WHEN prest.rb_ordre IS NULL THEN prest.name 
        ELSE prest.rb_ordre 
    END 'NOM_BENEFICIAIRE_DU_REMB',
    cgni.nom 'COMPAGNIE',
    b.code 'CODE_SOCIETE',
    rg_act.libelle 'REGROUPEMENT_INOV',
    qtbf.libelle 'TYPE_ASSURE'


    FROM sinistres sin
    LEFT JOIN formulegarantie fg on fg.id = sin.formulegarantie_id
    LEFT JOIN aliments alm ON alm.id = sin.aliment_id
    LEFT JOIN aliments alm_adh ON alm_adh.id=alm.adherent_principal_id
    LEFT JOIN cartes cart on cart.aliment_id=alm.id AND cart.statut='ACTIF'
    LEFT JOIN qualite_beneficiaire qtbf on qtbf.id = alm.qualite_beneficiaire_id
    LEFT JOIN actes act on act.id = sin.acte_id
    LEFT JOIN regroupement_acte rg_act on rg_act.id  = act.regroupement_acte_id
    LEFT JOIN prescripteur presteur ON presteur.id = sin.prescripteur_id
    LEFT JOIN specialite spte on spte.id = presteur.specialite_id
    LEFT JOIN dossier_sinistre doss_sin on doss_sin.id = sin.dossier_sinistre_id
    LEFT JOIN prestataires centre_prest on centre_prest.id = doss_sin.centre_prescripteur_id
    LEFT JOIN polices pl on sin.police_id=pl.id
    LEFT JOIN compagnies cgni on sin.compagnie_id=cgni.id
    LEFT JOIN clients clt on clt.id=pl.client_id
    LEFT JOIN factures_prestataires f_pres on f_pres.id = sin.facture_prestataire_id
    LEFT JOIN affections affect on affect.id = sin.affection_id
    LEFT JOIN prestataires prest on sin.prestataire_id=prest.id
    LEFT JOIN bureau b on b.id=prest.bureau_id
    LEFT JOIN type_prestataires tp_prest on tp_prest.id=prest.type_prestataire_id
    LEFT JOIN bordereau_ordonnancement  br_ord on br_ord.id=sin.bordereau_ordonnancement_id
    LEFT JOIN prestataires br_ord_prest on br_ord_prest.id=br_ord.prestataire_id
    LEFT JOIN type_prefinancement typ_pref on typ_pref.id=sin.type_prefinancement_id
    LEFT JOIN reseaux_soins rss on rss.id=fg.reseau_soin_id
    WHERE sin.bordereau_ordonnancement_id IS NOT NULL 
    AND sin.dossier_sinistre_id is not null
    AND sin.statut_validite = 'VALIDE'
    AND prest.bureau_id={bureau_id}
    AND DATE(br_ord.created_at) BETWEEN '{date_debut}' AND '{date_fin}'
    AND alm_adh.id={beneficiaire_id}
    """
    
    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete 
    
def requete_demandes_accords_prealables_traitees_par_les_medecins_conseil(bureau_id, date_debut_survenance_demandes_accords_prealables_traitees, date_fin_survenance_demandes_accords_prealables_traitees):
    requete = f"""
    SELECT 
    CONCAT_WS(' ', cus.first_name, cus.last_name) AS medecin_conseil,
    CONCAT_WS(' ', COALESCE(cl.nom, ''), COALESCE(cl.prenoms, '')) AS client,
    cie.nom AS garant,
    po.numero AS numero_police,
    s.numero AS numero_dossier,
    c.numero AS numero_carte,
    adh.nom AS nom_patient,
    adh.prenoms AS prenom_patient,
    (SELECT libelle FROM qualite_beneficiaire WHERE id = adh.qualite_beneficiaire_id) AS qualite_patient,
    DATE_FORMAT(adh.date_naissance, "%d/%m/%Y") AS date_naissance_patient,
    CONCAT(ap.nom, ' ', ap.prenoms) AS adherent_principal,
    DATE_FORMAT(s.date_survenance, "%d/%m/%Y") AS date_sinistre,
    actes.libelle AS acte,
    aff.libelle AS affection,
    aff.code_cim_10 AS code_affection,
    pr.name AS prestataire,
    (SELECT tp.name FROM type_prestataires tp WHERE tp.id = pr.type_prestataire_id) AS type_prestataires,
    (SELECT presc.name FROM prestataires presc WHERE presc.id = ds.centre_prescripteur_id) AS centre_prescripteur,
    CONCAT(medp.nom, ' ', medp.prenoms) AS medecin_traitant,
    spec.name AS specialite_medecin_traitant,
    s.frais_reel,
    s.part_assure,
    s.part_compagnie AS part_inov,
	s.statut AS statut_pec,
	s.motif_rejet,
	s.statut_validite
	
    FROM sinistres s
    JOIN dossier_sinistre ds ON ds.id = s.dossier_sinistre_id
    JOIN actes ON actes.id = s.acte_id
    JOIN affections aff ON aff.id = s.affection_id
    JOIN prestataires pr ON pr.id = s.prestataire_id
    JOIN polices po ON po.id = s.police_id
    JOIN clients cl ON cl.id = po.client_id
    JOIN compagnies cie ON cie.id = po.compagnie_id
    JOIN bureau br ON br.id = pr.bureau_id
    JOIN aliments adh ON adh.id = s.aliment_id
    JOIN cartes c ON c.aliment_id = adh.id AND c.date_desactivation IS NULL
    JOIN aliments ap ON ap.id = adh.adherent_principal_id
    LEFT JOIN prescripteur medp ON medp.id = s.prescripteur_id
    LEFT JOIN specialite spec ON spec.id = medp.specialite_id
    LEFT JOIN configurations_user cus ON cus.id = s.approuved_by_id

    WHERE s.approuved_by_id IS NOT NULL 
        AND br.id = '{bureau_id}'
        AND (s.date_survenance BETWEEN '{date_debut_survenance_demandes_accords_prealables_traitees}' AND '{date_fin_survenance_demandes_accords_prealables_traitees}');
    """

    return requete

def requete_sinistres_traites_et_valides_par_les_gestionnaires(bureau_id, date_debut_survenance_sinistres_traites_par_les_gestionnaires, date_fin_survenance_sinistres_traites_par_les_gestionnaires, reference_facture):
    requete = f"""
    SELECT 
    br.code AS code_societe,
    CONCAT_WS(' ', cus.first_name, cus.last_name) AS gestionnaire,
    CONCAT_WS(' ', COALESCE(cl.nom, ''), COALESCE(cl.prenoms, '')) AS client,
    cie.nom AS compagnie,
    po.numero AS numero_police,
    s.numero AS numero_dossier,
    c.numero AS numero_carte,
    adh.nom AS nom_patient,
    adh.prenoms AS prenom_patient,
    (SELECT libelle FROM qualite_beneficiaire WHERE id = adh.qualite_beneficiaire_id) AS qualite_patient,
    DATE_FORMAT(adh.date_naissance, "%d/%m/%Y") AS date_naissance_patient,
    CONCAT(ap.nom, ' ', ap.prenoms) AS adherent_principal,
    DATE_FORMAT(s.date_survenance, "%d/%m/%Y %H:%i") AS date_sinistre,
    s.date_reception_facture,
    actes.libelle AS acte,
    aff.libelle AS affection,
    aff.code_cim_10 AS code_affection,
    pr.name AS prestataire,
    (SELECT tp.name FROM type_prestataires tp WHERE tp.id = pr.type_prestataire_id) AS type_prestataires,
    (SELECT presc.name FROM prestataires presc WHERE presc.id = ds.centre_prescripteur_id) AS centre_prescripteur,
    CONCAT(medp.nom, ' ', medp.prenoms) AS medecin_traitant,
    spec.name AS specialite_medecin_traitant,
    s.frais_reel,
    s.part_assure,
    s.part_compagnie AS part_inov,
    s.statut AS statut_pec,
    s.statut_remboursement,
    s.montant_remboursement_accepte,
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) 'tps_aib_bnc',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) 'far',
    DATE_FORMAT(rsin.created_at, "%d/%m/%Y %H:%i") AS date_traitement,
    rsin.motif,
    s.statut_validite
        
    FROM sinistres s
    JOIN dossier_sinistre ds ON ds.id = s.dossier_sinistre_id
    JOIN actes ON actes.id = s.acte_id
    JOIN affections aff ON aff.id = s.affection_id
    JOIN prestataires pr ON pr.id = s.prestataire_id
    JOIN remboursement_sinistre rsin ON rsin.sinistre_id = s.id  AND rsin.is_invalid = 0 AND rsin.statut='ACCEPTE'
    JOIN polices po ON po.id = s.police_id
    JOIN clients cl ON cl.id = po.client_id
    JOIN compagnies cie ON cie.id = po.compagnie_id
    JOIN bureau br ON br.id = pr.bureau_id
    JOIN aliments adh ON adh.id = s.aliment_id
    JOIN cartes c ON c.aliment_id = adh.id AND c.date_desactivation IS NULL
    JOIN aliments ap ON ap.id = adh.adherent_principal_id
    LEFT JOIN prescripteur medp ON medp.id = s.prescripteur_id
    LEFT JOIN specialite spec ON spec.id = medp.specialite_id
    LEFT JOIN configurations_user cus ON cus.id = rsin.created_by_id 
    
    WHERE br.id = '{bureau_id}' 
        AND (s.date_survenance BETWEEN '{date_debut_survenance_sinistres_traites_par_les_gestionnaires}' AND '{date_fin_survenance_sinistres_traites_par_les_gestionnaires}');
        """

    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete


def requete_liste_sinistre_entre_2date(bureau_id, date_debut, date_fin, reference_facture):
    print("requete_liste_sinistre_entre_2date")
    print(bureau_id)
    print(date_debut)
    print(date_fin)
    requete = f"""
    With dataremb as (
	select 
	    s.id
	    ,DATE_FORMAT(pc.date_paiement,"%d/%m/%Y") as date_reg
	    ,pc.numero as numero_bordereau
	    ,case 
	        when bo.prestataire_id is not null then pr.name
	        when bo.adherent_principal_id is not null then concat(coalesce(ap.nom,''), ' ', coalesce(ap.prenoms,''))
	    end as Nom_Beneficiaire_du_remb
	    ,pc.numero_piece as numero_piece
	from sinistres s  
	join prestataires pr on pr.id = s.prestataire_id
	join aliments adh on adh.id = s.aliment_id
	join aliments ap on ap.id = adh.adherent_principal_id
	join bordereau_ordonnancement bo on bo.id = s.bordereau_ordonnancement_id
	join paiement_comptable pc on pc.bordereau_ordonnancement_id = bo.id
	where s.statut_validite = 'VALIDE' and pc.bureau_id={bureau_id}
    )
    ,
    contrat as (
        select 
            p.id,
            p.numero ,
            concat(coalesce(c.nom,''),' ',coalesce(c.prenoms,'')) as nom_client	 
        from polices p 
        join clients c on c.id = p.client_id
        where p.bureau_id={bureau_id}
    )
    select 
	(select contrat.nom_client from contrat where contrat.id = fg.police_id) 'NOM_CLIENT',
	sin.numero 'NUMERO_DOSSIER',
	(select contrat.numero from contrat where contrat.id = fg.police_id) 'NUMERO_POLICE',
	DATE_FORMAT(per_couv.date_debut_effet,"%d/%m/%Y") 'DATE_EFFET',
    DATE_FORMAT(per_couv.date_fin_effet,"%d/%m/%Y") 'DATE_ECHEANCE',
	cgni.nom 'NOM_CIE',
	CONCAT(COALESCE(alm_adh.nom),' ',COALESCE(alm_adh.prenoms)) 'ADHERENT_PRINCIPAL',
	alm_adh.numero_famille 'NUMERO_FAMILLE',
    fg.libelle 'FORMULE', 
    cart.numero 'NUMERO_CARTE',
    alm.nom 'NOM_PATIENT',
    alm.prenoms 'PRENOM_PATIENT',
    DATE_FORMAT(alm.date_affiliation,"%d/%m/%Y") 'DATE_EFFET_PATIENT',
    DATE_FORMAT(alm.date_naissance,"%d/%m/%Y") 'DATE_NAISSANCE_PATIENT',
    qtbf.libelle 'LIEN_PATIENT',
    DATE_FORMAT(sin.date_survenance,"%d/%m/%Y") 'DATE_SINISTRE',
    DATE_FORMAT(sin.created_at,"%d/%m/%Y") 'DATE_SAISIE',
    (select CONCAT(util.last_name,' ',util.first_name) from configurations_user util where util.id = sin.created_by_id) 'SAISI_PAR',
    act.libelle 'ACTE',
    affect.libelle 'LIB_AFFECTION',
    affect.code_cim_10 'CODE_CIMDIS',
    rg_act.libelle 'LIB_REGROUPEMENT',
    CONCAT(prest.code,'|',prest.name) 'PRESTATAIRE',
    sin.frais_reel 'FRAIS_REEL',
    sin.part_compagnie 'PART_INOV',
    sin.part_assure 'PART_ASSURE',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'TPS / AIB / BNC',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'FAR',
    sin.depassement 'DEPASSEMENT/EXCLUSION',
    CASE WHEN tpr.code = "PREF_TOUT" THEN sin.part_assure else 0 END 'TICKET_PREFINANCE',
    sin.part_compagnie 'PART_COMPAGNIE',
    # CASE WHEN dataremb.date_reg IS NULL THEN NULL ELSE rsin.montant END 'NET_REGLE',
    CASE WHEN dataremb.date_reg IS NULL THEN NULL ELSE ((select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id)) END 'NET_REGLE',
    dataremb.date_reg 'DATE_REGLEMENT',
    dataremb.numero_bordereau 'NUMERO_BORDEREAU',
    dataremb.Nom_Beneficiaire_du_remb 'BENEFICIAIRE_DU_REMB',
    sin.statut 'STATUT_PEC',
    sin.statut_remboursement 'STATUT_DEMAND_REMB',
    (select CONCAT(util.last_name,' ',util.first_name) from configurations_user util where util.id = rsin.created_by_id) 'DEMAND_REMB_VALIDEE_PAR',
	sin.statut_validite 'STATUT_SINISTRE',
    (select DATE_FORMAT(f_pres.created_at,"%d/%m/%Y") from factures_prestataires f_pres where f_pres.id = sin.facture_prestataire_id) 'DATE_RECEPTION_FACTURE',
    sin.reference_facture 'REFERENCE_FACTURE'
    FROM sinistres sin 
    JOIN periode_couverture per_couv on per_couv.id = sin.periode_couverture_id 
    JOIN aliments alm ON alm.id = sin.aliment_id
    JOIN cartes cart on cart.aliment_id=alm.id AND cart.statut='ACTIF'
    JOIN aliments alm_adh ON alm_adh.id=alm.adherent_principal_id
    JOIN formulegarantie fg on fg.id = sin.formulegarantie_id
    JOIN qualite_beneficiaire qtbf on qtbf.id = alm.qualite_beneficiaire_id
    JOIN actes act on act.id = sin.acte_id
    JOIN regroupement_acte rg_act on rg_act.id  = act.regroupement_acte_id
    JOIN prestataires prest on sin.prestataire_id=prest.id
    JOIN compagnies cgni on sin.compagnie_id=cgni.id
    LEFT JOIN type_prefinancement tpr on tpr.id = sin.type_prefinancement_id
    LEFT JOIN affections affect on affect.id = sin.affection_id
    LEFT JOIN remboursement_sinistre rsin ON rsin.sinistre_id = sin.id AND rsin.is_invalid = 0 AND rsin.statut = 'ACCEPTE' 
    LEFT JOIN dataremb on sin.id = dataremb.id
    WHERE prest.bureau_id={bureau_id}
    AND sin.dossier_sinistre_id is not null
    AND DATE(sin.date_survenance) BETWEEN '{date_debut}' AND '{date_fin}'
    """

    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete 


def requete_analyse_prime_compta(request):
    ActionLog.objects.create(done_by=request.user, action="execution_requete_excel_compta",
                             description="Extraction de la requete ANALYSE_PRIMES",
                             table="",
                             row=None, data_before=None,
                             data_after=None)

    # queryset = Sinistre.objects.filter(bordereau_ordonnancement__isnull=False).order_by('-id')
    queryset = Quittance.objects.select_related('police').filter(statut_validite=StatutValidite.VALIDE,
                                                                 bureau_id=request.user.bureau.id, import_stats=False).order_by('-id')
    default_taux_euro = get_taux_euro_by_devise(request.user.bureau.pays.devise.code)
    print(default_taux_euro)
    # dd(queryset)
    pprint(queryset.count())
    # Exportation excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ANALYSE_PRIMES_COMPTA.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'ANALYSE DES PRIMES COMPTA'

    # Write header row

    header = [
        'Pays', #1
        'Societe', #2
        'No client', #3
        'Nom du client', #4
        # 'Famille Branche', #5
        'Branche', #6
        # 'Groupe', #7
        # 'Groupe Compagnie', #8
        'Compagnie', #9
        'Numéro de la police', #10
        # 'Programme', #11
        'Effet Police', #12
        'Date Ech Police', #13
        'Date der mvt police', #14
        'Tx Com Court Police', #15
        'Tx Com Gest Police', #16
        'Devise', #17
        'Quittance', #18
        'Type quittance', #19
        "Date d'émission", #20
        'Date effet prime', #21
        'Date règlement client', #22
        'Date reversement', #23
        'date operation revers', #24
        'date encaissement commission', #25
        'Prime_nette', #26
        'Prime_nette EUR', #27
        'Prime TTC', #28
        'Prime TTC EUR', #29
        'Com Court émise', #30
        'Com Court émise EUR', #31
        'Com Gest émise', #32
        'Com Gest émise EUR', #33
        'Prime EVCAT', #34
        'COM EVCAT', #35
        'Com Court encaissée', #36
        'Com Court encaissée EUR', #37
        'Com Court Non encaissée', #38
        'Com Court Non encaissée EUR', #39
        'Com Gest encaissée',  # 40
        'Com Gest encaissée EUR',  # 41
        'Com Gest Non encaissée',  # 42
        'Com Gest Non encaissée EUR', # 43

        ##
        'Cout police courtier', #44
        'Cout police courtier EUR', #45

        'Frais Cabinet', #46
        'Frais Cabinet EUR', #47
        'Acc. Cie', #48
        'Acc. Cie EUR', #49
        'Acc Court', #50
        'Acc Court EUR', #51
        'ca_revers_aa', #52
        'ca_revers_aamm', #53
        'ca_emis_aa', #54
        'ca_emis_aa2', #55
        'ca_emis_aamm', #56
        'ca_emis_aamm2', #57
        'ca_type', #58
        'ca_aa', #59
        'ca_aamm', #60
        'EUROS' #61
    ]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    data = []
    for c in queryset:

        taux_euro = c.taux_euro if c.taux_euro and c.taux_euro > 0 else default_taux_euro

        nom = c.police.client.nom if c.police.client.nom else ''
        prenoms = c.police.client.prenoms if c.police.client.prenoms else ''
        nom_client = f"{nom if c.police.client.nom else ''} {prenoms if c.police.client.prenoms else ''}"
        date_der_mvt_police = ''
        date_reglement_client = ''  # c.date_reglement_client

        last_mouvement_police = MouvementPolice.objects.filter(police=c.police).order_by('-id').first()
        date_der_mvt_police = last_mouvement_police.created_at if last_mouvement_police else ''

        last_reglement = Reglement.objects.filter(quittance=c).order_by('-id').first()
        date_reglement_client = last_reglement.date_paiement if last_reglement else ''

        last_reglement_reverse_cie = Reglement.objects.filter(quittance=c,
                                                              statut_reversement_compagnie=StatutReversementCompagnie.REVERSE).order_by(
            '-date_reversement_compagnie').first()
        date_reversement_compagnie = last_reglement_reverse_cie.date_reversement_compagnie if last_reglement_reverse_cie else ''

        last_reglement = Reglement.objects.filter(quittance=c).order_by('-id').first()
        last_com_encaisse = None
        if last_reglement:
            last_com_encaisse = EncaissementCommission.objects.filter(reglement_id=last_reglement.id).order_by('-id').first()
        date_encaissement_commission = last_com_encaisse.created_at if last_com_encaisse else ''

        reglements_quittance = Reglement.objects.filter(quittance=c)
        com_court_encaissee = 0
        com_gest_encaissee = 0


        for r in reglements_quittance:
            com_court_encaissee += float(r.montant_com_courtage_encaisse())
            com_gest_encaissee += float(r.montant_com_gestion_encaisse())

        com_gest_non_encaissee = c.commission_gestion - com_gest_encaissee
        com_court_non_encaissee = c.commission_courtage - com_court_encaissee

        ca_revers_aa = 0
        ca_revers_aamm = 0
        ca_emis_aa = c.date_emission.strftime("%Y") if hasattr(c, 'date_emission') else None
        ca_emis_aa2 = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
        ca_emis_aamm = c.date_emission.strftime("%Y%m") if hasattr(c, 'date_emission') else None
        ca_emis_aamm2 = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
        ca_type = c.police.bureau.ca_type
        ca_aa = None
        ca_aamm = None

        # ca_revers_aa
        if date_reversement_compagnie:
            ca_revers_aa = date_reversement_compagnie.strftime(
                "%Y") if c.police.date_debut_effet <= date_reversement_compagnie.date() <= c.police.date_fin_effet else 0
            ca_revers_aamm = date_reversement_compagnie.strftime(
                "%Y%m") if c.police.date_debut_effet <= date_reversement_compagnie.date() <= c.police.date_fin_effet else 0

        # ca_emis_aa
        if int(ca_emis_aa2) <= int(ca_emis_aa):
            ca_emis_aa2 = ca_emis_aa
            ca_emis_aamm2 = ca_emis_aamm

        # ca_aa
        if ca_type == 'EMISSION':
            if int(c.date_debut.strftime("%Y")) >= int(c.date_emission.strftime("%Y")):
                ca_aa = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
                ca_aamm = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
            else:
                ca_aa = c.date_emission.strftime("%Y") if hasattr(c, 'date_emission') else None
                ca_aamm = c.date_emission.strftime("%Y%m") if hasattr(c, 'date_emission') else None
        else:
            if date_reversement_compagnie:
                date_reversement = date_reversement_compagnie.strftime("%Y")
                if int(c.date_debut.strftime("%Y")) <= int(date_reversement):
                    ca_aa = date_reversement_compagnie.strftime("%Y")
                    ca_aamm = date_reversement_compagnie.strftime("%Y%m")
                elif int(c.date_debut.strftime("%Y")) > int(date_reversement):
                    ca_aa = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
                    ca_aamm = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
                else:
                    ca_aa = 0
                    ca_aamm = 0
            else:
                ca_aa = 0
                ca_aamm = 0

        # pprint(date_der_mvt_police)

        data_iten = [
            c.police.bureau.pays.nom if hasattr(c.police, 'bureau') else '', #1
            c.police.bureau.code if hasattr(c.police, 'bureau') else '', #2
            c.police.client.code if hasattr(c.police.client, 'code') else '', #3
            nom_client, #4
            # c.police.famille_branche if hasattr(c.police, 'famille_branche') else '', #5
            c.police.produit.branche if hasattr(c.police, 'branche') else 'SANTE',  #6
            # c.police.client.groupe if hasattr(c.police.client, 'groupe') else '',#7
            # c.police.groupe_compagnie if hasattr(c.police, 'groupe_compagnie') else '', #8
            c.police.compagnie.nom if hasattr(c.police, 'compagnie') else '', #9
            c.police.numero if hasattr(c.police, 'numero') else '', #10
            # c.police.programme if hasattr(c.police, 'programme') else '', #11
            c.police.date_debut_effet.strftime("%d/%m/%Y") if hasattr(c.police, 'date_debut_effet') else '', #12
            c.police.date_fin_effet.strftime("%d/%m") if hasattr(c.police, 'date_fin_effet') else '', #13
            date_der_mvt_police.strftime("%d/%m/%Y") if date_der_mvt_police else '', #14
            c.police.taux_com_courtage if hasattr(c.police, 'taux_com_courtage') else '', #15
            c.police.taux_com_gestion if hasattr(c.police, 'taux_com_gestion') else '', #16
            c.police.bureau.pays.devise.code if c.police.bureau.pays.devise else '', #17
            c.numero if hasattr(c, 'numero') else '', #18
            c.type_quittance.libelle if hasattr(c, 'type_quittance') else '', #19
            c.date_emission.strftime("%Y-%m") if hasattr(c, 'date_emission') else '', #20
            c.date_debut.strftime("%Y-%m") if hasattr(c, 'date_debut') else '', #21
            date_reglement_client.strftime("%Y-%m") if date_reglement_client else '', #22
            date_reversement_compagnie.strftime("%Y-%m") if date_reversement_compagnie else '', #23
            date_reversement_compagnie.strftime("%d/%m/%Y") if date_reversement_compagnie else '', #24
            # en attendant de voir si les dates sont différentes
            date_encaissement_commission.strftime("%Y-%m") if date_encaissement_commission else '', #25
            c.prime_ht if hasattr(c, 'prime_ht') else '', #26
            round((c.prime_ht * taux_euro), 2) if hasattr(c, 'prime_ht') else '',  # prime_ht_eur #27
            c.prime_ttc if hasattr(c, 'prime_ttc') else '', #28
            round((c.prime_ttc * taux_euro), 2) if hasattr(c, 'prime_ttc') else '',  # prime_ttc_eur #29
            c.commission_courtage if hasattr(c, 'commission_courtage') else '',  # ajouter appport gest #30
            round((c.commission_courtage * taux_euro), 2) if hasattr(c, 'commission_courtage') else '', #31
            # commission_courtage_euro
            c.commission_gestion if hasattr(c, 'commission_gestion') else '',  # ajouter appport gest #32
            round((c.commission_gestion * taux_euro), 2) if hasattr(c, 'commission_gestion') else '', #33
            # commission_gestion_euro
            c.prime_EVCAT if hasattr(c, 'prime_EVCAT') else '',  #34
            c.COM_EVCAT if hasattr(c, 'COM_EVCAT') else '',  #35

            com_court_encaissee if com_court_encaissee else 0, #36
            round((com_court_encaissee * taux_euro), 2) if com_court_encaissee else 0,  # com_encaissee_EUR #37
            com_court_non_encaissee if com_court_non_encaissee else 0, # com_non_encaissee #38
            round((com_court_non_encaissee * taux_euro), 2) if com_court_non_encaissee else 0, #39

            com_gest_encaissee if com_gest_encaissee else 0,  # 40
            round((com_gest_encaissee * taux_euro), 2) if com_gest_encaissee else 0,  # com_encaissee_EUR #41
            com_gest_non_encaissee if com_gest_non_encaissee else 0,  # com_non_encaissee #42
            round((com_gest_non_encaissee * taux_euro), 2) if com_gest_non_encaissee else 0,  # com_non_encaissee_EUR #43

            ##
            c.cout_police_courtier if hasattr(c, 'cout_police_courtier') else '', #44
            round((c.cout_police_courtier * taux_euro), 2) if hasattr(c, 'cout_police_courtier') else '',  # cout_police_courtier_EUR #45

            # com_non_encaissee_EUR
            c.frais_cabinet if hasattr(c, 'frais_cabinet') else '',  # accessoire courtier #46
            round((c.frais_cabinet * taux_euro), 2) if hasattr(c, 'frais_cabinet') else '',  # frais_cabinet_EUR #47
            c.acc_cie if hasattr(c, 'acc_cie') else '', #48
            round((c.acc_cie * taux_euro), 2) if hasattr(c, 'acc_cie') else '',  # acc_cie_EUR #49
            c.acc_court if hasattr(c, 'acc_court') else '', #50
            round((c.acc_court * taux_euro), 2) if hasattr(c, 'acc_court') else '',  # acc_court_EUR #51
            ca_revers_aa if ca_revers_aa else 0,  # annee de reversement comprise dans la période de couverture, sinon 0 #52
            ca_revers_aamm if ca_revers_aamm else 0, #53
            ##mois et année de reversement comprise dans 0la période de couverture, sinon 0
            ca_emis_aa if ca_emis_aa else '',  # année d'émission de la quittance (date emission) #54
            ca_emis_aa2 if ca_emis_aa2 else '', #55
            # année d'émission de la quittance (date emission) année/mois  d'émission de la prime sauf si la prime concerne l'année suivante (période de garantie sur l'année suivante) auquel cas l'année suivante
            ca_emis_aamm if ca_emis_aamm else '',  # année mois émission prime #56
            ca_emis_aamm2 if ca_emis_aamm2 else '', #57
            # année/mois  d'émission de la prime sauf si la prime concerne l'année suivante (période de garantie sur l'année suivante) auquel cas l'année suivante
            ca_type if ca_type else '', #58
            # indique si le pays est en compta à l'émission ou au reversement (A paramétrer sur le bureau)
            ca_aa if ca_aa else '0',  # année pour laquelle le ca va compter #59
            ca_aamm if ca_aamm else '0', #60
            # année de comptabilisation du CA selon si on est à l'émission ou au reversement
            taux_euro if taux_euro else '',  # taux de change (A paramétrer de façon générale pour l'application) #61
        ]
        data.append(data_iten)

    for row_num, row in enumerate(data, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def requete_analyse_prime_compta_apporteur(request):
    ActionLog.objects.create(done_by=request.user, action="execution_requete_excel_compta",
                             description="Extraction de la requete ANALYSE_PRIMES",
                             table="",
                             row=None, data_before=None,
                             data_after=None)

    # queryset = Sinistre.objects.filter(bordereau_ordonnancement__isnull=False).order_by('-id')
    queryset = Quittance.objects.select_related('police','compagnie').filter(statut_validite=StatutValidite.VALIDE,
                                                                 bureau_id=request.user.bureau.id, import_stats=False).order_by('-id')
    default_taux_euro = get_taux_euro_by_devise(request.user.bureau.pays.devise.code)
    print(default_taux_euro)
    # dd(queryset)
    pprint(queryset.count())
    # Exportation excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ANALYSE_PRIMES_AVEC_APPORTEUR.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'ANALYSE AVEC APPORTEURS'

    # Write header row

    header = [
        'Pays', #1
        'Societe', #2
        'Groupe', #3 --
        'No client', #4
        'Nom du client', #5
        'Famille Branche', #6 --
        'Branche', #7
        'Groupe Compagnie', #8 --
        'Compagnie', #9
        'Numéro de la police', #10
        'Programme', #11 --
        'Effet Police', #12
        'Date Ech Police', #13
        'Date der mvt police', #14
        'Tx Com Court Police', #15
        'Tx Com Gest Police', #16
        'Devise', #17
        'Quittance', #18
        'Type quittance', #19
        "Date d'émission", #20
        'Date effet prime', #21
        'Date règlement client', #22
        'Date reversement', #23
        'date operation revers', #24
        'date encaissement commission', #25
        'Prime_nette', #26
        'Prime_nette EUR', #27
        'Prime TTC', #28
        'Prime TTC EUR', #29
        'Com Court émise', #30
        'Com Court émise EUR', #31
        'Com Gest émise', #32
        'Com Gest émise EUR', #33
        'Prime EVCAT', #34
        'COM EVCAT', #35
        'Com Court encaissée', #36
        'Com Court encaissée EUR', #37
        'Com Court Non encaissée', #38
        'Com Court Non encaissée EUR', #39
        'Com Gest encaissée',  # 40
        'Com Gest encaissée EUR',  # 41
        'Com Gest Non encaissée',  # 42
        'Com Gest Non encaissée EUR', # 43
        'Retro emise', #44 --
        'Retro emise EUR', #45 --
        'Retro Payée', #46 --
        'Retro Payée EUR', #47 --
        'Frais Cabinet', #48
        'Frais Cabinet EUR', #49
        'Acc. Cie', #50
        'Acc. Cie EUR', #51
        'Acc Court', #52
        'Acc Court EUR', #53
        'ca_revers_aa', #54
        'ca_revers_aamm', #55
        'ca_emis_aa', #56
        'ca_emis_aa2', #57
        'ca_emis_aamm', #58
        'ca_emis_aamm2', #59
        'ca_type', #60
        'ca_aa', #61
        'ca_aamm', #62
        #'EUROS' #63,
        'Apporteur police', #64 --
        'Apporteur quittance', #65 --
        'Apporteur Début', #66 --
        'Apporteur Fin', #67 --
        'Apporteur Tx Terme', #68 --
        'Apporteur Tx Cpt', #69 --
        'Apporteur Mnt Fixe', #70 --

    ]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    data = []
    for c in queryset:

        taux_euro = c.taux_euro if c.taux_euro and c.taux_euro > 0 else default_taux_euro

        nom = c.police.client.nom if c.police.client.nom else ''
        prenoms = c.police.client.prenoms if c.police.client.prenoms else ''
        nom_client = f"{nom if c.police.client.nom else ''} {prenoms if c.police.client.prenoms else ''}"
        date_der_mvt_police = ''
        date_reglement_client = ''  # c.date_reglement_client

        last_mouvement_police = MouvementPolice.objects.filter(police=c.police).order_by('-id').first()
        date_der_mvt_police = last_mouvement_police.created_at if last_mouvement_police else ''

        last_reglement = Reglement.objects.filter(quittance=c).order_by('-id').first()
        date_reglement_client = last_reglement.date_paiement if last_reglement else ''

        last_reglement_reverse_cie = Reglement.objects.filter(quittance=c,
                                                              statut_reversement_compagnie=StatutReversementCompagnie.REVERSE).order_by(
            '-date_reversement_compagnie').first()
        date_reversement_compagnie = last_reglement_reverse_cie.date_reversement_compagnie if last_reglement_reverse_cie else ''

        last_reglement = Reglement.objects.filter(quittance=c).order_by('-id').first()
        last_com_encaisse = None
        if last_reglement:
            last_com_encaisse = EncaissementCommission.objects.filter(reglement_id=last_reglement.id).order_by('-id').first()
        date_encaissement_commission = last_com_encaisse.created_at if last_com_encaisse else ''

        reglements_quittance = Reglement.objects.filter(quittance=c)
        com_court_encaissee = 0
        com_gest_encaissee = 0
        retro_paye = 0


        for r in reglements_quittance:
            com_court_encaissee += float(r.montant_com_courtage_encaisse())
            com_gest_encaissee += float(r.montant_com_gestion_encaisse())
            retro_paye += float(r.montant_com_intermediaire)

        com_gest_non_encaissee = c.commission_gestion - com_gest_encaissee
        com_court_non_encaissee = c.commission_courtage - com_court_encaissee

        ca_revers_aa = 0
        ca_revers_aamm = 0
        ca_emis_aa = c.date_emission.strftime("%Y") if hasattr(c, 'date_emission') else None
        ca_emis_aa2 = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
        ca_emis_aamm = c.date_emission.strftime("%Y%m") if hasattr(c, 'date_emission') else None
        ca_emis_aamm2 = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
        ca_type = c.police.bureau.ca_type
        ca_aa = None
        ca_aamm = None

        # ca_revers_aa
        if date_reversement_compagnie:
            ca_revers_aa = date_reversement_compagnie.strftime(
                "%Y") if c.police.date_debut_effet <= date_reversement_compagnie.date() <= c.police.date_fin_effet else 0
            ca_revers_aamm = date_reversement_compagnie.strftime(
                "%Y%m") if c.police.date_debut_effet <= date_reversement_compagnie.date() <= c.police.date_fin_effet else 0

        # ca_emis_aa
        if int(ca_emis_aa2) <= int(ca_emis_aa):
            ca_emis_aa2 = ca_emis_aa
            ca_emis_aamm2 = ca_emis_aamm

        # ca_aa
        if ca_type == 'EMISSION':
            if int(c.date_debut.strftime("%Y")) >= int(c.date_emission.strftime("%Y")):
                ca_aa = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
                ca_aamm = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
            else:
                ca_aa = c.date_emission.strftime("%Y") if hasattr(c, 'date_emission') else None
                ca_aamm = c.date_emission.strftime("%Y%m") if hasattr(c, 'date_emission') else None
        else:
            if date_reversement_compagnie:
                date_reversement = date_reversement_compagnie.strftime("%Y")
                if int(c.date_debut.strftime("%Y")) <= int(date_reversement):
                    ca_aa = date_reversement_compagnie.strftime("%Y")
                    ca_aamm = date_reversement_compagnie.strftime("%Y%m")
                elif int(c.date_debut.strftime("%Y")) > int(date_reversement):
                    ca_aa = c.date_debut.strftime("%Y") if hasattr(c, 'date_debut') else None
                    ca_aamm = c.date_debut.strftime("%Y%m") if hasattr(c, 'date_debut') else None
                else:
                    ca_aa = 0
                    ca_aamm = 0
            else:
                ca_aa = 0
                ca_aamm = 0

        # pprint(date_der_mvt_police)
        # apporteur police
        apporteur_police_query = ApporteurPolice.objects.filter(police=c.police, statut_validite=StatutValidite.VALIDE).first()
        apporteur_police = None
        if apporteur_police_query:
            apporteur_police = f"{apporteur_police_query.apporteur.nom if apporteur_police_query.apporteur else ''} {apporteur_police_query.apporteur.prenoms if apporteur_police_query.apporteur.prenoms else ''}"

        # Programme
        programme_international = c.police.programme_international if c.police else None
        programme = None
        if programme_international:
            programme = "LOCAL" if programme_international == 'NON' else "INTERNATIONAL"

        # groups inter
        groupe_international = c.police.client.groupe_international if c.police.client.groupe_international else None

        # groupe compagnie
        groupe_compagnie = c.compagnie.groupe_compagnie.code if c.compagnie.groupe_compagnie else None

        data_iten = [
            c.police.bureau.pays.nom if hasattr(c.police, 'bureau') else '', #1
            c.police.bureau.code if hasattr(c.police, 'bureau') else '', #2
            groupe_international.code if groupe_international else '',#3 -- Groupe
            c.police.client.code if hasattr(c.police.client, 'code') else '', #4
            nom_client, #5
            c.police.produit.famille_branche if hasattr(c.police, 'famille_branche') else 'ASSURANCE DE PERSONNES', #6 --
            c.police.produit.branche if hasattr(c.police, 'branche') else 'SANTE',  #7
            groupe_compagnie if groupe_compagnie else '', #8 -- groupe_compagnie
            c.police.compagnie.nom if hasattr(c.police, 'compagnie') else '', #9
            c.police.numero if hasattr(c.police, 'numero') else '', #10
            programme if programme else '', #11 --
            c.police.date_debut_effet.strftime("%d/%m/%Y") if hasattr(c.police, 'date_debut_effet') else '', #12
            c.police.date_fin_effet.strftime("%d/%m") if hasattr(c.police, 'date_fin_effet') else '', #13
            date_der_mvt_police.strftime("%d/%m/%Y") if date_der_mvt_police else '', #14
            c.police.taux_com_courtage if hasattr(c.police, 'taux_com_courtage') else '', #15
            c.police.taux_com_gestion if hasattr(c.police, 'taux_com_gestion') else '', #16
            c.police.bureau.pays.devise.code if c.police.bureau.pays.devise else '', #17
            c.numero if hasattr(c, 'numero') else '', #18
            c.type_quittance.libelle if hasattr(c, 'type_quittance') else '', #19
            c.date_emission.strftime("%Y-%m") if hasattr(c, 'date_emission') else '', #20
            c.date_debut.strftime("%Y-%m") if hasattr(c, 'date_debut') else '', #21
            date_reglement_client.strftime("%Y-%m") if date_reglement_client else '', #22
            date_reversement_compagnie.strftime("%Y-%m") if date_reversement_compagnie else '', #23
            date_reversement_compagnie.strftime("%d/%m/%Y") if date_reversement_compagnie else '', #24
            # en attendant de voir si les dates sont différentes
            date_encaissement_commission.strftime("%Y-%m") if date_encaissement_commission else '', #25
            c.prime_ht if hasattr(c, 'prime_ht') else '', #26
            round((c.prime_ht * taux_euro), 2) if hasattr(c, 'prime_ht') else '',  # prime_ht_eur #27
            c.prime_ttc if hasattr(c, 'prime_ttc') else '', #28
            round((c.prime_ttc * taux_euro), 2) if hasattr(c, 'prime_ttc') else '',  # prime_ttc_eur #29
            c.commission_courtage if hasattr(c, 'commission_courtage') else '',  # ajouter appport gest #30
            round((c.commission_courtage * taux_euro), 2) if hasattr(c, 'commission_courtage') else '', #31
            # commission_courtage_euro
            c.commission_gestion if hasattr(c, 'commission_gestion') else '',  # ajouter appport gest #32
            round((c.commission_gestion * taux_euro), 2) if hasattr(c, 'commission_gestion') else '', #33
            # commission_gestion_euro
            c.prime_EVCAT if hasattr(c, 'prime_EVCAT') else '',  #34
            c.COM_EVCAT if hasattr(c, 'COM_EVCAT') else '',  #35

            com_court_encaissee if com_court_encaissee else 0, #36
            round((com_court_encaissee * taux_euro), 2) if com_court_encaissee else 0,  # com_encaissee_EUR #37
            com_court_non_encaissee if com_court_non_encaissee else 0, # com_non_encaissee #38
            round((com_court_non_encaissee * taux_euro), 2) if com_court_non_encaissee else 0, #39

            com_gest_encaissee if com_gest_encaissee else 0,  # 40
            round((com_gest_encaissee * taux_euro), 2) if com_gest_encaissee else 0,  # com_encaissee_EUR #41
            com_gest_non_encaissee if com_gest_non_encaissee else 0,  # com_non_encaissee #42
            round((com_gest_non_encaissee * taux_euro), 2) if com_gest_non_encaissee else 0,  # com_non_encaissee_EUR #43

            c.commission_intermediaires if c.commission_intermediaires else 0,  #44 --
            round((c.commission_intermediaires * taux_euro), 2) if c.commission_intermediaires else 0,  #45 --
            retro_paye if retro_paye else 0,  #46 --
            round((retro_paye * taux_euro), 2) if retro_paye else 0,  #47 --

            c.frais_cabinet if hasattr(c, 'frais_cabinet') else '',  #48
            round((c.frais_cabinet * taux_euro), 2) if hasattr(c, 'frais_cabinet') else '',  # frais_cabinet_EUR #49
            c.acc_cie if hasattr(c, 'acc_cie') else '', #50
            round((c.acc_cie * taux_euro), 2) if hasattr(c, 'acc_cie') else '',  # acc_cie_EUR #51
            c.acc_court if hasattr(c, 'acc_court') else '', #52
            round((c.acc_court * taux_euro), 2) if hasattr(c, 'acc_court') else '',  # acc_court_EUR #53
            ca_revers_aa if ca_revers_aa else 0,  # annee de reversement comprise dans la période de couverture, sinon 0 #54
            ca_revers_aamm if ca_revers_aamm else 0, #55
            ##mois et année de reversement comprise dans la période de couverture, sinon 0
            ca_emis_aa if ca_emis_aa else '',  # année d'émission de la quittance (date emission) #56
            ca_emis_aa2 if ca_emis_aa2 else '', #57
            # année d'émission de la quittance (date emission) année/mois  d'émission de la prime sauf si la prime concerne l'année suivante (période de garantie sur l'année suivante) auquel cas l'année suivante
            ca_emis_aamm if ca_emis_aamm else '',  # année mois émission prime #58
            ca_emis_aamm2 if ca_emis_aamm2 else '', #59
            # année/mois  d'émission de la prime sauf si la prime concerne l'année suivante (période de garantie sur l'année suivante) auquel cas l'année suivante
            ca_type if ca_type else '', #60
            # indique si le pays est en compta à l'émission ou au reversement (A paramétrer sur le bureau)
            ca_aa if ca_aa else '0',  # année pour laquelle le ca va compter #61
            ca_aamm if ca_aamm else '0', #62
            # année de comptabilisation du CA selon si on est à l'émission ou au reversement
            # taux_euro if taux_euro else '',  # taux de change (A paramétrer de façon générale pour l'application) #63
            apporteur_police if apporteur_police else '', #64 --
            apporteur_police if apporteur_police else '', #65 --
            # apporteur_police_query.date_effet.strftime("%d/%m/%Y") if apporteur_police_query else '', #66 --
            apporteur_police_query.date_effet.strftime("%d/%m/%Y") if apporteur_police_query and apporteur_police_query.date_effet else '',
            c.apporteur_fin if hasattr(c, 'apporteur_fin') else '', #67 --
            apporteur_police_query.taux_com_renouvellement if apporteur_police_query else 0, #68 -- apporteur_tx_terme
            apporteur_police_query.taux_com_affaire_nouvelle if apporteur_police_query else 0, #69 -- apporteur_tx_cpt
            c.apporteur_mnt_fixe if hasattr(c, 'apporteur_mnt_fixe') else 0, #70 --
        ]
        data.append(data_iten)

    for row_num, row in enumerate(data, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response




def exportation_en_excel_text_brute(title, header, queryset):
    # Exportation excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    # Appliquer un style au titre des colonnes
    bold_font = Font(bold=True)
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = bold_font

    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response


def exportation_en_excel_avec_style(title, header, queryset):
    # Exportation excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SUIVI SP CLIENT PAR FILIALE"
    worksheet.title = title

    # Personnalisation de l'entête
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color='ffebcd', end_color='ffebcd', fill_type='solid')  # Couleur de remplissage pour les entêtes
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = bold_font
        cell.fill = header_fill

    # Remplissage du tableau avec les données
    for row_num, row in enumerate(queryset, 2):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Appliquer un style de tableau (en spécifiant les cellules à inclure dans le tableau)
    table_range = f"A1:{openpyxl.utils.get_column_letter(len(header))}{len(queryset) + 1}"  # Range pour le tableau
    table_style = openpyxl.worksheet.table.Table(displayName="Table11", ref=table_range)
    worksheet.add_table(table_style)

    # Ajuster la largeur des colonnes en fonction du contenu pour faciliter la lecture du tableau
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response

def exportation_en_excel_avec_style_background_task(title, header, queryset):
    # Exportation excel
    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, row in enumerate(queryset, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    return workbook.save(f"{title}.xlsx")

def requete_liste_des_sp_client_par_filiale(code_bureau, sp_a_la_date_du):
    return f"""
    SELECT 
    REQ.nom_cie as "ASSUREUR",
    REQ.nom2_per as "SOUSCRIPTEUR",
    REQ.num_pol as "N° POLICE GARANT",
    REQ.dateEffet as "DEBUT EFFET POLICE",
    REQ.proch_echeance as "ECHEANCE POLICE",
    REQ.mt_regle as "SINISTRE COURANT",
    CASE WHEN (REQ.Period_totale > 0 AND REQ.Period_consommee >0) THEN round((REQ.prime_net * REQ.Period_consommee /     REQ.Period_totale),2) ELSE 0 END as "PRIME TOTALE PERIODE",
    REQ.prime_net as "PRIME TOTALE",
    CASE WHEN (REQ.Period_totale > 0 AND REQ.Period_consommee >0 AND REQ.prime_net>0) THEN round((REQ.mt_regle /        (REQ.prime_net * REQ.Period_consommee / REQ.Period_totale))*100,2) ELSE 0 END as "RAPPORT S/P (%)"
    FROM 
    (
        SELECT
            (
                SELECT 
                round(coalesce(sum(rsin.montant),0),2)
                FROM sinistres sinn
                join polices on polices.id = sinn.police_id 
                join remboursement_sinistre rsin ON rsin.sinistre_id = sinn.id 
                AND rsin.statut = 'ACCEPTE' AND rsin.is_invalid = 0 
                WHERE sinn.bordereau_ordonnancement_id is not NULL 
                AND (sinn.date_survenance between B.dtmvt_pol and '{sp_a_la_date_du}')
                AND sinn.statut_validite = 'VALIDE'
                AND sinn.police_id = B.id_npol
            ) as mt_regle,
            round(coalesce(po.prime_ht * (1-coalesce(po.taux_charge, 0)/100 ),0),2) prime_net,
            B.num_per,
            B.nom2_per,
            B.nom_cie,
            B.id_npol,
            B.num_pol,
            B.id_pol,
            date_format(B.dtmvt_pol,'%d/%m/%Y') as dateEffet,
            date_format(B.dtfin_pol,'%d/%m/%Y') as proch_echeance,
            datediff('{sp_a_la_date_du}',B.dtmvt_pol) as Period_consommee,
            datediff(B.dtfin_pol,B.dtmvt_pol) as Period_totale
        FROM
        (
            select 
                max(pc.id) as id_pol,
                po.numero as num_pol,
                po.id as id_npol,
                pc.date_debut_effet as dtmvt_pol,
                pc.date_fin_effet as dtfin_pol,
                cl.code as num_per,
                concat_ws(' ', cl.nom, cl.prenoms) as nom2_per,
                cie.nom as nom_cie
            from polices po
            join periode_couverture pc on pc.police_id = po.id
            join clients cl on cl.id = po.client_id
            join compagnies cie on cie.id = po.compagnie_id
            join bureau soc on soc.id = po.bureau_id
            join produit pr on pr.id = po.produit_id and pr.code = '100991'
            where pc.date_debut_effet <= CURRENT_DATE()
            and ('{sp_a_la_date_du}' between pc.date_debut_effet and pc.date_fin_effet)
            and po.statut = 'ACTIF'
            and soc.code='{code_bureau}'
            group by 
            po.numero,
            pc.date_debut_effet,
            pc.date_fin_effet,
            cl.code,
            cl.nom, 
            cl.prenoms,
            cie.nom
        ) B 
        join polices po on po.id = B.id_npol
        where datediff(B.dtfin_pol,B.dtmvt_pol) <= 366
    ) REQ
    """

def load_backgroound_request_task():
    tasks = BackgroundQueryTask.objects.filter(status="ENATT").all()
    print("task")
    print(tasks)
    print(len(tasks))
    for task in tasks:
        try:
            # mise a jour du status de la tache
            task.status = "ENCOURS"
            task.save()

            queryset, header = execute_query(task.query)

            print("queryset")
            print(queryset)

            print("header")
            print(header)
            themp_name = datetime.now().strftime("%Y%m%d%H%M%S")
            fichier_excel = exportation_en_excel_avec_style_background_task(themp_name, header, queryset)

            with open(f'{themp_name}.xlsx', 'rb') as f:
                task.file.save("DONNEES_REQUETE_PLATEFORME_V2.xlsx", File(f), save=False)

            #  Suppression du fichier excel thempo
            if os.path.exists(f'{themp_name}.xlsx'):
                os.remove(f'{themp_name}.xlsx')
            # mise a jour du status de la tache
            task.status = "TERMINEE"
            task.error_message = ""
            task.save()

            send_notification_background_task_mail(task.created_by.email, task)

            CronLog.objects.create(action="export", table="background_query_task",
                                   description=f'requête {task.id} | {task.name} effectué avec succès pour {task.created_by.username}').save()

            # self.stdout.write(self.style.SUCCESS(
            #     f'requête {task.id} | {task.name} effectué avec succès pour {task.created_by.username}'))
        except Exception as e:
            print(e)
            try:
                # mise a jour du status de la tache
                task.status = "ECHOUEE"
                task.error_message = str(e)
                task.save()
            except Exception as e:
                print(e)
            send_notification_background_task_mail("a.tissi@inov.africa", task)

    # self.stdout.write(self.style.SUCCESS('Executions requêtes terminé avec succès'))

def requete_liste_paiement_sinistre_sante_entre_deux_dates(code_bureau, date_debut_paiment_sinisre, date_fin_paiment_sinisre, reference_facture, numero_police):
    requete = f"""
        select 
        concat_ws(COALESCE(cl.nom,''), ' ', COALESCE(cl.prenoms,'')) as client
        ,s.numero as numero_dossier
        ,po.numero as numero_police
        ,c.numero as numero_carte
        ,adh.nom as nom_patient
        ,adh.prenoms as prenom_patient
        ,(SELECT libelle from qualite_beneficiaire where id = adh.qualite_beneficiaire_id) as qualite_patient
        ,adh.matricule_employe as matricule
        ,DATE_FORMAT(adh.date_naissance,"%d/%m/%Y") as date_naissance_patient
        ,concat(ap.nom, ' ', ap.prenoms) as adherent_principal
        ,fg.libelle as formule
        ,DATE_FORMAT(s.date_survenance,"%d/%m/%Y") as date_sinistre
        ,s.date_reception_facture
        ,actes.libelle as acte
        ,aff.libelle as affection
        ,aff.code_cim_10 as code_affection
        ,pr.name as prestataire
        ,(select rs.nom from reseaux_soins rs where rs.id = fg.reseau_soin_id) as reseau_soin
        ,(select tp.name from type_prestataires tp where tp.id = pr.type_prestataire_id) as type_prestataires
        ,(select presc.name from prestataires presc where presc.id = ds.centre_prescripteur_id) as centre_prescripteur
        ,concat(medp.nom, ' ', medp.prenoms) as medecin_traitant 
        ,spec.name as specialite_medecin_traitant
        ,null as medecin_prescripteur -- concat(medp.nom, ' ', medp.prenoms) 
        ,null as specialite_medecin_prescripteur -- spec.name
        ,s.frais_reel
        ,s.part_assure
        ,s.part_compagnie as part_inov
        ,s.part_compagnie
        ,(select COALESCE(sum(montant),0) from remboursement_sinistre rs where rs.sinistre_id = s.id and rs.statut = 'TAXT' and  rsin.is_invalid = 0)  as tps_aib_bnc
        ,(select COALESCE(sum(montant),0) from remboursement_sinistre rs where rs.sinistre_id = s.id and rs.statut = 'TAXE FAR' and  rsin.is_invalid = 0)  as far
        ,s.depassement as depassement_exclusion	
        ,(case tpr.id when 1 THEN s.part_assure ELSE NULL END) as ticket_prefinance	
        ,(rsin.montant + (select COALESCE(sum(montant),0) from remboursement_sinistre rs where rs.sinistre_id = s.id and rs.statut = 'TAXT' and  rsin.is_invalid = 0) + (select COALESCE(sum(montant),0) from remboursement_sinistre rs where rs.sinistre_id = s.id and rs.statut = 'TAXE FAR' and  rsin.is_invalid = 0)) as net_regle
        ,DATE_FORMAT(rsin.created_at,"%d/%m/%Y") as date_ord
        ,DATE_FORMAT(pc.date_paiement,"%d/%m/%Y") as date_reg
        ,pc.numero as numero_bordereau
        ,fact_prest.numero as num_facture
        ,case 
            when bo.prestataire_id is not null then pr.code
            when bo.adherent_principal_id is not null then ap.numero
        end as Num_Beneficiaire_du_remb
        ,case 
            when bo.prestataire_id is not null then pr.name
            when bo.adherent_principal_id is not null then concat(ap.nom, ' ', ap.prenoms)
        end as Nom_Beneficiaire_du_remb
        ,cie.nom as compagnie
        ,br.code as code_societe
        ,rub.libelle as regroupement_inov
        ,(select libelle from mode_reglements where mode_reglements.id =  pc.mode_reglement_id) as mode_reglement
        ,pc.numero_piece as numero_piece
        ,bo.numero as numero_bordereau_ordonnancement
    from sinistres s  
    join dossier_sinistre ds on ds.id = s.dossier_sinistre_id
    left join type_prefinancement tpr on tpr.id = s.type_prefinancement_id 
    join prestataires pr on pr.id = s.prestataire_id
    join remboursement_sinistre rsin ON rsin.sinistre_id = s.id AND rsin.statut = 'ACCEPTE' AND rsin.is_invalid = 0
    join polices po on po.id = s.police_id
    join clients cl on cl.id = po.client_id
    join compagnies cie on cie.id = po.compagnie_id
    join formulegarantie fg on fg.id = s.formulegarantie_id
    join aliments adh on adh.id = s.aliment_id
    join cartes c on c.aliment_id = adh.id and c.date_desactivation is null
    join aliments ap on ap.id = adh.adherent_principal_id
    join factures_prestataires fact_prest on fact_prest.id = s.facture_prestataire_id
    join paiement_comptable pc on pc.id = s.paiement_comptable_id
    join bordereau_ordonnancement bo on bo.id = pc.bordereau_ordonnancement_id
    join bureau br on br.id = pc.bureau_id
    left join actes on actes.id = s.acte_id
    left join affections aff on aff.id = s.affection_id
    left join rubriques rub on rub.id = actes.rubrique_id
    left join prescripteur medp on medp.id = s.prescripteur_id
    left join specialite spec on spec.id = medp.specialite_id
    where s.statut_validite = 'VALIDE'  
    and br.code in ('{code_bureau}')
    and ('{numero_police}' = '' or po.numero = '{numero_police}')
    and (pc.date_paiement between '{date_debut_paiment_sinisre}' and '{date_fin_paiment_sinisre}')
    """

    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete 



#
def requete_liste_sinistre_saisies_entre_2date(bureau_id, date_debut, date_fin, reference_facture):
    print("requete_liste_sinistre_saisies_entre_2date @@@@@@@@@@@@@@@@@@@@@@@@@")
    print(bureau_id)
    print(date_debut)
    print(date_fin)
    requete = f"""
    With dataremb as (
	select 
	    s.id
	    ,DATE_FORMAT(pc.date_paiement,"%d/%m/%Y") as date_reg
	    ,pc.numero as numero_bordereau
	    ,case 
	        when bo.prestataire_id is not null then pr.name
	        when bo.adherent_principal_id is not null then concat(coalesce(ap.nom,''), ' ', coalesce(ap.prenoms,''))
	    end as Nom_Beneficiaire_du_remb
	    ,pc.numero_piece as numero_piece
	from sinistres s  
	join prestataires pr on pr.id = s.prestataire_id
	join aliments adh on adh.id = s.aliment_id
	join aliments ap on ap.id = adh.adherent_principal_id
	join bordereau_ordonnancement bo on bo.id = s.bordereau_ordonnancement_id
	join paiement_comptable pc on pc.bordereau_ordonnancement_id = bo.id
	where s.statut_validite = 'VALIDE' and pc.bureau_id={bureau_id}
    )
    ,
    contrat as (
        select 
            p.id,
            p.numero ,
            concat(coalesce(c.nom,''),' ',coalesce(c.prenoms,'')) as nom_client	 
        from polices p 
        join clients c on c.id = p.client_id
        where p.bureau_id={bureau_id}
    )
    select 
	(select contrat.nom_client from contrat where contrat.id = fg.police_id) 'NOM_CLIENT',
	sin.numero 'NUMERO_DOSSIER',
	(select contrat.numero from contrat where contrat.id = fg.police_id) 'NUMERO_POLICE',
	DATE_FORMAT(per_couv.date_debut_effet,"%d/%m/%Y") 'DATE_EFFET',
    DATE_FORMAT(per_couv.date_fin_effet,"%d/%m/%Y") 'DATE_ECHEANCE',
	cgni.nom 'NOM_CIE',
	CONCAT(COALESCE(alm_adh.nom),' ',COALESCE(alm_adh.prenoms)) 'ADHERENT_PRINCIPAL',
	alm_adh.numero_famille 'NUMERO_FAMILLE',
    fg.libelle 'FORMULE', 
    cart.numero 'NUMERO_CARTE',
    alm.nom 'NOM_PATIENT',
    alm.prenoms 'PRENOM_PATIENT',
    DATE_FORMAT(alm.date_affiliation,"%d/%m/%Y") 'DATE_EFFET_PATIENT',
    DATE_FORMAT(alm.date_naissance,"%d/%m/%Y") 'DATE_NAISSANCE_PATIENT',
    qtbf.libelle 'LIEN_PATIENT',
    DATE_FORMAT(sin.date_survenance,"%d/%m/%Y") 'DATE_SINISTRE',
    DATE_FORMAT(sin.created_at,"%d/%m/%Y") 'DATE_SAISIE',
    (select CONCAT(util.last_name,' ',util.first_name) from configurations_user util where util.id = sin.created_by_id) 'SAISI_PAR',
    act.libelle 'ACTE',
    affect.libelle 'LIB_AFFECTION',
    affect.code_cim_10 'CODE_CIMDIS',
    rg_act.libelle 'LIB_REGROUPEMENT',
    CONCAT(prest.code,'|',prest.name) 'PRESTATAIRE',
    sin.frais_reel 'FRAIS_REEL',
    sin.part_compagnie 'PART_INOV',
    sin.part_assure 'PART_ASSURE',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'TPS_AIB_BNC',
    (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) 'FAR',
    sin.depassement 'DEPASSEMENT/EXCLUSION',
    CASE WHEN tpr.code = "PREF_TOUT" THEN sin.part_assure else 0 END 'TICKET_PREFINANCE',
    sin.part_compagnie 'PART_COMPAGNIE',
    # CASE WHEN dataremb.date_reg IS NULL THEN NULL ELSE rsin.montant END 'NET_REGLE',
    CASE WHEN dataremb.date_reg IS NULL THEN NULL ELSE ((select SUM(remb_sin.montant) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='ACCEPTE' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id) + (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=sin.id)) END 'NET_REGLE',
    dataremb.date_reg 'DATE_REGLEMENT',
    dataremb.numero_bordereau 'NUMERO_BORDEREAU',
    dataremb.Nom_Beneficiaire_du_remb 'BENEFICIAIRE_DU_REMB',
    sin.statut 'STATUT_PEC',
    sin.statut_remboursement 'STATUT_DEMAND_REMB',
    (select CONCAT(util.last_name,' ',util.first_name) from configurations_user util where util.id = rsin.created_by_id) 'DEMAND_REMB_VALIDEE_PAR',
	sin.statut_validite 'STATUT_SINISTRE',
    (select DATE_FORMAT(f_pres.created_at,"%d/%m/%Y") from factures_prestataires f_pres where f_pres.id = sin.facture_prestataire_id) 'DATE_RECEPTION_FACTURE',
    sin.reference_facture 'REFERENCE_FACTURE'
    FROM sinistres sin 
    JOIN periode_couverture per_couv on per_couv.id = sin.periode_couverture_id 
    JOIN aliments alm ON alm.id = sin.aliment_id
    JOIN cartes cart on cart.aliment_id=alm.id AND cart.statut='ACTIF'
    JOIN aliments alm_adh ON alm_adh.id=alm.adherent_principal_id
    JOIN formulegarantie fg on fg.id = sin.formulegarantie_id
    JOIN qualite_beneficiaire qtbf on qtbf.id = alm.qualite_beneficiaire_id
    JOIN actes act on act.id = sin.acte_id
    JOIN regroupement_acte rg_act on rg_act.id  = act.regroupement_acte_id
    JOIN prestataires prest on sin.prestataire_id=prest.id
    JOIN compagnies cgni on sin.compagnie_id=cgni.id
    JOIN dossier_sinistre doss_sin on doss_sin.id = sin.dossier_sinistre_id
    LEFT JOIN type_prefinancement tpr on tpr.id = sin.type_prefinancement_id
    LEFT JOIN affections affect on affect.id = sin.affection_id
    LEFT JOIN remboursement_sinistre rsin ON rsin.sinistre_id = sin.id AND rsin.is_invalid = 0 AND rsin.statut = 'ACCEPTE' 
    LEFT JOIN dataremb on sin.id = dataremb.id
    WHERE prest.bureau_id={bureau_id}
    AND sin.dossier_sinistre_id is not null
    AND doss_sin.mode_creation_id=2
    AND DATE(sin.created_at) BETWEEN '{date_debut}' AND '{date_fin}'
    """

    if reference_facture is not None and reference_facture != "":
        return requete + f""" AND sin.reference_facture='{reference_facture}' """
    else:
        return requete


##
def extraction_des_sinistres_traites_valides(bureau_id, date_debut, date_fin):
    print("requete_extraction_des_sinistres_traites_valides")
    print(bureau_id)
    print(date_debut)
    print(date_fin)
    requete = f"""
            SELECT   
            CONCAT_WS(' ', COALESCE(cl.nom, ''), COALESCE(cl.prenoms, '')) AS client,
            s.numero AS numero_dossier,
            po.numero AS numero_police,
            c.numero AS numero_carte,
            adh.nom AS nom_patient,
            adh.prenoms AS prenom_patient,
            (SELECT libelle FROM qualite_beneficiaire WHERE id = adh.qualite_beneficiaire_id) AS qualite_patient,
            DATE_FORMAT(adh.date_naissance, "%d/%m/%Y") AS date_naissance_patient,
            CONCAT(ap.nom, ' ', ap.prenoms) AS adherent_principal,
            DATE_FORMAT(s.date_survenance, "%d/%m/%Y") AS date_sinistre,
            s.date_reception_facture,
            actes.libelle AS acte,
            aff.libelle AS affection,
            aff.code_cim_10 AS code_affection,
            pr.name AS prestataire,
            (SELECT tp.name FROM type_prestataires tp WHERE tp.id = pr.type_prestataire_id) AS type_prestataires,
            (SELECT presc.name FROM prestataires presc WHERE presc.id = ds.centre_prescripteur_id) AS centre_prescripteur,
            CONCAT(medp.nom, ' ', medp.prenoms) AS medecin_traitant,
            spec.name AS specialite_medecin_traitant,
            s.frais_reel,
            s.part_assure,
            s.part_compagnie AS part_inov,
            (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) as tps_aib_bnc,
            (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) as far,
            cie.nom AS compagnie,
            -- br.code AS code_societe,
            CONCAT_WS(' ', COALESCE(cus.first_name, ''), COALESCE(cus.last_name, '')) AS gestionnaire
            
        FROM sinistres s
        JOIN dossier_sinistre ds ON ds.id = s.dossier_sinistre_id
        JOIN actes ON actes.id = s.acte_id
        LEFT JOIN affections aff ON aff.id = s.affection_id
        JOIN prestataires pr ON pr.id = s.prestataire_id
        JOIN remboursement_sinistre rsin ON rsin.sinistre_id = s.id  AND rsin.is_invalid = 0 AND rsin.statut='ACCEPTE'
        JOIN polices po ON po.id = s.police_id
        JOIN clients cl ON cl.id = po.client_id
        JOIN compagnies cie ON cie.id = po.compagnie_id
        JOIN aliments adh ON adh.id = s.aliment_id
        JOIN cartes c ON c.aliment_id = adh.id AND c.date_desactivation IS NULL
        JOIN aliments ap ON ap.id = adh.adherent_principal_id
        LEFT JOIN prescripteur medp ON medp.id = s.prescripteur_id
        LEFT JOIN specialite spec ON spec.id = medp.specialite_id
        JOIN configurations_user cus ON cus.id = rsin.created_by_id  and cus.bureau_id={bureau_id}
        
        WHERE pr.bureau_id = {bureau_id} AND s.statut_validite = 'VALIDE'
        AND (DATE(s.date_survenance) BETWEEN '{date_debut}' AND '{date_fin}');
    """

    return requete

##
def extraction_demandes_accords_prealables_traitees_par_medecins_conseil(bureau_id, date_debut, date_fin):
    print("requete_extraction_demandes_accords_prealables_traitees_par_medecins_conseil")
    print(bureau_id)
    print(date_debut)
    print(date_fin)
    requete = f"""
    SELECT   
        CONCAT_WS(' ', COALESCE(cl.nom, ''), COALESCE(cl.prenoms, '')) AS client,
        s.numero AS numero_dossier,
        po.numero AS numero_police,
        c.numero AS numero_carte,
        adh.nom AS nom_patient,
        adh.prenoms AS prenom_patient,
        (SELECT libelle FROM qualite_beneficiaire WHERE id = adh.qualite_beneficiaire_id) AS qualite_patient,
        DATE_FORMAT(adh.date_naissance, "%d/%m/%Y") AS date_naissance_patient,
        CONCAT(ap.nom, ' ', ap.prenoms) AS adherent_principal,
        DATE_FORMAT(s.date_survenance, "%d/%m/%Y") AS date_sinistre,
        actes.libelle AS acte,
        aff.libelle AS affection,
        aff.code_cim_10 AS code_affection,
        pr.name AS prestataire,
        (SELECT tp.name FROM type_prestataires tp WHERE tp.id = pr.type_prestataire_id) AS type_prestataires,
        (SELECT presc.name FROM prestataires presc WHERE presc.id = ds.centre_prescripteur_id) AS centre_prescripteur,
        CONCAT(medp.nom, ' ', medp.prenoms) AS medecin_traitant,
        spec.name AS specialite_medecin_traitant,
        s.frais_reel,
        s.part_assure,
        s.part_compagnie AS part_inov,
        (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXT' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) as tps_aib_bnc,
        (select COALESCE(SUM(remb_sin.montant),0) FROM remboursement_sinistre remb_sin WHERE remb_sin.statut='TAXE FAR' AND remb_sin.is_invalid=0 and remb_sin.sinistre_id=s.id) as far,
        cie.nom AS compagnie,
        CONCAT_WS(' ', cus.first_name, cus.last_name) AS medecin_conseil,
        DATE_FORMAT(s.created_at, "%d/%m/%Y") AS date_demande,
        DATE_FORMAT(s.created_at, "%H:%i") AS heure_demande,
        DATE_FORMAT(s.reviewed_at, "%d/%m/%Y") AS date_traitement,
        DATE_FORMAT(s.reviewed_at, "%H:%i") AS heure_traitement

    FROM sinistres s
    JOIN dossier_sinistre ds ON ds.id = s.dossier_sinistre_id
    JOIN actes ON actes.id = s.acte_id
    LEFT JOIN affections aff ON aff.id = s.affection_id
    JOIN prestataires pr ON pr.id = s.prestataire_id
    JOIN polices po ON po.id = s.police_id
    JOIN clients cl ON cl.id = po.client_id
    JOIN compagnies cie ON cie.id = po.compagnie_id
    JOIN bureau br ON br.id = pr.bureau_id
    JOIN aliments adh ON adh.id = s.aliment_id
    JOIN cartes c ON c.aliment_id = adh.id AND c.date_desactivation IS NULL
    JOIN aliments ap ON ap.id = adh.adherent_principal_id
    LEFT JOIN prescripteur medp ON medp.id = s.prescripteur_id
    LEFT JOIN specialite spec ON spec.id = medp.specialite_id
    LEFT JOIN configurations_user cus ON cus.id = s.approuved_by_id

    WHERE s.approuved_by_id IS NOT NULL 
        AND br.id = {bureau_id}
        AND s.statut_validite = 'VALIDE' 
        AND (DATE(s.date_survenance) BETWEEN '{date_debut}' AND '{date_fin}');
    """

    return requete



#retourne la retenue selon contexte
def get_retenue_selon_contexte(prestataire_id):
    prestataire = Prestataire.objects.filter(pk=prestataire_id).first()
    # taux_retenues = sum(retenue.taux or 0 for retenue in Retenue.objects.filter(bureau=prestataire.bureau).filter(Q(secteur=prestataire.secteur) | Q(secteur__isnull=True)).filter(Q(type_prestataire__contains=prestataire.type_prestataire)) | Q(type_prestataire__isnull=True)) if prestataire else 0
    
    if not prestataire:
        return 0
    
    retenues = Retenue.objects.filter(
        bureau=prestataire.bureau
    ).filter(
        Q(secteur=prestataire.secteur) | Q(secteur__isnull=True)
    ).filter(
        Q(type_prestataire=prestataire.type_prestataire) | Q(type_prestataire__isnull=True)
    )
    
    taux_retenues = sum(retenue.taux or 0 for retenue in retenues)

    return (taux_retenues*(-1)) if taux_retenues > 0 else taux_retenues
