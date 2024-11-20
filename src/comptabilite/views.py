from ast import literal_eval
import calendar
import datetime
from collections import defaultdict
from decimal import Decimal
from functools import reduce
import locale
import os
from pathlib import Path
from pprint import pprint
from sqlite3 import Date

import PyPDF2
from django.urls import reverse
import openpyxl
from django.contrib import admin
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.http import FileResponse, Http404, JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from api.api_helper import send_cron_state_mail
from babel.numbers import format_number, NumberFormatError
from django.core.mail import send_mail

from django.core.paginator import Paginator

from inov import settings
from shared.enum import StatutFacture, StatutSinistre, TypeAlerte

from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django_dump_die.middleware import dd
from django.db import transaction

from comptabilite.models import BordereauOrdonnance, CompteComptable, EncaissementCommission, Journal, ReglementReverseCompagnie
from configurations.helper_config import create_query_background_task, execute_query
from configurations.models import Bureau, Caution, Compagnie, MailingList, NatureOperation, Devise, ModeReglement, Banque, PeriodeComptable, \
    CompteTresorerie, ActionLog, TypeRemboursement, Prestataire
from configurations.models import Compagnie, NatureOperation, Devise, ModeReglement, Banque, PeriodeComptable, \
    CompteTresorerie, ActionLog, TypeRemboursement, Prestataire, ModelLettreCheque, \
    BordereauLettreCheque
from production.models import Aliment, Reglement, Police, Quittance, Operation, OperationReglement, MouvementPolice, \
    Client
from production.templatetags.my_filters import money_field
from shared.enum import MoyenPaiement, SatutBordereauDossierSinistres, StatutPaiementSinistre, \
    StatutReversementCompagnie, StatutEncaissementCommission, StatutReglementApporteurs, \
    StatutValidite, Statut
from shared.helpers import generate_random_string, render_pdf
from shared.veos import get_taux_euro_by_devise
from sinistre.helper_sinistre import requete_analyse_prime_compta
from sinistre.models import FactureCompagnie, FacturePrestataire, ReglementCompagnie, Sinistre, BordereauOrdonnancement, PaiementComptable
from sinistre.helper_sinistre import requete_analyse_prime_compta, requete_analyse_prime_compta_apporteur
from sinistre.models import FacturePrestataire, Sinistre, BordereauOrdonnancement, PaiementComptable
from sinistre.helper_sinistre import requete_analyse_prime_compta
from sinistre.models import FactureCompagnie, FacturePrestataire, ReglementCompagnie, Sinistre, BordereauOrdonnancement, PaiementComptable
from sinistre.helper_sinistre import exportation_en_excel_avec_style, requete_analyse_prime_compta, requete_liste_paiement_sinistre_sante_entre_deux_dates
from sinistre.models import FacturePrestataire, Sinistre, BordereauOrdonnancement

from configurations.models import Compagnie, User
from django.db.models import OuterRef, Subquery, Count, Sum, Min, Max

from django.core.files.storage import FileSystemStorage

from django.db.models import Sum, DateTimeField
from django.db.models.functions import TruncMonth
# If no file exists, create a new PDF
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


#locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
# Create your views here.


def uploaded_file_url(file):
    # Simulation de l'upload du fichier sur le serveur
    fs = FileSystemStorage()
    file_path = fs.save(file.name, file)
    return fs.url(file_path)


@method_decorator(login_required, name='dispatch')
class BordereauxOrdonnancesView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/bordereaux_ordonnances.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.all()

        prestataire_ids = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE).values_list('prestataire_id', flat=True)
        prestataires = Prestataire.objects.filter(id__in=prestataire_ids).order_by('name')

        adherent_principal_ids = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE, type_remboursement__code="RD").values_list('adherent_principal_id', flat=True)
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')

        assures_ids = BordereauOrdonnancement.objects.filter(assure_id__isnull=False, bureau=request.user.bureau,statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE, type_remboursement__code="RD").values_list('assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()
        context = self.get_context_data(**kwargs)
        
        # MONTANT TOTAL BDX ORDONANCES
        total_amount = BordereauOrdonnancement.objects.filter(
            bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.ORDONNANCE,
            statut_validite=StatutValidite.VALIDE
        ).aggregate(total=Sum('montant_accepte_total'))['total'] or 0
        #
        context['total_amount'] = money_field(total_amount)

        # context['dossiers_sinistres'] = dossiers_sinistres
        context['prestataires'] = prestataires
        context['assures'] = assures
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        context['type_remboursements'] = type_remboursements
        context['adhs'] = adhs

        today = timezone.now()
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

######### new fonction ######### 
def export_bordereaux_ordonnances(request):
    try:
        # Retrieve the BordereauOrdonnancement instances based on filters and user's bureau
        queryset = BordereauOrdonnancement.objects.filter(
            bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.ORDONNANCE,  # Filter by statut_paiement = 'ORDONNANCE'
        ).order_by('-id')

        # Create an Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="LISTE DES BORDEREAUX ORDONNANC√âS.xlsx"'

        # Initialize the workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'BORDEREAUX'

        # Write header row
        header = [
            'N¬∞ BORDEREAU',
            'TYPE REMBOURSEMENT',
            'BENEFICIAIRES',
            'P√âRIODE COMPTABLE',
            'MONTANT ACCEPT√â TOTAL',
            'MONTANT REJET PAY√â',
            'MONTANT NET A PAY√â',
            'DATE DE CR√âATION',
            'STATUT ORDONNANCE '
        ]
        for col_num, column_title in enumerate(header, 1):
            worksheet.cell(row=1, column=col_num).value = column_title

        # Write data rows
        for row_num, bordereau in enumerate(queryset, 2):
            # Retrieve related data
            type_remboursement = bordereau.type_remboursement.libelle if bordereau.type_remboursement else ""
            periode_comptable = bordereau.periode_comptable.libelle if bordereau.periode_comptable else ""

            row = [
                bordereau.numero,  # Bordereau numero
                type_remboursement,  # Type Remboursement
                bordereau.ordre_de,  # Beneficiaries (or other related field if needed)
                periode_comptable,  # Periode Comptable
                bordereau.montant_remb_total,  # Montant total accepte
                bordereau.montant_rejet_total,  # Montant total rejet
                bordereau.montant_accepte_total,  # Montant total impaye
                bordereau.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                bordereau.statut_paiement,  # Statut Paiement

            ]
            
            # Write each cell value to the worksheet
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value

        # Save the workbook to the response
        workbook.save(response)
        return response

    except Exception as e:
        return JsonResponse({"message": str(e)}, status=500)
    
###

def bordereaux_ordonnances_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')


    queryset = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.ORDONNANCE, statut_validite=StatutValidite.VALIDE).order_by('-id')

    # la recherche
    if search_type_remboursement:
        queryset = queryset.filter(type_remboursement__code=search_type_remboursement)

    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_assure:
        queryset = queryset.filter(assure__id=search_assure)

    if search_adherent_principal:
        queryset = queryset.filter(adherent_principal__id=search_adherent_principal)
        
    # FILTRED MONTANT TOTAL DES BDX ORDONANCES
    total_amount_filtered = queryset.aggregate(total=Sum('montant_accepte_total'))['total'] or 0

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'prestataire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_bordereau_ordonnancement', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> D√©tails</span></a>'

        # nom_prestataire = c.prestataire.name if c.prestataire else f'{c.adherent_principal.nom} {c.adherent_principal.prenoms} ({c.adherent_principal.carte_active()})'

        if c.type_remboursement.code == "TP":
            nom_prestataire = c.prestataire.name if c.prestataire else ""
        elif c.type_remboursement.code == "RD" :
            if c.assure:
                nom_prestataire = (c.assure.nom + " " + (c.assure.prenoms if c.assure.prenoms else "")) if c.assure else ""
            else:
                nom_prestataire = f'{c.adherent_principal.nom} {c.adherent_principal.prenoms} ({c.adherent_principal.carte_active()})' if c.adherent_principal else ""


        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_remboursement": c.type_remboursement.code,
            "prestataire__name": nom_prestataire,
            "periode_comptable": periode_comptable_libelle,
            "rembourse": f'<div style="text-align:right;">{money_field(c.montant_accepte_total)}</div>',
            "rejet": f'<div style="text-align:right;">{money_field(c.montant_rejet_total)}</div>',
            "net_a_payer": f'<div style="text-align:right;">{money_field(c.montant_remb_total)}</div>',
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
        #
        "total_amount_filtered": money_field(total_amount_filtered)
    })


@method_decorator(login_required, name='dispatch')
class InitialisationFondRoulementView(TemplateView):
    template_name = 'comptabilite/initialisation_fonds_de_roulements.html'

    def get_context_data(self, *args, **kwargs):
        bureau = self.request.user.bureau
        cautions = Caution.par_bureau(bureau).filter(date_fin_effet__isnull=True, status=True)

        # Liste des garants sans caution (pour le modal et en ajouter)
        garants_sans_caution = Compagnie.par_bureau(bureau).annotate(
            caution_exists=Subquery(
                Caution.objects.filter(compagnie_id=OuterRef('pk'), date_fin_effet__isnull=True).values('compagnie_id')
            )
        ).filter(caution_exists__isnull=True)

        context_data = {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            'cautions': cautions,
            'garants_sans_caution' : garants_sans_caution,
        }

        return context_data

def get_fdr_data(request):
    # Vue appel√©e via AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        bureau = request.user.bureau
        cautions = Caution.par_bureau(bureau).filter(date_fin_effet__isnull=True, status=True)

        # Effectuez les calculs ou les requ√™tes n√©cessaires
        total_global_caution = cautions.aggregate(total_montant=Sum('montant'))['total_montant'] or 0

        total_global_en_attente_de_reglement = FactureCompagnie.objects.filter(
            bureau=bureau,
            statut=StatutFacture.NON_SOLDE
        ).aggregate(total_montant_restant=Sum('montant_restant'))['total_montant_restant'] or 0

        # total_global_a_refacturer = BordereauOrdonnancement.objects.filter(
        total_global_a_refacturer_global = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.PAYE,
            statut_validite=StatutValidite.VALIDE,
            dossier_sinistre_id__isnull=False,
            facture_compagnie_id__isnull=True,
        ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0

        data = {
            'total_global_caution': money_field(total_global_caution),
            'total_global_en_attente_de_reglement': money_field(total_global_en_attente_de_reglement),
            'total_global_a_refacturer': money_field(total_global_a_refacturer_global),
        }
        return JsonResponse(data)
    else:
        return render(request, 'comptabilite/initialisation_fonds_de_roulements.html')


def add_mise_en_initialiser_fdr_garant(request):
    if request.method == 'POST':
        compagnie_id = request.POST.get('compagnie_id')

        saved_caution = save_caution(request=request, compagnie_id=compagnie_id)
        if saved_caution:

            return JsonResponse(
                {
                    'statut': 1,
                    'message': 'Caution mise √† jour avec succ√®s !',
                }
            )
        else:
            return JsonResponse(
                {
                    'statut': 0,
                    'message': 'Veuillez fournir toutes les donn√©es requises'
                }
            )
    else:
        return JsonResponse(
            {
                'statut': 0,
                'message': 'M√©thode non autoris√©e'
            }
        )


def save_caution(request, compagnie_id):
    bureau = request.user.bureau
    compagnie = Compagnie.objects.get(id=compagnie_id)
    montant = request.POST.get('montant')
    # Nettoyer la valeur du montant
    montant = int(montant.replace(" ", ""))

    date_debut_effet_str = request.POST.get('date_debut_effet')
    date_debut_effet = datetime.datetime.strptime(date_debut_effet_str, '%Y-%m-%d').date()

    # date_fin_effet = date_debut_effet - datetime.timedelta(days=1) + heure 23h59
    date_fin_effet = datetime.datetime.combine(date_debut_effet - datetime.timedelta(days=1), datetime.time(23, 59, 59))

    # V√©rification des donn√©es requises
    if compagnie and montant and date_debut_effet:

        # Si la compagnie a d√©j√† une ligne dans la table de caution
        existing_caution = Caution.par_bureau(bureau).filter(compagnie=compagnie, date_fin_effet__isnull=True,
                                                             status=True).first()

        # Si la compagnie √† d√©j√† une ligne dans la table caution set date_fin et RENEW
        if existing_caution:

            # Mettre date fin  = request.POST.get('date_debut_effet') -1j
            existing_caution.date_fin_effet = date_fin_effet
            existing_caution.status = False
            existing_caution.save();

            new_caution = Caution.objects.create(
                bureau=compagnie.bureau,
                compagnie=compagnie,
                montant=montant,
                date_debut_effet=date_debut_effet,
                created_by=request.user,
                status=True
            )

            return new_caution

        else:
            # Cr√©er et enregistrer une instance de Caution dans la base de donn√©es
            new_caution = Caution.objects.create(
                bureau=compagnie.bureau,
                compagnie=compagnie,
                montant=montant,
                date_debut_effet=date_debut_effet,
                created_by=request.user,
                status=True
            )

            return new_caution
    else:
        return False


def update_caution_garant(request, garant_id):
    if request.method == 'POST':
        bureau = request.user.bureau
        compagnie = Compagnie.objects.first(id=garant_id)
        montant = request.POST.get('montant')
        updated_at = request.POST.get('updated_at')

        # Nettoyer la valeur du montant
        montant = int(montant.replace(" ", ""))

        # V√©rification des donn√©es requises
        if compagnie and montant and updated_at:

            # Si la compagnie a d√©j√† une ligne dans la table de caution
            existing_caution = Caution.par_bureau(bureau).filter(compagnie=compagnie, date_fin_effet__isnull=True, status=True).first()



            if not existing_caution:
                return JsonResponse(
                    {
                        'statut': 1,
                        'message': 'Aucune caution n\'existe pour cette compagnie pour √™tre modifier'
                    }
                )
            else:
                # Cr√©er et enregistrer une instance de Caution dans la base de donn√©es
                caution = Caution.objects.create(
                    bureau=compagnie.bureau,
                    compagnie=compagnie,
                    montant=montant,
                    updated_at=updated_at,
                    created_by=request.user  # l'utilisateur actuel
                )

                return JsonResponse(
                    {
                        'statut': 1,
                        'message': 'Caution actualis√© avec succ√®s !',
                    }
                )
        else:
            return JsonResponse(
                {
                    'statut': 0,
                    'message': 'Veuillez fournir toutes les donn√©es requises'
                }
            )
    else:
        return JsonResponse(
            {
                'statut': 0,
                'message': 'M√©thode non autoris√©e'
            }
        )

def edition_caution_compagnie(request, compagnie_id):
    compagnie = Compagnie.objects.get(id=compagnie_id)

    with transaction.atomic():
        if request.method == 'POST':

            saved_caution = save_caution(request=request, compagnie_id=compagnie_id)

            if saved_caution:

                return JsonResponse(
                    {
                        'statut': 1,
                        'message': 'Caution mise √† jour avec succ√®s !',
                    }
                )
            else:
                return JsonResponse(
                    {
                        'statut': 0,
                        'message': 'Veuillez fournir toutes les donn√©es requises'
                    }
                )
        else:
            bureau = request.user.bureau
            compagnie = Compagnie.objects.get(id=compagnie_id)
            caution = Caution.par_bureau(bureau).filter(compagnie=compagnie, date_fin_effet__isnull=True, status=True).first()

            return render(request, 'modals/editer-fdr-garant.html', {'compagnie': compagnie, 'caution': caution, 'now':timezone.now().date()})


#Script / Not for interface
def init_fonds_de_roulements(request):
    user = request.user
    bureau = request.user.bureau

    garants = Compagnie.par_bureau(bureau)

    for garant in garants:

        caution_exists = Caution.objects.filter(compagnie=garant).exists()

        # SET FDR TO 0 IF NOT INIT
        if not caution_exists:
            new_caution = Caution.objects.create(
                    bureau=bureau,
                    compagnie=garant,
                    montant=0,
                    created_by=user,
                    date_debut_effet=datetime.datetime.now(tz=timezone.utc)
                )
            # pprint("Nouvelle caution set 0 => " + str(new_caution.compagnie))

    return JsonResponse(
        {
            'statut': 1,
            'message': 'Fonds de roulement initialis√© avec succ√®s !'
        }
    )





def add_mise_en_reglement_factures_garant(request):
    return


method_decorator(login_required, name="dispatch")
class RefacturationAssureurView(TemplateView):
    template_name = 'comptabilite/refacturation-assureur.html'

    def get_context_data(self, **kwargs):
        bureau = self.request.user.bureau

        # R√©cup√©rer les compagnies avec des sinistres facturables
        '''garants = Compagnie.objects.filter(bureau=bureau,
            sinistre__in=Sinistre.objects.filter(
                statut=StatutSinistre.ACCORDE,
                statut_paiement=StatutPaiementSinistre.PAYE,
                statut_validite=StatutValidite.VALIDE,
                facture_compagnie__isnull=True,
                dossier_sinistre__isnull=False,
            )
        ).distinct()
        '''

        # Tenir compte du bureau du prestataire
        # R√©cup√©rer les compagnies avec des sinistres facturables
        compagnie_ids = Sinistre.objects.filter(prestataire__bureau=bureau,
                                               statut=StatutSinistre.ACCORDE,
                                               statut_paiement=StatutPaiementSinistre.PAYE,
                                               statut_validite=StatutValidite.VALIDE,
                                               facture_compagnie__isnull=True,
                                               dossier_sinistre__isnull=False,
                                           ).distinct().values_list('compagnie_id', flat=True)

        garants = Compagnie.objects.filter(id__in=compagnie_ids, bureau=bureau)

        type_remboursements = TypeRemboursement.objects.all()



        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            'garants': garants,
            'type_remboursements': type_remboursements,
        }


@login_required
def generate_facture_assureur_datatable(request):
    bureau = request.user.bureau
    items_per_page = 100
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = request.GET.get('order[0][column]')
    sort_direction = request.GET.get('order[0][dir]')
    select_all = request.GET.get('all', '')


    search_garant = request.GET.get('search_garant', '')
    search_date_debut_paiement = request.GET.get('search_date_debut_paiement', '')
    search_date_fin_paiement = request.GET.get('search_date_fin_paiement', '')


    pprint("search_date_fin_paiement")
    pprint(search_date_fin_paiement)

    search = request.GET.get('search[value]', '')

    queryset = Sinistre.par_bureau(bureau).filter(
        statut=StatutSinistre.ACCORDE,
        statut_paiement=StatutPaiementSinistre.PAYE,
        statut_validite=StatutValidite.VALIDE,
        facture_compagnie__isnull=True,
        dossier_sinistre__isnull=False, #pour ne pas g√©n√©rer de factures pour les sinistres de la v1
    )

    if search_garant:
        garant = Compagnie.objects.get(id=search_garant)

        # Filter par garant
        queryset = queryset.filter(compagnie=garant)

    if search_date_debut_paiement:

        # Filter par date debut paiement
        #1==1
        queryset = queryset.filter(date_paiement__gte=search_date_debut_paiement)

    if search_date_fin_paiement:

        # Filter par date fin paiement
        #1==1
        queryset = queryset.filter(date_paiement__lte=search_date_fin_paiement)


    # On ne peut pas acceder au property par instance comme queryset.aggregate(total_a_refacturer=Sum('montant_remb_accepte'))
    total_a_refacturer = sum(
        sinistre.montant_remb_accepte for sinistre in queryset
    )

        # if search:
        #     queryset = queryset.filter(
        #         Q(numero__icontains=search) |
        #         Q(dossier_sinistre__numero__icontains=search) |
        #         Q(prestataire__name__icontains=search) |
        #         Q(aliment__nom__icontains=search) |
        #         Q(aliment__prenoms__icontains=search) |
        #         Q(aliment__cartes__numero__icontains=search) |
        #         Q(acte__libelle__icontains=search)
        #     )


    if select_all:
        data_ids = [x.id for x in queryset]
        print(data_ids)
        print(len(data_ids))
        return JsonResponse({
            "data": data_ids,
        })

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_dossier_sinistre_url = reverse('details_dossier_sinistre',
                                              args=[
                                                  c.dossier_sinistre.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_dossier_sinistre_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> D√©tails</span></a>&nbsp;&nbsp;'

        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        statut_html = f'<span class="badge badge-{c.statut.lower()}">{c.statut}</span>'

        # total_facture = c.total_frais_reel if c.total_frais_reel else 0
        # total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        # total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie

        total_frais_reel = c.total_frais_reel
        total_part_compagnie = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        # checkbox = f'<input type="checkbox" name="sinistres_ids[]" value="{c.id}" data-sinistre_id="{c.id}" class ="select-row sinistres">'

        data_iten = {
            "id": c.id,
            "numero_sinistre": c.numero,
            "acte": c.acte.libelle,
            "beneficiaire": c.aliment.nom + ' ' + c.aliment.prenoms,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "total_part_compagnie": money_field(total_part_compagnie),
            "rejete": money_field(c.montant_remboursement_refuse),
            "net_paye": money_field(c.montant_remb_accepte),
            "date_paiement": c.date_paiement.strftime("%d/%m/%Y") if c.date_paiement else '',
            "statut": statut_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "total_a_refacturer": total_a_refacturer,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })
    
    
    # NB: Cette fonction ne fait que regenerer le fichier facture lui-m√™me. elle ne peut pas remplacer la fonction de generer facture qui stocke des calculs.
def generate_facture_compagnie_pdf(request, facture_compagnie_id):
    
    if request.method == "POST":
        devise = request.user.bureau.pays.devise.code
        facture_compagnie = FactureCompagnie.objects.filter(id=facture_compagnie_id).get()
        sinistres = Sinistre.objects.filter(facture_compagnie=facture_compagnie)
        
        # sinistres_grouped = defaultdict(list)
        sinistres_grouped = defaultdict(lambda: {
            'sinistres': [],
            'total_frais_reel': 0,
            'total_part_assure': 0,
            'total_base_remboursement': 0,
            'total_rejete': 0,
            'total_net_a_payer': 0,
            'nbre_sinistres':0,
        })
        
        
        # for sinistre in sinistres:
        #     souscripteur = sinistre.police.client  # This should be the key
        #     sinistres_grouped[souscripteur].append(sinistre)
            
        for sinistre in sinistres:
            souscripteur = sinistre.police.client
            sinistres_grouped[souscripteur]['sinistres'].append(sinistre)
            
            sinistres_grouped[souscripteur]['total_frais_reel'] += sinistre.total_frais_reel or 0
            sinistres_grouped[souscripteur]['total_part_assure'] += sinistre.total_part_assure or 0
            sinistres_grouped[souscripteur]['total_base_remboursement'] += sinistre.total_part_compagnie or 0
            sinistres_grouped[souscripteur]['total_rejete'] += sinistre.montant_remb_refuse or 0
            sinistres_grouped[souscripteur]['total_net_a_payer'] += sinistre.montant_remboursement_accepte or 0
            sinistres_grouped[souscripteur]['nbre_sinistres'] += 1

            
        grouped_items = list(sinistres_grouped.items())
        pprint("-----GROUPED ITEMS-----")
        pprint(grouped_items)
        pprint("-----GROUPED ITEMS-----")
        
                
        # Obtenir les dates les plus anciennes et les plus r√©centes
        dates = sinistres.aggregate(
            date_debut=Min('date_paiement'),
            date_fin=Max('date_paiement')
        )
        date_debut = dates['date_debut']
        date_fin = dates['date_fin']
        periode = "MM YYYY"
        # V√©rifier si les dates sont pr√©sentes
        if date_debut and date_fin :

            # Extraire le mois et l'ann√©e
            mois_debut = date_debut.strftime("%B")  # Nom complet du mois
            mois_fin = date_fin.strftime("%B")
            annee_debut = date_debut.strftime("%Y")
            annee_fin = date_fin.strftime("%Y")


            # V√©rification si le mois de d√©but est diff√©rent du mois de fin
            if date_debut.month != date_fin.month or date_debut.year != date_fin.year:
                periode = f"{mois_debut} - {mois_fin} {annee_fin}"
            else:
                periode = f"{mois_debut} {annee_fin}"

        nombre_total_des_sinistres = sinistres.count()
        montant_total_des_sinistres = sinistres.aggregate(total=Sum('montant_remboursement_accepte'))['total'] or 0


        total_frais_reel = sum(s.total_frais_reel for s in sinistres) or 0
        total_global_part_assure = sum(s.total_part_assure for s in sinistres) or 0
        total_global_part_compagnie = sum(s.total_part_compagnie for s in sinistres) or 0
        total_global_base_remboursement = sum(s.montant_remb_accepte for s in sinistres) or 0
        total_global_rejete = sum(s.montant_remb_refuse for s in sinistres) or 0
        total_global_net_a_payer = sum(s.montant_remb_accepte for s in sinistres) or 0



        facture_data = {
            'date_debut': dates['date_debut'],
            'date_fin': dates['date_fin'],
            'periode': periode,
            'sinistres': sinistres,
            'nombre_total_des_sinistres': nombre_total_des_sinistres,
            'montant_total_des_sinistres': montant_total_des_sinistres,
            'total_frais_reel': total_frais_reel,
            'total_global_part_assure': total_global_part_assure,
            'total_global_part_compagnie': total_global_part_compagnie,
            'total_global_base_remboursement': total_global_base_remboursement,
            'total_global_rejete': total_global_rejete,
            'total_global_net_a_payer': total_global_net_a_payer,
            'grouped_items': grouped_items
        }
        

        context_dict = {
            'facture_data': facture_data,
            'facture_compagnie': facture_compagnie,
            # 'sinistres': sinistres_pdf,
            'montant_total_des_sinistres': montant_total_des_sinistres,
            'bureau': request.user.bureau,
            'devise': devise
        }
        
        # generer le pdf
        pdf = render_pdf("courriers/facture_compagnie_avec_liste_sinistres.html", context_dict)

        pdf_file = PyPDF2.PdfReader(pdf)
        nombre_pages = len(pdf_file.pages)

        #ajout du nombre de page obtenu au contexte pour le rendu final
        context_dict['nombre_pages'] = nombre_pages
        pdf = render_pdf('courriers/facture_compagnie_avec_liste_sinistres.html', context_dict)
        
        # Enregistrement du bordereau
        facture_compagnie.fichier.save(f'facture_compagnie_{facture_compagnie.numero}.pdf', File(pdf))
        facture_compagnie.save()
        
        return facture_compagnie
    
        
    

@transaction.atomic
def submit_generate_facture_assureur(request):
    # TODO : CODE_PAYS_YY_MM
    # try:
    with transaction.atomic():

        compagnie = Compagnie.objects.get(id=int(request.POST.get('search_garant')))
        devise = request.user.bureau.pays.devise.code

        sinistres_payes_ids = literal_eval(request.POST.get('selectedItems'))
        # print("sinistres_payes_ids")
        # print(sinistres_payes_ids)

        search_garant = request.GET.get('search_garant', '')
        search_date_debut_paiement = request.POST.get('search_date_debut_paiement', '')
        search_date_fin_paiement = request.POST.get('search_date_fin_paiement', '')

        sinistres = Sinistre.objects.filter(id__in=sinistres_payes_ids)
        # pprint(sinistres)

        if sinistres and compagnie:

            # numero_facture =
            # Total des montants de remboursement accept√©s pour les sinistres s√©lectionn√©s avec 2 chiffres avec la vigule
            montant_total = round(sinistres.aggregate(Sum('montant_remboursement_accepte'))['montant_remboursement_accepte__sum'], 2)
            montant_total_int = int(montant_total)

            # Insertion dans FactureCompagnie
            facture_compagnie = FactureCompagnie.objects.create(
                bureau=request.user.bureau,
                created_by=request.user,
                compagnie=compagnie,
                # devise=None,
                # numero=numero_facture, #Generer auto
                montant_total=montant_total_int,
                # montant_regle="compagnie",
                montant_restant=montant_total_int,
                date_emission=timezone.now().date(),
                statut=StatutFacture.NON_SOLDE,
                statut_validite=StatutValidite.VALIDE,
                observation="Cr√©√© par interface le " + str(timezone.now().date()),
            )

            sinistres_pdf = []

            details_fac = []

            # Topper tous les sinistres concern√©es comme rentr√©e en facturation
            for sinistre in sinistres: # Mise √† jour des sinistres
                sinistre.facture_compagnie = facture_compagnie
                sinistre.save()

            
            # generation de la facture qui prends
            # facture_compagnie_id permet de recuperer les sinistres concernees.
            facture_compagnie = generate_facture_compagnie_pdf(request=request, facture_compagnie_id=facture_compagnie.id)

            return JsonResponse(
                {'statut': 1, 'message': 'Facture g√©n√©r√© avec succ√®s', 'facture_pdf': facture_compagnie.fichier.url,
                 'facture_id': facture_compagnie.pk}, status=200)


        else:
            return JsonResponse({'statut': 0, 'message': 'Aucun sinistre s√©lectionn√©'})

def get_refacturation_assureur_data(request):
    # Vue appel√©e via AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        bureau = request.user.bureau
        cautions = Caution.par_bureau(bureau).filter(date_fin_effet__isnull=True, status=True)
        # Effectuez les calculs ou les requ√™tes n√©cessaires
        total_global_caution = cautions.aggregate(total_montant=Sum('montant'))['total_montant'] or 0
        total_global_en_attente_de_reglement = FactureCompagnie.objects.filter(
            bureau=bureau,
            statut=StatutFacture.NON_SOLDE
        ).aggregate(total_montant_restant=Sum('montant_restant'))['total_montant_restant'] or 0

        # total_global_a_refacturer = BordereauOrdonnancement.objects.filter(
        total_global_a_refacturer_global = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.PAYE,
            statut_validite=StatutValidite.VALIDE,
            dossier_sinistre_id__isnull=False,
            facture_compagnie_id__isnull=True,
        ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0
        cautions = list(Caution.objects.all().values())

        data = {
            'total_global_a_refacturer': money_field(total_global_a_refacturer_global),
        }
        return JsonResponse(data)
    else:
        return render(request, 'comptabilite/refacturation-assureur.html')

def get_garant_selectionne_data(request, compagnie_id):
    # Vue appel√©e via AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        bureau = request.user.bureau
        compagnie = Compagnie.objects.get(id=compagnie_id)

        # total_garant_a_refacturer = BordereauOrdonnancement.objects.filter( #methode 2
        total_garant_a_refacturer_global = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.PAYE,
            statut_validite=StatutValidite.VALIDE,
            dossier_sinistre_id__isnull=False,
            facture_compagnie_id__isnull=True,
            compagnie=compagnie,
        ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0

        data = {
            'compagnie_name': compagnie.nom,
            'total_garant_selectionne_a_refacturer': money_field(total_garant_a_refacturer_global),
        }
        return JsonResponse(data)
    else:
        return render(request, 'comptabilite/refacturation-assureur.html')


def obtenir_montant_sinistre(request):
    if request.method == 'GET' and 'sinistre_id' in request.GET:
        sinistre_id = request.GET.get('sinistre_id')
        try:
            sinistre = Sinistre.objects.get(id=sinistre_id)
            montant_total = int(sinistre.montant_remb_accepte)
            return JsonResponse({'montant_total': montant_total})
        except Sinistre.DoesNotExist:
            return JsonResponse({'error': 'Le sinistre sp√©cifi√©e n\'existe pas.'}, status=404)
    else:
        return JsonResponse({'error': 'Requ√™te invalide.'}, status=400)


@method_decorator(login_required, name='dispatch')
class  FactureCompagnieView(TemplateView):
    template_name = 'comptabilite/factures-compagnies.html'
    model = FactureCompagnie

    def get_context_data(self, **kwargs):
        bureau = self.request.user.bureau
        context = super().get_context_data(**kwargs)
        factures = FactureCompagnie.objects.all()
        modes_reglements = ModeReglement.objects.all().order_by('libelle')


        # for facture in factures:

        #     if facture.montant_restant <= 0:
        #         facture.statut = StatutFacture.SOLDE
        #     else:
        #         facture.statut = StatutFacture.NON_SOLDE
        #     facture.save()

        factures_impayees = FactureCompagnie.par_bureau(bureau).filter(statut=StatutFacture.NON_SOLDE, statut_validite=StatutValidite.VALIDE).order_by('date_emission') # Du plus ancien au plus r√©centes
        factures_impayees_numeros = factures_impayees.values_list('numero', flat=True) # Facture compagnie pointe plutot sur numero au lieu de #ID üò≤


        ids_compagnies_impayees = factures_impayees.values('compagnie_id').distinct()

        sinistres_factures_impayees = Sinistre.objects.filter(facture_compagnie__numero__in=factures_impayees_numeros)


        compagnies_impayees = Compagnie.par_bureau(bureau).filter(id__in=ids_compagnies_impayees)

        factures_payees = FactureCompagnie.par_bureau(bureau).filter(statut=StatutFacture.SOLDE, statut_validite=StatutValidite.VALIDE).order_by('-date_emission') # plus r√©centes
        compagnies = Compagnie.objects.all()

        context['compagnies_impayees'] = compagnies_impayees
        context['factures_impayees'] = factures_impayees
        context['sinistres_factures_impayees'] = sinistres_factures_impayees
        context['factures_payees'] = factures_payees
        context['compagnies'] = compagnies
        context['modes_reglements'] = modes_reglements
        context['now'] = timezone.now().date()
        context.update(admin.site.each_context(self.request))

        return context

def facture_compagnie_datatable(request):
    items_per_page = int(request.GET.get('length', 10))
    page_number = int(request.GET.get('start', 0)) // items_per_page + 1
    sort_column_index = int(request.GET.get('order[0][column]', 0))
    sort_direction = request.GET.get('order[0][dir]', 'asc')

    search_numero = request.GET.get('search_numero', '').strip()
    search_compagnie = request.GET.get('search_compagnie', '').strip()
    search_date_debut = request.GET.get('search_date_debut', '').strip()
    search_date_fin = request.GET.get('search_date_fin', '').strip()
    filter_status = request.GET.get('filter_status', '').strip()

    sort_column = ["date_emission", "numero", "compagnie__nom", "montant_total", "statut", "updated_at"][sort_column_index]
    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    queryset = FactureCompagnie.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE)




    if filter_status == 'impayee':
        queryset = queryset.filter(montant_restant__gt=0)
    elif filter_status == 'payee':
        queryset = queryset.filter(montant_restant__lte=0)

    if search_numero:
        queryset = queryset.filter(numero__contains=search_numero)
    if search_compagnie:
        queryset = queryset.filter(compagnie_id=search_compagnie)
    if search_date_debut:
        queryset = queryset.filter(date_emission__gte=search_date_debut)
    if search_date_fin:
        queryset = queryset.filter(date_emission__lte=search_date_fin)

    queryset = queryset.order_by(sort_column)
    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)

    data = []

    for facture in page_obj:

        voir_url = reverse('detail_facture_garant', args=[facture.id])
        fichier = facture.fichier.url if facture.fichier else "#"

        if facture.statut == "SOLDE":
            statut_facture_html = f'<span class="badge badge-success p-2 mb-2">{facture.statut}</span>'
        elif facture.statut == "NON SOLDE":

            statut_facture_html = f'<span class="badge badge-warning p-2 mb-2">{facture.statut}</span>'


        url_regler_facture_compagnie_treso = reverse('reglement_facture_simple', args=[facture.id])
        url_annuler_facture_compagnie_treso = reverse('annulation_facture_simple', args=[facture.id])

        modal_title = "R√âGL√âMENT DE FACTURE " + str(facture.compagnie.nom)

        if facture.statut == "SOLDE":
            actions_html = f'<a title="Voir les paiements et sinistres" href="{voir_url}" class="badge badge-warning p-2 mr-2 mb-2"><i class="fa fa-eye"></i> Voir</a>'\
                f'<a title="T√©l√©charger la facture" href="{fichier}" target="_blank" class="badge badge-secondary text-white p-2 mb-2"><i class="fas fa-download"></i> T√©l√©charger</a>'
        else:
            if int(facture.montant_total) != int(facture.montant_restant):
                actions_html = f'<a title="Faire un r√©gl√©ment" class="badge badge-success btn_regler_facture_compagnie_treso text-white p-2 mr-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="{modal_title}" style="cursor: pointer;" data-href="{url_regler_facture_compagnie_treso}"><i class="fas fa-file-invoice"></i> R√©gler</a>'\
                                f'<a title="Voir les paiements et sinistres" href="{voir_url}" class="badge badge-warning p-2 mr-2 mb-2"><i class="fa fa-eye"></i></a>' \
                                f'<a title="T√©l√©charger la facture" href="{fichier}" target="_blank" class="badge badge-secondary text-white p-2 mb-2"><i class="fas fa-download"></i></a>'
            else:
                actions_html = f'<a title="Faire un r√©gl√©ment" class="badge badge-success btn_regler_facture_compagnie_treso text-white p-2 mr-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="{modal_title}" style="cursor: pointer;" data-href="{url_regler_facture_compagnie_treso}"><i class="fas fa-file-invoice"></i> R√©gler</a>'\
                            f'<a title="Voir les paiements et sinistres" href="{voir_url}" class="badge badge-warning p-2 mr-2 mb-2"><i class="fa fa-eye"></i></a>' \
                            f'<a title="T√©l√©charger la facture" href="{fichier}" target="_blank" class="badge badge-secondary text-white mr-2 p-2 mb-2"><i class="fas fa-download"></i></a>'\
                            f'<a title="Annuler la facture" class="badge o-bg-primary btn_annuler_une_facture_compagnie_treso text-white p-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="ANNULER LA FACTURE" style="cursor: pointer;" data-href="{url_annuler_facture_compagnie_treso}"><i class="fas fa-trash"></i> </a>'

        nombre_versement = ReglementCompagnie.objects.filter(numero=facture.numero).count()
        badge_class = 'badge-success' if facture.montant_restant <= 0 else 'badge-danger'
        nombre_versements = f'<span class="badge {badge_class}">{nombre_versement}</span>'

        data.append({
            "id": facture.id,
            "numero": f'{facture.numero}', # {nombre_versements}
            "date_emission": facture.date_emission.strftime("%d/%m/%Y"),
            "compagnie": facture.compagnie.nom,
            "montant_total": f'<div style="text-align:right;">{money_field(facture.montant_total)}</div>',
            "montant_regle": f'<div style="text-align:right;">{money_field(facture.montant_regle)}</div>',
            "montant_restant": f'<div style="text-align:right;">{money_field(facture.montant_restant)}</div>',
            "statut": statut_facture_html,
            "updated_at": facture.updated_at.strftime("%d/%m/%Y"),
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": FactureCompagnie.objects.filter(bureau=request.user.bureau).count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def fetch_factures(request):
    bureau = request.user.bureau
    compagnie_id = request.GET.get('compagnie_id', None)

    if compagnie_id is not None:
        # R√©cup√©rer les factures pour la compagnie sp√©cifi√©e
        factures = FactureCompagnie.par_bureau(bureau).filter(compagnie_id=compagnie_id).values('numero', 'compagnie', 'montant_total', 'montant_regle', 'montant_restant')

        # Convertir en liste de dictionnaires pour la r√©ponse JSON
        factures_list = list(factures)

        return JsonResponse(factures_list, safe=False)

    return JsonResponse([], safe=False)

@transaction.atomic
def reglement_facture_garant_simple(request, facture_id):
    facture = FactureCompagnie.objects.get(id=facture_id)

    with transaction.atomic():
        if request.method == 'POST':

            facture_compagnie = FactureCompagnie.objects.get(id=facture_id)
            date_reglement = request.POST.get('date_reglement')
            mode_reglement = ModeReglement.objects.get(id=request.POST.get('mode_reglement_id'))
            banque_emettrice = request.POST.get('banque_id')
            # banque = Banque.objects.get(id=request.POST.get('banque_id'))
            numero_piece = request.POST.get('numero_piece')
            compte_tresorerie = CompteTresorerie.objects.get(id=request.POST.get('compte_tresorerie_id'))
            devise = request.user.bureau.pays.devise

            # montant a regler avec format espacer money_field
            montant_a_regler = request.POST.get('montant_a_regler')
            # Nettoyer la valeur du montant
            montant_a_regler = int(montant_a_regler.replace(" ", ""))

            if montant_a_regler > facture_compagnie.montant_restant:
                return JsonResponse(
                    {
                        'statut': 0,
                        'message': 'Erreur, vous ne pouvez pas r√©gler ' + str(montant_a_regler) + ' sur cette facture.',
                    }
                )

            # V√©rification des donn√©es requises
            if facture_compagnie and montant_a_regler:

                # Cr√©er et enregistrer une instance de reglement compagnie dans la base de donn√©es
                reglement_compagnie = ReglementCompagnie.objects.create(
                    bureau = request.user.bureau,
                    created_by = request.user,
                    compagnie = facture_compagnie.compagnie,
                    numero = facture_compagnie.numero,
                    numero_piece = numero_piece,
                    mode_reglement = mode_reglement,
                    banque_emettrice=banque_emettrice,
                    compte_tresorerie = compte_tresorerie,
                    devise = devise,
                    montant = montant_a_regler,
                    observation = "Cr√©√© par interface le " + str(timezone.now().date()),
                    date_reglement = date_reglement,
                    statut_validite = StatutValidite.VALIDE,
                )

                # Mise √† jour le la ligne facture
                if reglement_compagnie:
                    facture_compagnie.montant_regle += reglement_compagnie.montant
                    facture_compagnie.montant_restant -= reglement_compagnie.montant
                    facture_compagnie.save()

                    if int(facture_compagnie.montant_restant) == 0:
                        facture_compagnie.statut = StatutFacture.SOLDE
                        facture_compagnie.save()

                return JsonResponse(
                    {
                        'statut': 1,
                        'message': 'Facture r√©gl√©e avec success actualis√© avec succ√®s !',
                    }
                )

            else:
                return JsonResponse(
                    {
                        'statut': 0,
                        'message': 'Veuillez fournir toutes les donn√©es requises'
                    }
                )
        else:
            facture = FactureCompagnie.objects.get(id=facture_id)
            modes_reglements = ModeReglement.objects.all().order_by('libelle')
            comptes_tresoreries = CompteTresorerie.objects.all().order_by('libelle')
            banques = Banque.objects.filter(bureau=request.user.bureau).order_by('libelle')

            return render(request, 'modals/creation-reglement-facture-unique-garant.html', {'facture': facture, 'modes_reglements': modes_reglements, 'comptes_tresoreries': comptes_tresoreries, 'banques': banques, 'now':timezone.now().date()})

@transaction.atomic
def annulation_facture_simple(request, facture_id):
    bureau = request.user.bureau

    with transaction.atomic():
        if request.method == "POST":

            observation = "facture annul√© d√©puis d'interface par " + request.user.first_name + request.user.last_name

            facture_compagnie = FactureCompagnie.objects.get(id=facture_id)

            sinistres = Sinistre.objects.filter(facture_compagnie=facture_compagnie)

            # Set facture compagnie = Null
            for sinistre in sinistres:
                        sinistre.facture_compagnie = None
                        sinistre.save()

            # Supprimer le bordereau
            facture_compagnie.fichier = None
            facture_compagnie.statut_validite = StatutValidite.SUPPRIME
            facture_compagnie.observation = observation
            facture_compagnie.save()
            return JsonResponse(
                {
                    "statut": 1,
                    "message": "Facture annuler avec succ√®s ! keep a cool head"
                }
            )

        else:
            facture = FactureCompagnie.objects.get(id=facture_id)
            sinistres_factures_impayees = Sinistre.objects.filter(facture_compagnie=facture)

            return render(request, 'modals/annuler-facture-garant.html', {'facture': facture, 'sinistres_factures_impayees': sinistres_factures_impayees})

class DetailFactureGarant(TemplateView):
    template_name = 'comptabilite/detail_facture_garant.html'
    model = ReglementCompagnie

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # R√©cup√©rer l'utilisateur actuel √† partir de la requ√™te
        user = self.request.user

        # R√©cup√©rer le facture_id depuis les kwargs
        facture_id = self.kwargs.get('facture_id')

        # R√©cup√©rer la facture sp√©cifique en utilisant son ID
        try:
            facture = FactureCompagnie.objects.get(id=facture_id)
        except FactureCompagnie.DoesNotExist:
            facture = None

        # Calculer le montant r√©gl√© et restant pour la facture
        if facture:
            sinistres_factures_impayees = Sinistre.objects.filter(facture_compagnie=facture)

            paiements = ReglementCompagnie.objects.filter(numero=facture.numero).order_by('date_reglement')
            montant_restant = facture.montant_total
            montant_regle = 0
            nombre_versements = 0

            for paiement in paiements:
                montant_restant -= paiement.montant
                paiement.montant_restant = montant_restant
                nombre_versements += 1
                montant_regle += paiement.montant
                paiement.save()

            facture.montant_restant = montant_restant
            facture.montant_regle = montant_regle
            facture.save()
        else:
            paiements = []
            facture = None
            nombre_versements = 0

        context.update(admin.site.each_context(self.request))
        context['facture'] = facture
        context['paiements'] = paiements
        context['nombre_versements'] = nombre_versements
        context['sinistres_factures_impayees'] = sinistres_factures_impayees

        return context


# Reg√©n√©rer la facture compagnie / !!! actualisation du fichier pdf uniquement : calcul au niveau de la table !!!
def regenerateFactureGarantpdf(request, facture_id):
    
    if request.method == "POST":
        
        # R√©cup√©rer l'instance de FactureCompagnie
        try:
            facture_compagnie = FactureCompagnie.objects.get(id=facture_id)
        except FactureCompagnie.DoesNotExist:
            return JsonResponse({
                'statut': 0,
                'message': 'Facture non trouv√©e',
                'status': 404
            })
            
        facture_compagnie = generate_facture_compagnie_pdf(request, facture_compagnie.id)
        
        return JsonResponse(
            {
                'statut':1,
                'message': 'Facture ref√©n√©rer avec succ√®s',
                'facture_pdf': facture_compagnie.fichier.url,
                'facture_id': facture_compagnie.pk,
                'status':200
            }
        )
    
    
    
@method_decorator(login_required, name="dispatch")
class SuiviTresorerie(TemplateView):
    template_name = 'comptabilite/suivi-tresorerie.html'
    model = Sinistre

    def get(self, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)
        bureau = self.request.user.bureau

        # garants_cautionne = Compagnie.par_bureau(bureau).filter(caution__isnull=False).distinct().annotate(
        #     montant_caution=Sum('caution__montant', distinct=True),
        #     created_at_caution=Min('caution__created_at'), # Oblig√© d'utiliser une aggr√©gation Min ou Max üòï
        # )

        # Pr√©parer les donn√©es pour le camembert par bureau uniquement(global et non par compagnie)
        data_camembert = prepare_camembert_data(bureau=bureau)
        data_chart_bar = prepare_chart_bar_data(bureau=bureau)
        data_chart_line = prepare_chart_line_data(bureau=bureau)

        # treso = []
        # # Parcourez chaque compagnie avec caution
        # for compagnie in garants_cautionne:

        #     # Calculer le total en attente de r√®glement pour cette compagnie
        #     total_en_attente_compagnie = FactureCompagnie.objects.filter(
        #         # bureau=compagnie.bureau,
        #         compagnie=compagnie,
        #         statut=StatutFacture.NON_SOLDE
        #     ).aggregate(total_montant_restant=Sum('montant_restant'))['total_montant_restant'] or 0

        #     # Calculez le total √† refacturer pour cette compagnie
        #     total_a_refacturer_compagnie = Sinistre.objects.filter(
        #         prestataire__bureau=self.request.user.bureau,
        #         compagnie=compagnie,
        #         statut_paiement=StatutPaiementSinistre.PAYE,
        #         statut_validite=StatutValidite.VALIDE,
        #         dossier_sinistre_id__isnull=False,
        #         facture_compagnie_id__isnull=False,
        #     ).aggregate(montant_remboursement_accepte=Sum('montant_remboursement_accepte'))['montant_remboursement_accepte'] or 0

        #     total_provisionne_compagnie = Sinistre.objects.filter(
        #         prestataire__bureau=self.request.user.bureau,
        #         compagnie=compagnie,
        #         statut_validite=StatutValidite.VALIDE,
        #         dossier_sinistre_id__isnull=False,
        #     ).aggregate(montant_remboursement_accepte=Sum('montant_remboursement_accepte'))['montant_remboursement_accepte'] or 0

        #     total_refacture_compagnie = FactureCompagnie.objects.filter(
        #         statut=StatutFacture.NON_SOLDE,
        #         compagnie=compagnie,
        #     ).aggregate(montant_restant=Sum('montant_restant'))['montant_restant'] or 0

        #     total_rembourse_compagnie = FactureCompagnie.objects.filter(
        #         compagnie=compagnie,
        #     ).aggregate(montant_regle=Sum('montant_regle'))['montant_regle'] or 0

        #     total_montant_ordonnance_compagnie = Sinistre.objects.filter(
        #         compagnie=compagnie,
        #         statut_paiement=StatutPaiementSinistre.ORDONNANCE,
        #         statut_validite=StatutValidite.VALIDE,
        #         dossier_sinistre_id__isnull=False,
        #         facture_compagnie_id__isnull=True,
        #     ).aggregate(montant_remboursement_accepte=Sum('montant_remboursement_accepte'))['montant_remboursement_accepte'] or 0

        #     # Ajoutez les valeurs calcul√©es √† la compagnie actuelle
        #     compagnie.total_en_attente_de_reglement = total_en_attente_compagnie
        #     compagnie.montant_ordonnance = total_montant_ordonnance_compagnie
        #     compagnie.total_provisionne = total_provisionne_compagnie
        #     compagnie.total_a_refacturer = total_a_refacturer_compagnie
        #     compagnie.total_refacture = total_refacture_compagnie

        #     # treso =  caution - montant √† r√©facturer - montant r√©factur√© + montant rembours√©
        #     compagnie.tresorerie = compagnie.montant_caution - (compagnie.total_a_refacturer + compagnie.total_refacture) + total_rembourse_compagnie

        #     if compagnie.montant_caution != 0:
        #         compagnie.dispo = int((compagnie.tresorerie/compagnie.montant_caution) * 100)
        #     else:
        #         compagnie.dispo = 0

        #     # Ajoutez la compagnie avec les valeurs calcul√©es √† la liste finale
        #     treso.append(compagnie)

        context_perso = {
            # "garants": garants_cautionne,
            # "treso": treso,
            "data_camembert": data_camembert,
            "data_chart_bar": data_chart_bar,
            "data_chart_line": data_chart_line,
        }

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# @require_GET
def suivi_treso_datatable(request):
    bureau = request.user.bureau
    # R√©cup√©ration des param√®tres de recherche et de pagination
    search_value = request.GET.get('search[value]', '')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 100))
    #dd(bureau)
    treso = []
    garants = Compagnie.par_bureau(bureau=bureau)

    for compagnie in garants:
        dispo = 0

        caution = Caution.par_bureau(bureau).filter(compagnie=compagnie, date_fin_effet__isnull=True,
                                                    status=True).first()
        mcaution = caution.montant if caution else 0
        montant_caution_formated = money_field(mcaution) + " " + request.user.bureau.pays.devise.code
        
        # Montant ordonnanc√© 
        montant_ordonnances = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.ORDONNANCE,
            statut_validite=StatutValidite.VALIDE,
            dossier_sinistre_id__isnull=False,
            # facture_compagnie__isnull=False, # Jointure unitile ici
            compagnie=compagnie,
        ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0
        montant_ordonnances_formated = money_field(montant_ordonnances) + " " + request.user.bureau.pays.devise.code
        
        # Fin Montant ordonnanc√© 

        # Sinistre r√©gl√© par INOV non encore r√©clam√© √† la compagnie
        montant_regle_non_reclame = Sinistre.objects.filter(
            prestataire__bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.PAYE,
            statut_validite=StatutValidite.VALIDE,
            dossier_sinistre_id__isnull=False,
            facture_compagnie__isnull=True,
            compagnie=compagnie,
        ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0
        montant_regle_non_reclame_formate = money_field(
            montant_regle_non_reclame) + " " + request.user.bureau.pays.devise.code
        # Fin / Sinistre r√©gl√© par INOV non encore r√©clam√© √† la compagnie

        # Sinistre r√©clam√© non encore r√©gl√© par le garant
        montant_reclame_non_regle_par_la_compagnie = FactureCompagnie.objects.filter(
            compagnie=compagnie,
            statut=StatutFacture.NON_SOLDE
        ).aggregate(total_montant_restant=Sum('montant_restant'))['total_montant_restant'] or 0
        montant_reclame_non_regle_par_la_compagnie_formate = money_field(
            montant_reclame_non_regle_par_la_compagnie) + " " + request.user.bureau.pays.devise.code
        # Fin / Sinistre r√©clam√© non encore r√©gl√© par le garant

        # Tr√©sorerie
        tresorerie = mcaution - (montant_regle_non_reclame + montant_reclame_non_regle_par_la_compagnie)
        tresorerie_formate = money_field(tresorerie) + " " + request.user.bureau.pays.devise.code
        # tresorerie = compagnie.montant_caution - (total_a_refacturer_compagnie + total_refacture_compagnie) + total_rembourse_compagnie

        if mcaution != 0:
            dispo = int((tresorerie / mcaution) * 100)

        disponibilite_html = ''

        state = 'secondary'

        if dispo < 11:  # Plus de 90% de consomation
            state = "danger"
        elif dispo < 41:  # entre de 60 a 90 % de consomation
            state = "warning"
        else:  #
            state = "success"

        disponibilite_html = f'<div class="progress rounded"><div class="progress-bar progress-bar-striped progress-bar-animated bg-{state}" role="progressbar" style="width: {dispo}%;" aria-valuenow="{dispo}" aria-valuemin="0" aria-valuemax="100"><span class="p-2">{dispo}%</span></div></div>'

        detail_url = reverse('detail_suivi_treso',
                             args=[compagnie.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a class="badge o-bg-primary text-white" href="{detail_url}" ><i class="fa fa-eye"></i> D√©tails</a>&nbsp;&nbsp;'

        treso.append({
            # 'code_du_garant': compagnie.code,  # Adapte ces cl√©s aux champs de ton mod√®le
            'nom_garant': compagnie.nom,
            'fonds_de_roulement': montant_caution_formated,
            'montant_ordonnances': montant_ordonnances_formated,
            'sinistre_regle_non_reclame': montant_regle_non_reclame_formate,
            'sinistre_reclame_non_regle': montant_reclame_non_regle_par_la_compagnie_formate,
            'tresorerie': tresorerie_formate,
            'niveau_de_dispo': disponibilite_html,
            'actions_html': actions_html
        })

    response = {
        'draw': int(request.GET.get('draw', 0)),
        'recordsTotal': len(treso),
        'recordsFiltered': len(treso),
        'data': treso
    }

    return JsonResponse(response)


@method_decorator(login_required, name="dispatch")
class DetailTresorerie(TemplateView):
    template_name = 'comptabilite/detail_tresorerie_par_garant.html'
    model = ReglementCompagnie

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # R√©cup√©rer l'utilisateur actuel √† partir de la requ√™te
        user = self.request.user
        compagnie_id = self.kwargs.get('compagnie_id')
        compagnie = Compagnie.objects.get(id=compagnie_id)

        data_camembert = prepare_camembert_data(bureau=compagnie.bureau, compagnie=compagnie)
        data_chart_bar = prepare_chart_bar_data(bureau=compagnie.bureau, compagnie=compagnie)
        data_chart_line = prepare_chart_line_data(bureau=compagnie.bureau, compagnie=compagnie)

        context.update(admin.site.each_context(self.request))
        context['compagnie'] = compagnie
        context['data_camembert'] = data_camembert
        context['data_chart_bar'] = data_chart_bar
        context['data_chart_line'] = data_chart_line
        return context


def datatable_facture_compagnie_specifique(request):
    items_per_page = int(request.GET.get('length', 10))
    page_number = int(request.GET.get('start', 0)) // items_per_page + 1
    sort_column_index = int(request.GET.get('order[0][column]', 0))
    sort_direction = request.GET.get('order[0][dir]', 'asc')

    compagnie_id = request.GET.get('compagnie_id', '').strip()
    search_numero = request.GET.get('search_numero', '').strip()
    search_compagnie = request.GET.get('search_compagnie', '').strip()
    search_date_debut = request.GET.get('search_date_debut', '').strip()
    search_date_fin = request.GET.get('search_date_fin', '').strip()
    filter_status = request.GET.get('filter_status', '').strip()

    sort_column = ["date_emission", "numero", "compagnie__nom", "montant_total", "statut", "updated_at"][
        sort_column_index]
    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    queryset = FactureCompagnie.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE)

    compagnie = Compagnie.objects.get(id=compagnie_id)

    if compagnie:
        queryset = queryset.filter(compagnie=compagnie)

    if filter_status == 'impayee':
        queryset = queryset.filter(montant_restant__gt=0)
    elif filter_status == 'payee':
        queryset = queryset.filter(montant_restant__lte=0)

    if search_numero:
        queryset = queryset.filter(numero__contains=search_numero)
    if search_compagnie:
        queryset = queryset.filter(compagnie_id=search_compagnie)
    if search_date_debut:
        queryset = queryset.filter(date_emission__gte=search_date_debut)
    if search_date_fin:
        queryset = queryset.filter(date_emission__lte=search_date_fin)

    queryset = queryset.order_by(sort_column)
    paginator = Paginator(queryset, items_per_page)
    page_obj = paginator.get_page(page_number)

    data = []

    for facture in page_obj:

        voir_url = reverse('detail_facture_garant', args=[facture.id])
        fichier = facture.fichier.url if facture.fichier else "#"

        if facture.statut == "SOLDE":
            statut_facture_html = f'<span class="badge badge-success p-2 mb-2">{facture.statut}</span>'
        elif facture.statut == "NON SOLDE":

            statut_facture_html = f'<span class="badge badge-warning p-2 mb-2">{facture.statut}</span>'

        url_regler_facture_compagnie_treso = reverse('reglement_facture_simple', args=[facture.id])
        url_annuler_facture_compagnie_treso = reverse('annulation_facture_simple', args=[facture.id])
        modal_title = "R√âGL√âMENT DE FACTURE " + str(facture.compagnie.nom)

        if facture.statut == "SOLDE":
            actions_html = f'<a href="{voir_url}" class="badge badge-warning p-2 mb-2"><i class="fa fa-eye"></i> Voir</a>' \
                           f'<a class="badge o-bg-primary text-white p-2 ml-2 mb-2"><i class="fas fa-download"></i> T√©l√©charger</a>'
        else:
            if int(facture.montant_total) != int(facture.montant_restant):
                actions_html = f'<a href="{voir_url}" class="badge badge-warning p-2 mb-2"><i class="fa fa-eye"></i> Voir</a> ' \
                               f'<a href="{fichier}" target="_blank" class="badge badge-secondary text-white p-2 mx-2 mb-2"><i class="fas fa-download"></i></a>' \
                               f'<a class="badge badge-success btn_regler_facture_compagnie_treso text-white p-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="{modal_title}" style="cursor: pointer;" data-href="{url_regler_facture_compagnie_treso}"><i class="fas fa-file-invoice"></i> R√©gler</a>'
            else:
                actions_html = f'<a class="badge badge-success btn_regler_facture_compagnie_treso text-white p-2 mr-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="{modal_title}" style="cursor: pointer;" data-href="{url_regler_facture_compagnie_treso}"><i class="fas fa-file-invoice"></i> R√©gler</a>' \
                               f'<a href="{voir_url}" class="badge badge-warning p-2 mr-2 mb-2"><i class="fa fa-eye"></i></a>' \
                               f'<a href="{fichier}" target="_blank" class="badge badge-secondary text-white mr-2 p-2 mb-2"><i class="fas fa-download"></i></a>' \
                               f'<a class="badge o-bg-primary btn_annuler_une_facture_compagnie_treso text-white p-2 mb-2" data-facture_id="{facture.id}" data-model_name="facture_compagnie" data-modal_title="ANNULER LA FACTURE" style="cursor: pointer;" data-href="{url_annuler_facture_compagnie_treso}"><i class="fas fa-trash"></i> </a>'

        nombre_versement = ReglementCompagnie.objects.filter(numero=facture.numero).count()
        badge_class = 'badge-success' if facture.montant_restant <= 0 else 'badge-danger'
        nombre_versements = f'<span class="badge {badge_class}">{nombre_versement}</span>'

        data.append({
            "id": facture.id,
            "numero": f'{facture.numero} {nombre_versements}',
            "date_emission": facture.date_emission.strftime("%d/%m/%Y"),
            "compagnie": facture.compagnie.nom,
            "montant_total": f'<div style="text-align:right;">{money_field(facture.montant_total)}</div>',
            "montant_regle": f'<div style="text-align:right;">{money_field(facture.montant_regle)}</div>',
            "montant_restant": f'<div style="text-align:right;">{money_field(facture.montant_restant)}</div>',
            "statut": statut_facture_html,
            "updated_at": facture.updated_at.strftime("%d/%m/%Y"),
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": FactureCompagnie.objects.filter(bureau=request.user.bureau).count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
class BordereauxPayesView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/bordereaux_payes.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.all()

        prestataire_ids = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.PAYE, statut_validite=StatutValidite.VALIDE).values_list('prestataire_id', flat=True)
        prestataires = Prestataire.objects.filter(id__in=prestataire_ids).order_by('name')

        adherent_principal_ids = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.PAYE, statut_validite=StatutValidite.VALIDE, type_remboursement__code="RD").values_list('adherent_principal_id', flat=True)
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')

        assures_ids = BordereauOrdonnancement.objects.filter(assure_id__isnull=False, bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.PAYE, statut_validite=StatutValidite.VALIDE, type_remboursement__code="RD").values_list('assure_id',flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()
        context = self.get_context_data(**kwargs)
        # context['dossiers_sinistres'] = dossiers_sinistres
        context['prestataires'] = prestataires
        context['assures'] = assures
        # context['facture_prestataires'] = facture_prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        context['type_remboursements'] = type_remboursements
        context['adhs'] = adhs

        today = timezone.now()
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


### NEW CODE 
def export_bordereaux_ordonnances_paye(request):
    try:
        # Retrieve the BordereauOrdonnancement instances with statut_paiement = 'PAYE'
        queryset = BordereauOrdonnancement.objects.filter(
            bureau=request.user.bureau,
            statut_paiement=StatutPaiementSinistre.PAYE,  # Use the choice constant
        ).order_by('-id')

        # Create an Excel response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_DES_BORDEREAUX_PAYES.xlsx"'

        # Initialize the workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'BORDEREAUX PAY√âS'

        # Write header row
        header = [
            'N¬∞ BORDEREAU',
            'TYPE REMBOURSEMENT',
            'BENEFICIAIRES',
            'P√âRIODE COMPTABLE',
            'MONTANT ACCEPT√â TOTAL',
            'MONTANT REJET PAY√â',
            'MONTANT NET A PAY√â',
            'DATE DE PAYEMENT',
            'STATUT PAIEMENT'
            
        ]
        for col_num, column_title in enumerate(header, 1):
            worksheet.cell(row=1, column=col_num).value = column_title

        # Write data rows
        for row_num, bordereau in enumerate(queryset, 2):
            # Retrieve related data
            type_remboursement = bordereau.type_remboursement.libelle if bordereau.type_remboursement else ""
            periode_comptable = bordereau.periode_comptable.libelle if bordereau.periode_comptable else ""

            row = [
                bordereau.numero,  # Bordereau numero
                type_remboursement,  # Type Remboursement
                bordereau.ordre_de,  # Beneficiaries or other related field
                periode_comptable,  # Periode Comptable
                bordereau.montant_accepte_total,  # Montant total accepte
                bordereau.montant_rejet_total,  # Montant total rejet
                bordereau.montant_total_impaye,  # Montant total impaye
                bordereau.created_at.strftime('%Y-%m-%d %H:%M:%S'),  # Format datetime to string
                bordereau.statut_paiement,  # Statut Paiement
            ]
            
            # Write each cell value to the worksheet
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value

        # Save the workbook to the response
        workbook.save(response)
        return response

    except Exception as e:
        return JsonResponse({"message": str(e)}, status=500)
#
def bordereaux_payes_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')

    queryset = BordereauOrdonnancement.objects.filter(bureau=request.user.bureau, statut_paiement=StatutPaiementSinistre.PAYE, statut_validite=StatutValidite.VALIDE).order_by('-id')

    if search_type_remboursement:
        queryset = queryset.filter(type_remboursement__code=search_type_remboursement)

    if search_numero_bordereau:
        queryset = queryset.filter(numero__contains=search_numero_bordereau)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_assure:
        queryset = queryset.filter(assure__id=search_assure)

    if search_adherent_principal:
        queryset = queryset.filter(adherent_principal__id=search_adherent_principal)


    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'prestataire__name',
        2: 'periode_comptable',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_bordereau_ordonnancement', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> D√©tails</span></a>'

        if c.type_remboursement.code == "TP":
            nom_prestataire = c.prestataire.name if c.prestataire else ""
        elif c.type_remboursement.code == "RD":
            if c.assure:
                nom_prestataire = (
                            c.assure.nom + " " + (c.assure.prenoms if c.assure.prenoms else "")) if c.assure else ""
            else:
                nom_prestataire = f'{c.adherent_principal.nom} {c.adherent_principal.prenoms} ({c.adherent_principal.carte_active()})' if c.adherent_principal else ""


        periode_comptable_libelle = c.periode_comptable.libelle if c.periode_comptable else ""

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "type_remboursement": c.type_remboursement.code,
            "prestataire__name": nom_prestataire,
            "periode_comptable": periode_comptable_libelle,
            "net_a_payer": f'<div style="text-align:right;">{money_field(c.montant_remb_total)}</div>',
            "rejet": f'<div style="text-align:right;">{money_field(c.montant_rejet_total)}</div>',
            "rembourse": f'<div style="text-align:right;">{money_field(c.montant_accepte_total)}</div>',
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })

#######################################################
def regenerate_bordereau_pdf(request, paiement_comptable_id):
    # Fetch the PaiementComptable instance
    paiement_comptable = get_object_or_404(PaiementComptable, id=paiement_comptable_id)

    # Retrieve the associated BordereauOrdonnancement
    bordereau = paiement_comptable.bordereau_ordonnancement

    # Retrieve all sinistres associated with the bordereau
    liste_sinistres = Sinistre.objects.filter(bordereau_ordonnancement=bordereau, statut_validite=StatutValidite.VALIDE)

    # Get unique compagnies concerned by the sinistres
    liste_garant_concernes_ids = liste_sinistres.values_list('compagnie', flat=True).distinct()
    liste_garant_concernes = Compagnie.objects.filter(id__in=liste_garant_concernes_ids)

    sinistres_par_garants = []
    total_global_net_a_payer = Decimal(0)  # Initialize total net to pay

    for garant in liste_garant_concernes:
        sinistres_compagnie_x = liste_sinistres.filter(compagnie=garant.id)

        # Calculate totals for the current compagnie
        total_frais_reel = sum((sinistre.frais_reel or 0) for sinistre in sinistres_compagnie_x)
        total_part_benef = sum((sinistre.part_assure or 0) for sinistre in sinistres_compagnie_x)
        total_base_remb = sum((Decimal(sinistre.total_frais_reel) if sinistre.tm_prefinanced else Decimal(sinistre.total_part_compagnie) or 0) for sinistre in sinistres_compagnie_x)
        total_rejete = sum((sinistre.montant_remb_refuse or 0) for sinistre in sinistres_compagnie_x)
        total_base_taxable = sum((sinistre.montant_remb_accepte or 0) for sinistre in sinistres_compagnie_x)
        total_taxe_tbs = sum((sinistre.montant_taxe_tbs or 0) for sinistre in sinistres_compagnie_x)
        total_taxe_far = sum((sinistre.montant_taxe_far or 0) for sinistre in sinistres_compagnie_x)
        total_taxes = total_taxe_tbs + total_taxe_far
        total_net_a_payer = total_base_taxable + total_taxes

        # Update the global total
        total_global_net_a_payer += total_net_a_payer

        garant_x = {
            'name': garant,
            'sinistres': sinistres_compagnie_x,
            'nbre_total_sinistres': len(sinistres_compagnie_x),
            'total_frais_reel': total_frais_reel,
            'total_part_benef': total_part_benef,
            'total_base_remb': total_base_remb,
            'total_rejete': total_rejete,
            'total_base_taxable': total_base_taxable,
            'total_taxes': total_taxes,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxe_far': total_taxe_far,
            'total_net_a_payer': total_net_a_payer,
        }
        sinistres_par_garants.append(garant_x)

    # Update the Montant total in the PaiementComptable instance
    paiement_comptable.montant_total = total_global_net_a_payer
    paiement_comptable.save()

    total_global_nombre_sinistres = sum(r['nbre_total_sinistres'] for r in sinistres_par_garants)
    total_global_part_assure = sum(r['total_part_benef'] for r in sinistres_par_garants)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in sinistres_par_garants)
    total_global_rejete = sum(r['total_rejete'] for r in sinistres_par_garants)
    total_global_base_remboursement = sum(r['total_base_remb'] for r in sinistres_par_garants)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in sinistres_par_garants)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in sinistres_par_garants)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in sinistres_par_garants)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    # total_global_net_a_payer = sum(r['total_net_a_payer'] for r in sinistres_par_garants)
    
    # Generate the PDF
    pdf = render_pdf("pdf/bordereau_reglement_garant.html", context_dict={
        'sinistres_par_garants': sinistres_par_garants,
        'bordereau': bordereau,
        'paiement_comptable': paiement_comptable,
        'total_global_net_a_payer': total_global_net_a_payer,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxe_far': total_global_taxe_far,
        'total_global_taxes': total_global_taxes,
    })

    # Save the regenerated PDF in the PaiementComptable model
    paiement_comptable.fichier.save(f'bordereau_paiement_comptable_{paiement_comptable.numero}.pdf', File(pdf))
    paiement_comptable.save()

    return HttpResponse(File(pdf), content_type='application/pdf')

##########################################################################


def paiements_comptables_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_paiement_comptable = request.GET.get('search_numero_paiement_comptable', '')
    search_numero_bordereau_ord = request.GET.get('search_numero_bordereau_ord', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')
    search_assure = request.GET.get('search_assure', '')

    queryset = PaiementComptable.objects.filter(bureau=request.user.bureau, statut_validite='VALIDE').order_by('-id')

    if search_type_remboursement:
        queryset = queryset.filter(bordereau_ordonnancement__type_remboursement__code=search_type_remboursement)

    if search_numero_paiement_comptable:
        queryset = queryset.filter(numero__contains=search_numero_paiement_comptable)

    if search_numero_bordereau_ord:
        queryset = queryset.filter(bordereau_ordonnancement__numero__contains=search_numero_bordereau_ord)

    if search_periode_comptable:
        queryset = queryset.filter(bordereau_ordonnancement__periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_assure:
        queryset = queryset.filter(bordereau_ordonnancement__assure__id=search_assure)

    if search_adherent_principal:
        queryset = queryset.filter(bordereau_ordonnancement__adherent_principal__id=search_adherent_principal)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order
            
    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Determine if the user is a superuser
    is_superuser = request.user.is_superuser

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        actions_html = f'''
            <a href="{c.fichier.url}" target="_blank" title="T√©l√©charger le bordereau de paiement">
                <span class="badge btn-sm btn-warning rounded-pill">
                    <i class="fa fa-download"></i> T√©l√©charger
                </span>
            </a>
        '''
        
    # Only add the "Reg√©n√©rer PDF" button if the user is a superuser
        if is_superuser:
            actions_html += f'''
                <a href="{reverse('regenerate_bordereau_pdf', args=[c.id])}" target="_blank" title="Reg√©n√©rer le PDF">
                    <span class="badge btn-sm btn-success rounded-pill">
                        <i class="fa fa-file-pdf"></i> Reg√©n√©rer PDF
                    </span>
                </a>
            '''


        nom_compagnie = c.compagnie.nom if c.compagnie else ""
        nom_beneficiaire = c.nom_beneficiaire
        
        
        # if c.bordereau_ordonnancement.type_remboursement.code == "RD":
        #     # nom_beneficiaire = "RD"
        #     # nom_beneficiaire = f'{c.adherent_principal.nom} {c.adherent_principal.prenoms} ({c.adherent_principal.carte_active()})'
        # elif c.bordereau_ordonnancement.type_remboursement.code == "TP":
        #     # nom_beneficiaire = "TP"
        #     nom_beneficiaire = c.prestataire.name

        data_item = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "bordereau_ordonnancement__numero": c.bordereau_ordonnancement.numero,
            "compagnie__nom": nom_compagnie,
            "nom_beneficiaire": nom_beneficiaire,
            "montant": f'<div style="text-align:right;">{money_field(c.montant_total)}</div>',
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "created_by": c.created_by.username,
            "actions": actions_html,
        }

        data.append(data_item)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })



# @method_decorator(csrf_exempt, name='dispatch')
def export_paiements_comptables(request):
    if request.method == 'POST':
        # R√©cup√©rer les param√®tres de filtre
        search_numero_paiement_comptable = request.POST.get('search_numero_paiement_comptable', '')
        search_numero_bordereau_ord = request.POST.get('search_numero_bordereau_ord', '')
        search_periode_comptable = request.POST.get('search_periode_comptable', '')
        search_type_remboursement = request.POST.get('search_type_remboursement', '')
        search_prestataire = request.POST.get('search_prestataire', '')
        search_adherent_principal = request.POST.get('search_adherent_principal', '')

        # Filtrer les donn√©es selon les param√®tres
        paiements = PaiementComptable.objects.filter(bureau=request.user.bureau)

        if search_periode_comptable:
            paiements = paiements.filter(bordereau_ordonnancement__periode_comptable__id=search_periode_comptable)
        if search_numero_paiement_comptable:
            paiements = paiements.filter(numero__icontains=search_numero_paiement_comptable)
        if search_numero_bordereau_ord:
            paiements = paiements.filter(bordereau_ordonnancement__numero__icontains=search_numero_bordereau_ord)
        if search_type_remboursement:
            paiements = paiements.filter(bordereau_ordonnancement__type_remboursement__code=search_type_remboursement)
        if search_prestataire:
            paiements = paiements.filter(prestataire__id=search_prestataire)
        if search_adherent_principal:
            paiements = paiements.filter(adherent_principal__id=search_adherent_principal)

        # Pr√©paration de l'exportation Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="paiements_filtr√©s.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Paiements Filtr√©s'

        # D√©finir les en-t√™tes
        headers = ['NÀö Paiement', 'NÀö Bordereau ord.', 'B√©n√©ficiaire', 'Garant', 'Net pay√©', 'Date paiement',
                   'Effectu√© par']
        for col_num, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Ajouter les donn√©es filtr√©es
        for row_num, paiement in enumerate(paiements, 1):
            row = [
                paiement.numero,
                paiement.bordereau_ordonnancement.numero,
                paiement.nom_beneficiaire,
                paiement.nom_beneficiaire,
                paiement.montant_total,
                paiement.created_at.strftime("%d/%m/%Y %H:%M"),
                paiement.created_by.username,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    else:
        return HttpResponse(status=405)

def generer_bordereau_reglement_ordonnancement_pdf(request, operation_id):
    operation = Operation.objects.get(id=operation_id)

    reglements = OperationReglement.objects.filter(operation=operation)

    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', {'operation': operation, 'reglements': reglements})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'operation': operation,
    }
    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', contexte)

    # Update bordereau data and save
    #operation.fichier.save(f'bordereau_reglement_compagnie_{operation.numero}.pdf', File(pdf))
    #operation.save()


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')


@method_decorator(login_required, name='dispatch')
class DetailBordereauOrdonnancementView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/details_bordereau_ordonnancement.html'
    model = Sinistre

    def get(self, request, bordereau_id, *args, **kwargs):
        #TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau
        bordereau = BordereauOrdonnancement.objects.filter(id=bordereau_id, bureau=request.user.bureau).first()
        # dd(bordereau)
        
        paiement_comptable = PaiementComptable.objects.filter(bordereau_ordonnancement_id=bordereau.id).first() if bordereau else None

        liste_sinistres_bordereau = Sinistre.objects.filter(bordereau_ordonnancement=bordereau, statut_validite=StatutValidite.VALIDE)

        compagnies_dans_liste = liste_sinistres_bordereau.values('compagnie').distinct()
        compagnies = Compagnie.objects.filter(id__in=compagnies_dans_liste)

        compagnies_non_payes_liste = liste_sinistres_bordereau.filter(statut_paiement=StatutPaiementSinistre.ORDONNANCE).values('compagnie').distinct()
        compagnies_non_payes = Compagnie.objects.filter(id__in=compagnies_non_payes_liste)

        prestataire_ids = liste_sinistres_bordereau.values_list('prestataire_id', flat=True)
        prestataires = Prestataire.objects.filter(id__in=prestataire_ids)

        modes_reglements = ModeReglement.objects.all().order_by('libelle')

        montant_remb_total = bordereau.montant_remb_total
        montant_rejet_total = bordereau.montant_rejet_total
        montant_accepte_total = bordereau.montant_total_impaye

        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in liste_sinistres_bordereau) 
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in liste_sinistres_bordereau)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)

        #montant_remb_total = sum(sinistre.frais_reel if sinistre.tm_prefinanced else sinistre.total_part_compagnie for sinistre in liste_sinistres_bordereau if sinistre)
        montant_total_paye = bordereau.montant_total_paye #0 #sum(sinistre.montant_remb_accepte for sinistre in liste_sinistres_bordereau.filter(statut_paiement=StatutPaiementSinistre.PAYE))
        montant_total_impaye = bordereau.montant_total_impaye + total_taxes #sum(sinistre.montant_remb_accepte for sinistre in liste_sinistres_bordereau.filter(statut_paiement=StatutPaiementSinistre.ORDONNANCE))

        context = self.get_context_data(**kwargs)
        context['bordereau'] = bordereau
        context['paiement_comptable'] = paiement_comptable
        context['liste_sinistres_bordereau'] = liste_sinistres_bordereau
        context['montant_remb_total'] = montant_remb_total
        context['montant_rejet_total'] = montant_rejet_total
        context['montant_accepte_total'] = montant_accepte_total
        context['montant_autres_taxes'] = total_taxes
        context['montant_total_paye'] = montant_total_paye
        context['montant_total_impaye'] = montant_total_impaye
        context['prestataires'] = prestataires
        context['garants'] = compagnies
        context['garants_non_paye'] = compagnies_non_payes
        context['modes_reglements'] = modes_reglements
        context['moyens_paiements'] = modes_reglements

        return self.render_to_response(context)


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



def detail_bordereau_ordonnancement_datatable(request, bordereau_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    # start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    # sort_column_index = int(request.GET.get('order[0][column]'))
    # sort_direction = request.GET.get('order[0][dir]')

    search_numero_bordereau = request.GET.get('search_numero_bordereau', '')
    search_numero_sinistre = request.GET.get('search_numero_sinistre', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_compagnie = request.GET.get('search_compagnie', '')
    search_statut_paiement = request.GET.get('search_statut_paiement', '')

    queryset = Sinistre.objects.filter(bordereau_ordonnancement__id=bordereau_id, statut_validite=StatutValidite.VALIDE).order_by('-id')

    if search_numero_bordereau:
        queryset = queryset.filter(dossier_sinistre__numero__contains=search_numero_bordereau)

    if search_numero_sinistre:
        queryset = queryset.filter(numero__contains=search_numero_sinistre)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=int(search_prestataire))

    if search_compagnie:
        queryset = queryset.filter(compagnie=search_compagnie)

    if search_statut_paiement:
        queryset = queryset.filter(statut_paiement=search_statut_paiement)


    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        if not c.aliment:
            c.aliment.nom = ''
        if not c.aliment:
            c.aliment.prenoms = ''

        statut_html = f'<span class="badge badge-{c.statut_paiement.lower()}">{c.statut_paiement}</span>'

        popup_details_sinistre = reverse('popup_details_sinistre', args=[c.id])
        actions_html = f'<span title="D√©tails" data-href="{popup_details_sinistre}" class="btn badge btn-sm btn-details rounded-pill btn-popup_details_sinistre"><i class="fa fa-eye"></i> D√©tails</span>'

        total_facture = c.total_frais_reel if c.total_frais_reel else 0
        total_part_assure = 0 if c.tm_prefinanced else c.total_part_assure
        total_base_remb = c.total_frais_reel if c.tm_prefinanced else c.total_part_compagnie
        total_rejet = c.montant_remb_refuse if c.montant_remb_refuse else 0
        total_net_payer = c.montant_remb_accepte if c.montant_remb_accepte else 0

        # collecte des montants totaux

        cartes = c.aliment.cartes.filter(statut=Statut.ACTIF) if c.aliment else None
        numero_carte = cartes.first().numero if cartes else None

        data_iten = {
            "id": c.id,
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M") if c.date_survenance else '',
            "numero": c.numero,
            "dossier_sinistre__numero": c.dossier_sinistre.numero,
            "prestataire": c.prestataire.name,
            "beneficiaire": c.aliment.nom + ' ' + c.aliment.prenoms,
            "carte_active": numero_carte,
            "compagnie": c.compagnie.nom,
            "total_facture": money_field(total_facture),
            "total_part_assure": money_field(total_part_assure),
            "total_base_remb": money_field(total_base_remb),
            "total_rejet": money_field(total_rejet),
            "total_net_payer": money_field(total_net_payer),
            "statut": statut_html,
            "actions": actions_html,
        }
        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })




def calculer_montant_accepte_total(garant_id, bordereau_id):
    bordereau = BordereauOrdonnance.objects.get(id=bordereau_id)

    query_params = {'bordereau_ordonnancement': bordereau}

    if garant_id:
        query_params['compagnie'] = Compagnie.objects.get(id=garant_id)

    liste_sinistres_bordereau = Sinistre.objects.filter(**query_params)

    montant_accepte_total = sum(item.montant_remb_accepte for item in liste_sinistres_bordereau)
    if not garant_id:
        montant_accepte_total = bordereau.montant_total_impaye

    # Formatage du montant_accepte_total avec 0 d√©cimales
    montant_accepte_total_formatted = "{:.0f}".format(montant_accepte_total)

    montant_accepte_total_formatted = format_number(montant_accepte_total, locale='fr')

    return montant_accepte_total_formatted


def update_montant_accepte_total(request):
    garant_id = request.GET.get('garant_id')
    bordereau_id = request.GET.get('bordereau_id')

    # Effectuez le traitement n√©cessaire pour r√©cup√©rer le montant_accepte_total en fonction des variables
    # Dans cet exemple, je suppose que vous avez une m√©thode ou une logique pour calculer le montant_accepte_total
    montant_accepte_total = calculer_montant_accepte_total(garant_id, bordereau_id)

    # Renvoyez la r√©ponse JSON
    return JsonResponse({'montant_accepte_total': montant_accepte_total})

def add_mise_en_reglement_ordonnancement(request):

    if request.method == 'POST':
        devise = request.POST.get('devise')
        mode_reglement = request.POST.get('mode_reglement')
        compte_tresorerie = request.POST.get('compte_tresorerie')

        nature_operation_code = "REGSIN"
        nature_operation = NatureOperation.objects.filter(code=nature_operation_code).first()

        #enregistrer les infos dans operation
        nombre_quittances = 0
        montant_total_regle = 0
        operation = Operation.objects.create(nature_operation=nature_operation,
                                             montant_total=montant_total_regle,
                                             compte_tresorerie_id=compte_tresorerie,
                                             devise_id=devise,
                                             mode_reglement_id=mode_reglement,
                                             created_by=request.user)
        operation.save()

        #enregistrer le details dans reglements (liste les quittances regl√©es avec chaque montant)
        nombre_reglements_selectionnes = 0
        montant_total_reglements_selectionne = 0
        i = 0

        response = {
            'statut': 1,
            'message': "R√®glement effectu√© avec succ√®s !",
            'data': {'montant_total_reglements_selectionne': montant_total_reglements_selectionne, 'nombre_reglements_selectionnes': nombre_reglements_selectionnes, 'operation_id': operation.pk}
        }

        return JsonResponse(response)


    else:
        return render(request, 'modals.creation-mise-en-reglement.html')

@transaction.atomic
def add_mise_en_reglement_ordonnancement_par_garant(request):
    with transaction.atomic():
        if request.method == 'POST':
            compagnie_id = request.POST.get('compagnie_id', '')

            code_mode_de_reglement = request.POST.get('garant_moyens_paiement_id', '')

            numero_piece = request.POST.get('numero_piece', '')

            ordre_de = request.POST.get('ordre_de', '')

            iban = request.POST.get('iban', '')

            date_de_paiement = request.POST.get('date_paiement', '')

            numero_bordereau = request.POST.get('numero', '')

            montant_a_regler = request.POST.get('montant_a_regler', '')
            montant_a_regler = montant_a_regler.replace(',', '.')
            montant_a_regler = float(montant_a_regler)

            if numero_bordereau:
                bordereau = BordereauOrdonnance.objects.filter(numero=numero_bordereau, bureau=request.user.bureau).first()
                if bordereau is not None:
                    if bordereau.montant_total_paye is not None:
                        bordereau.montant_total_paye = bordereau.montant_total_paye + montant_a_regler
                    else:
                        bordereau.montant_total_paye = montant_a_regler

                    bordereau.montant_total_impaye = bordereau.montant_accepte_total - bordereau.montant_total_paye

                bordereau.save()



            # GENERER LE PDF
            if bordereau.assure:
                pdf = genreate_bordereau_reglement_assure_par_garant(request, bordereau.id, bordereau.assure.id, compagnie_id)
            else:
                pdf = genreate_bordereau_reglement_par_garant(request, bordereau.id, compagnie_id)
            # END GENERERATION PDF

            bordereau_apres_paiement = BordereauOrdonnance.objects.filter(numero=numero_bordereau, bureau=request.user.bureau).first()

            statut_bordereau_pour_redirection = bordereau_apres_paiement.statut_paiement if bordereau_apres_paiement else ''

            pprint("ON VA PASSER A L'AFFICHAGE")

            # upload file on server and return its url
            # file_url = uploaded_file_url(pdf)

            return JsonResponse(
                {'statut': 1, 'message': 'Bordereau de paiement g√©n√©r√© avec succ√®ss',
                    'bordereau_pdf': pdf.url, 'statut_bordereau_pour_redirection': statut_bordereau_pour_redirection}, status=200)

        else:
            return render(request, 'modals.creation-mise-en-reglement-par-garant.html')


def genreate_bordereau_reglement_par_garant(request, bordereau_id, compagnie_id):

    mode_reglement_id = request.POST.get('garant_moyens_paiement_id', '')
    numero_piece = request.POST.get('numero_piece', '')
    ordre_de = request.POST.get('ordre_de', '')
    numero_iban = request.POST.get('iban', '')
    date_paiement = request.POST.get('date_paiement', '')
    numero_bordereau = request.POST.get('numero', '')
    montant_a_regler = request.POST.get('montant_a_regler', '').replace(" ", "")



    sinistres_par_garants = []

    bordereau = BordereauOrdonnance.objects.get(id=bordereau_id)
    liste_sinistres = Sinistre.objects.filter(bordereau_ordonnancement=bordereau, statut_validite=StatutValidite.VALIDE)

    nombre_total_sinistres = liste_sinistres.count()

    liste_sinistres_ordonnances = liste_sinistres.filter(statut_paiement=StatutPaiementSinistre.ORDONNANCE)

    compagnie = Compagnie.objects.get(id=compagnie_id) if compagnie_id else None
    mode_reglement = ModeReglement.objects.get(id=mode_reglement_id) if mode_reglement_id else None


    if compagnie:
        liste_sinistres_ordonnances = liste_sinistres_ordonnances.filter(compagnie=compagnie)

    liste_garant_concernes_ids = liste_sinistres_ordonnances.values_list('compagnie', flat=True)
    liste_garant_concernes = Compagnie.objects.filter(id__in=liste_garant_concernes_ids)


    for garant in liste_garant_concernes:
        sinistres_compagnie_x = liste_sinistres_ordonnances.filter(compagnie=garant.id)


        # Calcul de  la somme du champ montant total a payer par garant
        total_frais_reel = sum(sinistre.frais_reel for sinistre in sinistres_compagnie_x)

        total_part_benef = sum(sinistre.part_assure for sinistre in sinistres_compagnie_x)
        total_base_remb = sum((Decimal(s.total_frais_reel) if s.tm_prefinanced else Decimal(s.total_part_compagnie)) or 0 for s in sinistres_compagnie_x)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres_compagnie_x)
        total_base_taxable = sum(s.montant_remb_accepte or 0 for s in sinistres_compagnie_x)
        # total_net_a_payer = total_base_taxable
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres_compagnie_x) 
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres_compagnie_x)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes
        garant_x = {
            'name': garant,
            'sinistres': sinistres_compagnie_x,
            'nbre_total_sinistres': len(sinistres_compagnie_x),
            'total_frais_reel': total_frais_reel,
            'total_part_benef': total_part_benef,
            'total_base_remb': total_base_remb,
            'total_rejete': total_rejete,
            'total_base_taxable': total_base_taxable,
            'total_taxes': total_taxes,
            'total_net_a_payer': total_net_a_payer,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxe_far': total_taxe_far,
        }
        sinistres_par_garants.append(garant_x)


    total_global_nombre_sinistres = sum(r['nbre_total_sinistres'] for r in sinistres_par_garants)
    total_global_part_assure = sum(r['total_part_benef'] for r in sinistres_par_garants)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in sinistres_par_garants)
    total_global_rejete = sum(r['total_rejete'] for r in sinistres_par_garants)
    total_global_base_remboursement = sum(r['total_base_remb'] for r in sinistres_par_garants)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in sinistres_par_garants)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in sinistres_par_garants)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in sinistres_par_garants)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(r['total_net_a_payer'] for r in sinistres_par_garants)



    # montant_total = sum(x.montant_remb_accepte for x in liste_sinistres_ordonnances)

    montant_total = total_global_net_a_payer

    paiement_comptable = PaiementComptable.objects.create(
        created_by=request.user,
        bureau=request.user.bureau,
        bordereau_ordonnancement=bordereau,
        compagnie=compagnie,
        #numero=generate_random_string(length=12),
        prestataire=bordereau.prestataire,
        nom_beneficiaire=bordereau.ordre_de,
        numero_iban=numero_iban,
        nombre_sinistres=liste_sinistres_ordonnances.count(),
        date_paiement=date_paiement,
        mode_reglement=mode_reglement,
        numero_piece=numero_piece,
        montant_total=montant_total
    )

    paiement_comptable.numero = 'PC' + str(Date.today().year) + str(paiement_comptable.pk).zfill(6)
    paiement_comptable.save()

    for sinistre in liste_sinistres_ordonnances:
        # Update paiement status
        sinistre.statut_paiement = StatutPaiementSinistre.PAYE
        sinistre.date_paiement = date_paiement
        sinistre.paiement_comptable = paiement_comptable
        sinistre.facture_prestataire.statut = SatutBordereauDossierSinistres.PAYE
        sinistre.facture_prestataire.save()
        sinistre.save()



    # generer le pdf
    pdf = render_pdf("pdf/bordereau_reglement_garant.html", context_dict={'sinistres_par_garants': sinistres_par_garants, 'bordereau': bordereau, 'paiement_comptable': paiement_comptable})
    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    pdf = render_pdf("pdf/bordereau_reglement_garant.html", context_dict={
        'sinistres_par_garants': sinistres_par_garants,
        'bordereau': bordereau,
        'paiement_comptable': paiement_comptable,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
    })
    pprint(pdf)

    # Enregistrement du bordereau

    paiement_comptable.fichier.save(f'bordereau_paiement_comptable_{paiement_comptable.numero}.pdf', File(pdf))
    paiement_comptable.save()

    #a conditionner si tous les sinistres payes
    nombre_sinistres_payes = liste_sinistres.filter(statut_paiement=StatutPaiementSinistre.PAYE).count()
    if nombre_total_sinistres == nombre_sinistres_payes:
        bordereau.statut_paiement = StatutPaiementSinistre.PAYE
        bordereau.save()


    return paiement_comptable.fichier

def genreate_bordereau_reglement_assure_par_garant(request, bordereau_id, assure_id, compagnie_id):

    mode_reglement_id = request.POST.get('garant_moyens_paiement_id', '')
    numero_piece = request.POST.get('numero_piece', '')
    ordre_de = request.POST.get('ordre_de', '')
    numero_iban = request.POST.get('iban', '')
    date_paiement = request.POST.get('date_paiement', '')
    numero_bordereau = request.POST.get('numero', '')
    montant_a_regler = request.POST.get('montant_a_regler', '').replace(" ", "")



    sinistres_par_garants = []

    bordereau = BordereauOrdonnance.objects.get(id=bordereau_id)
    liste_sinistres = Sinistre.objects.filter(bordereau_ordonnancement=bordereau, facture_prestataire__assure__id=assure_id, statut_validite=StatutValidite.VALIDE)

    nombre_total_sinistres = liste_sinistres.count()

    liste_sinistres_ordonnances = liste_sinistres.filter(statut_paiement=StatutPaiementSinistre.ORDONNANCE)

    compagnie = Compagnie.objects.get(id=compagnie_id) if compagnie_id else None
    mode_reglement = ModeReglement.objects.get(id=mode_reglement_id) if mode_reglement_id else None


    if compagnie:
        liste_sinistres_ordonnances = liste_sinistres_ordonnances.filter(compagnie=compagnie)


    sinistres_par_adherent = defaultdict(list)
    for sinistre in liste_sinistres_ordonnances:
        adherent_id = sinistre.adherent_principal.id
        sinistres_par_adherent[adherent_id].append(sinistre)

    # liste_garant_concernes_ids = liste_sinistres_ordonnances.values_list('compagnie', flat=True)
    # liste_garant_concernes = Compagnie.objects.filter(id__in=liste_garant_concernes_ids)


    for adherent_id, sinistres_compagnie_x in sinistres_par_adherent.items():
        # sinistres_compagnie_x = liste_sinistres_ordonnances.filter(compagnie=garant.id)


        # Calcul de  la somme du champ montant total a payer par garant
        total_frais_reel = sum(sinistre.frais_reel for sinistre in sinistres_compagnie_x)

        total_part_benef = sum(sinistre.part_assure for sinistre in sinistres_compagnie_x)
        total_base_remb = sum((Decimal(s.total_frais_reel) if s.tm_prefinanced else Decimal(s.total_part_compagnie)) or 0 for s in sinistres_compagnie_x)
        total_rejete = sum(s.montant_remb_refuse or 0 for s in sinistres_compagnie_x)
        total_base_taxable = sum(s.montant_remb_accepte or 0 for s in sinistres_compagnie_x)
        # total_net_a_payer = total_base_taxable
        total_taxe_tbs = sum(s.montant_taxe_tbs or 0 for s in sinistres_compagnie_x)
        total_taxe_far = sum(s.montant_taxe_far or 0 for s in sinistres_compagnie_x)
        total_taxes = int(total_taxe_tbs) + int(total_taxe_far)
        total_net_a_payer = total_base_taxable + total_taxes
        garant_x = {
            'name': sinistres_compagnie_x[0].adherent_principal.nom_prenoms,
            'sinistres': sinistres_compagnie_x,
            'nbre_total_sinistres': len(sinistres_compagnie_x),
            'total_frais_reel': total_frais_reel,
            'total_part_benef': total_part_benef,
            'total_base_remb': total_base_remb,
            'total_rejete': total_rejete,
            'total_base_taxable': total_base_taxable,
            'total_taxes': total_taxes,
            'total_net_a_payer': total_net_a_payer,
            'total_taxe_tbs': total_taxe_tbs,
            'total_taxe_far': total_taxe_far,
        }
        sinistres_par_garants.append(garant_x)


    total_global_nombre_sinistres = sum(r['nbre_total_sinistres'] for r in sinistres_par_garants)
    total_global_part_assure = sum(r['total_part_benef'] for r in sinistres_par_garants)
    total_global_frais_reel = sum(r['total_frais_reel'] for r in sinistres_par_garants)
    total_global_rejete = sum(r['total_rejete'] for r in sinistres_par_garants)
    total_global_base_remboursement = sum(r['total_base_remb'] for r in sinistres_par_garants)
    total_global_base_taxable = sum(r['total_base_taxable'] for r in sinistres_par_garants)
    total_global_taxe_tbs = sum(r['total_taxe_tbs'] for r in sinistres_par_garants)
    total_global_taxe_far = sum(r['total_taxe_far'] for r in sinistres_par_garants)
    total_global_taxes = total_global_taxe_far + total_global_taxe_tbs
    total_global_net_a_payer = sum(r['total_net_a_payer'] for r in sinistres_par_garants)



    # montant_total = sum(x.montant_remb_accepte for x in liste_sinistres_ordonnances)

    montant_total = total_global_net_a_payer

    paiement_comptable = PaiementComptable.objects.create(
        created_by=request.user,
        bureau=request.user.bureau,
        bordereau_ordonnancement=bordereau,
        compagnie=compagnie,
        #numero=generate_random_string(length=12),
        prestataire=bordereau.prestataire,
        nom_beneficiaire=bordereau.ordre_de,
        numero_iban=numero_iban,
        nombre_sinistres=liste_sinistres_ordonnances.count(),
        date_paiement=date_paiement,
        mode_reglement=mode_reglement,
        numero_piece=numero_piece,
        montant_total=montant_total
    )

    paiement_comptable.numero = 'PC' + str(Date.today().year) + str(paiement_comptable.pk).zfill(6)
    paiement_comptable.save()

    for sinistre in liste_sinistres_ordonnances:
        # Update paiement status
        sinistre.statut_paiement = StatutPaiementSinistre.PAYE
        sinistre.date_paiement = date_paiement
        sinistre.paiement_comptable = paiement_comptable
        sinistre.facture_prestataire.statut = SatutBordereauDossierSinistres.PAYE
        sinistre.facture_prestataire.save()
        sinistre.save()



    # generer le pdf
    pdf = render_pdf("pdf/bordereau_reglement_assure_garant.html", context_dict={'sinistres_par_garants': sinistres_par_garants, 'bordereau': bordereau, 'paiement_comptable': paiement_comptable})
    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    pdf = render_pdf("pdf/bordereau_reglement_assure_garant.html", context_dict={
        'sinistres_par_garants': sinistres_par_garants,
        'bordereau': bordereau,
        'paiement_comptable': paiement_comptable,
        'total_global_nombre_sinistres': total_global_nombre_sinistres,
        'total_global_part_assure': total_global_part_assure,
        'total_global_frais_reel': total_global_frais_reel,
        'total_global_rejete': total_global_rejete,
        'total_global_base_remboursement': total_global_base_remboursement,
        'total_global_base_taxable': total_global_base_taxable,
        'total_global_taxe_tbs': total_global_taxe_tbs,
        'total_global_taxe_far': total_global_taxe_far,
        'total_global_taxes': total_global_taxes,
        'total_global_net_a_payer': total_global_net_a_payer,
    })
    pprint(pdf)

    # Enregistrement du bordereau

    paiement_comptable.fichier.save(f'bordereau_paiement_comptable_{paiement_comptable.numero}.pdf', File(pdf))
    paiement_comptable.save()

    #a conditionner si tous les sinistres payes
    nombre_sinistres_payes = liste_sinistres.filter(statut_paiement=StatutPaiementSinistre.PAYE).count()
    if nombre_total_sinistres == nombre_sinistres_payes:
        bordereau.statut_paiement = StatutPaiementSinistre.PAYE
        bordereau.save()


    return paiement_comptable.fichier

def generer_bordereau_reglement_ordonnancement_par_garant_pdf(request, operation_id):
    operation = Operation.objects.get(id=operation_id)

    reglements = OperationReglement.objects.filter(operation=operation)

    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', {'operation': operation, 'reglements': reglements})

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    contexte = {
        'operation': operation,
    }
    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', contexte)

    # Update bordereau data and save
    #operation.fichier.save(f'bordereau_reglement_compagnie_{operation.numero}.pdf', File(pdf))
    #operation.save()


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')

def bordereau_ordonnancement_pdf(request):
    bordereau = BordereauOrdonnancement.objects.get(id=15747)
    compagnie_id = None
    pdf = genreate_bordereau_reglement_assure_par_garant(request, bordereau.id, bordereau.assure.id, compagnie_id)
    return JsonResponse(
        {'statut': 1, 'message': 'Bordereau de paiement g√©n√©r√© avec succ√®ss',
         'pdf': pdf.url}, status=200)

@method_decorator(login_required, name="dispatch")
class PaiementsRealises(TemplateView):
    template_name = 'comptabilite/paiements-realises.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.all()

        prestataire_ids = PaiementComptable.objects.filter(bureau=request.user.bureau).values_list('prestataire_id', flat=True)
        prestataires = Prestataire.objects.filter(id__in=prestataire_ids).order_by('name')

        bordereau_ordonnancement_ids = PaiementComptable.objects.filter(bureau=request.user.bureau).values_list('bordereau_ordonnancement_id', flat=True)
        adherent_principal_ids = BordereauOrdonnancement.objects.filter(id__in=bordereau_ordonnancement_ids).values_list('adherent_principal_id', flat=True)
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')

        assures_ids = BordereauOrdonnancement.objects.filter(assure_id__isnull=False, id__in=bordereau_ordonnancement_ids).values_list('assure_id', flat=True).distinct()
        assures = Client.objects.filter(id__in=assures_ids)
        pprint("assures_ids")
        pprint(assures_ids)

        periodes_comptables = PeriodeComptable.objects.all()
        context = self.get_context_data(**kwargs)
        context['prestataires'] = prestataires
        context['assures'] = assures
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        context['type_remboursements'] = type_remboursements
        context['adhs'] = adhs

        today = timezone.now().date()
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


method_decorator(login_required, name="dispatch")
class EditionLettreCheque(TemplateView):
    template_name = 'comptabilite/edition_lettre_cheque.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        user = User.objects.get(id=request.user.id)

        type_remboursements = TypeRemboursement.objects.all()

        prestataire_ids = PaiementComptable.objects.filter(bureau=request.user.bureau, mode_reglement_id=5, bordereau_lettre_cheque__isnull=True).values_list('prestataire_id',
                                                                                                   flat=True)
        prestataires = Prestataire.objects.filter(id__in=prestataire_ids).order_by('name')

        bordereau_ordonnancement_ids = PaiementComptable.objects.filter(bureau=request.user.bureau, mode_reglement_id=5, bordereau_lettre_cheque__isnull=True).values_list('bordereau_ordonnancement_id', flat=True)
        adherent_principal_ids = BordereauOrdonnancement.objects.filter(id__in=bordereau_ordonnancement_ids).values_list('adherent_principal_id', flat=True)
        adhs = Aliment.objects.filter(id__in=adherent_principal_ids).order_by('nom')

        model_lettre_cheques = ModelLettreCheque.objects.filter(bureau=request.user.bureau).order_by('-id')

        periodes_comptables = PeriodeComptable.objects.all()
        context = self.get_context_data(**kwargs)
        context['prestataires'] = prestataires
        context['periodes_comptables'] = periodes_comptables
        context['user'] = user

        context['type_remboursements'] = type_remboursements
        context['adhs'] = adhs
        context['model_lettre_cheques'] = model_lettre_cheques

        today = timezone.now().date()
        context['today'] = today
        context['breadcrumbs'] = 'today'

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def edition_lettre_cheque_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_paiement_comptable = request.GET.get('search_numero_paiement_comptable', '')
    search_numero_bordereau_ord = request.GET.get('search_numero_bordereau_ord', '')
    search_periode_comptable = request.GET.get('search_periode_comptable', '')
    search_type_remboursement = request.GET.get('search_type_remboursement', '')
    search_prestataire = request.GET.get('search_prestataire', '')
    search_adherent_principal = request.GET.get('search_adherent_principal', '')

    queryset = PaiementComptable.objects.filter(bureau=request.user.bureau, mode_reglement_id=5, bordereau_lettre_cheque__isnull=True).order_by('-id')
    print(queryset)

    if search_type_remboursement:
        queryset = queryset.filter(bordereau_ordonnancement__type_remboursement__code=search_type_remboursement)

    if search_numero_paiement_comptable:
        queryset = queryset.filter(numero__contains=search_numero_paiement_comptable)

    if search_numero_bordereau_ord:
        queryset = queryset.filter(bordereau_ordonnancement__numero__contains=search_numero_bordereau_ord)

    if search_periode_comptable:
        queryset = queryset.filter(periode_comptable__id=search_periode_comptable)

    if search_prestataire:
        queryset = queryset.filter(prestataire__id=search_prestataire)

    if search_adherent_principal:
        queryset = queryset.filter(bordereau_ordonnancement__adherent_principal__id=search_adherent_principal)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        3: 'created_at',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = "#"  # reverse('details_paiement_comptable', args=[c.id])  # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{c.fichier.url}" target="_blank" title="T√©l√©charger le bordereau de paiement"><span class="badge btn-sm btn-warning rounded-pill"><i class="fa fa-download"></i> T√©l√©charger</span></a>'

        nom_compagnie = c.compagnie.nom if c.compagnie else ""
        nom_beneficiaire = c.nom_beneficiaire

        # if c.bordereau_ordonnancement.type_remboursement.code == "RD":
        #     # nom_beneficiaire = "RD"
        #     # nom_beneficiaire = f'{c.adherent_principal.nom} {c.adherent_principal.prenoms} ({c.adherent_principal.carte_active()})'
        # elif c.bordereau_ordonnancement.type_remboursement.code == "TP":
        #     # nom_beneficiaire = "TP"
        #     nom_beneficiaire = c.prestataire.name

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "bordereau_ordonnancement__numero": c.bordereau_ordonnancement.numero,
            "compagnie__nom": nom_compagnie,
            "nom_beneficiaire": nom_beneficiaire,
            "montant": f'<div style="text-align:right;">{money_field(c.montant_total)}</div>',
            "created_at": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "created_by": c.created_by.username,
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count() if not request.user.is_med else len(queryset),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@transaction.atomic
def submit_edition_lettre_cheque(request):
    # TODO : CODE_PAYS_YY_MM
    # try:
    with transaction.atomic():

        # print("periode " + request.POST.get('periode_id'))
        # print("search adh " + request.POST.get('search_adh'))
        # print("prestataire " + request.POST.get('prestataire_id'))
        # print("selectedItems" + request.POST.get('selectedItems'))
        # pprint("ADH ID")
        # print(request.POST.get('search_adh'))

        Paiement_ids = literal_eval(request.POST.get('selectedItems'))
        model_lettre_cheque_id = request.POST.get('modelLettreCheque')

        paiements = PaiementComptable.objects.filter(id__in=Paiement_ids).order_by('-id')
        model_lettre_cheque = ModelLettreCheque.objects.get(id=int(model_lettre_cheque_id))

        if len(paiements) > 0 and model_lettre_cheque:
            contexte = {
                'paiements': paiements,
                'date_edition': timezone.now().strftime("%d/%m/%Y")
            }
            pdf = render_pdf(model_lettre_cheque.model, contexte)

            # Cr√©er une r√©ponse avec le type MIME appropri√© pour le PDF
            libelle = f'Edition_lettre_cheque_{timezone.now().strftime("%d_%m_%Y")}'
            bordereau_lettre_cheque = BordereauLettreCheque.objects.create(
                created_by=request.user,
                bureau=request.user.bureau,
                model_lettre_cheque=model_lettre_cheque,
                libelle=libelle,
                nombre=len(paiements),
            )

            # save file on server and return its url
            bordereau_lettre_cheque.fichier.save(
                f'{libelle}.pdf', File(pdf))
            bordereau_lettre_cheque.save()

            for paiement in paiements:
                paiement.bordereau_lettre_cheque = bordereau_lettre_cheque
                paiement.save()


            return JsonResponse(
                {'statut': 1, 'message': 'lettre cheque g√©n√©r√© avec succ√®s',
                 'lettreChequePdf': bordereau_lettre_cheque.fichier.url}, status=200)
        else:
            return JsonResponse({'statut': 0, 'message': 'Aucune donn√©e trouv√©', 'Paiement_ids':Paiement_ids}, status=200)

def edition_lettre_cheque_pdf(request):
    paiements = PaiementComptable.objects.filter(bureau=request.user.bureau, mode_reglement_id=5)[:2]

    contexte = {
        'paiements': paiements,
        'date_edition': timezone.now().strftime("%d/%m/%Y")
    }
    pdf = render_pdf('courriers/lettre_cheque/sn_cboa_model.html', contexte)

    # Update bordereau data and save
    #paiement.fichier.save(f'lettre_cheque_{paiement.numero}.pdf', File(pdf))
    #paiement.save()

    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')


@method_decorator(login_required, name='dispatch')
class ReversesementCompagniesView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/reglements_compagnies.html'
    model = Reglement

    def get(self, request, *args, **kwargs):

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')

        for compagnie in compagnies:
            pprint(compagnie.nombre_reglements_a_reverser_cie)
            if compagnie.nombre_reglements_a_reverser_cie == 0:
                compagnies = compagnies.exclude(id=compagnie.id)


        context = self.get_context_data(**kwargs)
        context['compagnies'] = compagnies

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required
def ajax_reglements_a_reverser_compagnie(request, compagnie_id):

    polices = Police.objects.filter(compagnie_id=compagnie_id)

    reglements_compagnies = ReglementReverseCompagnie.objects.filter(quittance__police__in=polices, statut_reversement_compagnie=StatutReversementCompagnie.NON_REVERSE, statut_validite=StatutValidite.VALIDE).exclude(quittance__nature_quittance__code="Ristourne").exclude(quittance__type_quittance__code="HONORAIRE")

    pprint(polices)

    return render(request, 'reglements_a_reverser_by_compagnie.html', {'reglements_compagnies':reglements_compagnies})


@login_required
def add_reglement_compagnie(request):

    if request.method == 'POST':
        devise = request.POST.get('devise')
        mode_reglement = request.POST.get('mode_reglement')
        compte_tresorerie = request.POST.get('compte_tresorerie')
        banque = request.POST.get('banque')
        numero_piece = request.POST.get('numero_piece')
        date_paiement = request.POST.get('date_paiement')
        reglement = request.POST.getlist('reglement')
        reglements_selectionnes = request.POST.getlist('reglement_selectionne')

        nature_operation_code = "REGCIE"
        nature_operation = NatureOperation.objects.filter(code=nature_operation_code).first()

        #enregistrer les infos dans operation
        nombre_quittances = 0
        montant_total_regle = 0
        operation = Operation.objects.create(nature_operation=nature_operation,
                                             numero_piece=numero_piece,
                                             montant_total=montant_total_regle,
                                             compte_tresorerie_id=compte_tresorerie,
                                             devise_id=devise,
                                             mode_reglement_id=mode_reglement,
                                             date_operation=date_paiement,
                                             created_by=request.user)
        operation.save()

        #enregistrer le details dans reglements (liste les quittances regl√©es avec chaque montant)
        nombre_reglements_selectionnes = 0
        montant_total_reglements_selectionne = 0
        i = 0
        for reglement_id in reglements_selectionnes:
            #reglement_id = reglement[i]
            i = i + 1
            print(reglement_id)
            if reglement_id is not None:
                reglement = Reglement.objects.get(id=reglement_id)
                reglement.statut_reversement_compagnie = StatutReversementCompagnie.REVERSE
                reglement.date_reversement_compagnie = datetime.datetime.now(tz=timezone.utc)
                reglement.save()

                devise = reglement.devise if devise is None else devise

                montant_total_reglements_selectionne+= reglement.montant
                nombre_reglements_selectionnes = nombre_reglements_selectionnes + 1

                #Lier l'op√©ration au r√®glement
                operation_reglement = OperationReglement.objects.create(operation=operation, reglement=reglement, created_by=request.user)
                operation_reglement.save()


        #mettre √† jour le total dans operation
        operation.montant_total = montant_total_reglements_selectionne
        operation.devise = devise
        operation.nombre_quittances = nombre_reglements_selectionnes
        operation.numero = 'OP' + str(Date.today().year) + str(operation.pk).zfill(6)
        operation.save()


        response = {
            'statut': 1,
            'message': "R√®glement effectu√© avec succ√®s !",
            'data': {'montant_total_reglements_selectionne': montant_total_reglements_selectionne, 'nombre_reglements_selectionnes': nombre_reglements_selectionnes, 'operation_id': operation.pk}
        }

        return JsonResponse(response)


    else:
        natures_operations = NatureOperation.objects.all()
        devises = Devise.objects.all()
        modes_reglements = ModeReglement.objects.all()
        banques = Banque.objects.filter(bureau=request.user.bureau).order_by('libelle')
        comptes_tresoreries = CompteTresorerie.objects.filter(code="REGCIE").order_by('libelle')
        reglements_compagnies = Reglement.objects.filter(statut_reversement_compagnie=StatutReversementCompagnie.NON_REVERSE, statut_validite=StatutValidite.VALIDE)

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')


        for compagnie in compagnies:
            if compagnie.nombre_reglements_a_reverser_cie == 0:
                compagnies = compagnies.exclude(id=compagnie.id)


        today = datetime.datetime.now(tz=timezone.utc)
        return render(request, 'modal_add_reglement_compagnie.html',
                      {'reglements_compagnies': reglements_compagnies,  'compagnies': compagnies, 'today': today, 'devises': devises, 'natures_operations': natures_operations, 'modes_reglements': modes_reglements, 'banques': banques, 'comptes_tresoreries': comptes_tresoreries, })



def generer_bordereau_reglement_compagnie_pdf(request, operation_id):
    operation = Operation.objects.get(id=operation_id)

    option_reglements = OperationReglement.objects.filter(operation=operation)

    compagnie = option_reglements.first().reglement.quittance.compagnie if option_reglements.first() and option_reglements.first().reglement and option_reglements.first().reglement.quittance else None
    bureau = option_reglements.first().reglement.bureau if option_reglements.first() and option_reglements.first().reglement else None

    # dd(option_reglements.first())
    total_montant_compagnie = 0
    total_montant_com_courtage = 0
    total_montant_com_gestion = 0
    total_montant_com_intermediaire = 0

    for option_reglement in option_reglements:
        total_montant_compagnie += option_reglement.reglement.montant_compagnie
        total_montant_com_courtage += option_reglement.reglement.montant_com_courtage
        total_montant_com_gestion += option_reglement.reglement.montant_com_gestion
        total_montant_com_intermediaire += option_reglement.reglement.montant_com_intermediaire

    contexte = {
        'operation': operation,
        'option_reglements': option_reglements,
        # 'nombre_pages': nombre_pages,
        'compagnie': compagnie,
        'bureau': bureau,
        'total_montant_compagnie': total_montant_compagnie,
        'total_montant_com_courtage': total_montant_com_courtage,
        'total_montant_com_gestion': total_montant_com_gestion,
        'total_montant_com_intermediaire': total_montant_com_intermediaire,
    }
    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', contexte)

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    #ajout du nombre de page obtenu au contexte pour le rendu final
    contexte['nombre_pages'] = nombre_pages
    pdf = render_pdf('courriers/bordereau_reglement_compagnie.html', contexte)

    # Update bordereau data and save
    #operation.fichier.save(f'bordereau_reglement_compagnie_{operation.numero}.pdf', File(pdf))
    #operation.save()


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')


@method_decorator(login_required, name='dispatch')
class EncaissementCommissionsView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/encaissement_commissions.html'
    model = Reglement

    def get(self, request, *args, **kwargs):

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')


        for compagnie in compagnies:
            if compagnie.nombre_reglements_a_recevoir_com == 0:
                compagnies = compagnies.exclude(id=compagnie.id)

        reglements_compagnies = Reglement.objects.filter(
            statut_reversement_compagnie=StatutReversementCompagnie.REVERSE,
            statut_commission=StatutEncaissementCommission.NON_ENCAISSEE, )

        context = self.get_context_data(**kwargs)
        context['compagnies'] = compagnies
        #context['reglements_compagnies'] = reglements_compagnies

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@method_decorator(login_required, name='dispatch')
class EncaissementCommissionsCourtGestView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/encaissement_commissions_court_gest.html'
    model = Reglement

    def get(self, request, *args, **kwargs):

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')

        type = self.kwargs.get('type')

        for compagnie in compagnies:
            if type == "courtage" and compagnie.nombre_reglements_a_recevoir_com_court == 0:
                compagnies = compagnies.exclude(id=compagnie.id)
            if type == "gestion" and compagnie.nombre_reglements_a_recevoir_com_gest == 0:
                compagnies = compagnies.exclude(id=compagnie.id)


        reglements_compagnies = Reglement.objects.filter(
            statut_reversement_compagnie=StatutReversementCompagnie.REVERSE,
            statut_commission=StatutEncaissementCommission.NON_ENCAISSEE, )

        context = self.get_context_data(**kwargs)
        context['compagnies'] = compagnies
        #context['reglements_compagnies'] = reglements_compagnies

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required
def ajax_encaissement_commissions(request, compagnie_id):

    polices = Police.objects.filter(compagnie_id=compagnie_id)

    reglements_compagnies = ReglementReverseCompagnie.objects.filter(quittance__police__in=polices, statut_reversement_compagnie=StatutReversementCompagnie.NON_REVERSE, statut_validite=StatutValidite.VALIDE)

    return render(request, 'reglements_a_reverser_by_compagnie.html', {'reglements_compagnies':reglements_compagnies})


@login_required
def ajax_reglements_reverses(request, compagnie_id):

    polices = Police.objects.filter(compagnie_id=compagnie_id)

    reglements_compagnies = ReglementReverseCompagnie.objects.filter(quittance__police__in=polices, statut_reversement_compagnie=StatutReversementCompagnie.REVERSE, statut_validite=StatutValidite.VALIDE).exclude(statut_commission=StatutEncaissementCommission.ENCAISSEE)

    pprint(reglements_compagnies)

    return render(request, 'reglements_reverses.html', {'reglements_compagnies':reglements_compagnies})


@login_required
def add_encaissement_commission(request):

    if request.method == 'POST':
        # dd(request)
        devise = request.POST.get('devise')
        mode_reglement = request.POST.get('mode_reglement')
        compte_tresorerie = request.POST.get('compte_tresorerie')
        #banque_id = request.POST.get('banque')
        banque_emettrice = request.POST.get('banque_emettrice')
        numero_piece = request.POST.get('numero_piece')
        date_paiement = request.POST.get('date_paiement')
        #reglement = request.POST.getlist('reglement')
        reglements_selectionnes = request.POST.getlist('reglement_selectionne')

        compte_difference = request.POST.get('compte_difference')
        debit_difference = request.POST.get('debit_difference').replace(" ", "")
        debit_difference = 0 if debit_difference == "" else float(debit_difference)
        credit_difference = request.POST.get('credit_difference').replace(" ", "")
        credit_difference = 0 if credit_difference == "" else float(credit_difference)
        libelle_difference = request.POST.get('libelle_difference')

        #dd(request.POST)
        #dd(montant_encaisse_court_selectionnes[0])

        nature_operation_code = "ENCCOM"
        nature_operation = NatureOperation.objects.filter(code=nature_operation_code).first()

        #compte COMPTABLE si utilis√© dans l'operation
        compte_comptable = CompteComptable.objects.filter(id=compte_difference).first() if compte_difference and compte_difference != "" else None

        #banque si utilis√©e
        # banque = Banque.objects.filter(id=banque_id).first() if banque_id and banque_id != "" else None

        #enregistrer les infos dans operation
        nombre_quittances = 0
        montant_total_regle = request.POST.get('montant_total_regle').replace(" ", "")
        operation = Operation.objects.create(nature_operation=nature_operation,
                                             numero_piece=numero_piece,
                                             montant_total=montant_total_regle,
                                             compte_tresorerie_id=compte_tresorerie,
                                             devise_id=devise,
                                             #banque=banque,
                                             banque_emettrice=banque_emettrice,
                                             mode_reglement_id=mode_reglement,
                                             date_operation=date_paiement,
                                             created_by=request.user)
        operation.save()

        #enregistrer le details dans reglements (liste les quittances regl√©es avec chaque montant)
        nombre_reglements_selectionnes = 0
        montant_total_reglements_selectionne = 0
        i = 0
        for reglement_id in reglements_selectionnes:

            montant_encaisse_court_selectionne = request.POST.get('montant_encaisse_court{}'.format(reglement_id))
            montant_encaisse_gest_selectionne = request.POST.get('montant_encaisse_gest{}'.format(reglement_id))

            montant_com_courtage = montant_encaisse_court_selectionne.replace(" ", "")
            montant_com_courtage = 0 if montant_com_courtage == "" else float(montant_com_courtage)
            montant_com_gestion = montant_encaisse_gest_selectionne.replace(" ", "")
            montant_com_gestion = 0 if montant_com_gestion == "" else float(montant_com_gestion)
            #reglement_id = reglement[i]
            i = i + 1
            #print(reglement_id)
            if reglement_id is not None:
                reglement = Reglement.objects.get(id=reglement_id)
                montant_total_reglements_selectionne += montant_com_courtage + montant_com_gestion
                nombre_reglements_selectionnes = nombre_reglements_selectionnes + 1

                """                 #Lier l'op√©ration au r√®glement
                operation_reglement = OperationReglement.objects.create(operation=operation, reglement=reglement, created_by=request.user)
                operation_reglement.save()
                """
                encaiss_com = EncaissementCommission.objects.create(operation=operation,
                                                             reglement=reglement,
                                                             created_by=request.user,
                                                             montant_com_courtage=montant_com_courtage,
                                                             montant_com_gestion=montant_com_gestion)
                encaiss_com.save()

                # application du debit au reste si operation correspond par le journal comptabale
                if debit_difference > 0 and debit_difference == reglement.montant_com_solde() and compte_comptable:
                    journal = Journal.objects.create(bureau=request.user.bureau,
                                                  compte_comptable=compte_comptable,
                                                  encaissement_commission=encaiss_com,
                                                  sens="D",
                                                  created_by=request.user,
                                                  montant=debit_difference,
                                                  designation=libelle_difference)
                    journal.save()

                # on constate l'encaissement total pour mettre a jour ledit statut
                if reglement.etat_encaisse() == True:
                    reglement.statut_commission = StatutEncaissementCommission.ENCAISSEE
                    reglement.save()

                nombre_reglements_selectionnes = i
                devise = reglement.devise

        # application d'un credit par le journal si l'operation est concern√©e par un depassement 
        if credit_difference > 0 and reglement_id and encaiss_com and compte_comptable:
            journal = Journal.objects.create(bureau=request.user.bureau,
                                            compte_comptable=compte_comptable,
                                            encaissement_commission=encaiss_com,
                                            sens="C",
                                            created_by=request.user,
                                            montant=debit_difference,
                                            designation=libelle_difference)
            journal.save()

            montant_total_reglements_selectionne += journal.montant


        #mettre √† jour le total dans operation
        operation.montant_total = montant_total_reglements_selectionne
        operation.nombre_quittances = nombre_reglements_selectionnes
        operation.numero = 'OP' + str(Date.today().year) + str(operation.pk).zfill(6)
        operation.devise = devise
        operation.save()

        pdf_url = reverse('generer_bordereau_encaissement_compagnie_pdf', args=[operation.pk])

        response = {
            'statut': 1,
            'message': "Encaissement effectu√© avec succ√®s !",
            'data': {'montant_total_reglements_selectionne': montant_total_reglements_selectionne, 'nombre_reglements_selectionnes': nombre_reglements_selectionnes, 'operation_id': operation.pk, 'pdf_url': pdf_url}
        }

        return JsonResponse(response)


    else:
        natures_operations = NatureOperation.objects.all()
        devises = Devise.objects.all()
        modes_reglements = ModeReglement.objects.all()
        banques = Banque.objects.filter(bureau=request.user.bureau).order_by('libelle')
        comptes_tresoreries = CompteTresorerie.objects.exclude(code="REGCIE").order_by('libelle') #filter(code="REGCIE").order_by('libelle')
        reglements_compagnies = Reglement.objects.filter(statut_reversement_compagnie=StatutReversementCompagnie.REVERSE, statut_validite=StatutValidite.VALIDE).exclude(statut_commission=StatutEncaissementCommission.NON_ENCAISSEE)

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')

        comptes_exercices = CompteComptable.objects.all()


        for compagnie in compagnies:
            if compagnie.nombre_reglements_a_recevoir_com == 0:
                compagnies = compagnies.exclude(id=compagnie.id)


        today = datetime.datetime.now(tz=timezone.utc)
        return render(request, 'modal_add_encaissement_commission.html',
                      {'reglements_compagnies': reglements_compagnies,  'compagnies': compagnies, 'today': today, 'devises': devises, 'natures_operations': natures_operations, 'modes_reglements': modes_reglements, 'banques': banques, 'comptes_tresoreries': comptes_tresoreries, 'comptes_exercices': comptes_exercices,})


@login_required
def ajax_reglements_reverses_court_gest(request, compagnie_id, type):

    polices = Police.objects.filter(compagnie_id=compagnie_id)

    reglements_compagnies = ReglementReverseCompagnie.objects.filter(quittance__police__in=polices, statut_reversement_compagnie=StatutReversementCompagnie.REVERSE, statut_validite=StatutValidite.VALIDE).exclude(statut_commission=StatutEncaissementCommission.ENCAISSEE)

    for reglement_compagnie in reglements_compagnies:
        #dd(reglement_compagnie)
        #reglements_compagnies.exclude(id=reglement_compagnie.id)
        #print(f"view 1 - {reglement_compagnie.montant_com_courtage_solde()}")
        #print(f"view 2 - {reglement_compagnie.montant_journal_debit() - reglement_compagnie.montant_journal_credit()}")
        if type == "courtage":
            if reglement_compagnie.montant_com_courtage_solde() != 0 and reglement_compagnie.montant_com_courtage_solde() != (reglement_compagnie.montant_journal_debit_courtage() - reglement_compagnie.montant_journal_credit_courtage()):
                pass
            else:
                reglements_compagnies = reglements_compagnies.exclude(id=reglement_compagnie.id)
        else:
            if reglement_compagnie.montant_com_gestion_solde() != 0 and reglement_compagnie.montant_com_gestion_solde() != (reglement_compagnie.montant_journal_debit_gestion() - reglement_compagnie.montant_journal_credit_gestion()):
                pass
            else:
                reglements_compagnies = reglements_compagnies.exclude(id=reglement_compagnie.id)

    # pprint(reglements_compagnies)

    return render(request, 'reglements_reverses_court_gest.html', {'reglements_compagnies':reglements_compagnies, 'type':type})


@login_required
def add_encaissement_com_court_gest(request, type):

    if request.method == 'POST':
        # dd(request)
        devise = request.POST.get('devise')
        mode_reglement = request.POST.get('mode_reglement')
        compte_tresorerie = request.POST.get('compte_tresorerie')
        #banque_id = request.POST.get('banque')
        banque_emettrice = request.POST.get('banque_emettrice')
        numero_piece = request.POST.get('numero_piece')
        date_paiement = request.POST.get('date_paiement')
        #reglement = request.POST.getlist('reglement')
        reglements_selectionnes = request.POST.getlist('reglement_selectionne')

        compte_difference = request.POST.get('compte_difference')
        debit_difference = request.POST.get('debit_difference').replace(" ", "")
        debit_difference = 0 if debit_difference == "" else float(debit_difference)
        credit_difference = request.POST.get('credit_difference').replace(" ", "")
        credit_difference = 0 if credit_difference == "" else float(credit_difference)
        libelle_difference = request.POST.get('libelle_difference')

        #dd(request.POST)
        #dd(montant_encaisse_court_selectionnes[0])

        nature_operation_code = "ENCCOM"
        nature_operation = NatureOperation.objects.filter(code=nature_operation_code).first()

        #compte COMPTABLE si utilis√© dans l'operation
        compte_comptable = CompteComptable.objects.filter(id=compte_difference).first() if compte_difference and compte_difference != "" else None

        #banque si utilis√©e
        # banque = Banque.objects.filter(id=banque_id).first() if banque_id and banque_id != "" else None

        #enregistrer les infos dans operation
        nombre_quittances = 0
        montant_total_regle = request.POST.get('montant_total_regle').replace(" ", "")
        operation = Operation.objects.create(nature_operation=nature_operation,
                                             numero_piece=numero_piece,
                                             montant_total=montant_total_regle,
                                             compte_tresorerie_id=compte_tresorerie,
                                             devise_id=devise,
                                             #banque=banque,
                                             banque_emettrice=banque_emettrice,
                                             mode_reglement_id=mode_reglement,
                                             date_operation=date_paiement,
                                             created_by=request.user)
        operation.save()

        #enregistrer le details dans reglements (liste les quittances regl√©es avec chaque montant)
        nombre_reglements_selectionnes = 0
        montant_total_reglements_selectionne = 0
        i = 0
        already_credited = False
        already_debited = False
        for reglement_id in reglements_selectionnes:

            montant_encaisse_court_selectionne = request.POST.get('montant_encaisse_court{}'.format(reglement_id)) if type == "courtage" else 0
            montant_encaisse_gest_selectionne = request.POST.get('montant_encaisse_gest{}'.format(reglement_id)) if type == "gestion" else 0

            montant_com_courtage = montant_encaisse_court_selectionne.replace(" ", "") if type == "courtage" else 0
            montant_com_courtage = 0 if montant_com_courtage == "" else float(montant_com_courtage)
            montant_com_gestion = montant_encaisse_gest_selectionne.replace(" ", "") if type == "gestion" else 0
            montant_com_gestion = 0 if montant_com_gestion == "" else float(montant_com_gestion)
            type_commission = "COURTAGE" if type == "courtage" else "GESTION"
            #reglement_id = reglement[i]
            i = i + 1
            #print(reglement_id)
            if reglement_id is not None:
                reglement = Reglement.objects.get(id=reglement_id)
                montant_total_reglements_selectionne += montant_com_courtage + montant_com_gestion
                nombre_reglements_selectionnes = nombre_reglements_selectionnes + 1

                """                 #Lier l'op√©ration au r√®glement
                operation_reglement = OperationReglement.objects.create(operation=operation, reglement=reglement, created_by=request.user)
                operation_reglement.save()
                """
                encaiss_com = EncaissementCommission.objects.create(operation=operation,
                                                             reglement=reglement,
                                                             created_by=request.user,
                                                             montant_com_courtage=montant_com_courtage,
                                                             montant_com_gestion=montant_com_gestion,
                                                             type_commission=type_commission)
                encaiss_com.save()

                # application du debit au reste si operation correspond par le journal comptabale
                if debit_difference > 0 and already_debited == False and ((debit_difference == reglement.montant_com_courtage_solde() and type == "courtage") or (debit_difference == reglement.montant_com_gestion_solde() and type == "gestion")) and compte_comptable:
                    journal = Journal.objects.create(bureau=request.user.bureau,
                                                  compte_comptable=compte_comptable,
                                                  encaissement_commission=encaiss_com,
                                                  sens="D",
                                                  created_by=request.user,
                                                  montant=debit_difference,
                                                  designation=libelle_difference)
                    journal.save()
                    already_debited = True

                # application d'un credit par le journal si l'operation est concern√©e par un depassement
                if credit_difference > 0 and already_credited == False and ((credit_difference == ((-1)*reglement.montant_com_courtage_solde()) and type == "courtage") or (((-1)*credit_difference == reglement.montant_com_gestion_solde()) and type == "gestion")) and compte_comptable:
                    journal = Journal.objects.create(bureau=request.user.bureau,
                                                    compte_comptable=compte_comptable,
                                                    encaissement_commission=encaiss_com,
                                                    sens="C",
                                                    created_by=request.user,
                                                    montant=credit_difference,
                                                    designation=libelle_difference)
                    journal.save()
                    already_credited = True
            
                # on constate l'encaissement total pour mettre a jour ledit statut
                if reglement.etat_encaisse() == True:
                    reglement.statut_commission = StatutEncaissementCommission.ENCAISSEE
                    reglement.save()

                nombre_reglements_selectionnes = i
                devise = reglement.devise

        # application d'un credit par le journal si l'operation est concern√©e par un depassement
       
        if credit_difference > 0 and already_credited == False and reglement_id and encaiss_com and compte_comptable:
            journal = Journal.objects.create(bureau=request.user.bureau,
                                            compte_comptable=compte_comptable,
                                            encaissement_commission=encaiss_com,
                                            sens="C",
                                            created_by=request.user,
                                            montant=debit_difference,
                                            designation=libelle_difference)
            journal.save()

            montant_total_reglements_selectionne += journal.montant
        


        #mettre √† jour le total dans operation
        operation.montant_total = montant_total_reglements_selectionne
        operation.nombre_quittances = nombre_reglements_selectionnes
        operation.numero = 'OP' + str(Date.today().year) + str(operation.pk).zfill(6)
        operation.devise = devise
        operation.save()

        pdf_url = reverse('generer_bordereau_encaissement_compagnie_pdf', args=[operation.pk])+"?type="+type

        response = {
            'statut': 1,
            'message': "Encaissement effectu√© avec succ√®s !",
            'data': {'montant_total_reglements_selectionne': montant_total_reglements_selectionne, 'nombre_reglements_selectionnes': nombre_reglements_selectionnes, 'operation_id': operation.pk, 'pdf_url': pdf_url}
        }

        return JsonResponse(response)


    else:
        natures_operations = NatureOperation.objects.all()
        devises = Devise.objects.all()
        modes_reglements = ModeReglement.objects.all()
        banques = Banque.objects.filter(bureau=request.user.bureau).order_by('libelle')
        comptes_tresoreries = CompteTresorerie.objects.exclude(code="REGCIE").order_by('libelle') #filter(code="REGCIE").order_by('libelle')
        reglements_compagnies = Reglement.objects.filter(statut_reversement_compagnie=StatutReversementCompagnie.REVERSE, statut_validite=StatutValidite.VALIDE).exclude(statut_commission=StatutEncaissementCommission.NON_ENCAISSEE)

        compagnies = Compagnie.objects.filter(bureau=request.user.bureau).order_by('nom')

        comptes_exercices = CompteComptable.objects.all()


        for compagnie in compagnies:
            if type == "courtage" and compagnie.nombre_reglements_a_recevoir_com_court == 0:
                compagnies = compagnies.exclude(id=compagnie.id)
            if type == "gestion" and compagnie.nombre_reglements_a_recevoir_com_gest == 0:
                compagnies = compagnies.exclude(id=compagnie.id)

        today = datetime.datetime.now(tz=timezone.utc)

        return render(request, 'modal_add_encaissement_com_court_gest.html',
                      {'reglements_compagnies': reglements_compagnies,  'compagnies': compagnies, 'today': today, 'devises': devises, 'natures_operations': natures_operations, 'modes_reglements': modes_reglements, 'banques': banques, 'comptes_tresoreries': comptes_tresoreries, 'comptes_exercices': comptes_exercices, 'type': type})


# a la fois pour le courtage ou la gestion determiner par ?type en get
def generer_bordereau_encaissement_compagnie_pdf(request, operation_id):
    operation = Operation.objects.get(id=operation_id)

    encaissement_commissions = EncaissementCommission.objects.filter(operation=operation)

    compagnie = encaissement_commissions.first().reglement.quittance.compagnie if encaissement_commissions.first() and encaissement_commissions.first().reglement and encaissement_commissions.first().reglement.quittance else None
    bureau = encaissement_commissions.first().reglement.bureau if encaissement_commissions.first() and encaissement_commissions.first().reglement else None

    # dd(option_reglements.first())
    total_montant_compagnie = 0
    total_montant_com_courtage = 0
    total_montant_com_gestion = 0
    total_montant_com_intermediaire = 0
    total_prime_ttc = 0
    total_montant_com_encaisse = 0
    op_div = 0
    op_sens = None
    op_designation = None

    type = request.GET.get('type', None)

    for encaissement_commission in encaissement_commissions:
        total_montant_compagnie += encaissement_commission.reglement.montant_compagnie
        total_montant_com_courtage += encaissement_commission.montant_com_courtage
        total_montant_com_gestion += encaissement_commission.montant_com_gestion
        total_prime_ttc += encaissement_commission.reglement.quittance.prime_ttc
        total_montant_com_encaisse += encaissement_commission.montant()
        for journal in encaissement_commission.journals.all():
            op_div = journal.montant
            op_designation = journal.designation
            if journal.sens == "D":
                op_sens = "D"
            else:
                op_sens = "C"
        # total_montant_com_intermediaire += encaissement_commission.montant_com_intermediaire

    total_montant_percu_final = total_montant_com_encaisse - op_div if op_sens == "D" else total_montant_com_encaisse + op_div

    contexte = {
        'operation': operation,
        'encaissement_commissions': encaissement_commissions,
        # 'nombre_pages': nombre_pages,
        'compagnie': compagnie,
        'bureau': bureau,
        'total_montant_compagnie': total_montant_compagnie,
        'total_montant_com_courtage': total_montant_com_courtage,
        'total_montant_com_gestion': total_montant_com_gestion,
        'total_prime_ttc': total_prime_ttc,
        'total_montant_com_encaisse': total_montant_com_encaisse,
        'op_div': op_div,
        'op_sens': op_sens,
        'op_designation': op_designation,
        'total_montant_percu_final': total_montant_percu_final,
        'type': type,
        # 'total_montant_com_intermediaire': total_montant_com_intermediaire,
    }
    pdf = render_pdf('courriers/bordereau_encaissement_compagnie.html', contexte) if type is None else render_pdf('courriers/bordereau_encaissement_compagnie_court_gest.html', contexte)

    pdf_file = PyPDF2.PdfReader(pdf)
    nombre_pages = len(pdf_file.pages)

    #ajout du nombre de page obtenu au contexte pour le rendu final
    contexte['nombre_pages'] = nombre_pages
    pdf = render_pdf('courriers/bordereau_encaissement_compagnie.html', contexte) if type is None else render_pdf('courriers/bordereau_encaissement_compagnie_court_gest.html', contexte)

    # Update bordereau data and save
    operation.fichier.save(f'bordereau_encaissement_{type if type is not None else ""}_compagnie{operation.numero}.pdf', File(pdf))
    operation.save()


    #return pdf

    #AFFICHER DIRECTEMENT
    return HttpResponse(File(pdf), content_type='application/pdf')



@method_decorator(login_required, name='dispatch')
class ReglementApporteursView(TemplateView):
    # permission_required = "comptabilite.view_reglement"
    template_name = 'comptabilite/reglements_apporteurs.html'
    model = Reglement

    def get(self, request, *args, **kwargs):

        reglements_compagnies = Reglement.objects.filter(
            statut_reversement_compagnie=StatutReversementCompagnie.REVERSE,
            statut_commission=StatutEncaissementCommission.ENCAISSEE,
            statut_reglement_apporteurs=StatutReglementApporteurs.NON_REGLE)
        compagnies = Compagnie.objects.all().order_by('nom')

        context = self.get_context_data(**kwargs)
        context['compagnies'] = compagnies
        context['reglements_compagnies'] = reglements_compagnies

        return self.render_to_response(context)


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def get_montant_caution(bureau, compagnie=None):
    queryset = Caution.par_bureau(bureau).filter(date_fin_effet__isnull=True, status=True)
    if compagnie:
       queryset = queryset.filter(compagnie=compagnie)

    mcaution = queryset.aggregate(mcaution=Sum('montant'))['mcaution'] or 0
    return mcaution

    # return queryset.annotate(
    #     montant_caution=Sum('montant'),
    #     created_at_caution=Min('caution__created_at')
    # )

def get_montant_sinistre_regle(bureau, compagnie=None, month=None):
    queryset = Sinistre.objects.filter(
        prestataire__bureau=bureau,
        statut_paiement=StatutPaiementSinistre.PAYE,
        statut_validite=StatutValidite.VALIDE,
        dossier_sinistre_id__isnull=False,
        facture_compagnie_id__isnull=True,
    )

    if month:
        queryset = queryset.filter(date_paiement__month=month)

    if compagnie:
        queryset = queryset.filter(facture_compagnie__compagnie=compagnie)

    return queryset.aggregate(total_remboursement=Sum('montant_remboursement_accepte'))['total_remboursement'] or 0

def get_montant_sinistre_reclame(bureau, compagnie=None):
    queryset = FactureCompagnie.par_bureau(bureau)
    if compagnie:
        queryset = queryset.filter(compagnie=compagnie)
    return queryset.aggregate(total_restant=Sum('montant_restant'))['total_restant'] or 0

def prepare_camembert_data(bureau, compagnie=None):

    # Preparer les donn√©es pour le camembert
    label_global_fdr = "FDR"
    valeur_global_fdr = get_montant_caution(bureau, compagnie)
    couleur_global_fdr = "#B7482B"
    stroke_couleur_global_fdr = "#B7482B"

    label_global_sinistre_regle = "Sinistres r√©gl√©s"
    valeur_global_sinistre_regle = int(get_montant_sinistre_regle(bureau=bureau, compagnie=compagnie))
    couleur_global_sinistre_regle = "#B8B8B8"
    stroke_couleur_global_sinistre_regle = "#B8B8B8"

    label_global_sinistre_reclame = "Sinistres r√©clam√©s"
    valeur_global_sinistre_reclame = int(get_montant_sinistre_reclame(bureau=bureau,compagnie=compagnie))
    couleur_global_sinistre_reclame = "orange"
    stroke_couleur_global_sinistre_reclame = "orange"

    return {
        "label_global_fdr": label_global_fdr,
        "valeur_global_fdr": valeur_global_fdr,
        "couleur_global_fdr": couleur_global_fdr,
        "stroke_couleur_global_fdr": stroke_couleur_global_fdr,
        "label_global_sinistre_regle": label_global_sinistre_regle,
        "valeur_global_sinistre_regle": valeur_global_sinistre_regle,
        "couleur_global_sinistre_regle": couleur_global_sinistre_regle,
        "stroke_couleur_global_sinistre_regle": stroke_couleur_global_sinistre_regle,
        "label_global_sinistre_reclame": label_global_sinistre_reclame,
        "valeur_global_sinistre_reclame": valeur_global_sinistre_reclame,
        "couleur_global_sinistre_reclame": couleur_global_sinistre_reclame,
        "stroke_couleur_global_sinistre_reclame": stroke_couleur_global_sinistre_reclame,
    }

def get_camembert_data_detail_par_garant(request, compagnie_id):

    compagnie_id = compagnie_id
    compagnie = Compagnie.objects.get(id=compagnie_id)


    bureau = compagnie.bureau if compagnie else None;

    data = prepare_camembert_data(bureau, compagnie)

    pie_data = [
        {
            "label": data["label_global_fdr"],
            "value": data["valeur_global_fdr"],
            "color": data["couleur_global_fdr"]
        },
        {
            "label": data["label_global_sinistre_regle"],
            "value": data["valeur_global_sinistre_regle"],
            "color": data["couleur_global_sinistre_regle"]
        },
        {
            "label": data["label_global_sinistre_reclame"],
            "value": data["valeur_global_sinistre_reclame"],
            "color": data["couleur_global_sinistre_reclame"]
        }
    ]

    return JsonResponse({"pieData": pie_data})


def get_sum_fdr_per_month(bureau, month, compagnie=None):
    queryset = Caution.par_bureau(bureau).filter(date_fin_effet__isnull=True, status=True)
    if compagnie:
        queryset = queryset.filter(compagnie=compagnie)

    total_montant = queryset.filter(
        created_at__month=month,
        # created_at__year=year
    ).aggregate(
        total_montant=Sum('montant')
    )['total_montant']

    return total_montant if total_montant else 0



def get_consumption_per_month(bureau, month, compagnie=None):
    queryset_sinistres_reclames_non_regles = FactureCompagnie.par_bureau(bureau)
    queryset_sinistres_regles = Sinistre.par_bureau(bureau)

    if compagnie:
        queryset_sinistres_reclames_non_regles = queryset_sinistres_reclames_non_regles.filter(compagnie=compagnie)
        queryset_sinistres_regles = queryset_sinistres_regles.filter(compagnie=compagnie)

    # sinistres a refacturer
    sinistre_regles = queryset_sinistres_regles.filter(
                date_paiement__month=month,
                statut_paiement=StatutPaiementSinistre.PAYE,
                statut_validite=StatutValidite.VALIDE,
                dossier_sinistre_id__isnull=False,
                # facture_compagnie_id__isnull=False
            ).aggregate(montant_remboursement_accepte=Sum('montant_remboursement_accepte'))['montant_remboursement_accepte'] or 0

    # en gros  total refactures
    sinistres_reclames = queryset_sinistres_reclames_non_regles.filter(
                created_at__month=month,
                statut=StatutFacture.NON_SOLDE,
            ).aggregate(montant_restant=Sum('montant_restant'))['montant_restant'] or 0


    depenses_total_par_mois = int(sinistre_regles + sinistres_reclames)

    return depenses_total_par_mois if depenses_total_par_mois else 0

def get_treso_per_month(bureau, month, compagnie=None):


    return 0


def prepare_chart_bar_data(bureau, compagnie=None):

    current_month = datetime.datetime.now().month  # Mois courant (1 pour janvier, 12 pour d√©cembre)

    # R√©cup√©rer les donn√©es pour chaque mois
    months = [calendar.month_name[i] for i in range(1, current_month + 1)]  # ["January", "February", ..., "June"]

    # Convertir les mois en fran√ßais si n√©cessaire (certaines configurations peuvent ne pas utiliser locale)
    months_fr = [calendar.month_name[i].capitalize() for i in range(1, current_month + 1)]

    sum_fdr_data = []
    sum_consumption_data = []
    sum_tresorerie_data = []

    for month in range(1, current_month + 1):
        sum_fdr = get_sum_fdr_per_month(bureau=bureau, compagnie=compagnie, month=month)
        sum_consumption = get_consumption_per_month(bureau=bureau, compagnie=compagnie, month=month)


        # sum_tresorerie = get_treso_per_month(bureau=bureau, month=month)

        sum_fdr_data.append(sum_fdr)
        sum_consumption_data.append(sum_consumption)
        sum_tresorerie_data.append(0)
        # sum_tresorerie_data.append(sum_fdr-sum_consumption if sum_fdr > sum_consumption else 0)

    #sum_tresorerie en faisant la diff√©rence entre sum_fdr et sum_consumption
    # sum_tresorerie_data = [d1 - d2 for d1, d2 in zip(sum_fdr_data, sum_consumption_data)]
    sum_tresorerie_data = [d1 - d2 for d1, d2 in zip(sum_fdr_data, sum_consumption_data)]

    # dd(months)
    context = {
        'months': months_fr,  # Mois en fran√ßais
        'months': months,
        'sum_fdr_data': sum_fdr_data,
        'sum_consumption_data': sum_consumption_data,
        'sum_tresorerie_data': sum_tresorerie_data,
        'sum_fdr_label': 'FDR',
        'sum_fdr_color': "#FFEA88",
        'sum_fdr_stroke_color': "#FFEA88",

        'sum_consumption_label': 'Conso',
        'sum_consumption_color': "orange",
        'sum_consumption_stroke_color': "orange",

        'sum_tresorerie_label': 'Tr√©so',
        'sum_tresorerie_color': "#a3dc33",
        'sum_tresorerie_stroke_color': "#b4ed44",
    }
    return context

def prepare_chart_line_data(bureau, compagnie=None):

    sin_regles_data = []
    current_month = datetime.datetime.now().month

    # R√©cup√©rer les donn√©es pour chaque mois
    months = [calendar.month_name[i] for i in range(1, current_month + 1)]
    months_fr = [calendar.month_name[i].capitalize() for i in range(1, current_month + 1)]

    for month in range(1, current_month + 1):
        sini_regles = int(get_montant_sinistre_regle(bureau=bureau, compagnie=compagnie, month=month))


        # sum_tresorerie = get_treso_per_month(bureau=bureau, month=month)

        sin_regles_data.append(sini_regles)

    # dd(months)
    context = {
        'months': months_fr,  # Mois en fran√ßais
        'sin_regles_data': sin_regles_data,
        # 'couleur_remplissage_sin_regles': "#FFEA88",
        # "stroke_color_sin_regles": "#B7482B",
        # "point_stroke_color_sin_regles": "#a3dc33",


        # 'couleur_remplissage_sin_regles': "#aFEA88",
        # "stroke_color_sin_regles": "#B7482B",
        # "point_stroke_color_sin_regles": "#FFEA88",

        'couleur_remplissage_sin_regles': "#9ACD32",  # Vert Jaune P√¢le
        # 'stroke_color_sin_regles': "#228B22",  # Vert For√™t
        'point_stroke_color_sin_regles': "#FFD700",  # Or
    }

    return context

@login_required
def ajax_reglements_apporteurs(request, compagnie_id):

    polices = Police.objects.filter(compagnie_id=compagnie_id)

    reglements_compagnies = ReglementReverseCompagnie.objects.filter(quittance__police__in=polices, statut_reversement_compagnie=StatutReversementCompagnie.NON_REVERSE, statut_validite=StatutValidite.VALIDE)

    return render(request, 'reglements_a_reverser_by_compagnie.html', {'reglements_compagnies':reglements_compagnies})




@method_decorator(login_required, name='dispatch')
class ExecutionRequeteExcelComptaView(TemplateView):
    template_name = 'comptabilite/execution_requete_excel_compta.html'
    model = Sinistre

    def get(self, request, *args, **kwargs):
        #TODO , filtrer sur le bureau : prestataire__bureau=request.user.bureau
        periode_comptable = PeriodeComptable.objects.all()
        query_datas = [
            {
                "query_label": "ANALYSE D√âTAILL√âE DES PRIMES COMPTA",
                "query_name": "ANALYSE_PRIMES",
                "query_param_data": periode_comptable,
                "query_param_label" :"P√©riode comptable",
                "query_param_name" :"period_comptable",
                "query_param_type" :"select",
                "query_param_isRequired" : False
            },
            {
                "query_label": "ANALYSE D√âTAILL√âE DES PRIMES COMPTA AVEC APPORTEURS",
                "query_name": "ANALYSE_PRIMES_APPORTEURS",
                "query_param_data": periode_comptable,
                "query_param_label": "P√©riode comptable",
                "query_param_name": "period_comptable",
                "query_param_type": "select",
                "query_param_isRequired": False
            },
            {
                "query_label": "PAIEMENT SINISTRE SANTE ENTRE DEUX DATES",
                "query_name": "PAIEMENT_SINISTRE_ENTRE_DEUX_DATES",
                "query_param_label" :"PAIEMENT SINISTRES SANT√â ENTRE DEUX DATES",
                "query_param_name" :"paiement_sinistre_sante_entre_deux_dates",
                "query_param_type" :"date",
                "query_param_isRequired" : True
            }
        ]
        #query_datas = []
        context = self.get_context_data(**kwargs)
        context['query_datas'] = query_datas
        context['periode_comptable'] = periode_comptable
        context['prestataires'] = None

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        print("----- fn= post -----")
        print(request.POST)
        query_name = request.POST.get('query_name')

        date_debut_paiment_sinisre = request.POST.get("date_debut_paiment_sinisre")
        date_fin_paiment_sinisre = request.POST.get("date_fin_paiment_sinisre")

        if query_name == "ANALYSE_PRIMES":

            # save log
            response = requete_analyse_prime_compta(request)
            return response

        elif query_name == "ANALYSE_PRIMES_APPORTEURS":
            # save log
            response = requete_analyse_prime_compta_apporteur(request)
            return response

        elif query_name == "PAIEMENT_SINISTRE_ENTRE_DEUX_DATES":

            reference_facture = request.POST.get("reference_facture_paiment_sinisre")
            numero_police = request.POST.get("numero_police")

            title = "PAIEMENT_SINISTRE_ENTRE_2_DATES"

            query = requete_liste_paiement_sinistre_sante_entre_deux_dates(request.user.bureau.code, date_debut_paiment_sinisre, date_fin_paiment_sinisre, reference_facture, numero_police)

            # creation de tache background
            create_query_background_task(name=f'PAIEMENT_SINISTRE_ENTRE_DEUX_DATES {date_debut_paiment_sinisre} _ {date_fin_paiment_sinisre}', query=query, request=request)
            queryset, header = execute_query(query)

            pprint("query")
            pprint(query)


            return exportation_en_excel_avec_style(title, header, queryset)

        else:
            return JsonResponse({
                "message": "La requ√™te n'est pas prise en charge"
            }, status=404)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }







def alert_consumption():
    try:
        bureaux = Bureau.objects.filter(mailinglist__statut=True).distinct()
    except Exception as e:
        print(f"Erreur lors de la r√©cup√©ration des bureaux : {e}")
        return False

    holding_mail = None
    try:
        holding_mail = MailingList.objects.filter(type_alerte=TypeAlerte.HOLDING).first()
    except Exception as e:
        print(f"Erreur lors de la r√©cup√©ration des e-mails de holding : {e}")

    for bureau in bureaux:
        garants = Compagnie.par_bureau(bureau=bureau)
        alert_list = []
        dest_mail = []

        for compagnie in garants:
            try:
                dispo = 0

                caution = Caution.par_bureau(bureau).filter(compagnie=compagnie, date_fin_effet__isnull=True, status=True).first()
                mcaution = caution.montant if caution else 0
                
                montant_regle_non_reclame = Sinistre.objects.filter(
                    prestataire__bureau=bureau,
                    statut_paiement=StatutPaiementSinistre.PAYE,
                    statut_validite=StatutValidite.VALIDE,
                    dossier_sinistre_id__isnull=False,
                    facture_compagnie_id__isnull=True,
                    compagnie=compagnie,
                ).aggregate(total_paye=Sum('montant_remboursement_accepte'))['total_paye'] or 0

                montant_reclame_non_regle_par_la_compagnie = FactureCompagnie.objects.filter(
                    compagnie=compagnie,
                    statut=StatutFacture.NON_SOLDE
                ).aggregate(total_montant_restant=Sum('montant_restant'))['total_montant_restant'] or 0

                tresorerie = mcaution - (montant_regle_non_reclame + montant_reclame_non_regle_par_la_compagnie)
                
                if mcaution != 0:
                    dispo = int((tresorerie / mcaution) * 100)
                
                consommation = dispo

                if consommation >= 60:  # Si la consommation atteint ou d√©passe 60%
                    alert_list.append({
                        'compagnie': compagnie.nom,
                        'consommation': consommation,  # Conserver comme nombre pour faciliter la comparaison
                        'bureau': bureau.nom,
                    })
            except Exception as e:
                print(f"Erreur lors du traitement de la compagnie {compagnie.nom} pour le bureau {bureau.nom} : {e}")

        if alert_list:
            subject = "Alerte : Niveau de consommation des garants"
            message = "Bonjour,\n\nVoici la liste des garants dont le niveau de consommation a atteint la limite :\n\n"
            
            for alert in alert_list:
                message += f"- Compagnie: {alert['compagnie']}, Bureau: {alert['bureau']}, Consommation: {alert['consommation']}%\n"
            
            message += "\nCordialement,\nL'application Sant√© V2"
            
            # Ajouter l'adresse e-mail du PDG si consommation d√©passe 90%
            if any(alert['consommation'] >= 90 for alert in alert_list):
                if holding_mail and holding_mail.mail_de_diffusion:
                    if holding_mail.mail_de_diffusion not in dest_mail:
                        dest_mail.append(holding_mail.mail_de_diffusion)
            
            # Ajouter l'adresse e-mail g√©n√©rale pour les alertes de consommation >= 60%
            if settings.DEFAULT_ALERT_EMAIL and settings.DEFAULT_ALERT_EMAIL not in dest_mail:
                dest_mail.append(settings.DEFAULT_ALERT_EMAIL)

            # Envoyer l'e-mail si des adresses sont d√©finies
            if dest_mail:
                try:
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        dest_mail,
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Erreur lors de l'envoi de l'e-mail : {e}")

    return True














def create_periode_comptable(request):
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')
    email = 'equipedev.os@inov.africa'
    cron_name = "CR√âATION DE LA P√âRIODE COMPTABLE"
    frequency = " √Ä 00:00 de chaque 1er jour du mois"
    state = "En attente"
    message = ""
    
    # Obtenir le mois et l'ann√©e actuels
    now = timezone.now()
    
    
    mois = now.month
    annee = now.year

    # V√©rifier si la p√©riode existe d√©j√†
    if not PeriodeComptable.objects.filter(mois=mois, annee=annee).exists():
        # Calculer les dates de d√©but et de fin
        date_debut = datetime.datetime(annee, mois, 1)  # Premier jour du mois
        if mois == 12:
            date_fin = datetime.datetime(annee + 1, 1, 1) - datetime.timedelta(days=1)  # Dernier jour de d√©cembre
        else:
            date_fin = datetime.datetime(annee, mois + 1, 1) - datetime.timedelta(days=1)  # Dernier jour du mois courant
        
        libelle = f"{now.strftime('%B').upper()} {annee}"

        # Cr√©er une nouvelle p√©riode comptable
        PeriodeComptable.objects.create(
            libelle=libelle,
            mois=mois,
            annee=annee,
            date_debut=date_debut,
            date_fin=date_fin
        )
        
        state = "<span style='color: green;'>Succ√®ss</span>"
        message = f'''<br>La p√©riode comptable {libelle} a √©t√© cr√©e.
        <br>
        Date d√©but not√© : {date_debut}
        <br>
        Date de fin not√© : {date_fin}
        '''

    else:
        state = "<span style='color: blue;'>Non ex√©cut√©e</span>"
        message = "La p√©riode comptable pour ce mois existe d√©j√†"
        print("p√©riode comptable exists d√©j√†")
    
    send_cron_state_mail(email=email, cron_name=cron_name, frequency=frequency, state=state, message=message)
    
    response = {
        'statut': 1,
        'message': 'TACHE CRON EXECUT√âE',
    }
    return JsonResponse(response, json_dumps_params={'ensure_ascii': False})
    
    
