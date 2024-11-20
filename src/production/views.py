# Create your views here.
import datetime
from datetime import timedelta
import json
import os
import uuid
from ast import literal_eval
from io import BytesIO
from pprint import pprint
import math
from sqlite3 import Date
from datetime import date
from django.db.models import Max

import openpyxl
import pandas as pd
# import qrcode
# import qrcode.image.svg
# from dateutil.relativedelta import relativedelta
from django.contrib import admin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.models import Group
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Q, OuterRef, Subquery, ExpressionWrapper, F, DurationField, IntegerField, Max
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.template.loader import get_template
from django.urls import reverse
from django.utils import timezone
from django.utils.datastructures import MultiValueDictKeyError
from django.utils.decorators import method_decorator
from django.utils.translation import gettext as _
from django.views import View
from django.views.decorators.cache import never_cache
from django.views.generic import TemplateView, ListView
from django_dump_die.middleware import dd
# from django_dump_die.middleware import dd
from fpdf import FPDF
from xhtml2pdf import pisa

from configurations.models import Compagnie, MarqueVehicule, Pays, Civilite, QualiteBeneficiaire, Profession, \
    Produit, \
    Territorialite, ModeCalcul, Duree, TicketModerateur, TypeCarosserie, User, Fractionnement, ModeReglement, \
    Regularisation, Bureau, BusinessUnit, \
    Devise, Taxe, BureauTaxe, Apporteur, BaseCalcul, TypeQuittance, NatureQuittance, TypeClient, TypePersonne, Langue, \
    Branche, ParamProduitCompagnie, CategorieVehicule, Banque, \
    NatureOperation, TypeTarif, Prestataire, Acte, Rubrique, ReseauSoin, Periodicite, PrescripteurPrestataire, \
    AuthGroup, ActionLog, SousRubrique, RegroupementActe, TypePrefinancement, CompteTresorerie, SousRegroupementActe, \
    GroupeInter
from grh.models import Prospect, Campagne, CampagneProspect
from inov import settings
from production.forms import ContactForm, FilialeForm, AcompteForm, DocumentForm, PoliceForm, PhotoUploadForm
from production.helper_production import create_alimet_helper
from production.models import FormuleRubriquePrefinance, ModePrefinancement, Motif, Mouvement, Aliment, Client, Police, \
    Acompte, Document, Filiale, \
    Contact, Quittance, SecteurActivite, TypeDocument, AlimentFormule, Statut, FormuleGarantie, MouvementPolice, StatutQuittance, \
    Genre, StatutFamilial, PlacementEtGestion, ModeRenouvellement, CalculTM, ApporteurPolice, TaxePolice, \
    TaxeQuittance, Reglement, OptionYesNo, Carte, TypeMajorationContrat, Vehicule, VehiculePolice, Energie, \
    StatutPolice, Operation, TarifPrestataireClient, PeriodeCouverture, Bareme, AlimentTemporaire, MouvementAliment, \
    OperationReglement, HistoriquePolice, HistoriqueApporteurPolice, HistoriqueTaxePolice
from production.templatetags.my_filters import money_field
from shared.enum import StatutIncorporation, StatutValidite, StatutSinistre, StatutEnrolement, StatutTraitement, \
    StatutReversementCompagnie, StatutValiditeQuittance
from shared.helpers import generer_qrcode_carte, generate_numero_famille, generate_numero_carte, render_pdf, \
    generer_numero_ordre, generer_nombre_famille_du_mois, custom_model_to_dict
from shared.veos import get_taux_euro_by_devise, get_taux_usd_by_devise, send_client_to_veos
from sinistre.models import Sinistre, DossierSinistre
import traceback
from django.core.files.base import File


## INOV API MOBILE
# from django.views.decorators.csrf import csrf_exempt
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status, generics
# from .models import CarteDigitalDematerialisee
# from rest_framework.permissions import AllowAny, IsAuthenticated

def is_in_group(user, groupe_name):
    return user.groups.filter(name=groupe_name).exists()


def is_in_commercial(user):
    return user.groups.filter(name='COMMERCIAL').exists()


def is_in_production(user):
    return user.groups.filter(name='PRODUCTION').exists()


def todo_manuel(request):
    # IMPORT ADHERENT PRINCIPAUX
    # IMPORT ADHERENT PRINCIPAUX
    adherents_principaux = Aliment.objects.filter(adherent_principal__isnull=True)
    for a in adherents_principaux:
        adp = Aliment.objects.filter(
            veos_code_aliment=a.veos_adherent_principal).first()  # Utilisez .first() pour obtenir le premier résultat
        if adp:
            aliment = Aliment.objects.get(id=a.id)
            aliment.adherent_principal = adp
            aliment.observation = "UPDATED ON 09102023"
            aliment.save()  # Enregistrez l'objet modifié dans la base de données

    response = {
        'statut': 0,
        'message': "UPDATE EFFECTUE",
    }

    return JsonResponse(response)


@method_decorator(login_required, name='dispatch')
class DetailsClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/details_client.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()

            pprint(client.pays.devise)
            polices = Police.objects.filter(client_id=client_id, statut=StatutPolice.ACTIF, statut_contrat='CONTRAT', statut_validite=StatutValidite.VALIDE).order_by('-id')

            derniere_police = polices.first()
            nouvelles_polices = [derniere_police] if derniere_police and not derniere_police.has_beneficiaires else []
            #nouvelles_polices = [derniere_police] if derniere_police else []

            #les anciennes polices qui un mouvement_police de résiliation
            #anciennes_polices = polices.filter(id__in=select police_id from mouvemet_police where date_debut >=today).exclude(id=derniere_police.id)
            anciennes_polices = polices.filter(
                id__in=MouvementPolice.objects.filter(
                    mouvement__code="RESIL",
                    statut_validite=StatutValidite.VALIDE,
                    #date_effet__gte=datetime.datetime.now(tz=timezone.utc).date(),
                    police_id__in=polices.values_list('id', flat=True)
                ).values_list('police_id', flat=True)
            )

            statut_contrat = "CONTRAT"

            quittances = []
            for police in polices:
                quittances_of_police = Quittance.objects.filter(police_id=police.id)
                quittances.extend(quittances_of_police)

            acomptes = Acompte.objects.filter(client_id=client_id)

            filiales = Filiale.objects.filter(client_id=client_id)

            documents = Document.objects.filter(client_id=client_id)

            contacts = Contact.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')

            types_documents = TypeDocument.objects.all().order_by('libelle')

            types_prefinancements = TypePrefinancement.objects.filter(statut=Statut.ACTIF).order_by('libelle')

            # pour la creation de police
            branches = Branche.objects.filter(status=True).order_by('nom')
            produits = Produit.objects.all().order_by('nom')
            bureaux = Bureau.objects.all().order_by('nom')
            utilisateurs = None  # User.objects.all().order_by('last_name')
            apporteurs = Apporteur.objects.filter(status=True).order_by('nom')
            tickets_moderateurs = TicketModerateur.objects.all().order_by('libelle')
            fractionnements = Fractionnement.objects.all().order_by('libelle')
            modes_reglements = ModeReglement.objects.all().order_by('libelle')
            regularisations = Regularisation.objects.all().order_by('libelle')
            compagnies = Compagnie.objects.filter(bureau=request.user.bureau, status=True).order_by('nom')
            durees = Duree.objects.all().order_by('libelle')
            devises = Devise.objects.filter(id=client.pays.devise_id).order_by('libelle')
            taxes = Taxe.objects.all().order_by('libelle')
            bureau_taxes = BureauTaxe.objects.filter(bureau_id=client.bureau_id)
            bases_calculs = BaseCalcul.objects.all().order_by('libelle')
            modes_calculs = ModeCalcul.objects.all().order_by('libelle')

            placement_gestion = PlacementEtGestion
            mode_renouvellement = ModeRenouvellement
            calcul_tm = CalculTM
            type_majoration_contrat = TypeMajorationContrat
            # statut_contrat = StatutContrat

            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {'client': client, 'contacts': contacts, 'polices': polices, 'quittances': quittances,
                             'acomptes': acomptes,
                             'filiales': filiales, 'documents': documents, 'types_documents': types_documents,
                             'branches': branches, 'produits': produits, 'pays': pays,
                             'compagnies': compagnies, 'durees': durees, 'placement_gestion': placement_gestion,
                             'mode_renouvellement': mode_renouvellement, 'tickets_moderateurs': tickets_moderateurs,
                             'calcul_tm': calcul_tm,
                             'fractionnements': fractionnements, 'modes_reglements': modes_reglements,
                             'regularisations': regularisations,
                             'devises': devises, 'utilisateurs': utilisateurs, 'bureaux': bureaux, 'taxes': taxes,
                             'bureau_taxes': bureau_taxes,
                             'apporteurs': apporteurs, 'bases_calculs': bases_calculs,
                             'type_majoration_contrat': type_majoration_contrat, 'modes_calculs': modes_calculs,
                             'statut_contrat': statut_contrat,
                             'types_prefinancements': types_prefinancements,
                             'anciennes_polices': anciennes_polices,
                             'nouvelles_polices': nouvelles_polices
                             }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@login_required
def add_contact(request, client_id):
    if request.method == "POST":

        form = ContactForm(request.POST)

        if form.is_valid():

            client_id = request.POST.get('client_id')

            contact = form.save(commit=False)
            contact.client = Client.objects.get(id=client_id)
            contact.save()

            response = {
                'statut': 1,
                'message': _("Enregistrement effectué avec succès !"),
                'data': {
                    'id': contact.pk,
                    'nom': contact.nom,
                    'prenoms': contact.prenoms,
                    'fonction': contact.fonction,
                    'telephone': contact.telephone,
                    'email': contact.email,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)


# modification d'un contact
@login_required
def modifier_contact(request, contact_id):
    from datetime import datetime
    contact = Contact.objects.get(id=contact_id)

    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': contact.pk,
                    'nom': contact.nom,
                    'prenoms': contact.prenoms,
                    'fonction': contact.fonction,
                    'telephone': contact.telephone,
                    'email': contact.email,
                }
            }

            return JsonResponse(response)
        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)

    else:
        contact = Contact.objects.get(id=contact_id)
        pays = Pays.objects.all().order_by('nom')

        form = ContactForm()

        context = {
            'contact': contact,
            'pays': pays,
            'form': form,
        }

        return render(request, 'client/modification_contact.html', context)

@login_required
def supprimer_contact(request):
    if request.method == "POST":

        contact_id = request.POST.get('contact_id')

        contact = Contact.objects.get(id=contact_id)
        if contact.pk is not None:
            contact.delete()

            response = {
                'statut': 1,
                'message': "Contact supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Contact non trouvé !",
            }

        return JsonResponse(response)


@login_required
def add_filiale(request, client_id):
    if request.method == "POST":

        form = FilialeForm(request.POST)

        if form.is_valid():

            client_id = request.POST.get('client_id')

            filiale = form.save(commit=False)
            filiale.client = Client.objects.get(id=client_id)
            filiale.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': filiale.pk,
                    'nom': filiale.nom,
                    'adresse': filiale.adresse,
                    'pays': filiale.pays.nom,
                    'ville': filiale.ville,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)


def modifier_filiale(request, filiale_id):
    filiale = Filiale.objects.get(id=filiale_id)

    if request.method == 'POST':
        form = FilialeForm(request.POST, instance=filiale)
        if form.is_valid():
            form.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': filiale.pk,
                    'nom': filiale.nom,
                    'adresse': filiale.adresse,
                    'pays': filiale.pays.nom,
                    'ville': filiale.ville,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)

    else:

        filiale = Filiale.objects.get(id=filiale_id)
        pays = Pays.objects.all().order_by('nom')

        form = FilialeForm()

        return render(request, 'client/modification_filiale.html',
                      {'filiale': filiale, 'pays': pays, 'form': form})


def supprimer_filiale(request):
    if request.method == "POST":

        filiale_id = request.POST.get('filiale_id')

        filiale = Filiale.objects.get(id=filiale_id)
        if filiale.pk is not None:
            filiale.delete()

            response = {
                'statut': 1,
                'message': "Filiale supprimée avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Filiale non trouvée !",
            }

        return JsonResponse(response)


def add_document(request, client_id):
    if request.method == "POST":

        form = DocumentForm(request.POST, request.FILES)

        if form.is_valid():

            client = Client.objects.get(id=client_id)
            type_document_id = request.POST.get('type_document')

            #   type_document = get_object_or_404(TypeDocument, id=request.POST.get('type_document'))

            # Use the ORM to create and update the Document instance
            document = form.save(commit=False)
            document.client = client
            #   document.type_document = type_document
            document.type_document = TypeDocument.objects.get(id=type_document_id)
            document.save()

            pprint("document.fichier")
            pprint(document.fichier.path)

            # response = {
            #     'id': document.pk,
            #     'nom': document.nom,
            #     'fichier': document.fichier.url,
            #     'type_document': document.type_document.libelle,
            #     'confidentialite': document.confidentialite,
            # }

            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': document.pk,
                    'nom': document.nom,
                    'fichier': '<a href="' + document.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                    'type_document': document.type_document.libelle,
                    'confidentialite': document.confidentialite,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire !",
                'errors': form.errors,
            }

            return JsonResponse(response)


def handle_uploaded_document(f, filename):
    path_ot_db = '/clients/documents/'
    dirname = settings.MEDIA_URL.replace('/', '') + path_ot_db
    path = os.path.join(dirname)

    if not os.path.exists(path):
        os.makedirs(path)

    with open(dirname + '/' + filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

    return path_ot_db + '/' + filename


def modifier_document(request, document_id):
    document = Document.objects.get(id=document_id)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=document)

        if form.is_valid():

            document_update = form.save(commit=False)
            document_update.save()
            #fichier = request.FILES['fichier']
            #filename_old = document.fichier.url
            #filename = "doc_" + str(uuid.uuid4()) + "_" + fichier.name

            #pprint(filename_old)

            # uploader le fichier
            #uploaded_file_name = ''
            #if fichier is not None:
            #    uploaded_file_name = handle_uploaded_document(fichier, filename)

            #type_document_id = request.POST.get('type_document')
            #document.nom = request.POST.get('nom')
            #document.type_document_id = type_document_id
            #document.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': document_update.pk,
                    'nom': document_update.nom,
                    'fichier': '<a href="' + document_update.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                    'type_document': document_update.type_document.libelle,
                    'confidentialite': document_update.confidentialite,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire !",
                'errors': form.errors,
            }

            return JsonResponse(response)


    else:

        document = Document.objects.get(id=document_id)
        types_documents = TypeDocument.objects.all().order_by('libelle')

        form = DocumentForm()

        return render(request, 'client/modification_document.html',
                      {'document': document, 'types_documents': types_documents, 'form': form})


def supprimer_document(request):
    if request.method == "POST":

        document_id = request.POST.get('document_id')

        document = Document.objects.get(id=document_id)
        if document.pk is not None:
            document.delete()

            response = {
                'statut': 1,
                'message': "Document supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Document non trouvé !",
            }

        return JsonResponse(response)


def add_acompte(request, client_id):
    errors = {}
    if request.method == "POST":

        montant = request.POST.get('montant')
        date_versement = request.POST.get('date_versement')

        if not montant:
            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': errors
            }

            return JsonResponse(response)

        if not date_versement:
            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': errors
            }

            return JsonResponse(response)

        if not errors:

            client_id = request.POST.get('client_id')

            acompte = Acompte(
                montant=request.POST.get('montant', '').replace(' ', ''),
                date_versement=request.POST.get('date_versement', ''),
            )
            acompte.client = Client.objects.get(id=client_id)
            acompte.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': acompte.pk,
                    'montant': acompte.montant,
                    'date_versement': acompte.date_versement,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors
            }

            return JsonResponse(response)


def modifier_acompte(request, acompte_id):
    acompte = Acompte.objects.get(id=acompte_id)

    if request.method == 'POST':
        form = AcompteForm(request.POST, instance=acompte)
        if form.is_valid():
            form.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': acompte.pk,
                    'nom': acompte.montant,
                    'nom': acompte.date_versement,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)

    else:

        acompte = Acompte.objects.get(id=acompte_id)

        form = AcompteForm()

        return render(request, 'client/modification_acompte.html',
                      {'acompte': acompte, 'form': form})


def supprimer_acompte(request):
    if request.method == "POST":

        acompte_id = request.POST.get('acompte_id')

        acompte = Acompte.objects.get(id=acompte_id)
        if acompte.pk is not None:
            acompte.delete()

            response = {
                'statut': 1,
                'message': "Acompte supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Acompte non trouvé !",
            }

        return JsonResponse(response)


# ajout de police
@transaction.atomic  # open a transaction
def add_police(request, client_id):
    taxes = request.COOKIES.get('taxes')
    client = Client.objects.get(id=client_id)

    if request.method == 'POST':

        # print(request.POST)

        form = PoliceForm(request.POST)

        if form.is_valid():

            #produit = Produit.objects.get(id=request.POST.get('produit'))
            produit = Produit.objects.get(code="100991")#SANTE
            branche = produit.branche
            compagnie = Compagnie.objects.get(id=request.POST.get('compagnie'))
            numero = request.POST.get('numero')
            apporteur = request.POST.get('apporteur')
            programme_international = request.POST.get('programme_international')
            placement_gestion = request.POST.get('placement_gestion')
            date_debut_effet = request.POST.get('date_debut_effet')
            date_fin_effet = request.POST.get('date_fin_effet')
            preavis_de_resiliation = request.POST.get('preavis_de_resiliation')
            mode_renouvellement = request.POST.get('mode_renouvellement')
            fractionnement_id = request.POST.get('fractionnement')
            mode_reglement_id = request.POST.get('mode_reglement')
            regularisation_id = request.POST.get('regularisation')
            date_prochaine_facture = request.POST.get('date_prochaine_facture')
            participation = request.POST.get('participation')
            taux_participation = request.POST.get('taux_participation').replace(' ', '')
            if taux_participation == "": taux_participation = 0
            prime_ht = request.POST.get('prime_ht').replace(' ', '')
            if prime_ht == "": prime_ht = 0
            prime_ttc = request.POST.get('prime_ttc').replace(' ', '')
            if prime_ttc == "": prime_ttc = 0
            taxe = request.POST.get('taxe').replace(' ', '')
            if taxe == "": taxe = 0
            autres_taxes = request.POST.get('autres_taxes').replace(' ', '')
            if autres_taxes == "": autres_taxes = 0
            taux_com_courtage = request.POST.get('taux_com_courtage').replace(' ', '')
            if taux_com_courtage == "": taux_com_courtage = 0
            taux_com_courtage_terme = request.POST.get('taux_com_courtage_terme').replace(' ', '')
            if taux_com_courtage_terme == "": taux_com_courtage_terme = 0
            taux_com_gestion = request.POST.get('taux_com_gestion').replace(' ', '')
            if taux_com_gestion == "": taux_com_gestion = 0
            commission_gestion = request.POST.get('commission_gestion').replace(' ', '')
            if commission_gestion == "": commission_gestion = 0
            commission_courtage = request.POST.get('commission_courtage').replace(' ', '')
            if commission_courtage == "": commission_courtage = 0
            commission_intermediaires = request.POST.get('commission_intermediaire').replace(' ', '')
            if commission_intermediaires == "": commission_intermediaires = 0
            cout_police_compagnie = request.POST.get('cout_police_compagnie').replace(' ', '')
            if cout_police_compagnie == "": cout_police_compagnie = 0
            cout_police_courtier = request.POST.get('cout_police_courtier').replace(' ', '')
            if cout_police_courtier == "": cout_police_courtier = 0
            ticket_moderateur_id = request.POST.get('ticket_moderateur')
            calcul_tm = request.POST.get('calcul_tm')
            type_prefinancement_id = request.POST.get('type_prefinancement')

            mode_calcul_id = request.POST.get('mode_calcul')
            prime_famille = request.POST.get('prime_famille').replace(' ', '')
            if prime_famille == "": prime_famille = 0
            nombre_max_enfants_famille = request.POST.get('nombre_max_enfants_famille').replace(' ', '')
            if nombre_max_enfants_famille == "": nombre_max_enfants_famille = 0
            nombre_max_personne_famille = request.POST.get('nombre_max_personne_famille').replace(' ', '')
            if nombre_max_personne_famille == "": nombre_max_personne_famille = 0
            age_max_enfants = request.POST.get('age_max_enfants').replace(' ', '')
            if age_max_enfants == "": age_max_enfants = 0
            age_max_adultes = request.POST.get('age_max_adultes').replace(' ', '')
            if age_max_adultes == "": age_max_adultes = 0
            surprime_personne_sup = request.POST.get('surprime_personne_sup').replace(' ', '')
            if surprime_personne_sup == "": surprime_personne_sup = 0
            surprime_enfant_sup = request.POST.get('surprime_enfant_sup').replace(' ', '')
            if surprime_enfant_sup == "": surprime_enfant_sup = 0
            surprime_age_adulte = request.POST.get('surprime_age_adulte').replace(' ', '')
            if surprime_age_adulte == "": surprime_age_adulte = 0
            surprime_ascendant = request.POST.get('surprime_ascendant').replace(' ', '')
            if surprime_ascendant == "": surprime_ascendant = 0
            prime_personne = request.POST.get('prime_personne').replace(' ', '')
            if prime_personne == "": prime_personne = 0
            prime_adulte = request.POST.get('prime_adulte').replace(' ', '')
            if prime_adulte == "": prime_adulte = 0
            prime_enfant = request.POST.get('prime_enfant').replace(' ', '')
            if prime_enfant == "": prime_enfant = 0
            taux_cotisation = request.POST.get('taux_cotisation').replace(' ', '')
            if taux_cotisation == "": taux_cotisation = 0
            part_employeur = request.POST.get('part_employeur').replace(' ', '')
            if part_employeur == "": part_employeur = 0
            cotisation_minimale = request.POST.get('cotisation_minimale').replace(' ', '')
            if cotisation_minimale == "": cotisation_minimale = 0
            cotisation_maximale = request.POST.get('cotisation_maximale').replace(' ', '')
            if cotisation_maximale == "": cotisation_maximale = 0
            type_majoration = request.POST.get('type_majoration')

            autofinancement = request.POST.get('autofinancement')
            devise_id = request.POST.get('devise')
            taux_charge = request.POST.get('taux_charge').replace(' ', '')
            if taux_charge == "": taux_charge = 0
            coefficient_n = request.POST.get('coefficient_n').replace(' ', '')
            if coefficient_n == "": coefficient_n = 0
            coefficient_n1 = request.POST.get('coefficient_n1').replace(' ', '')
            if coefficient_n1 == "": coefficient_n1 = 0
            coefficient_n2 = request.POST.get('coefficient_n2').replace(' ', '')
            if coefficient_n2 == "": coefficient_n2 = 0
            coefficient_n3 = request.POST.get('coefficient_n3').replace(' ', '')
            if coefficient_n3 == "": coefficient_n3 = 0

            statut_contrat = request.POST.get('statut_contrat')
            statut_contrat = "CONTRAT"

            police_created = Police.objects.create(bureau_id=client.bureau_id,
                                           client_id=client_id,
                                           compagnie_id=compagnie.id,
                                           produit_id=produit.id,
                                           numero=numero,
                                           apporteur=apporteur,
                                           #programme_international=programme_international,
                                           #placement_gestion=placement_gestion,
                                           date_souscription=date_debut_effet,
                                           date_debut_effet=date_debut_effet,
                                           date_fin_effet=date_fin_effet,
                                           preavis_de_resiliation=preavis_de_resiliation,
                                           mode_renouvellement=mode_renouvellement,
                                           fractionnement_id=fractionnement_id,
                                           mode_reglement_id=mode_reglement_id,
                                           regularisation_id=regularisation_id,
                                           date_prochaine_facture=date_prochaine_facture,
                                           participation=participation,
                                           taux_participation=taux_participation,
                                           prime_ht=prime_ht,
                                           prime_ttc=prime_ttc,
                                           taxe=taxe,
                                           autres_taxes=autres_taxes,
                                           taux_com_courtage=taux_com_courtage,
                                           taux_com_courtage_terme=taux_com_courtage_terme,
                                           #taux_com_gestion=taux_com_gestion,
                                           commission_gestion=commission_gestion,
                                           commission_courtage=commission_courtage,
                                           commission_intermediaires=commission_intermediaires,
                                           cout_police_compagnie=cout_police_compagnie,
                                           cout_police_courtier=cout_police_courtier,
                                           ticket_moderateur_id=ticket_moderateur_id,
                                           type_prefinancement_id=type_prefinancement_id,
                                           calcul_tm=calcul_tm,

                                           #mode_calcul_id=mode_calcul_id,
                                           #prime_famille=prime_famille,
                                           #nombre_max_enfants_famille=nombre_max_enfants_famille,
                                           #nombre_max_personne_famille=nombre_max_personne_famille,
                                           #age_max_enfants=age_max_enfants,
                                           #age_max_adultes=age_max_adultes,
                                           #surprime_personne_sup=surprime_personne_sup,
                                           #surprime_enfant_sup=surprime_enfant_sup,
                                           #surprime_age_adulte=surprime_age_adulte,
                                           #surprime_ascendant=surprime_ascendant,
                                           #prime_personne=prime_personne,
                                           #prime_adulte=prime_adulte,
                                           #prime_enfant=prime_enfant,
                                           #taux_cotisation=taux_cotisation,
                                           #part_employeur=part_employeur,
                                           #cotisation_minimale=cotisation_minimale,
                                           #cotisation_maximale=cotisation_maximale,
                                           #type_majoration=type_majoration,

                                           #autofinancement=autofinancement,
                                           #devise_id=devise_id,
                                           #taux_charge=taux_charge,
                                           #coefficient_n=coefficient_n,
                                           #coefficient_n1=coefficient_n1,
                                           #coefficient_n2=coefficient_n2,
                                           #coefficient_n3=coefficient_n3,

                                           statut_contrat=statut_contrat,
                                           statut=Statut.ACTIF,
                                           created_by=request.user
                                           )

            police_created.save()

            code_bureau = request.user.bureau.code
            police_created.numero_provisoire = str(code_bureau)+'P' + str(Date.today().year)[-2:] + str(police_created.pk).zfill(6)
            if police_created.numero == "": police_created.numero = police_created.numero_provisoire

            police_created.save()

            # Handle logo_partenaire upload
            logo_partenaire_file = request.FILES.get('logo_partenaire')
            #dd(logo_partenaire_file)
            if logo_partenaire_file:
                police_created.logo_partenaire.save(logo_partenaire_file.name, logo_partenaire_file)
                police_created.save()
                pprint("Logo partenaire joint")
            else:
                pprint("Pas de logo partenaire joint")

            police = Police.objects.get(id=police_created.pk)

            # enregistrer les intermédiaires si existants
            intermediaires = request.POST.getlist('intermediaires')
            base_calcul_taux_retrocession = request.POST.getlist('base_calcul_taux_retrocession')
            taux_com_affaire_nouvelle = request.POST.getlist('taux_com_affaire_nouvelle')
            taux_com_renouvelement = request.POST.getlist('taux_com_renouvelement')

            pprint('len(intermediaires) : ' + str(len(intermediaires)))
            if len(intermediaires) > 0:  # pourquoi j'ai mis 3: à vérifier, en attendant je met à 0
                i = 0
                for apporteur_id in intermediaires:
                    apporteur_id = int('0' + apporteur_id)
                    base_calcul = int('0' + base_calcul_taux_retrocession[i])
                    taux_com_an = float('0' + taux_com_affaire_nouvelle[i])
                    taux_com_renew = float('0' + taux_com_renouvelement[i])

                    pprint({'apporteur_id': apporteur_id, 'base_calcul': base_calcul, 'taux_com_an': taux_com_an,
                            'taux_com_renew': taux_com_renew})

                    # Insérer la ligne si renseignée
                    if apporteur_id > 0 and base_calcul > 0 and (taux_com_an > 0 or taux_com_renew > 0):
                        ApporteurPolice.objects.create(police_id=police.id, apporteur_id=apporteur_id,
                                                       base_calcul_id=base_calcul,
                                                       taux_com_affaire_nouvelle=taux_com_an,
                                                       taux_com_renouvellement=taux_com_renew, ).save()
                    i += 1

            # enregistrer les autres taxes
            taxes = request.COOKIES.get('taxes')
            if taxes:
                taxes = json.loads(taxes)

                for taxe in taxes:
                    taxe = list(taxe.values())
                    taxe_id = taxe[0]
                    taxe_montant = taxe[1]

                    # Insérer la ligne
                    #TaxePolice.objects.create(police_id=police.id, taxe_id=taxe_id, montant=taxe_montant).save()

            # créer une ligne de mouvement_police avec le mouvement affaire nouvelle et le motif affaire nouvelle
            mp = MouvementPolice()
            mp.police = police
            mp.mouvement = Mouvement.objects.get(code='AN')
            mp.motif = Motif.objects.get(code='AN')
            mp.date_effet = police.date_debut_effet
            mp.date_fin_periode_garantie = police.date_fin_effet
            mp.save()

            # créer une ligne dans période de couverture
            periode_couverture = PeriodeCouverture.objects.create(
                police_id=police.id,
                date_debut_effet=date_debut_effet,
                date_fin_effet=date_fin_effet,
            ).save()

            response = {
                'statut': 1,
                'message': "Police enregistrée avec succès !",
                'data': {
                    'id': police.pk,
                    'numero': police.numero,
                    'produit': police.produit.nom,
                    'compagnie': police.compagnie.nom,
                    'prime_ht': police.prime_ht,
                    'prime_ttc': police.prime_ttc,
                    'commission_gestion': police.commission_gestion,
                    'commission_courtage': police.commission_courtage,
                    'date_debut_effet': police.date_debut_effet,
                    'date_fin_effet': police.date_fin_effet,
                    'statut': police.statut,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire",
                'errors': form.errors,
            }

            return JsonResponse(response)


# modification de police
@transaction.atomic  # open a transaction
def modifier_police(request, police_id):
    if request.method == 'POST':

        produit = Produit.objects.get(id=request.POST.get('produit'))
        branche = produit.branche
        compagnie = Compagnie.objects.get(id=request.POST.get('compagnie'))
        numero = request.POST.get('numero')
        apporteur = request.POST.get('apporteur')
        programme_international = request.POST.get('programme_international')
        placement_gestion = request.POST.get('placement_gestion')
        date_debut_effet = request.POST.get('date_debut_effet')
        date_fin_effet = request.POST.get('date_fin_effet')
        preavis_de_resiliation = request.POST.get('preavis_de_resiliation')
        mode_renouvellement = request.POST.get('mode_renouvellement')
        fractionnement_id = request.POST.get('fractionnement')
        mode_reglement_id = request.POST.get('mode_reglement')
        regularisation_id = request.POST.get('regularisation')
        date_prochaine_facture = request.POST.get('date_prochaine_facture')
        participation = request.POST.get('participation')
        taux_participation = request.POST.get('taux_participation').replace(' ', '')
        if taux_participation == "": taux_participation = 0
        prime_ht = request.POST.get('prime_ht').replace(' ', '')
        if prime_ht == "": prime_ht = 0
        prime_ttc = request.POST.get('prime_ttc').replace(' ', '')
        if prime_ttc == "": prime_ttc = 0
        taxe = request.POST.get('taxe').replace(' ', '')
        if taxe == "": taxe = 0
        autres_taxes = request.POST.get('autres_taxes').replace(' ', '')
        if autres_taxes == "": autres_taxes = 0
        #taux_com_courtage = compagnie.taux_com_courtage
        #taux_com_gestion = compagnie.taux_com_gestion
        taux_com_courtage = request.POST.get('taux_com_courtage').replace(' ', '')
        if taux_com_courtage == "": taux_com_courtage = 0
        taux_com_courtage_terme = request.POST.get('taux_com_courtage_terme').replace(' ', '')
        if taux_com_courtage_terme == "": taux_com_courtage_terme = 0
        taux_com_gestion = request.POST.get('taux_com_gestion').replace(' ', '')
        if taux_com_gestion == "": taux_com_gestion = 0
        commission_gestion = request.POST.get('commission_gestion').replace(' ', '')
        if commission_gestion == "": commission_gestion = 0
        commission_courtage = request.POST.get('commission_courtage').replace(' ', '')
        if commission_courtage == "": commission_courtage = 0
        commission_intermediaires = request.POST.get('commission_intermediaire').replace(' ', '')
        if commission_intermediaires == "": commission_intermediaires = 0
        cout_police_compagnie = request.POST.get('cout_police_compagnie').replace(' ', '')
        if cout_police_compagnie == "": cout_police_compagnie = 0
        cout_police_courtier = request.POST.get('cout_police_courtier').replace(' ', '')
        if cout_police_courtier == "": cout_police_courtier = 0
        ticket_moderateur_id = request.POST.get('ticket_moderateur')
        calcul_tm = request.POST.get('calcul_tm')

        type_prefinancement_id = request.POST.get('type_prefinancement')

        mode_calcul_id = request.POST.get('mode_calcul')
        prime_famille = request.POST.get('prime_famille').replace(' ', '')
        if prime_famille == "": prime_famille = 0
        nombre_max_enfants_famille = request.POST.get('nombre_max_enfants_famille').replace(' ', '')
        if nombre_max_enfants_famille == "": nombre_max_enfants_famille = 0
        nombre_max_personne_famille = request.POST.get('nombre_max_personne_famille').replace(' ', '')
        if nombre_max_personne_famille == "": nombre_max_personne_famille = 0
        age_max_enfants = request.POST.get('age_max_enfants').replace(' ', '')
        if age_max_enfants == "": age_max_enfants = 0
        age_max_adultes = request.POST.get('age_max_adultes').replace(' ', '')
        if age_max_adultes == "": age_max_adultes = 0
        surprime_personne_sup = request.POST.get('surprime_personne_sup').replace(' ', '')
        if surprime_personne_sup == "": surprime_personne_sup = 0
        surprime_enfant_sup = request.POST.get('surprime_enfant_sup').replace(' ', '')
        if surprime_enfant_sup == "": surprime_enfant_sup = 0
        surprime_age_adulte = request.POST.get('surprime_age_adulte').replace(' ', '')
        if surprime_age_adulte == "": surprime_age_adulte = 0
        surprime_ascendant = request.POST.get('surprime_ascendant').replace(' ', '')
        if surprime_ascendant == "": surprime_ascendant = 0
        prime_personne = request.POST.get('prime_personne').replace(' ', '')
        if prime_personne == "": prime_personne = 0
        prime_adulte = request.POST.get('prime_adulte').replace(' ', '')
        if prime_adulte == "": prime_adulte = 0
        prime_enfant = request.POST.get('prime_enfant').replace(' ', '')
        if prime_enfant == "": prime_enfant = 0
        taux_cotisation = request.POST.get('taux_cotisation').replace(' ', '')
        if taux_cotisation == "": taux_cotisation = 0
        part_employeur = request.POST.get('part_employeur').replace(' ', '')
        if part_employeur == "": part_employeur = 0
        cotisation_minimale = request.POST.get('cotisation_minimale').replace(' ', '')
        if cotisation_minimale == "": cotisation_minimale = 0
        cotisation_maximale = request.POST.get('cotisation_maximale').replace(' ', '')
        if cotisation_maximale == "": cotisation_maximale = 0
        type_majoration = request.POST.get('type_majoration')

        autofinancement = request.POST.get('autofinancement')
        devise_id = request.POST.get('devise')
        taux_charge = request.POST.get('taux_charge').replace(' ', '')
        if taux_charge == "": taux_charge = 0
        coefficient_n = request.POST.get('coefficient_n').replace(' ', '')
        if coefficient_n == "": coefficient_n = 0
        coefficient_n1 = request.POST.get('coefficient_n1').replace(' ', '')
        if coefficient_n1 == "": coefficient_n1 = 0
        coefficient_n2 = request.POST.get('coefficient_n2').replace(' ', '')
        if coefficient_n2 == "": coefficient_n2 = 0
        coefficient_n3 = request.POST.get('coefficient_n3').replace(' ', '')
        if coefficient_n3 == "": coefficient_n3 = 0

        # Récupérer la police à mettre à jour
        police_old = Police.objects.get(id=police_id)

        # Créer l'historique avant la mise à jour
        histtorique_police = HistoriquePolice.objects.create(
            police=police_old,
            numero=numero,
            created_by=police_old.created_by,
            updated_by=request.user,
            produit=police_old.produit,
            type_assurance=police_old.type_assurance,
            bureau=police_old.bureau,
            compagnie=police_old.compagnie,
            client=police_old.client,
            veos_code_client=police_old.veos_code_client,
            veos_code_cie=police_old.veos_code_cie,
            veos_id_npol=police_old.veos_id_npol,
            veos_id_pol=police_old.veos_id_pol,
            devise=police_old.devise,
            apporteur=police_old.apporteur,
            programme_international=police_old.programme_international,
            placement_gestion=police_old.placement_gestion,
            date_souscription=police_old.date_souscription,
            date_debut_effet=police_old.date_debut_effet,
            date_fin_effet=police_old.date_fin_effet,
            date_fin_police=police_old.date_fin_police,
            preavis_de_resiliation=police_old.preavis_de_resiliation,
            mode_renouvellement=police_old.mode_renouvellement,
            fractionnement=police_old.fractionnement,
            mode_reglement=police_old.mode_reglement,
            regularisation=police_old.regularisation,
            date_prochaine_facture=police_old.date_prochaine_facture,
            taux_com_courtage=police_old.taux_com_courtage,
            taux_com_courtage_terme=police_old.taux_com_courtage_terme,
            taux_com_gestion=police_old.taux_com_gestion,
            participation=police_old.participation,
            taux_participation=police_old.taux_participation,
            prime_ht=police_old.prime_ht,
            prime_ttc=police_old.prime_ttc,
            commission_gestion=police_old.commission_gestion,
            commission_courtage=police_old.commission_courtage,
            commission_intermediaires=police_old.commission_intermediaires,
            cout_police_compagnie=police_old.cout_police_compagnie,
            cout_police_courtier=police_old.cout_police_courtier,
            taxe=police_old.taxe,
            autres_taxes=police_old.autres_taxes,
            type_prefinancement=police_old.type_prefinancement,
            ticket_moderateur=police_old.ticket_moderateur,
            mode_calcul=police_old.mode_calcul,
            prime_famille=police_old.prime_famille,
            nombre_max_personne_famille=police_old.nombre_max_personne_famille,
            nombre_max_enfants_famille=police_old.nombre_max_enfants_famille,
            age_max_adultes=police_old.age_max_adultes,
            age_max_enfants=police_old.age_max_enfants,
            surprime_personne_sup=police_old.surprime_personne_sup,
            surprime_enfant_sup=police_old.surprime_enfant_sup,
            surprime_age_adulte=police_old.surprime_age_adulte,
            surprime_ascendant=police_old.surprime_ascendant,
            prime_personne=police_old.prime_personne,
            prime_adulte=police_old.prime_adulte,
            prime_enfant=police_old.prime_enfant,
            taux_cotisation=police_old.taux_cotisation,
            part_employeur=police_old.part_employeur,
            cotisation_minimale=police_old.cotisation_minimale,
            cotisation_maximale=police_old.cotisation_maximale,
            type_majoration=police_old.type_majoration,
            autofinancement=police_old.autofinancement,
            taux_charge=police_old.taux_charge,
            coefficient_n=police_old.coefficient_n,
            coefficient_n1=police_old.coefficient_n1,
            coefficient_n2=police_old.coefficient_n2,
            coefficient_n3=police_old.coefficient_n3,
            statut=police_old.statut,
            created_at=police_old.created_at,
            updated_at=police_old.updated_at,
        )
        # # Historique apporteur police
        apporteurs_old = ApporteurPolice.objects.filter(police_id=police_id)
        for apporteur_old in apporteurs_old:
            HistoriqueApporteurPolice.objects.create(
                taux_com_affaire_nouvelle=apporteur_old.taux_com_affaire_nouvelle,
                taux_com_renouvellement=apporteur_old.taux_com_renouvellement,
                base_calcul=apporteur_old.base_calcul,
                apporteur=apporteur_old.apporteur,
                historique_police=histtorique_police,
                date_effet=apporteur_old.date_effet,
                statut_validite=apporteur_old.statut_validite,
                created_at=apporteur_old.created_at,
                updated_at=apporteur_old.updated_at,
                deleted_at=apporteur_old.deleted_at,
                added_by_id=apporteur_old.added_by_id,
            )
        # Historique taxe police
        taxes_old = TaxePolice.objects.filter(police_id=police_id)
        for taxe in taxes_old:
            HistoriqueTaxePolice.objects.create(
                montant=taxe.montant,
                taxe=taxe.taxe,
                historique_police=histtorique_police,
                created_at=taxe.created_at,
                updated_at=taxe.updated_at,
            )

        # relier l'historique police au mouvement police
        # get 2nd last mouvement police

        mouvement_police = MouvementPolice.objects.filter(police_id=police_id, historique_police_id__isnull=True).order_by('-id').first()
        if mouvement_police:
            mouvement_police.historique_police = histtorique_police
            mouvement_police.save()

        # creation du monvement police
        movement_data_save = request.session.get('add_avenant')
        print(movement_data_save)
        if movement_data_save:
            date_fin_periode_garantie = movement_data_save.get('date_fin_periode_garantie')
            mouvement_police = MouvementPolice.objects.create(police_id=police_id,
                                                              mouvement_id=movement_data_save.get('mouvement'),
                                                              motif_id=movement_data_save.get('motif'),
                                                              date_effet=movement_data_save.get('date_effet'),
                                                              date_fin_periode_garantie=date_fin_periode_garantie if date_fin_periode_garantie else None,
                                                              created_by=request.user
                                                              )
            mouvement_police.save()
            mouvement = Mouvement.objects.get(id=mouvement_police.mouvement_id)

            # si c'est un renouvellement, créer une période de couverture
            if mouvement.code == "AVENANT":
                # créer une ligne dans période de couverture
                periode_couverture = PeriodeCouverture.objects.create(
                    police_id=police_old.id,
                    date_debut_effet=mouvement_police.date_effet,
                    date_fin_effet=mouvement_police.date_fin_periode_garantie,
                ).save()

        # Mise à jour de la police
        police = Police.objects.filter(id=police_id).update(
            produit_id=produit.id,
            compagnie_id=compagnie.id,
            numero=numero,
            apporteur=apporteur,
            programme_international=programme_international,
            placement_gestion=placement_gestion,
            date_souscription=date_debut_effet,
            date_debut_effet=date_debut_effet,
            date_fin_effet=date_fin_effet,
            preavis_de_resiliation=preavis_de_resiliation,
            mode_renouvellement=mode_renouvellement,
            fractionnement_id=fractionnement_id,
            mode_reglement_id=mode_reglement_id,
            regularisation_id=regularisation_id,
            date_prochaine_facture=date_prochaine_facture,
            participation=participation,
            taux_participation=taux_participation,
            prime_ht=prime_ht,
            prime_ttc=prime_ttc,
            taxe=taxe,
            autres_taxes=autres_taxes,
            taux_com_courtage=taux_com_courtage,
            taux_com_courtage_terme=taux_com_courtage_terme,
            taux_com_gestion=taux_com_gestion,
            commission_gestion=commission_gestion,
            commission_courtage=commission_courtage,
            commission_intermediaires=commission_intermediaires,
            cout_police_compagnie=cout_police_compagnie,
            cout_police_courtier=cout_police_courtier,
            ticket_moderateur_id=ticket_moderateur_id,
            calcul_tm=calcul_tm,
            type_prefinancement_id=type_prefinancement_id,

            mode_calcul_id=mode_calcul_id, #champ non présent dans le formulaire de modification
            prime_famille=prime_famille,
            nombre_max_enfants_famille=nombre_max_enfants_famille,
            nombre_max_personne_famille=nombre_max_personne_famille,
            age_max_enfants=age_max_enfants,
            age_max_adultes=age_max_adultes,
            surprime_personne_sup=surprime_personne_sup,
            surprime_enfant_sup=surprime_enfant_sup,
            surprime_age_adulte=surprime_age_adulte,
            surprime_ascendant=surprime_ascendant,
            prime_personne=prime_personne,
            prime_adulte=prime_adulte,
            prime_enfant=prime_enfant,
            taux_cotisation=taux_cotisation,
            part_employeur=part_employeur,
            cotisation_minimale=cotisation_minimale,
            cotisation_maximale=cotisation_maximale,
            type_majoration=type_majoration,

            autofinancement=autofinancement,
            devise_id=devise_id,
            taux_charge=taux_charge,
            coefficient_n=coefficient_n,
            coefficient_n1=coefficient_n1,
            coefficient_n2=coefficient_n2,
            coefficient_n3=coefficient_n3,
            statut=Statut.ACTIF,
            updated_by=request.user
        )

        police = Police.objects.get(id=police_id)

        # Handle logo_partenaire upload
        logo_partenaire_file = request.FILES.get('logo_partenaire')
        if logo_partenaire_file:
            police.logo_partenaire.save(logo_partenaire_file.name, logo_partenaire_file)
            police.save()
            pprint("Logo partenaire joint")
        else:
            pprint("Pas de logo partenaire joint")


        #TODO METTRE A JOUR LA PERIODE DE COUVERTURE SI LA DATE DEBUT OU FIN A CHANGÉ
        #OU CRÉER UNE NOUVELLE LIGNE - NON, VU QUE LES SINISTRES TIENNES COMPTE DE ÇA



        # enregistrer les intermédiaires si existants
        intermediaires = request.POST.getlist('intermediaires')

        base_calcul_taux_retrocession = request.POST.getlist('base_calcul_taux_retrocession')
        taux_com_affaire_nouvelle = request.POST.getlist('taux_com_affaire_nouvelle')
        taux_com_renouvelement = request.POST.getlist('taux_com_renouvelement')

        #Commented on 25062024: bien revoir la procédure pour ne pas ajouter et supprimer
        if len(intermediaires) > 0:  # pourquoi j'ai mis 3: à vérifier, en attendant je met à 0
            i = 0
            for apporteur_id in intermediaires:

                pprint("apporteur_id")
                pprint(apporteur_id)
                pprint("base_calcul_taux_retrocession")
                pprint(base_calcul_taux_retrocession)
                pprint("taux_com_affaire_nouvelle")
                pprint(taux_com_affaire_nouvelle)
                pprint("taux_com_renouvelement")
                pprint(taux_com_renouvelement)

                # # Vider la table intermédiaire pour ajouter les nouveaux
                # ApporteurPolice.objects.filter(police_id=police_id).update(statut_validite=StatutValidite.SUPPRIME, updated_at=datetime.datetime.now(tz=timezone.utc))
                # dd("ok")


                apporteur_id = int('0' + apporteur_id)

                # base_calcul = int('0' + base_calcul_taux_retrocession[i])
                # taux_com_an = float('0' + taux_com_affaire_nouvelle[i])
                # taux_com_renew = float('0' + taux_com_renouvelement[i])

                taux_retro_str = base_calcul_taux_retrocession[i].strip()  # Remove any leading or trailing whitespace

                if taux_retro_str:  # Check if the string is not empty
                    base_calcul = float(taux_retro_str.replace(' ', '').replace(',', '.'))
                else:
                    base_calcul = 0.0  # or handle the empty case as needed

                taux_com_str = taux_com_affaire_nouvelle[i].strip()  # Remove any leading or trailing whitespace

                if taux_com_str:  # Check if the string is not empty
                    taux_com_an = float(taux_com_str.replace(' ', '').replace(',', '.'))
                else:
                    taux_com_an = 0.0  # or handle the empty case as needed

                taux_renouvellement_str = taux_com_renouvelement[i].strip()  # Remove any leading or trailing whitespace

                if taux_renouvellement_str:  # Check if the string is not empty
                    taux_com_renew = float(taux_renouvellement_str.replace(' ', '').replace(',', '.'))
                else:
                    taux_com_renew = 0.0  # or handle the empty case as needed


                # Insérer la ligne si renseignée
                if apporteur_id > 0 and base_calcul > 0 and (taux_com_an > 0 or taux_com_renew > 0):
                    apporteur_existant = ApporteurPolice.objects.filter(police_id=police.id, apporteur_id=apporteur_id,base_calcul_id=base_calcul, taux_com_affaire_nouvelle=taux_com_an, taux_com_renouvellement=taux_com_renew).first()
                    if not apporteur_existant:

                        # Vider la table intermédiaire pour ajouter les nouveaux
                        ApporteurPolice.objects.filter(police_id=police_id).update(statut_validite=StatutValidite.SUPPRIME, updated_at=datetime.datetime.now(tz=timezone.utc))

                        ApporteurPolice.objects.create(police_id=police.id, apporteur_id=apporteur_id,
                                                    base_calcul_id=base_calcul, taux_com_affaire_nouvelle=taux_com_an,
                                                    taux_com_renouvellement=taux_com_renew, ).save()
                i += 1


        # enregistrer les autres taxes
        taxes = request.POST.get('liste_autres_taxes_modification')

        if len(taxes) > 0:
            taxes = json.loads(taxes)

            for taxe in taxes:
                taxe = list(taxe.values())
                taxe_id = taxe[0]
                taxe_montant = taxe[1]

                # Insérer la ligne
                TaxePolice.objects.create(police_id=police.id, taxe_id=taxe_id, montant=taxe_montant).save()

        response = {
            'statut': 1,
            'message': "Police modifiée avec succès !",
            'data': {
                'id': police.pk,
                'numero': police.numero,
                'compagnie': police.compagnie.nom,
                'prime_ht': police.prime_ht,
                'prime_ttc': police.prime_ttc,
                'commission_gestion': police.commission_gestion,
                'commission_courtage': police.commission_courtage,
                'date_debut_effet': police.date_debut_effet,
                'date_fin_effet': police.date_fin_effet,
                'statut': police.statut,
            }
        }

        return JsonResponse(response)

    else:

        police = Police.objects.get(id=police_id)

        periode_couverture = PeriodeCouverture.objects.filter(police_id=police_id).order_by('-id').first()

        pays = Pays.objects.all().order_by('nom')

        types_documents = TypeDocument.objects.all().order_by('libelle')
        apporteurs_police = ApporteurPolice.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).all()
        # taxes_police = TaxePolice.objects.filter(police_id=police_id).all()

        # pour la creation de police
        produits = Produit.objects.all().order_by('nom')
        utilisateurs = None #User.
        # .all().order_by('last_name')
        apporteurs = Apporteur.objects.filter(Q(bureau=request.user.bureau) | Q(bureau__isnull=True)).filter(status=True).order_by('nom')
        tickets_moderateurs = TicketModerateur.objects.all().order_by('libelle')
        fractionnements = Fractionnement.objects.all().order_by('libelle')
        modes_reglements = ModeReglement.objects.all().order_by('libelle')
        regularisations = Regularisation.objects.all().order_by('libelle')
        compagnies = Compagnie.objects.filter(bureau=request.user.bureau, status=True).order_by('nom')
        durees = Duree.objects.all().order_by('libelle')
        devises = Devise.objects.all().order_by('libelle')
        taxes = Taxe.objects.all().order_by('libelle')
        bureau_taxes = BureauTaxe.objects.filter(bureau_id=police.bureau_id)
        bases_calculs = BaseCalcul.objects.all().order_by('libelle')
        modes_calculs = ModeCalcul.objects.all().order_by('libelle')
        branches = Branche.objects.filter(status=True).order_by('nom')
        types_prefinancements = TypePrefinancement.objects.filter(statut=Statut.ACTIF).order_by('libelle')

        placement_gestion = PlacementEtGestion
        mode_renouvellement = ModeRenouvellement
        calcul_tm = CalculTM
        type_majoration_contrat = TypeMajorationContrat
        optionYesNo = OptionYesNo

        user = User.objects.get(id=request.user.id)

        # Injecter les valeurs pour autres taxes dejà renseignées
        for bt in bureau_taxes:
            taxe_police = TaxePolice.objects.filter(police_id=police_id).filter(taxe_id=bt.taxe_id).first()

            if taxe_police is not None:
                bt.montant_existant = taxe_police.montant
            else:
                bt.montant_existant = 0

        return render(request, 'police/modal_police_modification.html',
                      {'police': police, 'periode_couverture':periode_couverture, 'types_documents': types_documents, 'branches': branches, 'produits': produits,
                       'pays': pays,
                       'compagnies': compagnies, 'durees': durees, 'placement_gestion': placement_gestion,
                       'mode_renouvellement': mode_renouvellement, 'tickets_moderateurs': tickets_moderateurs,
                       'calcul_tm': calcul_tm, 'optionYesNo': optionYesNo,
                       'fractionnements': fractionnements, 'modes_reglements': modes_reglements,
                       'regularisations': regularisations,
                       'devises': devises, 'utilisateurs': utilisateurs, 'taxes': taxes,
                       'bureau_taxes': bureau_taxes,
                       'apporteurs': apporteurs, 'bases_calculs': bases_calculs, 'modes_calculs': modes_calculs,
                       'apporteurs_police': apporteurs_police, 'type_majoration_contrat': type_majoration_contrat, 'types_prefinancements': types_prefinancements,
                       })


@login_required
def list_polices(request, client_id):
    # request.session['client_id_for_new_police'] = client_id

    return redirect('/production/clien/?client__id__exact=' + str(client_id))


@login_required
def add_carte(request, aliment_id):
    aliment = Aliment.objects.get(id=aliment_id)

    if request.method == 'POST':

        # générer le pdf

        response = {
            'statut': 1,
            'message': "Carte enregistrée avec succès !",
            'data': {
            }
        }

        return JsonResponse(response)

    else:

        return render(request, 'police/modal_carte.html', {'aliment': aliment, })


@login_required
# edition des cartes des adhérents
def editer_cartes(request, police_id):
    return render(request, 'admin/form_edition_cartes.html')


# IMPRESSION DE LA CARTE
class CARTE_SANTE(FPDF):
    # marron (167, 79, 51)
    # jaune (234, 173, 82)
    start_x = 0
    start_y = 0

    def set_origin(self, x, y):
        self.start_x = x
        self.start_y = y

    def set_cadre(self):
        self.set_draw_color(234, 173, 82)
        self.set_fill_color(167, 79, 51)
        self.set_text_color(0, 0, 0)
        self.set_line_width(.1)

        self.rect(self.start_x + 4, self.start_y + 4, 102, 55)  # cadre de la carte
        self.rect(self.start_x + 5, self.start_y + 4, 4, 55, 'F')  # bande verticale

        self.rect(self.start_x + 5, self.start_y + 4, 4, 55, 'F')

    def set_logo(self, logo_src):
        # LE LOGO
        self.image(logo_src, self.start_x + 10, self.start_y + 5, 0, 10)

    def set_photo(self, photo_src):
        self.set_draw_color(128, 128, 128)  # gray
        self.rect(self.start_x + 90, self.start_y + 5, 15, 17)
        self.image(photo_src, self.start_x + 90, self.start_y + 5, 15, 17)

    def set_numero_carte(self, numero):
        # Arial bold 15
        self.set_font('Arial', '', 12)
        self.set_draw_color(0, 80, 180)
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)
        self.set_line_width(.1)

        self.set_font('Arial', '', 10)
        self.set_xy(self.start_x + 40, self.start_y + 5)
        self.cell(20, 5, "N° CARTE:", 0, 0, 'L', 1)

        self.set_x(self.start_x + 60)
        self.set_font('Arial', 'B', 12)
        # Title
        self.cell(30, 5, numero, 0, 0, 'L', 1)
        # Line break
        self.ln(10)

        # si inactive, imprimé le statut dessus

    def set_statut(self, statut):
        if statut == Statut.INACTIF:
            self.set_font('Arial', 'B', 10)
            self.set_xy(self.start_x + 50, self.start_y + 10)
            self.set_text_color(255, 0, 0)
            self.cell(55, 5, "I N A C T I V E", 0, 0, 'L', 0)
            self.set_text_color(0, 0, 0)

    def set_nom_beneficiaire(self, nom, date_naissance, numero_cmu):
        if nom is None: nom = ''
        if date_naissance is None: date_naissance = ''
        if numero_cmu is None: numero_cmu = ''

        self.set_y(self.start_y + 22)

        self.set_fill_color(167, 79, 51)
        self.rect(self.start_x + 11, self.start_y + 22, 1, 4, 'F')
        self.set_xy(self.start_x + 12, self.start_y + 22)
        self.set_font('Arial', '', 8)
        self.cell(95, 4, "BÉNÉFICIAIRE", 0, 1, 'L', 0)

        self.ln(1)
        self.set_x(self.start_x + 10)
        self.set_font('Arial', 'B', 10)
        self.cell(95, 4, nom, 0, 1, 'L', 0)

        self.set_font('Arial', '', 8)
        self.ln(1)
        self.set_x(self.start_x + 10)
        self.cell(17, 4, "Né(e) le:", 0, 0, 'L', 0)
        self.set_font('Arial', 'B', 10)
        self.set_x(self.start_x + 27)
        self.cell(30, 4, str(date_naissance), 0, 1, 'L', 0)

        self.set_font('Arial', '', 8)
        self.ln(1)
        self.set_x(self.start_x + 10)
        self.cell(17, 4, "N° CMU:", 0, 0, 'L', 0)
        self.set_font('Arial', 'B', 10)
        self.set_x(self.start_x + 27)
        self.cell(30, 4, str(numero_cmu), 0, 0, 'L', 0)

    def set_nom_adherent_principal(self, adherent_principal, numero_police):
        if adherent_principal:
            nom = adherent_principal.nom + " " + adherent_principal.prenoms
            self.set_y(self.start_y + 43)
            self.set_fill_color(167, 79, 51)
            self.rect(self.start_x + 11, self.start_y + 45, 1, 4, 'F')
            self.set_xy(self.start_x + 12, self.start_y + 45)

            self.set_font('Arial', '', 8)
            self.cell(60, 4, "ADHÉRENT PRINCIPAL", 0, 1, 'L', 0)
            self.ln(1)
            self.set_x(self.start_x + 10)
            self.set_font('Arial', 'B', 10)
            self.cell(60, 4, nom, 0, 1, 'L', 0)

            self.set_font('Arial', '', 8)
            self.ln(1)
            self.set_x(self.start_x + 10)
            self.cell(17, 4, "N° POLICE:", 0, 0, 'L', 0)
            self.set_font('Arial', 'B', 10)
            self.set_x(self.start_x + 27)
            self.cell(43, 4, str(numero_police), 0, 0, 'L', 0)

    def set_qrcode(self, photo_src):
        self.set_draw_color(128, 128, 128)
        self.rect(self.start_x + 62, self.start_y + 40, 15, 15)
        self.image(photo_src, self.start_x + 62, self.start_y + 40, 15, 15)

    def set_signature_filialle(self, photo_src):
        # self.rect(self.start_x + 89, self.start_y + 40, 15, 15)
        self.image(photo_src, self.start_x + 89, self.start_y + 40, 15, 15)

    def set_origin_verso(self, x, y):
        self.start_x = x
        self.start_y = y

    def set_cadre_verso(self):
        self.set_draw_color(234, 173, 82)
        self.set_fill_color(167, 79, 51)
        self.set_text_color(0, 0, 0)
        self.set_line_width(.1)

        self.rect(self.start_x + 4, self.start_y + 4, 102, 55)  # cadre de la carte

    def set_informations_verso(self, informations, nom_bureau, whatsapp, telephone, email, adresse):
        if nom_bureau is None: nom_bureau = ''
        if whatsapp is None: whatsapp = ''
        if telephone is None: telephone = ''
        if email is None: email = ''
        if adresse is None: adresse = ''

        self.set_xy(self.start_x + 8, self.start_y + 8)

        self.set_draw_color(167, 79, 51)
        self.set_fill_color(234, 173, 82)
        self.set_text_color(255, 255, 255)
        self.set_draw_color(255, 255, 255)
        self.set_fill_color(255, 255, 255)
        self.set_text_color(0, 0, 0)

        self.set_font('Arial', '', 9)
        self.multi_cell(94, 4, informations, 1, 'L', 1)
        self.ln(4)
        self.set_x(self.start_x + 8)
        self.set_font('Arial', 'B', 14)
        self.multi_cell(94, 5, nom_bureau, 1, 'C', 1)
        self.ln(5)

        self.set_x(self.start_x + 8)
        self.set_font('Arial', '', 10)
        self.cell(24, 4, "WhatsApp: ", 1, 0, 'L', 1)
        self.set_font('Arial', '', 10)
        self.set_x(self.start_x + 32)
        self.cell(70, 4, str(whatsapp), 1, 1, 'L', 1)

        self.set_x(self.start_x + 8)
        self.set_font('Arial', '', 10)
        self.cell(24, 4, "Fixe: ", 1, 0, 'L', 1)
        self.set_font('Arial', '', 10)
        self.set_x(self.start_x + 32)
        self.cell(70, 4, str(telephone), 1, 1, 'L', 1)

        self.set_x(self.start_x + 8)
        self.set_font('Arial', '', 10)
        self.cell(24, 4, "E-mail: ", 1, 0, 'L', 1)
        self.set_font('Arial', '', 10)
        self.set_x(self.start_x + 32)
        self.cell(70, 4, str(email), 1, 1, 'L', 1)

        self.ln(4)
        self.set_x(self.start_x + 8)
        self.set_font('Arial', '', 9)
        self.multi_cell(94, 5, str(adresse), 1, 'L', 1)


# defining the function to convert an HTML file to a PDF file
def html_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


@login_required
def imprimer_carte(request, police_id, aliment_id):
    # selectionner la carte active de l'aliment pour l'imprimer

    police = Police.objects.get(id=police_id)

    cartes = Carte.objects.filter(aliment_id=aliment_id, statut=Statut.ACTIF).order_by('-id')

    aliment = Aliment.objects.get(id=aliment_id)

    date_du_jour_gmt_ = datetime.datetime.now(tz=datetime.timezone.utc).date()
    formulegarantie = aliment.formule_atdate(date_du_jour_gmt_)
    pprint(formulegarantie)

    pprint(cartes)
    if cartes:
        carte_active = cartes.first()
        aliments = [carte_active.aliment]
        logo = settings.JAZZMIN_SETTINGS.get('logo_for_carte')
        pdf = render_pdf('police/courriers/cartes.html', {"logo": logo, "aliments": aliments,"carte_active":carte_active, "police": police, "formulegarantie": formulegarantie, "request": request})
        return HttpResponse(File(pdf), content_type='application/pdf')

    else:
        return HttpResponse(404)

# TODO FIX GENERATION CARTE
@login_required
def imprimer_cartes(request, police_id):
    # selectionner la carte active des aliments de la police pour les imprimer

    if request.method == "POST":

        parametre_impression = request.POST.get('parametre_impression')
        numero_debut = int("0" + request.POST.get('numero_debut').replace(" ",""))
        numero_fin = int("0" + request.POST.get('numero_fin').replace(" ",""))

        pprint(numero_debut)
        pprint(numero_fin)

        police = Police.objects.get(id=police_id)

        # Recuperation des beneficiaires
        aliments_formules = AlimentFormule.objects.filter(formule_id__in=[p.id for p in police.formules], statut=Statut.ACTIF).order_by('-id')

        cartes = Carte.objects.filter(statut=Statut.ACTIF, aliment_id__in=[af.aliment_id for af in aliments_formules])

        if parametre_impression == 'plage':
            if numero_fin > cartes.count(): numero_fin = cartes.count()

            cartes = cartes[numero_debut-1:numero_fin]
            pprint('Plage à imprimer: ' + str(numero_debut) + ' - ' + str(numero_fin))

        if cartes:

            aliments = []
            logo = settings.JAZZMIN_SETTINGS.get('logo_for_carte')

            for carte_active in cartes:
                aliments.append(carte_active.aliment)

            pdf = render_pdf('police/courriers/cartes.html',
                             {"logo": logo, "aliments": aliments, "police": police,
                              "request": request})
            return HttpResponse(File(pdf), content_type='application/pdf')

        else:

            response = {
                'statut': 0,
                'message': "AUCUNE CARTE ACTIVE TOUVÉE",
            }

        return JsonResponse(response, status=400)


# TODO ORDONNER COMME LA LISTE DES BENEFICIAIRES
@login_required
def imprimer_cartes_new(request, police_id):
    # selectionner la carte active des aliments de la police pour les imprimer

    if request.method == "POST":

        parametre_impression = request.POST.get('parametre_impression')
        numero_debut = int("0" + request.POST.get('numero_debut').replace(" ",""))
        numero_fin = int("0" + request.POST.get('numero_fin').replace(" ",""))

        pprint(numero_debut)
        pprint(numero_fin)

        police = Police.objects.get(id=police_id)

        # Pour éviter les doublons
        # Récupérer les IDs des aliments dans aliment_formule
        aliment_ids = AlimentFormule.objects.filter(
            formule_id__in=[p.id for p in police.formules],
            statut=Statut.ACTIF,
            statut_validite=StatutValidite.VALIDE
        ).values_list('aliment_id', flat=True)

        aliments_all = Aliment.objects.filter(id__in=aliment_ids).order_by('adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

        #cartes = Carte.objects.filter(statut=Statut.ACTIF, aliment_id__in=[af.aliment_id for af in aliments_formules])

        if parametre_impression == 'plage':
            if numero_fin > aliments_all.count(): numero_fin = aliments_all.count()

            aliments_all = aliments_all[numero_debut-1:numero_fin]
            pprint('Plage à imprimer: ' + str(numero_debut) + ' - ' + str(numero_fin))

        if aliments_all:

            aliments = []
            logo = settings.JAZZMIN_SETTINGS.get('logo_for_carte')

            for aliment in aliments_all:
                aliments.append(aliment)

            pdf = render_pdf('police/courriers/cartes.html',
                             {"logo": logo, "aliments": aliments, "police": police,
                              "request": request})
            return HttpResponse(File(pdf), content_type='application/pdf')

        else:

            response = {
                'statut': 0,
                'message': "AUCUNE CARTE ACTIVE TOUVÉE",
            }

        return JsonResponse(response, status=400)


def download(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, "cartes/" + filename)
    # file_path = os.path.join(filename)

    response = FileResponse(open(file_path, 'rb'))
    return response


def dateFromDB(date_naissance):
    formated_date = None

    if date_naissance:
        date = datetime.datetime.strptime(str(date_naissance), "%Y-%m-%d")

        if date.day > 1:
            jour = str(date.day)
        else:
            jour = "0" + str(date.day)

        if date.month > 1:
            mois = str(date.month)
        else:
            mois = "0" + str(date.month)

        annee = str(date.year)

        formated_date = jour + "/" + mois + "/" + annee

    return formated_date


@login_required
# liste des adhérents
def adherents(request, police_id):
    adherents = Aliment.objects.select_related("formule").all()

    return render(request, 'admin/liste_adherents.html', {'adherents': adherents})


@login_required
def ajax_apporteurs(request):
    apporteurs = Apporteur.objects.filter(status=True)
    apporteurs_serialize = serializers.serialize('json', apporteurs)

    return HttpResponse(apporteurs_serialize, content_type='application/json')

@login_required
def ajax_produits(request, branche_id):
    produits = Produit.objects.filter(branche_id=branche_id)
    produits_serialize = serializers.serialize('json', produits)

    return HttpResponse(produits_serialize, content_type='application/json')


@login_required
def sous_rubriques_by_rubrique(request, rubrique_id):
    actes = SousRubrique.objects.filter(rubrique_id=rubrique_id, status=1).order_by('libelle')
    actes_serialize = serializers.serialize('json', actes)

    return HttpResponse(actes_serialize, content_type='application/json')


@login_required
def regroupements_actes_by_rubrique(request, rubrique_id):
    actes = RegroupementActe.objects.filter(rubrique_id=rubrique_id, status=1).order_by('libelle')
    actes_serialize = serializers.serialize('json', actes)

    return HttpResponse(actes_serialize, content_type='application/json')


@login_required
def sous_regroupements_actes_by_rubrique(request, rubrique_id):
    data = SousRegroupementActe.objects.filter(rubrique_id=rubrique_id, status=1).order_by('libelle')
    data_serialize = serializers.serialize('json', data)

    return HttpResponse(data_serialize, content_type='application/json')


@login_required
def actes_by_rubrique(request, rubrique_id):
    rubrique = Rubrique.objects.get(id=rubrique_id)

    if rubrique.code != 'PHARMACIE':
        actes = Acte.objects.filter(rubrique_id=rubrique_id, status=1).order_by('libelle')
    else:
        actes = []

    actes_serialize = serializers.serialize('json', actes)
    return HttpResponse(actes_serialize, content_type='application/json')


@login_required
def actes_by_regroupement_acte(request, regroupement_acte_id):
    regroupement_acte = RegroupementActe.objects.get(id=regroupement_acte_id)
    if regroupement_acte and regroupement_acte.rubrique.code != 'PHARMACIE':
        actes = Acte.objects.filter(regroupement_acte_id=regroupement_acte_id, status=1).order_by('libelle')
    else:
        actes = []

    actes_serialize = serializers.serialize('json', actes)

    return HttpResponse(actes_serialize, content_type='application/json')



@login_required
def formules_by_police(request, police_id):
    formules = FormuleGarantie.objects.filter(police_id=police_id, statut=Statut.ACTIF)
    formules_serialize = serializers.serialize('json', formules)

    return HttpResponse(formules_serialize, content_type='application/json')

@login_required
def polices_restantes(request, police_id):
    police = Police.objects.get(id=police_id)
    polices_restantes = Police.objects.filter(client_id=police.client_id, statut_validite=StatutValidite.VALIDE).exclude(id=police_id)
    polices_restantes = [x for x in polices_restantes if x.etat_police=="En cours" ]

    polices_restantes_serialize = serializers.serialize('json', polices_restantes)

    return HttpResponse(polices_restantes_serialize, content_type='application/json')


@login_required
# récupère le taux paramétré sur le produit en fonction de la compagnie
def ajax_infos_compagnie(request, compagnie_id, produit_id):
    param_produit_compagnie = ParamProduitCompagnie.objects.filter(compagnie_id=compagnie_id,
                                                                   produit_id=produit_id).first()
    print(param_produit_compagnie)
    if (param_produit_compagnie is not None):
        response = {
            'id': param_produit_compagnie.compagnie.id,
            'code': param_produit_compagnie.compagnie.code,
            'nom': param_produit_compagnie.compagnie.nom,
            'taux_com_courtage': param_produit_compagnie.taux_com_courtage,
            'taux_com_courtage_terme': param_produit_compagnie.taux_com_courtage_terme,
            'taux_com_gestion': param_produit_compagnie.taux_com_gestion,
        }

    else:
        response = {
            'taux_com_courtage': '',
            'taux_com_courtage_terme': '',
            'taux_com_gestion': '',
        }

    return JsonResponse(response)


def motifs_by_mouvement(request, mouvement_id):
    motifs = Motif.objects.filter(mouvement_id=mouvement_id)
    motifs_serialize = serializers.serialize('json', motifs)
    return HttpResponse(motifs_serialize, content_type='application/json')


# détails d'une police, du bureau de l'utisateur
@method_decorator(login_required, name='dispatch')
class DetailsPoliceView(TemplateView):
    template_name = 'police/index.html'
    model = Police

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police_id = kwargs['police_id']
        polices = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE)
        if polices:
            police = polices.first()
            #dd(police)
            reseaux_soins = ReseauSoin.objects.filter(bureau=police.bureau)

            # duree_police = police.date_fin_effet - police.date_debut_effet
            # duree_police_en_jours = duree_police.days

            if police.date_debut_effet and police.date_fin_effet:

                # Calculer la différence en mois
                '''duree_police_en_mois = (police.date_fin_effet.year - police.date_debut_effet.year) * 12 + (
                        police.date_fin_effet.month - police.date_debut_effet.month)

                duree = str(duree_police_en_mois) + ' mois'

                if duree_police_en_mois == 0:
                    duree = str((police.date_fin_effet - police.date_debut_effet).days) + ' jours'
                '''

                #nouveau
                duree_police_en_mois = Police.objects.filter(id=police.id).annotate(
                    duree_police_en_mois=ExpressionWrapper(
                        F('date_fin_effet') - F('date_debut_effet'),
                        output_field=DurationField()
                    )
                ).values('id', 'duree_police_en_mois').first()['duree_police_en_mois']


                nombre_total_mois = duree_police_en_mois.days // 30

                duree = f"{nombre_total_mois} {'mois'}"


            # etat police = dernier motif
            etat_police = "" #MouvementPolice.objects.filter(police_id=police_id).order_by('-id')[:1].get().motif.libelle

            #mouvement_police = MouvementPolice.objects.filter(police_id=police_id).order_by('-id')[:1].get()
            mouvement_police = MouvementPolice.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).order_by('-id').first()

            apporteurs_police = ApporteurPolice.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE)

            context_perso = {'police': police, 'etat_police': etat_police, 'duree_police': duree,
                             'mouvement_police': mouvement_police,
                             'apporteurs_police': apporteurs_police, 'reseaux_soins': reseaux_soins, }
            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            #liste_clients_url = reverse('clients')
            return redirect("clients")

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


@method_decorator(login_required, name='dispatch')
class DetailsHistoriquePoliceView(TemplateView):
    template_name = 'police/historique_police_detail.html'
    model = HistoriquePolice

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        hist_police_id = kwargs['police_id']
        hist_polices = HistoriquePolice.objects.filter(id=hist_police_id)
        if hist_polices:
            hist_police = hist_polices.first()
            #dd(police)
            reseaux_soins = ReseauSoin.objects.filter(bureau=hist_police.bureau)

            # duree_police = police.date_fin_effet - police.date_debut_effet
            # duree_police_en_jours = duree_police.days

            if hist_police.date_debut_effet and hist_police.date_fin_effet:

                # Calculer la différence en mois
                '''duree_police_en_mois = (police.date_fin_effet.year - police.date_debut_effet.year) * 12 + (
                        police.date_fin_effet.month - police.date_debut_effet.month)

                duree = str(duree_police_en_mois) + ' mois'

                if duree_police_en_mois == 0:
                    duree = str((police.date_fin_effet - police.date_debut_effet).days) + ' jours'
                '''

                #nouveau
                duree_police_en_mois = HistoriquePolice.objects.filter(id=hist_police.id).annotate(
                    duree_police_en_mois=ExpressionWrapper(
                        F('date_fin_effet') - F('date_debut_effet'),
                        output_field=DurationField()
                    )
                ).values('id', 'duree_police_en_mois').first()['duree_police_en_mois']


                nombre_total_mois = duree_police_en_mois.days // 30

                duree = f"{nombre_total_mois} {'mois'}"


            # etat police = dernier motif
            etat_police = "" #MouvementPolice.objects.filter(police_id=police_id).order_by('-id')[:1].get().motif.libelle

            #mouvement_police = MouvementPolice.objects.filter(police_id=police_id).order_by('-id')[:1].get()
            hist_mouvement_police = MouvementPolice.objects.filter(historique_police_id=hist_police_id, statut_validite=StatutValidite.VALIDE).order_by('-id').first()

            hist_apporteurs_police = HistoriqueApporteurPolice.objects.filter(historique_police_id=hist_police_id)

            context_perso = {'police': hist_police, 'etat_police': etat_police, 'duree_police': duree,
                             'mouvement_police': hist_mouvement_police,
                             'apporteurs_police': hist_apporteurs_police, 'reseaux_soins': reseaux_soins, }
            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            #liste_clients_url = reverse('clients')
            return redirect("clients")

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



# get all police quittances
class PoliceQuittancesView(TemplateView):
    template_name = 'police/quittances.html'
    model = Quittance

    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.get(id=police_id)

        quittances = Quittance.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE, import_stats=False)
        types_quittances = TypeQuittance.objects.all()

        '''
        for avenant in mouvements_police:
            quittances_of_avenant = Quittance.objects.filter(mouvement_police_id = avenant.id)
            quittances.extend(quittances_of_avenant)
        '''

        # filtrer les quittances impayés
        quittances_impayees = filter(lambda quittance: quittance.statut == StatutQuittance.IMPAYE, quittances)
        quittances_honoraires = filter(lambda quittance: quittance.type_quittance.code == "HONORAIRE", quittances)
        quittances_ristournes = filter(lambda quittance: quittance.nature_quittance.code == "Ristourne", quittances)

        quittances_annulees = Quittance.objects.filter(police_id=police_id, statut_validite=StatutValiditeQuittance.ANNULEE, import_stats=False)

        # etat police = dernier motif
        etat_police = police.etat_police

        documents = Document.objects.filter(quittance__in=quittances)


        context_perso = {'police': police, 'types_quittances': types_quittances, 'quittances': quittances, 'documents': documents,
                         'quittances_impayees': quittances_impayees, 'quittances_honoraires': quittances_honoraires,
                         'quittances_ristournes': quittances_ristournes, 'quittances_annulees': quittances_annulees, 'etat_police': etat_police}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

# new code
@login_required
def add_document_to_quittance(request, quittance_id, police_id):
    if request.method == "POST":
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            police = get_object_or_404(Police, id=police_id)
            quittance = get_object_or_404(Quittance, id=quittance_id)

            type_document_id = request.POST.get('type_document')

            document = form.save(commit=False)
            document.police = police
            document.quittance = quittance
            document.type_document = get_object_or_404(TypeDocument, id=type_document_id)
            document.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': document.pk,
                    'nom': document.nom,
                    'fichier': '<a href="' + document.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                    'type_document': document.type_document.libelle,
                    'confidentialite': document.confidentialite,
                }
            }
            return JsonResponse(response)
        else:
            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire !",
                'errors': form.errors,
            }
            return JsonResponse(response)

@login_required
def details_quittance(request, quittance_id):
    quittance = Quittance.objects.get(id=quittance_id)
    police = quittance.police

    natures_quittances = NatureQuittance.objects.all().order_by('libelle')
    types_quittances = TypeQuittance.objects.all().order_by('libelle')
    types_documents = TypeDocument.objects.all().order_by('libelle')

    taxes_quittances = TaxeQuittance.objects.filter(quittance_id=quittance_id)

    reglements = Reglement.objects.filter(quittance_id=quittance_id, statut_validite=StatutValidite.VALIDE)
    documents = Document.objects.filter(quittance_id=quittance)

    # reglements = Quittance.objects.all()

    return render(request, 'police/modal_details_quittance.html',
                  {'police': police, 'types_quittances': types_quittances, 'natures_quittances': natures_quittances,'types_documents':types_documents,
                   'taxes_quittances': taxes_quittances, 'quittance': quittance, 'reglements': reglements,'documents':documents })

@login_required
def add_quittance(request, police_id):
    police = Police.objects.get(id=police_id)

    error = False
    error_message = ""

    if request.method == 'POST':
        commission_intermediaires = request.POST.get('commission_intermediaire').replace(' ', '')
        nature_quittance_id = request.POST.get('nature_quittance')
        type_quittance_id = request.POST.get('type_quittance')
        prime_ht = request.POST.get('prime_ht').replace(' ', '')
        cout_police_courtier = request.POST.get('cout_police_courtier').replace(' ', '')
        cout_police_compagnie = request.POST.get('cout_police_compagnie').replace(' ', '')
        taxe = request.POST.get('taxe').replace(' ', '')
        autres_taxes = request.POST.get('autres_taxes').replace(' ', '')
        prime_ttc = request.POST.get('prime_ttc').replace(' ', '')
        taux_com_courtage = request.POST.get('taux_com_courtage').replace(' ', '')
        if taux_com_courtage == "": taux_com_courtage = 0
        taux_com_gestion = request.POST.get('taux_com_gestion').replace(' ', '')
        if taux_com_gestion == "": taux_com_gestion = 0
        commission_courtage = request.POST.get('commission_courtage').replace(' ', '')
        if commission_courtage == "": commission_courtage = 0
        commission_gestion = request.POST.get('commission_gestion').replace(' ', '')
        if commission_gestion == "": commission_gestion = 0
        date_emission = request.POST.get('date_emission')
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')

        # Convert numeric values to appropriate types or set default to 0
        prime_ht = int(prime_ht) if prime_ht else 0
        cout_police_courtier = int(cout_police_courtier) if cout_police_courtier else 0
        cout_police_compagnie = int(cout_police_compagnie) if cout_police_compagnie else 0
        taxe = int(taxe) if taxe else 0
        autres_taxes = int(autres_taxes) if autres_taxes else 0
        prime_ttc = int(prime_ttc) if prime_ttc else 0
        solde = prime_ttc  # set solde directly

        taux_com_gestion_formatted = str(taux_com_gestion).replace(',', '.') if taux_com_gestion is not None else ''
        taux_com_courtage_formatted = str(taux_com_courtage).replace(',', '.')

        try:
            taux_com_gestion = float(taux_com_gestion_formatted)
            taux_com_courtage = float(taux_com_courtage_formatted)
        except:
            # Handle the error (e.g., set an error message or use a default value)
            error = True
            error_message = "Ereeur sur le format des données. Ex: " + str(taux_com_gestion)


        commission_courtage = int(commission_courtage) if commission_courtage else 0
        commission_gestion = int(commission_gestion) if commission_gestion else 0
        commission_intermediaires = int(commission_intermediaires) if commission_intermediaires else 0

        devise = police.bureau.pays.devise

        # print(nature_quittance_id == '3')
        # dd(nature_quittance_id)

        # Create Quittance object
        quittance = Quittance.objects.create(police_id=police_id,
                                            compagnie=police.compagnie,
                                            devise=devise,
                                            nature_quittance_id=nature_quittance_id,
                                            type_quittance_id=type_quittance_id,
                                            prime_ht=prime_ht,
                                            cout_police_courtier=cout_police_courtier,
                                            cout_police_compagnie=cout_police_compagnie,
                                            taxe=taxe,
                                            autres_taxes=autres_taxes,
                                            prime_ttc=prime_ttc,
                                            montant_regle=0,
                                            solde=solde,
                                            taux_euro=get_taux_euro_by_devise(devise.code) if devise else None,
                                            taux_usd=get_taux_usd_by_devise(devise.code) if devise else None,
                                            taux_com_gestion=taux_com_gestion,
                                            taux_com_courtage=taux_com_courtage,
                                            commission_courtage=commission_courtage,
                                            commission_gestion=commission_gestion,
                                            commission_intermediaires=commission_intermediaires,
                                            date_emission=date_emission,
                                            date_debut=date_debut,
                                            date_fin=date_fin,
                                            statut=StatutQuittance.IMPAYE,
                                            created_by=request.user,
                                            bureau=request.user.bureau
                                        )

        # Mettre a jour le numero
        code_bureau = request.user.bureau.code
        numero = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(quittance.pk).zfill(7) + '-Q'
        quittance.numero = numero
        quittance.save()

        # enregistrer les autres taxes si existants
        taxes = request.COOKIES.get('taxes_quittance')

        print(taxes)

        if taxes:
            taxes = json.loads(taxes)

            for taxe in taxes:
                taxe = list(taxe.values())
                taxe_id = taxe[0]
                taxe_montant = taxe[1]

                # Insérer la ligne
                TaxeQuittance.objects.create(quittance_id=quittance.id, taxe_id=taxe_id, montant=taxe_montant).save()

        # il s'agit d'une quittance ristourne alors reglons la automatiquement
        if nature_quittance_id == '3':
            # Caclculer le pourcentage des coms qui se trouvent sur la quittance pour déterminer les montants des coms sur les règlements

            date_paiement = datetime.datetime.now(tz=timezone.utc)

            montant_compagnie = prime_ttc - commission_courtage - commission_gestion

            pprint('tx_com_courtage' + str(taux_com_courtage) + 'montant_com_courtage' + str(commission_courtage))

            reglement = Reglement.objects.create(quittance_id=quittance.id,
                                                    montant=prime_ttc,
                                                    montant_compagnie=montant_compagnie,
                                                    compagnie=quittance.compagnie,
                                                    devise_id=devise.pk,
                                                    banque_id=None,
                                                    compte_tresorerie_id=None,
                                                    numero_piece=None,
                                                    montant_com_courtage=commission_courtage,
                                                    montant_com_gestion=commission_gestion,
                                                    montant_com_intermediaire=commission_intermediaires,
                                                    mode_reglement_id=None,
                                                    date_paiement=date_paiement,
                                                    statut_reversement_compagnie=StatutReversementCompagnie.REVERSE,
                                                    created_by=request.user,
                                                    bureau=request.user.bureau)
            reglement.save()
            # mettre à jour son numéro
            reglement.numero = 'R' + str(Date.today().year) + str(reglement.pk).zfill(6)
            reglement.save()

            # mise à jour du solde de la quittance
            quittance.montant_regle = prime_ttc
            quittance.solde = 0
            quittance.statut = StatutQuittance.PAYE
            quittance.updated_at = date_paiement
            quittance.save()

            operation = Operation.objects.create(nature_operation=None,
                                                 numero_piece=None,
                                                 montant_total=prime_ttc,
                                                 compte_tresorerie_id=None,
                                                 devise_id=devise.pk,
                                                 mode_reglement_id=None,
                                                 date_operation=date_paiement,
                                                 created_by=request.user,
                                                 uuid=uuid.uuid4())
            operation.save()

            nombre_quittances = 1

            # Lier l'opération au règlement
            operation_reglement = OperationReglement.objects.create(operation=operation, reglement=reglement, created_by=request.user)
            operation_reglement.save()


            # mettre à jour le total dans operation
            operation.montant_total = prime_ttc
            operation.nombre_quittances = nombre_quittances
            operation.numero = 'OP' + str(Date.today().year) + str(operation.pk).zfill(6)
            operation.save()


        response = {
            'statut': 1,
            'message': "Quittance enregistrée avec succès !",
            'data': {
                'id': quittance.pk,
                'numero': quittance.numero,
                'montant': quittance.prime_ttc,
            }
        }

        return JsonResponse(response)

    else:

        natures_quittances = NatureQuittance.objects.filter(status=True).order_by('libelle')
        types_quittances = TypeQuittance.objects.filter(status=True).order_by('libelle')

        bureau_taxes = BureauTaxe.objects.filter(bureau_id=police.bureau_id)
        for bureau_taxe in bureau_taxes:
            print(bureau_taxe)

        #staxes_police = TaxePolice.objects.filter(police=police)
        taxes_police = BureauTaxe.objects.filter(bureau=police.bureau) #pour être plus flexible, pas obligé que la taxe ait été ajouté sur la police avant qu'elle apparaisse à la création de la quittance

        param_produit = ParamProduitCompagnie.objects.filter(produit=police.produit, compagnie=police.compagnie)[:1].get()

        police_dernier_mouvement = MouvementPolice.objects.filter(police=police, statut_validite=StatutValidite.VALIDE, motif__code__in=["AN", "RENOUV"]).last()

        #get apporteur : un seul apporteur par police
        apporteurs_polices = ApporteurPolice.objects.filter(police=police)

        today = timezone.now()
        return render(request, 'police/modal_add_quittance.html',
                      {'police': police, 'police_dernier_mouvement': police_dernier_mouvement, 'taxes_police': taxes_police, 'param_produit': param_produit, 'today': today,
                       'types_quittances': types_quittances, 'natures_quittances': natures_quittances,
                       'bureau_taxes': bureau_taxes, 'apporteurs_polices': apporteurs_polices})


@login_required
def add_reglement(request, police_id):
    if request.method == 'POST':
        uuid_reglement = request.POST.get('uuid_reglement')
        devise_id = request.POST.get('devise')
        # banque_id = request.POST.get('banque')
        banque_emettrice = request.POST.get('banque')
        compte_tresorerie_id = request.POST.get('compte_tresorerie')
        numero_piece = request.POST.get('numero_piece')
        nature_operation = request.POST.get('nature_operation')
        mode_reglement = request.POST.get('mode_reglement')
        date_paiement = request.POST.get('date_paiement')
        quittances_regles = request.POST.getlist('quittance_regle')
        montants_regles = request.POST.getlist('montant_regle')

        nature_operation_code = "REGCLT"
        nature_operation = NatureOperation.objects.filter(code=nature_operation_code).first()

        #Vérifier si l'uuid n'existe pas déjà dans opération pour s'assurer que l'utilisateur n'as pas cliqué 2 fois
        uuid_reglement_existant = Operation.objects.filter(uuid=uuid_reglement)
        if not uuid_reglement_existant:

            # enregistrer les infos dans operation
            nombre_quittances = 0
            montant_total_regle = 0
            operation = Operation.objects.create(nature_operation=nature_operation,
                                                 numero_piece=numero_piece,
                                                 montant_total=montant_total_regle,
                                                 compte_tresorerie_id=compte_tresorerie_id,
                                                 devise_id=devise_id,
                                                 mode_reglement_id=mode_reglement,
                                                 date_operation=date_paiement,
                                                 created_by=request.user,
                                                 uuid=uuid_reglement)
            operation.save()


            # enregistrer le details dans reglements (liste les quittances reglées avec chaque montant)
            i = 0
            for montant_regle in montants_regles:
                quittance_regle_id = quittances_regles[i]
                i = i + 1

                if montant_regle is not None and quittance_regle_id is not None:
                    montant_regle = float(montant_regle.replace(' ', ''))
                    quittance = Quittance.objects.get(id=quittance_regle_id)

                    if montant_regle > 0 and quittance is not None:

                        # Caclculer le pourcentage des coms qui se trouvent sur la quittance pour déterminer les montants des coms sur les règlements

                        tx_com_courtage = (quittance.commission_courtage * 100) / quittance.prime_ttc
                        tx_com_gestion = (quittance.commission_gestion * 100) / quittance.prime_ttc
                        tx_com_intermediaire = (quittance.commission_intermediaires * 100) / quittance.prime_ttc

                        montant_com_courtage = (tx_com_courtage / 100) * montant_regle
                        montant_com_gestion = (tx_com_gestion / 100) * montant_regle
                        montant_com_intermediaire = (tx_com_intermediaire / 100) * montant_regle
                        montant_compagnie = montant_regle - montant_com_courtage - montant_com_gestion

                        pprint('tx_com_courtage' + str(tx_com_courtage) + 'montant_com_courtage' + str(montant_com_courtage))

                        reglement = Reglement.objects.create(quittance_id=quittance_regle_id,
                                                             montant=montant_regle,
                                                             montant_compagnie=montant_compagnie,
                                                             compagnie=quittance.compagnie,
                                                             devise_id=devise_id,
                                                             #banque_id=banque_id,
                                                             banque_emettrice=banque_emettrice,
                                                             compte_tresorerie_id=compte_tresorerie_id,
                                                             numero_piece=numero_piece,
                                                             montant_com_courtage=montant_com_courtage,
                                                             montant_com_gestion=montant_com_gestion,
                                                             montant_com_intermediaire=montant_com_intermediaire,
                                                             mode_reglement_id=mode_reglement,
                                                             date_paiement=date_paiement,
                                                             created_by=request.user,
                                                             bureau=request.user.bureau)
                        reglement.save()
                        # mettre à jour son numéro
                        reglement.numero = 'R' + str(Date.today().year) + str(reglement.pk).zfill(6)
                        reglement.save()

                        # mise à jour du solde de la quittance
                        quittance.montant_regle = quittance.montant_regle + montant_regle
                        quittance.solde = quittance.solde - montant_regle
                        if quittance.solde == 0: quittance.statut = StatutQuittance.PAYE
                        quittance.updated_at = datetime.datetime.now(tz=timezone.utc)
                        quittance.save()

                        montant_total_regle += montant_regle
                        nombre_quittances = nombre_quittances + 1

                        # Lier l'opération au règlement
                        operation_reglement = OperationReglement.objects.create(operation=operation, reglement=reglement, created_by=request.user)
                        operation_reglement.save()


            # mettre à jour le total dans operation
            operation.montant_total = montant_total_regle
            operation.nombre_quittances = nombre_quittances
            operation.numero = 'OP' + str(Date.today().year) + str(operation.pk).zfill(6)
            operation.save()

            response = {
                'statut': 1,
                'message': "Règlement effectué avec succès !",
                'data': {}
            }

        else:
            response = {
                'statut': 0,
                'message': "Règlement déjà effectué, veuillez vérifier !",
                'data': {}
            }

        return JsonResponse(response)


    else:
        police = Police.objects.get(id=police_id)
        natures_operations = NatureOperation.objects.all()
        devises = Devise.objects.all()
        modes_reglements = ModeReglement.objects.all()
        comptes_tresoreries = CompteTresorerie.objects.filter(status=True)
        banques = Banque.objects.filter(bureau=request.user.bureau, status=True)
        quittances_impayees = Quittance.objects.filter(police_id=police_id, statut=StatutQuittance.IMPAYE, statut_validite=StatutValidite.VALIDE, import_stats=False)

        uuid_reglement = uuid.uuid4()
        today = datetime.datetime.now(tz=timezone.utc)
        return render(request, 'police/modal_add_reglement.html',
                      {'police': police, 'today': today, 'quittances_impayees': quittances_impayees, 'devises': devises,
                       'natures_operations': natures_operations, 'modes_reglements': modes_reglements,
                       'banques': banques, 'comptes_tresoreries': comptes_tresoreries, 'uuid_reglement': uuid_reglement})


# all police avenants
class PoliceAvenantsView(TemplateView):
    template_name = 'police/avenants.html'
    model = Police

    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).first()
        if police:
            mouvements = Mouvement.objects.filter(type="POLICE").exclude(code="AN").order_by('libelle')
            mouvements_police = MouvementPolice.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).order_by('-id') #, statut_validite=StatutValidite.VALIDE

            # etat police = dernier motif
            etat_police = police.etat_police

            pprint("(((((( ETAT POLICE )))))")
            pprint(etat_police)

            if etat_police != "Suspendu":
                # Retirer mise en vigueur (REMVIG) sauf cas de suspention
                mouvements = mouvements.exclude(code="REMVIG").order_by('libelle')


            context_perso = {'police': police, 'mouvements_police': mouvements_police, 'mouvements': mouvements,
                             'etat_police': etat_police}

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# all police garanties
class PoliceGedView(TemplateView):
    template_name = 'police/ged.html'
    model = Police

    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).first()
        if police:
            types_documents = TypeDocument.objects.all().order_by('libelle')
            documents = Document.objects.filter(police_id=police_id)

            # etat police = dernier motif
            etat_police = police.etat_police

            context_perso = {'police': police, 'documents': documents, 'types_documents': types_documents,
                             'etat_police': etat_police, }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)
        else:
            return redirect("clients")

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def police_add_document(request, police_id):
    if request.method == "POST":

        form = DocumentForm(request.POST, request.FILES)

        if form.is_valid():

            police = Police.objects.get(id=police_id)
            type_document_id = request.POST.get('type_document')

            document = form.save(commit=False)
            document.client = police.client
            document.police = police
            document.type_document = TypeDocument.objects.get(id=type_document_id)
            document.save()

            pprint("document.fichier")
            pprint(document.fichier.path)



            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': document.pk,
                    'nom': document.nom,
                    'fichier': '<a href="' + document.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                    'type_document': document.type_document.libelle,
                    'confidentialite': document.confidentialite,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire !",
                'errors': form.errors,
            }

            return JsonResponse(response)


def aliment_add_document(request, aliment_id):
    if request.method == "POST":

        form = DocumentForm(request.POST, request.FILES)

        if form.is_valid():

            aliment = Aliment.objects.get(id=aliment_id)
            type_document_id = request.POST.get('type_document')

            document = form.save(commit=False)
            document.aliment = aliment
            document.type_document = TypeDocument.objects.get(id=type_document_id)
            document.save()

            pprint(document)

            response = {
                'statut': 1,
                'message': "Enregistrement effectué avec succès !",
                'data': {
                    'id': document.pk,
                    'nom': document.nom,
                    'fichier': '<a href="' + document.fichier.url + '"><i class="fa fa-file" title="Aperçu"></i> Afficher</a>',
                    'type_document': document.type_document.libelle,
                    'confidentialite': document.confidentialite,
                }
            }

            return JsonResponse(response)

        else:

            response = {
                'statut': 0,
                'message': "Veuillez renseigner correctement le formulaire !",
                'errors': form.errors,
            }

            return JsonResponse(response)


# get all sinistres for police
class PoliceSinistresView(TemplateView):
    template_name = 'police/sinistres.html'
    model = DossierSinistre

    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).first()
        if police:
            prestataires_ids = Sinistre.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).values_list('prestataire_id', flat=True)
            prestataires = Prestataire.objects.filter(id__in=prestataires_ids).order_by('name')

            # etat police = dernier motif
            etat_police = police.etat_police

            context_perso = {'police': police, 'etat_police': etat_police, 'dossiers_sinistres': None,
                             'sinistres': None, 'prestataires': prestataires}

            context = {**context_original, **context_perso}

            return self.render_to_response(context)
        else:
            return redirect("clients")

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def police_sinistres_datatable(request, police_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    search_numero_assure = request.GET.get('num_assure', '')
    search_numero_sinistre = request.GET.get('num_sinistre', '')
    search_numero_dossier_sinistre = request.GET.get('num_feuille_soins', '')
    search_date_survenance = request.GET.get('date_prestation', '')
    prestataire = request.GET.get('prestataire', '')
    search_value = request.GET.get('search[value]')

    queryset = Sinistre.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).order_by('id')
    # dd(queryset)

    if prestataire:
        queryset = queryset.filter(prestataire_id=prestataire)

    if search_numero_assure:
        cartes = Carte.objects.filter(numero=search_numero_assure)
        carte = cartes.first() if cartes else None
        aliment = carte.aliment if carte else None
        queryset = queryset.filter(aliment_id=aliment.pk) if aliment else queryset.filter(numero="nexisterajamais")

    if search_numero_dossier_sinistre:
        queryset = queryset.filter(dossier_sinistre__numero__contains=search_numero_dossier_sinistre)

    if search_numero_sinistre:
        queryset = queryset.filter(numero__contains=search_numero_sinistre)

    if search_date_survenance:
        queryset = queryset.filter(date_survenance__contains=search_date_survenance)

    if search_value is not None and search_value != '':
        queryset = queryset.filter(
        Q(numero__contains=search_value) |
        Q(dossier_sinistre__numero__contains=search_value) |
        Q(aliment__nom__icontains=search_value) | Q(aliment__prenoms__icontains=search_value) |
        Q(prestataire__name__icontains=search_value) |
        Q(date_survenance__contains=search_value)).distinct()


    # Map column index to corresponding model field for sorting
    sort_columns = {
        0: '-numero',
        1: 'aliment__nom',
        2: 'statut',
        # Add more columns as needed
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')

    if sort_direction == 'desc':
        sort_column = '-' + sort_column  # For descending order

    # Apply sorting
    queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:
        detail_url = reverse('details_dossier_sinistre', args=[c.dossier_sinistre.id]) if c.dossier_sinistre else None # URL to the detail view# URL to the detail view
        actions_html = f'<a href="{detail_url}" class="text-center"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> {_("Détails")}</span></a>&nbsp;&nbsp;'

        statut_html = f'<span class="badge badge-{c.statut.lower()}">{c.statut}</span>'

        nom_beneficiaire_soins = ''
        if c.aliment:
            nom_beneficiaire_soins = c.aliment.nom + ' ' + c.aliment.prenoms

        data_iten = {
            "id": c.id,
            "numero": c.numero if c.numero else "",
            "numero_dossier_sinistre": c.dossier_sinistre.numero if c.dossier_sinistre else '',
            "nom_beneficiaire_soins": nom_beneficiaire_soins,
            "prestataire": c.prestataire.name if c.prestataire else '',
            "acte": c.acte.libelle,
            "total_frais_reel": money_field(c.total_frais_reel),
            "total_part_compagnie": money_field(c.total_part_compagnie),
            "total_part_assure": money_field(c.total_part_assure),
            "date_survenance": c.date_survenance.strftime("%d/%m/%Y %H:%M"),
            "date_saisie": c.created_at.strftime("%d/%m/%Y %H:%M"),
            "statut": statut_html,
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


@method_decorator(login_required, name='dispatch')
# get all beneficiaires for police
class PoliceBeneficiairesView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'police/beneficiaires.html'
    model = Aliment

    def get(self, request, police_id, aliment_id=None, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).first()
        if police:

            liste_des_formules = [] #FormuleGarantie.objects.filter(police_id=police_id)
            polices_du_bureau_actif = None #Police.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE)

            # aliments = getAliments(police_id)

            adherents_principaux = None #getAdherentsPrincipaux(police_id)

            qualites_beneficiaires = [] #QualiteBeneficiaire.objects.all().order_by('libelle')
            civilites = [] #Civilite.objects.all().exclude(code="STE").order_by('name')

            formules = [] #police.formules.order_by('libelle')

            pays = [] #Pays.objects.all().order_by('nom')
            professions = [] #Profession.objects.all().order_by('name')

            # etat police = dernier motif
            etat_police = MouvementPolice.objects.filter(police_id=police_id, statut_validite=StatutValidite.VALIDE).order_by('-id').first().motif.libelle

            aliment = Aliment.objects

            option_export_beneficiaires = police.bureau.option_export_beneficiaires

            today = timezone.now().date()

            prospect_to_open = self.request.GET.get('prospect_to_open', None)

            context_perso = {'police': police, 'adherents_principaux': adherents_principaux,
                             'qualites_beneficiaires': qualites_beneficiaires, 'civilites': civilites, 'aliments': None,
                             'formules': formules, 'pays': pays, 'professions': professions, 'etat_police': etat_police,
                             'liste_des_formules': liste_des_formules, 'polices_du_bureau_actif': polices_du_bureau_actif,
                             'option_export_beneficiaires': option_export_beneficiaires, 'aliment_id': aliment_id,
                             'today': today, 'prospect_to_open': prospect_to_open}

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def get_formules(request, police_id):
    # Récupérez les formules associées à la police_id (ajustez selon votre modèle)
    formules = FormuleGarantie.objects.filter(police_id=police_id).values('code', 'libelle')

    # Convertissez les résultats en un format JSON approprié
    formules_list = [{'code': formule['code'], 'libelle': formule['libelle']} for formule in formules]

    return JsonResponse({'formules': formules_list})


def police_beneficiaires_datatable(request, police_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = request.GET.get('order[0][column]')
    sort_direction = request.GET.get('order[0][dir]')
    search_value = request.GET.get('search[value]')
    search_etat = request.GET.get('search_etat', '')
    # draw = int(request.GET.get('draw', 1))

    print('search_value')
    print(search_value)

    # TODO police_beneficiaires_datatable
    police = Police.objects.get(id=police_id)
    etat_police = police.etat_police

    #Pour éviter les doublons
    # Récupérer les IDs des aliments dans aliment_formule
    aliment_ids = AlimentFormule.objects.filter(
        formule_id__in=[p.id for p in police.formules],
        statut=Statut.ACTIF,
        statut_validite=StatutValidite.VALIDE
    ).values_list('aliment_id', flat=True)

    today = datetime.datetime.now(tz=timezone.utc).date()

    queryset = Aliment.objects.filter(id__in=aliment_ids).order_by('-adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

    if search_etat:

        if search_etat == 'ACTIF':

            #retirer les suspendus
            aliments_id_suspendus = MouvementAliment.objects.filter(
                Q(police_id=police.id, mouvement__code="SUSPENSION-BENEF", date_effet__lte=today, statut_validite=StatutValidite.VALIDE)
            ).values_list('aliment_id', flat=True)

            aliments_id_remis_en_vigueur = MouvementAliment.objects.filter(
                police_id=police.id, mouvement__code="REMISEVIGUEUR-BENEF", date_effet__lte=today, statut_validite=StatutValidite.VALIDE
            ).values_list('aliment_id', flat=True)

            aliments_id_suspendus_reel = Aliment.objects.filter(id__in=aliments_id_suspendus).exclude(id__in=aliments_id_remis_en_vigueur).values_list('id', flat=True)

            queryset = Aliment.objects.filter(id__in=aliment_ids, date_sortie__isnull=True, statut_incorporation=StatutIncorporation.INCORPORE).exclude(id__in=aliments_id_suspendus_reel).order_by('-adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')


        elif search_etat == 'SUSPENDU':# a revoir: si plusieurs suspensions et remise en vigueur, ça pourrais ne pas marcher: considérer les derniers ...
            aliments_id_suspendus = MouvementAliment.objects.filter(
                Q(police_id=police.id, mouvement__code="SUSPENSION-BENEF", date_effet__lte=today, statut_validite=StatutValidite.VALIDE)
            ).values_list('aliment_id', flat=True)

            aliments_id_remis_en_vigueur = MouvementAliment.objects.filter(
                police_id=police.id, mouvement__code="REMISEVIGUEUR-BENEF", date_effet__lte=today, statut_validite=StatutValidite.VALIDE
            ).values_list('aliment_id', flat=True)

            queryset = Aliment.objects.filter(id__in=aliments_id_suspendus).exclude(id__in=aliments_id_remis_en_vigueur).order_by('-adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

        elif search_etat == 'SORTI':
            queryset = Aliment.objects.filter(id__in=aliment_ids, date_sortie__isnull=False, date_sortie__lte=today).order_by('-adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

        elif search_etat == 'LISTE_VIDE':
            #renvoyer une liste vide
            queryset = Aliment.objects.none()

    if search_value is not None and search_value != '':
        queryset = queryset.filter(
        Q(nom__icontains=search_value) | Q(prenoms__icontains=search_value) |
        Q(matricule_cie__icontains=search_value) |
        Q(adherent_principal__numero_famille__icontains=search_value) |
        # Q(formule__libelle__icontains=search_value) |
        Q(qualite_beneficiaire__libelle__icontains=search_value) |
        Q(genre__icontains=search_value) |
        Q(date_naissance__icontains=search_value) |
        Q(cartes__numero__contains=search_value)).distinct()

    print(queryset)


    # Map column index to corresponding model field for sorting
    #sort_columns = {
    #     0: 'aliment__adherent_principal__numero_famille',
    #     1: 'aliment__nom',
    #     2: 'aliment__prenoms',
    #     3: '-aliment__cartes__numero',
    #     4: 'aliment__date_naissance',
    #     5: 'aliment__qualite_beneficiaire__libelle',
    #     6: 'aliment__matricule_cie',
    #     7: 'aliment__formule__libelle',
    #     # Add more columns as needed
    #}
    #
    #sort_column = sort_columns.get(sort_column_index, '-id')
    #
    #if sort_direction == 'desc':
    #    sort_column = '-' + sort_column  # For descending order
    #
    # # Apply sorting
    #queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    liste_aliments_ajoutes = []

    for c in page_obj:

        if c.id not in liste_aliments_ajoutes:
            liste_aliments_ajoutes.append(c.id)

            details_beneficiaire_url = reverse('details_beneficiaire', args=[police_id, c.id])
            update_beneficiaire_url = reverse('update_beneficiaire', args=[police_id, c.id])
            actions_html = f'<span style="cursor:pointer;" class="btn_details_beneficiaire" data-aliment_id="{c.id}" data-model_name="aliment" data-modal_title="FICHE ADHERENT - {c.nom} {c.prenoms}" data-href="{details_beneficiaire_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> {_("Détails")}</span></span>&nbsp;&nbsp; \
                            <span style="cursor:pointer;" class="btn_modifier_beneficiaire" data-aliment_id="{c.id}" data-model_name="aliment" data-modal_title="MODIFICATION D\'UN BENEFICIAIRE" data-href="{update_beneficiaire_url}"><span class="badge btn-sm btn-modifier rounded-pill"><i class="fa fa-edit"></i> {_("Modifier")}</span></span>&nbsp;&nbsp;<img id="loder-aliment-{c.id}" src="/static/admin_custom/img/loader.gif" style="width: 20px;display: none;">'

            today = timezone.now().date()

            # Convertir c.date_sortie en date si nécessaire
            date_sortie = c.date_sortie.date() if isinstance(c.date_sortie, datetime.datetime) else c.date_sortie

            #sa_formule_today = c.aliment.formule_atdate(today)
            last_formule = c.last_formule

            #Statut.ACTIF if sa_formule_today else Statut.INACTIF
            etat_beneficiaire = c.etat_beneficiaire
            if c.date_sortie and c.date_sortie <= today:
                etat_beneficiaire = 'SORTI'

            #pour le visuel si la police est résilié affiché INACTIF sur tous les bénéficiaires
            if etat_police == "Résilié":
                etat_beneficiaire = 'SORTI'
            if etat_police == "Suspendu":
                etat_beneficiaire = 'SUSPENDU'

            #si la police est échue
            if police.is_echue:
                etat_beneficiaire = 'SORTI'


            etat_beneficiaire_html = f'<span class="badge badge-'+etat_beneficiaire.lower()+'">'+etat_beneficiaire+'</span>'


            data_iten = {
                "id": c.id,
                #"aliment__cartes__numero": f'{c.cartes.filter(statut=Statut.ACTIF).latest("id")}' if c.carte_active else "",
                "aliment__cartes__numero": f'{c.carte_active()}' if c.carte_active else "",
                "aliment__nom": c.nom if c.nom else "",
                "aliment__prenoms": c.prenoms if c.prenoms else "",
                "aliment__date_naissance": c.date_naissance.strftime("%d/%m/%Y") if c.date_naissance else "",
                "aliment__genre": c.genre if c.genre else "",
                "aliment__qualite_beneficiaire__libelle": c.qualite_beneficiaire.libelle if c.qualite_beneficiaire else "",
                "aliment__matricule_employe": c.matricule_employe if c.matricule_employe else "",
                "aliment__adherent_principal__numero_famille": c.adherent_principal.numero_famille if c.adherent_principal else "",
                #"aliment__formule__libelle": sa_formule_today.libelle if sa_formule_today else last_formule.libelle if last_formule else 'd',
                "aliment__formule__libelle": last_formule.libelle if last_formule else '',
                "aliment__statut": etat_beneficiaire_html,
                "actions": actions_html,
            }


            data.append(data_iten)


    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


def transfert_beneficiaires_datatable(request, client_id):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = request.GET.get('order[0][column]')
    sort_direction = request.GET.get('order[0][dir]')
    search_value = request.GET.get('search[value]')
    # draw = int(request.GET.get('draw', 1))
    search_police_origine = request.GET.get('search_police_origine', 0)
    search_police_destination = request.GET.get('search_police_destination', 0)
    search_formule_origine = request.GET.get('search_formule_origine', 0)
    search_etat_beneficiaire = request.GET.get('search_etat_beneficiaire', 0)

    print('search_value')
    print(search_value)

    # TODO police_beneficiaires_datatable
    if search_police_origine:
        police = Police.objects.get(id=search_police_origine)

        #Pour éviter les doublons
        # Récupérer les IDs des aliments dans aliment_formule
        aliment_ids = AlimentFormule.objects.filter(
            formule_id__in=[p.id for p in police.formules],
            statut=Statut.ACTIF,
            statut_validite=StatutValidite.VALIDE
        ).values_list('aliment_id', flat=True)

        # Récupérer les aliments correspondants à ces IDs
        queryset = Aliment.objects.filter(id__in=aliment_ids).order_by('-adherent_principal_id', 'qualite_beneficiaire_id','nom','prenoms')

        #si police destination renseignée
        if search_police_destination:
            police_destination = Police.objects.filter(id=search_police_destination, statut_validite=StatutValidite.VALIDE).first()
            formules_sur_police_destination = FormuleGarantie.objects.filter(police=police_destination)

            pprint("formules_sur_police_destination")
            pprint(formules_sur_police_destination)

            # Exclure ceux déjà sur la police destination
            aliment_ids_a_exclure = AlimentFormule.objects.filter(
                formule_id__in=[f.id for f in formules_sur_police_destination],
                # statut=Statut.ACTIF,
                statut_validite=StatutValidite.VALIDE
            ).values_list('aliment_id', flat=True)

            pprint("aliment_ids_a_exclure")
            pprint(aliment_ids_a_exclure)

            queryset = queryset.exclude(id__in=aliment_ids_a_exclure) #added on 03062024


        if search_value is not None and search_value != '':
            queryset = queryset.filter(
            Q(nom__icontains=search_value) | Q(prenoms__icontains=search_value) |
            Q(matricule_cie__icontains=search_value) |
            Q(adherent_principal__numero_famille__icontains=search_value) |
            # Q(formule__libelle__icontains=search_value) |
            Q(qualite_beneficiaire__libelle__icontains=search_value) |
            Q(genre__icontains=search_value) |
            Q(date_naissance__icontains=search_value) |
            Q(cartes__numero__contains=search_value)).distinct()

        print(queryset)


        # Map column index to corresponding model field for sorting
        #sort_columns = {
        #     0: 'aliment__adherent_principal__numero_famille',
        #     1: 'aliment__nom',
        #     2: 'aliment__prenoms',
        #     3: '-aliment__cartes__numero',
        #     4: 'aliment__date_naissance',
        #     5: 'aliment__qualite_beneficiaire__libelle',
        #     6: 'aliment__matricule_cie',
        #     7: 'aliment__formule__libelle',
        #     # Add more columns as needed
        #}
        #
        #sort_column = sort_columns.get(sort_column_index, '-id')
        #
        #if sort_direction == 'desc':
        #    sort_column = '-' + sort_column  # For descending order
        #
        # # Apply sorting
        #queryset = queryset.order_by(sort_column)

        paginator = Paginator(queryset, length)
        page_obj = paginator.get_page(page_number)

        # Prepare the data in the expected format
        data = []

        for c in page_obj:

            today = datetime.datetime.now(tz=timezone.utc).date()
            #sa_formule_today = c.aliment.formule_atdate(today)
            last_formule = c.last_formule

            #Statut.ACTIF if sa_formule_today else Statut.INACTIF
            etat_beneficiaire = c.etat_beneficiaire
            if c.date_sortie and c.date_sortie <= today:
                etat_beneficiaire = 'SORTI'
            etat_beneficiaire_html = f'<span class="badge badge-'+etat_beneficiaire.lower()+'">'+etat_beneficiaire+'</span>'

            if etat_beneficiaire != "SORTI":
                actions_html = f'<input type="checkbox" class="checkbox_transfert_beneficiaire" value="{c.id}" />'
            else:
                actions_html = f'<input type="checkbox" class="" disabled/>'


            data_iten = {
                "checkbox": actions_html,
                "id": c.id,
                #"aliment__cartes__numero": f'{c.cartes.filter(statut=Statut.ACTIF).latest("id")}' if c.carte_active else "",
                "aliment__cartes__numero": f'{c.carte_active()}' if c.carte_active else "",
                "aliment__nom": c.nom if c.nom else "",
                "aliment__prenoms": c.prenoms if c.prenoms else "",
                "aliment__date_naissance": c.date_naissance.strftime("%d/%m/%Y") if c.date_naissance else "",
                "aliment__genre": c.genre if c.genre else "",
                "aliment__qualite_beneficiaire__libelle": c.qualite_beneficiaire.libelle if c.qualite_beneficiaire else "",
                "aliment__matricule_cie": c.matricule_cie if c.matricule_cie else "",
                "aliment__adherent_principal__numero_famille": c.adherent_principal.numero_famille if c.adherent_principal else "",
                #"aliment__formule__libelle": sa_formule_today.libelle if sa_formule_today else last_formule.libelle if last_formule else 'd',
                "aliment__formule__libelle": last_formule.libelle if last_formule else '',
                "aliment__statut": etat_beneficiaire_html,
            }

            if etat_beneficiaire != "SORTI":
                if search_formule_origine:
                    if last_formule.id == int(search_formule_origine):
                        data.append(data_iten)

                else:
                    data.append(data_iten)


        return JsonResponse({
            "data": data,
            "recordsTotal": queryset.count(),
            "recordsFiltered": paginator.count,
            "draw": int(request.GET.get('draw', 1)),
        })

    else:

        data = []
        return JsonResponse({
            "data": data,
            "recordsTotal": 0,
            "recordsFiltered": 0,
            "draw": int(request.GET.get('draw', 1)),
        })


@login_required()
def transfert_beneficiaires(request, client_id):
    rstatut = 0
    rmessage = ""
    rerrors = {}

    try:
        if request.method == 'POST':
            client = Client.objects.get(id=client_id)
            if client.bureau == request.user.bureau:

                police_origine_id = request.POST.get('police_origine_id')
                formule_origine_id = request.POST.get('formule_origine_id')
                police_destination_id = request.POST.get('police_destination_id')
                formule_destination_id = request.POST.get('formule_destination_id')
                formule_destination_id = request.POST.get('formule_destination_id')

                aliments_ids = literal_eval(request.POST.get('selectedItems'))
                print("aliments_ids")
                print(aliments_ids)

                police_origine = Police.objects.filter(id=police_origine_id, statut_validite=StatutValidite.VALIDE).first()

                police_destination = Police.objects.filter(id=police_destination_id, statut_validite=StatutValidite.VALIDE).first()
                formule_destination = FormuleGarantie.objects.filter(id=formule_destination_id).first()
                aliments = Aliment.objects.filter(id__in=aliments_ids, date_sortie__isnull=True)

                if police_destination and formule_destination:

                    #Pour chaque aliment de la liste

                    # Sortir l'aliment de la formule en ajoutant la date de fin
                    pprint("TRAITEMENT DES BENEFICIAIRES")
                    #Prendre ceux la date de sortie est vide ou supérieur à la date_debut_effet de la nouvelle police
                    #

                    for aliment in aliments:

                        #récupérer sa dernière formule
                        aliment_formule_old = AlimentFormule.objects.filter(aliment=aliment, formule__police=police_origine).filter(Q(date_fin__isnull=True)).first()

                        is_on_police_destination = AlimentFormule.objects.filter(aliment=aliment, formule__police=police_destination).first()

                        if not is_on_police_destination:

                            today = datetime.datetime.now(tz=timezone.utc)

                            date_fin_ancienne_formule = aliment_formule_old.date_debut #today
                            date_debut_nouvelle_formule = aliment_formule_old.date_debut #today

                            # Sortir les anciennes lignes
                            pprint("Le sortir de l'ancienne formule")
                            aliment_formule_old.date_fin = date_fin_ancienne_formule
                            aliment_formule_old.motif = f'Transféré sur la police {police_destination.numero}, formule:{formule_destination.code}'
                            #aliment_formule_old.statut = Statut.INACTIF
                            aliment_formule_old.save()

                            #Créer une niouvelle ligne avec la nouvelle formule
                            pprint("Créer une nouvelle ligne avec la nouvelle formule")
                            new_af = AlimentFormule.objects.create(aliment=aliment, formule=formule_destination, date_debut=date_debut_nouvelle_formule, statut=Statut.ACTIF, created_by=request.user)


                    #save log
                    ActionLog.objects.create(done_by=request.user, action="transfert_beneficiaires",
                                             description="Transfert de bénéficiaire: copie des bénéficiaires de la police " + str(police_origine.numero) + " vers la police " + str(police_destination.numero) + ", new_formule:"  + str(formule_destination.code), table="",
                                             row=None, data_before=None,
                                             data_after=None)

                    rstatut = 1
                    rmessage = "Opération effectuée avec succès."
                    rerrors = {}

                else:
                    rstatut = 0
                    message = "Vous n'êtes pas autorisé à faire cette opération"

            else:
                rstatut = 0
                message = "Polices non trouvées"


    except Exception as e:
        error_message = str(e)

        rmessage = "Erreur lors du traitement: " + error_message
        rerrors = {'Erreur': error_message}

    response = {
        'statut': rstatut,
        'message': rmessage,
        'errors': rerrors,
    }

    return JsonResponse(response)


@never_cache
def police_beneficiaires(request, police_id):
    return render(request, 'police/beneficiaires.html', )


def getAliments(police_id):
    aliments = []

    police = Police.objects.filter(pk=police_id, statut_validite=StatutValidite.VALIDE)
    for formule in police.formules:

        aliment_formule = AlimentFormule.objects.filter(formule_id=formule.id, statut=Statut.ACTIF).order_by('-id')

        for af in aliment_formule:
            if af.aliment not in aliments:
                aliments.append(af.aliment)

    return aliments


def getAdherentsPrincipaux(police_id):
    aliments = []

    police = Police.objects.get(id=police_id)

    for formule in police.formules:

        aliment_formule = AlimentFormule.objects.filter(formule_id=formule.id, statut=Statut.ACTIF, statut_validite=StatutValidite.VALIDE).order_by('-id')

        for af in aliment_formule:
            if af.aliment.statut == Statut.ACTIF and af.aliment.qualite_beneficiaire and af.aliment.qualite_beneficiaire.code == "AD":
                if af.aliment not in aliments:
                    aliments.append(af.aliment)

    return aliments


def check_pandas_value(value):
    return value if pd.notna(value) else None


def check_pandas_date_value(value):
    return pd.to_datetime(value) if pd.notna(value) else None


@transaction.atomic
def importation_beneficiaire_from_ancienne_police(request, police_id):

    from datetime import datetime

    # Date et l'heure actuelles
    date_daujourdhui = datetime.now()

    rstatut = 0
    rmessage = ""
    rerrors = {}

    try:
        if request.method == 'POST':

            formule_sortante_code = request.POST.get('formule_sortante_code')
            nouvelle_formule_code = request.POST.get('nouvelle_formule_code')

            print("FORM VALUE")
            print(formule_sortante_code)
            print(nouvelle_formule_code)

            formule_sortante = FormuleGarantie.objects.filter(code=formule_sortante_code).first()
            new_formule = FormuleGarantie.objects.filter(code=nouvelle_formule_code).first()

            print("AFFICHAGE FORMULE SORTANTE")
            print(formule_sortante.id)
            pprint(formule_sortante)
            print("FIN AFFICHAGE FORMULE SORTANTE")


            print("AFFICHAGE NOUVELLE FORMULE")
            print(new_formule.id)
            pprint(new_formule)
            print("FIN AFFICHAGE NOUVELLE FORMULE")

            aliments_formules = AlimentFormule.objects.filter(formule_id=formule_sortante.id)

            # Recuperer les ids des aliments qui sont sur cette formule
            ids_aliments_concernees = [alimentformule.aliment_id for alimentformule in aliments_formules]

            print("AFFICHAGE DES IDS DES ALIMENTS")
            pprint(ids_aliments_concernees)

            # Sortir l'aliment de la formule en ajoutant la date de fin
            for aliment_formule in aliments_formules:
                # Mettre date_fin a now()
                aliment_formule.date_fin = date_daujourdhui
                aliment_formule.motif = "Changement de formule"
                aliment_formule.statut = Statut.INACTIF
                aliment_formule.save()


            # Enregistrer les aliments et la formule dans AlimentFormules ou inserer les nouvelles lignes avec la nouvelle formule
            for aliment_id in ids_aliments_concernees:
                aliment = Aliment.objects.get(id=aliment_id)
                af = AlimentFormule.objects.create(aliment=aliment, formule=new_formule, date_debut=date_daujourdhui, statut=Statut.ACTIF)




            rstatut = 1
            rmessage = "Changement de formule reussi."
            rerrors = {}

    except Exception as e:
        error_message = str(e)

        rmessage = "Erreur lors du changement de formule =>  " + error_message
        rerrors = {'Erreur': error_message}

    response = {
        'statut': rstatut,
        'message': rmessage,
        'errors': rerrors,
    }

    return JsonResponse(response)


#TODO CHANGEMENT DE COMPAGNIE
@transaction.atomic
def changement_compagnie(request, client_id):
    #dd("Désactivé pour l'instant")

    rstatut = 0
    rmessage = ""
    rerrors = {}

    try:
        if request.method == 'POST':

            ancienne_police_id = request.POST.get('ancienne_police')
            nouvelle_police_id = request.POST.get('nouvelle_police')

            ancienne_police = Police.objects.filter(id=ancienne_police_id, statut_validite=StatutValidite.VALIDE).first()
            nouvelle_police = Police.objects.filter(id=nouvelle_police_id, statut_validite=StatutValidite.VALIDE).first()

            if ancienne_police and nouvelle_police:
                anciennes_formules = FormuleGarantie.objects.filter(police=ancienne_police)

                #Pour chaque formule
                pprint("Pour chaque formule")
                for old_formule in anciennes_formules:

                    # vérifier qu'il y a des bénéficiaires actif sur la formule
                    pprint("TRAITEMENT DES FORMULES DE GARANTIE ET SPECIFICITES")
                    pprint("vérifier qu'il y a des bénéficiaires actif sur la formule")

                    #Créer la nouvelle formule
                    pprint("Créer la nouvelle formule")
                    new_formule = FormuleGarantie.objects.create(police=nouvelle_police,
                                                                 created_by=request.user,
                                                                 territorialite=old_formule.territorialite,
                                                                 type_tarif=old_formule.type_tarif,
                                                                 libelle=old_formule.libelle,
                                                                 taux_couverture=old_formule.taux_couverture,
                                                                 plafond_conso_famille=old_formule.plafond_conso_famille,
                                                                 plafond_conso_individuelle=old_formule.plafond_conso_individuelle,
                                                                 taux_tm=old_formule.taux_tm,
                                                                 garantis_pharmacie=old_formule.garantis_pharmacie,
                                                                 date_debut=nouvelle_police.date_debut_effet,
                                                                 )
                    code_bureau = request.user.bureau.code
                    code_formule = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(new_formule.pk).zfill(7)

                    new_formule.code = 'F' + str(Date.today().year)[-2:] + '' + str(new_formule.pk).zfill(7)
                    new_formule.save()


                    #Réccupérer les barèmes de l'ancienne formule
                    pprint("Réccupérer les barèmes de l'ancienne formule")
                    anciens_baremes = Bareme.objects.filter(formulegarantie=old_formule)

                    #Copier les barèmes sur la nouvelle formule
                    pprint("Copier les barèmes sur la nouvelle formule")
                    for old_bareme in anciens_baremes:
                        new_bareme = Bareme.objects.create(formulegarantie=new_formule,
                                                           created_by=request.user,
                                                           rubrique_id=old_bareme.rubrique_id,
                                                           sous_rubrique_id=old_bareme.sous_rubrique_id,
                                                           regroupement_acte_id=old_bareme.regroupement_acte_id,
                                                           acte_id=old_bareme.acte_id,
                                                           is_garanti=old_bareme.is_garanti,
                                                           qualite_beneficiaire_id=old_bareme.qualite_beneficiaire_id,
                                                           taux_couverture=old_bareme.taux_couverture,
                                                           taux_tm=old_bareme.taux_tm,
                                                           plafond_individuel=old_bareme.plafond_individuel,
                                                           plafond_famille=old_bareme.plafond_famille,
                                                           plafond_rubrique=old_bareme.plafond_rubrique,
                                                           plafond_sous_rubrique=old_bareme.plafond_sous_rubrique,
                                                           plafond_regroupement_acte=old_bareme.plafond_regroupement_acte,
                                                           plafond_acte=old_bareme.plafond_acte,
                                                           nombre_acte=old_bareme.nombre_acte,
                                                           periodicite_id=old_bareme.periodicite_id,
                                                           # unite_frequence = unite_frequence,
                                                           # frequence = frequence,
                                                           age_minimum=old_bareme.age_minimum,
                                                           age_maximum=old_bareme.age_maximum,
                                                           date_debut=nouvelle_police.date_debut_effet,
                                                           )
                        new_bareme.save()
                        pprint("new_bareme crée")
                        pprint(new_bareme)


                    # Sortir l'aliment de la formule en ajoutant la date de fin
                    pprint("TRAITEMENT DES BENEFICIAIRES")
                    #Prendre ceux la date de sortie est vide ou supérieur à la date_debut_effet de la nouvelle police
                    aliments_formules_old = AlimentFormule.objects.filter(formule=old_formule).filter(Q(date_fin__isnull=True) | Q(date_fin__gte=nouvelle_police.date_debut_effet))

                    for old_af in aliments_formules_old:
                        aliment = old_af.aliment

                        # Sortir les anciennes lignes
                        pprint("Sortir les anciennes lignes")
                        old_af.date_fin = ancienne_police.date_fin_effet
                        old_af.motif = "Changement de compagnie"
                        #old_af.statut = Statut.INACTIF
                        old_af.save()


                        #Créer une niouvelle ligne avec la nouvelle formule
                        pprint("Créer une niouvelle ligne avec la nouvelle formule")
                        new_af = AlimentFormule.objects.create(aliment=aliment, formule=new_formule, date_debut=nouvelle_police.date_debut_effet, statut=Statut.ACTIF, created_by=request.user)

                #save log
                ActionLog.objects.create(done_by=request.user, action="changement_assureur",
                                         description="Changement d'assureur: copie des formules et bénéficiaires de la police " + str(ancienne_police.numero) + " vers la police " + str(nouvelle_police.numero), table="",
                                         row=None, data_before=None,
                                         data_after=None)

                rstatut = 1
                rmessage = "Opération effectuée avec succès."
                rerrors = {}

            else:
                rstatut = 0
                message = "Polices non trouvées"


    except Exception as e:
        error_message = str(e)

        rmessage = "Erreur lors du traitement: " + error_message
        rerrors = {'Erreur': error_message}

    response = {
        'statut': rstatut,
        'message': rmessage,
        'errors': rerrors,
    }

    return JsonResponse(response)


# TODO IMPORTATION DE BENEFICIAIRE
#Avec des entête personnalisées V2

@transaction.atomic
def verifier_beneficiaires(request, police_id):

    if request.method == 'POST' and request.FILES['fichier']:

        police = Police.objects.get(id=police_id)
        print("Police")
        pprint(police)

        try:
            fichier = request.FILES['fichier']
        except MultiValueDictKeyError:
            fichier = False

        try:
            fs = FileSystemStorage()
            file_name_renamed = fichier.name.replace(" ", "_")

            filename = fs.save(file_name_renamed, fichier)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            excel_file_full_path = fs.path(filename)

            pprint("excel_file_full_path")
            pprint(excel_file_full_path)

            sheet_name = "BENEFICIAIRES"
            empexceldata = pd.read_excel(excel_file_full_path, sheet_name=sheet_name)

            errors = []
            session_adherent_principal_id = None

            print("db Frame")
            print(empexceldata)
            pprint("VERIFICATION FICHIER EXCEL")

            for index, row in empexceldata.iterrows():
                date_affiliation = timezone.make_aware(check_pandas_date_value(row["DATE_ENTREE"]).to_pydatetime(), timezone.get_current_timezone()) if check_pandas_date_value(
                    row["DATE_ENTREE"]) else None
                civilite = Civilite.objects.filter(code=row['CIVILITE']).first()
                nom = check_pandas_value(row['NOM'])
                prenoms = check_pandas_value(row['PRENOMS'])
                genre = check_pandas_value(row['SEXE'])
                statut_familiale = check_pandas_value(row['SITUATION_MATRIMONIALE'])
                date_naissance = check_pandas_date_value(row['DATE_NAISSANCE'])
                date_naissance = timezone.make_aware(date_naissance.to_pydatetime(), timezone.get_current_timezone()) if date_naissance else None
                pays_naissance = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_NAISSANCE'])).first()
                qualite_beneficiaire = QualiteBeneficiaire.objects.filter(code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()
                numero_famille_excel = check_pandas_value(row['NO_FAMILLE'])
                formule = FormuleGarantie.objects.filter(code=check_pandas_value(row['CODE_FORMULE']), police=police).first()
                ville = check_pandas_value(row['VILLE'])
                telephone_mobile = check_pandas_value(row['TEL_MOBILE'])
                aliments = Aliment.objects.filter(nom=nom, prenoms=prenoms, date_naissance=date_naissance, genre=genre, qualite_beneficiaire=qualite_beneficiaire, bureau=request.user.bureau).first()

                if date_affiliation is None:
                    errors.append({"row": index + 2, "collunmName": "DATE_ENTREE", "message": "La date d'entrée est obligatoire"})
                if civilite is None:
                    errors.append({"row": index + 2, "collunmName": "CIVILITE", "message": "La civilité est obligatoire"})
                if nom is None:
                    errors.append({"row": index + 2, "collunmName": "NOM", "message": "Le nom est obligatoire"})
                if prenoms is None:
                    errors.append({"row": index + 2, "collunmName": "PRENOMS", "message": "Le prénom est obligatoire"})
                if genre is None:
                    errors.append({"row": index + 2, "collunmName": "SEXE", "message": "Le genre est obligatoire"})
                if statut_familiale is None:
                    errors.append({"row": index + 2, "collunmName": "SITUATION_MATRIMONIALE", "message": "La situation matrimoniale est obligatoire"})
                if date_naissance is None:
                    errors.append({"row": index + 2, "collunmName": "DATE_NAISSANCE", "message": "La date de naissance est obligatoire"})
                if pays_naissance is None:
                    errors.append({"row": index + 2, "collunmName": "CODE_PAYS_NAISSANCE", "message": "Le pays de naissance est obligatoire"})
                if qualite_beneficiaire is None:
                    errors.append({"row": index + 2, "collunmName": "CODE_LIEN_PARENTE", "message": "Le lien de parenté est obligatoire"})
                if numero_famille_excel is None:
                    errors.append({"row": index + 2, "collunmName": "NO_FAMILLE", "message": "Le numéro de famille est obligatoire"})
                if check_pandas_value(row['CODE_FORMULE']) is None:
                    errors.append({"row": index + 2, "collunmName": "CODE_FORMULE", "message": "La formule est obligatoire"})
                if formule is None:
                    errors.append({"row": index + 2, "collunmName": "CODE_FORMULE", "message": "Veillez rensigner une formule valide."})
                if ville is None:
                    errors.append({"row": index + 2, "collunmName": "VILLE", "message": "La ville est obligatoire"})
                if telephone_mobile is None:
                    errors.append({"row": index + 2, "collunmName": "TEL_MOBILE", "message": "Le téléphone mobile est obligatoire"})
                if aliments is not None:
                    errors.append({"row": index + 2, "collunmName": "NOM, PRENOMS, SEXE, DATE_NAISSANCE et CODE_LIEN_PARENTE", "message": "Un bénéficiaire existe déjà avec le même nom, prénom, date de naissance, sexe et lien de parenté."})

                # Verification par ligne si il y a des erreurs
                if len(errors) > 0:
                    break

            # Si il y a des erreurs
            if len(errors) > 0:
                response = {
                    'statut': 0,
                    'errors': errors
                }
                return JsonResponse(response, status=200)

            pprint("FIN VERIFICATION FICHIER EXCEL")


            response = {
                'statut': 1,
                'message': "Vérification terminée avec succès",
                #   'data': "serializers.serialize('json', liste_aliments)"
                'data': {}
            }

            return JsonResponse(response, status=200)
        except Exception as e:
            pprint("ERREUR DANS LE FICHIER EXCEL")
            pprint(e)
            response = {
                'statut': 0,
                'errors_details': str(e),
                'errors': [
                    {"row": "##", "collunmName": "##", "message": "Erreur lors de la lecture du fichier excel"}]
            }
            return JsonResponse(response, status=200)


@transaction.atomic
def import_beneficiaires(request, police_id):

    if request.method == 'POST' and request.FILES['fichier']:

        police = Police.objects.get(id=police_id)
        print("Police")
        pprint(police)

        try:
            fichier = request.FILES['fichier']
        except MultiValueDictKeyError:
            fichier = False

        try:
            fs = FileSystemStorage()
            file_name_renamed = fichier.name.replace(" ", "_")

            filename = fs.save(file_name_renamed, fichier)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            excel_file_full_path = fs.path(filename)

            pprint("excel_file_full_path")
            pprint(excel_file_full_path)

            # Nom de la fuille a parcourir
            sheet_name = "BENEFICIAIRES"

            # Nombre de ligne à sauter avant
            # de commencer la lecture du tableau
            # skip_rows = 1 # CAS DE LA FEUILLE EXEMPLE

            empexceldata = pd.read_excel(excel_file_full_path, sheet_name=sheet_name)
            adherent_principal_frame_data = empexceldata[empexceldata['CODE_LIEN_PARENTE'] == 'AD']
            beneficiaires_frame_data = empexceldata[empexceldata['CODE_LIEN_PARENTE'] != 'AD']


            liste_adherent_principal = []

            errors = []

            #créer une variable qui va sauvegarder les adh_principaux pendant la boucle
            session_adherent_principal_id = None

            print("db Frame")
            print(empexceldata)
#            pprint("VERIFICATION FICHIER EXCEL")
#            for index, row in empexceldata.iterrows():
#                date_affiliation = timezone.make_aware(check_pandas_value(row["DATE_ENTREE"]).to_pydatetime(),
#                                                       timezone.get_current_timezone()) if check_pandas_value(
#                    row["DATE_ENTREE"]) else None
#                civilite = Civilite.objects.filter(code=row['CIVILITE']).first()
#                nom = check_pandas_value(row['NOM'])
#                prenoms = check_pandas_value(row['PRENOMS'])
#                # nom_jeune_fille = check_pandas_value(row['NOM_JEUNE_FILLE'])
#                genre = check_pandas_value(row['SEXE'])
#                statut_familiale = check_pandas_value(row['SITUATION_MATRIMONIALE'])
#                # numero_piece = check_pandas_value(row["NO_CNI"])
#                date_naissance = check_pandas_value(row['DATE_NAISSANCE'])
#                date_naissance = timezone.make_aware(date_naissance.to_pydatetime(),
#                                                     timezone.get_current_timezone()) if date_naissance else None
#                # lieu_naissance = check_pandas_value(row['LIEU_NAISSANCE'])

#                pays_naissance = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_NAISSANCE'])).first()
#                # pays_residence = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_RESIDENCE'])).first()
#                # pays_activite_professionnelle = Pays.objects.filter(
#                #     code=check_pandas_value(row['CODE_PAYS_ACTIVITE_PROFESSIONNELLE'])).first()
#                # profession = Profession.objects.filter(name=check_pandas_value(row['CODE_PROFESSION'])).first()
#                qualite_beneficiaire = QualiteBeneficiaire.objects.filter(
#                    code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()

#                numero_famille_excel = check_pandas_value(row['NO_FAMILLE'])

#                formule = FormuleGarantie.objects.filter(code=check_pandas_value(row['CODE_FORMULE']),
#                                                         police=police).first()
                # surprime_ht = check_pandas_value(row['SURPRIME_HT'])
                # surprime_ttc = check_pandas_value(row['SURPRIME_TTC'])
                # plafond_extra = check_pandas_value(row['PLAFOND_EXTRA'])
                # matricule_employe = check_pandas_value(row['MATRICULE_EMPLOYE'])
                # matricule_cie = check_pandas_value(row['MATRICULE_COMPAGNIE'])
                # numero_securite_sociale = check_pandas_value(row['NO_SECURITE_SOCIALE'])
#                ville = check_pandas_value(row['VILLE'])
                # adresse = check_pandas_value(row['ADRESSE'])
                # code_postal = check_pandas_value(row['CODE_POSTAL'])
                # rib = check_pandas_value(row['RIB'])
                # telephone_fixe = check_pandas_value(row['TEL_FIXE'])
#                telephone_mobile = check_pandas_value(row['TEL_MOBILE'])
                # email = check_pandas_value(row['EMAIL'])
                # commentaire = check_pandas_value(row['COMMENTAIRE'])

                # MISE EN PLACE MESSAGE D'ERREUR
#                if date_affiliation is None:
#                    errors.append({"row": index + 2, "collunmName": "DATE_ENTREE", "message": "La date d'entrée est obligatoire"})
#                if civilite is None:
#                    errors.append({"row": index + 2, "collunmName": "CIVILITE", "message": "La civilité est obligatoire"})
#                if nom is None:
#                    errors.append({"row": index + 2, "collunmName": "NOM", "message": "Le nom est obligatoire"})
#                if prenoms is None:
#                    errors.append({"row": index + 2, "collunmName": "PRENOMS", "message": "Le prénom est obligatoire"})
#                if genre is None:
#                    errors.append({"row": index + 2, "collunmName": "SEXE", "message": "Le genre est obligatoire"})
#                if statut_familiale is None:
#                    errors.append({"row": index + 2, "collunmName": "SITUATION_MATRIMONIALE", "message": "La situation matrimoniale est obligatoire"})
#                if date_naissance is None:
#                    errors.append({"row": index + 2, "collunmName": "DATE_NAISSANCE", "message": "La date de naissance est obligatoire"})
#                if pays_naissance is None:
#                    errors.append({"row": index + 2, "collunmName": "CODE_PAYS_NAISSANCE", "message": "Le pays de naissance est obligatoire"})
#                if qualite_beneficiaire is None:
#                    errors.append({"row": index + 2, "collunmName": "CODE_LIEN_PARENTE", "message": "Le lien de parenté est obligatoire"})
#                if numero_famille_excel is None:
#                    errors.append({"row": index + 2, "collunmName": "NO_FAMILLE", "message": "Le numéro de famille est obligatoire"})
#                if check_pandas_value(row['CODE_FORMULE']) is None:
#                    errors.append({"row": index + 2, "collunmName": "CODE_FORMULE", "message": "La formule est obligatoire"})
#                if formule is None:
#                    errors.append({"row": index + 2, "collunmName": "CODE_FORMULE", "message": "Veillez rensigner une formule valide."})
#                if ville is None:
#                    errors.append({"row": index + 2, "collunmName": "VILLE", "message": "La ville est obligatoire"})
#                if telephone_mobile is None:
#                    errors.append({"row": index + 2, "collunmName": "TEL_MOBILE", "message": "Le téléphone mobile est obligatoire"})

                # Verification par ligne si il y a des erreurs
#                if len(errors) > 0:
#                    break

#            # Si il y a des erreurs
#            if len(errors) > 0:
#                response = {
#                    'statut': 0,
#                    'errors': errors
#                }
#                return JsonResponse(response, status=200)

#            pprint("FIN VERIFICATION FICHIER EXCEL")

            pprint("IMPORT DES ADHÉRENT PRINCIPAUX")
            # Parcourir les lignes et insérer adhérents principaux

            for index, row in adherent_principal_frame_data.iterrows():

                pprint(index)


                if check_pandas_value(row["CODE_LIEN_PARENTE"]) == "AD":

                    qualite_beneficiaire = QualiteBeneficiaire.objects.filter(
                        code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()

                    pprint("qualite_beneficiaire")
                    pprint(qualite_beneficiaire)

                    # numero = None  # check_pandas_value(row['No de matricule'])

                    date_affiliation = timezone.make_aware(check_pandas_date_value(row["DATE_ENTREE"]).to_pydatetime(), timezone.get_current_timezone()) if check_pandas_date_value(row["DATE_ENTREE"]) else None
                    pprint(date_affiliation)
                    civilite = Civilite.objects.filter(code=row['CIVILITE']).first()
                    nom = check_pandas_value(row['NOM'])
                    prenoms = check_pandas_value(row['PRENOMS'])
                    nom_jeune_fille = check_pandas_value(row['NOM_JEUNE_FILLE'])
                    genre = check_pandas_value(row['SEXE'])
                    statut_familiale = check_pandas_value(row['SITUATION_MATRIMONIALE'])
                    numero_piece = check_pandas_value(row["NO_CNI"])
                    date_naissance = check_pandas_date_value(row['DATE_NAISSANCE'])
                    pprint("date_naissance")
                    pprint(date_naissance)
                    date_naissance = timezone.make_aware(date_naissance.to_pydatetime(), timezone.get_current_timezone()) if date_naissance else None
                    pprint("date_naissance")
                    pprint(date_naissance)
                    lieu_naissance = check_pandas_value(row['LIEU_NAISSANCE'])

                    pays_naissance = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_NAISSANCE'])).first()
                    pays_residence = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_RESIDENCE'])).first()
                    pays_activite_professionnelle = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_ACTIVITE_PROFESSIONNELLE'])).first()
                    profession = Profession.objects.filter(name=check_pandas_value(row['CODE_PROFESSION'])).first()
                    # qualite_beneficiaire = QualiteBeneficiaire.objects.filter(code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()

                    numero_famille_excel = check_pandas_value(row['NO_FAMILLE'])

                    # Auto generate
                    numero_famille = generate_numero_famille()

                    formule = FormuleGarantie.objects.filter(code=check_pandas_value(row['CODE_FORMULE']), police=police).first()

                    surprime_ht = check_pandas_value(row['SURPRIME_HT'])
                    surprime_ttc = check_pandas_value(row['SURPRIME_TTC'])
                    plafond_extra = check_pandas_value(row['PLAFOND_EXTRA'])
                    matricule_employe = check_pandas_value(row['MATRICULE_EMPLOYE'])
                    matricule_cie = check_pandas_value(row['MATRICULE_COMPAGNIE'])
                    numero_securite_sociale = check_pandas_value(row['NO_SECURITE_SOCIALE'])
                    ville = check_pandas_value(row['VILLE'])
                    adresse = check_pandas_value(row['ADRESSE'])
                    code_postal = check_pandas_value(row['CODE_POSTAL'])
                    rib = check_pandas_value(row['RIB'])
                    telephone_fixe = check_pandas_value(row['TEL_FIXE'])
                    telephone_mobile = check_pandas_value(row['TEL_MOBILE'])
                    email = check_pandas_value(row['EMAIL'])
                    commentaire = check_pandas_value(row['COMMENTAIRE'])


                    #la formule est obligatoire
                    if formule:

                        aliment = Aliment.objects.create(bureau=request.user.bureau,
                                                         civilite=civilite,
                                                         # numero=numero,
                                                         nom=nom,
                                                         prenoms=prenoms,
                                                         nom_jeune_fille=nom_jeune_fille,
                                                         date_naissance=date_naissance if date_naissance else None,
                                                         #  date_naissance= None if date_naissance is None else (datetime.datetime.strptime(date_naissance, '%A %d %B') if isinstance(date_naissance, str) else date_naissance),
                                                         lieu_naissance=lieu_naissance,
                                                         genre=genre,
                                                         telephone_mobile=telephone_mobile,
                                                         telephone_fixe=telephone_fixe,
                                                         email=email,
                                                         matricule_employe=matricule_employe,
                                                         numero_famille=numero_famille,
                                                         numero_securite_sociale=numero_securite_sociale,
                                                         date_affiliation=date_affiliation if date_affiliation else None,
                                                         photo="",
                                                         statut_familiale=statut_familiale,
                                                         qualite_beneficiaire=qualite_beneficiaire,
                                                         numero_piece=numero_piece,
                                                         profession=profession,
                                                         code_postal=code_postal,
                                                         adresse=adresse,
                                                         ville=ville,
                                                         pays_naissance=pays_naissance,
                                                         pays_residence=pays_residence,
                                                         pays_activite_professionnelle=pays_activite_professionnelle,
                                                         surprime_ht=surprime_ht,
                                                         surprime_ttc=surprime_ttc,
                                                         plafond_extra=plafond_extra,
                                                         matricule_cie=matricule_cie,
                                                         commentaire=commentaire,
                                                         rib=rib,
                                                         statut=Statut.ACTIF,
                                                         )

                        aliment.save()
                        aliment.adherent_principal_id = aliment.pk # comme c'est un adhérent
                        aliment.numero_ordre = 1
                        aliment.numero_famille_du_mois = generer_nombre_famille_du_mois()
                        aliment.save()

                        #Insérer la ligne dans AlimentTemporaire pour gérer l'import de ses bénéficiaires
                        liste_adherent_principal.append({"aliment" : aliment, "numero_famille_excel" : numero_famille_excel})
                        # session_import = str(request.user.id) + str(uuid.uuid4())
                        # aliment_temporaire = AlimentTemporaire.objects.create(aliment_id=aliment.pk,
                        #                                                       session_import=session_import,
                        #                                                       created_by=request.user,
                        #                                                       numero_famille_import=numero_famille
                        #                                                       )


                        #Insérer une ligne dans aliment_formule
                        aliment_formule = AlimentFormule.objects.create(formule=formule, date_debut=date_affiliation, aliment=aliment, statut=Statut.ACTIF).save()
                        pprint(aliment_formule)

                        # Créer le mouvement d'incorporation du bénéficiaire
                        mouvement = Mouvement.objects.filter(code="INCORPORATION").first()
                        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                                            aliment=aliment,
                                                                            mouvement=mouvement,
                                                                            police=police,
                                                                            date_effet=aliment.date_affiliation,
                                                                            motif="Nouvelle incorporation")
                        mouvement_aliment.save()


                        #TODO générer sa carte
                        # aliment = Aliment.objects.get(id=aliment.pk)

                        suffixe = 'A'

                        # désactiver ses cartes actives
                        # Carte.objects.filter(aliment_id=aliment.pk).filter(statut=Statut.ACTIF).update(statut=Statut.INACTIF, date_desactivation=datetime.datetime.now(tz=timezone.utc))

                        # enregistrer la nouvelle carte
                        carte = Carte.objects.create(aliment_id=aliment.pk,
                                                     date_edition=datetime.datetime.now(tz=timezone.utc),
                                                     motif_edition="Nouvelle incorporation",
                                                     statut=Statut.ACTIF
                                                     )

                        # carte = Carte.objects.get(id=carte.pk)

                        # METTRE A JOUR LE NUMERO
                        # prefixe = request.user.bureau.code
                        numero_carte = generate_numero_carte(aliment)
                        carte.numero = numero_carte

                        # générer le qrcode
                        qrcode_file = generer_qrcode_carte(numero_carte)
                        print("qrcode_img")
                        # print(qrcode_img)
                        carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
                        carte.save()

            pprint("FIN DE L'IMPORT DES ADHÉRENT PRINCIPAUX")


            pprint("IMPORT DES AYANTS-DROITS")
            # Parcourir les lignes et insérer les ayants-droits

            for index, row in beneficiaires_frame_data.iterrows():

                pprint(index)



                if check_pandas_value(row["CODE_LIEN_PARENTE"]) != "AD":

                    qualite_beneficiaire = QualiteBeneficiaire.objects.filter(
                        code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()
                    # numero = None  # check_pandas_value(row['No de matricule'])

                    date_affiliation = timezone.make_aware(check_pandas_date_value(row["DATE_ENTREE"]).to_pydatetime(),
                                                           timezone.get_current_timezone()) if check_pandas_date_value(
                        row["DATE_ENTREE"]) else None
                    pprint(date_affiliation)
                    civilite = Civilite.objects.filter(code=row['CIVILITE']).first()
                    nom = check_pandas_value(row['NOM'])
                    prenoms = check_pandas_value(row['PRENOMS'])
                    nom_jeune_fille = check_pandas_value(row['NOM_JEUNE_FILLE'])
                    genre = check_pandas_value(row['SEXE'])
                    statut_familiale = check_pandas_value(row['SITUATION_MATRIMONIALE'])
                    numero_piece = check_pandas_value(row["NO_CNI"])
                    date_naissance = check_pandas_date_value(row['DATE_NAISSANCE'])
                    pprint("date_naissance")
                    pprint(date_naissance)
                    date_naissance = timezone.make_aware(date_naissance.to_pydatetime(),
                                                         timezone.get_current_timezone()) if date_naissance else None
                    pprint("date_naissance")
                    pprint(date_naissance)
                    lieu_naissance = check_pandas_value(row['LIEU_NAISSANCE'])

                    pays_naissance = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_NAISSANCE'])).first()
                    pays_residence = Pays.objects.filter(code=check_pandas_value(row['CODE_PAYS_RESIDENCE'])).first()
                    pays_activite_professionnelle = Pays.objects.filter(
                        code=check_pandas_value(row['CODE_PAYS_ACTIVITE_PROFESSIONNELLE'])).first()
                    profession = Profession.objects.filter(name=check_pandas_value(row['CODE_PROFESSION'])).first()
                    # qualite_beneficiaire = QualiteBeneficiaire.objects.filter(code=check_pandas_value(row["CODE_LIEN_PARENTE"])).first()

                    numero_famille_excel = check_pandas_value(row['NO_FAMILLE'])


                    formule = FormuleGarantie.objects.filter(code=check_pandas_value(row['CODE_FORMULE']),
                                                             police=police).first()

                    surprime_ht = check_pandas_value(row['SURPRIME_HT'])
                    surprime_ttc = check_pandas_value(row['SURPRIME_TTC'])
                    plafond_extra = check_pandas_value(row['PLAFOND_EXTRA'])
                    matricule_employe = check_pandas_value(row['MATRICULE_EMPLOYE'])
                    matricule_cie = check_pandas_value(row['MATRICULE_COMPAGNIE'])
                    numero_securite_sociale = check_pandas_value(row['NO_SECURITE_SOCIALE'])
                    ville = check_pandas_value(row['VILLE'])
                    adresse = check_pandas_value(row['ADRESSE'])
                    code_postal = check_pandas_value(row['CODE_POSTAL'])
                    rib = check_pandas_value(row['RIB'])
                    telephone_fixe = check_pandas_value(row['TEL_FIXE'])
                    telephone_mobile = check_pandas_value(row['TEL_MOBILE'])
                    email = check_pandas_value(row['EMAIL'])
                    commentaire = check_pandas_value(row['COMMENTAIRE'])

                    # la formule est obligatoire
                    if formule:

                        #récupérer l'id de son adhérent principal
                        adherent_principal = None

                        ad_list = [ ad["aliment"] for ad in liste_adherent_principal if ad["numero_famille_excel"] == numero_famille_excel ]
                        if len(ad_list) > 0:
                            adherent_principal = ad_list[0]

                        if adherent_principal:
                            aliment = Aliment.objects.create(bureau=request.user.bureau,
                                                             adherent_principal=adherent_principal,
                                                            civilite=civilite,
                                                             # numero=numero,
                                                             nom=nom,
                                                             prenoms=prenoms,
                                                             nom_jeune_fille=nom_jeune_fille,
                                                             date_naissance=date_naissance if date_naissance else None,
                                                             #  date_naissance= None if date_naissance is None else (datetime.datetime.strptime(date_naissance, '%A %d %B') if isinstance(date_naissance, str) else date_naissance),
                                                             lieu_naissance=lieu_naissance,
                                                             genre=genre,
                                                             telephone_mobile=telephone_mobile,
                                                             telephone_fixe=telephone_fixe,
                                                             email=email,
                                                             matricule_employe=matricule_employe,
                                                             numero_famille=adherent_principal.numero_famille,
                                                             numero_securite_sociale=numero_securite_sociale,
                                                             date_affiliation=date_affiliation if date_affiliation else None,
                                                             photo="",
                                                             statut_familiale=statut_familiale,
                                                             qualite_beneficiaire=qualite_beneficiaire,
                                                             numero_piece=numero_piece,
                                                             profession=profession,
                                                             code_postal=code_postal,
                                                             adresse=adresse,
                                                             ville=ville,
                                                             pays_naissance=pays_naissance,
                                                             pays_residence=pays_residence,
                                                             pays_activite_professionnelle=pays_activite_professionnelle,
                                                             surprime_ht=surprime_ht,
                                                             surprime_ttc=surprime_ttc,
                                                             plafond_extra=plafond_extra,
                                                             matricule_cie=matricule_cie,
                                                             commentaire=commentaire,
                                                             rib=rib,
                                                             statut=Statut.ACTIF,
                                                             )

                            aliment.save()

                            # liste_aliments.append(aliment)

                            # créer le numéro d'ordre dans la famille
                            aliment.numero_ordre = generer_numero_ordre(aliment)
                            aliment.save()

                            # Insérer une ligne dans aliment_formule
                            aliment_formule = AlimentFormule.objects.create(formule=formule,
                                                                            date_debut=date_affiliation,
                                                                            aliment=aliment, statut=Statut.ACTIF).save()
                            pprint(aliment_formule)

                            # Créer le mouvement d'incorporatio du bénéficiaire
                            mouvement = Mouvement.objects.filter(code="INCORPORATION").first()
                            mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                                                aliment=aliment,
                                                                                mouvement=mouvement,
                                                                                police=police,
                                                                                date_effet=aliment.date_affiliation,
                                                                                motif="Nouvelle incorporation")
                            mouvement_aliment.save()


                            # TODO générer sa carte
                            # aliment = Aliment.objects.get(id=aliment.pk)

                            # suffixe = 'A'

                            # désactiver ses cartes actives
                            # Carte.objects.filter(aliment_id=aliment.pk).filter(statut=Statut.ACTIF).update(
                            #     statut=Statut.INACTIF, date_desactivation=datetime.datetime.now(tz=timezone.utc))

                            # enregistrer la nouvelle carte
                            carte = Carte.objects.create(aliment_id=aliment.pk,
                                                         date_edition=datetime.datetime.now(tz=timezone.utc),
                                                         motif_edition="Nouvelle incorporation",
                                                         statut=Statut.ACTIF
                                                         )

                            # carte = Carte.objects.get(id=carte.pk)

                            # METTRE A JOUR LE NUMERO
                            # prefixe = request.user.bureau.code
                            numero_carte = generate_numero_carte(aliment)
                            carte.numero = numero_carte

                            # générer le qrcode
                            qrcode_file = generer_qrcode_carte(numero_carte)
                            print("qrcode_img")
                            # print(qrcode_img)
                            carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
                            carte.save()

            pprint("FIN DE L'IMPORT DES AYANTS-DROITS")


            response = {
                'statut': 1,
                'message': "Importation effectuée avec succès",
                'data': "serializers.serialize('json', liste_aliments)"
            }

            return JsonResponse(response, status=200)
        except Exception as e:
            pprint("ERREUR DANS LE FICHIER EXCEL")
            pprint(e)
            response = {
                'statut': 0,
                'errors_details': str(e),
                'errors': [
                    {"row": "##", "collunmName": "##", "message": "Erreur d'importation"}]
            }
            return JsonResponse(response, status=200)


def serialize_sets(obj):
    if isinstance(obj, set):
        return list(obj)

    return obj

# json_str = json.dumps(set([1,2,3]), default=serialize_sets)

# ajout de bénéficiaire
@transaction.atomic  # open a transaction
def add_beneficiaire(request, police_id):
    police = Police.objects.get(id=police_id)
    file_upload_path = ''

    if request.method == 'POST':

        # upload de la photo
        if 'photo' in request.FILES:

            try:

                photo = request.FILES['photo']

                fs = FileSystemStorage()

                filebase, extension = photo.name.rsplit('.', 1)
                file_name = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                file_upload_path = 'aliments/police_%s/%s.%s' % (police.id, file_name, extension)

                fs.save(file_upload_path, photo)

            except MultiValueDictKeyError:
                file_upload_path = ''

        formule_id = request.POST.get('formule_id')
        adherent_principal_id = request.POST.get('adherent_principal')
        qualite_beneficiaire_id = request.POST.get('qualite_beneficiaire_id')
        qualite_beneficiaire = QualiteBeneficiaire.objects.get(id=qualite_beneficiaire_id)

        aliment = Aliment.objects.create(bureau=request.user.bureau,
                                         adherent_principal_id=adherent_principal_id,
                                         pays_naissance_id=request.POST.get('pays_naissance'),
                                         pays_residence_id=request.POST.get('pays_residence'),
                                         pays_activite_professionnelle_id=request.POST.get('pays_activite_professionnelle'),
                                         #profession_id=request.POST.get('profession'),
                                         profession_libelle=request.POST.get('profession'),
                                         plafond_extra=request.POST.get('plafond_extra').replace(' ','') if request.POST.get('plafond_extra').replace(' ', '') != '' else None,
                                         rib=request.POST.get('rib'),
                                         surprime_ht=request.POST.get('surprime_ht').replace(' ',
                                                                                             '') if request.POST.get(
                                             'surprime_ht').replace(' ', '') != '' else None,
                                         surprime_ttc=request.POST.get('surprime_ttc').replace(' ',
                                                                                               '') if request.POST.get(
                                             'surprime_ttc').replace(' ', '') != '' else None,
                                         code_postal=request.POST.get('code_postal'),
                                         ville=request.POST.get('ville'),
                                         adresse=request.POST.get('adresse'),
                                         commentaire=request.POST.get('commentaire'),
                                         civilite_id=request.POST.get('civilite'),
                                         lieu_naissance = request.POST.get('lieu_naissance'),
                                         numero_securite_sociale = request.POST.get('numero_securite_sociale'),
                                         nom=request.POST.get('nom'),
                                         prenoms=request.POST.get('prenoms'),
                                         nom_jeune_fille=request.POST.get('nom_jeune_fille'),
                                         date_naissance=request.POST.get('date_naissance'),
                                         genre=request.POST.get('genre'),
                                         telephone_fixe=request.POST.get('telephone_fixe'),
                                         telephone_mobile= request.POST.get('indicatif') + request.POST.get('telephone_mobile'),
                                         email=request.POST.get('email'),
                                         matricule_employe=request.POST.get('matricule_employe'),
                                         date_affiliation=request.POST.get('date_entree'),
                                         photo=file_upload_path,
                                         statut_familiale=request.POST.get('statut_familiale'),
                                         qualite_beneficiaire_id=qualite_beneficiaire_id,
                                         numero_piece=request.POST.get('numero_piece'),
                                         apci_ald = request.POST.get('apci_ald'),
                                         statut=Statut.ACTIF,
                                         )

        aliment.save()

        # génération des numéros
        aliment = Aliment.objects.get(id=aliment.pk)
        aliment.numero = 'A' + str(Date.today().year) + str(aliment.pk).zfill(6)
        if qualite_beneficiaire.code == "AD":
            aliment.adherent_principal = aliment
            aliment.numero_ordre = 1
            #générer un numéro de famille
            aliment.numero_famille = generate_numero_famille()
            aliment.numero_famille_du_mois = generer_nombre_famille_du_mois()

        aliment.save()

        if qualite_beneficiaire.code != "AD":
            #générer le numéro d'ordre dans la famille
            aliment.numero_ordre = generer_numero_ordre(aliment)
            aliment.save()


        # Renseigner la table association qui lie l'aliment à la police et à la formule
        aliment_formule = AlimentFormule.objects.create(formule_id=formule_id, aliment_id=aliment.pk,
                                                        date_debut=aliment.date_affiliation, statut=Statut.ACTIF, created_by=request.user)


        # Créer le mouvement d'incorporatio du bénéficiaire
        mouvement = Mouvement.objects.filter(code="INCORPORATION").first()

        # Créer le mouvement_aliment
        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                            aliment=aliment,
                                                            mouvement=mouvement,
                                                            police=police,
                                                            date_effet=aliment.date_affiliation,
                                                            motif="Nouvelle incorporation")
        mouvement_aliment.save()


        #TODO générer une carte en même temps

        aliment = Aliment.objects.get(id=aliment.pk)

        suffixe = 'A'

        # désactiver ses cartes actives
        Carte.objects.filter(aliment_id=aliment.pk).filter(statut=Statut.ACTIF).update(statut=Statut.INACTIF,
                                                                                       date_desactivation=datetime.datetime.now(tz=timezone.utc))

        # enregistrer la nouvelle carte
        carte = Carte.objects.create(aliment_id=aliment.pk,
                                     date_edition=datetime.datetime.now(tz=timezone.utc),
                                     motif_edition="Nouvelle incorporation",
                                     statut=Statut.ACTIF
                                     )

        carte = Carte.objects.get(id=carte.pk)

        # METTRE A JOUR LE NUMERO
        prefixe = request.user.bureau.code
        numero_carte = generate_numero_carte(aliment)
        carte.numero = numero_carte
        carte.save()

        # générer le qrcode
        qrcode_file = generer_qrcode_carte(numero_carte)
        print("qrcode_img")
        # print(qrcode_img)
        carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
        carte.save()




        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': aliment.pk,
                'numero': aliment.numero,
                'nom': aliment.nom,
                'prenoms': aliment.prenoms,
                'date_naissance': aliment.date_naissance,
                'genre': aliment.genre,
                'carte': "",
                'qualite_beneficiaire': "",
                'matricule_employe': aliment.matricule_employe,
                'matricule_cie': aliment.matricule_cie,
                'numero_famille': aliment.adherent_principal.numero_famille if aliment.adherent_principal else "",
                'statut': aliment.statut,
                'date_sortie': aliment.date_sortie,
                'formule': aliment_formule.formule.libelle
            }
        }

        return JsonResponse(response)
    else:

        police = Police.objects.get(id=police_id)
        #liste_des_formules = FormuleGarantie.objects.filter(police_id=police_id)
        polices_du_bureau_actif = Police.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE)

        # aliments = getAliments(police_id)

        adherents_principaux = getAdherentsPrincipaux(police_id)

        qualites_beneficiaires = QualiteBeneficiaire.objects.all().order_by('libelle')
        civilites = Civilite.objects.all().exclude(code="STE").order_by('name')

        #formules = police.formules.order_by('libelle')

        pays = Pays.objects.all().order_by('nom')
        professions = Profession.objects.all().order_by('name')

        # etat police = dernier motif
        #etat_police = MouvementPolice.objects.filter(police_id=police_id).order_by('-id')[:1].get().motif.libelle

        #option_export_beneficiaires = police.bureau.option_export_beneficiaires

        today = datetime.datetime.now(tz=timezone.utc).date()

        return render(request, 'police/modal_beneficiaire.html', {'police': police, 'adherents_principaux': adherents_principaux,
                         'qualites_beneficiaires': qualites_beneficiaires, 'civilites': civilites,
                         'pays': pays, 'professions': professions,
                        'polices_du_bureau_actif': polices_du_bureau_actif,'today': today})


def getVehicules(police_id):
    print(police_id)
    vehicules = []
    vehicule_police = VehiculePolice.objects.filter(police_id=police_id, statut=Statut.ACTIF).order_by('-id')

    for vp in vehicule_police:
        if vp.vehicule.statut == Statut.ACTIF:
            vehicules.append(vp.vehicule)

    return vehicules


@never_cache
def police_vehicules(request, police_id):
    police = Police.objects.get(id=police_id)

    vehicules = getVehicules(police_id)

    energies = Energie.choices

    types_carroserie = TypeCarosserie.objects.all().order_by('libelle')
    categories_vehicule = CategorieVehicule.objects.all().order_by('libelle')

    marques = MarqueVehicule.objects.all().order_by('libelle')

    pprint(vehicules)

    return render(request, 'police/vehicules.html',
                  {'police': police, 'vehicules': vehicules, 'energies': energies, 'marques': marques,
                   'types_carroserie': types_carroserie, 'categories_vehicule': categories_vehicule})


# ajout de véhicule
def add_vehicule(request, police_id):
    police = Police.objects.get(id=police_id)

    if request.method == 'POST':
        # x = 3
        formule_id = request.POST.get('formule_id')

        vehicule = Vehicule.objects.create(
            numero_immatriculation=request.POST.get('numero_immatriculation'),
            numero_immat_provisoire=request.POST.get('numero_immat_provisoire'),
            numero_serie=request.POST.get('numero_serie'),
            modele=request.POST.get('modele'),
            conducteur=request.POST.get('conducteur'),
            place=request.POST.get('place'),
            energie=request.POST.get('energie'),
            valeur_neuve=request.POST.get('valeur_neuve'),
            valeur_actuelle=request.POST.get('valeur_actuelle'),
            puissance=request.POST.get('puissance'),
            poids_a_vide=request.POST.get('poids_a_vide'),
            poids_a_charge=request.POST.get('poids_a_charge'),
            date_mis_en_circulation=request.POST.get('date_mis_en_circulation'),
            categorie_vehicule_id=request.POST.get('categorie_vehicule_id'),
            marque_id=request.POST.get('marque_id'),
            type_carosserie_id=request.POST.get('type_carosserie_id'),
        )
        vehicule.save()

        vehicule_police = VehiculePolice.objects.create(
            motif="vehicule",
            date_mouvement=datetime.now(),
            statut=Statut.ACTIF,
            formule_id=formule_id,
            police_id=police_id,
            vehicule_id=vehicule.id,
        )

        vehicule_police.save()

        response = {
            'statut': 1,
            'message': "Véhicule ajouté avec succès !",
            'data': {
                'id': "",
            }
        }
    return JsonResponse(response)


def update_vehicule(request, police_id, vehicule_id):
    police = Police.objects.get(id=police_id)
    vehicule = Vehicule.objects.get(id=vehicule_id)

    energies = Energie.choices

    types_carroserie = TypeCarosserie.objects.all().order_by('libelle')
    categories_vehicule = CategorieVehicule.objects.all().order_by('libelle')

    marques = MarqueVehicule.objects.all().order_by('libelle')

    if request.method == 'POST':

        Vehicule.objects.filter(id=vehicule_id).update(
            numero_immatriculation=request.POST.get('numero_immatriculation'),
            numero_immat_provisoire=request.POST.get('numero_immat_provisoire'),
            numero_serie=request.POST.get('numero_serie'),
            modele=request.POST.get('modele'),
            conducteur=request.POST.get('conducteur'),
            place=request.POST.get('place'),
            energie=request.POST.get('energie'),
            valeur_neuve=request.POST.get('valeur_neuve'),
            valeur_actuelle=request.POST.get('valeur_actuelle'),
            puissance=request.POST.get('puissance'),
            poids_a_vide=request.POST.get('poids_a_vide'),
            poids_a_charge=request.POST.get('poids_a_charge'),
            date_mis_en_circulation=request.POST.get('date_mis_en_circulation'),
            categorie_vehicule_id=request.POST.get('categorie_vehicule_id'),
            marque_id=request.POST.get('marque_id'),
            type_carosserie_id=request.POST.get('type_carosserie_id'),
        )

        # aliment.save()

        vehicule = Vehicule.objects.get(id=vehicule_id)

        response = {
            'statut': 1,
            'message': "Modification effectuée avec succès !",
            'data': {
                'id': vehicule.pk,
                'numero_immat_provisoire': vehicule.pk,
                'numero_immatriculation': vehicule.pk,
                'numero_serie': vehicule.pk,
                'modele': vehicule.pk,
                'conducteur': vehicule.pk,
                'place': vehicule.place,
                'energie': vehicule.energie,
                'valeur_neuve': vehicule.valeur_neuve,
                'valeur_actuelle': vehicule.pk,
                'puissance': vehicule.pk,
                'poids_a_vide': vehicule.pk,
                'poids_a_charge': vehicule.pk,
                'date_mis_en_circulation': vehicule.pk,
                'categorie_vehicule_id': vehicule.categorie_vehicule_id,
                'marque_id': vehicule.pk,
                'type_carosserie_id': vehicule.pk,
                'date_mis_en_circulation': vehicule.pk,
                'vehicule.id': vehicule.id,
            }
        }

        return JsonResponse(response)

    else:

        return render(request, 'police/modal_vehicule_modification.html', {
            'vehicule': vehicule,
            'police': police,
            'energies': energies,
            'types_carroserie': types_carroserie,
            'categories_vehicule': categories_vehicule,
            'marques': marques,
        }
                      )


# Afficher les details d'un vehicule
def details_vehicule(request, police_id, vehicule_id):
    police = Police.objects.get(id=police_id)
    vehicule = Vehicule.objects.get(id=vehicule_id)
    energies = Energie.choices

    sinistres = []

    tarifs = []

    return render(
        request,
        'police/modal_details_vehicule.html',
        {
            'police': police,
            'vehicule': vehicule,
            'tarifs': tarifs,
            'sinistres': sinistres,
            'energies': energies
        }
    )


# Supprimer un vehicule mais c'est resté en cours
def supprimer_vehicule(request, police_id, vehicule_id):
    vehicule_police = VehiculePolice.objects.get(id=police_id)
    police = Police.objects.get(id=police_id)
    vehicule = Vehicule.objects.get(id=vehicule_id)

    if request.method == "POSeT":

        vehicule_id = request.POST.get('vehicule_id')

        vehicule = Vehicule.objects.get(id=vehicule_id)
        if vehicule.pk is not None:
            # vehicule.delete()

            Vehicule.objects.filter(id=vehicule_id).update(statut=Statut.INACTIF)

            response = {
                'statut': 1,
                'message': "Vehicule supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Vehicule non trouvé !",
            }

        return JsonResponse(response)


# Importation des vehicules
def import_vehicules(request, police_id):
    # categorieVehicule = CategorieVehicule
    # marqueVehicule = MarqueVehicule
    # typeCarosserie = TypeCarosserie

    formule_id = request.POST.get('formule_id')

    # print(police_id)
    try:
        if request.method == 'POST' and request.FILES['fichier']:

            try:
                fichier = request.FILES['fichier']
            except MultiValueDictKeyError:
                fichier = False

            fs = FileSystemStorage()
            file_name_renamed = fichier.name.replace(" ", "_")

            filename = fs.save(file_name_renamed, fichier)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url

            # Lire le fichier CSV
            empexceldata = pd.read_csv("." + excel_file, delimiter=";")

            liste_vehicules = []

            police = Police.objects.get(id=police_id)

            dbframe = empexceldata

            # print(dbframe)
            for dbframe in dbframe.itertuples():

                try:
                    numero_immatriculation = dbframe.Immatriculation
                except:
                    numero_immatriculation = None

                try:
                    formule = FormuleGarantie.objects.get(id=formule_id)
                except:
                    formule = None

                try:
                    numero_immat_provisoire = dbframe.numero_immat_provisoire
                except:
                    numero_immat_provisoire = None

                try:
                    numero_serie = dbframe.numero_serie.strip()
                except:
                    numero_serie = None

                try:
                    modele = dbframe.modele
                except:
                    modele = None

                try:
                    conducteur = dbframe.conducteur
                except:
                    conducteur = None

                try:
                    place = dbframe.place.strip()
                except:
                    place = None

                try:
                    energie = dbframe.energie
                except:
                    energie = None

                try:
                    valeur_neuve = dbframe.valeur_neuve.strip()
                except:
                    valeur_neuve = None

                try:
                    valeur_actuelle = dbframe.valeur_actuelle.strip()
                except:
                    valeur_actuelle = None

                try:
                    puissance = dbframe.puissance.strip()
                except:
                    puissance = None

                try:
                    poids_a_vide = dbframe.poids_a_vide.strip()
                except:
                    poids_a_vide = None

                try:
                    poids_a_charge = dbframe.poids_a_charge.strip()
                except:
                    poids_a_charge = None

                try:
                    date_mis_en_circulation = dbframe.date_mis_en_circulation
                except:
                    date_mis_en_circulation = None

                try:
                    categorieVehicule = CategorieVehicule.objects.get(id=4)
                except:
                    categorieVehicule = None

                # print(numero_immat_provisoire)
                # print(numero_serie)
                # print(modele)
                # print(conducteur)
                # print(place)
                # print(energie)
                # print(valeur_neuve)
                # print(valeur_actuelle)
                # print(puissance)
                # print(poids_a_vide)
                # print(poids_a_charge)
                # print(date_mis_en_circulation)

                try:
                    marqueVehicule = MarqueVehicule.objects.get(libelle=dbframe.marque_id)
                except:
                    marqueVehicule = None

                try:
                    typeCarosserie = TypeCarosserie.objects.get(code=dbframe.Carrosserie)
                except:
                    typeCarosserie = None

                vehicule = Vehicule.objects.create(
                    numero_immatriculation=numero_immatriculation,
                    numero_immat_provisoire=numero_immat_provisoire,
                    numero_serie=numero_serie,
                    modele=modele,
                    conducteur=conducteur,
                    place=place,
                    energie=energie,
                    valeur_neuve=valeur_neuve,
                    valeur_actuelle=valeur_actuelle,
                    puissance=puissance,
                    poids_a_vide=poids_a_vide,
                    poids_a_charge=poids_a_charge,
                    date_mis_en_circulation=date_mis_en_circulation,
                    categorie_vehicule=categorieVehicule,
                    marque=marqueVehicule,
                    type_carosserie=typeCarosserie,
                )

                vehicule.save()
                liste_vehicules.append(vehicule)

                vehicule_police = VehiculePolice.objects.create(
                    motif="Par Importation",
                    date_mouvement=datetime.datetime.now(),
                    statut=Statut.ACTIF,
                    formule=formule,
                    police_id=police_id,
                    vehicule=vehicule
                ).save()

            # pprint(liste_vehicules)

            response = {
                'statut': 1,
                'message': "Importation des véhicules effectuée avec succès",
                'data': "serializers.serialize('json', liste_aliments)"
            }

            return JsonResponse(response)

    except Exception as identifier:

        response = {
            'statut': 0,
            'message': "Erreur lors de l'importation " + identifier.__str__(),
            'errors': {'Erreur': identifier.__str__()}
        }

        return JsonResponse(response)


# modification d'un bénéficiaire
def update_beneficiaire(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)
    file_upload_path = ''

    if request.method == 'POST':

        if aliment:

            # upload de la photo
            if 'photo' in request.FILES:

                try:

                    photo = request.FILES['photo']

                    fs = FileSystemStorage()

                    filebase, extension = photo.name.rsplit('.', 1)
                    file_name = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
                    file_upload_path = 'aliments/police_%s/%s.%s' % (police.id, file_name, extension)

                    fs.save(file_upload_path, photo)
                    # aliment.photo = file_upload_path  # Mettez à jour le chemin de la photo dans l'objet aliment

                except MultiValueDictKeyError:
                    file_upload_path = ''

            # Si aucune nouvelle photo n'est téléchargée, utilisez la photo actuelle
            else:
                if 'current_photo' in request.POST:
                    file_upload_path = request.POST['current_photo']



            #Données à modifier
            qualite_beneficiaire_id = request.POST.get('qualite_beneficiaire_id')
            qualite_beneficiaire = QualiteBeneficiaire.objects.filter(id=qualite_beneficiaire_id).first()

            if qualite_beneficiaire and qualite_beneficiaire.code == "AD":
                #aliment.numero_famille = generate_numero_famille() #si
                pprint("DEJA AD")
            else:
                aliment.adherent_principal_id = request.POST.get('adherent_principal')

            aliment.pays_naissance_id = request.POST.get('pays_naissance')
            aliment.pays_residence_id = request.POST.get('pays_residence')
            aliment.pays_activite_professionnelle_id = request.POST.get('pays_activite_professionnelle')
            #aliment.profession_id = request.POST.get('profession')
            aliment.profession_libelle = request.POST.get('profession')
            aliment.plafond_extra = request.POST.get('plafond_extra').replace(' ', '') if request.POST.get('plafond_extra') else 0
            aliment.rib = request.POST.get('rib')
            aliment.surprime_ht = request.POST.get('surprime_ht').replace(' ', '') if request.POST.get('surprime_ht') else 0
            aliment.surprime_ttc = request.POST.get('surprime_ttc').replace(' ', '') if request.POST.get('surprime_ttc') else 0
            aliment.code_postal = request.POST.get('code_postal')
            aliment.ville = request.POST.get('ville')
            aliment.adresse = request.POST.get('adresse')
            aliment.commentaire = request.POST.get('commentaire')
            aliment.civilite_id = request.POST.get('civilite')
            aliment.numero = request.POST.get('numero')
            aliment.nom = request.POST.get('nom')
            aliment.prenoms = request.POST.get('prenoms')
            aliment.nom_jeune_fille = request.POST.get('nom_jeune_fille')
            if request.POST.get('date_naissance'): aliment.date_naissance = request.POST.get('date_naissance')
            aliment.genre = request.POST.get('genre')
            aliment.telephone_fixe = request.POST.get('telephone_fixe')
            aliment.telephone_mobile = request.POST.get('indicatif') + request.POST.get('telephone_mobile')
            aliment.email = request.POST.get('email')
            aliment.matricule_employe = request.POST.get('matricule_employe')
            aliment.date_affiliation = request.POST.get('date_entree')
            aliment.photo = file_upload_path
            aliment.statut_familiale = request.POST.get('statut_familiale')
            aliment.numero_securite_sociale = request.POST.get('numero_securite_sociale')
            aliment.lieu_naissance = request.POST.get('lieu_naissance')
            aliment.qualite_beneficiaire_id = request.POST.get('qualite_beneficiaire_id')
            aliment.numero_piece = request.POST.get('numero_piece')
            aliment.apci_ald = request.POST.get('apci_ald')
            aliment.statut = Statut.ACTIF

            #dd(aliment)
            aliment.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': aliment.pk,
                    'numero': aliment.numero,
                    'nom': aliment.nom,
                    'prenoms': aliment.prenoms,
                    'date_naissance': aliment.date_naissance,
                    'genre': aliment.genre,
                    'carte': "",
                    'qualite_beneficiaire': "",
                    'matricule_employe': aliment.matricule_employe,
                    'numero_famille': aliment.numero_famille,
                    'statut': aliment.statut,
                    'date_sortie': aliment.date_sortie
                }
            }

        else:
            response = {
                'statut': 0,
                'message': "Bénéficiaire non trouvé!",
                'data': {
                }
            }

        return JsonResponse(response)

    else:

        # récupérer les adhérents principaux
        adherents_principaux = getAdherentsPrincipaux(police_id)

        # adherents_principaux = police.formules.aliments.filter(qualite_beneficiaire_id=1).order_by('nom')
        qualites_beneficiaires = QualiteBeneficiaire.objects.all().order_by('libelle')
        civilites = Civilite.objects.all().order_by('name')
        # formules = Formule.objects.all().order_by('libelle')
        # les formules associées à la police
        formules = police.formules.order_by('libelle')
        pays = Pays.objects.all().order_by('nom')
        professions = Profession.objects.all().order_by('name')

        genre = Genre
        situation_familial = StatutFamilial

        return render(request, 'police/modal_beneficiaire_modification.html',
                      {'genre': genre, 'situation_familial': situation_familial, 'police': police, 'aliment': aliment,
                       'adherents_principaux': adherents_principaux, 'qualites_beneficiaires': qualites_beneficiaires,
                       'civilites': civilites, 'formules': formules, 'pays': pays,
                       'professions': professions})


# modification d'un bénéficiaire
def details_beneficiaire(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    date_du_jour_gmt_ = datetime.datetime.now(tz=datetime.timezone.utc).date()
    formulegarantie = aliment.formule_atdate(date_du_jour_gmt_)

    types_documents = TypeDocument.objects.all().order_by('libelle')
    documents = Document.objects.filter(aliment_id=aliment_id).order_by('-id')

    cachet_bureau = police.bureau.cachet.url if police.bureau.cachet else ''

    cartes = Carte.objects.filter(aliment_id=aliment_id).order_by('-id')
    if len(cartes) > 0:
        carte_active = Carte.objects.filter(aliment_id=aliment_id).order_by('-id')[:1].get()

        carte_qrcode = os.path.join(settings.MEDIA_URL, 'cartes/' + carte_active.numero + ".png")
        pprint(carte_qrcode)

    else:
        carte_active = None
        carte_qrcode = ""

    sinistres = Sinistre.objects.filter(aliment_id=aliment_id, statut_validite=StatutValidite.VALIDE)

    # exclure le concerné: membres_famille = Aliment.objects.filter(adherent_principal_id=aliment.adherent_principal.id).filter(~Q(id=aliment_id))
    membres_famille = Aliment.objects.filter(adherent_principal_id=aliment.adherent_principal_id)

    formules = FormuleGarantie.objects.filter(police_id=police_id).order_by('libelle')
    #aliments_polices_formules = AlimentFormule.objects.filter(aliment_id=aliment_id).order_by('-id')
    historique_formules = aliment.historique_formules.filter(formule__police=police).exclude(statut_validite=StatutValidite.SUPPRIME).order_by('date_debut')

    date_entree_police = historique_formules.first().date_debut if historique_formules else None
    #dd(date_entree_police)

    cant_change_his_formule = historique_formules.filter(date_fin__isnull=True).exists()

    liste_mouvements = Mouvement.objects.filter(type="BENEFICIAIRE").order_by('libelle')
    mouvements_aliments = MouvementAliment.objects.filter(aliment=aliment, statut_validite=StatutValidite.VALIDE).order_by('id')

    tarifs = []

    last_mouvement = aliment.last_mouvement if aliment.last_mouvement else None

    today = date_du_jour_gmt_
    etat_beneficiaire = aliment.etat_beneficiaire
    #if aliment.date_sortie:
    #    etat_beneficiaire = 'SORTI'

    # si la police est échue
    police_echue = False
    if police.is_echue:
        etat_beneficiaire = 'SORTI'
        police_echue = True
    #    dd(police.date_fin_effet)

    carte_path = os.path.join(settings.MEDIA_URL, 'cartes/')

    today = datetime.datetime.now(tz=timezone.utc)

    #
    pays = Pays.objects.all().order_by('nom')
    civilites = Civilite.objects.all().exclude(code="STE").order_by('name')

    if request.user.bureau.pays.poligamie:
        qualite_beneficiaires = ['CONJOINT', 'ENFANT', 'ASCENDANT']
    else:

        has_conjoint = membres_famille.filter(qualite_beneficiaire__code='CO', date_sortie__isnull=True)
        print("has_conjoint")
        print(has_conjoint)
        if has_conjoint.exists():
            qualite_beneficiaires = ['ENFANT', 'ASCENDANT'] #
        else:
            qualite_beneficiaires = ['CONJOINT', 'ENFANT', 'ASCENDANT'] #

    # total_conjoints = membres_famille.filter(qualite_beneficiaire__code='CO').count()
    # if total_conjoints >= 2:
    #     qualite_beneficiaires = ['ENFANT', 'ASCENDANT']
    # else:
    # qualite_beneficiaires = ['CONJOINT', 'ENFANT', 'ASCENDANT']

    family_members = Aliment.objects.filter(adherent_principal_id=aliment.adherent_principal_id) #
    adherent_principal_famille = family_members.filter(qualite_beneficiaire__code='AD').first() #

    return render(request, 'police/modal_details_beneficiaire.html',
                  {'police': police, 'aliment': aliment, 'membres_famille': membres_famille,
                   'aliments_polices_formules': historique_formules, 'last_mouvement': last_mouvement, 'etat_beneficiaire': etat_beneficiaire, 'liste_mouvements': liste_mouvements, 'mouvements_aliments': mouvements_aliments, 'formules': formules, 'tarifs': tarifs,
                   'cartes': cartes, 'carte_active': carte_active, 'carte_qrcode': carte_qrcode,
                   cachet_bureau: cachet_bureau,
                   'carte_path': carte_path, 'date_entree_police': date_entree_police, 'sinistres': sinistres,
                   'types_documents': types_documents, 'documents': documents,
                   'cant_change_his_formule': cant_change_his_formule, 'today': today, 'formulegarantie': formulegarantie,
                   'pays':pays, 'civilites': civilites, 'qualites_beneficiaires':qualite_beneficiaires, 'adherent_principal_famille': adherent_principal_famille, 'police_echue': police_echue})

# ajouter membre famille d'un beneficiaire
def police_add_membre_famille_beneficiaire(request, police_id, aliment_id):

    police = Police.objects.get(id=police_id)
    file_upload_path = ''

    if request.method == 'POST':

        aliment = Aliment.objects.get(pk=aliment_id)

        # upload de la photo
        if 'photo' in request.FILES:
            try:
                photo = request.FILES['photo']
                fs = FileSystemStorage()
                file_name_renamed = 'ph_' + photo.name.replace(" ", "_")
                file_upload_path = 'aliments/police_' + str(police.id) + '/' + file_name_renamed
                fs.save(file_upload_path, photo)
            except MultiValueDictKeyError:
                file_upload_path = ''

        formule_id = request.POST.get('formule_id')
        adherent_principal_id = request.POST.get('adherent_principal')

        qualite_beneficiaire_libelle = request.POST.get('qualite_beneficiaire_libelle')
        qualite_beneficiaire = QualiteBeneficiaire.objects.get(libelle=qualite_beneficiaire_libelle)
        qualite_beneficiaire_id = qualite_beneficiaire.id

        aliment = Aliment.objects.create(bureau=request.user.bureau,
                                         #
                                         adherent_principal_id=adherent_principal_id,
                                         #
                                         pays_naissance_id=request.POST.get('pays_naissance'),
                                         pays_residence_id=request.POST.get('pays_residence'),
                                         pays_activite_professionnelle_id=request.POST.get('pays_activite_professionnelle'),
                                         #  profession_id=request.POST.get('profession'),
                                         plafond_extra=request.POST.get('plafond_extra').replace(' ','') if request.POST.get('plafond_extra').replace(' ', '') != '' else None,
                                         rib=request.POST.get('rib'),
                                         surprime_ht=request.POST.get('surprime_ht').replace(' ','') if request.POST.get('surprime_ht').replace(' ', '') != '' else None,
                                         surprime_ttc=request.POST.get('surprime_ttc').replace(' ','') if request.POST.get('surprime_ttc').replace(' ', '') != '' else None,
                                         code_postal=request.POST.get('code_postal'),
                                         ville=request.POST.get('ville'),
                                         adresse=request.POST.get('adresse'),
                                         commentaire=request.POST.get('commentaire'),
                                         civilite_id=request.POST.get('civilite'),
                                         lieu_naissance = request.POST.get('lieu_naissance'),
                                         numero_securite_sociale = request.POST.get('numero_securite_sociale'),
                                         nom=request.POST.get('nom'),
                                         prenoms=request.POST.get('prenoms'),
                                         nom_jeune_fille=request.POST.get('nom_jeune_fille'),
                                         date_naissance=request.POST.get('date_naissance'),
                                         genre=request.POST.get('genre'),
                                         telephone_fixe=request.POST.get('telephone_fixe'),
                                         telephone_mobile= request.POST.get('indicatif') + request.POST.get('telephone_mobile'),
                                         email=request.POST.get('email'),
                                         matricule_cie=request.POST.get('matricule_cie'),
                                         date_affiliation=request.POST.get('date_entree'),
                                         photo=file_upload_path,
                                         statut_familiale=request.POST.get('statut_familiale'),
                                         #
                                         qualite_beneficiaire_id=qualite_beneficiaire_id,
                                         numero_piece=request.POST.get('numero_piece'),
                                         apci_ald = request.POST.get('apci_ald'),
                                         statut=Statut.ACTIF,
                                         )

        aliment.save()

        aliment = Aliment.objects.get(id=aliment.pk)
        aliment.numero = 'A' + str(Date.today().year) + str(aliment.pk).zfill(6)
        if qualite_beneficiaire.code == "AD":
            aliment.adherent_principal = aliment
            aliment.numero_ordre = 1
            aliment.numero_famille = generate_numero_famille()
            aliment.numero_famille_du_mois = generer_nombre_famille_du_mois()

        aliment.save()

        if qualite_beneficiaire.code != "AD":
            aliment.numero_ordre = generer_numero_ordre(aliment)
            aliment.save()

        aliment_formule = AlimentFormule.objects.create(formule_id=formule_id, aliment_id=aliment.pk,
                                                        date_debut=aliment.date_affiliation, statut=Statut.ACTIF, created_by=request.user)

        mouvement = Mouvement.objects.filter(code="INCORPORATION").first()

        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                            aliment=aliment,
                                                            mouvement=mouvement,
                                                            police=police,
                                                            date_effet=aliment.date_affiliation,
                                                            motif="Nouvelle incorporation")
        mouvement_aliment.save()

        aliment = Aliment.objects.get(id=aliment.pk)
        suffixe = 'A'
        Carte.objects.filter(aliment_id=aliment.pk).filter(statut=Statut.ACTIF).update(statut=Statut.INACTIF,
                                                                                       date_desactivation=datetime.datetime.now(tz=timezone.utc))
        carte = Carte.objects.create(aliment_id=aliment.pk,
                                     date_edition=datetime.datetime.now(tz=timezone.utc),
                                     motif_edition="Nouvelle incorporation",
                                     statut=Statut.ACTIF
                                     )
        carte = Carte.objects.get(id=carte.pk)

        prefixe = request.user.bureau.code
        numero_carte = generate_numero_carte(aliment)
        carte.numero = numero_carte
        carte.save()

        qrcode_file = generer_qrcode_carte(numero_carte)
        print("qrcode_img")
        carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
        carte.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': aliment.pk,
                'numero': aliment.numero,
                'nom': aliment.nom,
                'prenoms': aliment.prenoms,
                'date_naissance': aliment.date_naissance,
                'genre': aliment.genre,
                'carte': "",
                'qualite_beneficiaire': "",
                'matricule_employe': aliment.matricule_employe,
                'matricule_cie': aliment.matricule_cie,
                'numero_famille': aliment.adherent_principal.numero_famille if aliment.adherent_principal else "",
                'statut': aliment.statut,
                'date_sortie': aliment.date_sortie,
                'formule': aliment_formule.formule.libelle
            }
        }

        return JsonResponse(response)
#

#
def add_carte(request, aliment_id):
    aliment = Aliment.objects.get(id=aliment_id)


    if request.method == 'POST':
        date_edition = request.POST.get('date_edition')
        motif = request.POST.get('motif')

    # désactiver ses cartes actives
    Carte.objects.filter(aliment_id=aliment_id).filter(statut=Statut.ACTIF).update(statut=Statut.INACTIF,
                                                                                   date_desactivation=datetime.datetime.now(
                                                                                       tz=timezone.utc))

    # enregistrer la nouvelle carte
    carte = Carte.objects.create(aliment_id=aliment_id,
                                 date_edition=date_edition,
                                 motif_edition=motif,
                                 statut=Statut.ACTIF,
                                 bureau=aliment.bureau
                                 )

    carte = Carte.objects.get(id=carte.pk)

    # METTRE A JOUR LE NUMERO
    numero_carte = generate_numero_carte(aliment)
    carte.numero = numero_carte
    carte.save()

    # générer le qrcode
    qrcode_file = generer_qrcode_carte(numero_carte)
    print("qrcode_img")
    # print(qrcode_img)
    carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
    carte.save()

    response = {
        'statut': 1,
        'message': "Enregistrement effectuée avec succès !",
        'data': {
            'id': carte.pk,
            'numero': carte.numero,
            'date_edition': carte.date_edition,
            'date_desactivation': '',
            'motif_edition': carte.motif_edition,
            'statut': carte.statut,
        }
    }

    return JsonResponse(response)


#
def change_formule(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    if request.method == 'POST':
        date_mouvement = request.POST.get('date_mouvement')
        motif = request.POST.get('motif')
        formule_id = request.POST.get('formule')

        # desactiver les autres lignes
        # AlimentFormule.objects.filter(aliment_id=aliment_id).update(statut=Statut.INACTIF)

        #recuperons la toute derniere formule l'aliment
        old_aliment_police_formule = AlimentFormule.objects.filter(aliment=aliment, date_fin__isnull=True).order_by("-pk").first()

        # enregistrer
        aliment_police_formule = AlimentFormule.objects.create(aliment=aliment,
                                                               formule_id=formule_id,
                                                               date_debut=date_mouvement,
                                                               # motif=motif,
                                                               statut=Statut.ACTIF,
                                                               created_by=request.user
                                                               )
        aliment_police_formule.save()

        #mettons fin à l'ancienne formule de l'aliment
        if old_aliment_police_formule:
            old_aliment_police_formule.date_fin = datetime.datetime.strptime(date_mouvement, '%Y-%m-%d').date() - timedelta(days=1)
            old_aliment_police_formule.save()

        aliment_police_formule = AlimentFormule.objects.get(id=aliment_police_formule.pk)


        mouvement = Mouvement.objects.filter(code="CHANGEFORMULE-BENEF").first()

        #Créer l'avenant
        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                            aliment=aliment,
                                                            mouvement=mouvement,
                                                            police=police,
                                                            date_effet=date_mouvement,
                                                            motif=motif)
        mouvement_aliment.save()

        formatted_date = f"{mouvement_aliment.date_effet[8:10]}/{mouvement_aliment.date_effet[5:7]}/{mouvement_aliment.date_effet[:4]}"

        response = {
            'statut': 1,
            'message': "Opération effectuée avec succès !",
            'data': {
                'formule_id': aliment_police_formule.pk,
                'formule_code_formule': aliment_police_formule.formule.code,
                'formule_libelle_formule': aliment_police_formule.formule.libelle,
                'formule_date_debut': "aliment_police_formule.date_debut",
                'formule_motif': aliment_police_formule.motif,
                'statut': aliment_police_formule.statut,
                'id': mouvement_aliment.pk,
                'libelle': mouvement_aliment.mouvement.libelle,
                'date_effet': formatted_date,
                'motif': mouvement_aliment.motif,
                'created_by': f"{mouvement_aliment.created_by.last_name} {mouvement_aliment.created_by.first_name}",
                'old_id': old_aliment_police_formule.pk,
                'old_date_fin': old_aliment_police_formule.date_fin.strftime("%d/%m/%Y") if old_aliment_police_formule.date_fin else ""
            }
        }

    return JsonResponse(response)


#
def suspension_beneficiaire(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    # noinspection LanguageDetectionInspection
    if request.method == 'POST':

        date_effet = request.POST.get('date_effet')
        motif = request.POST.get('motif')

        mouvement = Mouvement.objects.filter(code="SUSPENSION-BENEF").first()

        #Créer l'avenant
        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                            aliment=aliment,
                                                            mouvement=mouvement,
                                                            police=police,
                                                            date_effet=date_effet,
                                                            motif=motif)
        mouvement_aliment.save()
        pprint(mouvement_aliment.date_effet)

        formatted_date ="" # f"{mouvement_aliment.date_effet[8:10]}/{mouvement_aliment.date_effet[5:7]}/{mouvement_aliment.date_effet[:4]}"

        response = {
            'statut': 1,
            'message': "Opération effectuée avec succès !",
            'data': {
                'id': mouvement_aliment.pk,
                'libelle': mouvement_aliment.mouvement.libelle,
                'date_effet': formatted_date,
                'motif': mouvement_aliment.motif,
                'created_by': f"{mouvement_aliment.created_by.last_name} {mouvement_aliment.created_by.first_name}"
            }
        }

    return JsonResponse(response)


#
def remise_en_vigueur(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    # noinspection LanguageDetectionInspection
    if request.method == 'POST':
        id = request.POST.get('id')
        date_effet = request.POST.get('date_effet')
        motif = request.POST.get('motif')

        mouvement = Mouvement.objects.filter(code="REMISEVIGUEUR-BENEF").first()

        #Créer l'avenant
        mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                            aliment=aliment,
                                                            mouvement=mouvement,
                                                            police=police,
                                                            date_effet=date_effet,
                                                            motif=motif)
        mouvement_aliment.save()

        formatted_date = f"{mouvement_aliment.date_effet[8:10]}/{mouvement_aliment.date_effet[5:7]}/{mouvement_aliment.date_effet[:4]}"

        response = {
            'statut': 1,
            'message': "Opération effectuée avec succès !",
            'data': {
                'id': mouvement_aliment.pk,
                'libelle': mouvement_aliment.mouvement.libelle,
                'date_effet': formatted_date,
                'motif': mouvement_aliment.motif,
                'created_by': f"{mouvement_aliment.created_by.last_name} {mouvement_aliment.created_by.first_name}"
            }
        }

    return JsonResponse(response)


#
def sortie_police(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    if request.method == 'POST':

        date_sortie = request.POST.get('date_sortie')
        motif = request.POST.get('motif')
        type_retrait = motif = request.POST.get('type_retrait', None)

        if type_retrait == "famille":
            aliments = Aliment.objects.filter(adherent_principal_id=aliment_id)
        else:
            aliments = Aliment.objects.filter(id=aliment_id) # aliement lui-même


        for aliment in aliments:

            # Comparaison de dates
            last_sinistre = aliment.last_sinistre

            last_sinistre_date_survenance = last_sinistre.date_survenance if last_sinistre else None

            if last_sinistre_date_survenance and date_sortie < str(last_sinistre_date_survenance):

                response = {
                    'statut': 0,
                    'message': f"Le dernier sinistre (N°{last_sinistre.numero}) du bénéficiaire est survenu le {last_sinistre.date_survenance}",
                    'data': {
                        'id_sinistre': last_sinistre.pk,
                        'numero_sinistre': last_sinistre.numero,
                        'date_survenance': last_sinistre.date_survenance,
                        'prestataire': last_sinistre.prestataire.name
                    }
                }

            else:

                # Mettre fin à son ancienne ligne d'aliment formule
                last_aliment_formule = AlimentFormule.objects.filter(aliment_id=aliment_id, date_fin__isnull=True).last()
                pprint("last_aliment_formule")
                pprint(last_aliment_formule)

                if last_aliment_formule:
                    last_aliment_formule.date_fin=date_sortie
                    last_aliment_formule.motif=motif
                    last_aliment_formule.save()


                mouvement = Mouvement.objects.filter(code="SORTIE-BENEF").first()

                #Créer l'avenant
                mouvement_aliment = MouvementAliment.objects.create(created_by=request.user,
                                                                    aliment=aliment,
                                                                    mouvement=mouvement,
                                                                    police=police,
                                                                    date_effet=date_sortie,
                                                                    motif=motif)
                mouvement_aliment.save()

                #update de la date de sortie sur l'aliment
                aliment.date_sortie = date_sortie
                aliment.updated_at = datetime.datetime.now(tz=timezone.utc)
                aliment.save()

                #enregistrer dans les logs
                ActionLog.objects.create(done_by=request.user, action="sortie_beneficiaire",
                                        description="Sortie du bénéficiaires " + str(aliment.nom) + " " + str(aliment.prenoms) + " (" + str(aliment.carte_active()) + ") de la police " + str(police.numero),
                                        table="",
                                        row=None, data_before=None,
                                        data_after=None)

                formatted_date = f"{mouvement_aliment.date_effet[8:10]}/{mouvement_aliment.date_effet[5:7]}/{mouvement_aliment.date_effet[:4]}"


        # Pour la reponse json l'assure lui même ou adh
        mouvement_aliment = MouvementAliment.objects.filter(
            created_by=request.user,
            aliment=aliment,
            mouvement=mouvement,
            police=police,
            date_effet=date_sortie,
            motif=motif
        ).first()

        response = {
            'statut': 1,
            'message': "Opération effectuée avec succès !",
            'data': {
                'id': mouvement_aliment.pk,
                'libelle': mouvement_aliment.mouvement.libelle,
                'date_effet': formatted_date,
                'motif': mouvement_aliment.motif,
                'created_by': f"{mouvement_aliment.created_by.last_name} {mouvement_aliment.created_by.first_name}"
            }
        }

    return JsonResponse(response)



# get all beneficiaires for police
def prime_famille(request, police_id):
    police = Police.objects.get(id=police_id)

    aliments_principaux = getAdherentsPrincipaux(police_id)

    for aliment in aliments_principaux:
        aliment.nombre_enfants = Aliment.objects.filter(adherent_principal_id=aliment.id,
                                                        qualite_beneficiaire_id=3).count()
        aliment.prime_conjoint = \
            Aliment.objects.filter(adherent_principal_id=aliment.id, qualite_beneficiaire_id=2).aggregate(
                all_sum=Sum('surprime_ttc'))['all_sum']
        aliment.prime_enfants = \
            Aliment.objects.filter(adherent_principal_id=aliment.id, qualite_beneficiaire_id=3).aggregate(
                all_sum=Sum('surprime_ttc'))['all_sum']
        aliment.prime_totale = \
            Aliment.objects.filter(adherent_principal_id=aliment.id).aggregate(all_sum=Sum('surprime_ttc'))['all_sum']

    # etat police = dernier motif
    etat_police = police.etat_police

    return render(request, 'police/prime_famille.html',
                  {'police': police, 'aliments_principaux': aliments_principaux, 'etat_police': etat_police, })


# ajout d'avenant
def add_avenant(request, police_id):
    police = Police.objects.get(id=police_id)

    if request.method == 'POST':

        if request.POST.get('mouvement') in ["5", "16"]:
            request.session['add_avenant'] = request.POST
            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {}
            }

        else:
            date_fin_periode_garantie = request.POST.get('date_fin_periode_garantie')

            mouvement_police = MouvementPolice.objects.create(police_id=police_id,
                                                              mouvement_id=request.POST.get('mouvement'),
                                                              motif_id=request.POST.get('motif'),
                                                              date_effet=request.POST.get('date_effet'),
                                                              date_fin_periode_garantie=date_fin_periode_garantie if date_fin_periode_garantie else None,
                                                              created_by=request.user
                                                              )

            mouvement_police.save()

            mouvement = Mouvement.objects.get(id=mouvement_police.mouvement_id)

            motif = Motif.objects.get(id=mouvement_police.motif_id)

            #si c'est un renouvellement, créer une période de couverture
            if mouvement.code == "AVENANT":

                # créer une ligne dans période de couverture
                periode_couverture = PeriodeCouverture.objects.create(
                    police_id=police.id,
                    date_debut_effet=mouvement_police.date_effet,
                    date_fin_effet=mouvement_police.date_fin_periode_garantie,
                ).save()

                #Mettre à jour la date début_effet et fin_effet de la police
                # police.date_debut_effet = mouvement_police.date_effet
                # police.date_fin_effet = mouvement_police.date_fin_periode_garantie
                # police.save()


            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {
                    'id': mouvement_police.pk,
                    'mouvement': mouvement.libelle,
                    'motif': motif.libelle,
                    'date_effet': mouvement_police.date_effet,
                    'date_fin_periode_garantie': mouvement_police.date_fin_periode_garantie,
                }
            }

        return JsonResponse(response)


# permet de définir des tarifs entre le client et les prestataires

class PoliceTarifsSpecifiquesView(TemplateView):
    template_name = 'police/tarifs_specifiques.html'
    model = TarifPrestataireClient

    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police = Police.objects.get(id=police_id)
        formules = FormuleGarantie.objects.filter(police_id=police_id).order_by('libelle')
        prestataires = Prestataire.objects.filter(bureau=request.user.bureau, status=True).order_by('name')

        tarifs_specifiques = TarifPrestataireClient.objects.filter(client_id=police.client.pk, statut=1)

        context_perso = {'police': police, 'tarifs_specifiques': tarifs_specifiques, 'formules': formules,
                         'prestataires': prestataires}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def add_tarif_specifique(request, police_id):
    if request.method == 'POST':
        pprint(request.POST)
        fichier = request.FILES['fichier_tarif']
        prestataire_id = request.POST.get('prestataire_id')
        formule_id = request.POST.get('formule_police_id')

        police = Police.objects.get(id=police_id)

        fs = FileSystemStorage()
        file_name_renamed = fichier.name.replace(" ", "_")

        filename = fs.save(file_name_renamed, fichier)

        prestataire = Prestataire.objects.filter(id=prestataire_id).first()
        formule = FormuleGarantie.objects.filter(id=formule_id).first()

        TarifPrestataireClient.objects.create(
            prestataire_id=prestataire.pk,
            client_id=police.client.pk,
            formule_id=formule.pk,
            fichier_tarification=filename
        )

        response = {
            'statut': 1,
            'message': "Enregistrement effectué avec succès !",
            'data': {
            }
        }

        return JsonResponse(response)


def del_tarif_specifique(request):
    if request.method == 'POST':
        tarif_id = request.POST.get('tarif_id')

        TarifPrestataireClient.objects.filter(id=tarif_id).update(statut=0)

        response = {
            'statut': 1,
            'message': "Suppression effectué avec succès !",
            'data': {
            }
        }

        return JsonResponse(response)

class PhotosBeneficiairesView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'police/benefs_pictures.html'
    model = Aliment
    def get(self, request, police_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        # TODO police_beneficiaires_datatable
        police = Police.objects.get(id=police_id)

        #Pour éviter les doublons
        # Récupérer les IDs des aliments dans aliment_formule
        aliment_ids = AlimentFormule.objects.filter(
            formule_id__in=[p.id for p in police.formules],
            statut=Statut.ACTIF,
            statut_validite=StatutValidite.VALIDE
        ).values_list('aliment_id', flat=True)

        today = datetime.datetime.now(tz=timezone.utc).date()

        beneficiaires = Aliment.objects.filter(id__in=aliment_ids, photo="").order_by('-adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

        # filter(photo="").

        # etat police = dernier motif
        etat_police = police.etat_police

        today = datetime.datetime.now(tz=timezone.utc).date()

        context_perso = {'police': police, 'beneficiaires': beneficiaires, 'etat_police': etat_police, 'today': today}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def upload_photo(request, beneficiaire_id):
    if request.method == 'POST':
        beneficiaire = get_object_or_404(Aliment, pk=beneficiaire_id)
        form = PhotoUploadForm(request.POST, request.FILES)

        if form.is_valid():
            beneficiaire.photo = form.cleaned_data['file']
            beneficiaire.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'error': 'Une erreur est survenue lors de la sauvegarde de la photo.'})
    return JsonResponse({'success': False})

# def benefs_pictures(request, police_id):
#     police = Police.objects.get(id=police_id)

#     aliments_police = AlimentFormule.objects.filter(police_id=police_id).all()

#     # filter(photo="").

#     # etat police = dernier motif
#     etat_police = police.etat_police

#     return render(request, 'police/benefs_pictures.html',
#                   {'police': police, 'aliments_police': aliments_police, 'etat_police': etat_police, })


def upload_benef_picture(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    aliment = Aliment.objects.get(id=aliment_id)

    if request.method == 'POST':

        fichier = request.FILES['file']
        extesion = os.path.splitext(str(fichier))[1]

        # filename = "ph_" + str(uuid.uuid4()) + "_" + fichier.name
        filename = str(aliment.id) + extesion

        # uploader le fichier
        uploaded_photo_name = handle_uploaded_photo(fichier, filename, police_id)

        # aliment.photo = filename
        aliment.photo = uploaded_photo_name
        aliment.save()

        response = {
            'statut': 1,
            'message': "Photo mis à jour avec succès !",
            'data': {

            }
        }

        return JsonResponse(response)

    else:
        response = {
            'statut': 0,
            'message': "HTTP GET method not supported",
            'data': {}
        }

        return JsonResponse(response)


def motifs_by_mouvement(request, mouvement_id):
    motifs = Motif.objects.filter(mouvement_id=mouvement_id)
    motifs_serialize = serializers.serialize('json', motifs)
    return HttpResponse(motifs_serialize, content_type='application/json')


# upload du fichier
def handle_uploaded_photo(f, filename, police_id):
    path_ot_db = '/aliments/police_' + str(police_id)
    dirname = settings.MEDIA_URL.replace('/', '') + path_ot_db
    path = os.path.join(dirname)

    if not os.path.exists(path):
        os.makedirs(path)

    with open(dirname + '/' + filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

    return path_ot_db + '/' + filename


# upload fichier tarification
def handle_uploaded_fichier(f, filename):
    path_ot_db = '/tarifs/'
    dirname = settings.MEDIA_URL.replace('/', '') + path_ot_db
    path = os.path.join(dirname)

    if not os.path.exists(path):
        os.makedirs(path)

    with open(dirname + '/' + filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

    return path_ot_db + '/' + filename


# Générer le code pour le client
def generate_client_code():
    current_year = str(date.today().year)[-2:]

    # Trouver le dernier code créé dans la base de données
    last_code = Client.objects.aggregate(Max('code'))['code__max']

    # Extraire le numéro incrémental du dernier code
    if last_code:
        last_number = int(last_code.split('-')[0])  # Ex: "0001-CL24" -> 0001
        new_number = last_number + 1
    else:
        new_number = 1  # Si aucun code n'existe encore

    # Formatage du nouveau numéro pour garder 4 chiffres
    new_code = f"{str(new_number).zfill(4)}-CL{current_year}"

    return new_code

class ClientsView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/clients.html'
    model = Client

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        types_clients = TypeClient.objects.all().order_by('libelle')
        types_personnes = TypePersonne.objects.all().order_by('libelle')
        civilites = Civilite.objects.all().order_by('name')
        bureaux = Bureau.objects.all().order_by('nom')
        pays = Pays.objects.all().order_by('nom')
        business_units = BusinessUnit.objects.all().order_by('libelle')
        utilisateurs = User.objects.filter(bureau=request.user.bureau, type_utilisateur__code="INTERNE", is_active=True).order_by('last_name')
        secteurs_activite = SecteurActivite.objects.filter(status=True).order_by('libelle')

        context_perso = {'types_clients': types_clients, 'types_personnes': types_personnes, 'secteurs_activite': secteurs_activite,
                         'civilites': civilites, 'bureaux': bureaux, 'pays': pays, 'business_units': business_units,
                         'utilisateurs': utilisateurs}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


def clients_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')
    search_nom = request.GET.get('search_nom', '').strip()
    search_numero_police = request.GET.get('search_numero_police', '').strip()
    search_type_personne = request.GET.get('search_type_personne', '').strip()

    aliment_trouve = False
    id_aliment_trouve = 0

    user = request.user
    queryset = Client.objects.filter(statut=Statut.ACTIF, bureau_id=user.bureau_id)

    if search_nom:
        queryset = queryset.filter(
            Q(nom__icontains=search_nom) | Q(prenoms__icontains=search_nom)
        )

    #un client a plusieurs polices
    if search_numero_police:
        queryset = queryset.filter(
            Q(polices__numero__icontains=search_numero_police)
        ).distinct()

        pprint(search_numero_police)


    if search_type_personne:
        queryset = queryset.filter(
            Q(type_personne_id=search_type_personne)
        )


    # Apply sorting
    queryset = queryset.order_by('-code')

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data in the expected format
    data = []
    for c in page_obj:

        detail_url = reverse('client_details', args=[c.id])  # URL to the detail view
        modifier_client_url = reverse('modifier_client', args=[c.id])  # URL to the detail view
        actions_html = f'<a href="{detail_url}" class="text-center"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> {_("Détails")}</span></a>&nbsp;&nbsp;' \
                       f'<span style="cursor:pointer;" class="btn_modifier_client badge btn-sm btn-modifier rounded-pill text-center" data-client_id="{c.id}" data-model_name="client" data-modal_title="MODIFICATION D\'UN CLIENT" data-href="{modifier_client_url}"><i class="fas fa-edit"></i> {_("Modifier")}</span></a>&nbsp;&nbsp;'

        liste_numeros_polices = ''
        for p in c.polices.filter(statut_validite=StatutValidite.VALIDE):
            detail_police_url = reverse('police.details', args=[p.id])  #
            liste_numeros_polices += f'<a target="_blank" href="{detail_police_url}"><span class="bold">{p.numero}</span></a>, '

        info_beneficiaire_html = ''
        if id_aliment_trouve:
            detail_infos_beneficiaire_url = reverse('auto_open_beneficiaire', args=[p.id, id_aliment_trouve])  #
            info_beneficiaire_html = f'<a target="_blank" href="{detail_infos_beneficiaire_url}"><span class="badge btn-sm">{nom_aliment_trouve}</span></a>'

        if not c.nom: c.nom = ''
        if not c.prenoms: c.prenoms = ''

        data.append({
            "id": c.id,
            "nom": c.nom + ' ' + c.prenoms,
            "numero_police": liste_numeros_polices[:-2],
            "code": c.code,
            "info_beneficiaire": info_beneficiaire_html,
            "type_personne": c.type_personne.libelle if c.type_personne else "",
            "type_client": c.type_client.libelle if c.type_client else "",
            "business_unit": c.business_unit.libelle if c.business_unit else "",
            "telephone_mobile": c.telephone_mobile,
            "statut": c.statut,
            "actions": actions_html,
        })

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# ajout d'avenant
@login_required
def add_client(request):
    from datetime import datetime
    if request.method == 'POST':

        date_naissance = request.POST.get('date_naissance', None)
        if date_naissance:
            date_naissance = datetime.strptime(date_naissance, '%Y-%m-%d').date()
        else:
            date_naissance = None

        client_created = Client.objects.create(bureau_id=request.user.bureau.id,
                                       nom=request.POST.get('nom'),
                                       prenoms=request.POST.get('prenoms'),
                                       secteur_activite_id=request.POST.get('secteur_activite_id'),
                                       type_client_id=request.POST.get('type_client_id'),
                                       business_unit_id=request.POST.get('business_unit_id'),
                                       date_naissance=date_naissance,
                                       telephone_mobile=request.POST.get('telephone_mobile'),
                                       telephone_fixe=request.POST.get('telephone_fixe'),
                                       email=request.POST.get('email'),
                                       ville=request.POST.get('ville'),
                                       adresse_postale=request.POST.get('adresse_postale'),
                                       adresse=request.POST.get('adresse'),
                                       gestionnaire_id=request.user.id,
                                       site_web=request.POST.get('site_web'),
                                       twitter=request.POST.get('twitter'),
                                       instagram=request.POST.get('instagram'),
                                       facebook=request.POST.get('facebook'),
                                       civilite_id=request.POST.get('civilite_id'),
                                       sexe=request.POST.get('sexe'),
                                       created_by_id=request.user.id,
                                       #created_at=datetime.datetime.now(tz=timezone.utc),
                                       updated_at = timezone.now(),
                                       pays_id=request.POST.get('pays_id'),
                                       type_personne_id=request.POST.get('type_personne_id'),
                                       )

        #TODO : nomenclature du code client a trouver
        code_bureau = request.user.bureau.code
        client_created.code_provisoire = str(code_bureau) + str(Date.today().year)[-2:] + '-' + str(client_created.pk).zfill(7) + '-CL'
        client_created.code = generate_client_code()
        client_created.save()

        # Handle logo_client upload
        logo_client_file = request.FILES.get('logo_client')
        if logo_client_file:
            client_created.logo.save(logo_client_file.name, logo_client_file)
            client_created.save()

        try:
            creation_veos = send_client_to_veos(client_created)
        except:
            pass


        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': client_created.pk,
                'nom': client_created.nom,
                'prenoms': client_created.prenoms,
                'date_naissance': client_created.date_naissance,
                #'type_client': client_created.type_client.libelle if client_created.type_client else "",
                'type_personne': client_created.type_personne.libelle if client_created.type_personne else "",
                'ville': client_created.ville,
                'created_at': client_created.created_at,
                'statut': client_created.statut,
            }
        }

        return JsonResponse(response)


# modification d'un bénéficiaire
@login_required
def modifier_client(request, client_id):
    from datetime import datetime
    client = Client.objects.get(id=client_id)
    file_upload_path = ''

    if request.method == 'POST':
        user = User.objects.get(id=request.user.id)

        date_naissance = request.POST.get('date_naissance', None)
        if date_naissance:
            date_naissance = datetime.strptime(date_naissance, '%Y-%m-%d').date()
        else:
            date_naissance = None

        Client.objects.filter(id=client_id).update(nom=request.POST.get('nom'),
                                                   prenoms=request.POST.get('prenoms'),
                                                   secteur_activite_id=request.POST.get('secteur_activite_id'),
                                                   type_client_id=request.POST.get('type_client_id'),
                                                   business_unit_id=request.POST.get('business_unit_id'),
                                                   date_naissance=date_naissance,
                                                   telephone_mobile=request.POST.get('telephone_mobile'),
                                                   telephone_fixe=request.POST.get('telephone_fixe'),
                                                   email=request.POST.get('email'),
                                                   ville=request.POST.get('ville'),
                                                   adresse_postale=request.POST.get('adresse_postale'),
                                                   adresse=request.POST.get('adresse'),
                                                   site_web=request.POST.get('site_web'),
                                                   twitter=request.POST.get('twitter'),
                                                   instagram=request.POST.get('instagram'),
                                                   facebook=request.POST.get('facebook'),
                                                   ancienne_ref=request.POST.get('ancienne_ref'),
                                                   civilite_id=request.POST.get('civilite_id'),
                                                   sexe=request.POST.get('sexe'),
                                                   #updated_at=datetime.datetime.now(tz=timezone.utc),
                                                   updated_at = timezone.now(),
                                                   pays_id=request.POST.get('pays_id'),
                                                   type_personne_id=request.POST.get('type_personne_id'),
                                                   )
        print(request.POST)
        # Handle logo_client upload
        logo_client_file = request.FILES.get('logo')
        if logo_client_file:
            client.logo.save(logo_client_file.name, logo_client_file)
            client.save()
            pprint("PHOTO SAUVEGARDÉ")


        response = {
            'statut': 1,
            'message': "Modification effectuée avec succès !",
            'data': {
                'id': client.pk,
                'nom': client.nom,
                'prenoms': client.prenoms,
                'date_naissance': client.date_naissance,
                'type_client': client.type_client.libelle if client.type_client else "",
                'type_personne': client.type_personne.libelle if client.type_personne else "",
                'ville': client.ville,
                'created_at': client.created_at,
                'statut': client.statut,
            }
        }

        return JsonResponse(response)

    else:
        client = Client.objects.get(id=client_id)
        types_clients = TypeClient.objects.all().order_by('libelle')
        types_personnes = TypePersonne.objects.all().order_by('libelle')
        civilites = Civilite.objects.all().order_by('name')
        business_units = BusinessUnit.objects.all().order_by('libelle')
        bureaux = Bureau.objects.all().order_by('nom')
        pays = Pays.objects.all().order_by('nom')
        utilisateurs = User.objects.all().order_by('last_name')
        gestionnaires = User.objects.filter(bureau=request.user.bureau, type_utilisateur__code="INTERNE").order_by('last_name')
        genre = Genre
        groupes_internationaux = GroupeInter.objects.filter(status=True).order_by('nom')
        secteurs_activite = SecteurActivite.objects.filter(status=True).order_by('libelle')

        return render(request, 'client/modal_client_modification.html',
                      {'client': client, 'types_clients': types_clients, 'types_personnes': types_personnes, 'secteurs_activite': secteurs_activite,
                       'civilites': civilites, 'business_units': business_units, 'bureaux': bureaux, 'pays': pays,
                       'utilisateurs': utilisateurs, 'gestionnaires': gestionnaires, 'genre': genre, 'groupes_internationaux': groupes_internationaux, })


@login_required
def supprimer_client(request):
    if request.method == "POST":

        client_id = request.POST.get('client_id')

        client = Client.objects.get(id=client_id)
        if client.pk is not None:
            # client.delete()

            Client.objects.filter(id=client_id).update(statut=Statut.INACTIF)

            response = {
                'statut': 1,
                'message': "Client supprimé avec succès !",
            }

        else:

            response = {
                'statut': 0,
                'message': "Client non trouvé !",
            }

        return JsonResponse(response)


#Liste des polices du client
@method_decorator(login_required, name='dispatch')
class PoliceClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/client_polices.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()

            pprint(client.pays.devise)
            polices = Police.objects.filter(client_id=client_id, statut=StatutPolice.ACTIF, statut_contrat='CONTRAT', statut_validite=StatutValidite.VALIDE).order_by('-id')

            derniere_police = polices.first()
            nouvelles_polices = [derniere_police] if derniere_police and not derniere_police.has_beneficiaires else []

            #les anciennes polices qui un mouvement_police de résiliation
            anciennes_polices = polices.filter(
                id__in=MouvementPolice.objects.filter(
                    mouvement__code="RESIL",
                    statut_validite=StatutValidite.VALIDE,
                    #date_effet__gte=datetime.datetime.now(tz=timezone.utc).date(),
                    police_id__in=polices.values_list('id', flat=True)
                ).values_list('police_id', flat=True)
            )

            statut_contrat = "CONTRAT"

            quittances = []
            for police in polices:
                quittances_of_police = Quittance.objects.filter(police_id=police.id)
                quittances.extend(quittances_of_police)

            acomptes = Acompte.objects.filter(client_id=client_id)

            filiales = Filiale.objects.filter(client_id=client_id)

            documents = Document.objects.filter(client_id=client_id)

            contacts = Contact.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')

            types_documents = TypeDocument.objects.all().order_by('libelle')

            types_prefinancements = TypePrefinancement.objects.filter(statut=Statut.ACTIF).order_by('libelle')

            # pour la creation de police
            branches = Branche.objects.filter(status=True).order_by('nom')
            produits = Produit.objects.all().order_by('nom')
            bureaux = Bureau.objects.all().order_by('nom')
            utilisateurs = None  # User.objects.all().order_by('last_name')
            apporteurs = Apporteur.objects.filter(status=True).order_by('nom')
            tickets_moderateurs = TicketModerateur.objects.all().order_by('libelle')
            fractionnements = Fractionnement.objects.all().order_by('libelle')
            modes_reglements = ModeReglement.objects.all().order_by('libelle')
            regularisations = Regularisation.objects.all().order_by('libelle')
            compagnies = Compagnie.objects.filter(bureau=request.user.bureau, status=True).order_by('nom')
            durees = Duree.objects.all().order_by('libelle')
            devises = Devise.objects.filter(id=client.pays.devise_id).order_by('libelle')
            taxes = Taxe.objects.all().order_by('libelle')
            bureau_taxes = BureauTaxe.objects.filter(bureau_id=client.bureau_id)
            bases_calculs = BaseCalcul.objects.all().order_by('libelle')
            modes_calculs = ModeCalcul.objects.all().order_by('libelle')

            placement_gestion = PlacementEtGestion
            mode_renouvellement = ModeRenouvellement
            calcul_tm = CalculTM
            type_majoration_contrat = TypeMajorationContrat

            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {'client': client, 'contacts': contacts, 'polices': polices, 'quittances': quittances,
                             'acomptes': acomptes,
                             'filiales': filiales, 'documents': documents, 'types_documents': types_documents,
                             'branches': branches, 'produits': produits, 'pays': pays,
                             'compagnies': compagnies, 'durees': durees, 'placement_gestion': placement_gestion,
                             'mode_renouvellement': mode_renouvellement, 'tickets_moderateurs': tickets_moderateurs,
                             'calcul_tm': calcul_tm,
                             'fractionnements': fractionnements, 'modes_reglements': modes_reglements,
                             'regularisations': regularisations,
                             'devises': devises, 'utilisateurs': utilisateurs, 'bureaux': bureaux, 'taxes': taxes,
                             'bureau_taxes': bureau_taxes,
                             'apporteurs': apporteurs, 'bases_calculs': bases_calculs,
                             'type_majoration_contrat': type_majoration_contrat, 'modes_calculs': modes_calculs,
                             'statut_contrat': statut_contrat,
                             'types_prefinancements': types_prefinancements,
                             'anciennes_polices': anciennes_polices,
                             'nouvelles_polices': nouvelles_polices
                             }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


#Liste des contacts du client
@method_decorator(login_required, name='dispatch')
class ContactClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/client_contacts.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()


            statut_contrat = "CONTRAT"

            contacts = Contact.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')

            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {'client': client, 'contacts': contacts,
                             'pays': pays,
                             'bureaux': bureaux, 'statut_contrat': statut_contrat
                             }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


#Liste des filiales du client
@method_decorator(login_required, name='dispatch')
class FilialeClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/client_filiales.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()

            statut_contrat = "CONTRAT"

            filiales = Filiale.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')


            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {'client': client,
                             'filiales': filiales, 'pays': pays,
                             'bureaux': bureaux, 'statut_contrat': statut_contrat
                             }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


#Liste des acomptes du client
@method_decorator(login_required, name='dispatch')
class AcompteClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/client_acomptes.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()

            acomptes = Acompte.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')

            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {
                'client': client,
                'acomptes': acomptes,
                'pays': pays,
                'bureaux': bureaux
            }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


#Liste des documents électronique du client
@method_decorator(login_required, name='dispatch')
class GEDClientView(TemplateView):
    permission_required = "production.view_clients"
    template_name = 'client/client_documents.html'
    model = Client

    def get(self, request, client_id, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)


        clients = Client.objects.filter(id=client_id, bureau=request.user.bureau)
        if clients:
            client = clients.first()

            pprint(client.pays.devise)

            statut_contrat = "CONTRAT"

            documents = Document.objects.filter(client_id=client_id)

            pays = Pays.objects.all().order_by('nom')

            bureaux = Bureau.objects.filter(id=request.user.bureau.id)

            context_perso = {'client': client, 'documents': documents, 'pays': pays,
                             'bureaux': bureaux, 'statut_contrat': statut_contrat
                             }

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")


    def post(self):
        pass

    def get_context_data(self, **kwargs):

        pprint(kwargs)
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }



# Test excel file exploid

class ExcelFileView(View):
    def get(self, request, *args, **kwargs):
        formule = FormuleGarantie.objects.get(id=1)
        print(formule.tarfile.path)

        excel_data_df = pd.read_excel(formule.tarfile.path, sheet_name='actes')
        print(excel_data_df.head())
        print(excel_data_df)

        acte_row = excel_data_df.loc[excel_data_df['Code acte'] == 'CH0106']

        print(acte_row)
        print(acte_row.empty)
        print(acte_row.to_dict())
        print(acte_row.to_dict()['Code acte'])
        # print(acte_row.col)
        value = [i for i in acte_row.to_dict()['Code acte'] if acte_row.to_dict()['Code acte'][i] == "CH0106"]
        print("key by value:", value)
        print("key by value:", value[0])
        print(acte_row.to_dict()['Code acte'][value[0]])
        print(acte_row.to_dict().keys())
        acts = acte_row.to_dict()
        acts_kay = acte_row.to_dict().keys()

        dic = {}
        for i in list(acts_kay):
            print(i)
            print(type(i))
            dic[i] = acts[i][value[0]]

        print(dic)

        return HttpResponse("Get method is not allowed")


# liste des formules universelles pour assurance universelle
class FormulesUniversellesView(TemplateView):
    # permission_required = "production.view_formules"
    template_name = 'police/formules_universelles.html'
    model = FormuleGarantie

    # @never_cache
    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        # police_id = kwargs['police_id']

        # police = Police.objects.get(id=police_id)
        bureau = request.user.bureau
        formules = FormuleGarantie.objects.filter(police__isnull=True, bureau=request.user.bureau, statut=Statut.ACTIF)
        territorialites = Territorialite.objects.all().order_by('libelle')
        types_tarifs = TypeTarif.objects.all().order_by('libelle')
        reseaux_soins = ReseauSoin.objects.filter(bureau=bureau)
        rubriques = Rubrique.objects.filter(status=True).order_by('libelle')
        mode_prefinancements = ModePrefinancement.objects.all().order_by('libelle')

        today = datetime.datetime.now(tz=timezone.utc)

        context_perso = {'bureau': bureau, 'formules': formules, 'types_tarifs': types_tarifs,
                         'territorialites': territorialites, 'reseaux_soins': reseaux_soins, 'rubriques': rubriques, 'mode_prefinancements':mode_prefinancements, 'today': today}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# ajout de formule universelle pour assurance universelle
@login_required()
def add_formule_universelle(request):
    # police = Police.objects.get(id=police_id)

    # if police and request.method == 'POST':
    if request.method == 'POST':
        territorialite_id = request.POST.get('territorialite')
        type_tarif_id = request.POST.get('type_tarif')
        reseau_soin_id = request.POST.get('reseau_soin')
        taux_couverture = request.POST.get('taux_couverture')
        libelle = request.POST.get('libelle')
        plafond_conso_famille = request.POST.get('plafond_conso_famille').replace(' ', '')
        plafond_conso_individuelle = request.POST.get('plafond_conso_individuelle').replace(' ', '')
        garantis_pharmacie = request.POST.get('garantis_pharmacie')
        date_debut = request.POST.get('date_debut')

        #date_debut = police.date_debut_effet

        # Ajout des 5/6 champs complementaires
        infos_carte_consultation = request.POST.get('infos_carte_consultation')
        infos_carte_hospitalisation = request.POST.get('infos_carte_hospitalisation')
        infos_carte_ambulatoire = request.POST.get('infos_carte_ambulatoire')
        infos_carte_vitamine = request.POST.get('infos_carte_vitamine')
        infos_carte_vaccination = request.POST.get('infos_carte_vaccination')

        infos_carte_numero_police = request.POST.get('infos_carte_numero_police')
        infos_carte_show_numero_police = request.POST.get('infos_carte_show_numero_police',0)

        option_mode_prefinancement = request.POST.get('option_mode_prefinancement')
        selected_rubriques = request.POST.getlist('selected_rubriques') if option_mode_prefinancement == "TPP" else []

        # print(option_type_prefinancement)
        # print(selected_rubriques)

        # obsolete_formule_rubriques = FormuleRubriquePrefinance.objects.filter(statut_validite=StatutValidite.VALIDE).exclude(rubrique__id__in=selected_rubriques)            

        # print(obsolete_formule_rubriques)

        if plafond_conso_famille == "": plafond_conso_famille = 0
        if plafond_conso_individuelle == "": plafond_conso_individuelle = 0

        taux_couverture = int(taux_couverture)
        taux_tm = 100 - taux_couverture

        territorialite = Territorialite.objects.get(id=territorialite_id)
        type_tarif = TypeTarif.objects.get(id=type_tarif_id)
        reseau_soin = ReseauSoin.objects.get(id=reseau_soin_id) if reseau_soin_id else None

        mode_prefinancement = ModePrefinancement.objects.filter(code=option_mode_prefinancement).first()

        formule = FormuleGarantie.objects.create(
            created_by=request.user,
            # police_id=police_id,
            territorialite_id=territorialite.pk,
            type_tarif_id=type_tarif.pk,
            reseau_soin=reseau_soin,
            libelle=libelle,
            taux_couverture=taux_couverture,
            plafond_conso_famille=plafond_conso_famille,
            plafond_conso_individuelle=plafond_conso_individuelle,
            taux_tm=taux_tm,
            garantis_pharmacie=garantis_pharmacie,
            date_debut=date_debut,
            infos_carte_consultation=infos_carte_consultation,
            infos_carte_hospitalisation=infos_carte_hospitalisation,
            infos_carte_ambulatoire=infos_carte_ambulatoire,
            infos_carte_vitamine=infos_carte_vitamine,
            infos_carte_vaccination=infos_carte_vaccination,
            infos_carte_numero_police=infos_carte_numero_police,
            infos_carte_show_numero_police=infos_carte_show_numero_police,
            mode_prefinancement = mode_prefinancement,
            bureau = request.user.bureau
        )

        formule.save()

        code_bureau = request.user.bureau.code
        formule.code = 'F' + str(Date.today().year)[-2:] + '' + str(formule.pk).zfill(7) + str(code_bureau)
        formule.save()

        # associons des rubriques en cas de tiers payant partiel
        if option_mode_prefinancement == "TPP":
            for selected_rubrique in selected_rubriques:
                old_formule_rubrique = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, rubrique__id=int(selected_rubrique), statut_validite=StatutValidite.VALIDE).first()
                if old_formule_rubrique is None:
                    rubrique = Rubrique.objects.filter(id=int(selected_rubrique), status=True).first()
                    new_formule_rubrique = FormuleRubriquePrefinance.objects.create(
                        formulegarantie = formule,
                        rubrique = rubrique,
                        created_by=request.user
                    )
                    new_formule_rubrique.save()

        # suppressions des rubriques en dehors des selectionnées en cas de tiers payant partiel ou de l'ensemble dans le cas de tiers payant generalisé
        obsolete_formule_rubriques = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, statut_validite=StatutValidite.VALIDE).exclude(rubrique__id__in=selected_rubriques)
        for obsolete_formule_rubrique in obsolete_formule_rubriques:
            obsolete_formule_rubrique.statut_validite = StatutValidite.SUPPRIME
            obsolete_formule_rubrique.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': formule.pk,
                'libelle': formule.libelle,
                'territorialite': formule.territorialite.libelle,
                'taux_couverture': formule.taux_couverture,
                'type_tarif': formule.type_tarif.libelle,
                'date_debut': formule.date_debut,
                'date_fin': formule.date_fin,
                'statut': formule.statut,
            }
        }

        return JsonResponse(response)


class FormulesView(TemplateView):
    # permission_required = "production.view_formules"
    template_name = 'police/formules.html'
    model = FormuleGarantie

    # @never_cache
    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        police_id = kwargs['police_id']

        police = Police.objects.filter(id=police_id, bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).first()
        if police:
            formules = FormuleGarantie.objects.filter(police_id=police_id, statut=Statut.ACTIF)
            territorialites = Territorialite.objects.all().order_by('libelle')
            types_tarifs = TypeTarif.objects.all().order_by('libelle')
            reseaux_soins = ReseauSoin.objects.filter(bureau=police.bureau)
            rubriques = Rubrique.objects.filter(status=True).order_by('libelle')
            mode_prefinancements = ModePrefinancement.objects.all().order_by('libelle')

            today = timezone.now().date()

            context_perso = {'police': police, 'formules': formules, 'types_tarifs': types_tarifs,
                             'territorialites': territorialites, 'reseaux_soins': reseaux_soins, 'rubriques': rubriques, 'mode_prefinancements':mode_prefinancements, 'today': today}

            context = {**context_original, **context_perso}

            return self.render_to_response(context)

        else:
            return redirect("clients")

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# ajout de bareme
@login_required()
def add_formule(request, police_id):
    police = Police.objects.get(id=police_id)

    if police and request.method == 'POST':
        territorialite_id = request.POST.get('territorialite')
        type_tarif_id = request.POST.get('type_tarif')
        reseau_soin_id = request.POST.get('reseau_soin')
        taux_couverture = request.POST.get('taux_couverture')
        libelle = request.POST.get('libelle')
        plafond_conso_famille = request.POST.get('plafond_conso_famille').replace(' ', '')
        plafond_conso_individuelle = request.POST.get('plafond_conso_individuelle').replace(' ', '')
        garantis_pharmacie = request.POST.get('garantis_pharmacie')
        #date_debut = request.POST.get('date_debut')
        date_debut = police.date_debut_effet

        # Ajout des 5/6 champs complementaires
        infos_carte_consultation = request.POST.get('infos_carte_consultation')
        infos_carte_hospitalisation = request.POST.get('infos_carte_hospitalisation')
        infos_carte_ambulatoire = request.POST.get('infos_carte_ambulatoire')
        infos_carte_vitamine = request.POST.get('infos_carte_vitamine')
        infos_carte_vaccination = request.POST.get('infos_carte_vaccination')

        infos_carte_numero_police = request.POST.get('infos_carte_numero_police')
        infos_carte_show_numero_police = request.POST.get('infos_carte_show_numero_police')

        option_mode_prefinancement = request.POST.get('option_mode_prefinancement')
        selected_rubriques = request.POST.getlist('selected_rubriques') if option_mode_prefinancement == "TPP" else []


        exclusion = request.POST.get('exclusion')
        # print(option_type_prefinancement)
        # print(selected_rubriques)

        # obsolete_formule_rubriques = FormuleRubriquePrefinance.objects.filter(statut_validite=StatutValidite.VALIDE).exclude(rubrique__id__in=selected_rubriques)            

        # print(obsolete_formule_rubriques)

        if plafond_conso_famille == "": plafond_conso_famille = 0
        if plafond_conso_individuelle == "": plafond_conso_individuelle = 0

        taux_couverture = int(taux_couverture)
        taux_tm = 100 - taux_couverture

        territorialite = Territorialite.objects.get(id=territorialite_id)
        type_tarif = TypeTarif.objects.get(id=type_tarif_id)
        reseau_soin = ReseauSoin.objects.get(id=reseau_soin_id) if reseau_soin_id else None

        mode_prefinancement = ModePrefinancement.objects.filter(code=option_mode_prefinancement).first()

        formule = FormuleGarantie.objects.create(
            created_by=request.user,
            police_id=police_id,
            territorialite_id=territorialite.pk,
            type_tarif_id=type_tarif.pk,
            reseau_soin=reseau_soin,
            libelle=libelle,
            taux_couverture=taux_couverture,
            plafond_conso_famille=plafond_conso_famille,
            plafond_conso_individuelle=plafond_conso_individuelle,
            taux_tm=taux_tm,
            garantis_pharmacie=garantis_pharmacie,
            date_debut=date_debut,
            infos_carte_consultation=infos_carte_consultation,
            infos_carte_hospitalisation=infos_carte_hospitalisation,
            infos_carte_ambulatoire=infos_carte_ambulatoire,
            infos_carte_vitamine=infos_carte_vitamine,
            infos_carte_vaccination=infos_carte_vaccination,
            infos_carte_numero_police=infos_carte_numero_police,
            infos_carte_show_numero_police=infos_carte_show_numero_police,
            mode_prefinancement=mode_prefinancement,
            bureau=police.bureau,
            exclusion=exclusion
        )

        formule.save()

        code_bureau = request.user.bureau.code
        formule.code = 'F' + str(Date.today().year)[-2:] + '' + str(formule.pk).zfill(7) + str(code_bureau)
        formule.save()

        # associons des rubriques en cas de tiers payant partiel
        if option_mode_prefinancement == "TPP":
            for selected_rubrique in selected_rubriques:
                old_formule_rubrique = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, rubrique__id=int(selected_rubrique), statut_validite=StatutValidite.VALIDE).first()
                if old_formule_rubrique is None:
                    rubrique = Rubrique.objects.filter(id=int(selected_rubrique), status=True).first()
                    new_formule_rubrique = FormuleRubriquePrefinance.objects.create(
                        formulegarantie = formule,
                        rubrique = rubrique,
                        created_by=request.user
                    )
                    new_formule_rubrique.save()

        # suppressions des rubriques en dehors des selectionnées en cas de tiers payant partiel ou de l'ensemble dans le cas de tiers payant generalisé
        obsolete_formule_rubriques = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, statut_validite=StatutValidite.VALIDE).exclude(rubrique__id__in=selected_rubriques)
        for obsolete_formule_rubrique in obsolete_formule_rubriques:
            obsolete_formule_rubrique.statut_validite = StatutValidite.SUPPRIME
            obsolete_formule_rubrique.save()

        response = {
            'statut': 1,
            'message': "Enregistrement effectuée avec succès !",
            'data': {
                'id': formule.pk,
                'libelle': formule.libelle,
                'territorialite': formule.territorialite.libelle,
                'taux_couverture': formule.taux_couverture,
                'type_tarif': formule.type_tarif.libelle,
                'date_debut': formule.date_debut,
                'date_fin': formule.date_fin,
                'statut': formule.statut,
            }
        }

        return JsonResponse(response)


# dialog modifier formule et update réel
@login_required()
@never_cache
def modifier_formule(request, formule_id):
    formule = FormuleGarantie.objects.get(id=formule_id)
    police = formule.police if formule else None

    if request.method == 'POST':

        formule_before = formule
        pprint(formule_before)

        libelle = request.POST.get('libelle')
        taux_couverture = int(request.POST.get('taux_couverture').replace(" ", ""))
        type_tarif_id = request.POST.get('type_tarif')
        reseau_soin_id = request.POST.get('reseau_soin')
        territorialite_id = request.POST.get('territorialite')
        plafond_conso_individuelle = request.POST.get('plafond_conso_individuelle').replace(" ", "")
        plafond_conso_famille = request.POST.get('plafond_conso_famille').replace(" ", "")
        garantis_pharmacie = request.POST.get('garantis_pharmacie')
        date_debut = request.POST.get('date_debut') if formule_before.police is None else formule.date_debut
        exclusion = request.POST.get('exclusion')

        # Infos complementaires
        infos_carte_consultation = request.POST.get('infos_carte_consultation', '')
        infos_carte_hospitalisation = request.POST.get('infos_carte_hospitalisation', '')
        infos_carte_ambulatoire = request.POST.get('infos_carte_ambulatoire', '')
        infos_carte_vitamine = request.POST.get('infos_carte_vitamine', '')
        infos_carte_vaccination = request.POST.get('infos_carte_vaccination', '')
        infos_carte_show_numero_police = request.POST.get('infos_carte_show_numero_police', 0)

        option_mode_prefinancement = request.POST.get('option_mode_prefinancement')
        selected_rubriques = request.POST.getlist('selected_rubriques') if option_mode_prefinancement == "TPP" else []

        mode_prefinancement = ModePrefinancement.objects.filter(code=option_mode_prefinancement).first()

        if plafond_conso_famille == "": plafond_conso_famille = 0
        if plafond_conso_individuelle == "": plafond_conso_individuelle = 0

        if taux_couverture >= 0 and taux_couverture <= 100:
            formule.libelle = libelle
            formule.taux_couverture = taux_couverture
            formule.taux_tm = 100 - taux_couverture
            formule.type_tarif_id = type_tarif_id
            formule.territorialite_id = territorialite_id
            if reseau_soin_id: formule.reseau_soin_id = reseau_soin_id
            formule.plafond_conso_individuelle = plafond_conso_individuelle
            formule.plafond_conso_famille = plafond_conso_famille
            formule.garantis_pharmacie = garantis_pharmacie
            #formule.date_debut = date_debut
            formule.updated_at = datetime.datetime.now(tz=timezone.utc)
            formule.updated_by = request.user
            formule.exclusion=exclusion

            formule.infos_carte_consultation = infos_carte_consultation
            formule.infos_carte_hospitalisation = infos_carte_hospitalisation
            formule.infos_carte_ambulatoire = infos_carte_ambulatoire
            formule.infos_carte_vitamine = infos_carte_vitamine
            formule.infos_carte_vaccination = infos_carte_vaccination
            formule.infos_carte_show_numero_police = infos_carte_show_numero_police
            formule.mode_prefinancement = mode_prefinancement
            formule.date_debut = date_debut


            ActionLog.objects.create(done_by=request.user, action="update",
                                     description="Modification d'une formule de garantie", table="formulegarantie",
                                     row=formule.pk,
                                     #data_before=json.dumps(custom_model_to_dict(formule_before)),
                                     #data_after=json.dumps(custom_model_to_dict(formule))
                                     )

            formule.save()

            # associons des rubriques en cas de tiers payant partiel
            if option_mode_prefinancement == "TPP":
                for selected_rubrique in selected_rubriques:
                    old_formule_rubrique = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, rubrique__id=int(selected_rubrique), statut_validite=StatutValidite.VALIDE).first()
                    if old_formule_rubrique is None:
                        rubrique = Rubrique.objects.filter(id=int(selected_rubrique), status=True).first()
                        new_formule_rubrique = FormuleRubriquePrefinance.objects.create(
                            formulegarantie = formule,
                            rubrique = rubrique,
                            created_by=request.user
                        )
                        new_formule_rubrique.save()

            # suppressions des rubriques en dehors des selectionnées en cas de tiers payant partiel ou de l'ensemble dans le cas de tiers payant generalisé
            obsolete_formule_rubriques = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, statut_validite=StatutValidite.VALIDE).exclude(rubrique__id__in=selected_rubriques)
            for obsolete_formule_rubrique in obsolete_formule_rubriques:
                obsolete_formule_rubrique.statut_validite = StatutValidite.SUPPRIME
                obsolete_formule_rubrique.save()

            response = {
                'statut': 1,
                'message': "Modification effectuée avec succès !",
                'data': {
                    'id': formule.pk,
                    'libelle': formule.libelle,
                    'statut': formule.statut,
                }
            }

        else:
            response = {
                'statut': 0,
                'message': "Le taux de couverture doit être compris entre 0 et 100",
                'data': {
                }
            }

        return JsonResponse(response)


    else:

        formule = FormuleGarantie.objects.get(id=formule_id)
        formules = FormuleGarantie.objects.all()
        territorialites = Territorialite.objects.all().order_by('libelle')
        types_tarifs = TypeTarif.objects.all().order_by('libelle')
        reseaux_soins = ReseauSoin.objects.filter(bureau=request.user.bureau).order_by('nom')

        today =datetime.datetime.now(tz=timezone.utc)

        rubriques = Rubrique.objects.filter(status=True).order_by('libelle')
        mode_prefinancements = ModePrefinancement.objects.all().order_by('libelle')
        formule_rubriques = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, statut_validite=StatutValidite.VALIDE).order_by('rubrique__libelle')

        # formule.plafond_conso_famille = float(formule.plafond_conso_famille)
        # formule.plafond_conso_individuelle = float(formule.plafond_conso_individuelle)

        template = 'police/modal_formule_update.html' if police else 'police/modal_formule_update_universelle.html'

        return render(request, template, {'formule': formule, 'types_tarifs': types_tarifs,
                                                                    'territorialites': territorialites, 'reseaux_soins': reseaux_soins, 'rubriques': rubriques, 'mode_prefinancements':mode_prefinancements, 'formule_rubriques':formule_rubriques, 'today': today, 'police': police})

# update formule
def desactivate_formule(request):
    if request.method == 'POST':
        formule_id = request.POST.get('id')

        formule = FormuleGarantie.objects.get(id=formule_id)
        formule_before = formule

        #voir si aucun bénéficiaire dessus avant de supprimer
        aliments_formules = AlimentFormule.objects.filter(formule_id=formule_id)

        if not aliments_formules:
            date_fin = request.POST.get('date_fin')

            formule.date_fin = date_fin
            formule.statut = Statut.INACTIF
            formule.deleted_by = request.user

            ActionLog.objects.create(done_by=request.user, action="delete",
                                     description="supppression d'une formule de garantie", table="formulegarantie",
                                     row=formule.pk, data_before="", data_after="")

            formule.save()

            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {
                    'id': formule.pk,
                    'date_debut': formule.date_debut,
                    'statut': formule.statut,
                }
            }

        else:
            response = {
                'statut': 0,
                'message': "Il existe des bénéficiaires sur la police !",
                'data': {
                }
            }
        if formule.police:
            return redirect("police_formules", formule.police.id)
        else:
            return redirect("formules_universelles")


class DetailsFormuleView(PermissionRequiredMixin, TemplateView):
    permission_required = "production.view_formulegarantie"
    template_name = 'police/details_formule.html'
    model = Bareme

    @method_decorator(never_cache)
    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        formule_id = kwargs['formule_id']

        pprint(formule_id)
        formule = FormuleGarantie.objects.get(id=formule_id)
        baremes = Bareme.objects.filter(formulegarantie_id=formule_id, statut=Statut.ACTIF)
        actes = Acte.objects.filter(status=1).order_by('libelle')
        rubriques = Rubrique.objects.filter(status=1).order_by('libelle')
        sous_rubriques = SousRubrique.objects.filter(status=1).order_by('libelle')
        regroupements = RegroupementActe.objects.all().order_by('libelle')
        sous_regroupements = SousRegroupementActe.objects.all().order_by('libelle')
        qualites_beneficiaires = QualiteBeneficiaire.objects.all().order_by('libelle')
        territorialites = Territorialite.objects.all().order_by('libelle')
        periodicites = Periodicite.objects.all().order_by('id')
        rubriques_prefinancees = FormuleRubriquePrefinance.objects.filter(formulegarantie=formule, statut_validite=StatutValidite.VALIDE)

        date_debut_effet_police = formule.police.date_debut_effet if formule.police else None
        today = timezone.now().date()

        police = formule.police
        client = police.client if formule.police else None

        self.template_name = self.template_name if formule.police else 'police/details_formule_universelle.html'

        context_perso = {'client': client, 'police': police, 'formule': formule, 'baremes': baremes,
                         'rubriques': rubriques, 'sous_rubriques': sous_rubriques, 'regroupements': regroupements, 'sous_regroupements': sous_regroupements,
                         'actes': actes, 'qualites_beneficiaires': qualites_beneficiaires,
                         'territorialites': territorialites, 'periodicites': periodicites, 'date_debut_effet_police': date_debut_effet_police, 'today': today, 'rubriques_prefinancees': rubriques_prefinancees}

        context = {**context_original, **context_perso}

        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }


# ajout des spécificités du barème
def add_bareme(request, formule_id):
    # dd(request)
    formule = FormuleGarantie.objects.get(id=formule_id)

    pprint(formule)
    if formule and request.method == 'POST':

        rubrique_id = request.POST.get('rubrique')
        acte_id = request.POST.get('acte')
        sous_rubrique_id = request.POST.get('sous_rubrique')
        regroupement_acte_id = request.POST.get('regroupement_acte')
        sous_regroupement_acte_id = request.POST.get('sous_regroupement_acte')
        is_garanti = request.POST.get('is_garanti')
        qualite_beneficiaire_id = request.POST.get('qualite_beneficiaire')
        plafond_famille = request.POST.get('plafond_famille', '').replace(" ", "")
        plafond_individuel = request.POST.get('plafond_individuel', '').replace(" ", "")
        taux_tm = request.POST.get('taux_tm', '').replace(" ", "")
        taux_couverture = request.POST.get('taux_couverture', '').replace(" ", "")
        plafond_rubrique = request.POST.get('plafond_rubrique', '').replace(" ", "")
        plafond_acte = request.POST.get('plafond_acte', '').replace(" ", "")
        plafond_regroupement_acte = request.POST.get('plafond_regroupement_acte', '').replace(" ", "")
        plafond_sous_regroupement_acte = request.POST.get('plafond_sous_regroupement_acte', '').replace(" ", "")
        plafond_sous_rubrique = request.POST.get('plafond_sous_rubrique', '').replace(" ", "")
        nombre_acte = request.POST.get('nombre_acte', '').replace(" ", "")
        periodicite_id = request.POST.get('periodicite')
        age_minimum = request.POST.get('age_minimum', '').replace(" ", "")
        age_maximum = request.POST.get('age_maximum', '').replace(" ", "")
        date_debut = request.POST.get('date_debut')

        # if acte_id == "": acte_id = 0
        if taux_tm == "": taux_tm = 0
        if taux_couverture == "": taux_couverture = 0
        if plafond_famille == "": plafond_famille = 0
        if plafond_individuel == "": plafond_individuel = 0
        if plafond_rubrique == "": plafond_rubrique = 0
        if plafond_sous_rubrique == "": plafond_sous_rubrique = 0
        if plafond_regroupement_acte == "": plafond_regroupement_acte = 0
        if plafond_sous_regroupement_acte == "": plafond_sous_regroupement_acte = 0
        if plafond_acte == "": plafond_acte = 0
        if nombre_acte == "": nombre_acte = 0
        if age_minimum == "": age_minimum = 0
        if age_maximum == "": age_maximum = 0

        acte = Acte.objects.filter(id=acte_id).first() if acte_id else None
        acte_libelle = acte.libelle if acte else ''

        rubrique = Rubrique.objects.filter(id=rubrique_id).first() if rubrique_id else None
        rubrique_libelle = rubrique.libelle if rubrique else ''

        sous_rubrique = Rubrique.objects.filter(id=sous_rubrique_id).first() if sous_rubrique_id else None
        sous_rubrique_libelle = sous_rubrique.libelle if sous_rubrique else ''

        regroupement_acte = RegroupementActe.objects.filter(
            id=regroupement_acte_id).first() if regroupement_acte_id else None
        regroupement_acte_libelle = regroupement_acte.libelle if regroupement_acte else ''

        qualite_beneficiaire = QualiteBeneficiaire.objects.filter(
            id=qualite_beneficiaire_id).first() if qualite_beneficiaire_id else None
        qualite_beneficiaire_libelle = qualite_beneficiaire.libelle if qualite_beneficiaire else ''

        # qualite_beneficiaire = QualiteBeneficiaire.objects.filter(id=qualite_beneficiaire_id).first() if qualite_beneficiaire_id else None
        # qualite_beneficiaire_libelle = qualite_beneficiaire.libelle if qualite_beneficiaire else ''

        if taux_couverture == 0:
            # taux_couverture = formule.taux_couverture #retirer : mettre en champ obligatoire
            pass

        # Vérification de doublon
        existing_bareme = Bareme.objects.filter(
            formulegarantie_id=formule_id,
            rubrique_id=rubrique_id if rubrique_id else None,
            sous_rubrique_id=sous_rubrique_id if sous_rubrique_id else None,
            regroupement_acte_id=regroupement_acte_id if regroupement_acte_id else None,
            acte_id=acte_id if acte_id else None,
            is_garanti=is_garanti,
            qualite_beneficiaire_id=qualite_beneficiaire_id if qualite_beneficiaire_id else None,
            taux_couverture=taux_couverture,
            age_minimum=age_minimum,
            age_maximum=age_maximum,
            date_debut=date_debut,
            statut=Statut.ACTIF,
        ).exists()

        if existing_bareme:

            response = {
                'statut': 0,
                'message': "Attention, cette spécificité du bareme a été déjà parametré. Veuillez bien vérifier !",
                'errors': {
                }
            }

        else:

            bareme = Bareme.objects.create(created_by=request.user,
                                           formulegarantie_id=formule_id,
                                           rubrique_id=rubrique_id,
                                           sous_rubrique_id=sous_rubrique_id,
                                           regroupement_acte_id=regroupement_acte_id,
                                           sous_regroupement_acte_id=sous_regroupement_acte_id,
                                           acte_id=acte_id if acte_id else None,
                                           is_garanti=is_garanti,
                                           qualite_beneficiaire_id=qualite_beneficiaire_id,
                                           taux_couverture=taux_couverture,
                                           taux_tm=100 - int(taux_couverture),
                                           plafond_individuel=plafond_individuel,
                                           plafond_famille=plafond_famille,
                                           plafond_rubrique=plafond_rubrique,
                                           plafond_sous_rubrique=plafond_sous_rubrique,
                                           plafond_regroupement_acte=plafond_regroupement_acte,
                                           plafond_sous_regroupement_acte=plafond_sous_regroupement_acte,
                                           plafond_acte=plafond_acte,
                                           nombre_acte=nombre_acte,
                                           periodicite_id=periodicite_id,
                                           # unite_frequence = unite_frequence,
                                           # frequence = frequence,
                                           age_minimum=age_minimum,
                                           age_maximum=age_maximum,
                                           date_debut=date_debut,
                                           )

            bareme.save()

            # save in log table
            ActionLog.objects.create(done_by=request.user, action="add",
                                     description="ajout d'un barème", table="bareme",
                                     row=bareme.pk, data_before="",
                                     data_after=json.dumps(model_to_dict(bareme)))

            response = {
                'statut': 1,
                'message': "Enregistrement effectuée avec succès !",
                'data': {
                    'id': bareme.pk,
                    'acte': acte_libelle,
                    'is_garanti': 'GARANTI' if bareme.is_garanti else "EXCLU",
                    'taux_tm': bareme.taux_tm,
                    'taux_couverture': bareme.taux_couverture,
                    'rubrique': rubrique_libelle,
                    'sous_rubrique': sous_rubrique_libelle,
                    'regroupement_acte': regroupement_acte_libelle,
                    'plafond_rubrique': bareme.plafond_rubrique,
                    'plafond_sous_rubrique': bareme.plafond_sous_rubrique,
                    'plafond_regroupement_acte': bareme.plafond_regroupement_acte,
                    'plafond_acte': bareme.plafond_acte,
                    'age_minimum': bareme.age_maximum,
                    'age_maximum': bareme.age_maximum,
                    'qualite_beneficiaire': qualite_beneficiaire_libelle,
                    'date_debut': bareme.date_debut,
                    'date_fin': '',
                    'statut': bareme.statut,
                }
            }

        return JsonResponse(response)


def detail_bareme(request, bareme_id):
    bareme = Bareme.objects.get(id=bareme_id)
    return render(request, 'police/modal_details_formule.html', {
        'bareme': bareme
    })


def del_bareme(request):
    if request.method == 'POST':
        bareme_id = request.POST.get('bareme_id')
        date_fin = request.POST.get('date_fin')

        Bareme.objects.filter(id=bareme_id).update(statut=Statut.INACTIF, deleted_by=request.user, date_fin=date_fin,
                                                   deleted_at=datetime.datetime.now(tz=timezone.utc))

        response = {
            'statut': 1,
            'message': "Barème désactivé avec succès !",
            'data': {
            }
        }

        return JsonResponse(response)

def generate_qrcode(request):

    cartes = Carte.objects.filter(Q(qrcode_file__isnull=True) | Q(qrcode_file=""))[:5000] #Q(qrcode_file__isnull=True) | Q(qrcode_file__isnull=""))[:5000] #generate-qrcode-carte
    print("cartes")
    #dd(cartes)
    print(len(cartes))
    datas = []
    for carte in cartes:
        # générer le qrcode
        qrcode_file = generer_qrcode_carte(carte.numero)
        print("carte.numero")
        print(carte.numero)
        carte.qrcode_file.save(f'qrcode_img_{carte.numero}.png', qrcode_file)
        carte.save()
        info = "New"

        datas.append({"id": carte.id, "numero": carte.numero, "qrcode_file": carte.qrcode_file.url, "info": info})

    response = {
        'statut': 1,
        'message': "Generation qrcode effectué avec succès !",
        'data': datas
    }

    return JsonResponse(response)

def beneficiaire_carte_pdf(request):
    logo = settings.JAZZMIN_SETTINGS.get('logo_for_carte')
    print(logo)
    aliments = [1,2,3,4,5]
    pdf = render_pdf('police/courriers/cartes-model.html', {"logo": logo, "aliments": aliments})
    return HttpResponse(File(pdf), content_type='application/pdf')

def beneficiaire_carte_html(request):
    return render(request, 'police/courriers/cartes.html', {})


@login_required()
def export_beneficiaires(request, police_id):
    police = Police.objects.get(id=police_id)
    today = timezone.now().date()

    if police:
        aliment_ids = AlimentFormule.objects.filter(
            formule_id__in=[p.id for p in police.formules],
            statut=Statut.ACTIF,
            statut_validite=StatutValidite.VALIDE
        ).values_list('aliment_id', flat=True)

        queryset = Aliment.objects.filter(id__in=aliment_ids).order_by('adherent_principal_id', 'qualite_beneficiaire_id', 'nom', 'prenoms')

        #print(str(queryset.query))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_DES_BENEFICIAIRES-POLICE_'+str(police.numero)+'.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'LISTE DES BENEFICIAIRES'

        header = [
            'NUMERO_POLICE', 'NUMERO_CARTE', 'NOM', 'PRENOMS', 'DATE_NAISSANCE', 'GENRE',
            'MATRICULE', 'ADHERENT_PRINCIPAL', 'LIEN', 'NUMERO_FAMILLE', 'DATE_ENTREE',
            'DATE_SORTI', 'FORMULE', 'TELEPHONE', 'ETAT',
        ]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        data = []
        liste_aliments_ajoutes = []

        for aliment in queryset:
            if aliment.id not in liste_aliments_ajoutes:
                liste_aliments_ajoutes.append(aliment.id)

            historique_formules = aliment.historique_formules.filter(formule__police=police).exclude(
                statut_validite=StatutValidite.SUPPRIME).order_by('date_debut')

            date_entree_police = historique_formules.first().date_debut if historique_formules else None

            numero_carte = aliment.carte_active().numero if aliment.carte_active() else None
            derniere_formule = AlimentFormule.objects.filter(aliment=aliment).order_by('-id').first()
            derniere_formule_libelle = derniere_formule.formule.libelle if derniere_formule and derniere_formule.formule else ''
            etat_beneficiaire = aliment.etat_beneficiaire_atdate(today)
            sa_formule = aliment.formule_atdate(today)
            sa_formule_libelle = sa_formule.libelle if sa_formule else ''

            data_iten = [
                police.numero, numero_carte, aliment.nom, aliment.prenoms,
                aliment.date_naissance.strftime("%d/%m/%Y") if aliment.date_naissance else '',
                aliment.genre, aliment.matricule_employe if aliment.matricule_employe else '',
                aliment.adherent_principal.nom + ' ' + aliment.adherent_principal.prenoms if aliment.adherent_principal else '',
                aliment.qualite_beneficiaire.libelle if aliment.qualite_beneficiaire else '',
                aliment.adherent_principal.numero_famille if aliment.adherent_principal else aliment.numero_famille,
                date_entree_police.strftime("%d/%m/%Y") if date_entree_police else '',
                aliment.date_sortie.strftime("%d/%m/%Y") if aliment.date_sortie else '',
                sa_formule_libelle if sa_formule_libelle else derniere_formule_libelle,
                aliment.telephone_mobile, etat_beneficiaire,
            ]
            data.append(data_iten)

        for row_num, row in enumerate(data, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    else:
        return JsonResponse({"message": "Police non trouvée"}, status=404)


@login_required()
def export_sinistres_beneficiaire(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    #today = datetime.datetime.now(tz=timezone.utc)
    aliment = Aliment.objects.get(id=aliment_id)

    if aliment:
        #queryset = Aliment.objects.select_related('qualite_beneficiaire', 'police__client').filter(formulegarantie__police=police=police).order_by('-id')
        # Optimisation de la recuperation des beneficiaires
        #queryset = AlimentFormule.objects.filter(formule_id__in=[p.id for p in police.formules], statut=Statut.ACTIF).order_by('-id')

        queryset = Sinistre.objects.filter(aliment_id=aliment_id, statut_validite=StatutValidite.VALIDE).order_by('-id')

        print("queryset")
        print(queryset)
        print(queryset.count())

        numero_carte = aliment.carte_active().numero if aliment.carte_active() else ""
        beneficiaire = aliment.nom + " " + aliment.prenoms
        qualite = aliment.qualite_beneficiaire.libelle if aliment.qualite_beneficiaire else ''
        adherent_principal = aliment.adherent_principal.nom + " " + aliment.adherent_principal.prenoms

        # Exportation excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_SINISTRES_BENEFICIAIRE_'+str(numero_carte)+'__{:%d:%m:%Y}.xlsx"'.format(timezone.now())

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'SINISTRES - {}'.format(numero_carte)

        # Write header row
        header = [
            'NUMERO_SINISTRE',
            'BENEFICIAIRE',
            'NUMERO_CARTE',
            'QUALITE',
            'ADHERENT_PRINCIPAL',
            'ACTE',
            'PRESTATAIRE',
            'FRAIS_REEL',
            'PART_ASSUREUR',
            'PART_BENEFICIAIRE',
            'DATE_PRESTATION',
            'DATE_SAISIE',
            'REFERENCE_FACTURE',
            'DATE_RECEPTION_FACTURE',
            'ETAT',
        ]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        data = []

        for sinistre in queryset:

            data_item = [
                sinistre.numero if sinistre.numero else "",
                beneficiaire,
                numero_carte if numero_carte else "",
                qualite,
                adherent_principal,
                sinistre.acte.libelle if sinistre.acte and sinistre.acte.libelle else "",
                sinistre.prestataire.name if sinistre.prestataire and sinistre.prestataire.name else "",
                sinistre.total_frais_reel,
                sinistre.total_part_compagnie if sinistre.part_compagnie else "",
                sinistre.total_part_assure if sinistre.part_assure else "",
                sinistre.date_survenance.strftime("%d/%m/%Y %H:%m") if sinistre.date_survenance else "",
                sinistre.created_at.strftime("%d/%m/%Y %H:%m") if sinistre.created_at else "",
                sinistre.reference_facture if sinistre.reference_facture else "",
                sinistre.date_reception_facture.strftime("%d/%m/%Y") if sinistre.date_reception_facture else "",
                sinistre.statut if sinistre.statut else "",
            ]
            data.append(data_item)

        for row_num, row in enumerate(data, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    else:
        return JsonResponse({
            "message": "Bénéficiaire non trouvé"
        }, status=404)


@login_required()
def export_sinistres_famille(request, police_id, aliment_id):
    police = Police.objects.get(id=police_id)
    #today = datetime.datetime.now(tz=timezone.utc)
    aliment = Aliment.objects.get(id=aliment_id)

    if aliment:
        #queryset = Aliment.objects.select_related('qualite_beneficiaire', 'police__client').filter(formulegarantie__police=police=police).order_by('-id')
        # Optimisation de la recuperation des beneficiaires
        #queryset = AlimentFormule.objects.filter(formule_id__in=[p.id for p in police.formules], statut=Statut.ACTIF).order_by('-id')

        adherent_principal = aliment.adherent_principal.nom + " " + aliment.adherent_principal.prenoms

        queryset = Sinistre.objects.filter(adherent_principal=aliment.adherent_principal, statut_validite=StatutValidite.VALIDE).order_by('-id')

        #dd(aliment.adherent_principal)

        print("queryset")
        print(queryset)
        print(queryset.count())

        # Exportation excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_SINISTRES_FAMILLE_'+str(aliment.adherent_principal.numero_famille)+'__'+str(police.numero)+'__{:%d:%m:%Y}.xlsx"'.format(timezone.now())

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'SINISTRES - {}'.format(aliment.adherent_principal.numero_famille)

        # Write header row
        header = [
            'NUMERO_SINISTRE',
            'BENEFICIAIRE',
            'NUMERO_CARTE',
            'QUALITE',
            'ADHERENT_PRINCIPAL',
            'ACTE',
            'PRESTATAIRE',
            'FRAIS_REEL',
            'PART_ASSUREUR',
            'PART_BENEFICIAIRE',
            'DATE_PRESTATION',
            'DATE_SAISIE',
            'REFERENCE_FACTURE',
            'DATE_RECEPTION_FACTURE',
            'ETAT',
        ]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        data = []

        for sinistre in queryset:
            numero_carte = sinistre.aliment.carte_active().numero if sinistre.aliment.carte_active() else ""
            beneficiaire = sinistre.aliment.nom + " " + sinistre.aliment.prenoms
            qualite = sinistre.aliment.qualite_beneficiaire.libelle if sinistre.aliment.qualite_beneficiaire else ''

            data_item = [
                sinistre.numero if sinistre.numero else "",
                beneficiaire,
                numero_carte if numero_carte else "",
                qualite,
                adherent_principal,
                sinistre.acte.libelle if sinistre.acte and sinistre.acte.libelle else "",
                sinistre.prestataire.name if sinistre.prestataire and sinistre.prestataire.name else "",
                sinistre.total_frais_reel,
                sinistre.total_part_compagnie if sinistre.part_compagnie else "",
                sinistre.total_part_assure if sinistre.part_assure else "",
                sinistre.date_survenance.strftime("%d/%m/%Y %H:%m") if sinistre.date_survenance else "",
                sinistre.created_at.strftime("%d/%m/%Y %H:%m") if sinistre.created_at else "",
                sinistre.reference_facture if sinistre.reference_facture else "",
                sinistre.date_reception_facture.strftime("%d/%m/%Y") if sinistre.date_reception_facture else "",
                sinistre.statut if sinistre.statut else "",
            ]
            data.append(data_item)

        for row_num, row in enumerate(data, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    else:
        return JsonResponse({
            "message": "Bénéficiaire non trouvé"
        }, status=404)


@login_required()
def export_sinistres_police(request, police_id):
    police = Police.objects.get(id=police_id)
    #today = datetime.datetime.now(tz=timezone.utc)

    if police:
        #queryset = Aliment.objects.select_related('qualite_beneficiaire', 'police__client').filter(formulegarantie__police=police=police).order_by('-id')
        # Optimisation de la recuperation des beneficiaires
        #queryset = AlimentFormule.objects.filter(formule_id__in=[p.id for p in police.formules], statut=Statut.ACTIF).order_by('-id')


        queryset = Sinistre.objects.filter(police=police, statut_validite=StatutValidite.VALIDE).order_by('-id')

        #dd(aliment.adherent_principal)

        print("queryset")
        print(queryset)
        print(queryset.count())

        # Exportation excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="LISTE_SINISTRES_POLICE_'+str(police.numero)+'__{:%d:%m:%Y}.xlsx"'.format(timezone.now())

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'SINISTRES - {}'.format(police.numero)

        # Write header row
        header = [
            'NUMERO_SINISTRE',
            'NUMERO_FEUILLE_SOIN',
            'BENEFICIAIRE',
            'NUMERO_CARTE',
            'QUALITE',
            'ADHERENT_PRINCIPAL',
            'ACTE',
            'PRESTATAIRE',
            'FRAIS_REEL',
            'PART_ASSUREUR',
            'PART_BENEFICIAIRE',
            'DATE_PRESTATION',
            'DATE_SAISIE',
            'REFERENCE_FACTURE',
            'DATE_RECEPTION_FACTURE',
            'ETAT',
        ]
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        data = []

        for sinistre in queryset:

            adherent_principal = sinistre.aliment.adherent_principal.nom + " " + sinistre.aliment.adherent_principal.prenoms
            numero_carte = sinistre.aliment.carte_active().numero if sinistre.aliment.carte_active() else ""
            beneficiaire = sinistre.aliment.nom + " " + sinistre.aliment.prenoms
            qualite = sinistre.aliment.qualite_beneficiaire.libelle if sinistre.aliment.qualite_beneficiaire else ''

            data_item = [
                sinistre.numero if sinistre.numero else "",
                sinistre.dossier_sinistre.numero if sinistre.dossier_sinistre and sinistre.dossier_sinistre.numero else "",
                beneficiaire,
                numero_carte if numero_carte else "",
                qualite,
                adherent_principal,
                sinistre.acte.libelle if sinistre.acte and sinistre.acte.libelle else "",
                sinistre.prestataire.name if sinistre.prestataire and sinistre.prestataire.name else "",
                sinistre.total_frais_reel,
                sinistre.total_part_compagnie if sinistre.part_compagnie else "",
                sinistre.total_part_assure if sinistre.part_assure else "",
                sinistre.date_survenance.strftime("%d/%m/%Y %H:%m") if sinistre.date_survenance else "",
                sinistre.created_at.strftime("%d/%m/%Y %H:%m") if sinistre.created_at else "",
                sinistre.reference_facture if sinistre.reference_facture else "",
                sinistre.date_reception_facture.strftime("%d/%m/%Y") if sinistre.date_reception_facture else "",
                sinistre.statut if sinistre.statut else "",
            ]
            data.append(data_item)

        for row_num, row in enumerate(data, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    else:
        return JsonResponse({
            "message": "Bénéficiaire non trouvé"
        }, status=404)

#new code
@login_required()
def export_prospect(request):
    mouvements = MouvementAliment.objects.filter(
        statut_traitement=StatutTraitement.NON_TRAITE,
        police__bureau=request.user.bureau
    ).order_by('aliment__nom').distinct()

    # Create an HttpResponse object with the appropriate content type and headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="LISTE DES MOUVEMENTS GRH.xlsx"'

    # Create a new Workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'LISTE DES MOUVEMENTS'

    # Define headers for the columns
    header = [
        'Date Mouvement', 'Client', 'Numéro Police', 'Formule', 'Nom du bénéficiaire', 'Prénoms du bénéficiaire',
        'Date Naissance', 'Sexe', 'Qualité', 'État'
    ]
    worksheet.append(header)

    # Iterate over each mouvement and add data rows to the worksheet
    for mouvement in mouvements:
        formule_libelle = ""
        etat_beneficiaire = mouvement.aliment.etat_beneficiaire if mouvement.aliment else ""

        if mouvement.aliment:
            aliment_formules = AlimentFormule.objects.filter(aliment=mouvement.aliment)
            if aliment_formules.exists():
                formulegarantie = aliment_formules.first()
                formule_libelle = formulegarantie.formule.libelle if formulegarantie else ""

        data_row = [
            mouvement.created_at.strftime("%d/%m/%Y %H:%M") if mouvement.created_at else "",
            f"{mouvement.police.client.nom}" if mouvement.police and mouvement.police.client else "",
            mouvement.police.numero if mouvement.police else "",
            formule_libelle,
            mouvement.aliment.nom if mouvement.aliment else "",
            mouvement.aliment.prenoms if mouvement.aliment else "",
            mouvement.aliment.date_naissance.strftime("%d/%m/%Y") if mouvement.aliment and mouvement.aliment.date_naissance else "",
            mouvement.aliment.genre if mouvement.aliment else "",
            mouvement.aliment.qualite_beneficiaire.libelle if mouvement.aliment and mouvement.aliment.qualite_beneficiaire else "",
            etat_beneficiaire,
        ]
        worksheet.append(data_row)

    # Save the workbook to the HttpResponse object
    workbook.save(response)

    return response
@method_decorator(login_required, name='dispatch')
class ProspectsGrhView(TemplateView):
    template_name = 'grh/liste_prospect.html'
    model = Prospect

    def get(self, request, *args, **kwargs):
        context_original = self.get_context_data(**kwargs)

        # Get police IDs based on the user's bureau and mouvement status
        police_ids = MouvementAliment.objects.filter(
            statut_traitement=StatutTraitement.NON_TRAITE,
            police__bureau=request.user.bureau
        ).values_list('police_id', flat=True)

        # Fetch clients based on the retrieved police_ids
        clients = Client.objects.filter(polices__id__in=police_ids).order_by('nom').distinct()

        # Initialize context variables
        liste_des_formules = []
        polices_du_bureau_actif = None
        adherents_principaux = None
        qualites_beneficiaires = []
        formules = []
        nom = []
        prenoms = []

        # Get the current date
        today = timezone.now().date()

        # Additional context specific to this view
        context_persos = {
            'adherents_principaux': adherents_principaux,
            'qualites_beneficiaires': qualites_beneficiaires,
            'formules': formules,
            'liste_des_formules': liste_des_formules,
            'polices_du_bureau_actif': polices_du_bureau_actif,
            'today': today,
            'clients': clients,
            'nom': nom,
            'prenoms': prenoms
        }

        # Merge original context with the custom context
        context = {**context_original, **context_persos}
        return self.render_to_response(context)

    def post(self):
        pass

    def get_context_data(self, **kwargs):
        # Add additional context and admin site data
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

def prospect_grh_datatable(request):
    items_per_page = 10
    page_number = request.GET.get('page')
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', items_per_page))
    sort_column_index = int(request.GET.get('order[0][column]'))
    sort_direction = request.GET.get('order[0][dir]')

    # Retrieve search filters from request
    search_client_id = request.GET.get('client', None)
    search_num_police = request.GET.get('num_police', None)
    search_date = request.GET.get('date', None)
    search_nom_benef = request.GET.get('nom_benef', None)

    # Filter police based on search criteria
    if search_client_id:
        police_ids = Police.objects.filter(bureau=request.user.bureau, client_id=search_client_id, statut_validite=StatutValidite.VALIDE).values_list('id')
    else:
        police_ids = Police.objects.filter(bureau=request.user.bureau, statut_validite=StatutValidite.VALIDE).values_list('id')

    if search_num_police:
        police_ids = Police.objects.filter(bureau=request.user.bureau, numero__contains=search_num_police, statut_validite=StatutValidite.VALIDE).values_list('id')

    # Fetch aliment movements that are not processed yet
    aliment_ids = MouvementAliment.objects.filter(statut_traitement=StatutTraitement.NON_TRAITE, police_id__in=police_ids).values_list('aliment_id')
    queryset = Aliment.objects.filter(id__in=aliment_ids)
    print(queryset)
    print("@@@@@@ TISSI")
    print(search_nom_benef)
    if search_nom_benef:
        queryset = queryset.filter(Q(nom__icontains=search_nom_benef) | Q(prenoms__icontains=search_nom_benef))

    if search_date:
        queryset = queryset.filter(created_at__contains=search_date)


    print("@@@@@ queryset @@@@@@@")
    print(queryset)

    # Map column index to corresponding model field for sorting
    sort_columns = {
        2: '-date',
        3: 'client',
        5: 'formule',
        6: 'nom_prospect',
        13: 'statut',
    }

    # Default sorting by 'id' if column index is not found
    sort_column = sort_columns.get(sort_column_index, 'id')
    if sort_direction == 'desc':
        sort_column = '-' + sort_column

    # Apply sorting
    # add condition to avoid list has no attribute order_by
    # if not request.user.is_med and not request.user.is_pharm:
    # queryset = queryset.order_by(sort_column)

    paginator = Paginator(queryset, length)
    page_obj = paginator.get_page(page_number)

    # Prepare the data for the response
    data = []
    for c in page_obj:
        # detail_url = "#" #reverse('details_dossier_sinistre', args=[c.id])  # URL to the detail view# URL to the detail view
        # actions_html = f'<a href="{detail_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-eye"></i> Détails</span></a>&nbsp;&nbsp;'
        details_prospect_url = reverse('details_prospect', args=[c.id])
        # update_prospect_url = reverse('update_prospect', args=[c.id])
        actions_html = f'<span style="cursor:pointer;" class="btn_details_prospect" data-prospect_id="{c.id}" data-model_name="prospect" data-modal_title="FICHE ADHERENT - {c.nom} {c.prenoms}" data-href="{details_prospect_url}"><span class="badge btn-sm btn-details rounded-pill"><i class="fa fa-edit"></i> Traiter</span></span>'

        statut_html = f'<span class="badge badge-{c.etat_beneficiaire.lower().replace(" ", "-")}">{c.etat_beneficiaire}</span>'

        aliment_formule = AlimentFormule.objects.filter(aliment=c).first()
        formule = aliment_formule.formule if aliment_formule else None
        police = formule.police if formule else None

        nom_client = police.client.nom if police and police.client.nom else ""
        prenom_client = police.client.prenoms if police and police.client.prenoms else ""

        # Fetch the timestamp when the GRH initiated the movement for validation
        grh_movement_initiation = c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else ""

        data_iten = {
            "id": c.id,
            "date": c.updated_at.strftime("%d/%m/%Y %H:%M") if c.updated_at else "",
            "client": f"{nom_client} {prenom_client}",
            "numero_police": police.numero if police else "",
            "formule": "",# c.formulegarantie.libelle if c.formulegarantie else "",
            "nom_prospect": c.nom,
            "prenom_prospect": c.prenoms,
            "date_naiss": c.date_naissance.strftime("%d/%m/%Y") if c.date_naissance else "",
            "sexe": c.genre,
            "qualite": c.qualite_beneficiaire.libelle if c.qualite_beneficiaire else "",
            "statut": statut_html,
            "grh_movement_initiation": grh_movement_initiation,  # Add the GRH initiation date and time
            "actions": actions_html,
        }

        data.append(data_iten)

    return JsonResponse({
        "data": data,
        "recordsTotal": queryset.count(),
        "recordsFiltered": paginator.count,
        "draw": int(request.GET.get('draw', 1)),
    })


# modification d'un bénéficiaire
def details_prospect(request, aliment_grh_id):
    aliment_grh = Aliment.objects.filter(id=aliment_grh_id).first()
    famille = Aliment.objects.filter(adherent_principal=aliment_grh.adherent_principal)

    # last() cause pour changement de formule (aliment a plusieurs formules)
    aliment_formule = AlimentFormule.objects.filter(aliment=aliment_grh).last()
    police = aliment_formule.formule.police if aliment_formule else None

    historique_formules = aliment_grh.historique_formules.filter(formule__police=police).order_by('date_debut')
    mouvements_aliments = MouvementAliment.objects.filter(aliment=aliment_grh, statut_validite=StatutValidite.VALIDE).order_by('id')

    return render(request, 'grh/modal_details_prospect.html',
                  {'prospect': aliment_grh, 'famille':famille, 'historique_formules':historique_formules, 'mouvements_aliments': mouvements_aliments, 'police':police})


@transaction.atomic
def submit_prospect(request, aliment_grh_id, police_id):
    aliment_grh = Aliment.objects.filter(id=aliment_grh_id).first()
    police = Police.objects.get(id=police_id)

    # prospect.statut_enrolement = StatutEnrolement.INCORPORE
    if aliment_grh is not None:

        prospect = Prospect.objects.filter(aliment=aliment_grh).first()
        if prospect:
            prospect.statut_enrolement = StatutEnrolement.INCORPORE
            prospect.save()

        campagne_prospect = CampagneProspect.objects.filter(prospect=prospect).first()
        if campagne_prospect:
            campagne_prospect.statut_enrolement = StatutEnrolement.INCORPORE
            campagne_prospect.save()

        if aliment_grh.etat_beneficiaire == "ENTREE EN COURS":

            # enregistrer la nouvelle carte
            carte = Carte.objects.create(
                aliment_id=aliment_grh.pk,
                date_edition=datetime.datetime.now(tz=timezone.utc),
                motif_edition="Nouvelle incorporation",
                statut=Statut.ACTIF
            )

            # METTRE A JOUR LE NUMERO
            prefixe = request.user.bureau.code
            numero_carte = generate_numero_carte(aliment_grh)
            carte.numero = numero_carte
            carte.save()

            # générer le qrcode
            qrcode_file = generer_qrcode_carte(numero_carte)
            print("qrcode_img")
            # print(qrcode_img)
            carte.qrcode_file.save(f'qrcode_img_{numero_carte}.png', qrcode_file)
            carte.save()

            # Créer le mouvement d'incorporatio du bénéficiaire
            mouvement = Mouvement.objects.filter(code="INCORPORATION").first()
            mouvement_aliment = MouvementAliment.objects.create(
                created_by=request.user,
                aliment=aliment_grh,
                mouvement=mouvement,
                police=police,
                date_effet=aliment_grh.date_affiliation,
                motif="Nouvelle incorporation",
                statut_validite=StatutValidite.VALIDE,
                statut_traitement=StatutTraitement.TRAITE
            )
            mouvement_aliment.save()

            #mettre à jour le statut_incorporation de l'aliment
            aliment_grh.statut_incorporation = StatutIncorporation.INCORPORE
            aliment_grh.save()

        elif aliment_grh.etat_beneficiaire == "SORTIE EN COURS":

            mouvement_aliment_demande_sortie = MouvementAliment.objects.filter(aliment=aliment_grh, police=police, mouvement__code="DMDSORTIE", statut_validite=StatutValidite.VALIDE, statut_traitement=StatutTraitement.NON_TRAITE).first()

            # Créer le mouvement de sortie
            mouvement = Mouvement.objects.filter(code="SORTIE-BENEF").first()
            mouvement_aliment = MouvementAliment.objects.create(
                created_by=request.user,
                aliment=aliment_grh,
                mouvement=mouvement,
                police=police,
                date_effet=datetime.datetime.now(),
                motif="Sortie demandé par le GRH et validé par le gestionnaire",
                statut_validite=StatutValidite.VALIDE,
                statut_traitement=StatutTraitement.TRAITE
            )
            mouvement_aliment.save()

            aliment_grh.date_sortie = mouvement_aliment_demande_sortie.date_effet
            aliment_grh.save()

        elif aliment_grh.etat_beneficiaire == "SUSPENSION EN COURS":
            # Créer le mouvement de suspension
            mouvement = Mouvement.objects.filter(code="SUSPENSION-BENEF").first()
            mouvement_aliment = MouvementAliment.objects.create(
                created_by=request.user,
                aliment=aliment_grh,
                mouvement=mouvement,
                police=police,
                date_effet=datetime.datetime.now(),
                motif="Suspension demandé par le GRH et validé par le gestionnaire",
                statut_validite=StatutValidite.VALIDE,
                statut_traitement=StatutTraitement.TRAITE
            )
            mouvement_aliment.save()

        #Marquer les mouvements en attente de l'aliment comme traités
        MouvementAliment.objects.filter(aliment=aliment_grh, statut_traitement=StatutTraitement.NON_TRAITE).update(statut_traitement=StatutTraitement.TRAITE)

        response = {
            'status': 1,
            'message': "Opération effectuée avec succès !",
            'data': {
            }
        }

        return JsonResponse(response)
    else:
        return JsonResponse({
            "status": 0,
            "message": "L'identifiant du prospect est introuvable."
        }, status=404)


def rejet_prospect(request, prospect_id):
    prospect = Prospect.objects.filter(id=prospect_id).first()

    if prospect is not None:
        prospect.statut_enrolement = StatutEnrolement.SOUMIS
        prospect.save()

        campagne_prospect = CampagneProspect.objects.filter(prospect=prospect).first()
        if campagne_prospect:
            campagne_prospect.statut_enrolement = StatutEnrolement.SOUMIS
            campagne_prospect.save()

        return JsonResponse({
            "status": 1,
            "message" : "Renvoi effectué avec succès !"
        })

    return JsonResponse({
        "status": 0,
        "message": "L'identifiant du prospect est introuvable."
    }, status=404)


#Annulation de quittance
@method_decorator(login_required, name='dispatch')
class AnnulerQuittanceView(TemplateView):
    template_name = 'police/annuler_quittance.html'
    model = Quittance

    # traitement à l'appel du lien en get
    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)

        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        return self.render_to_response(context)

    # traitement à l'appel du lien en post pour la recherche de dossier et la suppresion de dossier ou police
    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        # recuperation de tout ce qui peut venir en post que ca soit pour la recherche ou la suppression
        btn_recherche = self.request.POST.get('recherche', None)
        submit_delete_item = self.request.POST.get('submit_delete_item', None)
        id_item = self.request.POST.get('id_item', None)
        motif_delete_item = self.request.POST.get('motif_delete_item', None)
        numero_quittance = self.request.POST.get('numero_quittance', None)
        context['breadcrumbs'] = [
            {'title': 'Factures', 'url': ''},
            {'title': 'Annulation', 'url': ''},
        ]
        dossier_reglement = None

        # cette condition précise que nous venons faire la recherche
        if btn_recherche and numero_quittance:
            quittance = Quittance.objects.filter(numero=numero_quittance, bureau=request.user.bureau, statut=StatutQuittance.IMPAYE).first()
            # dd(quittance)

            context['numero_quittance'] = numero_quittance
            context['quittance'] = quittance

        # cette condition précise que nous venons faire l'annulation de la quittance
        if submit_delete_item and id_item:

            quittance = Quittance.objects.filter(id=id_item, bureau=request.user.bureau, statut=StatutQuittance.IMPAYE).first()

            if quittance:

                #récupérer les règlements sur la quittance
                reglements = Reglement.objects.filter(quittance=quittance, bureau=request.user.bureau)

                if reglements:
                    context['reglements_existants'] = True

                    '''
                    #TODO: Traitement à discuter et confirmer avec Marius
                    #vérifier s'il existe un règlement déjà reversé compagnie
                    if reglements.filter(statut_reversement_compagnie=StatutReversementCompagnie.REVERSE).first():
                        context['reverse_reglement'] = True
                        #Avoir l'accord de la finance (GILDAS)
    
                    else:
                        context['reglements_existants_annules'] = True
                        #Annuler les règlements sur la quittances
                        for reglement in reglements:
                            # traitement reglement
                            reglement.reg_deleted_by = request.user
                            reglement.statut_validite = StatutValidite.SUPPRIME
                            reglement.observation = motif_delete_item
                            #reglement.save #décommenter après
    
                            ActionLog.objects.create(done_by=request.user, action="annulation_reglement", description="Annulation d'un règlement", table="reglement", row=reglement.pk)
                            #
    
                        # Annuler la quittance
                        quittance.deleted_by = request.user
                        quittance.statut_validite = StatutValidite.SUPPRIME
                        quittance.observation = motif_delete_item
                        quittance.save()
                        context['old_quittance'] = quittance.numero
                        #
                    '''

                else:
                    context['reglements_existants'] = False

                    # Annuler la quittance
                    quittance.deleted_by = request.user
                    quittance.statut_validite = StatutValiditeQuittance.ANNULEE
                    quittance.observation = motif_delete_item
                    quittance.save()
                    context['old_quittance'] = quittance.numero
                    #

                    #Créer une ligne dans mouvement_quittances



                # enregistrer dans les log
                ActionLog.objects.create(done_by=request.user, action="annulation_quittance",
                                         description="Annulation d'une quittances", table="quittances",
                                         row=quittance.pk)

            # print(code_dossier_police)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        return {
            **super().get_context_data(**kwargs),
            **admin.site.each_context(self.request),
            "opts": self.model._meta,
        }

