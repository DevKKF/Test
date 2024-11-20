# api_utils.py
import base64
import decimal
import time
import uuid
from datetime import datetime, timezone, timedelta, date
import json
from pprint import pprint

from django.core.files.base import ContentFile
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
import openpyxl
import requests
from django_dump_die.middleware import dd

from django.contrib.auth.models import AbstractUser, Group
from comptabilite.models import EncaissementCommission
from configurations.models import AuthGroup, ChangementFormule, ComptePrestataireVeos, Devise, GroupeInter, NatureQuittance, PeriodeVeos, QuittanceVeos, TypeApporteur, TypeQuittance, \
    User, \
    PrescripteurVeos, PrestataireVeos, Pays, Profession, QualiteBeneficiaire, Bureau, TypePersonne, TypeClient, \
    Civilite, Langue, Prescripteur, Prestataire, Compagnie, Specialite, PrescripteurPrestataire, Acte, Affection, \
    TypePrestataire, \
    CronLog, TypeGarant, Territorialite, TypeTarif, Rubrique, Produit, ParamProduitCompagnie, CompagnieVeos, ClientVeos, \
    PoliceVeos, FormuleVeos, SinistreVeos, Secteur, TypePrefinancement, AlimentVeos, UtilisateurGrhVeos, \
    UtilisateurVeos, ApporteurVeos, \
    Apporteur, BaseCalcul, ApporteurInternational
from production.models import Aliment, Client, Mouvement, Police, MouvementPolice, Motif, FormuleGarantie, AlimentFormule, \
    PeriodeCouverture, Carte, Quittance, ApporteurPolice, Reglement
from shared.enum import OptionYesNo, Statut, StatutContrat, Etat, StatutQuittance, StatutReversementCompagnie, StatutValidite
from sinistre.models import Sinistre
from pathlib import Path
from environ import environ
from unidecode import unidecode

# Dﾃｩfinir l'URL de l'API en tant que variable globale
API_URL_PROD = "https://inov.prod-veos.iga.fr/rs/rsExtranet2"
API_URL_PREPROD = "https://inov.preprod-veos.iga.fr/rs/rsExtranet2"#--pre
API_URL_PREPROD_UPD_QUI = "https://inov.preprod-veos.iga.fr/rs"#--pre
API_URL_REPORTING_PREPROD_UPD_QUI = "https://inovreporting.preprod-veos.iga.fr/rs"#--pre
API_URL_PROD_UPD_QUI = "https://inov.prod-veos.iga.fr/rs"#--pre
API_URL_REPORTING_PREPROD = "https://inovreporting.preprod-veos.iga.fr/rs/rsExtranet2"#--pre


# Build paths inside the project like this: BASE_DIR / 'subdir'.
BASE_DIR = Path(__file__).resolve().parent.parent

def get_api_url():
    env = environ.Env()
    environ.Env.read_env(env_file=str(BASE_DIR / "inov" / ".env"))

    ENVIRONMENT = env("ENVIRONMENT")
    url = API_URL_PROD if ENVIRONMENT == "PROD" else API_URL_PREPROD
    return url


API_URL = get_api_url()


def getToken():
    try:
        api_url = f"{API_URL}/login"
        pprint(api_url)
        # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
        data = {
            "uid": "WS",
            "passwd": "WS1234"
        }

        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data)

        if response.status_code == 200:
            # Si la requﾃｪte a rﾃｩussi, vous pouvez accﾃｩder aux donnﾃｩes renvoyﾃｩes par l'API
            api_response_data = json.loads(response.content)
            token = api_response_data['token']
            return token
        return None

    except:
        return None


def getPoliceIdPol(police):

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_GET_LAST_POLICE_IDPOL",
                "params": {
                    'id_npol': police
                }
            }
        ]
    }

    api_response = call_api_with_data(data)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        response = api_response['responses'][0]['beans']

        id_pol = response[0]['ID_POL']

        return id_pol

    else:
        return None

#rﾃｩcuppﾃｩrer un token de connexion
BEARER_TOKEN = getToken()


def call_api_with_data(data):
    # Construire l'URL de l'API en utilisant la variable API_URL
    api_url = f"{API_URL}/boBy/list"
    error_message = "Impossible de rﾃｩcupﾃｩrer les donnﾃｩes sur VEOS."

    # dd(error_message)

    try:
        # refresh token
        BEARER_TOKEN = getToken()
        print(f"Token : {BEARER_TOKEN}")

        # Dﾃｩfinir les en-tﾃｪtes de la requﾃｪte avec le Bearer token
        headers = {
            "Authorization": f"Bearer {BEARER_TOKEN}",
            "Content-Type": "application/json"
        }

        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        # dd(response)
        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:
            
            # Si la requﾃｪte a rﾃｩussi, vous pouvez accﾃｩder aux donnﾃｩes renvoyﾃｩes par l'API
            api_response_data = response.content
            return json.loads(api_response_data)
        else:
            # dd(response)
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez None ou une valeur par dﾃｩfaut selon votre cas
            return JsonResponse({"error": error_message}, status=500)

    except AttributeError as e:

        print(f"Erreur AttributeError : {e}")
        # Code ﾃ exﾃｩcuter en cas d'AttributeError
        print(f"Erreur AttributeError : {e}")

        # En cas d'erreur lors de la requﾃｪte (par exemple, problﾃｨme de connexion)
        return None

def query_builder(boby, param, start_row, end_row):
    status = None
    api_url = f"{API_URL}/boBy/list"
    try:
        headers = {'Authorization': 'Bearer ' + BEARER_TOKEN}
        payload = {
            'requests': [
                {
                    'name': boby,
                    'params': param,
                    'doTotalCount': True,
                    'startRowNum': start_row,
                    'endRowNum': end_row,
                }
            ]
        }

        response = requests.post(api_url, json=payload, headers=headers)
        response_data = response.json()

        return response_data['responses'][0]
    except requests.RequestException as e:
        print(e)
        status = e.response.status_code if e.response else 404

    return status


def query_builder_tools(boby, param):
    data = []
    start_row = 1
    end_row = 3000
    loop_step = 3000

    try:
        query1 = query_builder(boby, param, start_row, end_row)
        data = query1.get('beans', [])
        while len(data) < query1['totalCount']:
            start_row = end_row + 1
            end_row = end_row + loop_step
            query2 = query_builder(boby, param, start_row, end_row)
            if 'beans' in query2:
                data.extend(query2['beans'])

        return data
    except Exception as e:
        # Handle the exception as needed
        pass


def get_document(ID_ADOC):
    # Construire l'URL de l'API en utilisant la variable API_URL
    api_url = f"{API_URL}/document/v2/{ID_ADOC}"
    error_message = "Impossible de rﾃｩcupﾃｩrer les donnﾃｩes sur VEOS."

    #dd(api_url)
    try:
        # Dﾃｩfinir les en-tﾃｪtes de la requﾃｪte avec le Bearer token
        headers = {
            "Authorization": f"Bearer {BEARER_TOKEN}",
            "Content-Type": "application/json"
        }

        # Appel de l'API en utilisant la mﾃｩthode GET avec les en-tﾃｪtes
        response = requests.get(api_url, headers=headers)
        if response.status_code == 200:
            # Si la requﾃｪte a rﾃｩussi, vous pouvez accﾃｩder aux donnﾃｩes renvoyﾃｩes par l'API
            api_response_data = response.json()

            return api_response_data

        else:
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez unempoﾃｨU! rﾃｩponse avec un message d'erreur et un statut 500
            return JsonResponse({"error": error_message}, status=500)


    except requests.exceptions.RequestException as e:
        # En cas d'erreur lors de la requﾃｪte (par exemple, problﾃｨme de connexion)
        print(f"An error occurred: {e}")
        return JsonResponse({"error": error_message}, status=500)



def create_sinistre_veos(data):
    # Construire l'URL de l'API en utilisant la variable API_URL
    # api_url = f"{API_URL}/sinistre"
    api_url = f"{API_URL_PREPROD}/sinistre"

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:

            # Si la requﾃｪte a rﾃｩussi, on renvoie un objet de donnees
            api_response_data = response.content
            return json.loads(api_response_data)

        else:
            # La rﾃｩponse ne correspond pas ﾃ ce qui est attendu
            raise ValueError("Erreur lors de l'envoie du sinistre sur VEOS")
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez None ou une valeur par dﾃｩfaut selon votre cas
            return JsonResponse({"error": ""}, status=500)

    except json.JSONDecodeError as e:
            # Une erreur s'est produite lors du chargement JSON
            raise ValueError(f"Erreur lors du chargement JSON : {e}")

    except Exception as e:
        raise ValueError(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}")




def get_pays_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_LISTE_TOTALE_PAYS",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        pays = api_response['responses'][0]['beans']

        for p in pays:
            # Check if the record already exists based on a unique field (code)
            existing_record = Pays.objects.filter(code=p["CODE"]).first()

            if existing_record:
                existing_record.libelle = p['NOM']
                existing_record.updated_at = datetime.now()
                existing_record.save()

            else:
                new = Pays(nom=p['NOM'], code=p['CODE'])
                new.save()

        return True



def get_bureaux_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_BUREAU_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    #dd(api_response)
    if api_response['statusCode'] == 0:

        bureaux = api_response['responses'][0]['beans']

        for bureau in bureaux:
            # Check if the record already exists based on a unique field (code)
            existing_record = Bureau.objects.filter(code=bureau["code"]).first()

            pays = Pays.objects.filter(code=bureau['pays'])
            p = pays.first() if pays else None

            devises = Devise.objects.filter(code=bureau['devise'])
            devise = devises.first() if pays else None

            if existing_record:
                existing_record.pays = p
                existing_record.devise = devise
                existing_record.nom = bureau['nom']
                existing_record.telephone = bureau['telephone']
                existing_record.email = bureau['email']
                existing_record.adresse = bureau['adresse']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(update_fields=['pays', 'devise', 'nom', 'telephone', 'email', 'adresse', 'updated_at'])

            else:

                new = Bureau(pays=p,
                         devise=devise,
                         code=bureau['code'],
                         nom=bureau['nom'],
                         telephone=bureau['telephone'],
                         email=bureau['email'],
                         adresse=bureau['adresse'])

                new.save()


        return True


def get_devise_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_DIVISE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    # dd(api_response)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        devises = api_response['responses'][0]['beans']

        for devise in devises:
            # Check if the record already exists based on a unique field (code)
            existing_record = Devise.objects.filter(code=devise["CODE"]).first()

            if existing_record:
                existing_record.libelle = devise['LIBELLE']
                existing_record.updated_at = datetime.now()
                existing_record.save(update_fields=['libelle', 'updated_at'])

            else:
                new = Devise(libelle=devise['LIBELLE'], code=devise['CODE'])
                new.save()

        return True


def get_compagnie_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_COMPAGNIE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    #dd(api_response)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        compagnies = api_response['responses'][0]['beans']

        for compagnie in compagnies:
            # Check if the record already exists based on a unique field (code)
            existing_record = Compagnie.objects.filter(code=compagnie["code"]).first()

            #type_garants = TypeGarant.objects.filter(code=compagnie['type_garant'])
            #type_garant = type_garants.first() if type_garants else None

            bureaux = Bureau.objects.filter(code=compagnie['bureau'])
            bureau = bureaux.first() if bureaux else None

            if existing_record:
                #existing_record.bureau = bureau,
                #existing_record.type_garant = type_garant,
                existing_record.nom = compagnie['nom']
                existing_record.telephone = compagnie['telephone']
                existing_record.email = compagnie['email']
                existing_record.adresse = compagnie['adresse']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(update_fields=['nom', 'code', 'telephone', 'email', 'adresse', 'updated_at'])

            else:

                if bureau:
                    new = Compagnie(bureau=bureau,
                                    #type_garant=type_garant,
                                    nom=compagnie['nom'],
                                    code=compagnie['code'],
                                    telephone=compagnie['telephone'],
                                    email=compagnie['email'],
                                    adresse=compagnie['adresse'],
                                    #taux_com_gestion=compagnie['taux_com_gestion'],
                                    #taux_com_courtage=compagnie['taux_com_courtage']
                                    )
                    new.save()
                else:
                    #dd(compagnie)
                    pass

        return True



def get_clients_lists():
    cpt_anciens_clients = 0
    cpt_nouveaux_clients = 0
    cpt_tous_clients = 0
    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    today = datetime.now(tz=timezone.utc).strftime('%d/%m/%Y')
    yesterday = (datetime.now(tz=timezone.utc) - timedelta(days=3)).strftime('%d/%m/%Y')

    data = {
        "requests": [
            {
                "name": "WS_CLIENT_TACHE_CRON_V2", #WS_CLIENT_TACHE_CRON
                "params": {
                    'date_debut': yesterday,
                    'date_fin': yesterday
                }
            }
        ]
    }
    #dd(data);
    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    #dd(api_response)
    if api_response['statusCode'] == 0:

        clients = api_response['responses'][0]['beans']

        #dd(clients)
        for client in clients:
            # Check if the record already exists based on a unique field (code)
            existing_record = Client.objects.filter(code=client["code"]).first()

            bureaux = Bureau.objects.filter(code=client['bureau'])
            bureau = bureaux.first() if bureaux else None
            
            payss = Pays.objects.filter(code=client['cd_pays'])
            pays = payss.first() if payss else bureau.pays if bureau else None

            types_personnes = TypePersonne.objects.filter(code=client['type_per'])
            type_personne = types_personnes.first() if types_personnes else None

            type_clients = TypeClient.objects.filter(code=client['bureau'])
            type_client = type_clients.first() if type_clients else None

            civilites = Civilite.objects.filter(code=client['civilite'])
            civilite = civilites.first() if civilites else None

            langues = Langue.objects.filter(code=client['bureau'])
            langue = langues.first() if langues else None

            # dd(existing_record)
            if existing_record:
                existing_record.bureau = bureau
                if type_personne: existing_record.type_personne = type_personne
                if type_client: existing_record.type_client = type_client
                #existing_record.pays = pays
                existing_record.code = client['code']
                existing_record.nom = client['nom']
                existing_record.prenoms = client['prenoms']
                existing_record.civilite = civilite
                #existing_record.sexe = ""
                existing_record.date_naissance = client['date_naissance'][:10] if client['date_naissance'] else None
                existing_record.telephone_mobile = client['telephone_mobile']
                existing_record.telephone_fixe = client['telephone_princ']
                existing_record.email = client['email']
                existing_record.langue = langue
                existing_record.ville = client['ville']
                existing_record.adresse_postale = ""
                existing_record.adresse = client['adresse']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(update_fields=['bureau', 'type_personne', 'type_client', 'pays', 'code', 'nom', 'prenoms', 'civilite', 'sexe', 'date_naissance', 'telephone_mobile', 'telephone_fixe', 'email', 'langue', 'ville', 'adresse_postale', 'adresse', 'updated_at'])

                cpt_anciens_clients += 1

            else:

                new = Client(
                    bureau=bureau,
                    type_personne=type_personne,
                    type_client=type_client,
                    pays=pays,
                    code=client['code'],
                    nom=client['nom'],
                    prenoms=client['prenoms'],
                    civilite=civilite,
                    sexe="",
                    date_naissance=client['date_naissance'][:10] if client['date_naissance'] else None,
                    telephone_mobile=client['telephone_mobile'],
                    telephone_fixe=client['telephone_princ'],
                    email=client['email'],
                    langue=langue,
                    longitude="",
                    latitude="",
                    ville=client['ville'],
                    adresse_postale="",
                    adresse=client['adresse'],
                    site_web="",
                    twitter="",
                    instagram="",
                    facebook="",
                    ancienne_ref="",
                    statut="ACTIF",
                    statut_relation="CLIENT",
                    )

                new.save()

                cpt_nouveaux_clients += 1;

            cpt_tous_clients += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciens_clients': cpt_anciens_clients,
                'cpt_nouveaux_clients': cpt_nouveaux_clients,
                'cpt_tous_clients': cpt_tous_clients
            }
        }
        return response

    #return True


def import_compagnie_manuellement(request):
    cpt_anciens_compagnies = 0
    cpt_nouveaux_compagnies = 0
    cpt_tous_compagnies = 0

    compagnies_veos = CompagnieVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_PER')[:1000]

    #dd(compagnies_veos)
    if compagnies_veos:

        for compagnie in compagnies_veos:
            # Check if the record already exists based on a unique field (code)
            existing_record = Compagnie.objects.filter(code=compagnie.CODE).first()

            #type_garants = TypeGarant.objects.filter(code=compagnie['type_garant'])
            #type_garant = type_garants.first() if type_garants else None

            bureaux = Bureau.objects.filter(code=compagnie.BUREAU)
            bureau = bureaux.first() if bureaux else None

            if bureau:
                if existing_record:
                    #existing_record.bureau = bureau,
                    #existing_record.type_garant = type_garant,
                    existing_record.nom = compagnie.NOM
                    existing_record.telephone = compagnie.TELEPHONE
                    existing_record.email = compagnie.EMAIL
                    existing_record.adresse = compagnie.ADRESSE
                    existing_record.updated_at = datetime.now(tz=timezone.utc)
                    existing_record.save(update_fields=['nom', 'code', 'telephone', 'email', 'adresse', 'updated_at'])


                    #Added on 15122023: paramﾃｩtrer les taux de com, si ﾃｧa changﾃｩ
                    produit = Produit.objects.get(code="100991")

                    param_existant = ParamProduitCompagnie.objects.filter(compagnie=existing_record, produit=produit, status=True).first()
                    if not param_existant:
                        param = ParamProduitCompagnie.objects.create(compagnie=existing_record,
                                                                     produit=produit,
                                                                     taux_com_courtage=compagnie.COM_APPORT_COMPTANT,
                                                                     taux_com_courtage_terme=compagnie.COM_APPORT_TERME,
                                                                     taux_com_gestion=compagnie.COM_GESTION,
                                                                     )

                        param.save()

                    else:

                        # vﾃｩrifier si les taux ont changﾃｩ, dﾃｩsactiver l'ancienne et crﾃｩer une nouvelle ligne
                        if param_existant.taux_com_courtage != compagnie.COM_APPORT_COMPTANT or param_existant.taux_com_courtage_terme != compagnie.COM_APPORT_TERME or param_existant.taux_com_gestion != compagnie.COM_GESTION:

                            #dﾃｩsactiver la ligne existante
                            param_existant.status = False
                            param_existant.save()

                            #insﾃｩrer la nouvelle ligne
                            param = ParamProduitCompagnie.objects.create(compagnie=existing_record,
                                                                         produit=produit,
                                                                         taux_com_courtage=compagnie.COM_APPORT_COMPTANT,
                                                                         taux_com_courtage_terme=compagnie.COM_APPORT_TERME,
                                                                         taux_com_gestion=compagnie.COM_GESTION,
                                                                         )

                            param.save()

                        else:
                            pprint("Les taux paramﾃｩtrﾃｩs n'ont pas changﾃｩ")


                    '''
                    COM_APPORT_COMPTANT = models.CharField(max_length=100, null=True)
                    COM_APPORT_TERME = models.CharField(max_length=100, null=True)
                    COM_GESTION = models.CharField(max_length=100, null=True)
                    '''

                    cpt_anciens_compagnies += 1

                else:

                    new_compagnie = Compagnie(bureau=bureau,
                                    #type_garant=type_garant,
                                    nom=compagnie.NOM,
                                    code=compagnie.CODE,
                                    telephone=compagnie.TELEPHONE,
                                    email=compagnie.EMAIL,
                                    adresse=compagnie.ADRESSE,
                                    #taux_com_gestion=compagnie['taux_com_gestion'],
                                    #taux_com_courtage=compagnie['taux_com_courtage']
                                    )
                    new_compagnie.save()

                    # Added on 15122023: paramﾃｩtrer les taux de com
                    produit = Produit.objects.get(code="100991")
                    param = ParamProduitCompagnie.objects.create(compagnie=new_compagnie,
                                                                 produit=produit,
                                                                 taux_com_courtage=compagnie.COM_APPORT_COMPTANT,
                                                                 taux_com_courtage_terme=compagnie.COM_APPORT_TERME,
                                                                 taux_com_gestion=compagnie.COM_GESTION,
                                                                 )

                    param.save()

                    cpt_nouveaux_compagnies += 1

                # le marquﾃｩ comme importﾃｩ
                compagnie.STATUT_IMPORT = True
                compagnie.save()

                cpt_tous_compagnies += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciens_compagnies': cpt_anciens_compagnies,
                'cpt_nouveaux_compagnies': cpt_nouveaux_compagnies,
                'cpt_tous_compagnies': cpt_tous_compagnies
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)


def import_client_manuellement(request):
    cpt_anciens_clients = 0
    cpt_nouveaux_clients = 0
    cpt_tous_clients = 0

    clients_veos = ClientVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_PER')[:1000]
    cpt = 0

    #dd(clients_veos)
    if clients_veos:

        for cv in clients_veos:

            # Check if the record already exists based on a unique field (code)
            #existing_record = Client.objects.filter(veos_client_idper=cv.ID_PER).first()
            # Check if the record already exists based on a unique field (code)
            existing_record = Client.objects.filter(code=cv.CODE).first()

            bureaux = Bureau.objects.filter(code=cv.BUREAU)
            bureau = bureaux.first() if bureaux else None

            payss = Pays.objects.filter(code=cv.PAYS)
            pays = payss.first() if payss else bureau.pays if bureau else None

            types_personnes = TypePersonne.objects.filter(code=cv.TYPE_PER)
            type_personne = types_personnes.first() if types_personnes else None

            #type_clients = TypeClient.objects.filter(code=cv.TYPE_CLIENT) #A ajouter ﾃ la requete
            #type_client = type_clients.first() if type_clients else None
            type_client = None

            #civilites = Civilite.objects.filter(code=cv.CIVILITE) #A ajouter ﾃ la requete
            #civilite = civilites.first() if civilites else None
            civilite = None

            langues = Langue.objects.filter(code=cv.LANG)
            langue = langues.first() if langues else None

            # dd(existing_record)
            if existing_record:
                existing_record.bureau = bureau
                if type_personne: existing_record.type_personne = type_personne
                if type_client: existing_record.type_client = type_client
                # existing_record.pays = pays
                existing_record.code = cv.CODE
                existing_record.nom = cv.NOM
                existing_record.prenoms = cv.PRENOMS
                existing_record.civilite = civilite
                # existing_record.sexe = ""
                existing_record.date_naissance = cv.DATE_NAISSANCE[:10] if cv.DATE_NAISSANCE else None
                existing_record.telephone_mobile = cv.TELEPHONE_MOBILE
                existing_record.telephone_fixe = cv.TELEPHONE_FIXE
                existing_record.email = cv.EMAIL
                existing_record.langue = langue
                existing_record.ville = cv.VILLE
                existing_record.adresse_postale = ""
                existing_record.adresse = cv.ADRESSE
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(
                    update_fields=['bureau', 'type_personne', 'type_client', 'pays', 'code', 'nom', 'prenoms',
                                   'civilite', 'sexe', 'date_naissance', 'telephone_mobile', 'telephone_fixe', 'email',
                                   'langue', 'ville', 'adresse_postale', 'adresse', 'updated_at'])

                cpt_anciens_clients += 1

            else:

                new = Client(
                    veos_client_idper=cv.ID_PER,
                    bureau=bureau,
                    type_personne=type_personne,
                    type_client=type_client,
                    pays=pays,
                    code=cv.CODE,
                    nom=cv.NOM,
                    prenoms=cv.PRENOMS,
                    civilite=civilite,
                    sexe="",
                    date_naissance=cv.DATE_NAISSANCE[:10] if cv.DATE_NAISSANCE else None,
                    telephone_mobile=cv.TELEPHONE_MOBILE,
                    telephone_fixe=cv.TELEPHONE_FIXE,
                    email=cv.EMAIL,
                    langue=langue,
                    longitude="",
                    latitude="",
                    ville=cv.VILLE,
                    adresse_postale="",
                    adresse=cv.ADRESSE,
                    site_web="",
                    twitter="",
                    instagram="",
                    facebook="",
                    ancienne_ref="",
                    statut="ACTIF",
                    statut_relation="CLIENT",
                )

                new.save()

                cpt_nouveaux_clients += 1;

            cpt_tous_clients += 1


            # le marquﾃｩ comme importﾃｩ
            cv.STATUT_IMPORT = True
            cv.save()


        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciens_clients': cpt_anciens_clients,
                'cpt_nouveaux_clients': cpt_nouveaux_clients,
                'cpt_tous_clients': cpt_tous_clients
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)




def cron_create_police(data):

    api_url = f"{API_URL_PREPROD}/police"

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:

            # Si la requﾃｪte a rﾃｩussi, on renvoie un objet de donnees
            api_response_data = response.content
            return json.loads(api_response_data)

        else:
            # La rﾃｩponse ne correspond pas ﾃ ce qui est attendu
            raise ValueError("Erreur lors de l'envoie de la police sur VEOS")

    except json.JSONDecodeError as e:
            # Une erreur s'est produite lors du chargement JSON
            raise ValueError(f"Erreur lors du chargement JSON : {e}")

    except Exception as e:
        raise ValueError(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}")



def get_polices_lists():

    cpt_anciennes_polices = 0
    cpt_nouvelles_polices = 0
    cpt_toutes_polices=0

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    today = datetime.now(tz=timezone.utc).strftime('%d/%m/%Y')
    yesterday = (datetime.now(tz=timezone.utc) - timedelta(days=3)).strftime('%d/%m/%Y')
    jours_10 = (datetime.now(tz=timezone.utc) - timedelta(days=30)).strftime('%d/%m/%Y')
    jours_7 = (datetime.now(tz=timezone.utc) - timedelta(days=7)).strftime('%d/%m/%Y')

    data = {
        "requests": [
            {
                "name": "WS_POLICE_TACHE_CRON",
                "params": {
                    #'num_soc': 'CI01', #quand les autres pays vont dﾃｩmarrer, faire une boucler sur les
                    'date_debut': yesterday,
                    'date_fin': today
                }
            }
        ]
    }

    #dd(data)

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    #dd(api_response)
    if api_response and api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        polices = api_response['responses'][0]['beans']

        # dd(polices)
        for police in polices:

            #infos pour crﾃｩer la nouvelle police
            bureaux = Bureau.objects.filter(code=police['BUREAU_ID'])
            bureau = bureaux.first() if bureaux else None

            clients = Client.objects.filter(code=police['CLIENT_ID'])
            client = clients.first() if clients else None

            compagnies = Compagnie.objects.filter(code=police['COMPAGNIE_ID'])
            compagnie = compagnies.first() if compagnies else None

            type_prefinancements = TypePrefinancement.objects.filter(code=police['PEF_TMOD'])
            type_prefinancement = type_prefinancements.first() if type_prefinancements else None

            date_debut_effet = police['DATE_DEBUT_EFFET'][:10] if police['DATE_DEBUT_EFFET'] else "",
            date_fin_effet = police['DATE_FIN_EFFET'][:10] if police['DATE_FIN_EFFET'] else "",

            date_debut_effet = datetime.strptime(date_debut_effet[0], '%Y-%m-%d').date() if date_debut_effet[0] else None
            date_fin_effet = datetime.strptime(date_fin_effet[0], '%Y-%m-%d').date() if date_fin_effet[0] else None

            motifs = Motif.objects.filter(code=police['MOTIF'])
            motif = motifs.first() if motifs else None

            mouvement = motif.mouvement if motif else None

            statut = Statut.ACTIF if police['STATUS'] == 'EC' else Statut.INACTIF

            devise = bureau.pays.devise if bureau else None

            #dd(client)

            # Check if the record already exists based on a unique field (code)
            existing_record = Police.objects.filter(veos_id_npol=police["ID_NPOL"]).first()

            # dd(existing_record)
            if existing_record:

                #autres traitements ﾃ effectuer

                #si la date a changﾃｩ, mettre ﾃ jour dans mouvement_police et periode couverture
                if mouvement:

                    #si le motif est renouv et que les dates ont changﾃｩ
                    if motif.code == "RENOUV" and date_debut_effet != existing_record.date_debut_effet:

                        # crﾃｩer une ligne dans mouvement_police
                        new_mouvement = MouvementPolice(
                            police=existing_record,
                            mouvement_id=mouvement.id,
                            motif=motif,
                            date_effet=date_debut_effet,
                            date_fin_periode_garantie=date_fin_effet,
                        )
                        new_mouvement.save()
                        # dd(new_mouvement)

                        #cloturer les autres pﾃｩriodes en cours
                        PeriodeCouverture.objects.filter(police=existing_record).update(statut=Statut.INACTIF)

                        # crﾃｩer une nouvelle periode de couverture
                        new_periode_couverture = PeriodeCouverture(
                            police=existing_record,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                        )
                        new_periode_couverture.save()


                #Mettre ﾃ jour les infos de la police
                existing_record.updated_at = datetime.now()
                existing_record.veos_code_cie = police['COMPAGNIE_ID']
                #existing_record.compagnie = compagnie
                existing_record.numero = police['NUMERO']
                existing_record.numero_provisoire = police['NUMERO_PROVISOIRE']
                existing_record.prime_ttc = decimal.Decimal(police['PRIME_TTC']) if police['PRIME_TTC'] else 0
                existing_record.prime_net = decimal.Decimal(police['PNET_POL']) if police['PNET_POL'] else 0
                existing_record.prime_ht = decimal.Decimal(police['PNET_POL']) if police['PNET_POL'] else 0
                existing_record.date_debut_effet = date_debut_effet
                if date_fin_effet: existing_record.date_fin_effet = date_fin_effet
                existing_record.statut = statut
                existing_record.observation = "UPDATED BY CRON - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S')) + " statut= " + str(police['STATUS'])
                existing_record.save()

                cpt_anciennes_polices += 1

            else:
                try:
                    #dd(police)
                    #Crﾃｩer la police

                    if client and compagnie and bureau:

                        new_police = Police(
                            type_assurance_id=1,
                            produit_id=1,
                            veos_code_client=police['CLIENT_ID'],
                            veos_code_cie=police['COMPAGNIE_ID'],
                            veos_id_npol=police['ID_NPOL'],
                            veos_id_pol=police['ID_POL'],
                            devise=devise,
                            bureau=bureau,
                            client=client,
                            compagnie=compagnie,
                            numero=police['NUMERO'],
                            numero_provisoire=police['NUMERO_PROVISOIRE'],
                            prime_ttc=decimal.Decimal(police['PRIME_TTC']) if police['PRIME_TTC'] else 0,
                            prime_net=decimal.Decimal(police['PNET_POL']) if police['PNET_POL'] else 0,
                            prime_ht=decimal.Decimal(police['PNET_POL']) if police['PNET_POL'] else 0,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                            statut_contrat=StatutContrat.CONTRAT,
                            statut=statut,
                            type_prefinancement=type_prefinancement,
                            observation="CREATED BY CRON - " + str(datetime.now(tz=timezone.utc)),
                            )

                        new_police.save()
                        #dd(new_police)

                        #crﾃｩer une ligne dans mouvement_police
                        new_mouvement = MouvementPolice(
                            police=new_police,
                            mouvement_id=mouvement.id,
                            motif=motif,
                            date_effet=date_debut_effet,
                            date_fin_periode_garantie=date_fin_effet,
                            )
                        new_mouvement.save()
                        #dd(new_mouvement)

                        #crﾃｩer la periode de couverture
                        new_periode_couverture = PeriodeCouverture(
                            police=new_police,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                            )
                        new_periode_couverture.save()


                        cpt_nouvelles_polices += 1


                except Exception as e:
                    error_message = str(e)  # Store the exception message
                    result = None


            cpt_toutes_polices += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciennes_polices': cpt_anciennes_polices,
                'cpt_nouvelles_polices': cpt_nouvelles_polices,
                'cpt_toutes_polices': cpt_toutes_polices
            }
        }

        return response



def create_police_veos(data):
    # Construire l'URL de l'API en utilisant la variable API_URL
    # api_url = f"{API_URL}/sinistre"
    api_url = f"{API_URL_PREPROD}/sinistre"

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:

            # Si la requﾃｪte a rﾃｩussi, on renvoie un objet de donnees
            api_response_data = response.content
            return json.loads(api_response_data)

        else:
            # La rﾃｩponse ne correspond pas ﾃ ce qui est attendu
            raise ValueError("Erreur lors de l'envoie du sinistre sur VEOS")
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez None ou une valeur par dﾃｩfaut selon votre cas
            return JsonResponse({"error": ""}, status=500)

    except json.JSONDecodeError as e:
            # Une erreur s'est produite lors du chargement JSON
            raise ValueError(f"Erreur lors du chargement JSON : {e}")

    except Exception as e:
        raise ValueError(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}")




#send sinistres to veos
def send_polices():
    # Liste des polices ﾃ envoyer
    polices = Police.objects.filter(id=26)
    nombre_police = polices.count()

    # Dﾃｩmarrage du compte ﾃ rebourt pour la crﾃｩation des polices sur veos
    temps_debut = time.time()


    for police in polices:

        date_str = police.created_at.strftime("%Y-%m-%d %H:%M:%S")
        date_format = "%Y-%m-%d %H:%M:%S"

        timestamp = datetime.strptime(date_str, date_format).timestamp()

        dd(police)

        data = {
            "numPol": police.numero,
            "compagnie": police.compagnie.code,
            "societe": police.bureau.code,
            "dtEffet": police.date_debut_effet,
            "duree": police.duree_police_en_mois,
            "ttc": police.prime_ttc,
            "encaissement": "1",
            "cdProd": "100991",
            "cdFrac": "M",
            "devise": "XOF",
            "mouvement": "AN",
            "modeReglement": "CHQ",
            "txCom": 23,
            "typeFrais": None,
            "echeance": "15\/02\/2025",
            "primeNet": 12000000,
            "tacit": "SR",
            "type": "C",
            "personnes": {
                "souscripteur": {
                    "id": 11087459
                },
                "listBenefs": [
                    {
                        "id": "11090150",
                        "categPerpolice": "UTIL SOC",
                        "cdConPerpolice": "GEST"
                    }
                ]
            },
            "motif": "AN",
            "cdPos": "EC",
            "dtPrEch": "15\/02\/2025",
            "cdRegul": "1",
            "infos": {
                "MAXAGEAD": 50,
                "MAXAGEENF": 18,
                "MAXENF": 5,
                "MAXPERFAM": 7,
                "MODECALC": "",
                "PREFINTM": "AVCIE ded",
                "TYPEMAJO": "",
                "COM_POL": 2760000,
                "DTTXCOM_POL": 18,
                "GESTBIND": 600000,
                "COUT_POLICE": 50000,+
                "VALTXCOM": 2760000,
                "APPORTEURYN": "0",
                "BRANCHEGA": "0249",
                "GESTSIN": "FALI",
                "GESTPROD": "RBANO",
                "CLASSEAXA": "01",
                "PARTICIPATIONTAUX": "",
                "PARTICIPATIONYN": "0",
                "PLACEGESTION": "0",
                "DATE_PREAVIS": "2 MOIS",
                "PROG_INTER": "0",
                "AUTOFI": "0",
                "DTTXGESTBIND_POL": 5,
                "TXCOM_POL": 2160000,
                "BENEF_REMB": "",
                "PREFINDEP": "",
                "REFPOLCIE": "",
                "PREFIN_DOSSIER": "",
                "COUT_POLFRAC": "0",
                "PROCH_TERM": "0",
                "CALCULTM": "SURPLAFOND",
                "DELAIREMB": "",
                "DELAI_DECLA": "",
                "BLOCAGE_PRESC": ""
            },
            "dtFin": "14\/02\/2025",
            "valIndice": "1",
            "cdIndice": "SSI"
        }

        try:
            api_response = create_police_veos(data)
            CronLog.objects.create(action="create_police", table="police", description="Create police Nﾂｰ " + police.numero + " on veos with id = " + str(api_response['id']) + " et numero = " + str(api_response['num'])).save()
            print("笨 === Police envoyﾃｩ avec succﾃｨs === :", api_response)

        except ValueError as e:
            response = []
            CronLog.objects.create(action="create_police", table="police",
                                   description="Error when creating police Nﾂｰ " + police.numero).save()
            print("Erreur :", e)

    temps_fin = time.time()
    temps_execution = temps_fin - temps_debut

    message = f"沐 Envoie des ({nombre_police}) polices terminﾃｩs en {temps_execution:.7f} sﾃｩcondes. 沐"

    pprint(message)
    return message



def call_api_updt_mvquittances_veos(data):
    # Construire l'URL de l'API en utilisant la variable API_URL
    # api_url = f"{API_URL}/sinistre"
    api_url = f"{API_URL_PROD_UPD_QUI}/rsUpdateBoByTrs/json"

    # dd(data)

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        #return JsonResponse(response, safe=False)

        # return json.loads(response.content)

        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:

            # Si la requﾃｪte a rﾃｩussi, on renvoie un objet de donnees
            api_response_data = response.content
    
            # return JsonResponse({"content": response.content}, status=response.status_code)
            return json.loads(api_response_data)

        else:
            response = {
            'statut': 1,
            'message': 'TRAITEMENT ERREUR',
            'data': {
                'description': str(e)
                }
            }
            return response            
            # La rﾃｩponse ne correspond pas ﾃ ce qui est attendu
            return JsonResponse({"error": "Erreur lors de l'envoie du sinistre sur VEOS"}, status=500, safe=False)
            raise ValueError("Erreur lors de l'envoie du sinistre sur VEOS")
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez None ou une valeur par dﾃｩfaut selon votre cas
            return JsonResponse({"error": ""}, status=500)

    except json.JSONDecodeError as e:
        # Une erreur s'est produite lors du chargement 
        response = {
        'statut': 1,
        'message': 'TRAITEMENT ERREUR',
        'data': {
            'description': str(e)
            }
        }
        return response
        return JsonResponse(f"Erreur lors du chargement JSON : {e}", safe=False)
        raise ValueError(f"Erreur lors du chargement JSON : {e}")

    except Exception as e:
        response = {
        'statut': 1,
        'message': 'TRAITEMENT ERREUR',
        'data': {
            'description': str(e)
            }
        }
        return response        
        return JsonResponse(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}", safe=False)
        raise ValueError(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}")

#send quittances to veos
def updt_mvquittances():
    
    env = environ.Env()
    environ.Env.read_env(env_file=str(BASE_DIR / "inov" / ".env"))

    ENVIRONMENT = env("ENVIRONMENT")

    # if ENVIRONMENT:
    if ENVIRONMENT == "PROD":
        # Liste des quttances ﾃ envoyer
        quittances = Quittance.objects.filter(statut_validite=StatutValidite.VALIDE, import_stats=False)
        nombre_quittance = quittances.count()
        # Dﾃｩmarrage du compte ﾃ rebourt pour la crﾃｩation des quittances sur veos
        temps_debut = time.time()

        # print(quittances)
        # dd(quittances.count())

        data = {
            "uid": "WS",
            "passwd": "WS1234",
            "requests": []
        }
        
        devise = None
        defaut_taux_euro = None
        defaut_taux_usd = None

        counter = 0
        index = 0
        total_quittances = len(quittances)
        data_num_qui = []

        for quittance in quittances:

            index += 1 

            date_str = quittance.created_at.strftime("%Y-%m-%d %H:%M:%S")
            date_format = "%Y-%m-%d %H:%M:%S"

            timestamp = datetime.strptime(date_str, date_format).timestamp()

            # dd(quittance.police)
            
            apporteur_police = ApporteurPolice.objects.filter(police=quittance.police).first()
            mouvement_polices = MouvementPolice.objects.filter(police=quittance.police, statut_validite=StatutValidite.VALIDE).order_by("-date_effet")
            
            last_reglement = Reglement.objects.filter(quittance=quittance).order_by('-id').first()
            date_reglement_client = last_reglement.date_paiement if last_reglement else ''

            last_reglement_reverse_cie = Reglement.objects.filter(quittance=quittance,
                                                                statut_reversement_compagnie=StatutReversementCompagnie.REVERSE).order_by(
                '-date_reversement_compagnie').first()
            date_reversement_compagnie = last_reglement_reverse_cie.date_reversement_compagnie if last_reglement_reverse_cie else ''

            last_reglement = Reglement.objects.filter(quittance=quittance).order_by('-id').first()
            last_com_encaisse = None
            if last_reglement:
                last_com_encaisse = EncaissementCommission.objects.filter(reglement_id=last_reglement.id).order_by('-id').first()
            date_encaissement_commission = last_com_encaisse.created_at if last_com_encaisse else ''

            reglements_quittance = Reglement.objects.filter(quittance=quittance)
            com_court_encaissee = 0
            com_gest_encaissee = 0
            retro_paye = 0

            code_nature_quittance = ""
            if quittance.nature_quittance:
                if quittance.nature_quittance.code == "AA" or "Acompte":
                    code_nature_quittance = "AA"
                elif quittance.nature_quittance.code == "COMPT" or "CT":
                    code_nature_quittance = "COMPT"
                elif quittance.nature_quittance.code == "TERME" or "Terme":
                    code_nature_quittance = "TERME"
                elif quittance.nature_quittance.code == "R" or "Ristourne":
                    code_nature_quittance = "R"
                elif quittance.nature_quittance.code:
                    code_nature_quittance = quittance.nature_quittance.code


            for r in reglements_quittance:
                com_court_encaissee += float(r.montant_com_courtage_encaisse())
                com_gest_encaissee += float(r.montant_com_gestion_encaisse())
                retro_paye += float(r.montant_com_intermediaire)

            com_gest_non_encaissee = quittance.commission_gestion - com_gest_encaissee
            com_court_non_encaissee = quittance.commission_courtage - com_court_encaissee

            if devise != quittance.devise:
                devise = quittance.devise
                defaut_taux_euro = get_taux_euro_by_devise(devise.code)
                defaut_taux_usd = get_taux_usd_by_devise(devise.code)
                
            taux_euro = quittance.taux_euro if quittance.taux_euro else defaut_taux_euro
            taux_usd = quittance.taux_usd if quittance.taux_usd else defaut_taux_usd

            #convert_montant_from_devise_to_euro_or_usd(devise_from, montant, euro_or_usd, taux_euro_or_usd=None)
            #data_num_qui.append(quittance.numero)
            data["requests"].append(
                {
                    "name": "UPDT_MVQUITTANCES",
                    "params": {
                        "CD_PAYS": quittance.bureau.pays.code,
                        "LIB_PAYS": unidecode(quittance.bureau.pays.nom.upper()),
                        "LIB_PAYS_EN": unidecode(quittance.bureau.pays.nom.upper()),
                        "NUM_SOC": quittance.bureau.code,
                        "LIB_SOC": quittance.bureau.nom,
                        "GRINTER": quittance.police.client.groupe_international.code if quittance.police.client.groupe_international else "",
                        "GROUPE": quittance.police.client.groupe_international.nom if quittance.police.client.groupe_international else "",
                        "NUM_PER": quittance.police.client.code,
                        "NOM_PER": str(quittance.police.client.nom) + " " + str(quittance.police.client.prenoms if quittance.police.client.prenoms else ""),
                        "NOM2_PER": str(quittance.police.client.nom) + " " + str(quittance.police.client.prenoms if quittance.police.client.prenoms else ""),
                        "ID_PER": quittance.police.client.veos_client_idper,
                        "VILLE": quittance.police.client.ville,
                        "PAYS": quittance.police.client.pays.nom,
                        "PAYS_EN": quittance.police.client.pays.nom,
                        "CDPAYS_CLI": quittance.police.client.pays.code,
                        "ID_PER2": quittance.police.client.veos_client_idper,
                        "ID_NPOL": quittance.police.veos_id_npol,
                        "CD_FBRANCHE": "10",
                        "FBRANCHE": "ASSURANCE DE PERSONNE",
                        "FBRANCHE_EN": "ASSURANCE DE PERSONNE",
                        "LIB_BR": quittance.police.produit.branche.nom if hasattr(quittance.police.produit, 'branche') else 'SANTE',
                        "CD_BR": "10001",
                        "BRANCHE": quittance.police.produit.branche.nom if hasattr(quittance.police.produit, 'branche') else 'SANTE',
                        "BRANCHE_EN": quittance.police.produit.branche.nom if hasattr(quittance.police.produit, 'branche') else 'SANTE',
                        "PROGINT": "1" if quittance.police.programme_international == OptionYesNo.OUI else "0", #0 ou 1
                        "ASSUREUR_GRP": quittance.compagnie.groupe_compagnie.nom if quittance.compagnie.groupe_compagnie else "",
                        "CD_CIE": quittance.compagnie.code,
                        "ASSUREUR": quittance.compagnie.nom,
                        "NUMPOLICE": quittance.police.numero,
                        "CD_POS": "EC",
                        "DTEFF_POL": quittance.police.date_debut_effet.strftime("%d/%m/%Y") if quittance.police.date_debut_effet else "",
                        "DTFIN_POL": quittance.police.date_fin_police.strftime("%d/%m/%Y") if quittance.police.date_fin_police else "",
                        "ECHP_POL": quittance.police.date_prochaine_facture.strftime("%d/%m/%Y") if quittance.police.date_prochaine_facture else "",
                        "TTC_POL_SAISIE": str(quittance.police.prime_ttc),
                        "TTC_POL": str(quittance.police.prime_ttc),
                        "TTC_POL_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.police.prime_ttc, "EUR", taux_euro)), #pas pris en compte pour l'instant
                        "CD_MON": devise.code,
                        "DTMVT_POL": mouvement_polices.first().date_effet.strftime("%d/%m/%Y") if mouvement_polices else "",
                        "RENOUV": quittance.police.date_prochaine_facture.strftime("%d/%m/%Y") if quittance.police.date_prochaine_facture else "",
                        "RENOUV_MVT": quittance.police.date_prochaine_facture.strftime("%d/%m/%Y") if quittance.police.date_prochaine_facture else "",
                        "TAUX_LOC_POL": "1",
                        "ID_NPOL3": quittance.police.veos_id_npol if quittance.police.veos_id_npol and quittance.police.veos_id_npol != "" else str(quittance.police.pk),
                        "ID_POL3": quittance.police.veos_id_pol if quittance.police.veos_id_pol and quittance.police.veos_id_pol != "" else str(quittance.police.pk),
                        "ID_POL": quittance.police.veos_id_pol if quittance.police.veos_id_pol and quittance.police.veos_id_pol != "" else str(quittance.police.pk),
                        "POL_NUM_APP": apporteur_police.apporteur.code if apporteur_police else "",
                        "NUM_QUI": str(quittance.numero),
                        "ID_QUI": str(quittance.pk),
                        "DATEMI_QUI": quittance.date_emission.strftime("%d/%m/%Y") if quittance.date_emission else "",
                        "DATEMI_QUI_AA": quittance.date_emission.strftime("%Y") if quittance.date_emission else "",
                        "DATDEB_QUI": quittance.date_debut.strftime("%d/%m/%Y") if quittance.date_debut else "",
                        "DATDEB_QUI_AA": quittance.date_debut.strftime("%Y") if quittance.date_debut else "",
                        "DTFIN_QUI": quittance.date_fin.strftime("%d/%m/%Y") if quittance.date_fin else "",
                        "CD_MONQ": devise.code,
                        "TTC_QUI_SAISIE": str(quittance.prime_ttc),
                        "TAUX_LOC_QUIT": "1",
                        "EUROS": str(taux_euro), 
                        "TTC_QUI": str(quittance.prime_ttc),
                        "TTC_QUI_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.prime_ttc, "EUR", taux_euro)), 
                        "TTC_QUI_PAYE_SAISIE": str(quittance.montant_regle),
                        "TTC_QUI_PAYE": str(quittance.montant_regle),
                        "TTC_QUI_PAYE_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.montant_regle, "EUR", taux_euro)), 
                        "PNET_QUI_SAISIE": str(quittance.prime_ht),
                        "PNET_QUI": str(quittance.prime_ht),
                        "PNET_QUI_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.prime_ht, "EUR", taux_euro)), 
                        "MCOM_QUI_SAISIE": str(quittance.commission_courtage),
                        "MCOM_QUI": str(quittance.commission_courtage),
                        "MCOM_QUI_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.commission_courtage, "EUR", taux_euro)), 
                        "MNT_DRQ_RET_SAISIE": str(quittance.commission_intermediaires), 
                        "MNT_DRQ_RET": str(quittance.commission_intermediaires), 
                        "MRETRO_QUI_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, quittance.commission_intermediaires, "EUR", taux_euro)), 
                        "MNT_DRQ_RET_PAYE_SAISIE": str(retro_paye),
                        "MNT_DRQ_RET_PAYE": str(retro_paye),
                        "MRETRO_QUI_PAYE_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, retro_paye, "EUR", taux_euro)), 
                        "MRETRO_QUI_SAISIE": str(quittance.commission_intermediaires),
                        "MRETRO_QUI": str(quittance.commission_intermediaires),
                        "MRETRO_QUI_PAYE_SAISIE": str(retro_paye),
                        "MRETRO_QUI_PAYE": str(retro_paye),
                        "MCOM_QUI_PAYE_SAISIE": str(com_court_encaissee),
                        "MCOM_QUI_PAYE": str(com_court_encaissee), 
                        "MCOM_QUI_ENCAISS_SAISIE": str(com_court_encaissee),
                        "MCOM_QUI_ENCAISS": str(com_court_encaissee),
                        "MCOM_QUI_ENCAISS_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, com_court_encaissee, "EUR", taux_euro)), 
                        "MCOM_QUI_PAYE_CLIENT_SAISIE": str(com_court_encaissee),
                        "MCOM_QUI_PAYE_CLIENT": str(com_court_encaissee),
                        "MCOM_QUI_PAYE_CLIENT_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, com_court_encaissee, "EUR", taux_euro)), 
                        "MCOM_QUI_REVERS_SAISIE": str(com_court_encaissee),
                        "MCOM_QUI_REVERS": str(com_court_encaissee),
                        "MCOM_QUI_REVERS_EUR": str(convert_montant_from_devise_to_euro_or_usd(devise.code, com_court_encaissee, "EUR", taux_euro)), 
                        "MFRAI_QUI": "0", #pas pris en compte pour l'instant
                        "MFRAI_QUI_EUR": "0", #pas pris en compte pour l'instant
                        "MFRAI_QUI_PAYE_CLIENT": "0", #pas pris en compte pour l'instant
                        "MFRAI_QUI_PAYE_CLIENT_EUR": "0", #pas pris en compte pour l'instant
                        "DATREG_QUI": date_reglement_client.strftime("%d/%m/%Y") if date_reglement_client != '' else "",
                        "DATREG_QUI_AA": date_reglement_client.strftime("%Y") if date_reglement_client != '' else "",
                        "DATCIE_QUI": date_reversement_compagnie.strftime("%d/%m/%Y") if date_reversement_compagnie else '',
                        "DATCIE_QUI_AA": date_reversement_compagnie.strftime("%Y") if date_reversement_compagnie else '',
                        "DATCOM_QUI": date_encaissement_commission.strftime("%d/%m/%Y") if date_encaissement_commission else '',
                        "DATCOM_QUI_AA": date_encaissement_commission.strftime("%Y") if date_encaissement_commission else '',
                        "MON_SOC": devise.code,
                        "RENOUV_QUIT": quittance.police.date_prochaine_facture.strftime("%d/%m/%Y") if quittance.police.date_prochaine_facture else "",
                        "ZONE": "FRANCO",
                        "NATPOL": "SANTE",
                        "NATZONE": "FRANCOSANTE",
                        "DIRSOCNATZONE": quittance.bureau.code+"SANTE",
                        "ZONEM": "FRANCO",
                        "TAUX_APPORTEUR_COMPTANT": str(apporteur_police.taux_com_affaire_nouvelle) if apporteur_police else "",
                        "TAUX_APPORTEUR_TERME": str(apporteur_police.taux_com_renouvellement) if apporteur_police else "",
                        "CD_SQ": "", #pas pris en compte pour l'instant
                        "LIB_SQ": "", #pas pris en compte pour l'instant
                        "CD_NATQ": code_nature_quittance,
                        "USD": str(taux_usd), 
                        "POL_ID_APP": apporteur_police.apporteur.id_per if apporteur_police and apporteur_police.apporteur.id_per == "" else "",
                        "DTCRE_PER": quittance.police.client.created_at.strftime("%d/%m/%Y") if quittance.police.client.created_at != '' else "",
                        "ISFILIALE": "F",
                        "NEWBIZ": "NB" if quittance.nature_quittance and (quittance.nature_quittance.code == "COMPT" or "CT") else ""
                    }
                }
            )

            counter += 1

            # Si le compteur atteint 10 ou si c'est la derniﾃｨre quittance de la liste
            if counter == 100 or index == total_quittances:
                #json_data = json.dumps(data)
                #return (data_num_qui)
                api_response = call_api_updt_mvquittances_veos(data)
                data["requests"] = [] # Rﾃｩinitialiser la liste de requﾃｪtes
                counter = 0  # Rﾃｩinitialiser le compteur

        # Si aprﾃｨs la boucle il reste des quittances non envoyﾃｩes (si leur nombre ﾃｩtait infﾃｩrieur ﾃ 10), 
        # on envoie les donnﾃｩes restantes
        if data["requests"]:
            api_response = call_api_updt_mvquittances_veos(data)

        #print(len(data['requests']))
        
        #json_data = json.load(data)
        #dd(json_data)
        #return JsonResponse(json_data, safe=False) 
        #dd(len(data['requests']))    
        try:
            # return JsonResponse(data)
            #api_response = call_api_updt_mvquittances_veos(data)
            #dd(api_response)
            CronLog.objects.create(action="updt_mvquittances", table="quittance", description="Create quittances count " + str(total_quittances) + " on veos ").save() #with id = " + str(api_response['id']) + " et numero = " + str(api_response['num'])).save()
            print("Quittance envoyﾃｩ avec succﾃｨs")
            response = {
            'statut': 1,
            'message': 'TRAITEMENT TERMINE',
            'data': {
                'count_all': nombre_quittance
                }
            }
            return response
            return JsonResponse(f"Quittance envoyﾃｩ avec succﾃｨs", safe=False)

        except Exception as e:
            response = {
            'statut': 1,
            'message': 'TRAITEMENT ERREUR',
            'data': {
                'description': str(e)
                }
            }
            return response
            return JsonResponse(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}", safe=False)

        #except ValueError as e:
        #    response = []
        #    CronLog.objects.create(action="updt_mvquittances", table="quittance",
        #                            description="Error when creating quittance Nﾂｰ " + quittance.numero).save()
        #    print("Erreur :", e)

        

        temps_fin = time.time()
        temps_execution = temps_fin - temps_debut

        message = f"Envoi de ({nombre_quittance}) quittance(s) terminﾃｩ en {temps_execution:.7f} sﾃｩcondes."

        return message


def get_grinter_veos():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_BUREAU_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    response = {
        'statut': 1,
        'message': 'TRAITEMENT NON EFFECTUE',
    }

    count_new = 0
    count_old = 0

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    #dd(api_response)
    if api_response['statusCode'] == 0:

        grinters = api_response['responses'][0]['beans']

        """
        for grinter in grinters:
            # Check if the record already exists based on a unique field (code)
            existing_record = GroupeInter.objects.filter(code=grinter["code"]).first()

            pays = Pays.objects.filter(code=bureau['pays'])
            p = pays.first() if pays else None

            devises = Devise.objects.filter(code=bureau['devise'])
            devise = devises.first() if pays else None

            if existing_record:
                
                pass
                existing_record.code = p
                existing_record.apporteur = devise
                existing_record.nom = bureau['nom']
                existing_record.status = bureau['telephone']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(update_fields=['pays', 'devise', 'nom', 'telephone', 'email', 'adresse', 'updated_at'])
                count_old += 1

            else:

                new = GroupeInter(code=p,
                         apporteur=devise,
                         code=bureau['code'],
                         nom=bureau['nom'])
                new.save()
                count_new += 1

            response = {
                'statut': 1,
                'message': 'TRAITEMENT TERMINE',
                'data': {
                    'count_new': count_new,
                    'count_old': count_old,
                    'count_all': count_new + count_old
                }
            }
        """
    return response


def get_apporteur_veos():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_BUREAU_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    response = {
        'statut': 1,
        'message': 'TRAITEMENT NON EFFECTUE',
    }

    count_new = 0
    count_old = 0

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    #dd(api_response)
    if api_response['statusCode'] == 0:

        apporteurs = api_response['responses'][0]['beans']

        """
        for apporteur in apporteurs:
            # Check if the record already exists based on a unique field (code)
            existing_record = Apporteur.objects.filter(code=grinter["code"]).first()

            pays = Pays.objects.filter(code=bureau['pays'])
            p = pays.first() if pays else None

            devises = Devise.objects.filter(code=bureau['devise'])
            devise = devises.first() if pays else None

            if existing_record:
                
                pass
                existing_record.adresse = p
                existing_record.apporteur_international = devise
                existing_record.bureau = devise
                existing_record.code = devise
                existing_record.email = bureau['nom']
                existing_record.id_per = bureau['nom']
                existing_record.nom = bureau['nom']
                existing_record.prenoms = bureau['nom']
                existing_record.pays = bureau['nom']
                existing_record.telephone = bureau['nom']
                existing_record.type_apporteur = bureau['nom']
                existing_record.type_personne = bureau['nom']
                existing_record.status = bureau['telephone']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save(update_fields=['pays', 'devise', 'nom', 'telephone', 'email', 'adresse', 'updated_at'])
                count_old += 1

            else:

                new = Apporteur(adresse=p,
                         apporteur_international=devise,
                         bureau=devise,
                         code=devise,
                         email=devise,
                         id_per=bureau['code'],
                         nom=bureau['code'],
                         prenoms=bureau['code'],
                         pays=bureau['code'],
                         telephone=bureau['code'],
                         type_apporteur=bureau['code'],
                         type_personne=bureau['code'],
                         status=bureau['nom'])
                new.save()
                count_new += 1

            response = {
                'statut': 1,
                'message': 'TRAITEMENT TERMINE',
                'data': {
                    'count_new': count_new,
                    'count_old': count_old,
                    'count_all': count_new + count_old
                }
            }
        """

    return response


def get_clients_veos():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_BUREAU_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    response = {
        'statut': 1,
        'message': 'TRAITEMENT NON EFFECTUE',
    }

    count_new = 0
    count_old = 0
    count_update = 0

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    #dd(api_response)
    if api_response['statusCode'] == 0:

        clients = api_response['responses'][0]['beans']

        """
        #dd(clients)
        for client in clients:
            # Check if the record already exists based on a unique field (code)
            existing_record = Client.objects.filter(code=client["code"]).first()

            bureaux = Bureau.objects.filter(code=client['bureau'])
            bureau = bureaux.first() if bureaux else None
            
            payss = Pays.objects.filter(code=client['cd_pays'])
            pays = payss.first() if payss else bureau.pays if bureau else None

            types_personnes = TypePersonne.objects.filter(code=client['type_per'])
            type_personne = types_personnes.first() if types_personnes else None

            type_clients = TypeClient.objects.filter(code=client['bureau'])
            type_client = type_clients.first() if type_clients else None

            civilites = Civilite.objects.filter(code=client['civilite'])
            civilite = civilites.first() if civilites else None

            langues = Langue.objects.filter(code=client['bureau'])
            langue = langues.first() if langues else None

            # dd(existing_record)
            if existing_record:
                
                count_old += 1
                if existing_record.veos_client_idper is None or existing_record.veos_client_idper == '':
                    existing_record.veos_client_idper = client['veos_client_idper']
                    existing_record.updated_at = datetime.now(tz=timezone.utc)
                    existing_record.save()
                    count_update += 1

            else:

                new = Client(
                    bureau=bureau,
                    type_personne=type_personne,
                    type_client=type_client,
                    pays=pays,
                    code=client['code'],
                    nom=client['nom'],
                    prenoms=client['prenoms'],
                    civilite=civilite,
                    sexe="",
                    date_naissance=client['date_naissance'][:10] if client['date_naissance'] else None,
                    telephone_mobile=client['telephone_mobile'],
                    telephone_fixe=client['telephone_princ'],
                    email=client['email'],
                    langue=langue,
                    longitude="",
                    latitude="",
                    ville=client['ville'],
                    adresse_postale="",
                    adresse=client['adresse'],
                    site_web="",
                    twitter="",
                    instagram="",
                    facebook="",
                    ancienne_ref="",
                    statut="ACTIF",
                    statut_relation="CLIENT",
                    )

                new.save()
                count_new += 1
        """
        response = {
            'statut': 1,
            'message': 'TRAITEMENT TERMINE',
            'data': {
                'count_new': count_new,
                'count_old': count_old,
                'count_update': count_update,
                'count_all': count_new + count_old
            }
        }

    return response


def call_api_client_veos(data):
    # Construire l'URL de l'API en utilisant la variable API_URL
    # api_url = f"{API_URL}/sinistre"
    api_url = f"{API_URL_PROD}/personne"
    #api_url = f"{API_URL_PROD}/personne"

    # dd(data)

    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        # Appel de l'API en utilisant la mﾃｩthode POST avec les donnﾃｩes JSON et les en-tﾃｪtes
        response = requests.post(api_url, json=data, headers=headers)

        #return JsonResponse(response, safe=False)

        # return json.loads(response.content)

        # Vﾃｩrification du statut de la rﾃｩponse
        if response.status_code == 200:

            # Si la requﾃｪte a rﾃｩussi, on renvoie un objet de donnees
            api_response_data = response.content
    
            # return JsonResponse({"content": response.content}, status=response.status_code)
            return json.loads(api_response_data)

        else:
            response = {
            'statut': 1,
            'message': 'TRAITEMENT ERREUR',
            'data': {
                'description': str(e)
                }
            }
            return response            
            # La rﾃｩponse ne correspond pas ﾃ ce qui est attendu
            return JsonResponse({"error": "Erreur lors de l'envoie du sinistre sur VEOS"}, status=500, safe=False)
            raise ValueError("Erreur lors de l'envoie du sinistre sur VEOS")
            # Si la requﾃｪte n'a pas rﾃｩussi, retournez None ou une valeur par dﾃｩfaut selon votre cas
            return JsonResponse({"error": ""}, status=500)

    except json.JSONDecodeError as e:
        # Une erreur s'est produite lors du chargement 
        response = {
        'statut': 1,
        'message': 'TRAITEMENT ERREUR',
        'data': {
            'description': str(e)
            }
        }
        return response
        return JsonResponse(f"Erreur lors du chargement JSON : {e}", safe=False)
        raise ValueError(f"Erreur lors du chargement JSON : {e}")

    except Exception as e:
        response = {
        'statut': 1,
        'message': 'TRAITEMENT ERREUR',
        'data': {
            'description': str(e)
            }
        }
        return response        
        return JsonResponse(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}", safe=False)
        raise ValueError(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}")


# send client to veos
def send_list_clients_to_veos():
    env = environ.Env()
    environ.Env.read_env(env_file=str(BASE_DIR / "inov" / ".env"))

    ENVIRONMENT = env("ENVIRONMENT")

    if ENVIRONMENT  == "PROD":
    #if ENVIRONMENT == "PREPROD":
    # if ENVIRONMENT:
        # Liste des quttances ﾃ envoyer
        clients = Client.objects.filter(veos_client_idper__isnull=True, nom__isnull=False, polices__isnull=False, statut=Statut.ACTIF).exclude(nom='')
        nombre_clients = clients.count()
        # Dﾃｩmarrage du compte ﾃ rebourt pour la crﾃｩation des quittances sur veos
        temps_debut = time.time()

        # print(clients)
        # dd(clients)

        counter = 0
        counter_update = 0
        index = 0
        total_clients = len(clients)
        data_num_cli = []

        for client in clients:

            index += 1

            date_str = client.created_at.strftime("%Y-%m-%d %H:%M:%S")
            date_format = "%Y-%m-%d %H:%M:%S"

            timestamp = datetime.strptime(date_str, date_format).timestamp()

            adresse = ""
            if client.adresse and client.adresse != "":
                adresse = client.adresse
            elif client.ville and client.ville != "":
                adresse = str(client.ville) + ", " + str(client.pays.nom)
            else:
                adresse = str(client.pays.nom)


            # convert_montant_from_devise_to_euro_or_usd(devise_from, montant, euro_or_usd, taux_euro_or_usd=None)
            # data_num_qui.append(quittance.numero)
            data = {
                        "categorie":"ASSURE", #//Fixe
                        "numSoc": client.bureau.code, #//Code societe VEOS
                        "type": client.type_personne.code if client.type_personne else "P", #//Type personne (P - >Physique | M -> Morale)
                        "nom": str(client.nom) + " " + str(client.prenoms if client.prenoms else ""), #//Nom personne
                        "mail": client.email, #//Email personne
                        "mobile": client.telephone_mobile, #//Tel mobile personne
                        "tel": client.telephone_fixe, #//Autre tel personne
                        "adresses":[
                            {
                            "adresse1": adresse, #//Adresse personne
                            "ville": client.ville, #//Ville personne
                            "cdPays": client.pays.code, #//Code pays
                            "principale":"1", #//Fixe
                            "categorie":"PRINCENVOI" #//Fixe
                            }
                        ],
                        "listInfos": [
                            {
                                "key": "INFOPER.REFEXTERNE",
                                "value": client.groupe_international.code if client.groupe_international else "ZZZ" #//Code groupe inter (Par dﾃｩfaut ZZZ)
                            },
                            {
                                "key": "INFOPER.REFERENT",
                                "value": "WS" #//Fixe
                            },
                            {
                                "key": "INFOPER.SUIVIPAR",
                                "value": "WS" #//Fixe
                            }
                        ]
                    }

            api_response = call_api_client_veos(data)
            client.veos_client_idper = api_response["id"]
            client.updated_at = datetime.now(tz=timezone.utc)
            client.save()
            # CronLog.objects.create(action="send_clients_to_veos", table="clients",
                                # description="1 client IDPER : " + str(api_response["id"]) + " / Num : " + str(api_response["num"]) + " created on veos").save()
            counter_update += 1
            print("Client envoyﾃｩ avec succﾃｨs")

            """
            if ((client.adresse or client.ville) and client.bureau):
                api_response = call_api_client_veos(data)
                client.veos_client_idper = api_response["id"]
                client.updated_at = datetime.now(tz=timezone.utc)
                client.save()
                # CronLog.objects.create(action="send_clients_to_veos", table="clients",
                                   # description="1 client IDPER : " + str(api_response["id"]) + " / Num : " + str(api_response["num"]) + " created on veos").save()
                counter_update += 1
                print("Client envoyﾃｩ avec succﾃｨs")
            else:
                client.adresse = client.pays.nom
                api_response = call_api_client_veos(data)
                CronLog.objects.create(action="send_clients_to_veos", table="clients",
                                   description="1 failure creation client : " + str(client.pk) + " on veos (adresse ou ville manquante)").save()
                print("Echec envoie client : " + str(client.pk))
            """

            counter += 1


        # print(len(data['requests']))

        # json_data = json.load(data)
        # dd(json_data)
        # return JsonResponse(json_data, safe=False)
        # dd(len(data['requests']))
        try:
            # return JsonResponse(data)
            # api_response = call_api_updt_mvquittances_veos(data)
            # dd(api_response)
            #CronLog.objects.create(action="send_clients_to_veos", table="clients",
                                  # description="Create clients count " + str(nombre_clients) + " on veos ").save()  # with id = " + str(api_response['id']) + " et numero = " + str(api_response['num'])).save()
            print("Client envoyﾃｩ avec succﾃｨs")
            response = {
                'statut': 1,
                'message': 'TRAITEMENT TERMINE',
                'data': {
                    'count_all': total_clients,
                    'counter_update': counter_update
                }
            }
            return response
            return JsonResponse(f"Client envoyﾃｩ avec succﾃｨs", safe=False)

        except Exception as e:
            response = {
                'statut': 1,
                'message': 'TRAITEMENT ERREUR',
                'data': {
                    'description': str(e)
                }
            }
            return response
            return JsonResponse(f"Erreur, une exception ﾃ ﾃｩtﾃｩ lﾃｩvﾃｩe : {e}", safe=False)

        # except ValueError as e:
        #    response = []
        #    CronLog.objects.create(action="updt_mvquittances", table="quittance",
        #                            description="Error when creating quittance Nﾂｰ " + quittance.numero).save()
        #    print("Erreur :", e)

        temps_fin = time.time()
        temps_execution = temps_fin - temps_debut

        message = f"Envoi de ({nombre_quittance}) client(s) terminﾃｩ en {temps_execution:.7f} sﾃｩcondes."

        return message


def send_client_to_veos(client):
    env = environ.Env()
    environ.Env.read_env(env_file=str(BASE_DIR / "inov" / ".env"))

    ENVIRONMENT = env("ENVIRONMENT")

    #if ENVIRONMENT  == "PROD":
    if ENVIRONMENT == "PROD":
    # if ENVIRONMENT:
        # Liste des quttances ﾃ envoyer
        # Dﾃｩmarrage du compte ﾃ rebourt pour la crﾃｩation des quittances sur veos
        temps_debut = time.time()

        # print(clients)
        # dd(clients)


        date_str = client.created_at.strftime("%Y-%m-%d %H:%M:%S")
        date_format = "%Y-%m-%d %H:%M:%S"

        timestamp = datetime.strptime(date_str, date_format).timestamp()

        adresse = ""
        if client.adresse and client.adresse != "":
            adresse = client.adresse
        elif client.ville and client.ville != "":
            adresse = str(client.ville) + ", " + str(client.pays.nom)
        else:
            adresse = str(client.pays.nom)

        # convert_montant_from_devise_to_euro_or_usd(devise_from, montant, euro_or_usd, taux_euro_or_usd=None)
        # data_num_qui.append(quittance.numero)
        data = {
                    "categorie":"ASSURE", #//Fixe
                    "numSoc": client.bureau.code, #//Code societe VEOS
                    "type": client.type_personne.code if client.type_personne else "P", #//Type personne (P - >Physique | M -> Morale)
                    "nom": str(client.nom) + " " + str(client.prenoms if client.prenoms else ""), #//Nom personne
                    "mail": client.email, #//Email personne
                    "mobile": client.telephone_mobile, #//Tel mobile personne
                    "tel": client.telephone_fixe, #//Autre tel personne
                    "adresses":[
                        {
                        "adresse1": adresse, #//Adresse personne
                        "ville": client.ville, #//Ville personne
                        "cdPays": client.pays.code, #//Code pays
                        "principale":"1", #//Fixe
                        "categorie":"PRINCENVOI" #//Fixe
                        }
                    ],
                    "listInfos": [
                        {
                            "key": "INFOPER.REFEXTERNE",
                            "value": client.groupe_international.code if client.groupe_international else "ZZZ" #//Code groupe inter (Par dﾃｩfaut ZZZ)
                        },
                        {
                            "key": "INFOPER.REFERENT",
                            "value": "WS" #//Fixe
                        },
                        {
                            "key": "INFOPER.SUIVIPAR",
                            "value": "WS" #//Fixe
                        }
                    ]
                }

        api_response = call_api_client_veos(data)
        client.veos_client_idper = api_response["id"]
        client.updated_at = datetime.now(tz=timezone.utc)
        client.save()
        CronLog.objects.create(action="send_clients_to_veos", table="clients",
                            description="1 client IDPER : " + str(api_response["id"]) + " created on veos").save()
        counter_update += 1
        print("Client envoyﾃｩ avec succﾃｨs")

        return True



def import_police_manuellement(request):

    cpt_anciennes_polices = 0
    cpt_nouvelles_polices = 0
    cpt_toutes_polices=0


    polices_veos = PoliceVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_NPOL')[:1000]
    cpt = 0

    #dd(polices_veos)
    if polices_veos:

        for police in polices_veos:

            #infos pour crﾃｩer la nouvelle police
            bureaux = Bureau.objects.filter(code=police.BUREAU_ID)
            bureau = bureaux.first() if bureaux else None

            clients = Client.objects.filter(code=police.CLIENT_ID)
            client = clients.first() if clients else None
            #dd(client)

            compagnies = Compagnie.objects.filter(code=police.COMPAGNIE_ID)
            compagnie = compagnies.first() if compagnies else None
            #dd(compagnie)

            type_prefinancements = TypePrefinancement.objects.filter(code=police.PEF_TMOD)
            type_prefinancement = type_prefinancements.first() if type_prefinancements else None

            date_souscription = police.DATE_DEBUT_EFFET[:10] if police.DATE_DEBUT_EFFET else "",
            date_debut_effet = police.DATE_MVT[:10] if police.DATE_MVT else "",
            date_fin_effet = police.DATE_FIN_EFFET[:10] if police.DATE_FIN_EFFET else "",

            pprint("date_souscription")
            pprint(date_souscription)
            pprint("date_debut_effet")
            pprint(date_debut_effet)
            pprint("date_fin_effet")
            pprint(date_fin_effet)

            date_souscription = datetime.strptime(date_souscription[0], '%d/%m/%Y').date() if date_souscription[0] else None
            date_debut_effet = datetime.strptime(date_debut_effet[0], '%d/%m/%Y').date() if date_debut_effet[0] else None
            date_fin_effet = datetime.strptime(date_fin_effet[0], '%d/%m/%Y').date() if date_fin_effet[0] else None

            motifs = Motif.objects.filter(code=police.MOTIF)
            motif = motifs.first() if motifs else None

            mouvement = motif.mouvement if motif else None

            statut = Statut.ACTIF if police.STATUS == 'EC' else Statut.INACTIF

            devise = bureau.pays.devise if bureau else None


            # Check if the record already exists based on a unique field (code)
            existing_record = Police.objects.filter(veos_id_npol=police.ID_NPOL).first()

            #dd(existing_record)
            if existing_record:

                #autres traitements ﾃ effectuer

                #si la date a changﾃｩ, mettre ﾃ jour dans mouvement_police et periode couverture
                if mouvement:

                    pprint("if motif.code == RENOUV and date_debut_effet != existing_record.date_debut_effet:")
                    pprint("motif.code")
                    pprint(motif.code)
                    pprint("date_debut_effet")
                    pprint(date_debut_effet)
                    pprint("existing_record.date_debut_effet")
                    pprint(existing_record.date_debut_effet)

                    #si le motif est renouv et que les dates ont changﾃｩ
                    #if motif.code == "RENOUV" and date_debut_effet != existing_record.date_debut_effet:
                    if motif.code == "RENOUV" and date_debut_effet != existing_record.date_debut_effet:

                        # crﾃｩer une ligne dans mouvement_police
                        new_mouvement = MouvementPolice(
                            police=existing_record,
                            mouvement_id=mouvement.id,
                            motif=motif,
                            date_effet=date_debut_effet,
                            date_fin_periode_garantie=date_fin_effet,
                        )
                        new_mouvement.save()
                        # dd(new_mouvement)

                        #cloturer les autres pﾃｩriodes en cours
                        PeriodeCouverture.objects.filter(police=existing_record).update(statut=Statut.INACTIF)

                        # crﾃｩer une nouvelle periode de couverture
                        new_periode_couverture = PeriodeCouverture(
                            police=existing_record,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                        )
                        new_periode_couverture.save()


                #Mettre ﾃ jour les infos de la police
                existing_record.updated_at = datetime.now()
                existing_record.veos_code_cie = police.COMPAGNIE_ID
                #existing_record.compagnie = compagnie
                existing_record.numero = police.NUMERO
                existing_record.numero_provisoire = police.NUMERO_PROVISOIRE
                existing_record.prime_ttc = decimal.Decimal(police.PRIME_TTC) if police.PRIME_TTC else 0
                existing_record.prime_net = decimal.Decimal(police.PRIME_NET) if police.PRIME_NET else 0
                existing_record.prime_ht = decimal.Decimal(police.PRIME_NET) if police.PRIME_NET else 0
                if police.TAUXCOM: existing_record.taux_com_courtage = decimal.Decimal(police.TAUXCOM)
                if police.TAUXGEST: existing_record.taux_com_gestion = decimal.Decimal(police.TAUXGEST)
                if police.MONTANTCOM: existing_record.commission_courtage = decimal.Decimal(police.MONTANTCOM)
                if police.MONTANTGEST: existing_record.commission_gestion = decimal.Decimal(police.MONTANTGEST)
                existing_record.date_souscription = date_souscription
                existing_record.date_debut_effet = date_debut_effet
                if date_fin_effet: existing_record.date_fin_effet = date_fin_effet
                existing_record.statut = statut
                existing_record.observation = "UPDATED BY CRON - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S')) + " statut= " + str(police.STATUS)
                existing_record.save()

                cpt_anciennes_polices += 1

            else:
                try:
                    #dd(police)
                    #Crﾃｩer la police

                    if client and compagnie and bureau:

                        new_police = Police(
                            type_assurance_id=1,
                            produit_id=1,
                            veos_code_client=police.CLIENT_ID,
                            veos_code_cie=police.COMPAGNIE_ID,
                            veos_id_npol=police.ID_NPOL,
                            veos_id_pol=police.ID_POL,
                            devise=devise,
                            bureau=bureau,
                            client=client,
                            compagnie=compagnie,
                            numero=police.NUMERO,
                            numero_provisoire=police.NUMERO_PROVISOIRE,
                            prime_ttc=decimal.Decimal(police.PRIME_TTC) if police.PRIME_TTC else 0,
                            prime_net=decimal.Decimal(police.PRIME_NET) if police.PRIME_NET else 0,
                            prime_ht=decimal.Decimal(police.PRIME_NET) if police.PRIME_NET else 0,
                            taux_com_courtage = decimal.Decimal(police.TAUXCOM) if police.TAUXCOM else 0,
                            taux_com_gestion = decimal.Decimal(police.TAUXGEST) if police.TAUXGEST else 0,
                            commission_courtage = decimal.Decimal(police.MONTANTCOM) if police.MONTANTCOM else 0,
                            commission_gestion = decimal.Decimal(police.MONTANTGEST) if police.MONTANTGEST else 0,
                            date_souscription=date_souscription,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                            statut_contrat=StatutContrat.CONTRAT,
                            statut=statut,
                            type_prefinancement=type_prefinancement,
                            observation="CREATED BY CRON - " + str(datetime.now(tz=timezone.utc)),
                            )

                        new_police.save()
                        #dd(new_police)

                        #crﾃｩer une ligne dans mouvement_police
                        new_mouvement = MouvementPolice(
                            police=new_police,
                            mouvement_id=mouvement.id,
                            motif=motif,
                            date_effet=date_debut_effet,
                            date_fin_periode_garantie=date_fin_effet,
                            )
                        new_mouvement.save()
                        #dd(new_mouvement)

                        #crﾃｩer la periode de couverture
                        new_periode_couverture = PeriodeCouverture(
                            police=new_police,
                            date_debut_effet=date_debut_effet,
                            date_fin_effet=date_fin_effet,
                            )
                        new_periode_couverture.save()


                        cpt_nouvelles_polices += 1

                    else:
                        pprint("Client " + str(police.CLIENT_ID) + " ou compagnie " + str(police.COMPAGNIE_ID) + " non trouvﾃｩ pour la police " + str(police.NUMERO))


                except Exception as e:
                    error_message = str(e)  # Store the exception message
                    result = None

                    #dd(e)

            #marquer comme importﾃｩ
            police.STATUT_IMPORT = True
            police.save()

            cpt_toutes_polices += 1




        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciennes_polices': cpt_anciennes_polices,
                'cpt_nouvelles_polices': cpt_nouvelles_polices,
                'cpt_toutes_polices': cpt_toutes_polices
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)



def import_formule_manuellement(request):

    cpt_anciennes_formules = 0
    cpt_nouvelles_formules = 0
    cpt_toutes_formules=0


    formules_veos = FormuleVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_NPOL')[:1000]
    cpt = 0

    #dd(formules_veos)
    if formules_veos:

        for formule in formules_veos:

            existing_record = FormuleGarantie.objects.filter(code=formule.CODE_FORMULE).first()

            polices = Police.objects.filter(veos_id_npol=formule.ID_NPOL)
            police = polices.first() if polices else None

            # dd(polices)
            territorialites = Territorialite.objects.filter(code=formule.CODE_TERRITORIALITE)
            territorialite = territorialites.first() if territorialites else None

            #types_tarifs = TypeTarif.objects.filter(code=formule.CODE_TYPE_TARIF) #ce n'est pas ﾃｧa: il renvoie plutﾃｴt l'option calcul de la prime
            #type_tarif = types_tarifs.first() if types_tarifs else None
            type_tarif = None


            date_debut = formule.DATE_DEBUT[:10] if formule.DATE_DEBUT else "",
            date_fin = formule.DATE_FIN[:10] if formule.DATE_FIN else "",

            date_debut = datetime.strptime(date_debut[0], '%Y-%m-%d').date() if date_debut[0] else None
            date_fin = datetime.strptime(date_fin[0], '%Y-%m-%d').date() if date_fin[0] else None

            # dd(existing_record.libelle)
            if existing_record:
                existing_record.libelle = formule.LIBELLE_FORMULE
                # existing_record.date_debut = formule['date_debut'][:10] if formule['date_debut'] else None
                # existing_record.date_fin = formule['date_fin'][:10] if formule['date_fin'] else None
                if police: existing_record.police = police
                # existing_record.territorialite = territorialite
                # if formule['code_territorialite']: existing_record.code = formule['code_territorialite']
                existing_record.save()

                cpt_anciennes_formules += 1

            else:

                new_formule = FormuleGarantie(
                    police=police,
                    territorialite=territorialite,
                    type_tarif=type_tarif,
                    code=formule.CODE_FORMULE if formule.CODE_FORMULE else None,
                    libelle=formule.LIBELLE_FORMULE,
                    taux_couverture=0,
                    taux_tm=100,
                    plafond_conso_famille=None,
                    plafond_conso_individuelle=None,
                    date_debut=date_debut,
                    date_fin=date_fin,

                    infos_carte_consultation = formule.LIGNE1,
                    infos_carte_hospitalisation = formule.LIGNE2,
                    infos_carte_ambulatoire = formule.LIGNE3,
                    infos_carte_vitamine = formule.LIGNE4,
                    infos_carte_vaccination = formule.LIGNE5,

                    statut=Statut.ACTIF,
                )

                new_formule.save()

                print(formule.LIBELLE_FORMULE + " <<< === >>> " + formule.CODE_FORMULE + " (NEW)")
                print("\n")

                cpt_nouvelles_formules += 1


            #marquer comme importﾃｩ
            formule.STATUT_IMPORT = True
            formule.save()

            cpt_toutes_formules += 1



        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciennes_formules': cpt_anciennes_formules,
                'cpt_nouvelles_formules': cpt_nouvelles_formules,
                'cpt_toutes_formules': cpt_toutes_formules
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)




def get_formules_lists():

    cpt_anciennes_formules = 0
    cpt_nouvelles_formules = 0
    cpt_toutes_formules=0


    api_url = f"{API_URL}/boBy/list"

    data = {
        "requests": [
            {
                "name": "WS_FORMULE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    api_response = call_api_with_data(data)
    #dd(api_response)
    if api_response['statusCode'] == 0:

        formules = api_response['responses'][0]['beans']

        # dd(formules)
        for formule in formules:

            existing_record = FormuleGarantie.objects.filter(code=formule['code_formule']).first()

            polices = Police.objects.filter(veos_id_npol=formule['id_npol'])
            police = polices.first() if polices else None

            #dd(polices)
            territorialites = Territorialite.objects.filter(code=formule['code_territorialite'])
            territorialite = territorialites.first() if territorialites else None

            #types_tarifs = TypeTarif.objects.filter(code=formule['code_type_tarif']) #pas ﾃｧa : option calcul de la prime
            #type_tarif = types_tarifs.first() if types_tarifs else None
            type_tarif = None


            #dd(existing_record.libelle)
            if existing_record:
                #existing_record.libelle = formule['libelle_formule']
                #existing_record.date_debut = formule['date_debut'][:10] if formule['date_debut'] else None
                #existing_record.date_fin = formule['date_fin'][:10] if formule['date_fin'] else None
                if police: existing_record.police = police
                #existing_record.territorialite = territorialite
                #if formule['code_territorialite']: existing_record.code = formule['code_territorialite']
                existing_record.save()

                cpt_anciennes_formules += 1

            else:

                new_formule = FormuleGarantie(
                    police=police,
                    territorialite=territorialite,
                    type_tarif=type_tarif,
                    code=formule['code_formule'] if formule['code_formule'] else None,
                    libelle=formule['libelle_formule'],
                    taux_couverture=0,
                    taux_tm=100,
                    plafond_conso_famille=None,
                    plafond_conso_individuelle=None,
                    date_debut=formule['date_debut'][:10] if formule['date_debut'] else None,
                    date_fin=formule['date_fin'][:10] if formule['date_fin'] else None,
                    statut=Statut.ACTIF,
                )

                new_formule.save()

                cpt_nouvelles_formules += 1

            cpt_toutes_formules += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciennes_formules': cpt_anciennes_formules,
                'cpt_nouvelles_formules': cpt_nouvelles_formules,
                'cpt_toutes_formules': cpt_toutes_formules
            }
        }

        return response


def get_aliments_lists():
    today = datetime.now(tz=timezone.utc).strftime('%d/%m/%Y')
    yesterday = (datetime.now(tz=timezone.utc) - timedelta(days=3)).strftime('%d/%m/%Y')

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    liste_code_bureaux = ['SN01', 'TG01', 'BN01', 'BF01', 'CF01', 'CG01', 'CM01', 'GA01'] #'CI01': la ci est entiﾃｨrement sur la v2 dﾃｩsormais,

    for code_bureau in liste_code_bureaux:

        param = {
            'date_debut': yesterday,
            'date_fin': today,
            'num_soc': code_bureau
        }
        #dd(param)

        pprint("____GET ALIMENT OF BUREAU "+ str(code_bureau) +" _____")
        # Appel de la fonction pour effectuer l'appel ﾃ l'API
        aliments_to_sort = query_builder_tools("WS_ALIMENT_TACHE_CRON", param)

        if aliments_to_sort:
            #dd(aliments_to_sort)
            aliments = sorted(aliments_to_sort, key=lambda x: x.get('ID', ''), reverse=True)

            cpt_all = 0
            cpt_anciens_sans_changement_formule = 0
            cpt_anciens_avec_changement_formule = 0
            cpt_nouveaux = 0

            date_du_jour = datetime.now(tz=timezone.utc).date()

            pprint("____LOOP ON ALIMENT OF "+ str(code_bureau) +" _____")
            #dd(aliments)
            for aliment in aliments:

                # Check if the record already exists based on a unique field (code)
                existing_record = Aliment.objects.filter(veos_code_aliment=aliment["ID"]).first()

                adherent_principal = Aliment.objects.filter(veos_code_aliment=aliment["ADHERENT_PRINCIPAL_ID"])
                adherent_principal_id = adherent_principal.first().id if adherent_principal else None  # By Richmond / A Vﾃ嘘IFIER

                pays_naissance = Pays.objects.filter(code=aliment["NOM"])  # A ajouter au boBy PAYS_ID
                pays_naissance_id = pays_naissance.first().id if pays_naissance else None  # By Richmond / A Vﾃ嘘IFIER

                pays_residence = Pays.objects.filter(code=aliment["NOM"])  # A ajouter au boBy
                pays_residence_id = pays_residence.first().id if pays_residence else None  # By Richmond / A Vﾃ嘘IFIER

                pays_activite_professionnelle = Pays.objects.filter(code=aliment["NOM"])  # A ajouter au boBy
                pays_activite_professionnelle_id = pays_activite_professionnelle.first().id if pays_activite_professionnelle else None  # By Richmond / A Vﾃ嘘IFIER

                profession = Profession.objects.filter(code=aliment["NOM"])  # A ajouter au boBy PROFESSION_ID
                profession_id = profession.first().id if profession else None  # By Richmond / A Vﾃ嘘IFIER

                qualite_beneficiaire = QualiteBeneficiaire.objects.filter(code=aliment["CODE_QUALITE_BENEFICIAIRE"])  # A ajouter au boBy CODE_QUALITE_BENEFICIAIRE
                qualite_beneficiaire_id = qualite_beneficiaire.first().id if qualite_beneficiaire else None  # By Richmond / A Vﾃ嘘IFIER

                formules = FormuleGarantie.objects.filter(code=aliment["CD_FORMULE"], code__isnull=False)  # A ajouter au boBy CODE_FORMULE
                formule = formules.first() if formules else None

                # dd(aliment.CODE_QUALITE_BENEFICIAIRE)

                # dd(existing_record)
                if existing_record:
                    pprint("ALIMENT EXISTANT DEJA EN BASE")

                    existing_record.observation = "UPDATED BY CRON - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S'))
                    existing_record.qualite_beneficiaire_id = qualite_beneficiaire_id
                    if aliment["DATE_NAISSANCE"]: existing_record.date_naissance = aliment["DATE_NAISSANCE"][:10]
                    #si la date sortie est renseignﾃｩ c'est qu'il est sortie
                    if aliment["DATE_SORTIE"]:
                        existing_record.date_sortie = aliment["DATE_SORTIE"][:10]

                    existing_record.veos_code_qualite_beneficiaire = aliment["CODE_QUALITE_BENEFICIAIRE"]
                    existing_record.save()

                    # s'il a changﾃｩ de formule:

                    if formule:
                        aliment_formule_old = AlimentFormule.objects.filter(aliment_id=existing_record.pk, formule_id=formule.pk).order_by('-id').first()

                        if aliment_formule_old:

                            # il a sa meme formule, mettre
                            pprint("Il a toujours sa mﾃｪme formule:" + aliment_formule_old.formule.code)
                            #aliment_formule_old = aliment_formule_olds.first()

                            # si la date sortie est renseignﾃｩ c'est qu'il est sortie de la formule
                            if aliment["DATE_SORTIE"]:
                                aliment_formule_old.date_fin = aliment["DATE_SORTIE"][:10]
                                aliment_formule_old.save()
                                #dd(aliment_formule_old)

                            cpt_anciens_sans_changement_formule = cpt_anciens_sans_changement_formule + 1

                        else:

                            # sinon il a changﾃｩ de formule
                            pprint("IL A CHANGﾃ DE FORMULE"+ str(existing_record.pk) + " " + str(formule.pk))

                            # rﾃｩcupﾃｩrer son ancienne ligne aliment_formule et mettre la date de fin
                            formule_precedente = AlimentFormule.objects.filter(aliment_id=existing_record.pk, date_fin__isnull=True).order_by('-id').first()

                            if formule_precedente:

                                #pas bon
                                '''
                                date_entree = aliment["DATE_ENTREE"][:10]
                                date_entree_obj = datetime.strptime(date_entree, '%Y-%m-%d').date()
                                date_entree = date_entree_obj - timedelta(days=1)
                                date_fin = date_entree - timedelta(days=1)
                                '''

                                #repris
                                date_fin = date_du_jour - timedelta(days=1)

                                formule_precedente.date_fin = date_fin
                                formule_precedente.save()
                                #dd(formule_precedente)


                            # crﾃｩer une ligne dans aliment_formule
                            aliment_formule = AlimentFormule(
                                aliment_id=existing_record.pk,
                                formule_id=formule.pk,
                                date_debut=date_du_jour,
                                date_fin=aliment["DATE_SORTIE"][:10] if aliment["DATE_SORTIE"] else None,
                            )
                            aliment_formule.save()

                            cpt_anciens_avec_changement_formule = cpt_anciens_avec_changement_formule + 1

                else:

                    if formule:

                        bureau = formule.police.bureau if formule.police else None

                        # dd(aliment.DATE_ENTREE[:10])
                        new_aliment = Aliment(
                            bureau=bureau,
                            nom=aliment["NOM"],
                            prenoms=aliment["PRENOMS"],
                            date_naissance=aliment["DATE_NAISSANCE"][:10] if aliment["DATE_NAISSANCE"] else None,
                            lieu_naissance="",
                            genre=aliment["GENRE"],
                            email=aliment["EMAIL"],
                            numero_securite_sociale="",
                            numero=aliment["ID"],  # A verifier
                            numero_famille=aliment["NUMERO_FAMILLE"],
                            matricule_employe="",
                            matricule_cie="",
                            date_affiliation=aliment["DATE_ENTREE"][:10] if aliment["DATE_ENTREE"] else None,
                            date_sortie=aliment["DATE_SORTIE"][:10] if aliment["DATE_SORTIE"] else None,
                            photo="",
                            statut_familiale="",
                            numero_piece="",
                            code_postal="",
                            ville=aliment["VILLE"],
                            adresse=aliment["ADRESSE"],
                            telephone_fixe=aliment["TELEPHONE_FIXE"],
                            telephone_mobile=aliment["TELEPHONE_MOBILE"],
                            rib="",
                            surprime_ttc="0",
                            plafond_extra="0",
                            plafond_individuel="0",
                            plafond_famille="0",
                            commentaire="",
                            statut=Statut.ACTIF,
                            adherent_principal_id=adherent_principal_id,
                            civilite_id="",
                            pays_activite_professionnelle_id=pays_activite_professionnelle_id,
                            pays_naissance_id=pays_naissance_id,
                            pays_residence_id=pays_residence_id,
                            profession_id=profession_id,
                            qualite_beneficiaire_id=qualite_beneficiaire_id,
                            veos_id_npol=aliment["ID_NPOL"],
                            veos_adherent_principal=aliment["ADHERENT_PRINCIPAL_ID"],
                            veos_adherent_principal_id_per=aliment["ADHERENT_PRINCIPAL_IDPER"],
                            veos_code_aliment=aliment["ID"],
                            veos_code_college=aliment["CD_COLLEGE"],
                            veos_code_formule=aliment["CD_FORMULE"],
                            veos_code_qualite_beneficiaire=aliment["CODE_QUALITE_BENEFICIAIRE"],
                            veos_numero_carte=aliment["NUMERO_CARTE"],
                            observation="CREATED BY IMPORT A - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S'))
                        )
                        new_aliment.save()

                        #si il est lui mﾃｪme adhﾃｩrent principal
                        if aliment["ADHERENT_PRINCIPAL_ID"] == aliment["ID"]:
                            new_aliment.adherent_principal_id = new_aliment.pk
                            new_aliment.save()

                        # dd(formule)

                        if formule:
                            # crﾃｩer une ligne dans aliment_formule
                            aliment_formule = AlimentFormule(
                                aliment_id=new_aliment.pk,
                                formule_id=formule.pk,
                                date_debut=new_aliment.date_affiliation,
                                date_fin=new_aliment.date_sortie,
                            )
                            aliment_formule.save()

                        # Crﾃｩer sa carte
                        carte = Carte(
                            aliment_id=new_aliment.pk,
                            numero=new_aliment.veos_numero_carte,
                            date_edition=new_aliment.date_affiliation,
                        )
                        carte.save()

                        cpt_nouveaux = cpt_nouveaux + 1
                        # dd(new)

                ## le marquﾃｩ comme importﾃｩ
                #aliment.STATUT_IMPORT = 1
                #aliment.save()

                #incrﾃｩmenter le compteur
                cpt_all = cpt_all + 1

    response = {
                'statut': 1,
                'message': 'ALIMENTS IMPORTES AVEC SUCCﾃS',
                'data': {
                         'NOUVEAUX': cpt_nouveaux,
                         'ANCIENS_SANS_CHANGEMENT_FORMULE': cpt_anciens_sans_changement_formule,
                         'ANCIENS_AVEC_CHANGEMENT_FORMULE': cpt_anciens_avec_changement_formule,
                         'TOUS': cpt_all,
                     }
                }

    #return JsonResponse(response)
    return True



def import_aliments_manuellement(request):

    cpt_all = 0
    cpt_anciens_sans_changement_formule = 0
    cpt_anciens_avec_changement_formule = 0
    cpt_nouveaux = 0

    date_du_jour = datetime.now(tz=timezone.utc).date()

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    aliments = AlimentVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_ALIMENT')[:3000]

    #dd(aliments)
    for aliment in aliments:

        # Check if the record already exists based on a unique field (code)
        existing_record = Aliment.objects.filter(veos_code_aliment=aliment.ID_ALIMENT).first()

        adherent_principal = Aliment.objects.filter(veos_code_aliment=aliment.ADHERENT_PRINCIPAL_ID)
        adherent_principal_id = adherent_principal.first().id if adherent_principal else None  # By Richmond / A Vﾃ嘘IFIER

        pays_naissance = Pays.objects.filter(code=aliment.NOM)  # A ajouter au boBy PAYS_ID
        pays_naissance_id = pays_naissance.first().id if pays_naissance else None  # By Richmond / A Vﾃ嘘IFIER

        pays_residence = Pays.objects.filter(code=aliment.NOM)  # A ajouter au boBy
        pays_residence_id = pays_residence.first().id if pays_residence else None  # By Richmond / A Vﾃ嘘IFIER

        pays_activite_professionnelle = Pays.objects.filter(code=aliment.NOM)  # A ajouter au boBy
        pays_activite_professionnelle_id = pays_activite_professionnelle.first().id if pays_activite_professionnelle else None  # By Richmond / A Vﾃ嘘IFIER

        profession = Profession.objects.filter(code=aliment.NOM)  # A ajouter au boBy PROFESSION_ID
        profession_id = profession.first().id if profession else None  # By Richmond / A Vﾃ嘘IFIER

        qualite_beneficiaire = QualiteBeneficiaire.objects.filter(code=aliment.CODE_QUALITE_BENEFICIAIRE)  # A ajouter au boBy CODE_QUALITE_BENEFICIAIRE
        qualite_beneficiaire_id = qualite_beneficiaire.first().id if qualite_beneficiaire else None  # By Richmond / A Vﾃ嘘IFIER

        formules = FormuleGarantie.objects.filter(code=aliment.CD_FORMULE, code__isnull=False)  # A ajouter au boBy CODE_FORMULE
        formule = formules.first() if formules else None

        # dd(aliment.CODE_QUALITE_BENEFICIAIRE)

        # dd(existing_record)
        if existing_record:

            pprint("ALIMENT EXISTANT DEJA EN BASE")
            #dd("ALIMENT EXISTANT DEJA EN BASE")

            existing_record.observation = "UPDATED BY CRON - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S'))
            existing_record.qualite_beneficiaire_id = qualite_beneficiaire_id
            if aliment.DATE_NAISSANCE: existing_record.date_naissance = aliment.DATE_NAISSANCE
            #si la date sortie est renseignﾃｩ c'est qu'il est sortie
            if aliment.DATE_SORTIE:
                existing_record.date_sortie = aliment.DATE_SORTIE

            existing_record.veos_code_qualite_beneficiaire = aliment.CODE_QUALITE_BENEFICIAIRE
            existing_record.save()

            # s'il a changﾃｩ de formule:

            if formule:
                aliment_formule_old = AlimentFormule.objects.filter(aliment_id=existing_record.pk, formule_id=formule.pk).order_by('-id').first()

                if aliment_formule_old:

                    # il a sa meme formule, mettre
                    pprint(aliment_formule_old)
                    pprint("Il a toujours sa mﾃｪme formule:" + aliment_formule_old.formule.code)
                    #dd("Il a toujours sa mﾃｪme formule:" + aliment_formule_old.formule.code)
                    #aliment_formule_old = aliment_formule_olds.first()

                    # si la date sortie est renseignﾃｩ c'est qu'il est sortie de la formule
                    if aliment.DATE_SORTIE:
                        aliment_formule_old.date_fin = aliment.DATE_SORTIE
                        aliment_formule_old.save()
                        #dd(aliment_formule_old)

                    cpt_anciens_sans_changement_formule = cpt_anciens_sans_changement_formule + 1

                else:

                    # sinon il a changﾃｩ de formule
                    pprint("IL A CHANGﾃ DE FORMULE"+ str(existing_record.pk) + " " + str(formule.pk))
                    #dd("IL A CHANGﾃ DE FORMULE"+ str(existing_record.pk) + " " + str(formule.pk))

                    # rﾃｩcupﾃｩrer son ancienne ligne aliment_formule et mettre la date de fin
                    formule_precedente = AlimentFormule.objects.filter(aliment_id=existing_record.pk, date_fin__isnull=True).order_by('-id').first()

                    if formule_precedente:

                        #pas bon
                        '''date_entree = aliment.DATE_ENTREE
                        #date_entree_obj = datetime.strptime(date_entree, '%Y-%m-%d').date()
                        #date_entree = date_entree_obj - timedelta(days=1)
                        date_fin = date_entree - timedelta(days=1)
                        '''

                        #repris
                        date_fin = date_du_jour - timedelta(days=1)

                        formule_precedente.date_fin = date_fin
                        formule_precedente.save()
                        #dd(formule_precedente)


                    # crﾃｩer une ligne dans aliment_formule
                    aliment_formule = AlimentFormule(
                        aliment_id=existing_record.pk,
                        formule_id=formule.pk,
                        date_debut=date_du_jour,
                        date_fin=aliment.DATE_SORTIE if aliment.DATE_SORTIE else None,
                    )
                    aliment_formule.save()

                    cpt_anciens_avec_changement_formule = cpt_anciens_avec_changement_formule + 1

        else:

            pprint("NOUVEAU ALIMENT")
            #dd("NOUVEAU ALIMENT")

            if formule:

                bureau = formule.police.bureau if formule.police else None

                # dd(aliment.DATE_ENTREE[:10])
                new_aliment = Aliment(
                    bureau=bureau,
                    nom=aliment.NOM,
                    prenoms=aliment.PRENOMS,
                    date_naissance=aliment.DATE_NAISSANCE if aliment.DATE_NAISSANCE else None,
                    lieu_naissance="",
                    genre=aliment.GENRE,
                    email=aliment.EMAIL,
                    numero_securite_sociale="",
                    numero=aliment.ID_ALIMENT,  # A verifier
                    numero_famille=aliment.NUMERO_FAMILLE,
                    matricule_employe="",
                    matricule_cie="",
                    date_affiliation=aliment.DATE_ENTREE if aliment.DATE_ENTREE else None,
                    date_sortie=aliment.DATE_SORTIE if aliment.DATE_SORTIE else None,
                    photo="",
                    statut_familiale="",
                    numero_piece="",
                    code_postal="",
                    ville=aliment.VILLE,
                    adresse=aliment.ADRESSE,
                    telephone_fixe=aliment.TELEPHONE_FIXE,
                    telephone_mobile=aliment.TELEPHONE_MOBILE,
                    rib="",
                    surprime_ttc="0",
                    plafond_extra="0",
                    plafond_individuel="0",
                    plafond_famille="0",
                    commentaire="",
                    statut=Statut.ACTIF,
                    adherent_principal_id=adherent_principal_id,
                    civilite_id="",
                    pays_activite_professionnelle_id=pays_activite_professionnelle_id,
                    pays_naissance_id=pays_naissance_id,
                    pays_residence_id=pays_residence_id,
                    profession_id=profession_id,
                    qualite_beneficiaire_id=qualite_beneficiaire_id,
                    veos_id_npol=aliment.ID_NPOL,
                    veos_adherent_principal=aliment.ADHERENT_PRINCIPAL_ID,
                    veos_adherent_principal_id_per=aliment.ADHERENT_PRINCIPAL_IDPER,
                    veos_code_aliment=aliment.ID_ALIMENT,
                    veos_code_college=aliment.CD_COLLEGE,
                    veos_code_formule=aliment.CD_FORMULE,
                    veos_code_qualite_beneficiaire=aliment.CODE_QUALITE_BENEFICIAIRE,
                    veos_numero_carte=aliment.NUMERO_CARTE,
                    observation="CREATED BY IMPORT M - " + str(datetime.now(tz=timezone.utc).strftime('%d/%m/%Y %H:%M:%S'))
                )
                new_aliment.save()

                # si il est lui mﾃｪme adhﾃｩrent principal
                if aliment.ADHERENT_PRINCIPAL_ID == aliment.ID_ALIMENT:
                    new_aliment.adherent_principal_id = new_aliment.pk
                    new_aliment.save()

                # dd(formule)

                if formule:
                    # crﾃｩer une ligne dans aliment_formule
                    aliment_formule = AlimentFormule(
                        aliment_id=new_aliment.pk,
                        formule_id=formule.pk,
                        date_debut=new_aliment.date_affiliation,
                        date_fin=new_aliment.date_sortie,
                    )
                    aliment_formule.save()

                # Crﾃｩer sa carte
                carte = Carte(
                    aliment_id=new_aliment.pk,
                    numero=new_aliment.veos_numero_carte,
                    date_edition=new_aliment.date_affiliation,
                )
                carte.save()

                cpt_nouveaux = cpt_nouveaux + 1
                # dd(new)

        ## le marquﾃｩ comme importﾃｩ
        aliment.STATUT_IMPORT = 1
        aliment.save()

        #incrﾃｩmenter le compteur
        cpt_all = cpt_all + 1

    response = {
                'statut': 1,
                'message': 'ALIMENTS IMPORTES AVEC SUCCﾃS',
                'data': {
                         'NOUVEAUX': cpt_nouveaux,
                         'ANCIENS_SANS_CHANGEMENT_FORMULE': cpt_anciens_sans_changement_formule,
                         'ANCIENS_AVEC_CHANGEMENT_FORMULE': cpt_anciens_avec_changement_formule,
                         'TOUS': cpt_all,
                     }
                }

    return JsonResponse(response)


def get_prestataire_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_PRESTATAIRE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    #dd(api_response)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        prestataires = api_response['responses'][0]['beans']

        for prestataire in prestataires:
            # Check if the record already exists based on a unique field (code)
            existing_record = Prestataire.objects.filter(code=prestataire["CODE"]).first()

            bureaux = Bureau.objects.filter(code=prestataire['SOCIETE'])
            bureau = bureaux.first() if bureaux else None

            type_prestataires = TypePrestataire.objects.filter(code=prestataire['TYPE_PRESTATAIRE'])
            type_prestataire = type_prestataires.first() if type_prestataires else None

            secteurs = Secteur.objects.filter(code=prestataire['SECTEUR'])
            secteur = secteurs.first() if secteurs else None

            #dd(type_prestataire)
            if existing_record:
                existing_record.code = prestataire['CODE']
                existing_record.name = prestataire['NAME']
                existing_record.telephone = prestataire['TELEPHONE']
                existing_record.fax = prestataire['FAX']
                existing_record.email = prestataire['EMAIL']
                existing_record.addresse = prestataire['ADRESSES']
                if type_prestataire: existing_record.type_prestataire_id = type_prestataire
                existing_record.id_per = prestataire['ID_PER']
                existing_record.veos_code_soc = prestataire['SOCIETE']
                existing_record.veos_type_pres = prestataire['TYPE_PRESTATAIRE']
                existing_record.save()

            else:

                if bureau:

                    new = Prestataire(code=prestataire['CODE'],
                                    name=prestataire['NAME'],
                                    telephone=prestataire['TELEPHONE'],
                                    fax=prestataire['FAX'],
                                    email=prestataire['EMAIL'],
                                    addresse=prestataire['ADRESSES'],
                                    bureau=bureau,
                                    type_prestataire=type_prestataire,
                                    secteur=secteur,
                                    id_per=prestataire['ID_PER'],
                                    veos_code_soc=prestataire['SOCIETE'],
                                    veos_type_pres=prestataire['TYPE_PRESTATAIRE'],
                                    )
                    new.save()

                else:
                    CronLog.objects.create(action="add", table="prestataire", description="Prestataire  " + prestataire['CODE'] + " non importﾃｩ: code bureau " + prestataire['SOCIETE'] + " non trouvﾃｩ").save()

        return True


def import_prestataires_manuellement(request):

    prestataires_veos = PrestataireVeos.objects.filter(STATUT_IMPORT=False)

    cpt_anciens = 0
    cpt_nouveaux = 0

    for prestataire in prestataires_veos:
            # dd(prestataire)
            # Check if the record already exists based on a unique field (code)
            existing_record = Prestataire.objects.filter(id_per=prestataire.ID_PER).first()

            bureaux = Bureau.objects.filter(code=prestataire.SOCIETE)
            bureau = bureaux.first() if bureaux else None

            type_prestataires = TypePrestataire.objects.filter(veos_code=prestataire.TYPE_PRESTATAIRE)
            type_prestataire = type_prestataires.first() if type_prestataires else None


            secteurs = Secteur.objects.filter(code=prestataire.SECTEUR)
            secteur = secteurs.first() if secteurs else None

            if existing_record:
                existing_record.code = prestataire.CODE
                existing_record.name = prestataire.NAME
                existing_record.telephone = prestataire.TELEPHONE
                existing_record.fax = prestataire.FAX
                existing_record.email = prestataire.EMAIL
                existing_record.addresse = prestataire.ADRESSE
                if type_prestataire: existing_record.type_prestataire_id = type_prestataire
                existing_record.id_per = prestataire.ID_PER
                existing_record.veos_code_soc = prestataire.SOCIETE
                existing_record.veos_type_pres = prestataire.TYPE_PRESTATAIRE
                existing_record.save()
                cpt_anciens += 1

            else:

                if bureau:

                    new = Prestataire(code=prestataire.CODE,
                                    name=prestataire.NAME,
                                    telephone=prestataire.TELEPHONE,
                                    fax=prestataire.FAX,
                                    email=prestataire.EMAIL,
                                    addresse=prestataire.ADRESSE,
                                    bureau=bureau,
                                    type_prestataire=type_prestataire,
                                    secteur=secteur,
                                    id_per=prestataire.ID_PER,
                                    veos_code_soc=prestataire.SOCIETE,
                                    veos_type_pres=prestataire.TYPE_PRESTATAIRE,
                                    )
                    new.save()

                    cpt_nouveaux += 1


            prestataire.STATUT_IMPORT = True
            prestataire.save()

    cpt_all = cpt_anciens + cpt_nouveaux

    response = {
        'statut': 1,
        'message': 'PRESTATAIRES IMPORTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'ANCIENS MISE ﾃ JOUR': cpt_anciens,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
        }
    }

    return JsonResponse(response)


def import_prescripteurs_manuellement(request):

    #TODO DOES
    prescripteurs_veos = PrescripteurVeos.objects.filter(STATUT_IMPORT=False)

    cpt_anciens = 0
    cpt_nouveaux = 0

    # dd(prescripteurs_veos)

    for prescripteur in prescripteurs_veos:
        # Check if the record already exists based on a unique field (code)
        existing_prescripteur = Prescripteur.objects.filter(veos_id_per=prescripteur.id_per).first()


        # Prestataire
        prestataires = Prestataire.objects.filter(id_per=prescripteur.ID_PRESTA)
        prestataire = prestataires.first() if prestataires else None

        specialites = Specialite.objects.filter(code=prescripteur.specialite)
        specialite = specialites.first() if specialites else None

        # dd(existing_prescripteur)

        if prestataire:

            if existing_prescripteur:
                existing_prescripteur.nom = prescripteur.nom
                existing_prescripteur.prenoms = prescripteur.prenom
                existing_prescripteur.numero_ordre = prescripteur.numero
                existing_prescripteur.telephone = prescripteur.telephone
                existing_prescripteur.email = prescripteur.email
                existing_prescripteur.specialite = specialite
                existing_prescripteur.veos_code_specialite = prescripteur.specialite
                existing_prescripteur.veos_id_per = prescripteur.id_per

                prescripteur.save()

                #TODO, vﾃｩrifie qu'il n'est pas dﾃｩjﾃ liﾃｩ ﾃ ace prestataire avant de le lier encore: un prescripteur peut ﾃｪtre liﾃｩ a plusieurs prestatairs
                pp_existant = PrescripteurPrestataire.objects.filter(prescripteur=existing_prescripteur, prestataire=prestataire)
                if not pp_existant:
                    pp = PrescripteurPrestataire(prescripteur=existing_prescripteur, prestataire=prestataire)
                    pp.save()

                cpt_anciens += 1

            else:

                new_prescripteur = Prescripteur(
                    nom=prescripteur.nom,
                    prenoms=prescripteur.prenom,
                    numero_ordre=prescripteur.numero,
                    telephone=prescripteur.telephone,
                    email=prescripteur.email,
                    specialite = specialite,
                    veos_code_specialite = prescripteur.specialite,
                    veos_id_per = prescripteur.id_per

                )

                new_prescripteur.save()

                pp = PrescripteurPrestataire(prescripteur=new_prescripteur, prestataire=prestataire)
                pp.save()
                cpt_nouveaux += 1


        prescripteur.STATUT_IMPORT = True
        prescripteur.save()

    cpt_all = cpt_anciens + cpt_nouveaux

    response = {
        'statut': 1,
        'message': 'PRESTATAIRES IMPORTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'ANCIENS MISE ﾃ JOUR': cpt_anciens,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
        }
    }

    return JsonResponse(response)


def import_utilisateurs_manuellement(request):

    cpt_anciens = 0
    cpt_nouveaux = 0

    utilisateurs_veos = UtilisateurVeos.objects.filter(STATUT_IMPORT=False)

    #dd(utilisateurs_veos)
    for utilisateur in utilisateurs_veos:
        user_exist = User.objects.filter(username=utilisateur.LOGIN).exists()

        # dd(user_exist)

        if user_exist:
            cpt_anciens += 1

            # mettre le statut ﾃ importﾃｩ
            utilisateur.STATUT_IMPORT = True
            utilisateur.save()

        else:
            bureau = Bureau.objects.filter(code=utilisateur.SOCIETE).first()
            prestataire = Prestataire.objects.filter(id_per=utilisateur.ID_PRESTA).first()
            # dd(prestataire)
            if prestataire:
                if prestataire.veos_type_pres == "CSOIN":
                    group = Group.objects.get(id=4)
                elif prestataire.veos_type_pres == "PHARM":
                    group = Group.objects.get(id=7)
                elif prestataire.veos_type_pres == "LABOR":
                    group = Group.objects.get(id=8)
                elif prestataire.veos_type_pres == "IMAGE":
                    group = Group.objects.get(id=9)
                elif prestataire.veos_type_pres == "OPTIQ":
                    group = Group.objects.get(id=10)
                elif prestataire.veos_type_pres == "DENTA":
                    group = Group.objects.get(id=13)
                else:
                    group = None


                # group = AuthGroup.objects.filter()
                # group = Group.objects.filter()
                # group = Group.objects.get(code=prestataire.veos_type_pres)

                # dd(group)

                new_utilisateur = User.objects.create(
                    username = utilisateur.LOGIN,
                    password = "pbkdf2_sha256$260000$0735abKRc6b084e19bnGIk$Dhb6Wwa1b/iMreoN+mgielG/4BvS2ir80Rr0iLelHUo=",
                    is_superuser = False,
                    first_name = utilisateur.PRENOM if utilisateur.PRENOM else "",
                    last_name = utilisateur.NOM if utilisateur.NOM else "",
                    email = utilisateur.EMAIL,
                    is_staff = True,
                    is_active = True,
                    bureau_id = bureau.id,
                    prestataire_id = prestataire.id,
                    veos_code = utilisateur.CODE,
                    veos_code_prestataire = utilisateur.NUM_PRESTA,
                    type_utilisateur_id = 2,
                    password_type = 'CUSTOM',
                )

                new_utilisateur.save()
                # new_utilisateur.groups.add()

                # Renseigner configuration_user_groups
                new_utilisateur.groups.add(group)
                cpt_nouveaux += 1

                #mettre le statut ﾃ importﾃｩ
                utilisateur.STATUT_IMPORT = True
                utilisateur.save()

    cpt_all = cpt_anciens + cpt_nouveaux

    response = {
        'statut': 1,
        'message': 'UTILISATEURS IMPORTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'ANCIENS MISE ﾃ JOUR': cpt_anciens,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
        }
    }

    return JsonResponse(response)


def import_utilisateurs_grh_manuellement(request):

    cpt_anciens = 0
    cpt_nouveaux = 0

    utilisateurs_grh_veos = UtilisateurGrhVeos.objects.filter(STATUT_IMPORT=False)

    #dd(utilisateurs_veos)
    for utilisateur in utilisateurs_grh_veos:

        user = User.objects.filter(username=utilisateur.LOGIN).first()
        
        list_client_associated = user.client_grh.all() if user else None

        bureau = Bureau.objects.filter(code=utilisateur.BUREAU).first()

        list_code_client = [code.strip() for code in utilisateur.CODE_CLIENT.strip().replace(';', ',').split(',')] if utilisateur.CODE_CLIENT else ""

        for code_client in list_code_client:
            
            client = Client.objects.filter(code=code_client).first()

            if user:
                
                if list_client_associated and client in list_client_associated:
                    cpt_anciens += 1
                else:
                    user.client_grh.add(client)
                    #user.save()
                    cpt_nouveaux += 1

                # mettre le statut ﾃ importﾃｩ
                utilisateur.STATUT_IMPORT = True
                utilisateur.save()

            else:

                if client:

                    new_utilisateur = User.objects.create(
                        username = utilisateur.LOGIN,
                        password = "pbkdf2_sha256$260000$0735abKRc6b084e19bnGIk$Dhb6Wwa1b/iMreoN+mgielG/4BvS2ir80Rr0iLelHUo=",
                        is_superuser = False,
                        # first_name = utilisateur.PRENOM if utilisateur.PRENOM else "",
                        last_name = utilisateur.NOM if utilisateur.NOM else "",
                        # email = utilisateur.LOGIN,
                        is_staff = True,
                        is_active = True,
                        bureau_id = bureau.id,
                        utilisateur_grh = client,
                        # prestataire_id = prestataire.id,
                        # veos_code = utilisateur.CODE,
                        # veos_code_prestataire = utilisateur.NUM_PRESTA,
                        type_utilisateur_id = 2,
                        password_type = 'CUSTOM',
                    )
                    new_utilisateur.save()
                    new_utilisateur.client_grh.add(client)
                    
                    user = new_utilisateur

                    # dd(new_utilisateur)

                    #new_utilisateur.save()
                    # new_utilisateur.groups.add()

                    # Renseigner configuration_user_groups
                    # new_utilisateur.groups.add(group)
                    cpt_nouveaux += 1

                    #mettre le statut ﾃ importﾃｩ
                    utilisateur.STATUT_IMPORT = True
                    utilisateur.save()

    cpt_all = cpt_anciens + cpt_nouveaux

    response = {
        'statut': 1,
        'message': 'UTILISATEURS GRH IMPORTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'ANCIENS MISE ﾃ JOUR': cpt_anciens,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
        }
    }

    return JsonResponse(response)


def import_utilisateurs_prestataire_manuellement(request):

    cpt_anciens = 0
    cpt_nouveaux = 0

    utilisateurs_veos = ComptePrestataireVeos.objects.filter(STATUT_IMPORT=False)

    #dd(utilisateurs_veos)
    for utilisateur in utilisateurs_veos:
        list_email_user = [email.strip() for email in utilisateur.EMAIL.strip().replace(';', ',').split(',')] if utilisateur.EMAIL else ""

        prestataire = Prestataire.objects.filter(code=utilisateur.CODE_PRESTATAIRE).first()
        
        secteur_prestataire = Secteur.objects.filter(libelle=utilisateur.SECTEUR_PRESTATAIRE).first()
        
        type_prestataire = TypePrestataire.objects.filter(name=utilisateur.TYPE_PRESTATAIRE).first()
        if utilisateur.TYPE_PRESTATAIRE == "MEDECIN":
            type_prestataire = TypePrestataire.objects.filter(name="CENTRE DE SOINS").first()

        # dd(prestataire)
        if prestataire:
            if prestataire.veos_type_pres == "CSOIN" or utilisateur.TYPE_PRESTATAIRE == "CENTRE DE SOINS":
                group = Group.objects.get(id=4)
            elif prestataire.veos_type_pres == "PHARM" or utilisateur.TYPE_PRESTATAIRE == "PHARMACIE":
                group = Group.objects.get(id=7)
            elif prestataire.veos_type_pres == "LABOR" or utilisateur.TYPE_PRESTATAIRE == "LABORATOIRE":
                group = Group.objects.get(id=8)
            elif prestataire.veos_type_pres == "MEDEC" or utilisateur.TYPE_PRESTATAIRE == "MEDECIN":
                group = Group.objects.get(id=4)
            elif prestataire.veos_type_pres == "IMAGE" or utilisateur.TYPE_PRESTATAIRE == "IMAGERIE":
                group = Group.objects.get(id=9)
            elif prestataire.veos_type_pres == "OPTIQ" or utilisateur.TYPE_PRESTATAIRE == "OPTIQUE":
                group = Group.objects.get(id=10)
            elif prestataire.veos_type_pres == "DENTA" or utilisateur.TYPE_PRESTATAIRE == "CENTRE DENTAIRE":
                group = Group.objects.get(id=13)
            else:
                group = None

            # update prestataire 
            prestataire.secteur = secteur_prestataire
            prestataire.type_prestataire = type_prestataire
            prestataire.name = utilisateur.NOM_PRESTATAIRE
            prestataire.rb_ordre = utilisateur.REMBOURSEMENT_ORDRE_DE 
            prestataire.save()    

            for email_user in list_email_user:
                # dd(user_exist)
                user_exist = User.objects.filter(username=email_user).exists()

                if user_exist:
                    cpt_anciens += 1

                    # mettre le statut ﾃ importﾃｩ
                    utilisateur.STATUT_IMPORT = True
                    utilisateur.save()

                else:

                    # group = AuthGroup.objects.filter()
                    # group = Group.objects.filter()
                    # group = Group.objects.get(code=prestataire.veos_type_pres)

                    # dd(group)

                    new_utilisateur = User.objects.create(
                        username = email_user,
                        password = "pbkdf2_sha256$260000$0735abKRc6b084e19bnGIk$Dhb6Wwa1b/iMreoN+mgielG/4BvS2ir80Rr0iLelHUo=",
                        is_superuser = False,
                        #first_name = utilisateur.PRENOM if utilisateur.PRENOM else "",
                        #last_name = utilisateur.NOM if utilisateur.NOM else "",
                        email = email_user,
                        is_staff = True,
                        is_active = True,
                        bureau_id = prestataire.bureau.id,
                        prestataire_id = prestataire.id,
                        #veos_code = utilisateur.CODE,
                        veos_code_prestataire = utilisateur.CODE_PRESTATAIRE,
                        type_utilisateur_id = 2,
                        password_type = 'CUSTOM',
                    )

                    new_utilisateur.save()
                    # new_utilisateur.groups.add()

                    # Renseigner configuration_user_groups
                    new_utilisateur.groups.add(group)
                    cpt_nouveaux += 1

                    #mettre le statut ﾃ importﾃｩ
                    utilisateur.STATUT_IMPORT = True
                    utilisateur.save()

    cpt_all = cpt_anciens + cpt_nouveaux

    response = {
        'statut': 1,
        'message': 'UTILISATEURS IMPORTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'ANCIENS MISE ﾃ JOUR': cpt_anciens,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
        }
    }

    return JsonResponse(response)


def import_changement_formule_manuellement(request):

    cpt_anciens = 0 #ceci stockera le nombre des formules qui ne seront pas impactﾃｩes car dﾃｩjﾃ conforme
    cpt_nouveaux = 0 #ceci stockera le nombre des formules qui ont ete tout simplement crﾃｩﾃｩes nouvellement
    cpt_desactives = 0 #ceci stockera le nombre des formules qui sont a proscrire
    cpt_updates = 0#ceci stockera le nombre des formules qui ont subit une mise a jour 

    changement_formules = ChangementFormule.objects.filter(STATUT_IMPORT=False) #recuperer toutes les donnees pas encore importes

    new = []

    """
    # Exportation excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="LISTE_CHANGEMENT_FORMULE"__{:%d:%m:%Y}.xlsx"'.format(datetime.now())

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CHANGEMENT FORMULE'

    # Write header row
    header = [
        'CARTE',
        'CD FORMULE',
        'LBL FORMULE',
        'DATE FIN',
    ]
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data rows
    data = []
    """

    #dd(utilisateurs_veos)
    #parcours
    for aliment_formule in changement_formules:

        carte = Carte.objects.filter(numero=aliment_formule.NUMERO_CARTE).first() #la carte associee pour l'aliment

        formule = FormuleGarantie.objects.filter(code=aliment_formule.CD_FORMULE).first() #la formule associee

        if carte and formule:

            old_aliment_formules = AlimentFormule.objects.filter(aliment=carte.aliment, formule__police=formule.police).exclude(formule=formule, statut=Statut.ACTIF, statut_validite=StatutValidite.VALIDE) #recuperer tous les aliment_formules correspondant a cet aliment
            
            #checkons si l'aliment_formule n'est pas deja conforme en base
            old_same_aliment = AlimentFormule.objects.filter(aliment=carte.aliment, formule=formule, statut=Statut.ACTIF, statut_validite=StatutValidite.VALIDE).first()

            if old_same_aliment:
                #pas besoin de traitement si aliment_formule conforme
                cpt_anciens += 1
            else:
                for old_aliment_formule in old_aliment_formules:
                    if old_aliment_formule.statut==Statut.ACTIF and old_aliment_formule.statut_validite==StatutValidite.VALIDE:
                        old_aliment_formule.statut = Statut.INACTIF
                        if old_aliment_formule.date_fin is None:
                            old_aliment_formule.date_fin = aliment_formule.DATE_DEBUT - timedelta(1) #-1 jour
                        old_aliment_formule.motif = str(old_aliment_formule.motif) + str(aliment_formule.MOTIF)
                        old_aliment_formule.save()
                        cpt_desactives += 1
                
                new_aliment_formule = AlimentFormule.objects.create(
                    aliment = carte.aliment,
                    formule = formule,
                    date_debut = formule.date_debut,
                    statut = Statut.ACTIF,
                    statut_validite = StatutValidite.VALIDE,
                )
                new_aliment_formule.save()

                """
                data_item = [
                    carte.numero,
                    formule.code,
                    formule.libelle,
                    aliment_formule.DATE_DEBUT - timedelta(1),
                ]
                data.append(data_item)
                """

                if old_aliment_formules.count() > 0:
                    cpt_updates += 1 
                else:
                    cpt_nouveaux += 1 
                aliment_formule.STATUT_IMPORT = True
                aliment_formule.save()   
    
    """
    for row_num, row in enumerate(data, 1):
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
    """

    cpt_all = cpt_anciens + cpt_nouveaux + cpt_updates
    #dd(new)

    response = {
        'statut': 1,
        'message': 'CHANGEMENTS FORMULE EXﾃ韻UTﾃ唄 AVEC SUCCﾃS',
        'data': {
            'RAS': cpt_anciens,
            'MISES ﾃ JOUR': cpt_updates,
            'NOUVEAUX': cpt_nouveaux,
            'TOUS': cpt_all,
            'DECHETS': cpt_desactives,
        }
    }

    return JsonResponse(response)


def import_quittances_manuellement(request):

    cpt_anciens = 0
    cpt_nouveaux = 0
    cpt_failed = 0

    quittances_veos = QuittanceVeos.objects.filter(STATUT_IMPORT=False)[:1000]

    if quittances_veos:

        #Parcourrir les quittances
        for quittance_veos in quittances_veos:
            quittance_existant = Quittance.objects.filter(numero=quittance_veos.ID_NPOL).first()

            if quittance_existant:
                cpt_anciens += 1

                print("Mise ﾃ jour de la quittance " + quittance_veos.NUMERO_QUITTANCE)

                quittance_veos.STATUT_IMPORT = True
                quittance_veos.save()

            else:

                print("Crﾃｩation de quittance " + quittance_veos.NUMERO_QUITTANCE)

                police = Police.objects.filter(veos_id_npol=quittance_veos.ID_NPOL).first()

                if police:

                    type_quittance = TypeQuittance.objects.filter(code_veos=quittance_veos.CODE_TYPE_QUITTANCE).first()
                    nature_quittance = NatureQuittance.objects.filter(code_veos=quittance_veos.CODE_NATURE_QUITTANCE).first()
                    compagnie = Compagnie.objects.filter(code=quittance_veos.NUMERO_COMPAGNIE).first()

                    #dd(nature_quittance)

                    devise = police.client.pays.devise

                    status_quittance = StatutQuittance.PAYE if quittance_veos.LIBELLE_SITUATION_CLIENT == "Soldﾃｩe" else StatutQuittance.IMPAYE

                    new_quittance = Quittance(
                        bureau=police.bureau,
                        # created_by=request.user.id,
                        type_quittance=type_quittance,
                        nature_quittance=nature_quittance,
                        police=police,
                        compagnie=compagnie,
                        devise=devise,
                        taxe=quittance_veos.TAXES,
                        numero=quittance_veos.NUMERO_QUITTANCE,
                        prime_ht=quittance_veos.PRIME_NETTE,
                        cout_police_courtier=quittance_veos.ACCESSOIRES,
                        cout_police_compagnie=quittance_veos.ACCESSOIRES_CIE,
                        #autres_taxes=quittance_veos.TSVL +  quittance_veos.CSS +  CSS_INOV,
                        prime_ttc=quittance_veos.PRIME_TTC,
                        commission_courtage=quittance_veos.COMMISSION if quittance_veos.PRIME_TTC else 0,
                        commission_gestion=quittance_veos.COMMISSION if not quittance_veos.PRIME_TTC else 0,
                        # commission_intermediaires=quittance_veos,
                        montant_regle=quittance_veos.PRIME_TTC if quittance_veos.SOLDE == 0 else 0, #A confirmer
                        solde=quittance_veos.SOLDE,
                        date_emission=quittance_veos.DATE_EMIS[:10],
                        date_debut=quittance_veos.DATE_DEBUT[:10],
                        date_fin=quittance_veos.DATE_FIN[:10],
                        statut=status_quittance,
                        import_stats=True
                    )

                    new_quittance.save()


                    quittance_veos.STATUT_IMPORT = True
                    quittance_veos.save()


                    cpt_nouveaux += 1
                    print("Ok, Crﾃｩation reussi !")

                else:
                    error_message = "Police introuvable"

    else:

        response = {
            'statut': 0,
            'message': "PAS DE DONNﾃ右S DISPONIBLES",
        }

    dd(quittances_veos)
    return True


def get_prescripteurs_lists():

    today = datetime.now(tz=timezone.utc).strftime('%d/%m/%Y')
    yesterday = (datetime.now(tz=timezone.utc) - timedelta(days=30)).strftime('%d/%m/%Y')

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_PRESCRIPTEUR_TACHE_CRON",
                "params": {
                    'date_debut':yesterday,
                    'date_fin':today
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    #dd(api_response)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        prescripteurs = api_response['responses'][0]['beans']

        # dd(prescripteurs)
        for prescripteur in prescripteurs:
            # Check if the record already exists based on a unique field (code)
            existing_record = Prescripteur.objects.filter(veos_id_per=prescripteur["id_per"]).first()

            prestataires = Prestataire.objects.filter(code=prescripteur['id_presta'])
            prestataire = prestataires.first() if prestataires else None

            specialites = Specialite.objects.filter(code=prescripteur['specialite'])
            specialite = specialites.first() if specialites else None


            # dd(existing_record)
            if existing_record:
                #existing_record.code = client['CODE']
                #existing_record.libelle = client['NOM']
                #existing_record.save

                #L'ajouter sur d'autres prestataires
                if prestataire:

                    pp_existant = PrescripteurPrestataire.objects.filter(prescripteur_id=existing_record.pk, prestataire_id=prestataire.pk)

                    if not pp_existant:
                        # une ligne dans prescripteur_prestataire
                        prescripteur_of_new_presta = PrescripteurPrestataire(
                            prescripteur=prescripteur,
                            prestataire=prestataire,
                        )


            else:

                new_prescripteur = Prescripteur(
                        veos_code_specialite=prescripteur['specialite'],
                        veos_id_per=prescripteur['id_per'],
                        nom=prescripteur['nom'],
                        prenoms=prescripteur['prenom'],
                        specialite=specialite,
                        numero_ordre=prescripteur['numero'],
                        telephone=prescripteur['telephone'],
                        email=prescripteur['email'],
                        statut=Statut.ACTIF,
                    )

                new_prescripteur.save()

                if prestataire:
                    #une ligne dans prescripteur_prestataire
                    new_pp = PrescripteurPrestataire(
                            prescripteur=new_prescripteur,
                            prestataire=prestataire,
                        )

                    new_pp.save()

        return True




def get_sinistres_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    boby_name = "WS_SINISTRE_TACHE_CRON" #WS_CRON_SINISTRE_TEST WS_MAJ_TABLE_SINISTRE
    param = {}

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    sinistres = query_builder_tools(boby_name, param)

    pprint(sinistres)

    '''
    sinistres = [
        {
            'ID_SIN': '2315841',
            'NUM_SIN': '328202-BN0123-SP',
            'PART_INOV': '0',
            'PART_ASSURE': '5825',
            'FRAIS_REEL': '5825',
            'TICKET_MODERATEUR': '0',
            'DEPASSEMENT': '5825',
            'DATE_SOINS': '2023-12-04',
            'CD_ACTE': 'G38147BN01',
            'LIBELLE_ACTE': 'PYRANTOX 125MG B/6COMP',
            'CD_AFFECTION': None,
            'LIB_AFFECTION': None,
            'CD_SPECIALITE': None,
            'LIB_SPECIALITE': None,
            'CATEG_PRESTA': 'PHARM',
            'LIB_CATEG_PRESTA': 'PHARMACIE',
            'IDPER_PRESTA': '12429169',
            'NOM_PRESTA': 'PHARMACIE HEVIE',
            'STATUT_PEC': 'ACCEPTEE',
            'STATUT_REMB': None,
            'MONTANT_ORDON': '0',
            'MONTANT_REMB': None,
            'ACTE_CODE_GROUPE_INOV': 'PHARMACIE',
            'ACTE_LIBELLE_GROUPE_INOV': 'PHARMACIE / VITAMINE / VACCINS',
            'ID_MEDTRAITANT': None,
            'NOM_MEDTRAITANT': None,
            'ID_NPOL': '107718',
            'ID_ADH': '225625',
            'STATUS': 'VALIDE'
        },

    ]
    '''

    #dd(sinistres)
    if sinistres:
        cpt = 0

        for sv in sinistres:

            # Check if the record already exists based on a unique field (code)
            existing_record = Sinistre.objects.filter(veos_id_sin=sv['ID_SIN']).first()

            pprint("- Import du sinistre Nﾂｰ" + sv['NUM_SIN'])

            # dd(existing_record)
            if existing_record:

                pprint("-- LE SINISTRE NUMERO : " + str(sv['NUM_SIN']) + " EXISTE DEJA ")
                #logger.info("-- LE SINISTRE NUMERO : " + str(sv['NUM_SIN']) + " EXISTE DEJA ")
                #existing_record.code = client['CODE']
                #existing_record.libelle = client['NOM']
                #existing_record.save
                #pass

            else:

                #dd(sv)

                prestataires = Prestataire.objects.filter(id_per=sv['IDPER_PRESTA'])
                prestataire = prestataires.first() if prestataires else None

                prescripteurs = Prescripteur.objects.filter(veos_id_per=sv['ID_MEDTRAITANT'])
                prescripteur = prescripteurs.first() if prescripteurs else None

                compagnies = Compagnie.objects.filter(code=sv['ID_MEDTRAITANT'])
                compagnie = compagnies.first() if compagnies else None

                actes = Acte.objects.filter(code=sv['CD_ACTE'])
                acte = actes.first() if actes else None

                affections = Affection.objects.filter(code=sv['CD_AFFECTION'])
                affection = affections.first() if actes else None

                aliments = Aliment.objects.filter(veos_code_aliment=sv['ID_ADH'])
                aliment = aliments.first() if aliments else None

                adherent_principal = aliment.adherent_principal if aliment else None # a corriger renvoi si vide nonetype has no attr.

                polices = Police.objects.filter(veos_id_npol=sv['ID_NPOL'])
                police = polices.first() if polices else None


                criteres_dates = Q(date_debut_effet__lte=sv['DATE_SOINS']) & (Q(date_fin_effet__gte=sv['DATE_SOINS']) | Q(date_fin_effet__isnull=True))
                periodes_couvertures = PeriodeCouverture.objects.filter(police=police, statut=Statut.ACTIF).filter(criteres_dates)
                periode_couverture = periodes_couvertures.first() if periodes_couvertures else None

                #dd(periode_couverture.id)

                #pprint(police)
                #pprint(acte)
                #pprint(aliment)
                if police and acte and aliment and sv['STATUS'] == "VALIDE":

                    new = Sinistre(
                        veos_id_sin=sv['ID_SIN'],
                        veos_id_npol=sv['ID_NPOL'],
                        veos_code_aliment=sv['ID_ADH'],
                        #veos_code_cie=sv['NUM_COMPAGNIE'],
                        veos_code_acte=sv['CD_ACTE'],
                        veos_code_affection=sv['CD_AFFECTION'],
                        veos_code_prestataire=sv['IDPER_PRESTA'],
                        veos_code_prescripteur=sv['ID_MEDTRAITANT'],
                        created_by=None,
                        updated_price_by=None,
                        approuved_by=None,
                        served_by=None,
                        dossier_sinistre=None,
                        aliment=aliment,
                        adherent_principal=adherent_principal if adherent_principal else None,
                        compagnie_id=compagnie.pk if compagnie else 1, #compagnie,
                        police=police,
                        periode_couverture=periode_couverture,
                        formulegarantie=None,
                        bareme=None,
                        acte=acte,
                        medicament=None,
                        affection=affection,
                        prestataire=prestataire,
                        prescripteur=prescripteur,
                        numero=sv['NUM_SIN'],
                        type_sinistre="",
                        prix_unitaire=0,
                        frais_reel=sv['FRAIS_REEL'],
                        ticket_moderateur=sv['TICKET_MODERATEUR'],
                        depassement=sv['DEPASSEMENT'],
                        nombre_demande=1,
                        nombre_accorde=1,
                        part_assure=sv['PART_ASSURE'],
                        part_compagnie=sv['PART_INOV'],
                        date_survenance=sv['DATE_SOINS'],
                        statut='ACCORDE' if sv['STATUS'] == 'VALIDE' else None,
                        observation="CREATED BY CRON TASK"
                        )

                    new.save()

                    cpt = cpt + 1;

                    pprint(" -Le sinistre " + sv['NUM_SIN'] + " a ﾃｩtﾃｩ importﾃｩ avec succﾃｨs")

                else:
                    pprint(" -Le sinistre " + sv['NUM_SIN'] + " n'a pas ﾃｩtﾃｩ enregistrﾃｩ, absence de police, acte ou aliment")


        response = {
            'statut': 1,
            'message': "SINISTRES IMPORTﾃ唄",
            'data': {
                'nombre_sinistres_crees': cpt,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "PAS DE DONNﾃ右S DISPONIBLES",
        }




def import_sinistre_manuellement_old(request):

    sinistres_veos = SinistreVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_SIN')[:1000]
    cpt = 0

    #dd(sinistres_veos)
    if sinistres_veos:

        for sv in sinistres_veos:

            # Check if the record already exists based on a unique field (code)
            existing_record = Sinistre.objects.filter(veos_id_sin=sv.ID_SIN).first()

            # dd(existing_record)
            if existing_record:

                pprint("-- LE SINISTRE " + str(sv.NUMERO) + " EXISTE DEJA ")
                #existing_record.code = client['CODE']
                #existing_record.libelle = client['NOM']
                #existing_record.save
                pass

            else:

                #dd(sv)

                prestataires = Prestataire.objects.filter(id_per=sv.PRESTATAIRE_ID)
                prestataire = prestataires.first() if prestataires else None

                prescripteurs = Prescripteur.objects.filter(veos_id_per=sv.PRESCRIPTEUR_ID)
                prescripteur = prescripteurs.first() if prescripteurs else None

                compagnies = Compagnie.objects.filter(code=sv.NUM_COMPAGNIE)
                compagnie = compagnies.first() if compagnies else None

                actes = Acte.objects.filter(code=sv.ACTE_ID)
                acte = actes.first() if actes else None

                aliments = Aliment.objects.filter(veos_code_aliment=sv.ALIMENT_ID)
                aliment = aliments.first() if aliments else None

                adherent_principal = aliment.adherent_principal if aliment else None # a corriger renvoi si vide nonetype has no attr.

                polices = Police.objects.filter(veos_id_npol=sv.POLICE_ID)
                police = polices.first() if polices else None


                criteres_dates = Q(date_debut_effet__lte=sv.DATE_SURVENANCE) & (Q(date_fin_effet__gte=sv.DATE_SURVENANCE) | Q(date_fin_effet__isnull=True))
                periodes_couvertures = PeriodeCouverture.objects.filter(police=police, statut=Statut.ACTIF).filter(criteres_dates)
                periode_couverture = periodes_couvertures.first() if periodes_couvertures else None

                #dd(periode_couverture.id)

                if acte and aliment and sv.STATUT == "ACCEPTEE":

                    new = Sinistre(
                        veos_id_sin=sv.ID_SIN,
                        veos_id_npol=sv.POLICE_ID,
                        veos_code_aliment=sv.ALIMENT_ID,
                        veos_code_cie=sv.NUM_COMPAGNIE,
                        veos_code_acte=sv.ACTE_ID,
                        veos_code_affection=sv.AFFECTION_ID,
                        veos_code_prestataire=sv.PRESTATAIRE_ID,
                        veos_code_prescripteur=sv.PRESCRIPTEUR_ID,
                        created_by=None,
                        updated_price_by=None,
                        approuved_by=None,
                        served_by=None,
                        dossier_sinistre=None,
                        aliment=aliment,
                        adherent_principal=adherent_principal if adherent_principal else None,
                        compagnie=compagnie,
                        police=police,
                        periode_couverture=periode_couverture,
                        formulegarantie=None,
                        bareme=None,
                        acte=acte,
                        medicament=None,
                        affection=None,
                        prestataire=prestataire,
                        prescripteur=prescripteur,
                        numero=sv.NUMERO,
                        type_sinistre="",
                        prix_unitaire=0,
                        frais_reel=sv.FRAIS_REEL,
                        ticket_moderateur=sv.TICKET_MODERATEUR,
                        depassement=sv.DEPASSEMENT,
                        nombre_demande=1,
                        nombre_accorde=1,
                        part_assure=sv.PART_ASSURE,
                        part_compagnie=sv.PART_COMPAGNIE,
                        date_survenance=sv.DATE_SURVENANCE,
                        statut='ACCORDE' if sv.STATUT == 'ACCEPTEE' else None,
                        observation="CREATED BY IMPORT M"
                        )

                    new.save()

                    cpt = cpt + 1

                    #le marquﾃｩ comme importﾃｩ
                    sv.STATUT_IMPORT = True
                    sv.save()


        response = {
            'statut': 1,
            'message': "SINISTRES IMPORTﾃ唄",
            'data': {
                'nombre_sinistres_crees': cpt,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "PAS DE DONNﾃ右S DISPONIBLES",
        }

    return JsonResponse(response)


def import_sinistre_manuellement(request):
    # Gﾃｩnﾃｩrer un UUID unique pour la session
    session_uuid = str(uuid.uuid4())

    # Rﾃｩcupﾃｩrer les 10 000 premiers IDs des objets ﾃ mettre ﾃ jour
    ids_to_update = list(SinistreVeos.objects.filter(STATUT_IMPORT=False, SESSION_UUID__isnull=True).order_by('ID_SIN').values_list('id', flat=True)[:10000])

    # Mettre ﾃ jour les objets correspondant aux IDs rﾃｩcupﾃｩrﾃｩs
    SinistreVeos.objects.filter(id__in=ids_to_update).update(SESSION_UUID=session_uuid)

    # Rﾃｩcupﾃｩrer les objets mis ﾃ jour
    sinistres_veos = SinistreVeos.objects.filter(STATUT_IMPORT=False, SESSION_UUID=session_uuid).order_by('ID_SIN')[:10000]
    cpt = 0
    cpt_updated = 0

    dd(sinistres_veos)
    if sinistres_veos:

        for sv in sinistres_veos:

            # dd(sv)

            formulegaranties = FormuleGarantie.objects.filter(code=sv.CD_FORMULE)
            formulegarantie = formulegaranties.first() if formulegaranties else None

            prestataires = Prestataire.objects.filter(id_per=sv.ID_PER_PRESTA)
            prestataire = prestataires.first() if prestataires else None

            prescripteurs = Prescripteur.objects.filter(veos_id_per=sv.ID_MED_PRESC)
            prescripteur = prescripteurs.first() if prescripteurs else None

            compagnies = Compagnie.objects.filter(code=sv.CODE_COMPAGNIE)
            compagnie = compagnies.first() if compagnies else None

            affections = Affection.objects.filter(code_cim_10=sv.CODE_AFFECTION, status=True)
            affection = affections.first() if affections else None

            actes = Acte.objects.filter(code=sv.CODE_ACTE)
            acte = actes.first() if actes else None

            aliments = Aliment.objects.filter(veos_code_aliment=sv.ID_ADHERENT)
            aliment = aliments.first() if aliments else None

            adherent_principal = aliment.adherent_principal if aliment else None  # a corriger renvoi si vide nonetype has no attr.

            polices = Police.objects.filter(veos_id_npol=sv.ID_NPOL)
            police = polices.first() if polices else None

            criteres_dates = Q(date_debut_effet__lte=sv.DATE_SINISTRE) & (
                        Q(date_fin_effet__gte=sv.DATE_SINISTRE) | Q(date_fin_effet__isnull=True))
            periodes_couvertures = PeriodeCouverture.objects.filter(police=police, statut=Statut.ACTIF, statut_validite=StatutValidite.VALIDE).filter(
                criteres_dates)
            periode_couverture = periodes_couvertures.first() if periodes_couvertures else None

            # dd(periode_couverture.id)


            # Check if the record already exists based on a unique field (code)
            existing_record = Sinistre.objects.filter(veos_id_sin=sv.ID_SIN).first()

            # dd(existing_record)
            if existing_record:

                if not existing_record.dossier_sinistre:
                    pprint("-- LE SINISTRE " + str(sv.NUMERO_DOSSIER) + " EXISTE DEJA ")

                    #existing_record.veos_id_sin = sv.ID_SIN
                    #existing_record.veos_id_npol = sv.ID_NPOL
                    #existing_record.veos_code_aliment = sv.ID_ADHERENT
                    #existing_record.veos_code_cie = sv.CODE_COMPAGNIE
                    #existing_record.veos_code_acte = sv.CODE_ACTE
                    #existing_record.veos_code_affection = sv.CODE_AFFECTION
                    #existing_record.veos_code_prestataire = sv.ID_PER_PRESTA
                    #existing_record.veos_code_prescripteur = sv.ID_MED_PRESC
                    #existing_record.created_by = None
                    #existing_record.updated_price_by = None
                    #existing_record.approuved_by = None
                    #existing_record.served_by = None
                    #existing_record.dossier_sinistre = None
                    #existing_record.aliment = aliment
                    #existing_record.adherent_principal = adherent_principal if adherent_principal else None
                    #existing_record.compagnie = compagnie
                    #existing_record.police = police
                    #existing_record.periode_couverture = periode_couverture
                    existing_record.formulegarantie = formulegarantie
                    existing_record.bareme = None
                    existing_record.acte = acte
                    existing_record.medicament = None
                    existing_record.affection = affection
                    existing_record.prestataire = prestataire
                    existing_record.prescripteur = prescripteur
                    #existing_record.numero = sv.NUMERO_DOSSIER
                    existing_record.type_sinistre = ""
                    existing_record.prix_unitaire = 0
                    existing_record.frais_reel = sv.FRAIS_REEL
                    existing_record.ticket_moderateur = sv.TICKET_MODERATEUR
                    existing_record.depassement = sv.DEPASSEMENT_EXCLUSION
                    existing_record.nombre_demande = 1
                    existing_record.nombre_accorde = 1
                    existing_record.part_assure = sv.PART_ASSURE
                    existing_record.part_compagnie = sv.PART_COMPAGNIE
                    existing_record.tps = sv.TPS
                    existing_record.far = sv.FAR
                    existing_record.ticket_prefinance = sv.TICKET_PREFINANCE
                    existing_record.net_regle = sv.NET_REGLE
                    existing_record.montant_remboursement_accepte = sv.NET_REGLE
                    existing_record.montant_refacture_compagnie = sv.PART_COMPAGNIE #a refacture ﾃ la compagnie
                    existing_record.montant_refacture_client = sv.TICKET_PREFINANCE
                    existing_record.date_survenance = sv.DATE_SINISTRE
                    existing_record.date_ordonnancement = sv.DATE_ORD
                    existing_record.date_reglement = sv.DATE_REG
                    existing_record.numero_bordereau = sv.NUMERO_BORDEREAU
                    existing_record.numero_lettre_cheque = sv.NUMERO_LCHQ
                    existing_record.statut = 'ACCORDE'
                    existing_record.observation = "UPDATE SINISTRES IMPORTES AVANT DEMARRAGE 03/09/2024"
                    existing_record.import_stats = 1
                    existing_record.save()

                    # le marquﾃｩ comme importﾃｩ
                    sv.STATUT_IMPORT = True
                    sv.save()

                    cpt_updated = cpt_updated + 1

            else:

                if acte and aliment:

                    new = Sinistre(
                        veos_id_sin=sv.ID_SIN,
                        veos_id_npol=sv.ID_NPOL,
                        veos_code_aliment=sv.ID_ADHERENT,
                        veos_code_cie=sv.CODE_COMPAGNIE,
                        veos_code_acte=sv.CODE_ACTE,
                        veos_code_affection=sv.CODE_AFFECTION,
                        veos_code_prestataire=sv.ID_PER_PRESTA,
                        veos_code_prescripteur=sv.ID_MED_PRESC,
                        created_by=None,
                        updated_price_by=None,
                        approuved_by=None,
                        served_by=None,
                        dossier_sinistre=None,
                        aliment=aliment,
                        adherent_principal=adherent_principal if adherent_principal else None,
                        compagnie=compagnie,
                        police=police,
                        periode_couverture=periode_couverture,
                        formulegarantie=formulegarantie,
                        bareme=None,
                        acte=acte,
                        medicament=None,
                        affection=affection,
                        prestataire=prestataire,
                        prescripteur=prescripteur,
                        numero=sv.NUMERO_DOSSIER,
                        type_sinistre="",
                        prix_unitaire=0,
                        frais_reel=sv.FRAIS_REEL,
                        ticket_moderateur=sv.TICKET_MODERATEUR,
                        depassement=sv.DEPASSEMENT_EXCLUSION,
                        nombre_demande=1,
                        nombre_accorde=1,
                        part_assure=sv.PART_ASSURE,
                        part_compagnie=sv.PART_COMPAGNIE,
                        tps=sv.TPS,
                        far=sv.FAR,
                        ticket_prefinance=sv.TICKET_PREFINANCE,
                        net_regle=sv.NET_REGLE, #net reglﾃｩ au bﾃｩnﾃｩficiaire de remboursement
                        montant_remboursement_accepte=sv.NET_REGLE, # payﾃｩ au bﾃｩnﾃｩficiaire de remboursement
                        montant_refacture_compagnie=sv.PART_COMPAGNIE, # a refacture ﾃ la compagnie
                        montant_refacture_client=sv.TICKET_PREFINANCE,
                        date_survenance=sv.DATE_SINISTRE,
                        date_ordonnancement=sv.DATE_ORD,
                        date_reglement=sv.DATE_REG,
                        numero_bordereau=sv.NUMERO_BORDEREAU,
                        numero_lettre_cheque=sv.NUMERO_LCHQ,
                        statut='ACCORDE',
                        observation="CREATED BY IMPORT M HISTORIQUE STAT SIN PAYE 03/09/2024",
                        import_stats=1
                        )

                    new.save()

                    cpt = cpt + 1

                    #le marquﾃｩ comme importﾃｩ
                    sv.STATUT_IMPORT = True
                    sv.save()

                else:
                    if not acte:
                        sv.OBSERVATION = f'Echec importation: Acte {sv.CODE_ACTE} non trouvﾃｩ sur la V2'
                        sv.STATUT_IMPORT = True
                        sv.save()

                    if not aliment:
                        sv.OBSERVATION = f'Echec importation: Aliment {sv.ID_ADHERENT} non trouvﾃｩ sur la V2'
                        sv.STATUT_IMPORT = True
                        sv.save()


        response = {
            'statut': 1,
            'message': "SINISTRES IMPORTﾃ唄",
            'data': {
                'nombre_sinistres_updated': cpt_updated,
                'nombre_sinistres_crees': cpt,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "PAS DE DONNﾃ右S DISPONIBLES",
        }

    return response


def import_sinistre_manuellement_cron():
    # Gﾃｩnﾃｩrer un UUID unique pour la session
    session_uuid = str(uuid.uuid4())

    # Rﾃｩcupﾃｩrer les 10 000 premiers IDs des objets ﾃ mettre ﾃ jour
    ids_to_update = list(SinistreVeos.objects.filter(STATUT_IMPORT=False, SESSION_UUID__isnull=True).order_by('ID_SIN').values_list('id', flat=True)[:10000])

    # Mettre ﾃ jour les objets correspondant aux IDs rﾃｩcupﾃｩrﾃｩs
    SinistreVeos.objects.filter(id__in=ids_to_update).update(SESSION_UUID=session_uuid)

    # Rﾃｩcupﾃｩrer les objets mis ﾃ jour
    sinistres_veos = SinistreVeos.objects.filter(STATUT_IMPORT=False, SESSION_UUID=session_uuid).order_by('ID_SIN')[:10000]
    cpt = 0
    cpt_updated = 0

    #dd(sinistres_veos)
    if sinistres_veos:

        for sv in sinistres_veos:

            # dd(sv)

            formulegaranties = FormuleGarantie.objects.filter(code=sv.CD_FORMULE)
            formulegarantie = formulegaranties.first() if formulegaranties else None

            prestataires = Prestataire.objects.filter(id_per=sv.ID_PER_PRESTA)
            prestataire = prestataires.first() if prestataires else None

            prescripteurs = Prescripteur.objects.filter(veos_id_per=sv.ID_MED_PRESC)
            prescripteur = prescripteurs.first() if prescripteurs else None

            compagnies = Compagnie.objects.filter(code=sv.CODE_COMPAGNIE)
            compagnie = compagnies.first() if compagnies else None

            affections = Affection.objects.filter(code_cim_10=sv.CODE_AFFECTION, status=True)
            affection = affections.first() if affections else None

            actes = Acte.objects.filter(code=sv.CODE_ACTE)
            acte = actes.first() if actes else None

            aliments = Aliment.objects.filter(veos_code_aliment=sv.ID_ADHERENT)
            aliment = aliments.first() if aliments else None

            adherent_principal = aliment.adherent_principal if aliment else None  # a corriger renvoi si vide nonetype has no attr.

            polices = Police.objects.filter(veos_id_npol=sv.ID_NPOL)
            police = polices.first() if polices else None

            criteres_dates = Q(date_debut_effet__lte=sv.DATE_SINISTRE) & (
                        Q(date_fin_effet__gte=sv.DATE_SINISTRE) | Q(date_fin_effet__isnull=True))
            periodes_couvertures = PeriodeCouverture.objects.filter(police=police, statut=Statut.ACTIF).filter(
                criteres_dates)
            periode_couverture = periodes_couvertures.first() if periodes_couvertures else None

            # dd(periode_couverture.id)


            # Check if the record already exists based on a unique field (code)
            existing_record = Sinistre.objects.filter(veos_id_sin=sv.ID_SIN).first()

            # dd(existing_record)
            if existing_record:

                if not existing_record.dossier_sinistre:
                    pprint("-- LE SINISTRE " + str(sv.NUMERO_DOSSIER) + " EXISTE DEJA ")

                    #existing_record.veos_id_sin = sv.ID_SIN
                    #existing_record.veos_id_npol = sv.ID_NPOL
                    #existing_record.veos_code_aliment = sv.ID_ADHERENT
                    #existing_record.veos_code_cie = sv.CODE_COMPAGNIE
                    #existing_record.veos_code_acte = sv.CODE_ACTE
                    #existing_record.veos_code_affection = sv.CODE_AFFECTION
                    #existing_record.veos_code_prestataire = sv.ID_PER_PRESTA
                    #existing_record.veos_code_prescripteur = sv.ID_MED_PRESC
                    #existing_record.created_by = None
                    #existing_record.updated_price_by = None
                    #existing_record.approuved_by = None
                    #existing_record.served_by = None
                    #existing_record.dossier_sinistre = None
                    #existing_record.aliment = aliment
                    #existing_record.adherent_principal = adherent_principal if adherent_principal else None
                    #existing_record.compagnie = compagnie
                    #existing_record.police = police
                    #existing_record.periode_couverture = periode_couverture
                    existing_record.formulegarantie = formulegarantie
                    existing_record.bareme = None
                    existing_record.acte = acte
                    existing_record.medicament = None
                    existing_record.affection = affection
                    existing_record.prestataire = prestataire
                    existing_record.prescripteur = prescripteur
                    #existing_record.numero = sv.NUMERO_DOSSIER
                    existing_record.type_sinistre = ""
                    existing_record.prix_unitaire = 0
                    existing_record.frais_reel = sv.FRAIS_REEL
                    existing_record.ticket_moderateur = sv.TICKET_MODERATEUR
                    existing_record.depassement = sv.DEPASSEMENT_EXCLUSION
                    existing_record.nombre_demande = 1
                    existing_record.nombre_accorde = 1
                    existing_record.part_assure = sv.PART_ASSURE
                    existing_record.part_compagnie = sv.PART_COMPAGNIE
                    existing_record.tps = sv.TPS
                    existing_record.far = sv.FAR
                    existing_record.ticket_prefinance = sv.TICKET_PREFINANCE
                    existing_record.net_regle = sv.NET_REGLE
                    existing_record.montant_remboursement_accepte = sv.PART_INOV
                    existing_record.montant_refacture_compagnie = sv.PART_COMPAGNIE
                    existing_record.montant_refacture_client = sv.TICKET_PREFINANCE
                    existing_record.date_survenance = sv.DATE_SINISTRE
                    existing_record.date_ordonnancement = sv.DATE_ORD
                    existing_record.date_reglement = sv.DATE_REG
                    existing_record.date_paiement = sv.DATE_REG
                    existing_record.numero_bordereau = sv.NUMERO_BORDEREAU
                    existing_record.numero_lettre_cheque = sv.NUMERO_LCHQ
                    existing_record.statut = 'ACCORDE'
                    existing_record.observation = "UPDATE SINISTRES IMPORTES AVANT DEMARRAGE 03/09/2024"
                    existing_record.import_stats = 1
                    existing_record.save()

                    # le marquﾃｩ comme importﾃｩ
                    sv.STATUT_IMPORT = True
                    sv.save()

                    cpt_updated = cpt_updated + 1

            else:

                if acte and aliment:

                    new = Sinistre(
                        veos_id_sin=sv.ID_SIN,
                        veos_id_npol=sv.ID_NPOL,
                        veos_code_aliment=sv.ID_ADHERENT,
                        veos_code_cie=sv.CODE_COMPAGNIE,
                        veos_code_acte=sv.CODE_ACTE,
                        veos_code_affection=sv.CODE_AFFECTION,
                        veos_code_prestataire=sv.ID_PER_PRESTA,
                        veos_code_prescripteur=sv.ID_MED_PRESC,
                        created_by=None,
                        updated_price_by=None,
                        approuved_by=None,
                        served_by=None,
                        dossier_sinistre=None,
                        aliment=aliment,
                        adherent_principal=adherent_principal if adherent_principal else None,
                        compagnie=compagnie,
                        police=police,
                        #periode_couverture=periode_couverture, #IMPORT STATS
                        formulegarantie=formulegarantie,
                        bareme=None,
                        acte=acte,
                        medicament=None,
                        affection=affection,
                        prestataire=prestataire,
                        prescripteur=prescripteur,
                        numero=sv.NUMERO_DOSSIER,
                        type_sinistre="",
                        prix_unitaire=0,
                        frais_reel=sv.FRAIS_REEL,
                        ticket_moderateur=sv.TICKET_MODERATEUR,
                        depassement=sv.DEPASSEMENT_EXCLUSION,
                        nombre_demande=1,
                        nombre_accorde=1,
                        part_assure=sv.PART_ASSURE,
                        part_compagnie=sv.PART_COMPAGNIE,
                        tps=sv.TPS,
                        far=sv.FAR,
                        ticket_prefinance=sv.TICKET_PREFINANCE,
                        net_regle=sv.NET_REGLE,
                        montant_remboursement_accepte=sv.PART_INOV,
                        montant_refacture_compagnie=sv.PART_COMPAGNIE,
                        montant_refacture_client=sv.TICKET_PREFINANCE,
                        date_survenance=sv.DATE_SINISTRE,
                        date_ordonnancement=sv.DATE_ORD,
                        date_reglement=sv.DATE_REG,
                        date_paiement=sv.DATE_REG,
                        numero_bordereau=sv.NUMERO_BORDEREAU,
                        numero_lettre_cheque=sv.NUMERO_LCHQ,
                        statut='ACCORDE',
                        observation="CREATED BY IMPORT M HISTORIQUE STAT SIN PAYE 03/09/2024",
                        import_stats=1
                        )

                    new.save()

                    cpt = cpt + 1

                    #le marquﾃｩ comme importﾃｩ
                    sv.STATUT_IMPORT = True
                    sv.save()

                else:
                    if not acte:
                        sv.OBSERVATION = f'Echec importation: Acte {sv.CODE_ACTE} non trouvﾃｩ sur la V2'
                        sv.STATUT_IMPORT = True
                        sv.save()

                    if not aliment:
                        sv.OBSERVATION = f'Echec importation: Aliment {sv.ID_ADHERENT} non trouvﾃｩ sur la V2'
                        sv.STATUT_IMPORT = True
                        sv.save()


            print(f'- Importation: SinistreVeos {sv.NUMERO_DOSSIER} importﾃｩ')


        response = {
            'statut': 1,
            'message': "SINISTRES IMPORTﾃ唄",
            'data': {
                'nombre_sinistres_updated': cpt_updated,
                'nombre_sinistres_crees': cpt,
            }
        }

    else:

        response = {
            'statut': 0,
            'message': "PAS DE DONNﾃ右S DISPONIBLES",
        }

    return JsonResponse(response)


def get_specialites_lists():
    # Construire l'URL de l'API en utilisant la variable API_URL
    api_url = f"{API_URL}/boBy/list"

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_SPECIALITE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        specialites = api_response['responses'][0]['beans']

        #dd(specialites)
        for specialite in specialites:
            # Check if the record already exists based on a unique field (code)
            existing_record = FormuleGarantie.objects.filter(code=specialite["CODE"]).first()

            # dd(existing_record)
            if existing_record:
                existing_record.code = specialite['CODE']
                existing_record.libelle = specialite['NOM']
                existing_record.save()

            else:
                new = Specialite(nom=specialite['NOM'],
                                 code=specialite['CODE']
                                 )
                new.save()

        return True



def get_actes_lists():

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_ACTE_TACHE_CRON",
                "params": {
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        actes = api_response['responses'][0]['beans']

        for acte in actes:
            # Check if the record already exists based on a unique field (code)
            existing_record = Acte.objects.filter(code=acte["cd_gar"]).first()

            rubriques = Rubrique.objects.filter(code=acte['cd_group_inov'])
            rubrique = rubriques.first() if rubriques else None

            accord_automatique = True if acte['acc_pre'] == 0 else False

            if existing_record:
                #existing_record.libelle = acte['lib_gar']
                existing_record.updated_at = datetime.now(tz=timezone.utc)
                existing_record.save()


            else:
                new = Acte(rubrique=rubrique,
                           nom=acte['lib_gar'],
                           code=acte['cd_gar'],
                           accord_automatique=accord_automatique
                           )

                new.save()

        return True




def import_apporteurs_manuellement(request):
    cpt_anciens_apporteurs = 0
    cpt_nouveaux_apporteurs = 0
    cpt_tous_apporteurs = 0

    apporteurs_veos = ApporteurVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_PER_APPORTEUR')[:1000]

    # dd(apporteurs_veos)
    if apporteurs_veos:

        for apporteur in apporteurs_veos:
            # Check if the record already exists based on a unique field (code)
            existing_record = Apporteur.objects.filter(id_per=apporteur.ID_PER_APPORTEUR).first()

            bureau = Bureau.objects.filter(code=apporteur.NUM_SOC).first()
            police = Police.objects.filter(veos_id_npol=apporteur.ID_NPOL).first()
            pays = Pays.objects.filter(code=apporteur.CD_PAYS).first()
            apporteur_international = ApporteurInternational.objects.filter(code=apporteur.APPORTEUR_INTER).first()

            if police:

                if bureau:
                    if existing_record:

                        existing_record.nom = apporteur.NOM_APPORTEUR
                        existing_record.prenoms = apporteur.PRENOM_APPORTEUR
                        existing_record.code = apporteur.NUM_APPORTEUR
                        existing_record.telephone = apporteur.MOBILE
                        existing_record.email = apporteur.EMAIL
                        existing_record.adresse = apporteur.VILLE_ADR
                        if apporteur_international and not existing_record.apporteur_international:
                            existing_record.apporteur_international = apporteur_international

                        existing_record.save()

                        #Si pas de ligne dans apporteur_police avec le contrat
                        apporteur_police_old = ApporteurPolice.objects.filter(apporteur=existing_record, police=police, statut_validite=StatutValidite.VALIDE)
                        if not apporteur_police_old:
                            #Enregistrer les taux de retro
                            base_calcul_old = BaseCalcul.objects.get(code='COM_COURT')
                            appo_old = ApporteurPolice.objects.create(apporteur=existing_record,
                                                                  police=police,
                                                                  base_calcul=base_calcul_old,
                                                                  taux_com_affaire_nouvelle=apporteur.TXC_CINT,
                                                                  taux_com_renouvellement=apporteur.TXT_CINT,
                                                                  date_effet=apporteur.EFFET,
                                                                 )

                            appo_old.save()


                        cpt_anciens_apporteurs += 1

                    else:

                        new_apporteur = Apporteur(bureau=bureau,
                                                  nom=apporteur.NOM_APPORTEUR,
                                                  prenoms=apporteur.PRENOM_APPORTEUR,
                                                  code=apporteur.NUM_APPORTEUR,
                                                  telephone=apporteur.MOBILE,
                                                  email=apporteur.EMAIL,
                                                  adresse=apporteur.VILLE_ADR,
                                                  id_per=apporteur.ID_PER_APPORTEUR,
                                                  pays=pays,
                                                  apporteur_international=apporteur_international,
                                                  )
                        new_apporteur.save()

                        #Enregistrer les taux de retro
                        base_calcul = BaseCalcul.objects.get(code='COM_COURT')
                        appo = ApporteurPolice.objects.create(apporteur=new_apporteur,
                                                              police=police,
                                                              base_calcul=base_calcul,
                                                              taux_com_affaire_nouvelle=apporteur.TXC_CINT,
                                                              taux_com_renouvellement=apporteur.TXT_CINT,
                                                              date_effet=apporteur.EFFET,
                                                             )

                        appo.save()

                        cpt_nouveaux_apporteurs += 1

                    # le marquﾃｩ comme importﾃｩ
                    apporteur.STATUT_IMPORT = True
                    apporteur.save()

                    cpt_tous_apporteurs += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciens_apporteurs': cpt_anciens_apporteurs,
                'cpt_nouveaux_apporteurs': cpt_nouveaux_apporteurs,
                'cpt_tous_apporteurs': cpt_tous_apporteurs
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)



def import_apporteurs_manuellement_sans_contrat(request):
    cpt_anciens_apporteurs = 0
    cpt_nouveaux_apporteurs = 0
    cpt_tous_apporteurs = 0

    apporteurs_veos = ApporteurVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_PER_APPORTEUR')[:1000]

    # dd(apporteurs_veos)
    if apporteurs_veos:

        for apporteur in apporteurs_veos:
            # Check if the record already exists based on a unique field (code)
            existing_record = Apporteur.objects.filter(Q(id_per=apporteur.ID_PER_APPORTEUR) | Q(code=apporteur.NUM_APPORTEUR)).first()

            bureau = Bureau.objects.filter(code=apporteur.NUM_SOC).first()
            pays = Pays.objects.filter(code=apporteur.CD_PAYS).first()
            apporteur_international = ApporteurInternational.objects.filter(code=apporteur.APPORTEUR_INTER).first()


            if bureau:
                if existing_record:

                    """ existing_record.nom = apporteur.NOM_APPORTEUR
                    existing_record.prenoms = apporteur.PRENOM_APPORTEUR
                    existing_record.code = apporteur.NUM_APPORTEUR
                    existing_record.telephone = apporteur.MOBILE
                    existing_record.email = apporteur.EMAIL
                    existing_record.adresse = apporteur.VILLE_ADR
                    if apporteur_international and not existing_record.apporteur_international:
                        existing_record.apporteur_international = apporteur_international

                    existing_record.save() """

                    cpt_anciens_apporteurs += 1
                    pass

                else:

                    new_apporteur = Apporteur(bureau=bureau,
                                                nom=apporteur.NOM_APPORTEUR,
                                                prenoms=apporteur.PRENOM_APPORTEUR,
                                                code=apporteur.NUM_APPORTEUR,
                                                telephone=apporteur.MOBILE,
                                                email=apporteur.EMAIL,
                                                adresse=apporteur.VILLE_ADR,
                                                id_per=apporteur.ID_PER_APPORTEUR,
                                                pays=pays,
                                                apporteur_international=apporteur_international,
                                                )
                    new_apporteur.save()

                    cpt_nouveaux_apporteurs += 1

                # le marquﾃｩ comme importﾃｩ
                apporteur.STATUT_IMPORT = True
                apporteur.save()

                cpt_tous_apporteurs += 1

        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciens_apporteurs': cpt_anciens_apporteurs,
                'cpt_nouveaux_apporteurs': cpt_nouveaux_apporteurs,
                'cpt_tous_apporteurs': cpt_tous_apporteurs
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)



#send sinistres to veos
def send_sinistres():
    # rﾃｩcupﾃｩrer les sinistres et les envoyer vers veos
    # sinistres = Sinistre.objects.filter(id=342121)
    # sinistres = Sinistre.objects.filter(statut_synchro_veos=False)

    # sinistres = Sinistre.objects.filter(
    #     aliment__veos_adherent_principal_id_per__isnull=False,
    #     police__client__veos_assure_sante_idper__isnull=False,
    #     dossier_sinistre_id__isnull=False
    # )

    sinistres = Sinistre.objects.filter(id=347937)
    # sinistres = Sinistre.objects.filter(id=347937)
    # 32863



    #  dd(sinistres)
    temps_debut = time.time()

    # sinistres = Sinistre.objects.filter(bordereau_ordonnancement_id__isnull=False, statut_synchro_veos=False).order_by('-id')[:50]
    nombre_sinistre = sinistres.count()


    for sinistre in sinistres:

        date_str = sinistre.created_at.strftime("%Y-%m-%d %H:%M:%S")
        date_format = "%Y-%m-%d %H:%M:%S"

        timestamp = datetime.strptime(date_str, date_format).timestamp()

        data = {
            "idPol": getPoliceIdPol(sinistre.police.veos_id_npol), #905428,  #-- idpol police
            "numSoc": sinistre.police.bureau.code, #-- numsoc bureau inov
            "idElem": sinistre.aliment.veos_code_aliment, #91756,  #-- id benef PEC
            "typeElem": "ADHERENT", #--Fixe
            "mouvement": "OUVSIN", #--Fixe
            "motif": "PECVALIDEE", #-- Fixe si PEC validee
            "numAuto": "O", #-- Fixe
            "txResp": "0", #-- Fixe
            "circonstance": "100991", #-- Code produit
            "causeCirconstance": "",
            "dateOuverture": sinistre.created_at.strftime("%d/%m/%Y"), #, #--Date du jour
            "dateSurvenance": sinistre.date_survenance.strftime("%d/%m/%Y"), #, #-- Date de soins
            "heureSurvenance": int(timestamp), #1693867781, #-- Heure de soins (timestamp)
            "dateDeclaration": sinistre.date_survenance.strftime("%d/%m/%Y"), #sinistre.created_at, #-- Date du jour
            "codeProduit": "100991", #--Code produit
            "loadAssure": "1", #-- Fixe
            "risque1": "",
            "infos": {
                "GESTIONNAIRE": "",
                "CDACT": sinistre.acte.code, #-- code acte
                "CDACTDEF1": "",
                "CDACTDEF2": "",
                "CDGAR": "",
                "CDGARGROUPACT": "",
                "CDGARGROUPRUB": "",
                "CDRUB": "", #"G66735CI01" -- Code rubrique de garantie
                "ADENTREE": "",
                "ADSORTIE": "",
                "CDMOTPOL": "MODIFGAR", #-- Motif police client
                "CDMVTPOL": "AVENANT", #--Mvt police client
                "RUBAFF": "TOUTES", #-- Rubrique affection
                "CDAFF": "", #sinistre.affection.code_cim_10 if sinistre.affection else None, #-- code affection
                "MNT_PLAFONNE": "45200", #--
                "TYPE_PLAFOND_APPLIQUE": "AUCUN PLAFOND APPLIQUE",
                "MODE_TM_LIB": "aucun (TM d\u00e9duit des remboursement)",
                "TM_AVCIE": "0", #--Ticket moderateur prﾃｩfinancﾃｩ
                "TM_DED": "4520", #-- ticket modﾃｩrateur
                "PEC_INOV": "40680", #-- Part INOV
                "COMMENT_PRESTA": "TEST SINISTRE SANTE V2",
                "DEPASSEMENT": "0", #-- Depassement
                "CHARGE_BENEF": int(sinistre.part_assure), #-- Part beneficiaire
                "MNT": int(sinistre.frais_reel), #-- Frais rﾃｩels
                "ACTE_LIB": sinistre.acte.libelle, #--Libellﾃｩ acte
                "GAR_LIB": "", #-- Libﾃｩllﾃｩ garantie
                "CD_MON": "", #--devise
                "CD_GAR": "", #-- code garantie
                "CONSO_GAR": "6750", #-- ??
                "CONSO_RUB": "0", #-- ??
                "SIT_ADH_CD": "ACTIF", #-- code statut benef
                "SIT_ADH_LIB": "ACTIF", #-- lib statut benef
                "PEC_STATUT": "ACCEPTEE", #-- statut PEC
                "PEC_ACCEPTEE_LE": sinistre.reviewed_at.strftime("%d/%m/%Y") if sinistre.reviewed_at else None, #-- Date acceptation PEC
                "SIT_ADH_MOTIF": "",
                "HOSPI_NBJ": "", #-- Nbre jours hospit
                "HOSPI_DEB": "", #-- Date debut hospit
                "HOSPI_FIN": "", #-- date fin hospit
                "CSOINS": "",
                "MEDSIN": "12310891", #-- Medecin traitant
                "MPRESCRIPTEUR": "", #-- medecin prescripteur
                "COEFF_FORCE": "",
                "PREFIN_DOSSIER": "",
                "NATURE_DEMANDE": "Soins", #--Fixe
                "G_ID_GRPC": "97661", #--Id formule de formule de garantie
                "G_DUREE": "1",
                "G_UNITE": "ANNEE",
                "G_PLAFOND": "",
                "G_NBMAX": "",
                "RESUME_G_NB_MAX": "",
                "RESUME_G_PL_PER": "-",
                "G_PLAFOND_LETTRE": "",
                "RESUME_G_PL_LETTRE": "-",
                "G_PLAFOND_ACTE": "",
                "RESUME_G_PL_ACTE": "-",
                "G_FR_MT": "10", #-- ??
                "G_FR_TYPE": "3", #-- ??
                "G_FR_MAX": "",
                "G_FR_MIN": "",
                "G_FR_TYPE_LIB": "% sinistre",
                "G_MODE_TM": "DED", #-- Code mode de calcul du TM
                "G_MODE_TM_LIB": "aucun (TM d\u00e9duit des remboursement)", #-- libellﾃｩ mode de calcul du TM
                "RESUME_TM": "10 %", #--ticket modﾃｩrateur
                "ACT_COEFF": "",
                "ACT_LETTRE": "",
                "ACT_PLAFOND": "",
                "RESUME_ACTE": "Aucun param\u00e9trage dans la nomenclature de l`acte",
                "ACT_DEF2": "",
                "ACT_GROL_CODE": "INFIRMERIE", #-- Code regroupement acte INOV
                "ACT_GROL_LIB": "INFIRMIER ET ACTES D'INFIRMERIE", #-- Libellﾃｩ regroupement INOV
                "RUB_ID_GRPC": "",
                "RUB_DUREE": "",
                "RUB_UNITE": "",
                "RUB_PLAFOND": "",
                "RUB_NBMAX": "",
                "RESUME_RUB_NB_MAX": "",
                "RESUME_RUB_PL_PER": "-",
                "RUB_LIB": "",
                "GAR_GROL_CODE": "GINFIRMERIE", #-- Code regroupement garantie INOV
                "G_LETTRE": "",
                "TM_AVOL": "0",
                "CALCULTM": "SURPLAFOND",
                "BORD": "",
                "BS": "",
                "HOSPI_NBJ_VISU": "",
                "FRAIS_SAISIS": "",
                "FRAIS_PHARMA": 45200, #-- ??
                "PL_PERS": "4918915", #-- plafond personne restant
                "PL_PERS_CONSOMME": "81085", #-- plafond personne consommﾃｩ
                "PL_PERS_FORM": "5000000", #--Plafond personne formule
                "QUANTITE": "1" #-- ??
            },
            "listPer":[
                {"idPer": "11686031", "categ": "USEREXT", "cdCon": "CREATEUR"}, #--id du user extranet veos qui te permet de consommer ce ws
                {"idPer": sinistre.prestataire.id_per, "categ": "PRESTAMED", "cdCon": "PRESTAMED"}, #--id du prestataire veos
                {"nom": sinistre.aliment.nom, "prenom": sinistre.aliment.prenoms, "categ": "BENEF", "cdCon": "BENEF", "type": "P", "ref": sinistre.aliment.veos_code_aliment}, #--Benef PEC
                {"idPer": sinistre.police.veos_id_npol, "nom": sinistre.aliment.adherent_principal.nom, "prenom": sinistre.aliment.adherent_principal.prenoms, "categ": "ADHERENT", "cdCon": "ADHERENT", "type": "P", "ref": sinistre.aliment.adherent_principal.veos_code_aliment}, #-- Adherent principal du benef PEC
                {"idPer": sinistre.police.client.veos_assure_sante_idper, "nom": sinistre.police.client.nom, "prenom": sinistre.police.client.prenoms, "categ": "ASSA", "cdCon": "ASSA", "type": "M", "ref": "11291732"} #--Client police
            ],
            "risque4": "97661" # id de la formule de garantie
        }


        try:
            api_response = create_sinistre_veos(data)
            CronLog.objects.create(action="create_sinistre", table="sinistre", description="Create sinistre Nﾂｰ " + sinistre.numero + " on veos with id = " + str(api_response['id']) + " et numero = " + str(api_response['num'])).save()
            print("笨 === Sinistre envoyﾃｩ avec succﾃｨs === :", api_response)

        except ValueError as e:
            response = []
            CronLog.objects.create(action="create_sinistre", table="sinistre",
                                   description="Error when creating sinistre Nﾂｰ " + sinistre.numero).save()
            print("Erreur :", e)

    temps_fin = time.time()
    temps_execution = temps_fin - temps_debut

    message = f"沐 Envoie des ({nombre_sinistre}) sinistres terminﾃｩs en {temps_execution:.7f} sﾃｩcondes. 沐"

    pprint(message)
    return message


def get_photos_lists_of_bureau(bureau_id):
    polices = Police.objects.filter(bureau_id=bureau_id)
    for police in polices:
        get_photos_lists(police.numero)

    message = f"Rﾃｩcupﾃｩration des photos terminﾃｩe."

    pprint(message)
    return message


def get_photos_lists(numero_police):

    police = Police.objects.filter(numero=numero_police).first()
    aliments_ids = AlimentFormule.objects.filter(formule__police=police).values_list('aliment_id', flat=True)

    aliments = Aliment.objects.filter(id__in=aliments_ids).filter(veos_numero_carte__isnull=False, has_photo_veos=True, statut_import_photo_veos=False).filter(Q(photo__isnull=True) | Q(photo__exact=''))[:1000]

    #aliments = Aliment.objects.filter(veos_numero_carte__isnull=False, photo__isnull=True)[:1000]
    #dd(aliments_ids)

    list_aliment_with_photo_importees = []
    list_aliment_with_photo_non_importees = []

    for aliment in aliments:
        veos_id_aliment = aliment.carte_active().numero[:-1]
        ID_ADOC = get_aliment_id_adoc(veos_id_aliment)

        pprint("veos_id_aliment")
        pprint(veos_id_aliment)

        # Appel de la fonction pour effectuer l'appel ﾃ l'API
        api_response = get_document(ID_ADOC)

        if 'datas' in api_response:
            base64_img = api_response['datas']

            # Remove the prefix 'data:image/jpeg;base64,' from the base64 string
            img_data = base64_img.split(';base64,')[-1]

            # Decode base64 to binary data
            binary_data = base64.b64decode(img_data)

            # Create ContentFile object from binary data
            content_file = ContentFile(binary_data, name='photo.jpg')  # Adjust name as needed

            # Assign content file to the photo field of the Aliment instance
            photo_name = f'photo_{veos_id_aliment}.jpg'
            aliment.photo.save(photo_name, content_file, save=True)

            #indiquer que la photo a ﾃｩtﾃｩ importﾃｩe
            aliment.statut_import_photo_veos = True

            # mettre ﾃ jour le statut importation de la photo
            aliment.save()

            list_aliment_with_photo_importees.append(aliment)
        else:
            # indiquer que la photo a ﾃｩtﾃｩ importﾃｩe
            aliment.has_photo_veos = False

            # mettre ﾃ jour le statut importation de la photo
            aliment.save()

            list_aliment_with_photo_non_importees.append(aliment)

        response = {
                "statut": 1,
                "STATS": {
                    "PHOTOS YES": list_aliment_with_photo_importees,
                    "PHOTOS NO": list_aliment_with_photo_non_importees,
                }
            }

    return response



def get_aliment_id_adoc(veos_id_aliment):
    ID_ADOC = ''

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
        "requests": [
            {
                "name": "WS_SANTE_MOB_IDADOC_BY_ADHERENT",
                "params": {
                    "ID_ADH": veos_id_aliment
                }
            }
        ]
    }

    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        adocs = api_response['responses'][0]['beans']

        for adoc in adocs:
            ID_ADOC = adoc["ID_ADOC"]


    return ID_ADOC


# send quittance to veos
def send_quittances():
    # rﾃｩcupﾃｩrer les quittances et les envoyer vers veos
    # sinistres = Sinistre.objects.filter(id=342121)

    quittances = Quittance.objects.filter(id=1)


    temps_debut = time.time()

    nombre_quittances = quittances.count()

    for quittance in quittances:

        # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
        data = {
            "requests": [
                {
                    "name": "WS_",
                    "params": {

                    }
                }
            ]
        }

        date_str = quittance.created_at.strftime("%Y-%m-%d %H:%M:%S")
        date_format = "%Y-%m-%d %H:%M:%S"

        timestamp = datetime.strptime(date_str, date_format).timestamp()

        # dd(sinistre)
        data = {
            "idPol": getPoliceIdPol(quittance.police.veos_id_npol),
            "nature": "COMPT",
            "dtEmission": quittance.date_emission.strftime("%d/%m/%Y"), #"14\/02\/2024",
            "dtDebut": quittance.date_debut.strftime("%d/%m/%Y"), #"14\/02\/2024",
            "dtFin": quittance.date_fin.strftime("%d/%m/%Y"), #"31/12/2024",
            "dtEch": None,
            "devise": quittance.devise.code, #"{{devise}}",
            "type": "CIE",
            "dtProchaineEcheance": None,
            "bonus": 0,
            "mntTaxe": quittance.taxe,
            "mntCatNat": 0,
            "mntCP": quittance.cout_police_courtier, #"25000",
            "mntFraisCie": quittance.cout_police_compagnie,
            "mntDivers": quittance.autres_taxes,
            "mntPrimeTotale": quittance.prime_ttc,
            "mntPrimeNet": quittance.prime_ht,
            "mntFraisCbt": 0,#quittance....,
            "mvt": None,
            "motif": None,
            "mntCommission": "20000",
            "mntTTC": "1700000"
        }

        #dd(data)

        data_quittance_com_gestion = {
            "idPol": 1116847,
            "nature": "COMPT",
            "dtEmission": "14\/02\/2024",
            "dtDebut": "01\/01\/2024",
            "dtFin": "31\/12\/2024",
            "devise": "XOF",
            "type": "COM",
            "mntTaxe": "0",
            "mntCP": "0",
            "mntFraisCie": "0",
            "mntDivers": "0",
            "mntPrimeTotale": "0",
            "mntPrimeNet": "0",
            "mntFraisCbt": "0",
            "mntCommission": 5000,
            "mntTTC": "0"
        }

        # dd(data)
        # Appel de la fonction pour effectuer l'appel ﾃ l'API
        api_response = create_sinistre_veos(data)

        dd(api_response)
        if api_response['statusCode'] == 0:
            # Access the "beans" key to get its content
            response = api_response['responses'][0]
            CronLog.objects.create(action="create_quittance", table="quittance",
                                   description="Create quittance Nﾂｰ " + quittance.numero + " on veos").save()

        else:
            response = []
            CronLog.objects.create(action="create_quittance", table="quittance",
                                   description="Error when creating quittance Nﾂｰ " + quittance.numero).save()

    temps_fin = time.time()

    temps_execution = temps_fin - temps_debut
    message = f"Le temps d'envoi de {nombre_quittances} quittances est de {temps_execution:.7f} secondes."

def get_taux_euro_by_devise(devise):

    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    data = {
            "requests": [
                {
                    "name": "WS_GET_TAUX_EURO_BY_DEVISE",
                    "params":
                        {
                            "devise": devise
                        }
                }
                ]
            }

    api_response = call_api_with_data(data)
    print(api_response)
    if api_response['statusCode'] == 0:
        # Access the "beans" key to get its content
        response = api_response['responses'][0]['beans']

        taux_euro = response[0]['TAUX_EURO']

        return float(taux_euro)

    else:
        return None
    
def get_taux_usd_by_devise(devise):
    euro_by_usd = get_taux_euro_by_devise("USD") 
    euro_by_devise = get_taux_euro_by_devise(devise)
    taux_usd = euro_by_usd/euro_by_devise if euro_by_usd and euro_by_devise else None
    return taux_usd   

def convert_montant_from_devise_to_euro_or_usd(devise_from, montant, euro_or_usd, taux_euro_or_usd=None):
    if taux_euro_or_usd and taux_euro_or_usd > 0:
        montant_convert = montant*taux_euro_or_usd
    elif euro_or_usd == "USD":
        montant_convert = montant*get_taux_usd_by_devise(devise_from)
    else:
        montant_convert = montant*get_taux_euro_by_devise(devise_from)
    return montant_convert


def get_clients_from_veos():
    cpt_anciens_clients = 0
    cpt_nouveaux_clients = 0
    cpt_tous_clients = 0
    cpt_grinter = 0
    cpt_apporteur = 0
    cpt_apporteur_inter = 0
    cpt_grinter_old = 0
    cpt_apporteur_old = 0
    cpt_apporteur_inter_old = 0
    # Les donnﾃｩes que vous souhaitez envoyer ﾃ l'API
    today = datetime.now(tz=timezone.utc).strftime('%d/%m/%Y')
    # today = date(2024, 9, 20)
    # yesterday = date(2024, 1, 1)
    yesterday = (datetime.now(tz=timezone.utc) - timedelta(days=1)).strftime('%d/%m/%Y')

    iterationx = None

    """ Restaurer le gap de client inexistant """
    if date.today() <= date(2024, 9, 20):
        # Date de dﾃｩpart
        today = date(2024, 9, 20)
        yesterday = date(2024, 1, 1)

        last_cron_client = CronLog.objects.filter(action="import", table="client").order_by("-pk").first()

        iterationx = int(last_cron_client.row) if last_cron_client and last_cron_client.row else 0

        # Intervalle initial de 7 jours
        init = timedelta(days=(iterationx*7))

        if (yesterday + init) > date(2024, 9, 20):
            response = {
                'statut': 1,
                'message': 'TERMINﾃ',
                'data': {
                }
            }
            return response

        # Calcul de la date de dﾃｩbut
        yesterday = yesterday + init

        # Calcul de la date de fin (7 jours aprﾃｨs le dﾃｩbut)
        today = yesterday + timedelta(days=7)

        yesterday = yesterday.strftime("%d/%m/%Y")
        today = today.strftime("%d/%m/%Y")

        iterationx += 1

        print(str(yesterday))
        print(str(today))

    data = {
        "requests": [
            {
                "name": "WS_V2_CLIENT_UPDATE", #WS_CLIENT_TACHE_CRON
                "params": {
                    'start_creation_date': yesterday,
                    'end_creation_date': today
                }
            }
        ]
    }
    # Appel de la fonction pour effectuer l'appel ﾃ l'API
    api_response = call_api_with_data(data)

    # dd(api_response)
    if api_response['statusCode'] == 0:

        clients = api_response['responses'][0]['beans']

        for client in clients:

            try:
                # Gestion apporteur
                apporteur = None
                if client["VEOS_APPORTEUR_IDPER"] != None and client["VEOS_APPORTEUR_IDPER"] != "":
                
                    type_personne_apporteur = None
                    if client["TYPE_APPORTEUR"] != None and client["TYPE_APPORTEUR"] != "":
                        type_personne_apporteur = TypePersonne.objects.filter(code=client["TYPE_APPORTEUR"].upper()).first()

                    type_apporteur = None
                    if client["TYPE_PERSONNE"] != None and client["TYPE_PERSONNE"] != "":
                        type_apporteur = TypeApporteur.objects.filter(code=client["TYPE_PERSONNE"].capitalize()).first()

                    bureau_apporteur = None
                    if client["CODE_BUREAU_APPORTEUR"] != None and client["CODE_BUREAU_APPORTEUR"] != "":
                        type_personne_apporteur = Bureau.objects.filter(code=client["CODE_BUREAU_APPORTEUR"]).first()

                    pays_apporteur = None
                    if client["PAYS_APPORTEUR"] != None and client["PAYS_APPORTEUR"] != "":
                        pays_apporteur = Pays.objects.filter(code=client["PAYS_APPORTEUR"]).first()
                    
                    
                    # check apporteur
                    apporteur_old = Apporteur.objects.filter(id_per=client["VEOS_APPORTEUR_IDPER"]).first()
                    if apporteur_old:
                        apporteur = apporteur_old
                        # apporteur.nom = client["NOM_APPORTEUR"]
                        # apporteur.pays = pays_apporteur
                        apporteur.adresse = client["ADRESSE_APPORTEUR"][:255] if client['ADRESSE_APPORTEUR'] != None and client['ADRESSE_APPORTEUR'] != "" else ""
                        # apporteur.bureau = bureau_apporteur
                        apporteur.email = client["EMAIL_APPORTEUR"]
                        apporteur.telephone = client["TELEPHONE_APP"][:25] if client['TELEPHONE_APP'] != None and client['TELEPHONE_APP'] != "" else ""
                        # apporteur.id_per = client["VEOS_APPORTEUR_IDPER"]
                        # apporteur.prenoms = client["PRENOMS_APPORTEUR"]
                        # apporteur.type_apporteur = type_apporteur
                        # apporteur.type_personne = type_personne_apporteur
                        # apporteur.observation = "Update from VEOS {{:%d-%m-%Y ﾃ %H:%M:%S}} with client {}".format(datetime.now(), client["CODE_CLIENT"])
                        cpt_apporteur_old += 1
                    else:
                        apporteur = Apporteur(code = client["CODE_APPORTEUR"],
                        nom = client["NOM_APPORTEUR"],
                        prenoms = client["PRENOMS_APPORTEUR"],
                        adresse = client["ADRESSE_APPORTEUR"][:255] if client['ADRESSE_APPORTEUR'] != None and client['ADRESSE_APPORTEUR'] != "" else "",
                        bureau = bureau_apporteur,
                        email = client["EMAIL_APPORTEUR"],
                        telephone = client["TELEPHONE_APP"][:25] if client['TELEPHONE_APP'] != None and client['TELEPHONE_APP'] != "" else "",
                        id_per = client["VEOS_APPORTEUR_IDPER"],
                        pays = pays_apporteur,
                        type_apporteur = type_apporteur,
                        type_personne = type_personne_apporteur)
                        # observation = "Imported from VEOS {{:%d-%m-%Y ﾃ %H:%M:%S}} with client {}".format(datetime.now(), client["CODE_CLIENT"]))
                        cpt_apporteur += 1
                    apporteur.save()

                    # check apporteur inter
                    if client["TYPE_PERSONNE"] != None and client["TYPE_PERSONNE"] != "" and client["TYPE_PERSONNE"].capitalize() == "Inter" :
                        apporteur_inter_old = ApporteurInternational.objects.filter(code=client["CODE_APPORTEUR"]).first()
                        if apporteur_inter_old:
                            apporteur_inter = apporteur_inter_old
                            # apporteur_inter.nom = client["NOM_APPORTEUR"]
                            cpt_apporteur_inter_old += 1
                        else:
                            apporteur_inter = ApporteurInternational(code=client["CODE_APPORTEUR"],
                                                                        nom=client["NOM_APPORTEUR"]
                                                                        )
                            cpt_apporteur_inter += 1
                        apporteur_inter.save()
                        apporteur.apporteur_international = apporteur_inter
                        apporteur.save()
                    
                # Gestion groupe inter
                grinter = None
                if client["CODE_GROUPE_INTERNATIONAL"] != None and client["CODE_GROUPE_INTERNATIONAL"] != "":
                    # check groupe inter
                    grinter_old = GroupeInter.objects.filter(code=client["CODE_GROUPE_INTERNATIONAL"]).first()
                    if grinter_old is None or grinter_old == "":
                        grinter = GroupeInter(code = client["CODE_GROUPE_INTERNATIONAL"],
                        nom = client["NOM_GROUPE_INTERNATIONAL"])
                        cpt_grinter += 1
                    else:
                        grinter = grinter_old
                        cpt_grinter_old += 1
                    if grinter and apporteur:
                        grinter.apporteur = apporteur
                    grinter.save()

                # Check if the record already exists based on a unique field (code)
                existing_record = Client.objects.filter(veos_client_idper=client["VEOS_CLIENT_IDPER"]).first()

                bureaux = Bureau.objects.filter(code=client['CODE_BUREAU_CLIENT'])
                bureau = bureaux.first() if bureaux else None
                
                payss = Pays.objects.filter(code=client['PAYS_CLIENT'])
                pays = payss.first() if payss else bureau.pays if bureau else None

                types_personnes = TypePersonne.objects.filter(code=client['TYPE_PERSONNE_CLIENT'].upper()) if client['TYPE_PERSONNE_CLIENT'] and client['TYPE_PERSONNE_CLIENT'] != "" else None 
                type_personne = types_personnes.first() if types_personnes else None

                civilites = Civilite.objects.filter(code=client['CIVILITE_CLIENT'].replace(".","").capitalize()) if client['CIVILITE_CLIENT'] and client['CIVILITE_CLIENT'] != "" else None 
                civilite = civilites.first() if civilites else None

                langues = Langue.objects.filter(code=client['LANGUE_CLIENT'])
                langue = langues.first() if langues else None

                # dd(existing_record)
                if existing_record:
                    existing_record.bureau = bureau
                    if type_personne: existing_record.type_personne = type_personne
                    #  if type_client: existing_record.type_client = type_client
                    #existing_record.pays = pays
                    # existing_record.code = client['code']
                    #existing_record.nom = client['NOM_CLIENT']
                    # existing_record.prenoms = client['PRENOMS_CLIENT']
                    # existing_record.civilite = civilite
                    # existing_record.sexe = client["SEXE_CLIENT"][0].capitalize() if client["SEXE_CLIENT"] != None and client["SEXE_CLIENT"] != "" else existing_record.sexe
                    # existing_record.date_naissance = client['DATE_NAISSANCE_CLIENT'][:10] if client["DATE_NAISSANCE_CLIENT"] != None and client["DATE_NAISSANCE_CLIENT"] != "" else existing_record.date_naissance
                    existing_record.telephone_mobile = client['TELEPHONE_MOBILE_CLIENT'][:20] if client['TELEPHONE_MOBILE_CLIENT'] != None and client['TELEPHONE_MOBILE_CLIENT'] != "" else ""
                    existing_record.telephone_fixe = client['TELEPHONE_FIXE_CLIENT'][:20] if client['TELEPHONE_FIXE_CLIENT'] != None and client['TELEPHONE_FIXE_CLIENT'] != "" else ""
                    existing_record.email = client['EMAIL_CLIENT']
                    # existing_record.langue = langue
                    existing_record.ville = client['VILLE_CLIENT'] if client['VILLE_CLIENT'] else existing_record.ville
                    # existing_record.adresse_postale = client['ADRESSE_POSTALE_CLIENT'] 
                    existing_record.adresse = client['ADRESSE_CLIENT'][:100] if client['ADRESSE_CLIENT'] != None and client['ADRESSE_CLIENT'] != "" else ""
                    existing_record.updated_at = datetime.now(tz=timezone.utc)
                    existing_record.save(update_fields=['bureau', 'type_personne', 'type_client', 'pays', 'code', 'nom', 'prenoms', 'civilite', 'sexe', 'date_naissance', 'telephone_mobile', 'telephone_fixe', 'email', 'langue', 'ville', 'adresse_postale', 'adresse', 'updated_at'])

                    cpt_anciens_clients += 1
                    client_instance = existing_record

                else:

                    new = Client(
                        bureau=bureau,
                        type_personne=type_personne,
                        #type_client=type_client,
                        pays=pays,
                        veos_client_idper=client['VEOS_CLIENT_IDPER'],
                        code=client['CODE_CLIENT'],
                        nom=client['NOM_CLIENT'],
                        prenoms=client['PRENOMS_CLIENT'],
                        civilite=civilite,
                        sexe=client["SEXE_CLIENT"][ 0 ].capitalize() if client["SEXE_CLIENT"] != None and client["SEXE_CLIENT"] != "" else "",
                        date_naissance=client['DATE_NAISSANCE_CLIENT'][:10] if client["DATE_NAISSANCE_CLIENT"] != None and client["DATE_NAISSANCE_CLIENT"] != "" else None,
                        telephone_mobile=client['TELEPHONE_MOBILE_CLIENT'][:20] if client['TELEPHONE_MOBILE_CLIENT'] != None and client['TELEPHONE_MOBILE_CLIENT'] != "" else "",
                        telephone_fixe=client['TELEPHONE_FIXE_CLIENT'][:20] if client['TELEPHONE_FIXE_CLIENT'] != None and client['TELEPHONE_FIXE_CLIENT'] != "" else "",
                        email=client['EMAIL_CLIENT'],
                        langue=langue,
                        longitude=client['LONGITUDE_CLIENT'],
                        latitude=client['LATITUDE_CLIENT'],
                        ville=client['VILLE_CLIENT'],
                        adresse_postale=client['ADRESSE_POSTALE_CLIENT'],
                        adresse=client['ADRESSE_CLIENT'][:100] if client['ADRESSE_CLIENT'] != None and client['ADRESSE_CLIENT'] != "" else "",
                        site_web="",
                        twitter="",
                        instagram="",
                        facebook="",
                        ancienne_ref="",
                        statut="ACTIF",
                        statut_relation="CLIENT",
                        )

                    new.save()

                    client_instance = new

                    cpt_nouveaux_clients += 1

                cpt_tous_clients += 1
              
                if grinter:
                    client_instance.groupe_international = grinter
                    client_instance.save()

            except Exception as e:
                print(str(client['VEOS_CLIENT_IDPER'])+"-----"+str(e))
                pass
        response = {
            'statut': 1,
            'message': 'TERMINﾃ',
            'data': {
                'all_cli': len(clients),
                'cli_old': cpt_anciens_clients,
                'cli_new': cpt_nouveaux_clients,
                'cli_treat': cpt_tous_clients,
                # 'x': iterationx,
                # 'gri_new': cpt_grinter,
                # 'app_new': cpt_apporteur,
                # 'appint_new': cpt_apporteur_inter,
                # 'gri_old': cpt_grinter_old,
                # 'app_old': cpt_apporteur_old,
                # 'appint_old': cpt_apporteur_inter_old
            }
        }
        return response

    #return True


def import_periode_veos_manuellement(request):

    cpt_anciennes_polices = 0
    cpt_nouvelles_polices = 0
    cpt_toutes_polices=0


    periode_veos = PeriodeVeos.objects.filter(STATUT_IMPORT=False).order_by('ID_NPOL')[:1000]
    cpt = 0

    #dd(polices_veos)
    if periode_veos:

        for periode in periode_veos:
            
            police = Police.objects.filter(veos_id_npol=periode.ID_NPOL).first()

            if police:

                """                 
                date_fin_periode_garantie = request.POST.get('date_fin_periode_garantie')

                mouvement_police = MouvementPolice.objects.create(police_id=police.id,
                                                                mouvement_id=request.POST.get('mouvement'),
                                                                motif_id=request.POST.get('motif'),
                                                                date_effet=request.POST.get('date_effet'),
                                                                date_fin_periode_garantie=date_fin_periode_garantie if date_fin_periode_garantie else None,
                                                                created_by=request.user
                                                                )

                mouvement_police.save()

                mouvement = Mouvement.objects.get(id=mouvement_police.mouvement_id)

                motif = Motif.objects.get(id=mouvement_police.motif_id)
                """
                #test = 1

                #si c'est un renouvellement, crﾃｩer une pﾃｩriode de couverture
                # if mouvement.code == "AVENANT":
                if 1 == 1:

                    existing_periode = PeriodeCouverture.objects.filter(police_id=police.id, date_debut_effet=periode.DEBUTEFFETOLD, date_fin_effet=periode.FINEFFETOLD).first()
                    
                    if not existing_periode:
                        # crﾃｩer une ligne dans pﾃｩriode de couverture
                        periode_couverture = PeriodeCouverture.objects.create(
                            police_id=police.id,
                            date_debut_effet=periode.DEBUTEFFETOLD,
                            date_fin_effet=periode.FINEFFETOLD,
                            statut_validite= StatutValidite.CLOTURE
                        ).save()
                        periode.STATUT_IMPORT = True
                    else:
                        periode.OBSERVATION = "periode " + str(existing_periode.id) + " existe deja"
                        periode.STATUT_IMPORT = True
                    periode.save()
                
        response = {
            'statut': 1,
            'message': 'IMPORTATION TERMINﾃ右',
            'data': {
                'cpt_anciennes_polices': cpt_anciennes_polices,
                'cpt_nouvelles_polices': cpt_nouvelles_polices,
                'cpt_toutes_polices': cpt_toutes_polices
            }
        }

    else:
        response = {
            'statut': 0,
            'message': 'PAS DE DONNﾃ右S DISPONIBLES',
        }

    return JsonResponse(response)


